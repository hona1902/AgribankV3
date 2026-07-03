from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook

from agribank_v3.quiz.database import QuizDatabase, QuizDatabaseError


@dataclass(frozen=True, slots=True)
class ImportResult:
    imported: int
    skipped: int


HEADER_ALIASES = {
    "id": {"id", "ma cau hoi"},
    "question_number": {"stt", "so cau", "so cau hoi"},
    "topic_code": {"ma nghiep vu", "ma chuyen de"},
    "topic_name": {"nghiep vu", "ten nghiep vu", "chuyen de"},
    "subject_name": {"chuyen de", "ten chuyen de"},
    "question_text": {"cau hoi", "noi dung cau hoi"},
    "option_a": {"dap an a", "cau a", "a"},
    "option_b": {"dap an b", "cau b", "b"},
    "option_c": {"dap an c", "cau c", "c"},
    "option_d": {"dap an d", "cau d", "d"},
    "correct_answer": {"dap an dung", "cau dung"},
    "source_reference": {"nguon", "nguon tham khao", "goi y"},
    "locked_answers": {
        "dap an khong dao",
        "cac dap an khong dao",
        "khoa dap an",
    },
}


def _normalize(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"\s+", " ", text.replace("đ", "d"))


def import_questions_from_excel(
    path: Path,
    database: QuizDatabase,
) -> ImportResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise QuizDatabaseError("File Excel không có dòng tiêu đề.")
        normalized_header = [_normalize(value) for value in header]
        columns: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for index, value in enumerate(normalized_header):
                if value in aliases:
                    columns[field] = index
                    break
        required = {
            "topic_code", "topic_name", "subject_name", "question_text",
            "option_a", "option_b", "correct_answer",
        }
        missing = sorted(required - columns.keys())
        if missing:
            raise QuizDatabaseError(
                "File Excel thiếu các cột bắt buộc: " + ", ".join(missing)
            )

        imported = 0
        skipped = 0
        prepared_topics: set[str] = set()
        prepared_subjects: dict[tuple[str, str], int] = {}
        for row in rows:
            def value(field: str) -> object:
                index = columns.get(field)
                return row[index] if index is not None and index < len(row) else ""

            question_text = str(value("question_text") or "").strip()
            if not question_text:
                skipped += 1
                continue
            code = str(value("topic_code") or "").strip()
            name = str(value("topic_name") or "").strip()
            if code not in prepared_topics:
                database.save_business_topic(code, name)
                prepared_topics.add(code)
            subject_name = str(value("subject_name") or "").strip()
            subject_key = (code, subject_name)
            if subject_key not in prepared_subjects:
                prepared_subjects[subject_key] = database.save_question_topic(
                    code,
                    subject_name,
                )
            raw_id = value("id")
            question_id = int(raw_id) if raw_id not in (None, "") else None
            raw_number = value("question_number")
            question_number = (
                int(raw_number) if raw_number not in (None, "") else None
            )
            database.save_question(
                question_id=question_id,
                question_number=question_number,
                topic_code=code,
                subject_id=prepared_subjects[subject_key],
                text=question_text,
                options={
                    letter: str(value(f"option_{letter.lower()}") or "")
                    for letter in "ABCD"
                },
                correct_answer=str(value("correct_answer") or ""),
                locked_answers=str(value("locked_answers") or ""),
                source_reference=str(value("source_reference") or ""),
            )
            imported += 1
        return ImportResult(imported=imported, skipped=skipped)
    finally:
        workbook.close()
