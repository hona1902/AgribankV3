from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from agribank_v3.quiz.models import Question


TITLE = "Bộ Câu Hỏi Ôn Tập, Khảo Sát Kiến Thức Nghiệp Vụ"
BURGUNDY = "8B1743"
YELLOW = "FFF200"
GREEN = "008000"
LIGHT_GREEN = "E7F6EC"
LIGHT_RED = "FDECEC"
BLACK = "000000"
WHITE = "FFFFFF"


def _row_height(text: str, width: int = 92, minimum: float = 22) -> float:
    lines = max(1, sum(max(1, (len(part) + width - 1) // width) for part in text.splitlines()))
    return max(minimum, min(120, 8 + lines * 16))


def export_questions_to_excel(
    path: Path,
    questions: list[Question],
    subtitle: str | None = None,
) -> None:
    if not questions:
        raise ValueError("Không có câu hỏi để xuất.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bo Cau Hoi"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 92
    sheet.column_dimensions["C"].width = 9
    sheet.column_dimensions["D"].width = 10

    sheet.merge_cells("A1:D1")
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Arial", size=18, bold=True, color=BURGUNDY)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    if subtitle is None:
        business_names = sorted({question.topic_name for question in questions})
        if len(business_names) == 1:
            subtitle = (
                f"Nghiệp vụ: {business_names[0]} "
                f"({len(questions)} Câu ngẫu nhiên)"
            )
        else:
            subtitle = f"Đề tổng hợp nghiệp vụ ({len(questions)} Câu ngẫu nhiên)"
    sheet.merge_cells("A2:D2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Arial", size=15, bold=True, color=BLACK)
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[2].height = 26

    sheet.merge_cells("A4:B4")
    sheet["A4"] = "Câu Hỏi"
    sheet["C4"] = "Đúng"
    sheet["D4"] = "Chọn"
    header_fill = PatternFill("solid", fgColor=YELLOW)
    thin = Side(style="thin", color=BLACK)
    medium = Side(style="medium", color=BLACK)
    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=medium, bottom=thin, left=thin, right=thin)
    sheet.row_dimensions[4].height = 24

    answer_validation = DataValidation(
        type="list",
        formula1='"A,B,C,D"',
        allow_blank=True,
    )
    answer_validation.error = "Chỉ chọn A, B, C hoặc D."
    answer_validation.errorTitle = "Đáp án không hợp lệ"
    sheet.add_data_validation(answer_validation)

    current_row = 5
    choice_cells: list[str] = []
    for number, question in enumerate(questions, start=1):
        question_row = current_row
        subject_suffix = (
            f" ({question.subject_name})"
            if question.subject_name
            else ""
        )
        sheet.cell(question_row, 1, f"Câu {number}:")
        sheet.cell(question_row, 2, f"{question.text}{subject_suffix}")
        sheet.cell(question_row, 3, question.correct_answer)
        sheet.cell(question_row, 4, "")
        sheet.row_dimensions[question_row].height = _row_height(question.text)

        for column in range(1, 5):
            cell = sheet.cell(question_row, column)
            cell.font = Font(name="Arial", size=11, bold=True)
            cell.alignment = Alignment(
                horizontal="center" if column in (1, 3, 4) else "left",
                vertical="top",
                wrap_text=True,
            )
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        answer_validation.add(sheet.cell(question_row, 4))
        choice_cells.append(f"D{question_row}")

        current_row += 1
        for letter, option_text in question.options.items():
            sheet.cell(current_row, 1, f"{letter}:")
            sheet.cell(current_row, 2, option_text)
            sheet.row_dimensions[current_row].height = _row_height(option_text)
            is_correct = letter == question.correct_answer
            for column in range(1, 5):
                cell = sheet.cell(current_row, column)
                cell.font = Font(
                    name="Arial",
                    size=11,
                    bold=is_correct,
                    color=GREEN if is_correct else BLACK,
                )
                cell.alignment = Alignment(
                    horizontal="right" if column == 1 else "left",
                    vertical="top",
                    wrap_text=True,
                    indent=1 if column == 2 else 0,
                )
                cell.border = Border(left=thin, right=thin)
            current_row += 1

        for column in range(1, 5):
            sheet.cell(current_row - 1, column).border = Border(
                left=thin,
                right=thin,
                bottom=medium,
            )

    for address in choice_cells:
        row = sheet[address].row
        sheet.conditional_formatting.add(
            address,
            FormulaRule(
                formula=[f'AND({address}<>"",{address}=C{row})'],
                fill=PatternFill("solid", fgColor=LIGHT_GREEN),
                font=Font(color=GREEN, bold=True),
            ),
        )
        sheet.conditional_formatting.add(
            address,
            FormulaRule(
                formula=[f'AND({address}<>"",{address}<>C{row})'],
                fill=PatternFill("solid", fgColor=LIGHT_RED),
                font=Font(color="A00000", bold=True),
            ),
        )

    sheet.auto_filter.ref = f"A4:D{current_row - 1}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:4"
    sheet.print_area = f"A1:D{current_row - 1}"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
