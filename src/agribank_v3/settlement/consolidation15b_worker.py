from __future__ import annotations

from collections import OrderedDict
from copy import copy
from decimal import Decimal
import json
import sys
import time
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agribank_v3.file_merge import merge_same_structure_csv_to_csv
from agribank_v3.settings import BranchProfile
from agribank_v3.settlement import SETTLEMENT_SPECS, SettlementOptions, SettlementRequest
from agribank_v3.settlement.processors import Mau1516Processor


def _log(message: str) -> None:
    print(message, flush=True)


def _branch_from_customer_id(customer_id: str, fallback: str) -> str:
    value = str(customer_id or "").strip()
    return value[:4] if len(value) >= 4 and value[:4].isdigit() else fallback


def _add_summary_sheet(workbook, records, profile: BranchProfile) -> None:
    if "SoLieuTongHop" in workbook.sheetnames:
        del workbook["SoLieuTongHop"]
    sheet = workbook.create_sheet("TongHop_Mau15b")
    fallback_branch = profile.branch_code.strip()
    branches = tuple(
        OrderedDict.fromkeys(
            _branch_from_customer_id(record.customer_id, fallback_branch)
            for record in records
        )
    )
    cash_keys = tuple(
        sorted(
            {
                f"{record.account}_{str(record.currency or '').strip() or 'VND'}"
                for record in records
            }
        )
    )
    accrual_accounts = tuple(
        sorted(
            {
                str(record.interest_account).strip()
                for record in records
                if str(record.interest_account).strip() in {"941001", "941002", "941003"}
            }
        )
    )

    def branch_of(record) -> str:
        return _branch_from_customer_id(record.customer_id, fallback_branch)

    def cash_key_of(record) -> str:
        return f"{record.account}_{str(record.currency or '').strip() or 'VND'}"

    def cash_value(branch: str, key: str) -> Decimal:
        return sum(
            (
                record.converted_balance
                for record in records
                if branch_of(record) == branch and cash_key_of(record) == key
            ),
            Decimal(0),
        )

    def accrual_value(branch: str, account: str) -> Decimal:
        return sum(
            (
                record.accrued_interest
                for record in records
                if branch_of(record) == branch
                and str(record.interest_account).strip() == account
            ),
            Decimal(0),
        )

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    outer_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    header_fill = PatternFill("solid", fgColor="FFFF00")

    def write_table(start_row: int, title: str, headers: tuple[str, ...], rows: list[list[object]]) -> int:
        sheet.cell(start_row, 1, title)
        sheet.cell(start_row, 1).font = Font(name="Times New Roman", size=12, bold=True, color="000080")
        header_row = start_row + 1
        for column, value in enumerate(headers, 1):
            cell = sheet.cell(header_row, column, value)
            cell.font = Font(name="Times New Roman", size=11, bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_offset, values in enumerate(rows, header_row + 1):
            for column, value in enumerate(values, 1):
                sheet.cell(row_offset, column, value)
        end_row = header_row + len(rows)
        end_column = len(headers)
        for row in range(header_row, end_row + 1):
            for column in range(1, end_column + 1):
                cell = sheet.cell(row, column)
                updated_font = copy(cell.font)
                updated_font.name = "Times New Roman"
                updated_font.sz = 11
                cell.font = updated_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if column == 1 or row == header_row else "right",
                    vertical="center",
                )
                if row > header_row and column > 1:
                    cell.number_format = "#,##0"
        for column in range(1, end_column + 1):
            sheet.cell(header_row, column).border = outer_border
            sheet.cell(end_row, column).border = outer_border
        return end_row

    cash_headers = ("Mã Chi nhánh", *cash_keys, "Cộng chi nhánh")
    cash_rows: list[list[object]] = []
    for branch in branches:
        values = [cash_value(branch, key) for key in cash_keys]
        cash_rows.append([branch, *values, sum(values, Decimal(0))])
    cash_total = [
        "Tổng cộng",
        *[
            sum((row[column] for row in cash_rows), Decimal(0))
            for column in range(1, len(cash_keys) + 1)
        ],
    ]
    cash_total.append(sum((row[-1] for row in cash_rows), Decimal(0)))
    cash_rows.append(cash_total)
    cash_end = write_table(
        1,
        "Báo cáo tổng hợp số liệu quyết toán mẫu 15b",
        cash_headers,
        cash_rows,
    )

    accrual_headers = ("Mã Chi nhánh", *accrual_accounts, "Cộng chi nhánh")
    accrual_rows: list[list[object]] = []
    for branch in branches:
        values = [accrual_value(branch, account) for account in accrual_accounts]
        accrual_rows.append([branch, *values, sum(values, Decimal(0))])
    accrual_total = [
        "Tổng cộng",
        *[
            sum((row[column] for row in accrual_rows), Decimal(0))
            for column in range(1, len(accrual_accounts) + 1)
        ],
    ]
    accrual_total.append(sum((row[-1] for row in accrual_rows), Decimal(0)))
    accrual_rows.append(accrual_total)
    write_table(
        cash_end + 3,
        "Số liệu phải thu đến 31/12 mẫu 15b/QT",
        accrual_headers,
        accrual_rows,
    )

    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18 if column > 1 else 14
    sheet.freeze_panes = "B3"
    sheet.page_setup.paperSize = 9
    sheet.page_setup.orientation = "landscape"
    sheet.print_options.horizontalCentered = True


def main() -> int:
    if len(sys.argv) != 2:
        print(
            "Usage: python -m agribank_v3.settlement.consolidation15b_worker <request.json>",
            file=sys.stderr,
        )
        return 2
    request_path = Path(sys.argv[1])
    payload = json.loads(request_path.read_text(encoding="utf-8"))
    source_paths = tuple(Path(path) for path in payload["source_paths"])
    output_path = Path(payload["output_path"])
    merged_csv_path = Path(payload["merged_csv_path"])
    profile = BranchProfile(**payload["profile"])
    options = SettlementOptions(**payload["options"])
    spec = SETTLEMENT_SPECS["consolidation.15b"]
    started = time.perf_counter()
    merge_result = None
    records = []
    try:
        _log(f"[TongHop15b] start: {len(source_paths)} source files -> {output_path}")
        step = time.perf_counter()
        merge_result = merge_same_structure_csv_to_csv(source_paths, merged_csv_path)
        _log(
            f"[TongHop15b] merge done in {time.perf_counter() - step:.1f}s, "
            f"rows={merge_result.row_count}"
        )
        processor = Mau1516Processor()
        request = SettlementRequest(
            spec=spec,
            profile=profile,
            options=options,
            source_paths=(merged_csv_path,),
        )
        step = time.perf_counter()
        records, report_date = processor.read_source(request, merged_csv_path)
        _log(
            f"[TongHop15b] read/sort done in {time.perf_counter() - step:.1f}s, "
            f"records={len(records)}"
        )
        step = time.perf_counter()
        workbook = processor.build_workbook(request, records, report_date)
        workbook.active.title = "SoLieu_Mau15b"
        _add_summary_sheet(workbook, records, profile)
        try:
            workbook.save(output_path)
        finally:
            workbook.close()
        _log(
            f"[TongHop15b] save done in {time.perf_counter() - step:.1f}s; "
            f"total={time.perf_counter() - started:.1f}s"
        )
        _log(
            "[TongHop15b] completed: "
            f"output={output_path}; "
            f"sources={merge_result.source_count if merge_result else 0}; "
            f"merged_rows={merge_result.row_count if merge_result else 0}; "
            f"processed_rows={len(records)}"
        )
        return 0
    finally:
        if merged_csv_path.exists():
            try:
                merged_csv_path.unlink()
            except OSError:
                pass


if __name__ == "__main__":
    raise SystemExit(main())
