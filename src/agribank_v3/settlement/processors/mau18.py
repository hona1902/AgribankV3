from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import (
    SettlementOptions,
    SettlementRequest,
    SettlementResult,
)
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)
from agribank_v3.settlement.transforms import (
    normalize_customer_id,
    parse_yyyymmdd,
    vietnamese_report_date,
)


@dataclass(frozen=True, slots=True)
class GuaranteeRecord:
    account: str
    currency: str
    customer_id: str
    customer_name: str
    contract_number: str
    issue_date: object
    expiry_date: object
    guarantee_type: str
    original_balance: Decimal
    converted_balance: Decimal
    deposit_balance: Decimal
    approval_number: object
    approval_amount: object
    source_values: tuple[object, ...] = ()


class Mau18Processor:
    """Processor for Mẫu 18/QT: bảo lãnh và thư tín dụng."""

    REQUIRED_HEADERS = {
        "NGAY_GIAO_DICH",
        "MA_CN",
        "PGD",
        "TAI_KHOAN",
        "LOAI_TIEN_LC_BL_CK",
        "LOAI_KH",
        "MA_KH",
        "TEN_KH",
        "SO_LC_BL_CK",
        "NGAY_PHAT_HANH",
        "NGAY_DEN_HAN",
        "LOAI_LC_BL_CK",
        "SO_DU_LC_BL_CK",
        "SO_DU_LC_BL_CK_QUY_DOI",
        "LOAI_TIEN_KQ_CK",
        "SO_DU_KQ_CK",
        "SO_PHE_DUYET",
        "SOTIEN_PHE_DUYET",
        "NGUOI_TAO",
    }

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError("Mẫu 18/QT cần đúng một file CSV nguồn.")
        source_path = request.source_paths[0]
        records, report_date = self.read_source(request, source_path)
        workbook = self.build_workbook(request, records, report_date)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}{self._output_prefix(request.options)}18.xlsx"
        )
        workbook.save(output_path)
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name="18",
            processed_rows=len(records),
        )

    def read_source(
        self,
        request: SettlementRequest,
        source_path: Path,
    ) -> tuple[list[GuaranteeRecord], str]:
        if source_path.suffix.casefold() not in {".csv", ".txt"}:
            raise SettlementError("Nguồn Mẫu 18/QT phải là file CSV.")
        try:
            with source_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                source_headers = tuple(reader.fieldnames or ())
                rows = list(reader)
        except (OSError, UnicodeError, csv.Error) as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc

        missing = sorted(self.REQUIRED_HEADERS - set(source_headers))
        if missing:
            raise SettlementError(
                f"File nguồn thiếu cột: {', '.join(missing)}",
                code="invalid_mau18_headers",
            )
        self._source_headers = source_headers
        self._control_group_order = []
        seen_groups: set[str] = set()
        for row in rows:
            key = f"{row['TAI_KHOAN'].strip()}_{row['LOAI_TIEN_LC_BL_CK'].strip() or 'VND'}"
            if key not in seen_groups:
                seen_groups.add(key)
                self._control_group_order.append(key)
        records = [
            self._row_to_record(row, source_headers, request.options)
            for row in rows
        ]
        records.sort(
            key=lambda item: (
                self._sort_number(item.account),
                item.customer_id,
                item.issue_date or "",
            )
        )
        report_date = vietnamese_report_date(rows[0]["NGAY_GIAO_DICH"]) if rows else ""
        return records, report_date

    def build_workbook(
        self,
        request: SettlementRequest,
        records: list[GuaranteeRecord],
        report_date: str,
    ) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "18"
        self._write_report_header(sheet, request, report_date)
        final_row = self._write_detail_rows(
            sheet,
            records,
            keep_unused_columns=not request.options.remove_unused_columns,
        )
        self._write_signatures_and_notes(sheet, request, report_date, final_row)
        if not request.options.remove_unused_columns:
            self._write_unused_source_headers(sheet)
        self._format_report(sheet, final_row, request.options)
        if request.options.create_control_sheet:
            self._write_control_sheet(workbook, records)
        return workbook

    def _row_to_record(
        self,
        row: dict[str, str],
        source_headers: tuple[str, ...],
        options: SettlementOptions,
    ) -> GuaranteeRecord:
        return GuaranteeRecord(
            account=row["TAI_KHOAN"].strip(),
            currency=row["LOAI_TIEN_LC_BL_CK"].strip(),
            customer_id=normalize_customer_id(
                row["MA_KH"],
                row["MA_CN"],
                include_branch=options.include_branch_in_customer_id,
            ),
            customer_name=row["TEN_KH"].strip(),
            contract_number=row["SO_LC_BL_CK"].strip().lstrip("'").strip(),
            issue_date=parse_yyyymmdd(row["NGAY_PHAT_HANH"]),
            expiry_date=parse_yyyymmdd(row["NGAY_DEN_HAN"]),
            guarantee_type=row["LOAI_LC_BL_CK"].strip(),
            original_balance=self._parse_amount(row["SO_DU_LC_BL_CK"]),
            converted_balance=self._parse_amount(row["SO_DU_LC_BL_CK_QUY_DOI"]),
            deposit_balance=self._parse_amount(row["SO_DU_KQ_CK"]),
            approval_number=self._blank_or_text(row["SO_PHE_DUYET"]),
            approval_amount=self._parse_optional_amount(row["SOTIEN_PHE_DUYET"]),
            source_values=tuple(row.get(header, "") for header in source_headers),
        )

    def _write_detail_rows(
        self,
        sheet,
        records: list[GuaranteeRecord],
        *,
        keep_unused_columns: bool,
    ) -> int:
        if not records:
            return 11
        row = 12
        group_start = row
        current_account = records[0].account
        subtotal_rows: list[int] = []
        for record in records:
            if record.account != current_account:
                self._write_subtotal(sheet, row, current_account, group_start, row - 1)
                subtotal_rows.append(row)
                row += 1
                group_start = row
                current_account = record.account
            self._write_record(
                sheet,
                row,
                record,
                keep_unused_columns=keep_unused_columns,
            )
            row += 1
        self._write_subtotal(sheet, row, current_account, group_start, row - 1)
        subtotal_rows.append(row)
        row += 1
        self._write_grand_total(sheet, row, subtotal_rows)
        return row

    def _write_record(
        self,
        sheet,
        row: int,
        record: GuaranteeRecord,
        *,
        keep_unused_columns: bool,
    ) -> None:
        values = (
            self._numeric_code(record.account),
            record.currency,
            record.customer_id,
            record.customer_name,
            record.contract_number,
            record.issue_date,
            record.expiry_date,
            record.guarantee_type,
            record.original_balance,
            record.converted_balance,
            record.deposit_balance if record.deposit_balance else None,
            record.approval_number,
            record.approval_amount,
        )
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
        if keep_unused_columns:
            for offset, value in enumerate(record.source_values, 14):
                sheet.cell(row, offset, value)

    @staticmethod
    def _write_subtotal(sheet, row: int, account: str, start: int, end: int) -> None:
        sheet.cell(row, 1, f"Cộng TK: {account.strip()}")
        for column in (9, 10, 11, 13):
            letter = get_column_letter(column)
            sheet.cell(row, column, f"=SUM({letter}{start}:{letter}{end})")

    @staticmethod
    def _write_grand_total(sheet, row: int, subtotal_rows: list[int]) -> None:
        sheet.cell(row, 1, "TỔNG CỘNG:")
        for column in (9, 10, 11, 13):
            letter = get_column_letter(column)
            refs = "+".join(f"{letter}{subtotal_row}" for subtotal_row in subtotal_rows)
            sheet.cell(row, column, f"={refs}" if refs else 0)

    def _write_signatures_and_notes(
        self,
        sheet,
        request: SettlementRequest,
        report_date: str,
        final_row: int,
    ) -> None:
        base = final_row + 1
        location_date = f"{request.profile.report_location}, {report_date}" if report_date else ""
        sheet.cell(base, 11, location_date)
        sheet.cell(base + 1, 1, "LẬP BIỂU")
        sheet.cell(base + 2, 1, "(Ký, ghi rõ họ tên, số ĐT liên hệ)")
        sheet.cell(
            base + 9,
            1,
            f"({request.profile.report_preparer} - SĐT: {request.profile.phone})",
        )
        sheet.cell(base + 1, 5, "TRƯỞNG PHÒNG KHCN/KHDN")
        sheet.cell(base + 2, 5, "(Ký, ghi rõ họ tên)")
        sheet.cell(base + 1, 9, "TRƯỞNG PHÒNG KẾ TOÁN")
        sheet.cell(base + 2, 9, "(Ký, ghi rõ họ tên)")
        sheet.cell(base + 1, 11, " GIÁM ĐỐC")
        sheet.cell(base + 2, 11, "(Ký, đóng dấu, ghi rõ họ tên)")
        for start, end, offset in (
            (1, 4, 1),
            (1, 4, 2),
            (1, 4, 9),
            (5, 8, 1),
            (5, 8, 2),
            (9, 10, 1),
            (9, 10, 2),
            (11, 13, 0),
            (11, 13, 1),
            (11, 13, 2),
        ):
            self._merge_if_unmerged(sheet, base + offset, start, base + offset, end)
        notes_start = base + 11
        sheet.cell(notes_start, 1, "Ghi chú:")
        notes = (
            "1. Các chi nhánh báo cáo chi tiết cho từng tài khoản trong khoản mục bảo lãnh và thư tín dụng (Theo tài khoản cấp V)",
            '2. Tổng số dư của mỗi tài khoản (Số "Cộng TK") phải được đối chiếu khớp đúng với bảng cân đối tài khoản.',
            "3. Cột 6 và 7: Ghi theo thứ tự ngày/tháng/năm (không ghi tháng trước ngày sau)",
            "4. Cột 7: Đối với thư tín dụng, ngày hết hiệu lực là ngày hết hiệu lực xuất trình bộ chứng từ thanh toán.",
            "5. Cột 11: Yêu cầu chi nhánh thực hiện đối chiếu khớp đúng với TK 4272, 4274 trên cân đối nội bảng. Trường hợp có chênh lệch phải giải trình cụ thể nguyên nhân từng khoản chênh lệch và nêu rõ bút toán điều chỉnh (nếu có).",
        )
        for offset, note in enumerate(notes):
            sheet.cell(notes_start + offset, 2, note)

    def _write_report_header(
        self,
        sheet,
        request: SettlementRequest,
        report_date: str,
    ) -> None:
        profile = request.profile
        entries = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": "----------------------------------",
            "A4": f"Mã chi nhánh: {profile.branch_code}",
            "A5": f"Tên {profile.branch_name}",
            "A6": "SAO KÊ CHI TIẾT TÀI KHOẢN 92: BẢO LÃNH VÀ THƯ TÍN DỤNG CHO KHÁCH HÀNG",
            "A7": report_date,
            "M8": "ĐVT: VNĐ",
            "L1": "1. Mẫu số 18/QT",
            "L2": "2. CN loại I gửi file về TSC",
            "L3": "3. Lưu tại Chi nhánh",
        }
        for address, value in entries.items():
            sheet[address] = value
        headers = (
            "TÀI KHOẢN",
            "LOẠI TIỀN TỆ",
            "MÃ KHÁCH HÀNG",
            "TÊN KHÁCH HÀNG",
            "SỐ LC, HỢP ĐỒNG BẢO LÃNH",
            "NGÀY BẮT ĐẦU",
            "NGÀY HẾT HIỆU LỰC",
            "LOẠI BẢO LÃNH, LC",
            "SỐ DƯ CỦA LC, BẢO LÃNH TRÊN TÀI KHOẢN NGOẠI BẢNG NGÀY 31/12",
            None,
            "SỐ TIỀN KÝ QUỸ TRÊN TK 427 (Quy đổi)",
            "HẠN MỨC TÍN DỤNG CHO KHÁCH HÀNG (nếu có)",
            None,
        )
        subheaders = (
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "Nguyên tệ",
            "Quy đổi VNĐ",
            None,
            "Số hợp đồng hạn mức",
            "Số tiền (Quy đổi)",
        )
        for column, value in enumerate(headers, 1):
            sheet.cell(9, column, value)
        for column, value in enumerate(subheaders, 1):
            sheet.cell(10, column, value)
        for column in range(1, 14):
            sheet.cell(11, column, column)
        for row in range(1, 6):
            self._merge_if_unmerged(sheet, row, 1, row, 4)
        self._merge_if_unmerged(sheet, 6, 1, 6, 13)
        self._merge_if_unmerged(sheet, 7, 1, 7, 13)
        for merged in (
            (9, 1, 10, 1),
            (9, 2, 10, 2),
            (9, 3, 10, 3),
            (9, 4, 10, 4),
            (9, 5, 10, 5),
            (9, 6, 10, 6),
            (9, 7, 10, 7),
            (9, 8, 10, 8),
            (9, 9, 9, 10),
            (9, 11, 10, 11),
            (9, 12, 9, 13),
        ):
            self._merge_if_unmerged(sheet, *merged)

    def _write_unused_source_headers(self, sheet) -> None:
        for offset, header in enumerate(getattr(self, "_source_headers", ()), 14):
            sheet.cell(11, offset, header)

    def _write_control_sheet(
        self,
        workbook: Workbook,
        records: list[GuaranteeRecord],
    ) -> None:
        sheet = workbook.create_sheet("SoLieuTongHop")
        sheet.append(("SỐ LIỆU TỔNG HỢP MẪU 18",))
        sheet.append(("Tài Khoản", "Số Dư Nguyên Tệ", "Số Dư Qui Đổi VNĐ", "Số tiền ký quỹ"))
        groups: dict[str, list[GuaranteeRecord]] = {}
        for record in records:
            key = f"{record.account}_{record.currency or 'VND'}"
            groups.setdefault(key, []).append(record)
        first_row = 3
        ordered_keys = [
            key for key in getattr(self, "_control_group_order", ()) if key in groups
        ]
        ordered_keys.extend(key for key in groups if key not in ordered_keys)
        for key in ordered_keys:
            selected = groups[key]
            sheet.append(
                (
                    key,
                    sum((item.original_balance for item in selected), Decimal(0)),
                    sum((item.converted_balance for item in selected), Decimal(0)),
                    sum((item.deposit_balance for item in selected), Decimal(0)),
                )
            )
        total_row = sheet.max_row + 1
        sheet.cell(total_row, 1, "Tổng cộng:")
        for column in range(2, 5):
            letter = get_column_letter(column)
            sheet.cell(total_row, column, f"=SUM({letter}{first_row}:{letter}{total_row - 1})")
        self._format_control_sheet(sheet, total_row)

    @staticmethod
    def _format_control_sheet(sheet, total_row: int) -> None:
        thin = Side(style="thin", color="000000")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=12)
                cell.alignment = Alignment(vertical="center")
        sheet["A1"].font = Font(name="Times New Roman", size=12, bold=True, color="000080")
        for cell in sheet[2]:
            cell.fill = PatternFill("solid", fgColor="FFFF00")
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(3, total_row + 1):
            for column in range(2, 5):
                sheet.cell(row, column).number_format = "#,##0"
        widths = {"A": 18, "B": 20, "C": 20, "D": 20}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

    def _format_report(
        self,
        sheet,
        last_detail_row: int,
        options: SettlementOptions,
    ) -> None:
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=10)
                cell.alignment = Alignment(vertical="center")
        style_agency_header(sheet, start_column=1, end_column=4)
        style_currency_unit(sheet["M8"])
        sheet["A6"].font = Font(name="Times New Roman", size=16, bold=True)
        sheet["A7"].font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        for address in ("A1", "A2", "A3", "A4", "A5", "A6", "A7"):
            sheet[address].alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=9, max_row=last_detail_row, min_col=1, max_col=13):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(9, 12):
            for column in range(1, 14):
                cell = sheet.cell(row, column)
                cell.font = Font(name="Times New Roman", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(12, last_detail_row + 1):
            for column in (9, 10, 11, 13):
                sheet.cell(row, column).number_format = "#,##0"
            for column in (6, 7):
                sheet.cell(row, column).number_format = (
                    "dd/mm/yyyy" if options.four_digit_year else "dd/mm/yy"
                )
            for column in (2, 3, 5, 6, 7, 8, 12):
                sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
            if self._is_total_row(sheet, row):
                for column in range(1, 14):
                    sheet.cell(row, column).font = Font(name="Times New Roman", size=10, bold=True)
                if str(sheet.cell(row, 1).value or "").startswith("TỔNG CỘNG:"):
                    for column in range(1, 14):
                        sheet.cell(row, column).border = Border(
                            left=medium, right=medium, top=medium, bottom=medium
                        )
        for row in range(last_detail_row + 1, sheet.max_row + 1):
            for column in range(1, 14):
                sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
        for row in (last_detail_row + 2,):
            for column in range(1, 14):
                sheet.cell(row, column).font = Font(name="Times New Roman", bold=True)
        for row in (last_detail_row + 1, last_detail_row + 3, last_detail_row + 10):
            for column in range(1, 14):
                sheet.cell(row, column).font = Font(name="Times New Roman", italic=True)
        for row in range(last_detail_row + 12, sheet.max_row + 1):
            sheet.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center")
        widths = {
            1: 7,
            2: 5.8,
            3: 12.5 if options.include_branch_in_customer_id else 9.6,
            4: 45,
            5: 9.4,
            6: 9.4 if options.four_digit_year else 7.3,
            7: 9.4 if options.four_digit_year else 7.3,
            8: 5.6,
            9: 20,
            10: 20,
            11: 20,
            12: 20,
            13: 20,
        }
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = width
        if not options.remove_unused_columns:
            for column in range(14, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(column)].width = 15
        sheet.row_dimensions[5].height = 17
        sheet.row_dimensions[9].height = 43
        sheet.row_dimensions[10].height = 20
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "B12"
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:M{sheet.max_row}",
            title_rows="9:10",
        )

    @staticmethod
    def _is_total_row(sheet, row: int) -> bool:
        return str(sheet.cell(row, 1).value or "").startswith(("Cộng TK:", "TỔNG CỘNG:"))

    @staticmethod
    def _blank_or_text(value: Any) -> object:
        text = str(value or "").strip()
        return text.lstrip("'") if text else None

    @staticmethod
    def _numeric_code(value: Any) -> object:
        text = str(value or "").strip()
        return int(text) if text.isdigit() else value

    @staticmethod
    def _parse_amount(value: Any) -> Decimal:
        if value is None or isinstance(value, bool):
            return Decimal(0)
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip().replace(" ", "").replace("\u00a0", "")
        if not text or text == "-":
            return Decimal(0)
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return Decimal(text)
        except Exception:
            return Decimal(0)

    @classmethod
    def _parse_optional_amount(cls, value: Any) -> object:
        text = str(value or "").strip()
        return cls._parse_amount(value) if text else None

    @staticmethod
    def _sort_number(value: Any) -> tuple[int, str]:
        text = str(value or "").strip()
        return (int(text), text) if text.isdigit() else (999999, text)

    @staticmethod
    def _merge_if_unmerged(sheet, start_row: int, start_column: int, end_row: int, end_column: int) -> None:
        cell_range = (
            f"{get_column_letter(start_column)}{start_row}:"
            f"{get_column_letter(end_column)}{end_row}"
        )
        if cell_range not in {str(merged) for merged in sheet.merged_cells.ranges}:
            sheet.merge_cells(cell_range)

    @staticmethod
    def _output_prefix(options: SettlementOptions) -> str:
        prefix = (options.output_prefix or "QT").strip().upper()
        return "BN" if prefix == "BN" else "QT"
