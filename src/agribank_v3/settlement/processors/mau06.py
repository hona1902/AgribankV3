from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementRequest, SettlementResult
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)


class Mau06Processor:
    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError("Mẫu 06/QT cần đúng một file Mẫu 05/QT.")
        source_path = request.source_paths[0]
        workbook = load_workbook(source_path, data_only=False)
        if "05" not in workbook.sheetnames:
            raise SettlementError("File nguồn không có sheet 05.")
        values = load_workbook(source_path, data_only=True)["05"]
        output = self.build_workbook(request, values)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}QT06.xlsx"
        )
        output.save(output_path)
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name="06",
            processed_rows=12,
        )

    def build_workbook(self, request: SettlementRequest, source) -> Workbook:
        direct_start, direct_total, guarantee_start, guarantee_total = self._sections(source)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "06"
        report_date = str(source["A6"].value or "")
        self._write_header(sheet, request, report_date)
        categories = ((6, "Tài sản thế chấp bằng bất động sản"),
                      (7, "Tài sản thế chấp bằng động sản"),
                      (8, "Tài sản thế chấp bằng giấy tờ có giá"),
                      (9, "Tài sản thế chấp khác"))
        self._write_section(sheet, source, 12, direct_start, direct_total, categories)
        self._write_section(sheet, source, 18, guarantee_start, guarantee_total, categories)
        sheet["A12"] = "I. TSTC của khách hàng trực tiếp vay vốn"
        sheet["A18"] = "II. TSTC của đơn vị bảo lãnh cho bên thứ ba vay vốn"
        for offset, (_, label) in enumerate(categories, 1):
            sheet.cell(12 + offset, 1, label)
            sheet.cell(18 + offset, 1, label)
        sheet["A17"] = "Cộng I"
        sheet["A23"] = "Cộng II"
        sheet["A24"] = "TỔNG CỘNG (I + II)"
        for column in range(2, 12):
            letter = get_column_letter(column)
            sheet.cell(17, column, f"=SUM({letter}13:{letter}16)")
            sheet.cell(23, column, f"=SUM({letter}19:{letter}22)")
            sheet.cell(24, column, f"={letter}17+{letter}23")
        self._write_footer(sheet, request, report_date)
        self._format(sheet)
        return workbook

    @staticmethod
    def _sections(source) -> tuple[int, int, int, int]:
        direct_heading = direct_total = guarantee_heading = guarantee_total = 0
        for row in range(1, source.max_row + 1):
            value = str(source.cell(row, 1).value or "")
            if value.startswith("III. TSTC"):
                guarantee_heading = row
            elif value.startswith("II. TSTC"):
                direct_heading = row
            elif value.startswith("Cộng III"):
                guarantee_total = row
            elif value.startswith("Cộng II"):
                direct_total = row
        if not all((direct_heading, direct_total, guarantee_heading, guarantee_total)):
            raise SettlementError("Không xác định được các phần II/III trong Mẫu 05.")
        return direct_heading, direct_total, guarantee_heading + 1, guarantee_total

    @staticmethod
    def _write_section(sheet, source, target_heading, start, total, categories) -> None:
        for offset, (asset_column, _) in enumerate(categories, 1):
            target_row = target_heading + offset
            amount_total = sum(
                source.cell(row, asset_column).value
                for row in range(start, total)
                if isinstance(source.cell(row, asset_column).value, (int, float))
            )
            sheet.cell(target_row, 2, amount_total)
            for target_column, flag_column in enumerate(range(10, 19), 3):
                count = 0
                for row in range(start, total):
                    amount = source.cell(row, asset_column).value
                    flag = str(source.cell(row, flag_column).value or "").upper()
                    if isinstance(amount, (int, float)) and amount > 0 and flag == "X":
                        count += 1
                sheet.cell(target_row, target_column, count)

    @staticmethod
    def _write_header(sheet, request, report_date) -> None:
        profile = request.profile
        entries = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": f"Mã chi nhánh: {profile.branch_code}",
            "A4": f"Tên {profile.branch_name}",
            "A5": "BÁO CÁO TỔNG HỢP KIỂM KÊ HỒ SƠ, TÀI SẢN THẾ CHẤP, CẦM CỐ CỦA KHÁCH HÀNG",
            "A6": report_date,
            "K7": "Đơn vị: VNĐ",
            "I1": "1. Mẫu số 06/QT",
            "I2": "2. CN loại I gửi báo cáo giấy về TSC",
            "I3": "3. Lưu tại Chi nhánh",
            "A8": "TÊN CHỈ TIÊU",
            "B8": "GIÁ TRỊ TSBĐ (VNĐ)",
            "C8": "HIỆN TRẠNG TSĐB (SỐ LƯỢNG)",
            "C9": "Tính pháp lý",
            "E9": "Khả năng hoàn thiện hồ sơ tài sản",
            "I9": "Khả năng phát mại",
        }
        for address, value in entries.items():
            sheet[address] = value
        for address in ("A1:B1", "A2:B2", "A3:B3", "A4:B4", "A5:K5",
                        "A6:K6", "A8:A10", "B8:B10", "C8:K8",
                        "C9:D9", "E9:H9", "I9:K9"):
            sheet.merge_cells(address)
        labels = ("Hợp pháp, hợp lệ", "Khác", "Đầy đủ hồ sơ theo quy định",
                  "Đang hoàn chỉnh thủ tục", "Không thể hoàn chỉnh được", "Khác",
                  "Có khả năng phát mại", "Ít có khả năng phát mại",
                  "Không có khả năng phát mại")
        for column, label in enumerate(labels, 3):
            sheet.cell(10, column, label)
        for column in range(1, 12):
            sheet.cell(11, column, column)

    @staticmethod
    def _write_footer(sheet, request, report_date) -> None:
        profile = request.profile
        sheet["H25"] = f"{profile.report_location}, {report_date}"
        sheet.merge_cells("H25:K25")
        for address, value in {
            "A26": "LẬP BIỂU",
            "B26": "TRƯỞNG PHÒNG KHDN/KHCN",
            "E26": "TRƯỞNG PHÒNG KẾ TOÁN",
            "H26": "GIÁM ĐỐC",
            "A27": "(Ký, ghi rõ họ tên, số ĐT liên hệ)",
            "B27": "(Ký, ghi rõ họ tên)",
            "E27": "(Ký, ghi rõ họ tên)",
            "H27": "(Ký, đóng dấu, ghi rõ họ tên)",
            "A33": f"({profile.report_preparer} - SĐT: {profile.phone})",
            "A37": "Ghi chú:",
            "A38": "1. Cột 2: Dòng Tổng cộng khớp với số dư trên cân đối của tài khoản 994, 995, 996, 992003, 992005.",
            "A39": "2. Từ cột 3 đến cột 11 được phản ánh theo số lượng tổng hợp từ cột 10 đến cột 18 của mẫu số 05/QT (lưu ý trường hợp 01 TSBĐ được dùng để bảo đảm cho khoản vay của nhiều hơn 02 khách hàng thì chỉ tính số lượng là 01)",
        }.items():
            sheet[address] = value
        for address in ("B26:D26", "E26:G26", "H26:K26",
                        "B27:D27", "E27:G27", "H27:K27", "A39:K40"):
            sheet.merge_cells(address)

    @staticmethod
    def _format(sheet) -> None:
        thin = Side(style="thin", color="000000")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=10)
                cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(min_row=8, max_row=24, min_col=1, max_col=11):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(8, 12):
            for cell in sheet[row]:
                cell.font = Font(name="Times New Roman", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in (12, 17, 18, 23, 24):
            for cell in sheet[row]:
                cell.font = Font(name="Times New Roman", bold=True)
        sheet["A5"].font = Font(name="Times New Roman", size=14, bold=True)
        sheet["A6"].font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        for address in ("A5", "A6"):
            sheet[address].alignment = Alignment(
                horizontal="center", vertical="center"
            )
        style_currency_unit(sheet["K7"])
        for row in range(11, 25):
            for column in range(2, 12):
                sheet.cell(row, column).number_format = "#,###"
        sheet["H25"].font = Font(name="Times New Roman", italic=True)
        sheet["H25"].alignment = Alignment(horizontal="center", vertical="center")
        for address in ("A26", "B26", "E26", "H26"):
            sheet[address].font = Font(name="Times New Roman", bold=True)
            sheet[address].alignment = Alignment(
                horizontal="center", vertical="center"
            )
        for address in ("A27", "B27", "E27", "H27"):
            sheet[address].font = Font(name="Times New Roman", italic=True)
            sheet[address].alignment = Alignment(
                horizontal="center", vertical="center"
            )
        for merged in sheet.merged_cells.ranges:
            if merged.min_row >= 25:
                sheet.cell(merged.min_row, merged.min_col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["B"].width = 23
        for column in range(3, 11):
            sheet.column_dimensions[get_column_letter(column)].width = 9.4
        sheet.column_dimensions["K"].width = 13
        style_agency_header(sheet, start_column=1, end_column=2)
        sheet.row_dimensions[10].height = 60
        sheet.sheet_view.showGridLines = False
        setup_a4_print_layout(sheet, print_area="A1:K40")
