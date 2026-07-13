from __future__ import annotations

from collections import OrderedDict
from datetime import date
from decimal import Decimal
from io import StringIO
from pathlib import Path
from typing import Any, TypeAlias

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementRequest, SettlementResult
from agribank_v3.settlement.processors.formatting import setup_a4_print_layout


Mau30SummaryRow: TypeAlias = tuple[str, Decimal] | tuple[str, str, Decimal]
Mau30BalanceMap: TypeAlias = dict[str, Decimal] | dict[tuple[str, str], Decimal]


class Mau30Processor:
    """Create Mẫu 30/QT from a generated settlement workbook."""

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) not in {1, 2}:
            raise SettlementError("Mẫu 30/QT cần một file quyết toán nguồn và tùy chọn một file cân đối.")
        source_path = request.source_paths[0]
        balance_path = request.source_paths[1] if len(request.source_paths) == 2 else None
        suffix = source_path.suffix.casefold()
        if suffix not in {".xlsx", ".xlsm"}:
            raise SettlementError("File quyết toán nguồn phải là .xlsx hoặc .xlsm.")

        model, processed_rows = self._build_and_save(
            request,
            source_path,
            source_path,
            balance_path,
        )

        return SettlementResult(
            spec_key=request.spec.key,
            output_path=source_path,
            workbook_name=source_path.name,
            worksheet_name=f"Mau30QT-{model}",
            processed_rows=processed_rows,
        )

    def _build_and_save(
        self,
        request: SettlementRequest,
        source_path: Path,
        working_path: Path,
        balance_path: Path | None,
    ) -> tuple[str, int]:
        workbook = load_workbook(working_path, data_only=False)
        values = load_workbook(working_path, data_only=True)
        try:
            model = self._model_from_request(request, source_path, values)
            balance = self._read_balance(balance_path) if balance_path else {}
            rows = self._summary_rows(values, model, request)
            self.build_sheet(request, workbook, values, model, balance)
            workbook.save(working_path)
            return model, len(rows)
        finally:
            values.close()
            workbook.close()

    def build_sheet(
        self,
        request,
        workbook,
        values_workbook,
        model: str,
        balance: Mau30BalanceMap | None = None,
    ):
        if (
            model.casefold() not in {"20a", "22", "23"}
            and "SoLieuTongHop" not in values_workbook.sheetnames
            and not self._consolidation_sheet_name(values_workbook, model)
        ):
            raise SettlementError("File nguồn không có sheet SoLieuTongHop.")
        source_sheet_name = self._sheet_name(values_workbook, model)
        if not source_sheet_name:
            raise SettlementError(f"File nguồn không có sheet {model}.")

        sheet_name = f"Mau30QT-{model}"
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        sheet = workbook.create_sheet(sheet_name, 0)
        rows = self._summary_rows(values_workbook, model, request)
        final_row = 11 + len(rows)
        self._write_header(sheet, request, values_workbook[source_sheet_name], model)
        self._write_rows(sheet, request, rows, balance or {})
        self._write_footer(sheet, request, final_row)
        self._format(sheet, final_row)
        return sheet

    def _summary_rows(
        self,
        values_workbook,
        model: str,
        request: SettlementRequest,
    ) -> list[Mau30SummaryRow]:
        consolidation_sheet_name = self._consolidation_sheet_name(values_workbook, model)
        if request.spec.key == "consolidation.30a" and consolidation_sheet_name:
            return self._summary_rows_from_consolidation_sheet(
                values_workbook[consolidation_sheet_name],
                model=model,
                include_accrual_accounts=request.options.include_accrual_accounts,
            )

        if model.casefold() == "20a":
            source_sheet_name = self._sheet_name(values_workbook, model)
            if not source_sheet_name:
                raise SettlementError("File nguồn không có sheet 20a.")
            source = values_workbook[source_sheet_name]
            grouped: OrderedDict[str, Decimal] = OrderedDict()
            for row in range(12, source.max_row + 1):
                row_label = self._text(source.cell(row, 1).value).casefold()
                if row_label.startswith("tổng cộng"):
                    break
                for account_column, amount_column in ((9, 10), (11, 12)):
                    account = self._text(source.cell(row, account_column).value)
                    if not account:
                        continue
                    grouped[account] = grouped.get(account, Decimal(0)) + self._number(
                        source.cell(row, amount_column).value
                    )
            return list(grouped.items())

        if model == "22":
            source_sheet_name = self._sheet_name(values_workbook, model)
            if not source_sheet_name:
                raise SettlementError("File nguồn không có sheet 22.")
            source = values_workbook[source_sheet_name]
            grouped: OrderedDict[str, Decimal] = OrderedDict()
            for row in range(12, source.max_row + 1):
                label = self._text(source.cell(row, 1).value)
                if label.casefold().startswith("cộng"):
                    break
                account = self._text(source.cell(row, 3).value)
                if not account or not account[:1].isdigit():
                    continue
                grouped[account] = grouped.get(account, Decimal(0)) + self._number(
                    source.cell(row, 13).value
                )
            return list(grouped.items())

        if model == "23":
            source_sheet_name = self._sheet_name(values_workbook, model)
            if not source_sheet_name:
                raise SettlementError("File nguồn không có sheet 23.")
            source = values_workbook[source_sheet_name]
            rows: list[Mau30SummaryRow] = []
            current_account = ""
            running_total = Decimal(0)
            for row in range(12, source.max_row + 1):
                label = self._text(source.cell(row, 1).value)
                if label.casefold().startswith("cộng tk"):
                    account = "".join(ch for ch in label if ch.isdigit())
                    if account:
                        cached_total = self._number(source.cell(row, 4).value)
                        rows.append((account, cached_total if cached_total else running_total))
                    current_account = ""
                    running_total = Decimal(0)
                    continue
                if label and label.isdigit():
                    current_account = label
                    running_total = self._number(source.cell(row, 4).value)
                    continue
                if current_account:
                    running_total += self._number(source.cell(row, 4).value)
            return rows

        source = values_workbook["SoLieuTongHop"]
        if model == "05":
            rows = []
            for row in range(3, source.max_row + 1):
                account = self._text(source.cell(row, 1).value)
                if not account or account.casefold().startswith("tổng"):
                    continue
                rows.append((account, self._number(source.cell(row, 3).value)))
            return rows

        grouped: OrderedDict[str, Decimal] = OrderedDict()
        for row in range(3, source.max_row + 1):
            account = self._text(source.cell(row, 1).value)
            if not account or account.casefold().startswith("tổng"):
                continue
            key = account[:6]
            grouped[key] = grouped.get(key, Decimal(0)) + self._number(
                source.cell(row, 3).value
            )

        if model.casefold() in {"13", "15a", "15b"} and request.options.include_accrual_accounts:
            for column in range(6, source.max_column + 1):
                account = self._text(source.cell(2, column).value)
                if not account:
                    continue
                total = sum(
                    self._number(source.cell(row, column).value)
                    for row in range(3, source.max_row + 1)
                )
                grouped[account] = grouped.get(account, Decimal(0)) + total
        return list(grouped.items())

    def _summary_rows_from_consolidation_sheet(
        self,
        sheet,
        *,
        model: str,
        include_accrual_accounts: bool,
    ) -> list[Mau30SummaryRow]:
        rows: list[Mau30SummaryRow] = []
        table_starts = [
            row
            for row in range(1, sheet.max_row + 1)
            if "mã chi nhánh" in self._text(sheet.cell(row, 1).value).casefold()
        ]
        if model.casefold() == "15b":
            table_starts = [
                row
                for row in table_starts
                if "phải thu" in self._text(sheet.cell(row - 1, 1).value).casefold()
            ]
        if not include_accrual_accounts and table_starts:
            table_starts = table_starts[:1]
        for index, header_row in enumerate(table_starts):
            next_header = table_starts[index + 1] if index + 1 < len(table_starts) else sheet.max_row + 2
            first_data_row = header_row + 1
            if first_data_row >= next_header:
                continue
            if self._header_columns_are_branches(sheet, header_row):
                rows.extend(
                    self._summary_rows_from_branch_columns(
                        sheet,
                        header_row,
                        next_header - 1,
                    )
                )
            else:
                rows.extend(
                    self._summary_rows_from_branch_rows(
                        sheet,
                        header_row,
                        next_header - 1,
                    )
                )
        return self._group_branch_account_rows(rows)

    def _summary_rows_from_branch_rows(
        self,
        sheet,
        header_row: int,
        table_end_row: int,
    ) -> list[Mau30SummaryRow]:
        rows: list[Mau30SummaryRow] = []
        account_columns: list[tuple[int, str]] = []
        for column in range(2, sheet.max_column + 1):
            header = self._text(sheet.cell(header_row, column).value)
            if not header:
                break
            if self._is_branch_total_header(header):
                break
            account_columns.append((column, self._normalize_summary_account(header)))
        for row in range(header_row + 1, table_end_row + 1):
            branch = self._text(sheet.cell(row, 1).value)
            if not branch or not branch[:1].isdigit():
                break
            for column, account in account_columns:
                if not account:
                    continue
                rows.append((branch, account, self._number(sheet.cell(row, column).value)))
        return rows

    def _summary_rows_from_branch_columns(
        self,
        sheet,
        header_row: int,
        table_end_row: int,
    ) -> list[Mau30SummaryRow]:
        rows: list[Mau30SummaryRow] = []
        branch_columns: list[tuple[int, str]] = []
        for column in range(2, sheet.max_column + 1):
            header = self._text(sheet.cell(header_row, column).value)
            if not header:
                break
            if self._is_branch_total_header(header):
                break
            branch_columns.append((column, header))
        for row in range(header_row + 1, table_end_row + 1):
            account = self._normalize_summary_account(sheet.cell(row, 1).value)
            if not account or account.casefold().startswith("tổng"):
                break
            for column, branch in branch_columns:
                rows.append((branch, account, self._number(sheet.cell(row, column).value)))
        return self._order_branch_columns_by_branch(rows, branch_columns)

    @staticmethod
    def _group_branch_account_rows(rows: list[Mau30SummaryRow]) -> list[Mau30SummaryRow]:
        grouped: OrderedDict[tuple[str, str], Decimal] = OrderedDict()
        branch_order: list[str] = []
        for row in rows:
            if len(row) != 3:
                continue
            branch, account, amount = row
            if branch not in branch_order:
                branch_order.append(branch)
            key = (branch, account)
            grouped[key] = grouped.get(key, Decimal(0)) + amount
        ordered_rows: list[Mau30SummaryRow] = []
        for selected_branch in branch_order:
            ordered_rows.extend(
                (branch, account, amount)
                for (branch, account), amount in grouped.items()
                if branch == selected_branch
            )
        return ordered_rows

    @staticmethod
    def _order_branch_columns_by_branch(
        rows: list[Mau30SummaryRow],
        branch_columns: list[tuple[int, str]],
    ) -> list[Mau30SummaryRow]:
        ordered: list[Mau30SummaryRow] = []
        for _, branch in branch_columns:
            ordered.extend(
                row
                for row in rows
                if len(row) == 3 and row[0] == branch
            )
        return ordered

    def _write_header(self, sheet, request, source_sheet, model: str) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        sheet["A1"] = "NGÂN HÀNG NÔNG NGHIỆP"
        sheet["A2"] = "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM"
        sheet["A3"] = f"Mã chi nhánh: {profile.branch_code.strip()}"
        sheet["A4"] = f"Tên {branch_name}"
        sheet["A5"] = (
            "TỔNG HỢP SỐ LIỆU TOÀN CHI NHÁNH MẪU BIỂU QUYẾT TOÁN SỐ "
            f"{model}"
        )
        sheet["A6"] = f"Tên mẫu biểu: {self._source_title(source_sheet, model)}"
        sheet["A7"] = self._source_report_date(source_sheet, model, request)
        sheet["H8"] = "Đơn vị : VND"
        sheet["F1"] = "1. Mẫu số 30/QT"
        sheet["F2"] = "2. CN loại I gửi kèm trong file mẫu biểu QT về TSC"
        sheet["F3"] = "3. Lưu tại chi nhánh"
        for address in ("A1:D1", "A2:D2", "A3:D3", "A4:D4", "A5:H5", "A6:H6", "A7:H7"):
            sheet.merge_cells(address)

        headers = {
            "A9": "TT",
            "B9": "CHI NHÁNH TRỰC THUỘC",
            "C9": "MÃ CHI NHÁNH TRÊN HỆ THỐNG IPCAS",
            "D9": "TÀI KHOẢN (Cấp V)",
            "E9": "TÌNH TRẠNG BÁO CÁO",
            "E10": "SỐ DƯ TÀI KHOẢN TRÊN CÂN ĐỐI IPCAS",
            "F10": "SỐ LIỆU TỔNG CỘNG CỦA TÀI KHOẢN TRÊN BÁO CÁO QUYẾT TOÁN CỦA CHI NHÁNH",
            "G10": "CHÊNH LỆCH",
            "H9": "NGUYÊN NHÂN",
        }
        for address, value in headers.items():
            sheet[address] = value
        for address in ("A9:A10", "B9:B10", "C9:C10", "D9:D10", "E9:G9", "H9:H10"):
            sheet.merge_cells(address)
        for column in range(1, 9):
            sheet.cell(11, column).value = column
        sheet["G11"] = "7 = 5 - 6"

    def _write_rows(
        self,
        sheet,
        request,
        rows: list[Mau30SummaryRow],
        balance: Mau30BalanceMap,
    ) -> None:
        parent_code = request.profile.parent_branch_code.strip()
        branch_code = request.profile.branch_code.strip()
        for index, row_data in enumerate(rows, start=12):
            if len(row_data) == 3:
                row_branch, account, amount = row_data
            else:
                account, amount = row_data
                row_branch = branch_code
            sheet.cell(index, 1).value = index - 11
            sheet.cell(index, 2).value = parent_code
            sheet.cell(index, 3).value = row_branch
            sheet.cell(index, 4).value = account
            branch_balance_key = (row_branch, account)
            if branch_balance_key in balance:
                sheet.cell(index, 5).value = self._number_to_cell(balance[branch_balance_key])
            elif account in balance:
                sheet.cell(index, 5).value = self._number_to_cell(balance[account])
            sheet.cell(index, 6).value = self._number_to_cell(amount)
            sheet.cell(index, 7).value = f"=F{index}-E{index}"

    def _write_footer(self, sheet, request, final_row: int) -> None:
        sheet.cell(final_row + 2, 1).value = "LẬP BIỂU"
        sheet.cell(final_row + 3, 1).value = "(Ký, ghi rõ họ tên, số ĐT liên hệ)"
        sheet.cell(final_row + 2, 4).value = "TRƯỞNG PHÒNG KẾ TOÁN"
        sheet.cell(final_row + 3, 4).value = "(Ký, ghi rõ họ tên)"
        report_date = str(sheet["A7"].value or "").strip()
        report_location = request.profile.report_location.strip()
        sheet.cell(final_row + 1, 7).value = (
            f"{report_location}, {report_date}" if report_location else report_date
        )
        sheet.cell(final_row + 2, 7).value = "GIÁM ĐỐC"
        sheet.cell(final_row + 3, 7).value = "(Ký, ghi rõ họ tên)"
        sheet.cell(final_row + 8, 1).value = "Ghi chú"
        sheet.cell(final_row + 9, 2).value = (
            "1. Cột 8: trong trường hợp không khớp đúng tại cột 7, "
            "chi nhánh phải nêu rõ nguyên nhân."
        )
        sheet.cell(final_row + 10, 2).value = (
            "2. Chi nhánh loại I, II tổng hợp mẫu biểu này đối với các mẫu "
            "quyết toán và được lưu vào sheet 1 của file tổng hợp mẫu biểu quyết toán đó."
        )
        for address in (
            f"A{final_row + 2}:C{final_row + 2}",
            f"D{final_row + 2}:F{final_row + 2}",
            f"G{final_row + 1}:H{final_row + 1}",
            f"G{final_row + 2}:H{final_row + 2}",
            f"A{final_row + 3}:C{final_row + 3}",
            f"D{final_row + 3}:F{final_row + 3}",
            f"G{final_row + 3}:H{final_row + 3}",
        ):
            sheet.merge_cells(address)

    def _format(self, sheet, final_row: int) -> None:
        thin = Side(style="thin", color="000000")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=13)
                cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(min_row=9, max_row=final_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(9, 12):
            for column in range(1, 9):
                cell = sheet.cell(row, column)
                cell.font = Font(name="Times New Roman", size=13, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(12, final_row + 1):
            for column in (1, 2, 3, 4):
                sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
            for column in (5, 6, 7):
                sheet.cell(row, column).number_format = "#,##0"
        for row in range(12, final_row + 1):
            if sheet.cell(row, 7).value:
                sheet.cell(row, 7).fill = PatternFill("solid", fgColor="FFF2CC")

        for row in range(1, 5):
            cell = sheet.cell(row, 1)
            cell.font = Font(name="Times New Roman", size=13, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet["A5"].font = Font(name="Times New Roman", size=16, bold=True)
        sheet["A5"].alignment = Alignment(horizontal="center", vertical="center")
        for address in ("A6", "A7"):
            sheet[address].font = Font(
                name="Times New Roman",
                size=13,
                bold=(address == "A6"),
                italic=(address == "A7"),
            )
            sheet[address].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet["H8"].font = Font(name="Times New Roman", size=13, italic=True)
        sheet["H8"].alignment = Alignment(horizontal="right", vertical="center")
        for row in (final_row + 1, final_row + 3, final_row + 8):
            for column in range(1, 9):
                sheet.cell(row, column).font = Font(name="Times New Roman", size=13, italic=True)
        for row in (final_row + 2,):
            for column in range(1, 9):
                sheet.cell(row, column).font = Font(name="Times New Roman", size=13, bold=True)
        for row in range(final_row + 1, final_row + 4):
            for column in range(1, 9):
                sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
        report_date_cell = sheet.cell(final_row + 1, 7)
        report_date_cell.font = Font(name="Times New Roman", size=12, italic=True)
        report_date_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False,
            shrink_to_fit=True,
        )
        sheet.row_dimensions[final_row + 1].height = 21
        for row in range(final_row + 9, final_row + 11):
            sheet.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        widths = {"A": 5, "B": 15.5, "C": 19.5, "D": 13, "E": 25, "F": 25, "G": 20, "H": 18}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[9].height = 16
        sheet.row_dimensions[10].height = 65
        sheet.row_dimensions[5].height = 25
        sheet.row_dimensions[6].height = 40
        sheet.row_dimensions[5].font = Font(name="Times New Roman", size=16, bold=True)
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:H{final_row + 10}",
            orientation="landscape",
            title_rows="$11:$11",
        )
        sheet.page_setup.scale = 85

    @staticmethod
    def _source_title(source_sheet, model: str) -> str:
        if model == "05":
            return str(source_sheet["A5"].value or "")
        if model == "13":
            return f"{source_sheet['A5'].value or ''} {source_sheet['A6'].value or ''}".strip()
        return str(source_sheet["A6"].value or "")

    @staticmethod
    def _source_report_date(source_sheet, model: str, request: SettlementRequest) -> str:
        if model == "05":
            return str(source_sheet["A6"].value or "")
        if model == "13":
            return str(source_sheet["A7"].value or "")
        value = str(source_sheet["A7"].value or "")
        if model == "22":
            return Mau30Processor._normalize_report_date(value, request)
        return value

    @staticmethod
    def _normalize_report_date(value: str, request: SettlementRequest) -> str:
        text = value.strip()
        compact = text.replace("/", " ").replace("-", " ")
        parts = [part for part in compact.split() if part.isdigit()]
        if len(parts) >= 3:
            day, month, year = parts[-3:]
            return f"ngày {int(day)} tháng {int(month)} năm {int(year)}"
        if text.casefold().startswith("ngày "):
            return "ngày " + text[5:].strip()
        today = date.today()
        if request.options.output_prefix == "BN":
            return f"ngày 30 tháng 6 năm {today.year}"
        return f"ngày 31 tháng 12 năm {today.year}"

    @staticmethod
    def _model_from_request(request: SettlementRequest, source_path: Path, workbook) -> str:
        selected = request.options.source_report_code.strip()
        if selected:
            return selected
        for part in source_path.stem.upper().split("QT", 1)[-1:]:
            normalized = part.replace("A", "a")
            if normalized in {"05", "13", "15a", "15b", "18", "20a", "22", "23", "24"}:
                return normalized
        for candidate in ("05", "13", "15a", "15b", "18", "20a", "22", "23", "24"):
            if Mau30Processor._sheet_name(workbook, candidate):
                return candidate
        raise SettlementError("Không xác định được mã mẫu QT nguồn để tạo Mẫu 30/QT.")

    @staticmethod
    def _sheet_name(workbook, model: str) -> str:
        for sheet_name in workbook.sheetnames:
            if sheet_name.casefold() == model.casefold():
                return sheet_name
        detail_name = f"SoLieu_Mau{model}".casefold()
        for sheet_name in workbook.sheetnames:
            if sheet_name.casefold() == detail_name:
                return sheet_name
        return ""

    @staticmethod
    def _consolidation_sheet_name(workbook, model: str) -> str:
        expected = f"TongHop_Mau{model}".casefold()
        for sheet_name in workbook.sheetnames:
            if sheet_name.casefold() == expected:
                return sheet_name
        return ""

    @classmethod
    def _normalize_summary_account(cls, value: Any) -> str:
        account = cls._text(value)
        if "_" in account:
            account = account.split("_", 1)[0]
        return account.strip()

    @staticmethod
    def _is_branch_total_header(value: str) -> bool:
        text = value.casefold()
        return text.startswith(("cộng", "cong")) or (
            "chi nh" in text and ("cộng" in text or "cong" in text)
        )

    @classmethod
    def _header_columns_are_branches(cls, sheet, header_row: int) -> bool:
        for column in range(2, sheet.max_column + 1):
            header = cls._text(sheet.cell(header_row, column).value)
            if not header:
                break
            if cls._is_branch_total_header(header):
                break
            return header.isdigit() and len(header) == 4
        return False

    def _read_balance(self, balance_path: Path | None) -> Mau30BalanceMap:
        if balance_path is None:
            return {}
        if balance_path.is_dir():
            return self._read_balance_folder(balance_path)
        suffix = balance_path.suffix.casefold()
        if suffix == ".xls":
            rows_by_sheet = self._read_xls_balance_sheets(balance_path)
        elif suffix in {".xlsx", ".xlsm"}:
            rows_by_sheet = self._read_xlsx_balance_sheets(balance_path)
        else:
            raise SettlementError("File cân đối phải là .xls/.xlsx/.xlsm.")
        required = {"acctcd", "afterbal_dr", "afterbal_cr"}
        for rows in rows_by_sheet:
            if not rows:
                continue
            headers = {key.casefold() for key in rows[0]}
            if not required.issubset(headers):
                continue
            result: dict[str, Decimal] = {}
            for row in rows:
                normalized = {key.casefold(): value for key, value in row.items()}
                account = self._text(normalized.get("acctcd")).strip()
                if not account:
                    continue
                debit = self._number(normalized.get("afterbal_dr"))
                credit = self._number(normalized.get("afterbal_cr"))
                result[account] = debit if debit != 0 else credit
            return result
        raise SettlementError(
            "File cân đối không có các cột Acctcd, afterbal_dr, afterbal_cr."
        )

    def _read_balance_folder(self, folder_path: Path) -> dict[tuple[str, str], Decimal]:
        result: dict[tuple[str, str], Decimal] = {}
        for path in sorted(folder_path.iterdir(), key=lambda item: item.name.casefold()):
            if (
                not path.is_file()
                or path.name.startswith("~$")
                or path.suffix.casefold() not in {".xls", ".xlsx", ".xlsm"}
            ):
                continue
            branch_code = path.stem.strip()
            if not branch_code.isdigit():
                continue
            branch_balance = self._read_balance(path)
            for account, amount in branch_balance.items():
                if isinstance(account, tuple):
                    continue
                result[(branch_code, account)] = amount
        return result

    def _read_xls_balance_sheets(self, path: Path) -> list[list[dict[str, Any]]]:
        try:
            import xlrd
        except ImportError as exc:
            raise SettlementError(
                "Cần xlrd để đọc file cân đối .xls.",
                code="missing_xlrd",
            ) from exc
        workbook = xlrd.open_workbook(str(path), logfile=StringIO())
        sheets: list[list[dict[str, Any]]] = []
        for sheet in workbook.sheets():
            if sheet.nrows < 2:
                continue
            headers = [self._text(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
            rows = []
            for row_index in range(1, sheet.nrows):
                values = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
                rows.append(dict(zip(headers, values, strict=False)))
            sheets.append(rows)
        return sheets

    def _read_xlsx_balance_sheets(self, path: Path) -> list[list[dict[str, Any]]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheets: list[list[dict[str, Any]]] = []
        for sheet in workbook.worksheets:
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                headers = [self._text(value) for value in next(rows_iter)]
            except StopIteration:
                continue
            rows = [dict(zip(headers, values, strict=False)) for values in rows_iter]
            sheets.append(rows)
        return sheets

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _number(value: Any) -> Decimal:
        if value is None or value == "":
            return Decimal(0)
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).replace(".", "").replace(",", ".").strip()
        try:
            return Decimal(text)
        except Exception:
            return Decimal(0)

    @staticmethod
    def _number_to_cell(value: Decimal) -> int | float:
        if value == value.to_integral_value():
            return int(value)
        return float(value)
