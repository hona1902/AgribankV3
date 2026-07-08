from __future__ import annotations

from collections.abc import Iterable
from datetime import date
from decimal import Decimal, InvalidOperation
from io import StringIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.theme import theme_xml

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementOptions, SettlementRequest, SettlementResult
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)


class Mau04Processor:
    """Processor for Mẫu 04/QT, ported from QUYETTOAN_MAU04."""

    REQUIRED_HEADERS = {
        "df_seq",
        "df_code",
        "df_name",
        "df_unit",
        "df_sqty",
        "df_samt",
        "df_iqty",
        "df_tqty",
        "df_iamt",
        "df_eqty",
        "name_1",
        "df_uqty",
        "df_dqty",
        "df_eamt",
        "df_bqty",
        "df_bamt",
        "df_vqty",
        "df_vamt",
        "df_rqty",
        "df_ramt",
        "df_lqty",
        "df_lamt",
        "df_reason",
    }
    VALUE_HEADERS = (
        "df_sqty",
        "df_samt",
        "df_iqty",
        "df_tqty",
        "df_iamt",
        "df_eqty",
        "name_1",
        "df_uqty",
        "df_dqty",
        "df_eamt",
        "df_bqty",
        "df_bamt",
        "df_vqty",
        "df_vamt",
        "df_rqty",
        "df_ramt",
        "df_lqty",
        "df_lamt",
    )
    CHECKING_CODES = {"300", "310"}
    TERM_DEPOSIT_CODES = {"200", "210", "220", "230", "231", "240", "241", "242", "250", "710"}
    NON_TERM_DEPOSIT_CODES = {"100", "110", "120", "700"}

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError("Mẫu 04/QT cần đúng một file nguồn.")
        source_path = request.source_paths[0]
        rows = self.read_source(source_path)
        summary = self.summarize(rows)
        workbook = self.build_workbook(request, summary)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}{self._output_prefix(request.options)}04.xlsx"
        )
        workbook.save(output_path)
        workbook.close()
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name="04",
            processed_rows=len(rows),
            warnings=(
                "Vui lòng điền số lượng và số tiền ấn chỉ tồn đầu kỳ "
                f"để hoàn thiện báo cáo {self._period_label(request.options)}.",
            ),
        )

    def read_source(self, source_path: Path) -> list[dict[str, Any]]:
        suffix = source_path.suffix.casefold()
        if suffix == ".xls":
            rows = self._read_xls_rows(source_path)
        elif suffix in {".xlsx", ".xlsm"}:
            rows = self._read_xlsx_rows(source_path)
        else:
            raise SettlementError("Nguồn Mẫu 04/QT phải là file Excel .xls/.xlsx.")
        if not rows:
            return []
        missing = sorted(self.REQUIRED_HEADERS - set(rows[0]))
        if missing:
            raise SettlementError(
                f"File nguồn thiếu cột: {', '.join(missing)}",
                code="invalid_mau04_headers",
            )
        return rows

    def summarize(self, rows: Iterable[dict[str, Any]]) -> list[list[Decimal]]:
        summary = [[Decimal(0) for _ in self.VALUE_HEADERS] for _ in range(9)]
        for row in rows:
            group_index = self._group_index(self._text(row.get("df_code")))
            values = [self._decimal(row.get(header)) for header in self.VALUE_HEADERS]
            for index, value in enumerate(values):
                summary[group_index][index] += value

        for index in range(len(self.VALUE_HEADERS)):
            summary[2][index] = summary[3][index] + summary[4][index]
            summary[8][index] = (
                summary[0][index]
                + summary[1][index]
                + summary[2][index]
                + summary[5][index]
                + summary[6][index]
                + summary[7][index]
            )
        return summary

    def build_workbook(
        self,
        request: SettlementRequest,
        summary: list[list[Decimal]],
    ) -> Workbook:
        workbook = Workbook()
        self._set_default_font(workbook)
        sheet = workbook.active
        sheet.title = "04"
        report_date = self._report_date_text(request.options)
        self._write_header(sheet, request, report_date)
        self._write_table_header(sheet)
        self._write_data_rows(sheet, summary)
        self._write_signatures_and_notes(sheet, request, report_date)
        self._format_report(sheet)
        return workbook

    def _read_xls_rows(self, source_path: Path) -> list[dict[str, Any]]:
        try:
            import xlrd
        except ImportError as exc:
            raise SettlementError(
                "Mẫu 04/QT cần thư viện xlrd để đọc file .xls. "
                "Hãy chạy lại cài đặt môi trường hoặc pip install -e .",
                code="missing_xlrd",
            ) from exc
        try:
            workbook = xlrd.open_workbook(str(source_path), logfile=StringIO())
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc
        if sheet.nrows < 2:
            return []
        headers = [self._text(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        rows: list[dict[str, Any]] = []
        for row_index in range(1, sheet.nrows):
            values = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
            if any(self._text(value) for value in values):
                rows.append(dict(zip(headers, values, strict=False)))
        return rows

    def _read_xlsx_rows(self, source_path: Path) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(source_path, read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [self._text(value) for value in next(rows_iter)]
        except StopIteration:
            return []
        rows = []
        for values in rows_iter:
            if any(self._text(value) for value in values):
                rows.append(dict(zip(headers, values, strict=False)))
        workbook.close()
        return rows

    def _group_index(self, code: str) -> int:
        if code in self.CHECKING_CODES:
            return 0
        if code == "400":
            return 1
        if code in self.TERM_DEPOSIT_CODES:
            return 3
        if code in self.NON_TERM_DEPOSIT_CODES:
            return 4
        if code == "800":
            return 5
        if code == "600":
            return 6
        return 7

    def _write_header(self, sheet, request: SettlementRequest, report_date: str) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        header_values = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": f"Mã chi nhánh: {profile.branch_code.strip()}",
            "A4": f"Tên {branch_name}",
            "A5": "BÁO CÁO TÌNH HÌNH SỬ DỤNG ẤN CHỈ QUAN TRỌNG",
            "A6": report_date,
            "S1": "1. Mẫu số 04/QT",
            "S2": "2. Lưu tại chi nhánh",
        }
        for cell, value in header_values.items():
            sheet[cell] = value
        for row in range(1, 5):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        sheet.merge_cells("A5:W5")
        sheet.merge_cells("A6:W6")
        style_agency_header(sheet, start_column=1, end_column=5)
        sheet["A5"].font = Font(name="Times New Roman", size=16, bold=True)
        sheet["A6"].font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        for row in range(5, 7):
            sheet.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        sheet["W7"] = "Đơn vị : VNĐ"
        style_currency_unit(sheet["W7"])

    def _write_table_header(self, sheet) -> None:
        merges = (
            "A8:A11",
            "B8:B11",
            "C8:C11",
            "D8:E9",
            "D10:D11",
            "E10:E11",
            "F8:I9",
            "F10:H10",
            "I10:I11",
            "J8:N9",
            "J10:M10",
            "N10:N11",
            "O8:R9",
            "O10:P10",
            "Q10:R10",
            "S8:V9",
            "S10:T10",
            "U10:V10",
            "W8:W11",
        )
        for merge in merges:
            sheet.merge_cells(merge)
        header_values = {
            "A8": "TT",
            "B8": "LOẠI ẤN CHỈ",
            "C8": "ĐƠN VỊ TÍNH",
            "D8": "ĐẦU KỲ",
            "D10": "Số lượng",
            "E10": "Số tiền",
            "F8": "NHẬP TRONG KỲ",
            "F10": "Số lượng",
            "F11": "Nhận từ TSC",
            "G11": "Nhận khác CN",
            "H11": "Điều chuyển nội bộ",
            "I10": "Số tiền",
            "J8": "XUẤT TRONG KỲ",
            "J10": "Số lượng",
            "J11": "Điều chuyển khác CN",
            "K11": "Điều chuyển nội bộ",
            "L11": "Phát hành vào lưu thông",
            "M11": "Hỏng",
            "N10": "Số tiền",
            "O8": "TỒN CUỐI KỲ",
            "O10": "Sổ sách",
            "O11": "Số lượng",
            "P11": "Số tiền",
            "Q10": "Thực tế",
            "Q11": "Số lượng",
            "R11": "Số tiền",
            "S8": "CHÊNH LỆCH",
            "S10": "Thừa",
            "S11": "Số lượng",
            "T11": "Số tiền",
            "U10": "Thiếu",
            "U11": "Số lượng",
            "V11": "Số tiền",
            "W8": "NGUYÊN NHÂN",
        }
        for cell, value in header_values.items():
            sheet[cell] = value
        for column in range(1, 24):
            cell = sheet.cell(12, column)
            cell.value = (
                "15=(4+6+7+8)\n-(10-11-12-13)"
                if column == 15
                else column
            )
            cell.font = Font(name="Times New Roman", size=10, italic=True)

    def _write_data_rows(self, sheet, summary: list[list[Decimal]]) -> None:
        rows = (
            (1, "Séc", "Tờ", 0),
            (2, "Sổ tiết kiệm có kỳ hạn", "Sổ", 3),
            (3, "Sổ tiết kiệm", "Sổ", 4),
            (4, "Giấy tờ có giá", "Sổ", 5),
            (5, "Bảo lãnh", "Tờ", 6),
            (6, "Ấn chỉ quan trọng khác", None, 7),
        )
        for offset, (ordinal, label, unit, summary_index) in enumerate(rows):
            row = 13 + offset
            sheet.cell(row, 1, ordinal)
            sheet.cell(row, 2, label)
            sheet.cell(row, 3, unit)

            values = summary[summary_index]
            target_columns = (
                4,
                5,
                6,
                8,
                9,
                10,
                11,
                12,
                13,
                14,
                15,
                16,
                17,
                18,
                19,
                20,
                21,
                22,
            )
            for column, value in zip(target_columns, values, strict=True):
                if value != 0:
                    sheet.cell(row, column, self._number(value))
            sheet.cell(row, 15, f"=D{row}+F{row}+G{row}+H{row}-J{row}-K{row}-L{row}-M{row}")
            sheet.cell(row, 16, f"=E{row}+I{row}-N{row}")

        total_row = 19
        sheet.cell(total_row, 2, "TỔNG CỘNG")
        for column in range(4, 19):
            sheet.cell(
                total_row,
                column,
                f"=SUM({get_column_letter(column)}13:{get_column_letter(column)}18)",
            )

    def _write_signatures_and_notes(
        self,
        sheet,
        request: SettlementRequest,
        report_date: str,
    ) -> None:
        profile = request.profile
        report_location = profile.report_location.strip()
        signature_date = report_date
        if report_location:
            signature_date = f"{report_location}, {report_date}"
        sheet.merge_cells("A21:F21")
        sheet.merge_cells("A22:F22")
        sheet.merge_cells("A26:F26")
        sheet.merge_cells("A27:F27")
        sheet.merge_cells("I21:O21")
        sheet.merge_cells("I22:O22")
        sheet.merge_cells("R20:W20")
        sheet.merge_cells("R21:W21")
        sheet.merge_cells("R22:W22")
        sheet["A21"] = "LẬP BIỂU"
        sheet["A22"] = "(Ký, ghi rõ họ tên, số điện thoại)"
        sheet["A26"] = profile.report_preparer.strip()
        sheet["A27"] = f"({profile.phone.strip()})" if profile.phone.strip() else ""
        sheet["I21"] = "TRƯỞNG PHÒNG KẾ TOÁN"
        sheet["I22"] = "(Ký, ghi rõ họ tên)"
        sheet["R20"] = signature_date
        sheet["R21"] = "GIÁM ĐỐC"
        sheet["R22"] = "(Ký, đóng dấu, ghi rõ họ tên)"
        sheet["A28"] = "Ghi chú:"
        sheet["B29"] = (
            "1. Nguyên tắc số lượng xuất nội bộ = số lượng nhập nội bộ "
            "trong toàn Agribank chi nhánh loại I"
        )
        sheet["B30"] = "2. Chi nhánh sử dụng màn hình GA/IC/100435."

    def _format_report(self, sheet) -> None:
        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in sheet.iter_rows(min_row=1, max_row=30, min_col=1, max_col=23):
            for cell in row:
                self._apply_times_font(cell)
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical="center",
                    wrap_text=cell.alignment.wrap_text,
                )

        for row in range(8, 12):
            for column in range(1, 24):
                cell = sheet.cell(row, column)
                cell.font = Font(name="Times New Roman", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

        for row in range(12, 20):
            for column in range(1, 24):
                cell = sheet.cell(row, column)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center" if column in {1, 3, 4, 6, 7, 8, 10, 11, 12, 13, 15, 17, 19, 21} else "right",
                    vertical="center",
                    wrap_text=(column == 15 and row == 12),
                )
        for row in range(13, 20):
            sheet.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center")
            sheet.cell(row, 23).alignment = Alignment(horizontal="left", vertical="center")
        for row in (13, 14, 15, 16, 17, 18, 19):
            sheet.cell(row, 2).font = Font(name="Times New Roman", size=10, bold=True)
        for row in range(13, 20):
            for column in range(4, 23):
                sheet.cell(row, column).number_format = "#,##0"

        for column in range(1, 24):
            sheet.cell(12, column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.cell(12, column).border = thin_border
        sheet["O12"].font = Font(name="Times New Roman", size=8, italic=True)
        for column in range(1, 24):
            sheet.cell(19, column).font = Font(name="Times New Roman", size=10, bold=True)
            sheet.cell(19, column).border = thin_border

        for range_ref in ("A21:F22", "I21:O22", "R20:W22", "A26:F27"):
            for row in sheet[range_ref]:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell_ref in ("A21", "I21", "R21"):
            sheet[cell_ref].font = Font(name="Times New Roman", size=10, bold=True)
        for cell_ref in ("A22", "I22", "R20", "R22", "A27"):
            sheet[cell_ref].font = Font(name="Times New Roman", size=10, italic=True)
        sheet["A28"].font = Font(name="Times New Roman", size=10, bold=True)
        for cell_ref in ("B29", "B30"):
            sheet[cell_ref].alignment = Alignment(horizontal="left", vertical="center")

        widths = {
            "A": 3.86,
            "B": 24.5,
            "C": 7.5,
            "D": 7.57,
            "E": 10.3,
            "F": 6.6,
            "G": 8.9,
            "H": 7.2,
            "I": 12.9,
            "J": 6,
            "K": 6.6,
            "L": 8.9,
            "M": 5.4,
            "N": 12.6,
            "O": 14,
            "P": 11.5,
            "Q": 5.4,
            "R": 13,
            "S": 5.8,
            "T": 6,
            "U": 13,
            "V": 5.4,
            "W": 12,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[8].height = 19
        sheet.row_dimensions[9].height = 19
        sheet.row_dimensions[10].height = 24
        sheet.row_dimensions[11].height = 44
        sheet.row_dimensions[12].height = 34
        for row in range(13, 20):
            sheet.row_dimensions[row].height = 19
        sheet.sheet_view.showGridLines = False
        setup_a4_print_layout(sheet, print_area="A1:W30", title_rows="$8:$12")

    def _set_default_font(self, workbook: Workbook) -> None:
        default_font = Font(name="Times New Roman", size=10)
        normal_style = workbook._named_styles["Normal"]
        normal_style.font = default_font
        if workbook._fonts:
            workbook._fonts[0] = default_font
        workbook.loaded_theme = theme_xml.replace("Calibri", "Times New Roman").encode("utf-8")

    def _apply_times_font(
        self,
        cell,
        *,
        size: int | float | None = None,
        bold: bool | None = None,
        italic: bool | None = None,
    ) -> None:
        current = cell.font
        current_size = current.sz if current.sz not in (None, 11) else 10
        cell.font = Font(
            name="Times New Roman",
            size=size if size is not None else current_size,
            bold=current.bold if bold is None else bold,
            italic=current.italic if italic is None else italic,
            underline=current.underline,
            strike=current.strike,
            color=current.color,
        )

    def _text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _decimal(self, value: Any) -> Decimal:
        if value is None or value == "":
            return Decimal(0)
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value).strip().replace(",", ""))
        except (InvalidOperation, AttributeError):
            return Decimal(0)

    def _number(self, value: Decimal) -> int | float:
        if value == value.to_integral_value():
            return int(value)
        return float(value)

    def _output_prefix(self, options: SettlementOptions) -> str:
        return "BN" if options.output_prefix.strip().upper() == "BN" else "QT"

    def _report_date_text(self, options: SettlementOptions) -> str:
        year = date.today().year
        if self._output_prefix(options) == "BN":
            return f"Ngày 30 tháng 6 năm {year}"
        return f"Ngày 31 tháng 12 năm {year}"

    def _period_label(self, options: SettlementOptions) -> str:
        return "bán niên" if self._output_prefix(options) == "BN" else "Quyết toán"
