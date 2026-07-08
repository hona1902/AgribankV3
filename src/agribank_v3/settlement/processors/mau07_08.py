from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from io import StringIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.theme import theme_xml

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementOptions, SettlementRequest, SettlementResult
from agribank_v3.settlement.processors.formatting import setup_a4_print_layout, style_agency_header, style_currency_unit


@dataclass(frozen=True, slots=True)
class InventoryRecord:
    code: str
    name: str
    book_qty: Decimal
    book_amount: Decimal
    actual_qty: Decimal
    actual_amount: Decimal
    reason: str = ""


class Mau0708Processor:
    """Processor for accounting Mẫu 07a/QT and Mẫu 08/QT."""

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError(f"Mẫu {request.spec.report_code}/QT cần đúng một file nguồn.")
        source_path = request.source_paths[0]
        report_code = self._report_code(request)
        if report_code == "07a":
            records = self.read_mau07_source(source_path)
            workbook = self.build_mau07_workbook(request, records)
        else:
            records = self.read_mau08_source(request, source_path)
            workbook = self.build_mau08_workbook(request, records)
        output_code = "07A" if report_code == "07a" else "08"
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}{self._output_prefix(request.options)}{output_code}.xlsx"
        )
        workbook.save(output_path)
        workbook.close()
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name=report_code,
            processed_rows=len(records),
        )

    def read_mau07_source(self, source_path: Path) -> list[InventoryRecord]:
        rows = self._read_xls_rows(source_path)
        if not rows:
            return []
        headers = [self._text(value) for value in rows[0]]
        if headers[:9] == ["buscd", "brcd", "name_1", "name_2", "name_3", "name_4", "name_5", "name_6", "name_7"]:
            records: list[InventoryRecord] = []
            for row in rows[1:]:
                if not any(self._text(value) for value in row):
                    continue
                records.append(
                    InventoryRecord(
                        code=self._clean_code(row[3] if len(row) > 3 else ""),
                        name=self._text(row[4] if len(row) > 4 else ""),
                        book_qty=self._decimal(row[5] if len(row) > 5 else 0),
                        book_amount=self._decimal(row[6] if len(row) > 6 else 0),
                        actual_qty=self._decimal(row[5] if len(row) > 5 else 0),
                        actual_amount=self._decimal(row[6] if len(row) > 6 else 0),
                        reason="",
                    )
                )
            return records
        return self._read_formatted_inventory_rows(rows)

    def read_mau08_source(
        self,
        request: SettlementRequest,
        source_path: Path,
    ) -> list[InventoryRecord]:
        rows = self._read_xls_rows(source_path)
        if not rows:
            return []
        return self._read_formatted_inventory_rows(rows, branch_code=request.profile.branch_code.strip())

    def build_mau07_workbook(
        self,
        request: SettlementRequest,
        records: list[InventoryRecord],
    ) -> Workbook:
        workbook = Workbook()
        self._set_default_font(workbook)
        sheet = workbook.active
        sheet.title = "7a"
        self._write_mau07_header(sheet, request)
        self._write_mau07_table(sheet, records)
        total_row = 11 + len(records)
        self._write_mau07_summary(sheet, total_row)
        self._write_mau07_signatures_and_notes(sheet, request, total_row + 10)
        self._format_mau07(sheet, total_row)
        return workbook

    def build_mau08_workbook(
        self,
        request: SettlementRequest,
        records: list[InventoryRecord],
    ) -> Workbook:
        workbook = Workbook()
        self._set_default_font(workbook)
        sheet = workbook.active
        sheet.title = "8"
        groups = self._mau08_groups(records)
        self._write_mau08_header(sheet, request)
        self._write_mau08_table(sheet, groups)
        self._write_mau08_signatures_and_notes(sheet, request, 18)
        self._format_mau08(sheet)
        return workbook

    def _read_xls_rows(self, source_path: Path) -> list[list[Any]]:
        if source_path.suffix.casefold() not in {".xls", ".xlsx", ".xlsm"}:
            raise SettlementError("Nguồn mẫu 07a/08 phải là file Excel.")
        if source_path.suffix.casefold() != ".xls":
            raise SettlementError("Nguồn mẫu 07a/08 hiện cần file Excel .xls xuất từ IPCAS.")
        try:
            import xlrd
        except ImportError as exc:
            raise SettlementError("Cần thư viện xlrd để đọc file .xls.") from exc
        try:
            workbook = xlrd.open_workbook(str(source_path), logfile=StringIO())
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc
        return [
            [sheet.cell_value(row, col) for col in range(sheet.ncols)]
            for row in range(sheet.nrows)
        ]

    def _read_formatted_inventory_rows(
        self,
        rows: list[list[Any]],
        *,
        branch_code: str = "",
    ) -> list[InventoryRecord]:
        start = 0
        if branch_code:
            for index, row in enumerate(rows):
                if self._text(row[0] if row else "").startswith(branch_code):
                    start = index + 1
                    break
        else:
            for index, row in enumerate(rows):
                if self._text(row[0] if row else "").casefold() == "stt":
                    start = index + 2
                    break
        records: list[InventoryRecord] = []
        for row in rows[start:]:
            if len(row) < 7:
                continue
            if not self._is_number_like(row[0]):
                continue
            records.append(
                InventoryRecord(
                    code=self._clean_code(row[1]),
                    name=self._text(row[2]),
                    book_qty=self._decimal(row[3]),
                    book_amount=self._decimal(row[4]),
                    actual_qty=self._decimal(row[5]),
                    actual_amount=self._decimal(row[6]),
                    reason=self._text(row[11] if len(row) > 11 else ""),
                )
            )
        return records

    def _write_mau07_header(self, sheet, request: SettlementRequest) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        report_date = self._report_date_text(request.options)
        values = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": f"Mã chi nhánh: {profile.branch_code.strip()}",
            "A4": f"Tên {branch_name}",
            "A5": "BÁO CÁO KIỂM KÊ CÔNG CỤ DỤNG CỤ",
            "A6": report_date,
            "J1": "1. Mẫu số 07a/QT",
            "J2": "2. CN loại I gửi file về TSC",
            "J3": "3. Lưu tại chi nhánh",
        }
        for cell, value in values.items():
            sheet[cell] = value
        for row in range(1, 5):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        for row in range(1, 4):
            sheet.merge_cells(start_row=row, start_column=10, end_row=row, end_column=12)
        sheet.merge_cells("A5:L5")
        sheet.merge_cells("A6:L6")
        sheet.merge_cells("I7:L7")
        sheet["I7"] = "Đơn vị : VNĐ"
        style_agency_header(sheet, start_column=1, end_column=4)
        self._style_title(sheet["A5"])
        self._style_report_date(sheet["A6"])
        style_currency_unit(sheet["I7"])
        self._write_mau07_table_header(sheet)

    def _write_mau07_table_header(self, sheet) -> None:
        merges = ("A8:A9", "B8:B9", "C8:C9", "D8:E8", "F8:G8", "H8:I8", "J8:K8", "L8:L9")
        for ref in merges:
            sheet.merge_cells(ref)
        values = {
            "A8": "TT",
            "B8": "MÃ LOẠI CCDC",
            "C8": "TÊN CÔNG CỤ DỤNG CỤ",
            "D8": "KIỂM KÊ SỔ SÁCH",
            "F8": "THEO THỰC TẾ KIỂM TRA",
            "H8": "CHÊNH LỆCH THỪA",
            "J8": "CHÊNH LỆCH THIẾU",
            "L8": "NGUYÊN NHÂN",
            "D9": "Số lượng",
            "E9": "Thành tiền",
            "F9": "Số lượng",
            "G9": "Thành tiền",
            "H9": "Số lượng",
            "I9": "Thành tiền",
            "J9": "Số lượng",
            "K9": "Thành tiền",
        }
        for cell, value in values.items():
            sheet[cell] = value
        for col in range(1, 13):
            sheet.cell(10, col, col)

    def _write_mau07_table(self, sheet, records: list[InventoryRecord]) -> None:
        for index, record in enumerate(records, start=11):
            sheet.cell(index, 1, index - 10)
            sheet.cell(index, 2, record.code)
            sheet.cell(index, 3, record.name)
            sheet.cell(index, 4, self._number(record.book_qty))
            sheet.cell(index, 5, self._number(record.book_amount))
            sheet.cell(index, 6, self._number(record.actual_qty))
            sheet.cell(index, 7, self._number(record.actual_amount))
            sheet.cell(index, 8, f"=F{index}-D{index}")
            sheet.cell(index, 9, f"=G{index}-E{index}")
            sheet.cell(index, 10, f"=F{index}-D{index}")
            sheet.cell(index, 11, f"=G{index}-E{index}")
            if record.reason:
                sheet.cell(index, 12, record.reason)

    def _write_mau07_summary(self, sheet, total_row: int) -> None:
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        sheet.cell(total_row, 1, "TỔNG CỘNG:")
        for col in range(4, 12):
            letter = get_column_letter(col)
            sheet.cell(total_row, col, f"=SUM({letter}11:{letter}{total_row - 1})")
        labels = (
            ("D", "Tồn đầu kỳ:"),
            ("D", "Mua trong kỳ:"),
            ("D", "Nhận điều chuyển:"),
            ("D", "Thanh lý:"),
            ("D", "Điều chuyển đi:"),
            ("D", "Cuối kỳ:"),
        )
        for offset, (_, label) in enumerate(labels, start=3):
            row = total_row + offset
            sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
            sheet.cell(row, 4, label)
        end_row = total_row + 8
        sheet.cell(end_row, 6, f"=F{total_row + 3}+F{total_row + 4}+F{total_row + 5}-F{total_row + 6}-F{total_row + 7}")
        sheet.cell(end_row, 8, f"=H{total_row + 3}+H{total_row + 4}+H{total_row + 5}-H{total_row + 6}-H{total_row + 7}")
        for cell in (sheet.cell(end_row, 4), sheet.cell(end_row, 6), sheet.cell(end_row, 8)):
            cell.font = Font(name="Times New Roman", size=10, bold=True)

    def _write_mau07_signatures_and_notes(self, sheet, request: SettlementRequest, base_row: int) -> None:
        self._write_signatures(sheet, request, base_row, last_col=12, first_blocks=((1, 3), (4, 6), (7, 9), (10, 12)))
        note_row = base_row + 11
        sheet.cell(note_row, 1, "Ghi chú:")
        sheet.cell(note_row, 1).font = Font(name="Times New Roman", size=10, bold=True)
        notes = [
            "- Chương trình sẽ hỗ trợ in báo cáo chi tiết kiểm kê CCDC, yêu cầu chi nhánh phải thực hiện các bước sau:",
            "- Thực hiện kiểm kê CCDC tại màn hình GA/WT/100635;",
            "- Vấn tin và kiểm tra báo cáo kiểm kê tại màn hình GA/WT/100642;",
            "- Chi nhánh in báo cáo tại màn hình GA/WT/100642, ký xác nhận và lưu vào hồ sơ quyết toán năm;",
            "- Trường hợp có chênh lệch, phải giải thích rõ nguyên nhân.",
        ]
        for offset, note in enumerate(notes, start=1):
            sheet.cell(note_row + offset, 2, note)

    def _write_mau08_header(self, sheet, request: SettlementRequest) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        report_date = self._report_date_text(request.options)
        values = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": f"Mã chi nhánh: {profile.branch_code.strip()}",
            "A4": f"Tên {branch_name}",
            "A5": "BÁO CÁO KIỂM KÊ TÀI SẢN CỐ ĐỊNH",
            "A6": report_date,
            "I1": "1. Mẫu số 08/QT",
            "I2": "2. CN loại I gửi báo cáo giấy, file về TSC",
            "I3": "3. Lưu tại Chi nhánh",
        }
        for cell, value in values.items():
            sheet[cell] = value
        for row in range(1, 5):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.merge_cells("A5:K5")
        sheet.merge_cells("A6:K6")
        sheet.merge_cells("J7:K7")
        sheet["J7"] = "Đơn vị : VNĐ"
        style_agency_header(sheet, start_column=1, end_column=2)
        self._style_title(sheet["A5"])
        self._style_report_date(sheet["A6"])
        style_currency_unit(sheet["J7"])
        self._write_mau08_table_header(sheet)

    def _write_mau08_table_header(self, sheet) -> None:
        for ref in ("A8:A9", "B8:B9", "C8:D8", "E8:F8", "G8:H8", "I8:J8", "K8:K9"):
            sheet.merge_cells(ref)
        values = {
            "A8": "TT",
            "B8": "CHỈ TIÊU",
            "C8": "GIÁ TRỊ SỔ SÁCH",
            "E8": "GIÁ TRỊ THỰC TẾ",
            "G8": "CHÊNH LỆCH THỪA",
            "I8": "CHÊNH LỆCH THIẾU",
            "K8": "NGUYÊN NHÂN",
            "C9": "Số lượng",
            "D9": "Số tiền",
            "E9": "Số lượng",
            "F9": "Số tiền",
            "G9": "Số lượng",
            "H9": "Số tiền",
            "I9": "Số lượng",
            "J9": "Số tiền",
        }
        for cell, value in values.items():
            sheet[cell] = value
        for col in range(1, 12):
            sheet.cell(10, col, col)

    def _write_mau08_table(self, sheet, groups: list[InventoryRecord]) -> None:
        labels = [
            "Nhà cửa, vật kiến trúc (TK3012)",
            "Máy móc, thiết bị (TK 3013)",
            "Phương tiện vận tải, thiết bị truyền dẫn (TK 3014)",
            "Thiết bị, dụng cụ quản lý (TK 3015)",
            "Tài sản cố định hữu hình khác (TK 3019)",
            "Tài sản cố định vô hình khác (TK 302)",
            "Tài sản cố định thuê tài chính (TK 303)",
        ]
        for idx, label in enumerate(labels, start=1):
            row = 10 + idx
            rec = groups[idx - 1]
            sheet.cell(row, 1, idx)
            sheet.cell(row, 2, label)
            sheet.cell(row, 3, self._number(rec.book_qty))
            sheet.cell(row, 4, self._number(rec.book_amount))
            sheet.cell(row, 5, self._number(rec.actual_qty))
            sheet.cell(row, 6, self._number(rec.actual_amount))
            sheet.cell(row, 7, f"=E{row}-C{row}")
            sheet.cell(row, 8, f"=F{row}-D{row}")
            sheet.cell(row, 9, f"=E{row}-C{row}")
            sheet.cell(row, 10, f"=F{row}-D{row}")
        sheet.merge_cells("A18:B18")
        sheet["A18"] = "TỔNG CỘNG"
        for col in range(3, 11):
            letter = get_column_letter(col)
            sheet.cell(18, col, f"=SUM({letter}11:{letter}17)")

    def _write_mau08_signatures_and_notes(self, sheet, request: SettlementRequest, total_row: int) -> None:
        self._write_signatures(sheet, request, total_row + 2, last_col=11, first_blocks=((1, 2), (3, 5), (6, 8), (9, 11)))
        note_row = total_row + 10
        sheet.cell(note_row, 1, "Ghi chú:")
        sheet.cell(note_row, 1).font = Font(name="Times New Roman", size=10, bold=True)
        notes = [
            "Chương trình sẽ hỗ trợ in báo cáo chi tiết kiểm kê TSCĐ, yêu cầu chi nhánh phải thực hiện các bước sau:",
            "- Thực hiện kiểm kê tài sản cố định tại màn hình GA/FA/100546;",
            "- Vấn tin và kiểm tra báo cáo kiểm kê tại màn hình GA/FA/100586:",
            "+ Kiểm tra tổng số lượng TSCĐ trên báo cáo khớp với số lượng TSCĐ đã kiểm kê;",
            "+ Kiểm tra tổng số tiền (cột 4) trên báo cáo kiểm kê phải khớp với số dư tài khoản cấp III trên cân đối.",
            "- Chi nhánh in báo cáo tại màn hình GA/FA/100586, ký xác nhận và lưu vào hồ sơ quyết toán năm.",
            "- Trường hợp có chênh lệch, phải giải thích rõ nguyên nhân.",
        ]
        for offset, note in enumerate(notes, start=0):
            sheet.cell(note_row + offset, 2 if offset else 1, note)

    def _write_signatures(
        self,
        sheet,
        request: SettlementRequest,
        row: int,
        *,
        last_col: int,
        first_blocks: tuple[tuple[int, int], tuple[int, int], tuple[int, int], tuple[int, int]],
    ) -> None:
        profile = request.profile
        report_date = self._report_date_text(request.options)
        location = profile.report_location.strip()
        signature_date = f"{location}, {report_date}" if location else report_date
        labels = (
            ("LẬP BIỂU", "(Ký, ghi rõ họ tên, số ĐT liên hệ)"),
            ("TRƯỞNG PHÒNG TỔNG HỢP", "(Ký, ghi rõ họ tên)"),
            ("TRƯỞNG PHÒNG KẾ TOÁN", "(Ký, ghi rõ họ tên)"),
            ("GIÁM ĐỐC", "(Ký, đóng dấu, ghi rõ họ tên)"),
        )
        sheet.merge_cells(start_row=row - 1, start_column=first_blocks[-1][0], end_row=row - 1, end_column=first_blocks[-1][1])
        sheet.cell(row - 1, first_blocks[-1][0], signature_date)
        sheet.cell(row - 1, first_blocks[-1][0]).font = Font(name="Times New Roman", size=10, italic=True)
        for index, (start, end) in enumerate(first_blocks):
            sheet.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            sheet.merge_cells(start_row=row + 1, start_column=start, end_row=row + 1, end_column=end)
            sheet.cell(row, start, labels[index][0])
            sheet.cell(row + 1, start, labels[index][1])
            sheet.cell(row, start).font = Font(name="Times New Roman", size=10, bold=True)
            sheet.cell(row + 1, start).font = Font(name="Times New Roman", size=10, italic=True)
        if profile.report_preparer.strip():
            sheet.merge_cells(start_row=row + 6, start_column=1, end_row=row + 6, end_column=first_blocks[0][1])
            sheet.cell(row + 6, 1, profile.report_preparer.strip())
        if profile.phone.strip():
            sheet.merge_cells(start_row=row + 7, start_column=1, end_row=row + 7, end_column=first_blocks[0][1])
            sheet.cell(row + 7, 1, f"({profile.phone.strip()})")
            sheet.cell(row + 7, 1).font = Font(name="Times New Roman", size=10, italic=True)
        for merged_row in range(row - 1, row + 8):
            for col in range(1, last_col + 1):
                sheet.cell(merged_row, col).alignment = Alignment(horizontal="center", vertical="center")

    def _format_mau07(self, sheet, total_row: int) -> None:
        self._format_common(sheet, max_row=max(total_row + 23, sheet.max_row), max_col=12)
        widths = {"A": 5.5, "B": 8.5, "C": 23.5, "D": 7.5, "E": 18, "F": 7.5, "G": 18, "H": 9.5, "I": 9.5, "J": 9.5, "K": 9.5, "L": 12.5}
        self._apply_widths(sheet, widths)
        setup_a4_print_layout(sheet, print_area=f"A1:L{sheet.max_row}", title_rows="$8:$10")

    def _format_mau08(self, sheet) -> None:
        self._format_common(sheet, max_row=sheet.max_row, max_col=11)
        widths = {"A": 7, "B": 50, "C": 7.5, "D": 20, "E": 7.5, "F": 20, "G": 15, "H": 15, "I": 15, "J": 15, "K": 15}
        self._apply_widths(sheet, widths)
        setup_a4_print_layout(sheet, print_area=f"A1:K{sheet.max_row}", title_rows="$8:$10")

    def _format_common(self, sheet, *, max_row: int, max_col: int) -> None:
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = Font(name="Times New Roman", size=cell.font.sz if cell.font.sz not in (None, 11) else 10, bold=cell.font.bold, italic=cell.font.italic)
                cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical="center", wrap_text=cell.alignment.wrap_text)
        for row in range(8, 11):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row, col)
                cell.font = Font(name="Times New Roman", size=10, bold=True if row < 10 else False, italic=True if row == 10 else False)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
        data_last = 18 if max_col == 11 else next((r for r in range(11, max_row + 1) if str(sheet.cell(r, 1).value or "").startswith("TỔNG") or str(sheet.cell(r, 2).value or "").startswith("TỔNG")), max_row)
        for row in range(11, data_last + 1):
            for col in range(1, max_col + 1):
                sheet.cell(row, col).border = border
                if col >= 3:
                    sheet.cell(row, col).number_format = "#,##0"
        sheet.sheet_view.showGridLines = False

    def _apply_widths(self, sheet, widths: dict[str, float]) -> None:
        for col, width in widths.items():
            sheet.column_dimensions[col].width = width

    def _mau08_groups(self, records: list[InventoryRecord]) -> list[InventoryRecord]:
        totals = [[Decimal(0), Decimal(0), Decimal(0), Decimal(0)] for _ in range(7)]
        for record in records:
            idx = self._mau08_group_index(record.code)
            if idx is None:
                continue
            totals[idx][0] += record.book_qty
            totals[idx][1] += record.book_amount
            totals[idx][2] += record.actual_qty
            totals[idx][3] += record.actual_amount
        return [InventoryRecord("", "", *values) for values in totals]

    def _mau08_group_index(self, code: str) -> int | None:
        if code.startswith("101"):
            return 0
        if code.startswith(("102", "103")):
            return 1
        if code.startswith("104"):
            return 2
        if code.startswith("105"):
            return 3
        if code.startswith("2"):
            return 5
        if code.startswith("3"):
            return 6
        return None

    def _style_title(self, cell) -> None:
        cell.font = Font(name="Times New Roman", size=18, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _style_report_date(self, cell) -> None:
        cell.font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _set_default_font(self, workbook: Workbook) -> None:
        default = Font(name="Times New Roman", size=10)
        workbook._named_styles["Normal"].font = default
        if workbook._fonts:
            workbook._fonts[0] = default
        workbook.loaded_theme = theme_xml.replace("Calibri", "Times New Roman").encode("utf-8")

    def _report_code(self, request: SettlementRequest) -> str:
        return "07a" if request.spec.key == "accounting.07a" else "08"

    def _output_prefix(self, options: SettlementOptions) -> str:
        return "BN" if options.output_prefix.strip().upper() == "BN" else "QT"

    def _report_date_text(self, options: SettlementOptions) -> str:
        year = date.today().year
        if self._output_prefix(options) == "BN":
            return f"Ngày 30 tháng 6 năm {year}"
        return f"Ngày 31 tháng 12 năm {year}"

    def _text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _clean_code(self, value: Any) -> str:
        return self._text(value).replace(".0", "")

    def _decimal(self, value: Any) -> Decimal:
        if value in (None, ""):
            return Decimal(0)
        try:
            return Decimal(str(value).strip().replace(",", ""))
        except (InvalidOperation, AttributeError):
            return Decimal(0)

    def _number(self, value: Decimal) -> int | float:
        return int(value) if value == value.to_integral_value() else float(value)

    def _is_number_like(self, value: Any) -> bool:
        if isinstance(value, (int, float)) and value != "":
            return True
        try:
            Decimal(str(value).strip())
            return True
        except (InvalidOperation, ValueError):
            return False
