from __future__ import annotations

from collections import OrderedDict, defaultdict
import csv
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
import gc
import re
from pathlib import Path
import shutil
import tempfile
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import (
    SettlementOptions,
    SettlementRequest,
    SettlementResult,
)
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)
from agribank_v3.settlement.transforms import (
    normalize_customer_id,
    parse_yyyymmdd,
    vietnamese_report_date,
)


@dataclass(frozen=True, slots=True)
class DepositRecord:
    ledger_account: str
    customer_id: str
    customer_name: str
    deposit_account: str
    savings_book: str
    currency: str
    original_balance: Decimal
    converted_balance: Decimal
    term_months: str
    deposit_date: date | None
    maturity_date: date | None
    interest_rate: int | float | str | None
    last_interest_date: date | None
    prepaid_interest: Decimal
    accrued_interest: Decimal
    interest_account: str
    prepaid_interest_display: int | float | None
    accrued_interest_display: int | float | None
    extra_values: tuple[object, ...]


class Mau1314Processor:
    """Processor shared by Mẫu 13/QT and Mẫu 14/QT."""

    REQUIRED_HEADERS = {
        "NGAY_HE_THONG",
        "MA_CN",
        "SO_HIEU_TK_SO_CAI",
        "MA_KH",
        "TEN_KH",
        "SO_TK_TIEN_GUI",
        "SO_SO_TIEN_GUI",
        "LOAI_TIEN_TE",
        "SO_DU_NGUYEN_TE_TAI_3112",
        "SO_DU_QUY_DOI_TAI_3112",
        "LOAI_KY_HAN",
        "NGAY_GUI",
        "NGAY_DEN_HAN",
        "LAI_SUAT_DC(NAM)",
        "NGAY_TRA_LAI_CUOI_CUNG",
        "SO_LAI_TRA_TRUOC_TU_3112",
        "SO_LAI_CON_PHAI_TRA_DEN_HET_3112",
        "DPTPCD",
        "ACCTSEQ",
        "TRCD",
        "TK_LAI_DU_CHI",
        "LOAI_KH",
        "TEN_LOAI_KH",
    }
    EXTRA_HEADERS = (
        "DPTPCD",
        "ACCTSEQ",
        "TRCD",
        "LOAI_KH",
        "NGAY_HE_THONG",
        "MA_CN",
        "TEN_LOAI_KH",
    )
    HEADER_ALIASES = {
        "SO_DU_NGUYEN_TE_TAI_3112": ("SO_DU_NGUYEN_TE_3112",),
        "SO_DU_QUY_DOI_TAI_3112": ("SO_DU_QUY_DOI_3112",),
        "LAI_SUAT_DC(NAM)": ("LAI_SUAT_DC",),
        "SO_LAI_TRA_TRUOC_TU_3112": ("SO_TRA_LAI_TRUOC_TU_3112",),
    }
    MAU14_OPTIONAL_HEADERS = {
        "TK_LAI_DU_CHI",
        "LOAI_KH",
        "TEN_LOAI_KH",
    }

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError(
                f"Mẫu {request.spec.report_code}/QT cần đúng một file CSV nguồn.",
                code="invalid_source_count",
            )
        source_path = request.source_paths[0]
        records, report_date, currency_order = self.read_source(request, source_path)
        workbook = self.build_workbook(
            request,
            records,
            report_date,
            currency_order,
        )
        processed_rows = len(records)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}{self._output_prefix(request.options)}{request.spec.report_code}.xlsx"
        )
        try:
            workbook.save(output_path)
        finally:
            workbook.close()
        del workbook
        del records
        gc.collect()
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name=request.spec.report_code,
            processed_rows=processed_rows,
        )

    def read_source(
        self,
        request: SettlementRequest,
        source_path: Path,
    ) -> tuple[list[DepositRecord], str, tuple[str, ...]]:
        if source_path.suffix.casefold() not in {".csv", ".txt"}:
            raise SettlementError(
                f"Nguồn Mẫu {request.spec.report_code}/QT phải là file CSV.",
                code="invalid_source_type",
            )
        report_code = self._report_code(request)
        branch_code = request.profile.branch_code.strip()
        records: list[DepositRecord] = []
        currency_order: list[str] = []
        report_date = ""
        try:
            with source_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                headers = set(reader.fieldnames or ())
                normalized_headers = self._normalized_headers(headers)
                required_headers = (
                    self.REQUIRED_HEADERS
                    if report_code == "13"
                    else self.REQUIRED_HEADERS - self.MAU14_OPTIONAL_HEADERS
                )
                missing = sorted(required_headers - normalized_headers)
                if missing:
                    raise SettlementError(
                        f"File nguồn thiếu cột: {', '.join(missing)}",
                        code="invalid_mau1314_headers",
                    )
                for row in reader:
                    row = self._normalize_row(row)
                    account = self._clean_code(row["SO_HIEU_TK_SO_CAI"])
                    if report_code == "13" and account[:2] in {"40", "41"}:
                        continue
                    if not report_date:
                        report_date = vietnamese_report_date(row["NGAY_HE_THONG"])
                    currency = row["LOAI_TIEN_TE"].strip()
                    if currency and currency not in currency_order:
                        currency_order.append(currency)
                    records.append(
                        self._row_to_record(
                            row,
                            self._row_branch_code(row, branch_code),
                            request.options,
                        )
                    )
        except SettlementError:
            raise
        except (OSError, UnicodeError, csv.Error) as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc

        if report_code == "13":
            records.sort(
                key=lambda item: (
                    item.ledger_account,
                    item.currency,
                    item.customer_id,
                )
            )
        else:
            records.sort(
                key=lambda item: (
                    item.customer_id,
                    item.ledger_account,
                    item.currency,
                )
            )
        return records, report_date, tuple(currency_order)

    def build_workbook(
        self,
        request: SettlementRequest,
        records: list[DepositRecord],
        report_date: str,
        currency_order: tuple[str, ...] | None = None,
    ) -> Workbook:
        report_code = self._report_code(request)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = report_code
        self._write_header(sheet, request, report_code, report_date)
        if report_code == "13":
            final_row, summary_rows = self._write_mau13_rows(
                sheet,
                records,
                currency_order or self._currency_order(records),
                keep_unused=not request.options.remove_unused_columns,
            )
        else:
            final_row = self._write_mau14_rows(
                sheet,
                records,
                keep_unused=not request.options.remove_unused_columns,
            )
            summary_rows = set()
            if request.options.remove_unused_columns:
                sheet.delete_cols(16, 1)
        self._write_signatures_and_notes(
            sheet,
            request,
            report_code,
            report_date,
            final_row,
        )
        if not request.options.remove_unused_columns:
            self._write_extra_headers(sheet)
        self._format_report(
            sheet,
            request.options,
            report_code,
            final_row,
            summary_rows,
        )
        if request.options.create_control_sheet and report_code == "13":
            self._write_control_sheet(workbook, request.options, records)
        workbook.calculation.fullCalcOnLoad = False
        workbook.calculation.forceFullCalc = False
        workbook.calculation.calcMode = "auto"
        return workbook

    def save_mau13_streaming_workbook(
        self,
        request: SettlementRequest,
        records: list[DepositRecord],
        report_date: str,
        currency_order: tuple[str, ...],
        output_path: Path,
    ) -> None:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("SoLieu_Mau13")
        final_row = self._write_mau13_streaming_sheet(
            sheet,
            request,
            records,
            report_date,
            currency_order,
        )
        if request.options.create_control_sheet:
            self._write_consolidation_mau13_summary_sheet(
                workbook,
                request.options,
                records,
            )
        self._write_style_reference_sheet(workbook)
        workbook.save(output_path)
        workbook.close()
        self._apply_mau13_streaming_xml_styles(output_path, final_row)

    def _write_mau13_streaming_sheet(
        self,
        sheet,
        request: SettlementRequest,
        records: list[DepositRecord],
        report_date: str,
        currency_order: tuple[str, ...],
    ) -> int:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        final_column = 16
        styles = self._streaming_styles()
        widths = {
            "A": 9,
            "B": 15,
            "C": 28,
            "D": 14.5,
            "E": 20,
            "F": 6.5,
            "G": 20,
            "H": 20,
            "I": 7,
            "J": 12,
            "K": 12,
            "L": 7,
            "M": 12,
            "N": 12,
            "O": 15,
            "P": 11,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        def cell(value=None, style_name: str = "body"):
            item = WriteOnlyCell(sheet, value=value)
            style = styles[style_name]
            item.font = style["font"]
            item.alignment = style["alignment"]
            if style["border"] is not None:
                item.border = style["border"]
            if style["fill"] is not None:
                item.fill = style["fill"]
            if style["number_format"] is not None:
                item.number_format = style["number_format"]
            return item

        def append(values: tuple[object, ...], style_name: str = "body") -> None:
            sheet.append(
                [
                    cell(values[index] if index < len(values) else None, style_name)
                    for index in range(final_column)
                ]
            )

        def append_title_row(left_value: str, right_value: str) -> None:
            row = [cell(None) for _ in range(final_column)]
            row[0] = cell(left_value, "agency")
            row[12] = cell(right_value, "instruction")
            sheet.append(row)

        append_title_row("NGÂN HÀNG NÔNG NGHIỆP", "1. Mẫu số 13/QT")
        append_title_row(
            "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "2. CN loại I gửi file về TSC",
        )
        append_title_row(
            f"Mã chi nhánh: {profile.branch_code.strip()}",
            "3. Lưu tại Chi nhánh",
        )
        append((f"Tên {branch_name}",), "agency")
        append(("SAO KÊ CHI TIẾT TIỀN GỬI KHÁCH HÀNG",), "title")
        append((
            "(TIỀN GỬI THANH TOÁN, TIỀN GỬI CÓ KỲ HẠN, "
            "TIỀN GỬI TIẾT KIỆM, KỲ PHIẾU, TRÁI PHIẾU)",
        ), "subtitle")
        append((report_date,), "date_title")
        unit_row = [cell(None) for _ in range(final_column)]
        unit_row[14] = cell("Đơn vị: VNĐ", "italic_left")
        sheet.append(unit_row)
        headers = (
            "SỐ HIỆU TÀI KHOẢN",
            "MÃ KHÁCH HÀNG",
            "TÊN KHÁCH HÀNG",
            "SỐ TÀI KHOẢN KHÁCH HÀNG",
            "SỐ SỔ TIẾT KIỆM",
            "LOẠI TIỀN TỆ",
            "SỐ DƯ NGUYÊN TỆ TẠI THỜI ĐIỂM 31/12",
            "SỐ DƯ QUY ĐỔI VNĐ TẠI THỜI ĐIỂM 31/12",
            "KỲ HẠN (tháng)",
            "NGÀY GỬI",
            "NGÀY ĐÁO HẠN",
            "LÃI SUẤT (%/năm)",
            "NGÀY TRẢ LÃI CUỐI CÙNG",
            "SỐ LÃI TRẢ TRƯỚC ĐẾN 31/12",
            "SỐ LÃI CÒN PHẢI TRẢ ĐẾN 31/12",
            "TÀI KHOẢN LÃI DỰ CHI",
        )
        append(headers, "header")
        append(tuple(range(1, final_column + 1)), "header")

        row_number = 11
        grouped: OrderedDict[str, OrderedDict[str, list[DepositRecord]]] = OrderedDict()
        for record in records:
            grouped.setdefault(record.ledger_account, OrderedDict()).setdefault(
                record.currency,
                [],
            ).append(record)

        currency_subtotals: dict[str, list[int]] = defaultdict(list)
        all_currency_subtotals: list[int] = []
        for account, account_groups in grouped.items():
            account_subtotals: list[int] = []
            for currency, group_records in account_groups.items():
                start_row = row_number
                for record in group_records:
                    self._append_streaming_record(sheet, record, cell)
                    row_number += 1
                subtotal_row = row_number
                sheet.append(
                    self._streaming_summary_row(
                        sheet,
                        cell,
                        f"     Cộng theo loại tiền: {currency}",
                        subtotal_row,
                        start_row,
                        row_number - 1,
                        (7, 8, 14, 15),
                    )
                )
                account_subtotals.append(subtotal_row)
                currency_subtotals[currency].append(subtotal_row)
                all_currency_subtotals.append(subtotal_row)
                row_number += 1
            account_row = row_number
            sheet.append(
                self._streaming_sum_rows(
                    sheet,
                    cell,
                    f"Cộng TK: {account}",
                    account_subtotals,
                    (8, 14, 15),
                )
            )
            row_number += 1

        append(tuple(), "body")
        row_number += 1
        for currency in currency_order:
            subtotal_rows = currency_subtotals.get(currency)
            if not subtotal_rows:
                continue
            sheet.append(
                self._streaming_sum_rows(
                    sheet,
                    cell,
                    f"TỔNG CỘNG THEO LOẠI TIỀN: {currency}",
                    subtotal_rows,
                    (7, 8, 14, 15),
                )
            )
            row_number += 1
        sheet.append(
            self._streaming_sum_rows(
                sheet,
                cell,
                "TỔNG CỘNG:",
                all_currency_subtotals,
                (8, 14, 15),
            )
        )
        final_row = row_number

        location = profile.report_location.strip()
        append((
            None, None, None, None, None, None, None, None, None, None,
            f"{location}, {report_date}" if location else report_date,
        ), "italic")
        append((
            "LẬP BIỂU", None, None, None,
            "TRƯỞNG PHÒNG KẾ TOÁN", None, None, None, None, None,
            "GIÁM ĐỐC",
        ), "signature")
        append((
            "(Ký, ghi rõ họ tên, số ĐT liên hệ)", None, None, None,
            "(Ký, ghi rõ họ tên)", None, None, None, None, None,
            "(Ký, đóng dấu, ghi rõ họ tên)",
        ), "italic")
        for _ in range(5):
            append(tuple(), "body")
        append((f"({profile.report_preparer} - SĐT: {profile.phone})",), "italic_left")
        append(tuple(), "body")
        append(("Ghi chú:",), "note_title")
        for note in (
            "1. Sao kê đối với TK 42; 43",
            "2. Cột 8 được quy đổi từ cột 7 theo tỷ giá TSC thông báo ngày 31/12",
            "3. Cột 7, cột 8: Dòng tổng cộng khớp với tổng số dư Có của các tài khoản 42, 43.",
            "4. Cột 14: Dòng tổng cộng khớp với dòng lãi huy động trả trước trên Sao kê chi tiết tài khoản 388 - Mẫu số 22/QT",
            "5. Cột 15: Dòng tổng cộng + dự chi lãi tiền gửi (đi vay) của các TCTD khác (cột 11, mẫu 12/QT) + dự chi lãi vay TSC (cột 12, mẫu 26/QT) phải khớp với tổng số dư Có của tài khoản 491, 492; 493",
            "6. Cột 10, 11 & 13: ghi theo thứ tự ngày/tháng/năm (không ghi tháng trước ngày sau)",
            "7. Dòng tổng cộng tính bằng công thức, không nhập số liệu trực tiếp",
        ):
            append((note,), "body")

        sheet.freeze_panes = "B11"
        sheet.page_setup.paperSize = 9
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 0
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = False
        sheet.print_area = f"A1:P{final_row + 18}"
        sheet.print_title_rows = "$9:$10"
        sheet.print_title_cols = "$A:$P"
        sheet.page_setup.scale = 62
        sheet.print_options.horizontalCentered = True
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        sheet.page_margins.header = 0.25
        sheet.page_margins.footer = 0.25
        sheet.HeaderFooter.differentFirst = True
        sheet.oddHeader.center.text = "&P/&N"
        sheet.firstHeader.center.text = ""
        return final_row

    def _append_streaming_record(self, sheet, record: DepositRecord, cell) -> None:
        sheet.append((
            self._numeric_code(record.ledger_account),
            record.customer_id,
            record.customer_name,
            record.deposit_account or None,
            record.savings_book or None,
            record.currency,
            self._number_to_cell(record.original_balance),
            self._number_to_cell(record.converted_balance),
            record.term_months or None,
            record.deposit_date,
            record.maturity_date,
            record.interest_rate,
            record.last_interest_date,
            record.prepaid_interest_display,
            record.accrued_interest_display,
            record.interest_account or None,
        ))

    @staticmethod
    def _streaming_summary_row(
        sheet,
        cell,
        label: str,
        row_number: int,
        start_row: int,
        end_row: int,
        sum_columns: tuple[int, ...],
    ) -> list[WriteOnlyCell]:
        row = []
        for column in range(1, 17):
            value = label if column == 1 else None
            if column in sum_columns:
                letter = get_column_letter(column)
                value = f"=SUM({letter}{start_row}:{letter}{end_row})"
            row.append(cell(value, "summary_number" if column in sum_columns else "summary"))
        return row

    @staticmethod
    def _streaming_sum_rows(
        sheet,
        cell,
        label: str,
        rows: list[int],
        sum_columns: tuple[int, ...],
    ) -> list[WriteOnlyCell]:
        output = []
        for column in range(1, 17):
            value = label if column == 1 else None
            if column in sum_columns:
                letter = get_column_letter(column)
                value = "=" + "+".join(f"{letter}{row}" for row in rows) if rows else "=0"
            output.append(cell(value, "summary_number" if column in sum_columns else "summary"))
        return output

    @staticmethod
    def _streaming_styles() -> dict[str, dict[str, object]]:
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return {
            "body": {
                "font": Font(name="Times New Roman", size=11),
                "alignment": Alignment(vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "text": {
                "font": Font(name="Times New Roman", size=11),
                "alignment": Alignment(vertical="center"),
                "border": None,
                "fill": None,
                "number_format": "@",
            },
            "number": {
                "font": Font(name="Times New Roman", size=11),
                "alignment": Alignment(vertical="center"),
                "border": None,
                "fill": None,
                "number_format": "#,##0",
            },
            "date": {
                "font": Font(name="Times New Roman", size=11),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": "dd/mm/yyyy",
            },
            "header": {
                "font": Font(name="Times New Roman", size=11, bold=True),
                "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
                "border": border,
                "fill": PatternFill("solid", fgColor="FFFFFF"),
                "number_format": None,
            },
            "agency": {
                "font": Font(name="Times New Roman", size=11, bold=True),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "instruction": {
                "font": Font(name="Times New Roman", size=11),
                "alignment": Alignment(horizontal="left", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "title": {
                "font": Font(name="Times New Roman", size=18, bold=True),
                "alignment": Alignment(horizontal="centerContinuous", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "subtitle": {
                "font": Font(name="Times New Roman", size=16, bold=True),
                "alignment": Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "date_title": {
                "font": Font(name="Times New Roman", size=14, bold=True, italic=True),
                "alignment": Alignment(horizontal="centerContinuous", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "summary": {
                "font": Font(name="Times New Roman", size=11, bold=True),
                "alignment": Alignment(vertical="center"),
                "border": border,
                "fill": None,
                "number_format": None,
            },
            "summary_number": {
                "font": Font(name="Times New Roman", size=11, bold=True),
                "alignment": Alignment(vertical="center"),
                "border": border,
                "fill": None,
                "number_format": "#,##0",
            },
            "signature": {
                "font": Font(name="Times New Roman", size=11, bold=True),
                "alignment": Alignment(horizontal="centerContinuous", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "italic": {
                "font": Font(name="Times New Roman", size=11, italic=True),
                "alignment": Alignment(horizontal="centerContinuous", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "italic_left": {
                "font": Font(name="Times New Roman", size=11, italic=True),
                "alignment": Alignment(horizontal="left", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "note_title": {
                "font": Font(name="Times New Roman", size=11, bold=True),
                "alignment": Alignment(horizontal="left", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
        }

    def _write_style_reference_sheet(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("_styles")
        sheet.sheet_state = "hidden"
        thin = Side(style="thin", color="000000")
        qt_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_cell(value=None, *, number_format=None, alignment=None):
            cell = WriteOnlyCell(sheet, value=value)
            cell.font = Font(name="Times New Roman", size=11)
            cell.border = qt_border
            cell.alignment = alignment or Alignment(vertical="center")
            if number_format is not None:
                cell.number_format = number_format
            return cell

        sheet.append((
            style_cell("body"),
            style_cell("text", number_format="@"),
            style_cell(123, number_format="#,##0"),
            style_cell(date(2025, 12, 31), number_format="dd/mm/yyyy"),
        ))

    @staticmethod
    def _apply_mau13_streaming_xml_styles(output_path: Path, final_row: int) -> None:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_path = Path(tmp.name)
        style_ids: dict[str, str] = {}
        with ZipFile(output_path, "r") as source_zip:
            style_xml = ""
            for sheet_path in ("xl/worksheets/sheet3.xml", "xl/worksheets/sheet2.xml"):
                try:
                    candidate = source_zip.read(sheet_path).decode("utf-8")
                except KeyError:
                    continue
                if "body" in candidate and "text" in candidate:
                    style_xml = candidate
                    break
            for column, name in (("A", "body"), ("B", "text"), ("C", "number"), ("D", "date")):
                match = re.search(rf'<c r="{column}1"[^>]* s="([^"]+)"', style_xml)
                if match:
                    style_ids[name] = match.group(1)
            if set(style_ids) != {"body", "text", "number", "date"}:
                shutil.copyfile(output_path, temp_path)
            else:
                with ZipFile(temp_path, "w", ZIP_DEFLATED) as target_zip:
                    for info in source_zip.infolist():
                        data = source_zip.read(info.filename)
                        if info.filename == "xl/worksheets/sheet1.xml":
                            xml = data.decode("utf-8")
                            xml = Mau1314Processor._style_mau13_sheet_xml(
                                xml,
                                final_row,
                                style_ids,
                            )
                            data = xml.encode("utf-8")
                        target_zip.writestr(info, data)
        shutil.move(str(temp_path), output_path)

    @staticmethod
    def _style_mau13_sheet_xml(
        xml: str,
        final_row: int,
        style_ids: dict[str, str],
    ) -> str:
        text_columns = {"B", "D", "E", "I", "P"}
        number_columns = {"G", "H", "N", "O"}
        date_columns = {"J", "K", "M"}

        def replacement(match: re.Match[str]) -> str:
            column = match.group(1)
            row = int(match.group(2))
            attrs = match.group(3)
            if row < 11 or row > final_row:
                return match.group(0)
            if column in number_columns:
                style = style_ids["number"]
            elif column in date_columns:
                style = style_ids["date"]
            elif column in text_columns:
                style = style_ids["text"]
            else:
                style = style_ids["body"]
            if ' s="' in attrs:
                if column not in date_columns:
                    return match.group(0)
                attrs = re.sub(r' s="[^"]+"', "", attrs)
            return f'<c r="{column}{row}" s="{style}"{attrs}>'

        xml = re.sub(r'<c r="([A-P])(\d+)"([^>]*)>', replacement, xml)
        xml = Mau1314Processor._fill_missing_mau13_cells(
            xml,
            final_row,
            style_ids["body"],
        )
        refs = [
            "A1:D1",
            "A2:D2",
            "A3:D3",
            "A4:D4",
            "A5:O5",
            "A6:O6",
            "A7:O7",
            f"K{final_row + 1}:P{final_row + 1}",
            f"A{final_row + 2}:D{final_row + 2}",
            f"E{final_row + 2}:J{final_row + 2}",
            f"K{final_row + 2}:P{final_row + 2}",
            f"A{final_row + 3}:D{final_row + 3}",
            f"E{final_row + 3}:J{final_row + 3}",
            f"K{final_row + 3}:P{final_row + 3}",
            f"A{final_row + 9}:D{final_row + 9}",
        ]
        merge_xml = Mau1314Processor._merge_refs_xml(refs)
        if "<mergeCells" in xml:
            xml = re.sub(r"<mergeCells[^>]*>.*?</mergeCells>", merge_xml, xml, flags=re.S)
        else:
            xml = xml.replace("</sheetData>", f"</sheetData>{merge_xml}", 1)
        return xml

    @staticmethod
    def _fill_missing_mau13_cells(xml: str, final_row: int, body_style: str) -> str:
        columns = tuple("ABCDEFGHIJKLMNOP")

        def row_replacement(match: re.Match[str]) -> str:
            row_number = int(match.group(1))
            row_xml = match.group(0)
            if row_number < 11 or row_number > final_row:
                return row_xml
            cells_by_column = {
                cell_match.group(1): cell_match.group(0)
                for cell_match in re.finditer(
                    r'<c\b(?=[^>]* r="([A-P])'
                    + str(row_number)
                    + r'")(?:(?:[^>]*?/>)|(?:[^>]*>.*?</c>))',
                    row_xml,
                    flags=re.S,
                )
            }
            ordered_cells = "".join(
                cells_by_column.get(
                    column,
                    f'<c r="{column}{row_number}" s="{body_style}"/>',
                )
                for column in columns
            )
            return re.sub(
                r'(<row[^>]* r="' + str(row_number) + r'"[^>]*>).*?(</row>)',
                lambda row_match: f"{row_match.group(1)}{ordered_cells}{row_match.group(2)}",
                row_xml,
                count=1,
                flags=re.S,
            )

        return re.sub(
            r'<row[^>]* r="(\d+)"[^>]*>.*?</row>',
            row_replacement,
            xml,
            flags=re.S,
        )

    @staticmethod
    def _merge_refs_xml(refs: list[str]) -> str:
        merge_cells = "".join(f'<mergeCell ref="{ref}"/>' for ref in refs)
        return f'<mergeCells count="{len(refs)}">{merge_cells}</mergeCells>'

    def _write_consolidation_mau13_summary_sheet(
        self,
        workbook: Workbook,
        options: SettlementOptions,
        records: list[DepositRecord],
    ) -> None:
        sheet = workbook.create_sheet("TongHop_Mau13")
        branch_order = tuple(
            dict.fromkeys(self._record_branch(record, options) for record in records)
        )
        normal: OrderedDict[str, dict[str, Decimal]] = OrderedDict()
        interest: OrderedDict[str, dict[str, Decimal]] = OrderedDict()
        for record in records:
            key = f"{record.ledger_account}_{record.currency}"
            branch = self._record_branch(record, options)
            normal.setdefault(key, {}).setdefault(branch, Decimal(0))
            normal[key][branch] += record.converted_balance
            if record.interest_account:
                interest.setdefault(record.interest_account, {}).setdefault(
                    branch,
                    Decimal(0),
                )
                interest[record.interest_account][branch] += (
                    record.prepaid_interest
                    if record.interest_account.startswith("388")
                    else record.accrued_interest
                )

        styles = self._summary_streaming_styles()

        def cell(value=None, style_name: str = "body"):
            item = WriteOnlyCell(sheet, value=value)
            style = styles[style_name]
            item.font = style["font"]
            item.alignment = style["alignment"]
            if style["border"] is not None:
                item.border = style["border"]
            if style["fill"] is not None:
                item.fill = style["fill"]
            if style["number_format"] is not None:
                item.number_format = style["number_format"]
            return item

        def append_row(values: list[object], style_name: str = "body") -> None:
            sheet.append([cell(value, style_name) for value in values])

        total_label = "Cộng chi nhánh"
        header = [" Mã Chi nhánh", *branch_order, total_label]
        sheet.column_dimensions["A"].width = 18
        for column in range(2, len(header) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 22
        sheet.page_setup.paperSize = 9
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_options.horizontalCentered = True
        append_row(["Báo cáo tổng hợp số liệu quyết toán mẫu 13/QT"], "title")
        append_row(header, "header")
        for key in sorted(normal):
            branch_values = [
                normal[key].get(branch, Decimal(0)) for branch in branch_order
            ]
            append_row(
                [
                    key,
                    *[self._number_to_cell(value) for value in branch_values],
                    self._number_to_cell(sum(branch_values, Decimal(0))),
                ],
                "number",
            )
        totals = [
            sum((normal[key].get(branch, Decimal(0)) for key in normal), Decimal(0))
            for branch in branch_order
        ]
        append_row(
            [
                "Tổng cộng",
                *[self._number_to_cell(value) for value in totals],
                self._number_to_cell(sum(totals, Decimal(0))),
            ],
            "total",
        )

        append_row([])
        append_row(["Số liệu dự chi đến 31/12 mẫu 13/QT"], "title")
        append_row(header, "header")
        if options.include_accrual_accounts:
            for account in sorted(interest, reverse=True):
                branch_values = [
                    interest[account].get(branch, Decimal(0))
                    for branch in branch_order
                ]
                append_row(
                    [
                        self._numeric_code(account),
                        *[self._number_to_cell(value) for value in branch_values],
                        self._number_to_cell(sum(branch_values, Decimal(0))),
                    ],
                    "number",
                )
            interest_totals = [
                sum(
                    (
                        interest[account].get(branch, Decimal(0))
                        for account in interest
                        if not account.startswith("388")
                    ),
                    Decimal(0),
                )
                for branch in branch_order
            ]
            append_row(
                [
                    "Tổng cộng",
                    *[self._number_to_cell(value) for value in interest_totals],
                    self._number_to_cell(sum(interest_totals, Decimal(0))),
                ],
                "total",
            )

    @staticmethod
    def _summary_streaming_styles() -> dict[str, dict[str, object]]:
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        total_border = Border(left=thin, right=thin, top=medium, bottom=medium)
        return {
            "body": {
                "font": Font(name="Times New Roman", size=12),
                "alignment": Alignment(vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "title": {
                "font": Font(name="Times New Roman", size=12, bold=True, color="000080"),
                "alignment": Alignment(horizontal="left", vertical="center"),
                "border": None,
                "fill": None,
                "number_format": None,
            },
            "header": {
                "font": Font(name="Times New Roman", size=12, bold=True),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": border,
                "fill": PatternFill("solid", fgColor="FFFF00"),
                "number_format": None,
            },
            "number": {
                "font": Font(name="Times New Roman", size=12),
                "alignment": Alignment(vertical="center"),
                "border": border,
                "fill": None,
                "number_format": "#,##0",
            },
            "total": {
                "font": Font(name="Times New Roman", size=12, bold=True),
                "alignment": Alignment(vertical="center"),
                "border": total_border,
                "fill": None,
                "number_format": "#,##0",
            },
        }

    @staticmethod
    def _record_branch(record: DepositRecord, options: SettlementOptions) -> str:
        if options.include_branch_in_customer_id and len(record.customer_id) >= 4:
            return record.customer_id[:4]
        if len(record.extra_values) >= 6 and str(record.extra_values[5]).strip():
            return str(record.extra_values[5]).strip()
        return ""

    def _row_to_record(
        self,
        row: dict[str, str],
        branch_code: str,
        options: SettlementOptions,
    ) -> DepositRecord:
        prepaid_text = row["SO_LAI_TRA_TRUOC_TU_3112"]
        accrued_text = row["SO_LAI_CON_PHAI_TRA_DEN_HET_3112"]
        return DepositRecord(
            ledger_account=self._clean_code(row["SO_HIEU_TK_SO_CAI"]),
            customer_id=normalize_customer_id(
                row["MA_KH"],
                branch_code,
                include_branch=options.include_branch_in_customer_id,
            ),
            customer_name=self._clean_text(row["TEN_KH"]),
            deposit_account=self._clean_code(row["SO_TK_TIEN_GUI"]),
            savings_book=self._clean_code(row["SO_SO_TIEN_GUI"]),
            currency=self._clean_text(row["LOAI_TIEN_TE"]),
            original_balance=self._parse_source_amount(
                row["SO_DU_NGUYEN_TE_TAI_3112"]
            ),
            converted_balance=self._parse_source_amount(
                row["SO_DU_QUY_DOI_TAI_3112"]
            ),
            term_months=self._clean_code(row["LOAI_KY_HAN"]),
            deposit_date=parse_yyyymmdd(row["NGAY_GUI"]),
            maturity_date=parse_yyyymmdd(row["NGAY_DEN_HAN"]),
            interest_rate=self._number_or_text(row["LAI_SUAT_DC(NAM)"]),
            last_interest_date=parse_yyyymmdd(row["NGAY_TRA_LAI_CUOI_CUNG"]),
            prepaid_interest=self._parse_source_amount(prepaid_text),
            accrued_interest=self._parse_source_amount(accrued_text),
            interest_account=self._clean_code(row.get("TK_LAI_DU_CHI", "")),
            prepaid_interest_display=self._amount_display(prepaid_text),
            accrued_interest_display=self._amount_display(accrued_text),
            extra_values=tuple(
                self._clean_text(row.get(header, ""))
                for header in self.EXTRA_HEADERS
            ),
        )

    @staticmethod
    def _row_branch_code(row: dict[str, str], fallback_branch_code: str) -> str:
        return row.get("MA_CN", "").strip().lstrip("'").strip() or fallback_branch_code

    def _write_header(
        self,
        sheet,
        request: SettlementRequest,
        report_code: str,
        report_date: str,
    ) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        for address in ("A1:D1", "A2:D2", "A3:D3", "A4:D4"):
            sheet.merge_cells(address)
        sheet["A1"] = "NGÂN HÀNG NÔNG NGHIỆP"
        sheet["A2"] = "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM"
        sheet["A3"] = f"Mã chi nhánh: {profile.branch_code.strip()}"
        sheet["A4"] = f"Tên {branch_name}"
        style_agency_header(sheet, start_column=1, end_column=4)

        for address in ("A5:O5", "A6:O6", "A7:O7"):
            sheet.merge_cells(address)
        if report_code == "13":
            sheet["A5"] = "SAO KÊ CHI TIẾT TIỀN GỬI KHÁCH HÀNG"
            sheet["A6"] = (
                "(TIỀN GỬI THANH TOÁN, TIỀN GỬI CÓ KỲ HẠN, "
                "TIỀN GỬI TIẾT KIỆM, KỲ PHIẾU, TRÁI PHIẾU)"
            )
        else:
            sheet["A5"] = (
                "SAO KÊ CHI TIẾT SỐ DƯ TIỀN GỬI, TIẾT KIỆM, "
                "KỲ PHIẾU, TRÁI PHIẾU"
            )
            sheet["A6"] = "CÁC KHÁCH HÀNG CÓ SỐ DƯ TỪ 10 TỶ VNĐ TRỞ LÊN"
        sheet["A7"] = report_date
        sheet["M1"] = f"1. Mẫu số {report_code}/QT"
        sheet["M2"] = "2. CN loại I gửi file về TSC"
        sheet["M3"] = "3. Lưu tại Chi nhánh"
        sheet["O8"] = "Đơn vị: VNĐ"
        style_currency_unit(sheet["O8"])

        headers = (
            "SỐ HIỆU TÀI KHOẢN",
            "MÃ KHÁCH HÀNG",
            "TÊN KHÁCH HÀNG",
            (
                "SỐ TÀI KHOẢN KHÁCH HÀNG"
                if report_code == "13"
                else "SỐ TÀI KHOẢN TIỀN GỬI CỦA KHÁCH HÀNG"
            ),
            "SỐ SỔ TIẾT KIỆM",
            "LOẠI TIỀN TỆ",
            "SỐ DƯ NGUYÊN TỆ TẠI THỜI ĐIỂM 31/12",
            "SỐ DƯ QUY ĐỔI VNĐ TẠI THỜI ĐIỂM 31/12",
            "KỲ HẠN (tháng)",
            "NGÀY GỬI",
            "NGÀY ĐÁO HẠN",
            "LÃI SUẤT (%/năm)",
            "NGÀY TRẢ LÃI CUỐI CÙNG",
            "SỐ LÃI TRẢ TRƯỚC ĐẾN 31/12",
            "SỐ LÃI CÒN PHẢI TRẢ ĐẾN 31/12",
            "TÀI KHOẢN LÃI DỰ CHI",
        )
        for column, value in enumerate(headers, start=1):
            sheet.cell(9, column).value = value
            sheet.cell(10, column).value = column
            sheet.cell(10, column).number_format = "(0)"

    def _write_mau13_rows(
        self,
        sheet,
        records: list[DepositRecord],
        currency_order: tuple[str, ...],
        *,
        keep_unused: bool,
    ) -> tuple[int, set[int]]:
        grouped: OrderedDict[str, OrderedDict[str, list[DepositRecord]]] = OrderedDict()
        for record in records:
            grouped.setdefault(record.ledger_account, OrderedDict()).setdefault(
                record.currency, []
            ).append(record)

        row = 11
        summary_rows: set[int] = set()
        currency_subtotals: dict[str, list[int]] = defaultdict(list)
        all_currency_subtotals: list[int] = []
        for account, account_groups in grouped.items():
            account_subtotals: list[int] = []
            for currency, group_records in account_groups.items():
                start_row = row
                for record in group_records:
                    self._write_record(sheet, row, record, keep_unused=keep_unused)
                    row += 1
                subtotal_row = row
                summary_rows.add(subtotal_row)
                sheet.cell(subtotal_row, 1).value = (
                    f"     Cộng theo loại tiền: {currency}"
                )
                for column in (7, 8, 14, 15):
                    letter = get_column_letter(column)
                    sheet.cell(subtotal_row, column).value = (
                        f"=SUM({letter}{start_row}:{letter}{row - 1})"
                    )
                account_subtotals.append(subtotal_row)
                currency_subtotals[currency].append(subtotal_row)
                all_currency_subtotals.append(subtotal_row)
                row += 1

            account_row = row
            summary_rows.add(account_row)
            sheet.cell(account_row, 1).value = f"Cộng TK: {account}"
            for column in (8, 14, 15):
                sheet.cell(account_row, column).value = self._sum_rows_formula(
                    get_column_letter(column),
                    account_subtotals,
                )
            row += 1

        row += 1
        for currency in currency_order:
            subtotal_rows = currency_subtotals.get(currency)
            if not subtotal_rows:
                continue
            summary_rows.add(row)
            sheet.cell(row, 1).value = f"TỔNG CỘNG THEO LOẠI TIỀN: {currency}"
            for column in (7, 8, 14, 15):
                sheet.cell(row, column).value = self._sum_rows_formula(
                    get_column_letter(column),
                    subtotal_rows,
                )
            row += 1

        summary_rows.add(row)
        sheet.cell(row, 1).value = "TỔNG CỘNG:"
        for column in (8, 14, 15):
            sheet.cell(row, column).value = self._sum_rows_formula(
                get_column_letter(column),
                all_currency_subtotals,
            )
        return row, summary_rows

    def _write_mau14_rows(
        self,
        sheet,
        records: list[DepositRecord],
        *,
        keep_unused: bool,
    ) -> int:
        grouped: OrderedDict[str, list[DepositRecord]] = OrderedDict()
        for record in records:
            grouped.setdefault(record.customer_id, []).append(record)
        row = 11
        for customer_id, customer_records in grouped.items():
            start_row = row
            for record in customer_records:
                self._write_record(sheet, row, record, keep_unused=keep_unused)
                row += 1
            sheet.cell(row, 1).value = f"   Cộng khách hàng: {customer_id}"
            for column in (8, 14, 15):
                letter = get_column_letter(column)
                sheet.cell(row, column).value = f"=SUM({letter}{start_row}:{letter}{row - 1})"
            row += 1
        return max(10, row - 1)

    def _write_record(
        self,
        sheet,
        row: int,
        record: DepositRecord,
        *,
        keep_unused: bool,
    ) -> None:
        values = (
            self._numeric_code(record.ledger_account),
            record.customer_id,
            record.customer_name,
            record.deposit_account or None,
            record.savings_book or None,
            record.currency,
            self._number_to_cell(record.original_balance),
            self._number_to_cell(record.converted_balance),
            record.term_months or None,
            record.deposit_date,
            record.maturity_date,
            record.interest_rate,
            record.last_interest_date,
            record.prepaid_interest_display,
            record.accrued_interest_display,
            record.interest_account or None,
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column).value = value
        if keep_unused:
            for column, value in enumerate(record.extra_values, start=17):
                sheet.cell(row, column).value = value or None

    def _write_signatures_and_notes(
        self,
        sheet,
        request: SettlementRequest,
        report_code: str,
        report_date: str,
        final_row: int,
    ) -> None:
        profile = request.profile
        signature_end_column = (
            15
            if report_code == "14" and request.options.remove_unused_columns
            else 16
        )
        date_row = final_row + 1
        title_row = final_row + 2
        note_row = final_row + 3
        preparer_row = final_row + 9
        notes_row = final_row + 11

        sheet.merge_cells(
            start_row=date_row,
            start_column=11,
            end_row=date_row,
            end_column=signature_end_column,
        )
        location = profile.report_location.strip()
        sheet.cell(date_row, 11).value = (
            f"{location}, {report_date}" if location else report_date
        )
        for start_column, end_column, title, note in (
            (1, 4, "LẬP BIỂU", "(Ký, ghi rõ họ tên, số ĐT liên hệ)"),
            (5, 11, "TRƯỞNG PHÒNG KẾ TOÁN", "(Ký, ghi rõ họ tên)"),
            (
                12,
                signature_end_column,
                "GIÁM ĐỐC",
                "(Ký, đóng dấu, ghi rõ họ tên)",
            ),
        ):
            sheet.merge_cells(
                start_row=title_row,
                start_column=start_column,
                end_row=title_row,
                end_column=end_column,
            )
            sheet.merge_cells(
                start_row=note_row,
                start_column=start_column,
                end_row=note_row,
                end_column=end_column,
            )
            sheet.cell(title_row, start_column).value = title
            sheet.cell(note_row, start_column).value = note
        sheet.merge_cells(
            start_row=preparer_row,
            start_column=1,
            end_row=preparer_row,
            end_column=4,
        )
        sheet.cell(preparer_row, 1).value = (
            f"({profile.report_preparer} - SĐT: {profile.phone})"
        )

        sheet.cell(notes_row, 1).value = "Ghi chú:"
        if report_code == "13":
            notes = (
                "1. Sao kê đối với TK 42; 43",
                "2. Cột 8 được quy đổi từ cột 7 theo tỷ giá TSC thông báo ngày 31/12",
                "3. Cột 7, cột 8: Dòng tổng cộng khớp với tổng số dư Có "
                "của các tài khoản 42, 43.",
                "4. Cột 14: Dòng tổng cộng khớp với dòng lãi huy động trả trước "
                "trên Sao kê chi tiết tài khoản 388 - Mẫu số 22/QT",
                "5. Cột 15: Dòng tổng cộng + dự chi lãi tiền gửi (đi vay) của "
                "các TCTD khác (cột 11, mẫu 12/QT) + dự chi lãi vay TSC "
                "(cột 12, mẫu 26/QT) phải khớp với tổng số dư Có của tài khoản "
                "491, 492; 493",
                "6. Cột 10, 11 & 13: ghi theo thứ tự ngày/tháng/năm "
                "(không ghi tháng trước ngày sau)",
                "7. Dòng tổng cộng tính bằng công thức, không nhập số liệu trực tiếp",
            )
            for offset, text in enumerate(notes, start=1):
                sheet.cell(notes_row + offset, 1).value = text
        else:
            sheet.cell(notes_row, 2).value = (
                "Cột 8 được quy đổi từ cột 7 theo tỷ giá TSC thông báo ngày 31/12"
            )

    def _write_control_sheet(
        self,
        workbook: Workbook,
        options: SettlementOptions,
        records: list[DepositRecord],
    ) -> None:
        sheet = workbook.create_sheet("SoLieuTongHop")
        normal: OrderedDict[str, list[DepositRecord]] = OrderedDict()
        for record in records:
            key = f"{record.ledger_account}_{record.currency}"
            normal.setdefault(key, []).append(record)

        interest_accounts: list[str] = []
        if options.include_accrual_accounts:
            interest_accounts = sorted(
                {
                    record.interest_account
                    for record in records
                    if record.interest_account
                }
            )

        sheet["A1"] = "SỐ LIỆU TỔNG HỢP MẪU 13"
        headers: list[object] = [
            "Tài Khoản",
            "Số Dư Nguyên Tệ",
            "Số Dư Qui Đổi VNĐ",
            "Số Lãi trả trước đến 31/12",
            "Số Lãi còn phải trả đến 31/12",
            *[self._numeric_code(account) for account in interest_accounts],
        ]
        for column, value in enumerate(headers, start=1):
            sheet.cell(2, column).value = value

        row = 3
        for key, key_records in normal.items():
            sheet.cell(row, 1).value = key
            sheet.cell(row, 2).value = self._number_to_cell(
                sum((item.original_balance for item in key_records), Decimal(0))
            )
            sheet.cell(row, 3).value = self._number_to_cell(
                sum((item.converted_balance for item in key_records), Decimal(0))
            )
            sheet.cell(row, 4).value = self._number_to_cell(
                sum((item.prepaid_interest for item in key_records), Decimal(0))
            )
            sheet.cell(row, 5).value = self._number_to_cell(
                sum((item.accrued_interest for item in key_records), Decimal(0))
            )
            for column, account in enumerate(interest_accounts, start=6):
                amount = sum(
                    (
                        item.prepaid_interest
                        if account.startswith("388")
                        else item.accrued_interest
                    )
                    for item in key_records
                    if item.interest_account == account
                )
                sheet.cell(row, column).value = self._number_to_cell(amount)
            row += 1
        normal_end_row = row - 1

        for account in interest_accounts:
            account_records = [
                item for item in records if item.interest_account == account
            ]
            sheet.cell(row, 1).value = self._numeric_code(account)
            if account.startswith("388"):
                sheet.cell(row, 4).value = self._number_to_cell(
                    sum((item.prepaid_interest for item in account_records), Decimal(0))
                )
                sheet.cell(row, 5).value = 0
            else:
                sheet.cell(row, 4).value = 0
                sheet.cell(row, 5).value = self._number_to_cell(
                    sum((item.accrued_interest for item in account_records), Decimal(0))
                )
            for column in range(6, 6 + len(interest_accounts)):
                sheet.cell(row, column).value = 0
            row += 1

        total_row = row
        sheet.cell(total_row, 1).value = "Tổng cộng:"
        for column in range(2, len(headers) + 1):
            amount = sum(
                (
                    Decimal(str(sheet.cell(data_row, column).value or 0))
                    for data_row in range(3, normal_end_row + 1)
                ),
                Decimal(0),
            )
            sheet.cell(total_row, column).value = self._number_to_cell(amount)
        self._format_control_sheet(sheet, total_row, len(headers))

    @staticmethod
    def _format_control_sheet(sheet, total_row: int, final_column: int) -> None:
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in sheet.iter_rows(
            min_row=2,
            max_row=total_row,
            min_col=1,
            max_col=final_column,
        ):
            for cell in row:
                cell.border = border
                cell.font = Font(name="Times New Roman", size=12)
                cell.alignment = Alignment(vertical="center")
        sheet["A1"].font = Font(
            name="Times New Roman",
            size=12,
            bold=True,
            color="000080",
        )
        for cell in sheet[2][:final_column]:
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.fill = PatternFill("solid", fgColor="FFFF00")
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        for cell in sheet[total_row][:final_column]:
            cell.font = Font(name="Times New Roman", size=12, bold=True)
        for row in sheet.iter_rows(
            min_row=3,
            max_row=total_row,
            min_col=2,
            max_col=final_column,
        ):
            for cell in row:
                cell.number_format = "#,##0"
        sheet.column_dimensions["A"].width = 16
        for column in range(2, final_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 18.5
        sheet.row_dimensions[2].height = 50
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:{get_column_letter(final_column)}{total_row}",
            orientation="portrait",
        )

    def _format_report(
        self,
        sheet,
        options: SettlementOptions,
        report_code: str,
        final_row: int,
        summary_rows: set[int],
    ) -> None:
        signature_end = final_row + 9
        notes_end = final_row + (18 if report_code == "13" else 11)
        final_column = 16 if report_code == "13" else (
            15 if options.remove_unused_columns else sheet.max_column
        )
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font = Font(name="Times New Roman", size=11)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=notes_end,
            min_col=1,
            max_col=max(final_column, sheet.max_column),
        ):
            for cell in row:
                cell.font = font
                cell.alignment = Alignment(vertical="center")

        style_agency_header(sheet, start_column=1, end_column=4)
        for row in range(1, 5):
            sheet.cell(row, 1).font = Font(
                name="Times New Roman",
                size=11,
                bold=True,
            )
            sheet.cell(row, 1).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
        sheet["A5"].font = Font(name="Times New Roman", size=18, bold=True)
        sheet["A6"].font = Font(name="Times New Roman", size=16, bold=True)
        sheet["A7"].font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        for address in ("A5", "A6", "A7"):
            sheet[address].alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=(address == "A6"),
            )
        sheet.row_dimensions[5].height = 24
        sheet.row_dimensions[6].height = 30
        sheet.row_dimensions[7].height = 22

        for row in range(9, 11):
            for column in range(1, final_column + 1):
                cell = sheet.cell(row, column)
                cell.font = Font(name="Times New Roman", size=11, bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = Border(
                    left=medium if column == 1 else thin,
                    right=medium if column == final_column else thin,
                    top=medium if row == 9 else thin,
                    bottom=medium if row == 10 else thin,
                )
        sheet.row_dimensions[9].height = 76
        sheet.row_dimensions[10].height = 18

        for row in range(11, final_row + 1):
            is_summary = row in summary_rows or (
                report_code == "14"
                and str(sheet.cell(row, 1).value or "").strip().startswith("Cộng khách hàng:")
            )
            for column in range(1, final_column + 1):
                cell = sheet.cell(row, column)
                cell.border = body_border
                if is_summary:
                    cell.font = Font(name="Times New Roman", size=11, bold=True)
            for column in (2, 6, 10, 11, 13):
                if column <= final_column:
                    sheet.cell(row, column).alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                    )
            for column in (7, 8, 14, 15):
                if column <= final_column:
                    sheet.cell(row, column).number_format = "#,##0"
            for column in (2, 4, 5, 9, 16):
                if column <= final_column:
                    sheet.cell(row, column).number_format = "@"
            for column in (10, 11, 13):
                if column <= final_column:
                    sheet.cell(row, column).number_format = (
                        "dd/mm/yyyy" if options.four_digit_year else "dd/mm/yy"
                    )
            if report_code in {"13", "14"} and not is_summary:
                sheet.cell(row, 3).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )
                name_length = len(str(sheet.cell(row, 3).value or ""))
                if name_length > 28:
                    sheet.row_dimensions[row].height = min(
                        45,
                        15 * ((name_length // 28) + 1),
                    )
            if report_code == "13" and not is_summary:
                sheet.cell(row, 5).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False,
                )
            if not is_summary and sheet.cell(row, 6).value != "VND":
                sheet.cell(row, 7).number_format = "#,##0.00"

        date_row = final_row + 1
        title_row = final_row + 2
        note_row = final_row + 3
        for row in range(date_row, signature_end + 1):
            for column in range(1, final_column + 1):
                sheet.cell(row, column).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
        sheet.cell(date_row, 11).font = Font(
            name="Times New Roman",
            size=11,
            italic=True,
        )
        for column in (1, 5, 12):
            sheet.cell(title_row, column).font = Font(
                name="Times New Roman",
                size=11,
                bold=True,
            )
            sheet.cell(note_row, column).font = Font(
                name="Times New Roman",
                size=11,
                italic=True,
            )
        sheet.cell(final_row + 9, 1).font = Font(
            name="Times New Roman",
            size=11,
            italic=True,
        )
        sheet.cell(final_row + 11, 1).font = Font(
            name="Times New Roman",
            size=11,
            bold=True,
        )
        for row in range(final_row + 11, notes_end + 1):
            start_column = 2 if report_code == "14" and row == final_row + 11 else 1
            sheet.cell(row, start_column).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False,
            )

        widths = {
            "A": 7,
            "B": 15 if options.include_branch_in_customer_id else 11,
            "C": 35 if report_code in {"13", "14"} else 29,
            "D": 15,
            "E": max(12, min(35, self._max_text_width(sheet, 5, 11, final_row) + 2))
            if report_code == "13"
            else 12,
            "F": 6,
            "G": 19,
            "H": 19,
            "I": 8,
            "J": 11 if options.four_digit_year else 9,
            "K": 11 if options.four_digit_year else 9,
            "L": 7,
            "M": 11 if options.four_digit_year else 9,
            "N": 12,
            "O": 13,
            "P": 13,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        print_end = notes_end
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:{get_column_letter(final_column)}{print_end}",
            orientation="landscape",
            title_rows="$9:$10",
        )
        sheet.print_title_cols = f"$A:${get_column_letter(final_column)}"
        sheet.page_setup.scale = (
            68
            if report_code == "14"
            else 72
            if options.include_branch_in_customer_id and options.four_digit_year
            else 74
            if options.include_branch_in_customer_id or options.four_digit_year
            else 77
        )
        sheet.page_setup.fitToWidth = 0
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = False
        sheet.freeze_panes = "B11"
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.view = "normal"

    def _write_extra_headers(self, sheet) -> None:
        for column, header in enumerate(self.EXTRA_HEADERS, start=17):
            sheet.cell(10, column).value = header

    @staticmethod
    def _report_code(request: SettlementRequest) -> str:
        code = request.spec.report_code.casefold()
        if code not in {"13", "14"}:
            raise SettlementError(f"Processor Mẫu 13/14 không hỗ trợ mã {code}.")
        return code

    @staticmethod
    def _currency_order(records: list[DepositRecord]) -> tuple[str, ...]:
        return tuple(dict.fromkeys(record.currency for record in records if record.currency))

    @staticmethod
    def _sum_rows_formula(column: str, rows: list[int]) -> str:
        return "=" + "+".join(f"{column}{row}" for row in rows) if rows else "=0"

    @staticmethod
    def _output_prefix(options: SettlementOptions) -> str:
        prefix = (options.output_prefix or "QT").strip().upper()
        return "BN" if prefix == "BN" else "QT"

    @staticmethod
    def _max_text_width(sheet, column: int, start_row: int, end_row: int) -> int:
        return max(
            (
                len(str(sheet.cell(row, column).value or ""))
                for row in range(start_row, end_row + 1)
            ),
            default=0,
        )

    @staticmethod
    def _clean_code(value: Any) -> str:
        return ILLEGAL_CHARACTERS_RE.sub(
            "",
            str(value or "").strip().lstrip("'").strip(),
        )

    @staticmethod
    def _clean_text(value: Any) -> str:
        return ILLEGAL_CHARACTERS_RE.sub("", str(value or "").strip())

    @classmethod
    def _normalized_headers(cls, headers: set[str]) -> set[str]:
        normalized = set(headers)
        for canonical, aliases in cls.HEADER_ALIASES.items():
            if canonical in headers or any(alias in headers for alias in aliases):
                normalized.add(canonical)
        return normalized

    @classmethod
    def _normalize_row(cls, row: dict[str, str]) -> dict[str, str]:
        normalized = dict(row)
        for canonical, aliases in cls.HEADER_ALIASES.items():
            if canonical in normalized:
                continue
            for alias in aliases:
                if alias in normalized:
                    normalized[canonical] = normalized[alias]
                    break
        return normalized

    @staticmethod
    def _numeric_code(value: str) -> int | str | None:
        text = str(value or "").strip()
        if not text:
            return None
        return int(text) if text.isdigit() else text

    @staticmethod
    def _number_or_text(value: Any) -> int | float | str | None:
        text = str(value or "").strip()
        if not text:
            return None
        amount = Mau1314Processor._parse_source_amount(text)
        if amount == 0 and text not in {"0", "0.0", "0,0"}:
            return text
        return Mau1314Processor._number_to_cell(amount)

    @staticmethod
    def _amount_display(value: Any) -> int | float | None:
        text = str(value or "").strip()
        if not text:
            return None
        return Mau1314Processor._number_to_cell(
            Mau1314Processor._parse_source_amount(text)
        )

    @staticmethod
    def _parse_source_amount(value: Any) -> Decimal:
        text = str(value or "").replace("\u00a0", "").replace(" ", "").strip()
        if not text or text == "-":
            return Decimal(0)
        if "," in text and "." not in text:
            text = text.replace(",", ".")
        elif "," in text and "." in text:
            text = text.replace(",", "")
        try:
            return Decimal(text)
        except InvalidOperation:
            return Decimal(0)

    @staticmethod
    def _number_to_cell(value: Decimal | int | float) -> int | float:
        if not isinstance(value, Decimal):
            value = Decimal(str(value))
        if value == value.to_integral_value():
            return int(value)
        return float(value)
