from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import StringIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementOptions, SettlementRequest, SettlementResult
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)


@dataclass(frozen=True, slots=True)
class BadDebtRecord:
    customer_id: str
    customer_name: str
    contract_number: str
    disbursement_code: str
    account_name: str
    interest_rate: Decimal
    current_principal_balance: Decimal
    principal_balance: Decimal
    interest_balance: Decimal
    real_estate_value: Decimal
    movable_asset_value: Decimal
    other_asset_value: Decimal
    collateral_value: Decimal


class Mau20aProcessor:
    """Processor for Mẫu 20a/QT, ported from QUYETTOAN_MAU20a."""

    REQUIRED_HEADERS = {
        "custno",
        "custnm",
        "MaGiaiNgan",
        "MaHopDong",
        "TaiKhoan",
        "DuGocHienTai",
        "DuGocCuoi",
        "DuLaiCuoi",
        "BatDongSan",
        "DongSan",
        "TSKhac",
        "TongTS",
        "laisuat",
    }

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError("Mẫu 20a/QT cần đúng một file nguồn.")
        source_path = request.source_paths[0]
        records = self.read_source(request, source_path)
        workbook = self.build_workbook(request, records, source_path)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}{self._output_prefix(request.options)}20a.xlsx"
        )
        workbook.save(output_path)
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name="20a",
            processed_rows=len(records),
        )

    def read_source(
        self,
        request: SettlementRequest,
        source_path: Path,
    ) -> list[BadDebtRecord]:
        suffix = source_path.suffix.casefold()
        if suffix == ".xls":
            rows = self._read_xls_rows(source_path)
        elif suffix in {".xlsx", ".xlsm"}:
            rows = self._read_xlsx_rows(source_path)
        else:
            raise SettlementError("Nguồn Mẫu 20a/QT phải là file Excel .xls/.xlsx.")
        if not rows:
            return []

        missing = sorted(self.REQUIRED_HEADERS - set(rows[0]))
        if missing:
            raise SettlementError(
                f"File nguồn thiếu cột: {', '.join(missing)}",
                code="invalid_mau20a_headers",
            )

        branch_code = request.profile.branch_code.strip()
        branch_rows = [
            row
            for row in rows
            if self._text(row.get("custno")).startswith(branch_code)
        ]
        selected_rows = branch_rows if branch_rows else rows
        records = [self._row_to_record(row) for row in selected_rows]

        # VBA deletes rows where column AA equals zero after DL20aXls. In the
        # transformed layout AA maps back to source column DuGocHienTai.
        return [
            record
            for record in records
            if record.current_principal_balance != 0
        ]

    def build_workbook(
        self,
        request: SettlementRequest,
        records: list[BadDebtRecord],
        source_path: Path,
    ) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "20a"
        report_year = self._report_year(source_path)
        self._write_report_header(sheet, request, report_year)
        total_row = self._write_detail_rows(sheet, request, records)
        self._write_signatures_and_notes(sheet, request, total_row, report_year)
        self._format_report(sheet, total_row)
        return workbook

    def _read_xls_rows(self, source_path: Path) -> list[dict[str, Any]]:
        try:
            import xlrd
        except ImportError as exc:
            raise SettlementError(
                "Mẫu 20a/QT cần thư viện xlrd để đọc file .xls. "
                "Hãy chạy lại cài đặt môi trường hoặc pip install -e .",
                code="missing_xlrd",
            ) from exc
        try:
            workbook = xlrd.open_workbook(str(source_path), logfile=StringIO())
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc
        if sheet.nrows < 2:
            return []
        headers = [self._text(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        rows: list[dict[str, Any]] = []
        for row_index in range(1, sheet.nrows):
            values = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
            if any(self._text(value) for value in values):
                rows.append(dict(zip(headers, values, strict=False)))
        return rows

    def _read_xlsx_rows(self, source_path: Path) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(source_path, read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [self._text(value) for value in next(rows_iter)]
        except StopIteration:
            return []
        rows = []
        for values in rows_iter:
            if any(self._text(value) for value in values):
                rows.append(dict(zip(headers, values, strict=False)))
        return rows

    def _row_to_record(self, row: dict[str, Any]) -> BadDebtRecord:
        return BadDebtRecord(
            customer_id=self._text(row.get("custno")),
            customer_name=self._text(row.get("custnm")),
            contract_number=self._text(row.get("MaHopDong")),
            disbursement_code=self._text(row.get("MaGiaiNgan")),
            account_name=self._text(row.get("TaiKhoan")),
            interest_rate=self._decimal(row.get("laisuat")),
            current_principal_balance=self._decimal(row.get("DuGocHienTai")),
            principal_balance=self._decimal(row.get("DuGocCuoi")),
            interest_balance=self._decimal(row.get("DuLaiCuoi")),
            real_estate_value=self._decimal(row.get("BatDongSan")),
            movable_asset_value=self._decimal(row.get("DongSan")),
            other_asset_value=self._decimal(row.get("TSKhac")),
            collateral_value=self._decimal(row.get("TongTS")),
        )

    def _write_report_header(
        self,
        sheet,
        request: SettlementRequest,
        report_year: int,
    ) -> None:
        profile = request.profile
        branch_name = (
            profile.reporting_branch_name.strip()
            or profile.branch_name.strip()
            or f"Chi nhánh {profile.branch_code.strip()}"
        )
        sheet.merge_cells("A1:D1")
        sheet["A1"] = "NGÂN HÀNG NÔNG NGHIỆP"
        sheet.merge_cells("A2:D2")
        sheet["A2"] = "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM"
        sheet.merge_cells("A4:D4")
        sheet["A4"] = f"Mã chi nhánh: {profile.branch_code.strip()}"
        sheet.merge_cells("A5:D5")
        sheet["A5"] = f"Tên chi nhánh: {branch_name}"
        style_agency_header(sheet, start_column=1, end_column=4)

        sheet.merge_cells("A6:P6")
        sheet["A6"] = "BÁO CÁO NỢ ĐƯỢC XỬ LÝ BẰNG NGUỒN DỰ PHÒNG"
        sheet["A6"].font = Font(name="Times New Roman", size=16, bold=True)
        sheet["A6"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells("A7:P7")
        sheet["A7"] = f"Ngày 31 tháng 12 năm {report_year}"
        sheet["A7"].font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        sheet["A7"].alignment = Alignment(horizontal="center", vertical="center")

        sheet.merge_cells("N1:P1")
        sheet["N1"] = "1. Mẫu số 20a/QT"
        sheet.merge_cells("N2:P2")
        sheet["N2"] = "2. CN loại I gửi file về TSC"
        sheet.merge_cells("N3:P3")
        sheet["N3"] = "3. Lưu tại Chi nhánh"
        for row in range(1, 4):
            sheet.cell(row, 14).font = Font(name="Times New Roman", size=8)

        sheet["O8"] = "Đơn vị: VNĐ"
        style_currency_unit(sheet["O8"])

        headers = {
            "A9": "STT",
            "B9": "MÃ CHI NHÁNH",
            "C9": "TÊN CHI NHÁNH",
            "D9": "MÃ KHÁCH HÀNG",
            "E9": "TÊN KHÁCH HÀNG",
            "F9": "SỐ HỢP ĐỒNG TÍN DỤNG",
            "G9": "SỐ KHẾ ƯỚC",
            "H9": "LÃI SUẤT",
            "I9": f"DƯ NỢ GỐC ĐÃ XLRR TẠI NGÀY 31/12/{report_year}",
            "K9": f"DƯ NỢ LÃI ĐÃ XLRR TẠI NGÀY 31/12/{report_year}",
            "M9": "LOẠI TÀI SẢN ĐẢM BẢO",
            "N9": f"GIÁ TRỊ TSĐB TẠI NGÀY 31/12/{report_year} (hạch toán ngoại bảng)",
            "O9": "HIỆN TRẠNG CỦA TÀI SẢN ĐẢM BẢO",
            "P9": "GIÁ TRỊ THU HỒI ƯỚC TÍNH",
            "I10": "TÀI KHOẢN (TK9711xx)",
            "J10": "SỐ TIỀN (Quy đổi VNĐ)",
            "K10": "TÀI KHOẢN (TK9712xx)",
            "L10": "SỐ TIỀN (Quy đổi VNĐ)",
        }
        for address, value in headers.items():
            sheet[address] = value
        for cell_range in (
            "A9:A10",
            "B9:B10",
            "C9:C10",
            "D9:D10",
            "E9:E10",
            "F9:F10",
            "G9:G10",
            "H9:H10",
            "I9:J9",
            "K9:L9",
            "M9:M10",
            "N9:N10",
            "O9:O10",
            "P9:P10",
        ):
            sheet.merge_cells(cell_range)
        for col in range(1, 17):
            cell = sheet.cell(11, col)
            cell.value = col
            cell.number_format = "(0)"

    def _write_detail_rows(
        self,
        sheet,
        request: SettlementRequest,
        records: list[BadDebtRecord],
    ) -> int:
        branch_code = request.profile.branch_code.strip()
        branch_name = (
            request.profile.reporting_branch_name.strip()
            or request.profile.branch_name.strip()
        )
        row = 12
        for index, record in enumerate(records, start=1):
            customer_id = self._customer_id(record.customer_id, branch_code, request)
            sheet.cell(row, 1).value = index
            sheet.cell(row, 2).value = branch_code
            sheet.cell(row, 3).value = branch_name
            sheet.cell(row, 4).value = customer_id
            sheet.cell(row, 5).value = record.customer_name
            sheet.cell(row, 6).value = record.contract_number
            sheet.cell(row, 7).value = record.disbursement_code
            sheet.cell(row, 8).value = self._decimal_to_number(record.interest_rate)
            sheet.cell(row, 9).value = self._principal_account(record.account_name)
            sheet.cell(row, 10).value = self._decimal_to_number(record.principal_balance)
            sheet.cell(row, 11).value = ""
            sheet.cell(row, 12).value = self._decimal_to_number(record.interest_balance)
            sheet.cell(row, 13).value = self._collateral_type(record)
            sheet.cell(row, 14).value = self._decimal_to_number(record.collateral_value)
            sheet.cell(row, 15).value = ""
            sheet.cell(row, 16).value = self._decimal_to_number(
                record.principal_balance if record.collateral_value > 0 else Decimal(0)
            )
            row += 1

        total_row = row
        sheet.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=3,
        )
        sheet.cell(total_row, 1).value = "TỔNG CỘNG:"
        if records:
            first_data_row = 12
            last_data_row = total_row - 1
            for col in (10, 12, 14, 16):
                letter = get_column_letter(col)
                sheet.cell(total_row, col).value = (
                    f"=SUM({letter}{first_data_row}:{letter}{last_data_row})"
                )
        return total_row

    def _write_signatures_and_notes(
        self,
        sheet,
        request: SettlementRequest,
        total_row: int,
        report_year: int,
    ) -> None:
        date_row = total_row + 3
        title_row = total_row + 4
        note_row = total_row + 5
        sheet.merge_cells(
            start_row=date_row,
            start_column=14,
            end_row=date_row,
            end_column=16,
        )
        date_cell = sheet.cell(date_row, 14)
        report_location = request.profile.report_location.strip()
        date_text = f"Ngày 31 tháng 12 năm {report_year}"
        date_cell.value = (
            f"{report_location}, {date_text}" if report_location else date_text
        )
        date_cell.font = Font(name="Times New Roman", size=10, italic=True)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        signature_ranges = (
            (1, 4, "LẬP BIỂU", "(Ký, ghi rõ họ tên, số ĐT liên hệ)"),
            (5, 8, "TRƯỞNG PHÒNG KHCN/KHDN", "(Ký, ghi rõ họ tên)"),
            (9, 12, "TRƯỞNG PHÒNG KẾ TOÁN", "(Ký, ghi rõ họ tên)"),
            (14, 16, "GIÁM ĐỐC", "(Ký, đóng dấu, ghi rõ họ tên)"),
        )
        for start_col, end_col, title, note in signature_ranges:
            sheet.merge_cells(
                start_row=title_row,
                start_column=start_col,
                end_row=title_row,
                end_column=end_col,
            )
            sheet.merge_cells(
                start_row=note_row,
                start_column=start_col,
                end_row=note_row,
                end_column=end_col,
            )
            title_cell = sheet.cell(title_row, start_col)
            note_cell = sheet.cell(note_row, start_col)
            title_cell.value = title
            note_cell.value = note
            title_cell.font = Font(name="Times New Roman", size=11, bold=True)
            note_cell.font = Font(name="Times New Roman", size=10, italic=True)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            note_cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, 17):
            sheet.cell(note_row, col).font = Font(
                name="Times New Roman",
                size=10,
                italic=True,
            )
        sheet.cell(note_row, 1).alignment = Alignment(horizontal="center", vertical="center")

        notes_start = total_row + 13
        notes = [
            ("A", "Ghi chú:"),
            ("B", "1. Cột 4, 5, 6, 7, 8: Yêu cầu điền đầy đủ, rõ ràng thông tin đối với từng khoản nợ."),
            ("B", "2. Cột 9, 11: Ghi chi tiết đến tài khoản cấp V."),
            ("B", "3. Cột 10, 12: Dòng tổng cộng của từng tài khoản phải được đối chiếu khớp đúng với số dư từng tài khoản trên cân đối của chi nhánh ngày 31/12."),
            ("B", "4. Hiện trạng của tài sản đảm bảo đề nghị ghi rõ các giấy tờ hiện chi nhánh đang nắm giữ, khả năng phát mãi những tài sản này hoặc quyết định thi hành án (nếu có), ..."),
            ("B", "5. Trung tâm Công nghệ thông tin sẽ hỗ trợ gửi file về cho chi nhánh, chi nhánh thực hiện kiểm tra, đối chiếu, riêng cột 9, 11 và 12 chi nhánh phải nhập thủ công."),
        ]
        for offset, (column, text) in enumerate(notes):
            row = notes_start + offset
            col = 1 if column == "A" else 2
            cell = sheet.cell(row, col)
            cell.value = text
            cell.font = Font(name="Times New Roman", size=10, bold=(offset == 0))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _format_report(self, sheet, total_row: int) -> None:
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in sheet.iter_rows(min_row=9, max_row=total_row, min_col=1, max_col=16):
            for cell in row:
                cell.border = border
                cell.font = Font(
                    name="Times New Roman",
                    size=10,
                    bold=cell.row in {9, 10, 11},
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        for row in sheet.iter_rows(min_row=12, max_row=max(total_row - 1, 12), min_col=1, max_col=16):
            for cell in row:
                cell.font = Font(name="Times New Roman", size=10)
            for col in (3, 5, 13, 15):
                sheet.cell(row[0].row, col).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )
            for col in (6, 7):
                sheet.cell(row[0].row, col).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False,
                    shrink_to_fit=True,
                )

        for col in (8, 10, 12, 14, 16):
            number_format = "#,##0.00" if col == 8 else "#,##0"
            for row in range(12, total_row + 1):
                sheet.cell(row, col).number_format = number_format
        for col in (2, 4, 6, 7, 9, 11):
            for row in range(12, total_row):
                sheet.cell(row, col).number_format = "@"

        for col in range(1, 17):
            total_cell = sheet.cell(total_row, col)
            total_cell.font = Font(name="Times New Roman", size=10, bold=True)
            total_cell.alignment = Alignment(horizontal="center", vertical="center")

        widths = {
            "A": 5.5,
            "B": 5.5,
            "C": 22,
            "D": 15,
            "E": 22,
            "F": 22,
            "G": 22,
            "H": 11,
            "I": 13.1,
            "J": 17,
            "K": 11,
            "L": 13,
            "M": 13,
            "N": 24,
            "O": 13,
            "P": 13,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[9].height = 27
        for row in (10, 11):
            sheet.row_dimensions[row].height = 28
        for row in range(12, total_row):
            sheet.row_dimensions[row].height = 24
        sheet.row_dimensions[total_row].height = 22
        sheet.freeze_panes = "A12"
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:P{total_row + 17}",
            orientation="landscape",
            title_rows="$11:$11",
        )
        sheet.page_setup.scale = 95

    def _collateral_type(self, record: BadDebtRecord) -> str:
        if record.real_estate_value > 0:
            return "BĐS"
        if record.movable_asset_value > 0:
            return "ĐS"
        if record.other_asset_value > 0:
            return "Khac"
        return ""

    @staticmethod
    def _principal_account(account_name: str) -> str:
        mappings = {
            "(TK 211)": "971101",
            "(TK 212)": "791102",
            "(TK 252101)": "791101",
            "(TK 252102)": "791102",
            "(TK 252103)": "791103",
            "(TK 271101)": "791101",
            "(TK 271102)": "791102",
            "(TK 271103)": "791103",
        }
        normalized = " ".join(account_name.strip().split())
        for suffix, account in mappings.items():
            if normalized.endswith(suffix):
                return account
        return ""

    @staticmethod
    def _customer_id(
        source_customer_id: str,
        branch_code: str,
        request: SettlementRequest,
    ) -> str:
        customer_id = source_customer_id[1:] if source_customer_id.startswith("'") else source_customer_id
        if request.options.include_branch_in_customer_id and not customer_id.startswith(branch_code):
            return f"{branch_code}{customer_id}"
        return customer_id

    @staticmethod
    def _report_year(source_path: Path) -> int:
        try:
            source_time = datetime.fromtimestamp(source_path.stat().st_mtime)
        except OSError:
            source_time = datetime.now()
        return source_time.year - 1

    @staticmethod
    def _decimal(value: Any) -> Decimal:
        if value is None or value == "":
            return Decimal(0)
        if isinstance(value, Decimal):
            return value
        if isinstance(value, bool):
            return Decimal(0)
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip().replace("\u00a0", "").replace(" ", "")
        text = text.replace(".", "").replace(",", ".")
        try:
            return Decimal(text)
        except InvalidOperation:
            return Decimal(0)

    @staticmethod
    def _decimal_to_number(value: Decimal) -> int | float:
        if value == value.to_integral_value():
            return int(value)
        return float(value)

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _output_prefix(options: SettlementOptions) -> str:
        prefix = (options.output_prefix or "QT").strip().upper()
        return "BN" if prefix == "BN" else "QT"
