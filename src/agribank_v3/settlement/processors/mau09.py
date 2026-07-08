from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.theme import theme_xml

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementRequest, SettlementResult
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)


class Mau09Processor:
    """Processor for accounting Mẫu 09a/09b/09c from mshr32 TSCĐ reports."""

    HEADER_ROWS = {
        "9a": [
            "KHOẢN MỤC",
            "NHÀ CỬA VẬT KIẾN TRÚC",
            "MÁY MÓC\nTHIẾT BỊ",
            "PHƯƠNG TIỆN VẬN TẢI TRUYỀN DẪN",
            "THIẾT BỊ DỤNG CỤ QUẢN LÝ",
            "TSCĐ HỮU HÌNH KHÁC",
            "TỔNG CỘNG",
        ],
        "9b": [
            "KHOẢN MỤC",
            "QUYỀN SỬ DỤNG ĐẤT",
            "BẢN QUYỀN BẰNG PHÁT MINH SÁNG CHẾ",
            "NHÃN HIỆU HÀNG HOÁ",
            "PHẦN MỀM MÁY VI TÍNH",
            "TÀI SẢN CỐ ĐỊNH VÔ HÌNH KHÁC",
            "TỔNG CỘNG",
        ],
        "9c": [
            "KHOẢN MỤC",
            "NHÀ CỬA VẬT KIẾN TRÚC",
            "MÁY MÓC\nTHIẾT BỊ",
            "PHƯƠNG TIỆN VẬN TẢI TRUYỀN DẪN",
            "THIẾT BỊ DỤNG CỤ QUẢN LÝ",
            "TSCĐ KHÁC",
            "TỔNG CỘNG",
        ],
    }

    LABELS = {
        "9a": [
            "I. NGUYÊN GIÁ TSCĐ HỮU HÌNH",
            "1. Số đầu kỳ (Số dư ngày 31/12/{prev_year} mang sang)",
            "2. Tăng trong kỳ:                ",
            "- Mua trong kỳ",
            "- Điều chuyển nội bộ đến",
            "- Tăng khác",
            "3. Giảm trong kỳ:                ",
            "- Thanh lý, nhượng bán",
            "- Điều chuyển nội bộ đi",
            "- Giảm khác",
            "4. Số cuối kỳ (4 = 1 + 2 - 3)",
            "II. GIÁ TRỊ HAO MÒN",
            "1. Số đầu kỳ (Số dư ngày 31/12/{prev_year} mang sang)",
            "2. Tăng trong kỳ:                ",
            "- Khấu hao trong kỳ",
            "- Điều chuyển nội bộ đến",
            "- Tăng khác",
            "3. Giảm trong kỳ:                ",
            "- Thanh lý, nhượng bán",
            "- Điều chuyển nội bộ đi",
            "- Giảm khác",
            "4. Số cuối kỳ (4 = 1 + 2 - 3)",
            "III. GIÁ TRỊ CÒN LẠI TSCĐ",
            "1. Tại ngày đầu kỳ (1 = I.1 - II.1)",
            "2. Tại ngày cuối kỳ: (2 = I.4 - II.4)",
        ],
        "9b": [
            "I - NGUYÊN GIÁ TSCĐ VÔ HÌNH",
            "1. Số đầu kỳ (Số dư ngày 31/12/{prev_year} mang sang)",
            "2. Tăng trong kỳ:                ",
            "- Mua trong kỳ",
            "- Tạo ra từ nội bộ doanh nghiệp",
            "- Điều chuyển nội bộ đến",
            "- Tăng khác",
            "3. Giảm trong kỳ:                ",
            "- Thanh lý, nhượng bán",
            "- Điều chuyển nội bộ đi",
            "- Giảm khác",
            "4. Số cuối kỳ (4 = 1 + 2 - 3)",
            "II -  GIÁ TRỊ HAO MÒN",
            "1. Số đầu kỳ (Số dư ngày 31/12/{prev_year} mang sang)",
            "2. Tăng trong kỳ:                ",
            "- Khấu hao trong kỳ",
            "- Điều chuyển nội bộ đến",
            "- Tăng khác",
            "3. Giảm trong kỳ:                ",
            "- Thanh lý, nhượng bán",
            "- Điều chuyển nội bộ đi",
            "- Giảm khác",
            "4. Số cuối kỳ (4 = 1 + 2 - 3)",
            "III - GIÁ TRỊ CÒN LẠI TSCĐ",
            "1. Tại ngày đầu kỳ (1 = I.1 - II.1)",
            "2. Tại ngày cuối kỳ: (2 = I.4 - II.4)",
        ],
        "9c": [
            "I. NGUYÊN GIÁ TSCĐ THUÊ TÀI CHÍNH",
            "1. Số đầu kỳ (Số dư ngày 31/12/{prev_year} mang sang)",
            "2. Tăng trong kỳ:                ",
            "- Thuê tài chính trong kỳ",
            " - Tăng khác",
            "3. Giảm trong kỳ:                ",
            "- Trả lại TSCĐ thuê tài chính",
            "- Mua lại TSCĐ thuê tài chính",
            "- Giảm khác",
            "4. Số cuối kỳ (4 = 1 + 2 - 3)",
            "II. GIÁ TRỊ HAO MÒN",
            "1. Số đầu kỳ (Số dư ngày 31/12/{prev_year} mang sang)",
            "2. Tăng trong kỳ:                ",
            "- Khấu hao trong kỳ",
            "- Tăng khác",
            "3. Giảm trong kỳ:                ",
            "- Trả lại TSCĐ thuê tài chính",
            "- Mua lại TSCĐ thuê tài chính",
            "- Giảm khác",
            "4. Số cuối kỳ (4 = 1 + 2 - 3)",
            "III. GIÁ TRỊ CÒN LẠI TSCĐ",
            "1. Tại ngày đầu kỳ (1 = I.1 - II.1)",
            "2. Tại ngày cuối kỳ: (2 = I.4 - II.4)",
        ],
    }

    FORMULA_ROWS = {
        "9a": ((10, (11, 12, 13)), (14, (15, 16, 17)), (18, (9, 10, -14)),
               (21, (22, 23, 24)), (25, (26, 27, 28)), (29, (20, 21, -25)),
               (31, (9, -20)), (32, (18, -29))),
        "9b": ((10, (11, 12, 13, 14)), (15, (16, 17, 18)), (19, (9, 10, -15)),
               (22, (23, 24, 25)), (26, (27, 28, 29)), (30, (21, 22, -26)),
               (32, (9, -21)), (33, (19, -30))),
        "9c": ((10, (11, 12)), (13, (14, 15, 16)), (17, (9, 10, -13)),
               (20, (21, 22)), (23, (24, 25, 26)), (27, (19, 20, -23)),
               (29, (9, -19)), (30, (17, -27))),
    }

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError(f"Mẫu {request.spec.report_code}/QT cần đúng một file nguồn.")
        source_path = request.source_paths[0]
        report_code = self._report_code(request)
        rows = self.read_source(source_path)
        self._validate_source(report_code, rows)
        workbook = self.build_workbook(request, rows)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}{request.options.output_prefix}{report_code}.xlsx"
        )
        workbook.save(output_path)
        workbook.close()
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name=report_code,
            processed_rows=len(self.LABELS[report_code]),
        )

    def read_source(self, source_path: Path) -> list[list[Any]]:
        suffix = source_path.suffix.casefold()
        if suffix not in {".xls", ".xlsx", ".xlsm"}:
            raise SettlementError("Nguồn mẫu 09a/09b/09c phải là file Excel.")
        try:
            with source_path.open("rb") as stream:
                raw = stream.read()
            if raw.startswith(b"PK"):
                workbook = load_workbook(BytesIO(raw), data_only=True)
                sheet = workbook.active
                rows = [
                    [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
                    for row in range(1, sheet.max_row + 1)
                ]
                workbook.close()
                return rows
            import xlrd
            workbook = xlrd.open_workbook(file_contents=raw)
            sheet = workbook.sheet_by_index(0)
            return [
                [sheet.cell_value(row, col) for col in range(sheet.ncols)]
                for row in range(sheet.nrows)
            ]
        except Exception as exc:
            raise SettlementError(f"Không thể đọc file nguồn mẫu 09: {exc}") from exc

    def build_workbook(self, request: SettlementRequest, rows: list[list[Any]]) -> Workbook:
        report_code = self._report_code(request)
        workbook = Workbook()
        workbook.loaded_theme = theme_xml
        self._set_default_font(workbook)
        sheet = workbook.active
        sheet.title = report_code
        self._write_header(sheet, request, report_code)
        self._write_table(sheet, rows, report_code)
        self._write_signatures_and_notes(sheet, request, report_code)
        self._format_report(sheet, report_code)
        return workbook

    def _write_header(self, sheet, request: SettlementRequest, report_code: str) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        report_date = self._report_date_text(request)
        titles = {
            "9a": "BÁO CÁO TÌNH HÌNH TĂNG, GIẢM TSCĐ HỮU HÌNH",
            "9b": "BÁO CÁO TÌNH HÌNH TĂNG, GIẢM TSCĐ VÔ HÌNH",
            "9c": "BÁO CÁO TÌNH HÌNH TĂNG, GIẢM TSCĐ THUÊ TÀI CHÍNH",
        }
        values = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": f"Mã chi nhánh: {profile.branch_code.strip()}",
            "A4": f"Tên chi nhánh: {branch_name.removeprefix('Chi nhánh ').removeprefix('chi nhánh: ').strip()}",
            "A5": titles[report_code],
            "A6": report_date,
            "E1": f"1. Mẫu số 09{report_code[-1]}/QT",
            "E2": "2. CN loại I gửi báo cáo giấy, file về TSC",
            "E3": "3. Lưu tại chi nhánh",
            "G6": "Đơn vị : VNĐ",
        }
        for cell, value in values.items():
            sheet[cell] = value
        for row in range(1, 5):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        for row in range(1, 4):
            sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
            sheet.cell(row, 5).font = Font(name="Times New Roman", size=10)
            sheet.cell(row, 5).alignment = Alignment(horizontal="left", vertical="center")
        sheet.merge_cells("A5:G5")
        sheet.merge_cells("A6:F6")
        style_agency_header(sheet, start_column=1, end_column=4)
        sheet["A5"].font = Font(name="Times New Roman", size=14, bold=True)
        sheet["A5"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A6"].font = Font(name="Times New Roman", size=12, bold=True, italic=True)
        sheet["A6"].alignment = Alignment(horizontal="center", vertical="center")
        style_currency_unit(sheet["G6"])

    def _write_table(self, sheet, rows: list[list[Any]], report_code: str) -> None:
        for column, label in enumerate(self.HEADER_ROWS[report_code], start=1):
            sheet.cell(7, column, label)
        prev_year = date.today().year - 1
        labels = [label.format(prev_year=prev_year) for label in self.LABELS[report_code]]
        source_start_row = 11
        for offset, label in enumerate(labels, start=8):
            sheet.cell(offset, 1, label)
            source = rows[source_start_row - 1 + offset - 8] if len(rows) >= source_start_row + offset - 8 else []
            for column in range(2, 7):
                sheet.cell(offset, column, self._number(source[column] if len(source) > column else None))
            sheet.cell(offset, 7, f"=SUM(B{offset}:F{offset})")

        for target_row, source_rows in self.FORMULA_ROWS[report_code]:
            excel_row = target_row
            for column in range(2, 8):
                letter = get_column_letter(column)
                formula_parts = []
                for row in source_rows:
                    sign = "-" if row < 0 else "+"
                    formula_parts.append(f"{sign}{letter}{abs(row)}")
                formula = "=" + "".join(formula_parts).lstrip("+")
                sheet.cell(excel_row, column, formula)
        if self._has_no_data(rows, report_code):
            sheet.merge_cells(start_row=8, start_column=2, end_row=self._table_end_row(report_code), end_column=7)
            sheet.cell(8, 2, "CHI NHÁNH KHÔNG PHÁT SINH")
            sheet.cell(8, 2).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(8, 2).font = Font(name="Times New Roman", size=10, bold=True)

    def _write_signatures_and_notes(self, sheet, request: SettlementRequest, report_code: str) -> None:
        end_row = self._table_end_row(report_code)
        location = request.profile.report_location.strip()
        report_date = self._report_date_text(request)
        date_text = f"{location}, {report_date}" if location else report_date
        date_row = end_row + 1
        sign_row = end_row + 2
        note_row = end_row + 3
        sheet.merge_cells(start_row=date_row, start_column=5, end_row=date_row, end_column=7)
        sheet.cell(date_row, 5, date_text)
        sheet.cell(date_row, 5).font = Font(name="Times New Roman", size=10, italic=True)
        sheet.cell(date_row, 5).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(sign_row, 1, "LẬP BIỂU")
        sheet.merge_cells(start_row=sign_row, start_column=2, end_row=sign_row, end_column=4)
        sheet.cell(sign_row, 2, "TRƯỞNG PHÒNG KẾ TOÁN")
        sheet.merge_cells(start_row=sign_row, start_column=5, end_row=sign_row, end_column=7)
        sheet.cell(sign_row, 5, "GIÁM ĐỐC")
        sheet.cell(note_row, 1, "(Ký, ghi rõ họ tên, số ĐT liên hệ)")
        sheet.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=4)
        sheet.cell(note_row, 2, "(Ký, ghi rõ họ tên)")
        sheet.merge_cells(start_row=note_row, start_column=5, end_row=note_row, end_column=7)
        sheet.cell(note_row, 5, "(Ký, đóng dấu, ghi rõ họ tên)")
        for row in (sign_row, note_row):
            for col in range(1, 8):
                sheet.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        for col in (1, 2, 5):
            sheet.cell(sign_row, col).font = Font(name="Times New Roman", size=10, bold=True)
            sheet.cell(note_row, col).font = Font(name="Times New Roman", size=10, italic=True)

        preparer = request.profile.report_preparer.strip()
        phone = request.profile.phone.strip()
        if preparer or phone:
            sheet.cell(end_row + 8, 1, f"{preparer} ({phone})" if phone else preparer)
            sheet.cell(end_row + 8, 1).font = Font(name="Times New Roman", size=10, italic=True)

        notes = self._notes(report_code)
        start_note = end_row + 11
        for index, text in enumerate(notes, start=start_note):
            sheet.cell(index, 1, text)
            sheet.cell(index, 1).font = Font(name="Times New Roman", size=10, bold=index == start_note)
            sheet.cell(index, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _format_report(self, sheet, report_code: str) -> None:
        max_row = sheet.max_row
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        table_end = self._table_end_row(report_code)
        for row in range(7, table_end + 1):
            for column in range(1, 8):
                cell = sheet.cell(row, column)
                cell.border = border
                cell.font = Font(name="Times New Roman", size=10, bold=row == 7)
                cell.alignment = Alignment(horizontal="center" if row == 7 else "left", vertical="center", wrap_text=row == 7)
                if column >= 2 and row >= 8:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
        for row in self._bold_rows(report_code):
            for column in range(1, 8):
                sheet.cell(row, column).font = Font(name="Times New Roman", size=10, bold=True)
        for row in self._indented_rows(report_code):
            sheet.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center", indent=2)
        no_data_range = f"B8:G{table_end}"
        if no_data_range in {str(merged) for merged in sheet.merged_cells.ranges}:
            sheet.cell(8, 2).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(8, 2).font = Font(name="Times New Roman", size=10, bold=True)

        widths = {
            "9a": {"A": 47, "B": 18, "C": 17, "D": 18, "E": 16, "F": 17, "G": 16},
            "9b": {"A": 40, "B": 16, "C": 21, "D": 19, "E": 16, "F": 18, "G": 16},
            "9c": {"A": 42, "B": 16, "C": 16, "D": 19, "E": 16, "F": 16, "G": 18},
        }[report_code]
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[7].height = 42
        for row in range(8, table_end + 1):
            sheet.row_dimensions[row].height = 16
        setup_a4_print_layout(sheet, print_area=f"A1:G{max_row}", title_rows="$7:$7")

    def _validate_source(self, report_code: str, rows: list[list[Any]]) -> None:
        if len(rows) < 11:
            raise SettlementError("File nguồn mẫu 09 không đủ dòng dữ liệu.")
        expected = {
            "9a": ("Nguyên giá TSCĐ hữu hình", "Nhà cửa, vật kiến trúc"),
            "9b": ("Nguyên giá TSCĐ vô hình", "Quyền sử dụng đất có thời hạn"),
            "9c": ("Nguyên giá TSCĐ thuê tài chính", "Nhà cửa, vật kiến trúc"),
        }[report_code]
        row10 = rows[9]
        row11 = rows[10]
        if (
            self._text(row10[0] if len(row10) > 0 else "") != "STT"
            or self._text(row10[1] if len(row10) > 1 else "") != "CHỈ TIÊU"
            or self._text(row10[2] if len(row10) > 2 else "") != expected[1]
            or expected[0] not in self._text(row11[1] if len(row11) > 1 else "")
        ):
            source_codes = {"9a": "TMBCTC_TSCD001", "9b": "TMBCTC_TSCD002", "9c": "TMBCTC_TSCD003"}
            raise SettlementError(
                f"File nguồn không đúng mẫu {report_code}/QT. "
                f"Hãy chọn file mshr32 báo cáo {source_codes[report_code]}."
            )

    def _report_code(self, request: SettlementRequest) -> str:
        suffix = request.spec.key.rsplit(".", 1)[-1].casefold()
        if suffix in {"09a", "09b", "09c"}:
            return suffix.replace("09", "9")
        raise SettlementError("Mẫu 09 không xác định được loại 9a/9b/9c.")

    def _report_date_text(self, request: SettlementRequest) -> str:
        today = date.today()
        if request.options.output_prefix == "BN":
            return f"Ngày 30 tháng 6 năm {today.year}"
        return f"Ngày 31 tháng 12 năm {today.year}"

    def _set_default_font(self, workbook: Workbook) -> None:
        workbook._named_styles["Normal"].font = Font(name="Times New Roman", size=10)

    def _number(self, value: Any) -> int | float | None:
        if value in (None, ""):
            return None
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        if number == 0:
            return None
        return int(number) if number.is_integer() else number

    def _text(self, value: Any) -> str:
        return "" if value is None else str(value).strip()

    def _bold_rows(self, report_code: str) -> tuple[int, ...]:
        return {
            "9a": (8, 9, 10, 14, 18, 19, 20, 21, 25, 29, 30),
            "9b": (8, 9, 10, 15, 19, 20, 21, 22, 26, 30, 31),
            "9c": (8, 9, 10, 13, 17, 18, 19, 20, 23, 27, 28),
        }[report_code]

    def _indented_rows(self, report_code: str) -> tuple[int, ...]:
        return {
            "9a": (11, 12, 13, 15, 16, 17, 22, 23, 24, 26, 27, 28, 31, 32),
            "9b": (11, 12, 13, 14, 16, 17, 18, 23, 24, 25, 27, 28, 29, 32, 33),
            "9c": (11, 12, 14, 15, 16, 21, 22, 24, 25, 26, 29, 30),
        }[report_code]

    def _table_end_row(self, report_code: str) -> int:
        return {"9a": 32, "9b": 33, "9c": 30}[report_code]

    def _has_no_data(self, rows: list[list[Any]], report_code: str) -> bool:
        row_count = len(self.LABELS[report_code])
        for row in rows[10 : 10 + row_count]:
            for value in row[2:7]:
                number = self._number(value)
                if number not in (None, 0):
                    return False
        return True

    def _notes(self, report_code: str) -> list[str]:
        common = [
            "Ghi chú: ",
            "1. Số liệu tại từng cột chỉ tiêu phải khớp với số dư tài khoản chi tiết trên cân đối tài khoản (nguyên giá, hao mòn).",
            "2. Giá trị nguyên giá, hao mòn TSCĐ: Số dư cuối kỳ trước bằng số dư đầu kỳ này.",
            "3. Giá trị còn lại của TSCĐ tại mẫu 9a/QT + 9b/QT = số dư tài khoản 519105.",
            "4. Chỉ tiêu Khấu hao trong kỳ tại mẫu 9a/QT + 9b/QT + 9c/QT = Số dư cuối tài khoản 871001.",
        ]
        if report_code != "9c":
            return common
        return common[:3] + [
            "3. Số dư cuối kỳ = Số dư đầu kỳ + Tăng trong kỳ - Giảm trong kỳ.",
            "4. Chỉ tiêu Khấu hao trong kỳ tại mẫu 9a/QT + 9b/QT + 9c/QT = Số dư cuối tài khoản 871001.",
            "5. Các chỉ tiêu Tăng khác/Giảm khác, Điều chuyển nội bộ đi/đến phải bằng nhau; nếu chênh lệch phải giải trình rõ.",
            "6. Đối với TSCĐ thuê tài chính mua lại trong kỳ: phần nguyên giá ghi tăng ở TSCĐ hữu hình, phần hao mòn ghi tăng khác.",
            "7. IPCAS hỗ trợ lấy số liệu tại MIS/HQ Report (SBV)/Các báo cáo tài chính.",
        ]
