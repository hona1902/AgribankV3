from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import (
    SettlementOptions,
    SettlementRequest,
    SettlementResult,
)
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)
from agribank_v3.settlement.transforms import (
    normalize_customer_id,
    parse_yyyymmdd,
    vietnamese_report_date,
)


@dataclass(frozen=True, slots=True)
class CollateralRecord:
    sort_customer_id: str
    customer_id: str
    customer_name: str
    contract_number: str
    contract_date: object
    total: int
    real_estate: int
    movable_property: int
    valuable_papers: int
    other_assets: int
    legal_valid: str
    legal_other: str
    dossier_complete: str
    dossier_in_progress: str
    dossier_impossible: str
    dossier_other: str
    marketable: str
    less_marketable: str
    not_marketable: str
    account: str
    collateral_type: str
    source_values: tuple[object, ...] = ()


class Mau05Processor:
    ACCOUNT_NAMES = {
        "994001": "Tài sản thế chấp của khách hàng",
        "994003": "Tài sản cầm cố của khách hàng",
        "994009": "Nhap ten TK vao",
    }

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError(
                "Mẫu 05/QT cần đúng một file nguồn rt05.csv.",
                code="invalid_source_count",
            )
        source_path = request.source_paths[0]
        records, report_date = self.read_source(request, source_path)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}QT05.xlsx"
        )
        workbook = self.build_workbook(request, records, report_date)
        workbook.save(output_path)
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name="05",
            processed_rows=len(records),
        )

    def read_source(
        self,
        request: SettlementRequest,
        source_path: Path,
    ) -> tuple[list[CollateralRecord], str]:
        if source_path.suffix.casefold() != ".csv":
            raise SettlementError(
                "Nguồn Mẫu 05/QT phải là file CSV.",
                code="invalid_source_type",
            )
        try:
            with source_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                source_headers = tuple(reader.fieldnames or ())
                rows = list(reader)
        except (OSError, UnicodeError, csv.Error) as exc:
            raise SettlementError(f"Không thể đọc file rt05: {exc}") from exc
        required = {
            "NGAY", "MA_KH", "TEN_KH", "LOAI_THE_CHAP_BAO_LANH",
            "SO_HD", "NGAY_HD", "SO_TIEN", "BAT_DONG_SAN", "DONG_SAN",
            "GTCG", "TS_KHAC", "CO_KHA_NANG_PHAT_MAI",
            "IT_CO_KHA_NANG_PHAT_MAI", "KHONG_CO_KHA_NANG_PHAT_MAI",
            "TAI_KHOAN", "MA_CHU_TSDB", "TEN_CHU_TSDB",
        }
        headers = set(rows[0]) if rows else set()
        missing = sorted(required - headers)
        if missing:
            raise SettlementError(
                f"File rt05 thiếu cột: {', '.join(missing)}",
                code="invalid_rt05_headers",
            )

        branch_code = request.profile.branch_code.strip()
        options = request.options
        records: list[CollateralRecord] = []
        for row in rows:
            collateral_type = row["LOAI_THE_CHAP_BAO_LANH"].strip()
            customer_id = row["MA_KH"]
            sort_customer_id = normalize_customer_id(
                customer_id,
                branch_code,
                include_branch=options.include_branch_in_customer_id,
            )
            customer_name = row["TEN_KH"]
            raw_owner_id = row["MA_CHU_TSDB"]
            owner_id = raw_owner_id.strip().lstrip("'").strip()
            if (
                options.use_collateral_owner_for_guarantee
                and collateral_type.casefold() == "bao lanh"
                and raw_owner_id != ""
            ):
                customer_id = owner_id
                customer_name = row["TEN_CHU_TSDB"]
            customer_id = normalize_customer_id(
                customer_id,
                branch_code,
                include_branch=options.include_branch_in_customer_id,
            )
            not_marketable = row["KHONG_CO_KHA_NANG_PHAT_MAI"]
            less_marketable = row["IT_CO_KHA_NANG_PHAT_MAI"]
            contract_number = row["SO_HD"].strip()
            records.append(
                CollateralRecord(
                    sort_customer_id=sort_customer_id,
                    customer_id=customer_id,
                    customer_name=customer_name,
                    contract_number=(
                        int(contract_number)
                        if contract_number.isdigit()
                        else contract_number
                    ),
                    contract_date=parse_yyyymmdd(row["NGAY_HD"]),
                    total=self._integer(row["SO_TIEN"]),
                    real_estate=self._integer(row["BAT_DONG_SAN"]),
                    movable_property=self._integer(row["DONG_SAN"]),
                    valuable_papers=self._integer(row["GTCG"]),
                    other_assets=self._integer(row["TS_KHAC"]),
                    legal_valid="X",
                    legal_other="",
                    dossier_complete="X",
                    dossier_in_progress="",
                    dossier_impossible="",
                    dossier_other="",
                    marketable="" if not_marketable or less_marketable else "X",
                    less_marketable="X" if less_marketable and not not_marketable else "",
                    not_marketable="X" if not_marketable else "",
                    account=(row["TAI_KHOAN"].strip() or "994009"),
                    collateral_type=collateral_type,
                    source_values=tuple(row.get(header, "") for header in source_headers),
                )
            )
        records.sort(
            key=lambda item: (
                0 if item.collateral_type.casefold() == "the chap" else 1,
                item.account,
                item.sort_customer_id,
            )
        )
        report_date = vietnamese_report_date(rows[0]["NGAY"]) if rows else ""
        self._source_headers = source_headers
        return records, report_date

    def build_workbook(
        self,
        request: SettlementRequest,
        records: list[CollateralRecord],
        report_date: str,
    ) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "05"
        self._write_report_header(sheet, request, report_date)
        current_row = 12
        sheet.cell(current_row, 1, "I. TSTC của khách hàng là TCTD trực tiếp vay vốn:")
        sheet.cell(current_row + 1, 1, "Cộng I")
        sheet.cell(current_row + 2, 1, "II. TSTC của khách hàng trực tiếp vay vốn")
        current_row += 3

        direct = [r for r in records if r.collateral_type.casefold() != "bao lanh"]
        guaranteed = [r for r in records if r.collateral_type.casefold() == "bao lanh"]
        direct_start = current_row
        current_row = self._write_grouped_records(
            sheet, direct, current_row, request.options
        )
        direct_total = current_row
        self._write_total(
            sheet,
            direct_total,
            "Cộng II:",
            direct_start + 1,
            direct_total - 1,
            flag_start=14,
        )
        current_row += 1
        sheet.cell(
            current_row,
            1,
            "III. TSTC của đơn vị bảo lãnh cho bên thứ ba vay vốn:",
        )
        current_row += 1
        guaranteed_start = current_row
        current_row = self._write_grouped_records(
            sheet, guaranteed, current_row, request.options
        )
        guaranteed_total = current_row
        self._write_total(
            sheet,
            guaranteed_total,
            "Cộng III:",
            guaranteed_start + 1,
            guaranteed_total - 1,
            flag_start=guaranteed_start + 1,
        )
        grand_total = guaranteed_total + 1
        sheet.cell(grand_total, 1, "Tổng cộng (I+II+III):")
        for column in range(5, 19):
            letter = get_column_letter(column)
            sheet.cell(
                grand_total,
                column,
                f"={letter}{direct_total}+ {letter}{guaranteed_total}",
            )
        self._write_signatures_and_notes(
            sheet,
            request,
            report_date,
            grand_total,
        )
        if not request.options.remove_unused_columns:
            self._write_unused_source_columns(sheet, records)
        self._format_report(sheet, grand_total, request.options)
        if request.options.create_control_sheet:
            self._write_control_sheet(workbook, records)
        return workbook

    def _write_grouped_records(
        self,
        sheet,
        records: Iterable[CollateralRecord],
        row: int,
        options: SettlementOptions,
    ) -> int:
        current_account = None
        current_customer = None
        customer_start = row
        include_customer_totals = options.include_customer_totals
        for record in records:
            if record.account != current_account:
                if include_customer_totals and current_customer is not None:
                    row = self._write_customer_total(
                        sheet, row, customer_start, row - 1, current_customer
                    )
                    current_customer = None
                current_account = record.account
                sheet.cell(row, 1, int(record.account) if record.account.isdigit() else record.account)
                sheet.cell(row, 2, self.ACCOUNT_NAMES.get(record.account, "Nhap ten TK vao"))
                row += 1
                customer_start = row
            customer_key = (record.customer_id, record.customer_name)
            if include_customer_totals and (
                current_customer is not None and customer_key != current_customer
            ):
                row = self._write_customer_total(
                    sheet, row, customer_start, row - 1, current_customer
                )
                customer_start = row
            current_customer = customer_key
            values = (
                record.customer_id, record.customer_name, record.contract_number,
                record.contract_date, record.total, record.real_estate,
                record.movable_property, record.valuable_papers,
                record.other_assets, record.legal_valid, record.legal_other,
                record.dossier_complete, record.dossier_in_progress,
                record.dossier_impossible, record.dossier_other,
                record.marketable, record.less_marketable, record.not_marketable,
            )
            for column, value in enumerate(values, 1):
                sheet.cell(row, column, value)
            if not options.remove_unused_columns:
                for offset, value in enumerate(record.source_values, 19):
                    sheet.cell(row, offset, value)
            row += 1
        if include_customer_totals and current_customer is not None:
            row = self._write_customer_total(
                sheet, row, customer_start, row - 1, current_customer
            )
        return row

    @staticmethod
    def _write_customer_total(
        sheet,
        row: int,
        start: int,
        end: int,
        customer: tuple[str, str],
    ) -> int:
        customer_id, customer_name = customer
        sheet.cell(row, 1, f"   Cộng khách hàng: {customer_id}")
        sheet.cell(row, 2, customer_name)
        for column in range(5, 19):
            letter = get_column_letter(column)
            if column <= 9:
                formula = f"=SUBTOTAL(9,{letter}{start}:{letter}{end})"
            else:
                formula = f'=COUNTIF({letter}{start}:{letter}{end}, "X")'
            sheet.cell(row, column, formula)
        return row + 1

    def _write_unused_source_columns(
        self,
        sheet,
        records: list[CollateralRecord],
    ) -> None:
        del records
        headers = tuple(getattr(self, "_source_headers", ()))
        for offset, header in enumerate(headers, 19):
            cell = sheet.cell(11, offset, header)
            cell.font = Font(name="Times New Roman", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _write_total(
        sheet,
        row: int,
        label: str,
        start: int,
        end: int,
        *,
        flag_start: int,
    ) -> None:
        sheet.cell(row, 1, label)
        for column in range(5, 19):
            letter = get_column_letter(column)
            if column <= 9:
                formula = f"=SUBTOTAL(9,{letter}{start}:{letter}{end})"
            else:
                formula = f'=COUNTIF({letter}{flag_start}:{letter}{end}, "X")'
            sheet.cell(row, column, formula)

    @staticmethod
    def _write_report_header(sheet, request: SettlementRequest, report_date: str) -> None:
        profile = request.profile
        values = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": f"Mã chi nhánh: {profile.branch_code}",
            "A4": f"Tên {profile.branch_name}",
            "A5": "BÁO CÁO KIỂM KÊ HỒ SƠ, TÀI SẢN THẾ CHẤP, CẦM CỐ CỦA KHÁCH HÀNG",
            "A6": report_date,
            "R7": "Đơn vị: VNĐ",
            "P1": "1. Mẫu số 05/QT",
            "P2": "2. Lưu tại Chi nhánh",
        }
        for address, value in values.items():
            sheet[address] = value
        for address in ("A1:D1", "A2:D2", "A3:D3", "A4:D4"):
            sheet.merge_cells(address)
        sheet.merge_cells("A5:R5")
        sheet.merge_cells("A6:R6")
        sheet["A8"] = "MÃ KHÁCH HÀNG"
        sheet["B8"] = "TÊN KHÁCH HÀNG"
        sheet["C8"] = "SỐ HĐ THẾ CHẤP, CẦM CỐ"
        sheet["D8"] = "NGÀY HĐ THẾ CHẤP, CẦM CỐ"
        sheet["E8"] = "GIÁ TRỊ TSTC"
        sheet["J8"] = "HIỆN TRẠNG TSĐB"
        sheet.merge_cells("A8:A10")
        sheet.merge_cells("B8:B10")
        sheet.merge_cells("C8:C10")
        sheet.merge_cells("D8:D10")
        sheet.merge_cells("E8:I8")
        sheet.merge_cells("J8:R8")
        for column, value in enumerate(
            ("Tổng ", "Bất ĐS", "Động sản", "GTCG", "TS khác"), 5
        ):
            sheet.cell(9, column, value)
            sheet.merge_cells(start_row=9, start_column=column, end_row=10, end_column=column)
        sheet["J9"] = "Tính pháp lý"
        sheet["L9"] = "Khả năng hoàn thiện hồ sơ TS"
        sheet["P9"] = "Khả năng phát mại"
        sheet.merge_cells("J9:K9")
        sheet.merge_cells("L9:O9")
        sheet.merge_cells("P9:R9")
        headers = (
            "Hợp pháp, hợp lệ ", "Khác", "Đầy đủ hồ sơ theo quy định",
            "Đang hoàn chỉnh thủ tục", "Không thể hoàn chỉnh được", "Khác",
            "Có khả năng phát mại", "Ít có khả năng phát mại",
            "Không có khả năng phát mại",
        )
        for column, value in enumerate(headers, 10):
            sheet.cell(10, column, value)
        for column in range(1, 19):
            sheet.cell(11, column, column)

    @staticmethod
    def _write_signatures_and_notes(
        sheet,
        request: SettlementRequest,
        report_date: str,
        grand_total: int,
    ) -> None:
        profile = request.profile
        signature_row = grand_total + 2
        sheet.cell(signature_row - 1, 12, f"{profile.report_location}, {report_date}")
        sheet.merge_cells(start_row=signature_row - 1, start_column=12, end_row=signature_row - 1, end_column=18)
        labels = ((1, 3, "LẬP BIỂU"), (4, 6, "TRƯỞNG PHÒNG KHDN/KHCN"),
                  (7, 11, "TRƯỞNG PHÒNG KẾ TOÁN"), (12, 18, "GIÁM ĐỐC"))
        for start, end, label in labels:
            sheet.cell(signature_row, start, label)
            sheet.merge_cells(start_row=signature_row, start_column=start, end_row=signature_row, end_column=end)
            note = "(Ký, đóng dấu, ghi rõ họ tên)" if start == 12 else "(Ký, ghi rõ họ tên)"
            sheet.cell(signature_row + 1, start, note)
            sheet.merge_cells(start_row=signature_row + 1, start_column=start, end_row=signature_row + 1, end_column=end)
        sheet.cell(signature_row + 10, 1, f"({profile.report_preparer} - SĐT: {profile.phone})")
        sheet.merge_cells(start_row=signature_row + 10, start_column=1, end_row=signature_row + 10, end_column=3)
        note_row = signature_row + 11
        notes = (
            "1. Cột 5: Dòng tổng cộng khớp với số dư trên cân đối của tài khoản 994, 995, 996, 992003, 992005",
            "2. Cột 10 đến cột 18: Chi nhánh dùng dấu (x)",
            "3. Cột 4: Ghi theo thứ tự ngày/tháng/năm (không ghi tháng trước ngày sau)",
            "4. Trung tâm CNTT sẽ hỗ trợ chuyển file về cho chi nhánh, yêu cầu chi nhánh thực hiện kiểm tra, đối chiếu và điền tiếp đầy đủ thông tin vào từ cột 10 đến 18.",
        )
        sheet.cell(note_row, 1, "Ghi chú:")
        for offset, note in enumerate(notes):
            sheet.cell(note_row + offset, 2, note)

    @staticmethod
    def _format_report(
        sheet,
        grand_total: int,
        options: SettlementOptions,
    ) -> None:
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=10)
                cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(min_row=8, max_row=grand_total, min_col=1, max_col=18):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(8, 12):
            for cell in sheet[row]:
                cell.font = Font(name="Times New Roman", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet["A5"].font = Font(name="Times New Roman", size=16, bold=True)
        sheet["A6"].font = Font(name="Times New Roman", size=12, bold=True, italic=True)
        for address in ("A1", "A2", "A3", "A4", "A5", "A6"):
            sheet[address].alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
        style_currency_unit(sheet["R7"])
        for row in range(12, grand_total + 1):
            sheet.cell(row, 1).number_format = "@"
            sheet.cell(row, 4).number_format = (
                "dd/mm/yyyy" if options.four_digit_year else "dd/mm/yy"
            )
            for column in range(5, 10):
                sheet.cell(row, column).number_format = "#,###"
            for column in range(10, 19):
                sheet.cell(row, column).number_format = "0"
                sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
        for row in (12, 13, 14, grand_total):
            sheet.cell(row, 1).font = Font(name="Times New Roman", bold=True, color="000080")
        for row in range(12, grand_total + 1):
            value = sheet.cell(row, 1).value
            if isinstance(value, int) and 990000 <= value <= 999999:
                sheet.cell(row, 1).font = Font(name="Times New Roman", bold=True, color="FF0000")
                sheet.cell(row, 2).font = Font(name="Times New Roman", bold=True, color="FF0000")
            if (
                options.bold_customer_rows
                and isinstance(value, str)
                and value.startswith("   Cộng khách hàng:")
            ):
                for cell in sheet[row][:18]:
                    cell.font = Font(name="Times New Roman", bold=True)
            if isinstance(value, str) and value.startswith("   Cộng khách hàng:"):
                sheet.cell(row, 1).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                )
            if isinstance(value, str) and value.startswith(("Cộng II:", "Cộng III:")):
                for cell in sheet[row][:18]:
                    cell.font = Font(name="Times New Roman", bold=True)
                    cell.fill = PatternFill("solid", fgColor="FFFF99")
        for cell in sheet[grand_total][:18]:
            cell.font = Font(name="Times New Roman", bold=True)
            cell.fill = PatternFill("solid", fgColor="00FFFF")
            cell.border = Border(left=medium, right=medium, top=medium, bottom=medium)
        for total_row in (grand_total,):
            for column in range(5, 10):
                sheet.cell(total_row, column).font = Font(
                    name="Times New Roman", size=9, bold=True
                )
        for row in range(grand_total + 1, sheet.max_row + 1):
            sheet.cell(row, 1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        for merged in sheet.merged_cells.ranges:
            if merged.min_row > grand_total:
                sheet.cell(merged.min_row, merged.min_col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
        signature_row = grand_total + 2
        for column in (1, 4, 7, 12):
            sheet.cell(signature_row, column).font = Font(
                name="Times New Roman", bold=True
            )
            sheet.cell(signature_row + 1, column).font = Font(
                name="Times New Roman", italic=True
            )
        sheet.cell(signature_row - 1, 12).font = Font(
            name="Times New Roman", italic=True
        )
        sheet.cell(signature_row + 10, 1).font = Font(
            name="Times New Roman", italic=True
        )
        widths = {
            1: 26,
            2: 18,
            3: 20,
            4: 11.5,
            5: 17.5,
            6: 17.5,
            7: 17.5,
            8: 17.5,
            9: 16,
            10: 6,
            11: 6,
            12: 6,
            13: 6,
            14: 6,
            15: 6,
            16: 6,
            17: 6,
            18: 6,
        }
        for column in range(1, 19):
            sheet.column_dimensions[get_column_letter(column)].width = widths[column]
        if not options.remove_unused_columns:
            for column in range(19, sheet.max_column + 1):
                letter = get_column_letter(column)
                sheet.column_dimensions[letter].width = 15
                for row in range(11, grand_total + 1):
                    sheet.cell(row, column).font = Font(name="Times New Roman", size=9)
        style_agency_header(sheet, start_column=1, end_column=4)
        sheet.row_dimensions[5].height = 24
        sheet.row_dimensions[8].height = 24
        sheet.row_dimensions[9].height = 30
        sheet.row_dimensions[10].height = 75
        sheet.freeze_panes = "B12"
        sheet.sheet_view.showGridLines = False
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:R{sheet.max_row}",
            title_rows="8:10",
        )

    @staticmethod
    def _write_control_sheet(workbook: Workbook, records: list[CollateralRecord]) -> None:
        sheet = workbook.create_sheet("SoLieuTongHop")
        sheet.append(("SỐ LIỆU TỔNG HỢP MẪU 05", "", "", "", "", "", ""))
        sheet.append(("Tài Khoản", "Số Tiền", "Số Tiền Qui Đổi VNĐ",
                      "Bất Động Sản", "Động Sản", "CT Có Giá", "TS Khác"))
        accounts = list(dict.fromkeys(record.account for record in records))
        for account in accounts:
            selected = [record for record in records if record.account == account]
            sheet.append((
                int(account) if account.isdigit() else account,
                sum(item.total for item in selected),
                sum(item.total for item in selected),
                sum(item.real_estate for item in selected),
                sum(item.movable_property for item in selected),
                sum(item.valuable_papers for item in selected),
                sum(item.other_assets for item in selected),
            ))
        total_row = sheet.max_row + 1
        sheet.cell(total_row, 1, "Tổng cộng:")
        for column in range(2, 8):
            letter = get_column_letter(column)
            sheet.cell(total_row, column, f"=SUM({letter}3:{letter}{total_row - 1})")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=12)
        for cell in sheet[2]:
            cell.fill = PatternFill("solid", fgColor="FFFF00")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in range(2, 8):
            for row in range(3, total_row + 1):
                sheet.cell(row, column).number_format = "#,##0"
        sheet.column_dimensions["A"].width = 16
        for column in "BCDEFG":
            sheet.column_dimensions[column].width = 20
        sheet.row_dimensions[2].height = 50

    @staticmethod
    def _integer(value: str) -> int:
        try:
            return int(float(value.strip() or "0"))
        except ValueError:
            return 0
