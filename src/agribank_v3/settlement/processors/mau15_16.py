from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import (
    SettlementOptions,
    SettlementRequest,
    SettlementResult,
)
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)
from agribank_v3.settlement.transforms import (
    normalize_customer_id,
    parse_vietnamese_amount,
    parse_yyyymmdd,
    vietnamese_report_date,
)


@dataclass(frozen=True, slots=True)
class LoanScheduleRecord:
    account: str
    customer_id: str
    customer_name: str
    contract_number: str
    currency: str
    debt_group: str
    industry_code: object
    industry_name: str
    capital_source: str
    loan_purpose: str
    original_balance: Decimal
    converted_balance: Decimal
    term_months: str
    disbursement_date: object
    maturity_date: object
    interest_rate: str
    last_interest_collection_date: object
    prepaid_interest: Decimal
    accrued_interest: Decimal
    interest_account: object
    last_interest_collection_date_display: object = None
    disbursement_date_display: object = None
    maturity_date_display: object = None
    prepaid_interest_display: object = None
    accrued_interest_display: object = None
    interest_account_display: object = None
    source_values: tuple[object, ...] = ()


class Mau1516Processor:
    """Processor shared by Mẫu 15A/QT, 15B/QT and 16/QT."""

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError(
                f"Mẫu {request.spec.report_code}/QT cần đúng một file CSV nguồn.",
                code="invalid_source_count",
            )
        source_path = request.source_paths[0]
        records, report_date = self.read_source(request, source_path)
        workbook = self.build_workbook(request, records, report_date)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}{self._output_prefix(request.options)}{request.spec.report_code}.xlsx"
        )
        workbook.save(output_path)
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name=request.spec.report_code.casefold(),
            processed_rows=len(records),
        )

    def read_source(
        self,
        request: SettlementRequest,
        source_path: Path,
    ) -> tuple[list[LoanScheduleRecord], str]:
        if source_path.suffix.casefold() not in {".csv", ".txt"}:
            raise SettlementError(
                f"Nguồn Mẫu {request.spec.report_code}/QT phải là file CSV.",
                code="invalid_source_type",
            )
        try:
            with source_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                source_headers = tuple(reader.fieldnames or ())
                rows = list(reader)
        except (OSError, UnicodeError, csv.Error) as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc

        report_code = self._report_code(request)
        required = self._required_headers(report_code)
        missing = sorted(required - set(source_headers))
        if missing:
            raise SettlementError(
                f"File nguồn thiếu cột: {', '.join(missing)}",
                code="invalid_mau1516_headers",
            )
        branch_code = request.profile.branch_code.strip()
        options = request.options
        records = [
            self._row_to_record(row, source_headers, report_code, branch_code, options)
            for row in rows
        ]
        if report_code == "16":
            records.sort(
                key=lambda item: (
                    item.customer_id,
                    item.account,
                    self._sort_debt_group(item.debt_group),
                )
            )
        else:
            records.sort(
                key=lambda item: (
                    item.account,
                    item.currency,
                    self._sort_debt_group(item.debt_group),
                )
            )
        self._source_headers = source_headers
        report_date = vietnamese_report_date(rows[0]["NGAY"]) if rows else ""
        return records, report_date

    def build_workbook(
        self,
        request: SettlementRequest,
        records: list[LoanScheduleRecord],
        report_date: str,
    ) -> Workbook:
        report_code = self._report_code(request)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = report_code
        self._write_report_header(sheet, request, report_code, report_date)
        final_row = self._write_detail_rows(
            sheet,
            report_code,
            records,
            keep_unused_columns=not request.options.remove_unused_columns,
        )
        self._write_signatures_and_notes(sheet, request, report_code, report_date, final_row)
        if not request.options.remove_unused_columns:
            self._write_unused_source_headers(sheet, report_code)
        self._format_report(sheet, report_code, final_row, request.options)
        if request.options.create_control_sheet and report_code in {"15a", "15b"}:
            self._write_control_sheet(workbook, request, report_code, records)
        return workbook

    def _row_to_record(
        self,
        row: dict[str, str],
        source_headers: tuple[str, ...],
        report_code: str,
        branch_code: str,
        options: SettlementOptions,
    ) -> LoanScheduleRecord:
        use_lds = options.include_loan_deposit_schedule
        if report_code == "15b":
            prepaid = Decimal(0)
            accrued = parse_vietnamese_amount(row["SO_LAI_PHAI_THU_DEN_3112"])
            interest_account = row["TK_LAI_PHAI_THU"].strip()
            term_key = "KYHAN"
        else:
            prepaid = parse_vietnamese_amount(row["LAI_TRA_TRUOC_DEN_3112"])
            accrued = parse_vietnamese_amount(row["LAI_DU_THU_HET_3112"])
            interest_account = row.get("TK_LAI_DU_THU", "").strip()
            term_key = "KYHAN_THANG" if "KYHAN_THANG" in row else "KYHAN"
        last_interest_raw = row["NGAY_THU_LAI_CUOICUNG"]
        last_interest_date = parse_yyyymmdd(last_interest_raw)
        disbursement_raw = row["NGAY_GIAI_NGAN"]
        maturity_raw = row["NGAY_DAO_HAN"]
        disbursement_date = parse_yyyymmdd(disbursement_raw)
        maturity_date = parse_yyyymmdd(maturity_raw)
        prepaid_raw = (
            row["SO_LAI_PHAI_THU_DEN_3112"]
            if report_code == "15b"
            else row["LAI_TRA_TRUOC_DEN_3112"]
        )
        accrued_raw = (
            row["SO_LAI_PHAI_THU_DEN_3112"]
            if report_code == "15b"
            else row.get("LAI_DU_THU_HET_3112", "")
        )
        contract_number = (
            row["SO_GIAI_NGAN"].strip()
            if use_lds and row.get("SO_GIAI_NGAN")
            else row["SO_HOP_DONG"].strip()
        )
        return LoanScheduleRecord(
            account=(row["SO_HIEU_TK"].strip() or "211108"),
            customer_id=normalize_customer_id(
                row["MA_KH"],
                branch_code,
                include_branch=options.include_branch_in_customer_id,
            ),
            customer_name=row["TEN_KH"],
            contract_number=self._numeric_code(contract_number),
            currency=(
                row["CCY_THEO_LDS"]
                if use_lds and row.get("CCY_THEO_LDS")
                else row["CCY_THEOHOPDONG"]
            ),
            debt_group=row["NHOM_NO"],
            industry_code=self._numeric_code(row.get("MA_NGANH", "").strip()),
            industry_name=self._text_or_none(row.get("TEN_NGANH_KINH_TE", "")),
            capital_source=self._text_or_none(row.get("NGUON_VON", "")),
            loan_purpose=(
                self._text_or_none(row.get("MUC_DICH_SU_DUNG", ""))
                if report_code == "16"
                else self._text_or_none(row.get("TEN_NGANH_KINH_TE", ""))
            ),
            original_balance=parse_vietnamese_amount(row["SODU_NGUYENTE_TAI_3112"]),
            converted_balance=parse_vietnamese_amount(row["SODU_QUYDOI_VND_TAI_3112"]),
            term_months=row.get(term_key, ""),
            disbursement_date=disbursement_date,
            maturity_date=maturity_date,
            interest_rate=self._number_or_text(row["LAISUAT"]),
            last_interest_collection_date=last_interest_date,
            prepaid_interest=prepaid,
            accrued_interest=accrued,
            interest_account=self._numeric_code(interest_account),
            last_interest_collection_date_display=self._blank_preserving_display(
                last_interest_raw,
                last_interest_date,
            ),
            disbursement_date_display=self._blank_preserving_display(
                disbursement_raw,
                disbursement_date,
            ),
            maturity_date_display=self._blank_preserving_display(maturity_raw, maturity_date),
            prepaid_interest_display=self._blank_preserving_display(prepaid_raw, prepaid),
            accrued_interest_display=self._blank_preserving_display(accrued_raw, accrued),
            interest_account_display=self._blank_preserving_display(
                row.get("TK_LAI_PHAI_THU" if report_code == "15b" else "TK_LAI_DU_THU", ""),
                self._numeric_code(interest_account),
            ),
            source_values=tuple(row.get(header, "") for header in source_headers),
        )

    def _write_detail_rows(
        self,
        sheet,
        report_code: str,
        records: list[LoanScheduleRecord],
        *,
        keep_unused_columns: bool,
    ) -> int:
        if not records:
            return 10
        row = 11
        group_start = row
        current_key = self._group_key(report_code, records[0])
        subtotal_rows: list[int] = []
        for record in records:
            key = self._group_key(report_code, record)
            if key != current_key:
                self._write_subtotal(sheet, row, report_code, current_key, group_start, row - 1)
                subtotal_rows.append(row)
                row += 1
                group_start = row
                current_key = key
            self._write_record(
                sheet,
                row,
                report_code,
                record,
                keep_unused_columns=keep_unused_columns,
            )
            row += 1
        self._write_subtotal(sheet, row, report_code, current_key, group_start, row - 1)
        subtotal_rows.append(row)
        row += 1
        self._write_grand_total(sheet, row, report_code, subtotal_rows)
        return row

    def _write_record(
        self,
        sheet,
        row: int,
        report_code: str,
        record: LoanScheduleRecord,
        *,
        keep_unused_columns: bool,
    ) -> None:
        values = [
            self._numeric_code(record.account),
            record.customer_id,
            record.customer_name,
            record.contract_number,
            record.currency,
            record.debt_group,
        ]
        if report_code != "16":
            values.extend([record.industry_code, record.industry_name])
        values.extend([
            record.capital_source,
            record.loan_purpose,
            record.original_balance,
            record.converted_balance,
            record.term_months,
            record.disbursement_date_display,
            record.maturity_date_display,
            record.interest_rate,
            record.last_interest_collection_date_display,
            (
                record.accrued_interest_display
                if report_code == "15b"
                else record.prepaid_interest_display
            ),
        ])
        if report_code != "15b":
            values.append(record.accrued_interest_display)
        if report_code != "16":
            values.append(record.interest_account_display)
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
        if keep_unused_columns:
            for offset, value in enumerate(record.source_values, self._last_column(report_code) + 1):
                sheet.cell(row, offset, value)

    @staticmethod
    def _write_subtotal(
        sheet,
        row: int,
        report_code: str,
        key: str,
        start: int,
        end: int,
    ) -> None:
        if report_code == "16":
            sheet.cell(row, 2, f"Cộng KH: {key}")
        else:
            sheet.cell(row, 1, f"Cộng TK: {key}")
        for column in Mau1516Processor._subtotal_columns(report_code):
            letter = get_column_letter(column)
            sheet.cell(row, column, f"=SUM({letter}{start}:{letter}{end})")

    @staticmethod
    def _write_grand_total(
        sheet,
        row: int,
        report_code: str,
        subtotal_rows: list[int],
    ) -> None:
        sheet.cell(row, 1, "TỔNG CỘNG:")
        for column in Mau1516Processor._subtotal_columns(report_code):
            letter = get_column_letter(column)
            refs = "+".join(f"{letter}{subtotal_row}" for subtotal_row in subtotal_rows)
            sheet.cell(row, column, f"={refs}" if refs else 0)

    def _write_signatures_and_notes(
        self,
        sheet,
        request: SettlementRequest,
        report_code: str,
        report_date: str,
        final_row: int,
    ) -> None:
        base = final_row + 1
        location_date = f"{request.profile.report_location}, {report_date}" if report_date else ""
        sheet.cell(base, 13 if report_code == "16" else 16, location_date)
        sheet.cell(base + 1, 1, "LẬP BIỂU")
        sheet.cell(base + 2, 1, "(Ký, ghi rõ họ tên, số ĐT liên hệ)")
        sheet.cell(
            base + 8,
            1,
            f"({request.profile.report_preparer} - SĐT: {request.profile.phone})",
        )
        sheet.cell(base + 1, 4, "TRƯỞNG PHÒNG KHDN/KHCN")
        sheet.cell(base + 2, 4, "(Ký, ghi rõ họ tên)")
        accounting_col = 8 if report_code == "16" else 10
        director_col = 13 if report_code == "16" else 16
        sheet.cell(base + 1, accounting_col, "TRƯỞNG PHÒNG KẾ TOÁN")
        sheet.cell(base + 2, accounting_col, "(Ký, ghi rõ họ tên)")
        sheet.cell(base + 1, director_col, "GIÁM ĐỐC")
        sheet.cell(base + 2, director_col, "(Ký, đóng dấu, ghi rõ họ tên)")

        self._merge_if_unmerged(sheet, base + 1, 1, base + 1, 3)
        self._merge_if_unmerged(sheet, base + 2, 1, base + 2, 3)
        self._merge_if_unmerged(sheet, base + 8, 1, base + 8, 3)
        if report_code == "16":
            self._merge_if_unmerged(sheet, base, 13, base, 17)
            self._merge_if_unmerged(sheet, base + 1, 4, base + 1, 7)
            self._merge_if_unmerged(sheet, base + 2, 4, base + 2, 7)
            self._merge_if_unmerged(sheet, base + 1, 8, base + 1, 11)
            self._merge_if_unmerged(sheet, base + 2, 8, base + 2, 11)
            self._merge_if_unmerged(sheet, base + 1, 13, base + 1, 17)
            self._merge_if_unmerged(sheet, base + 2, 13, base + 2, 17)
        else:
            signature_end = 19 if report_code == "15b" else 20
            self._merge_if_unmerged(sheet, base + 1, 4, base + 1, 9)
            self._merge_if_unmerged(sheet, base + 2, 4, base + 2, 9)
            self._merge_if_unmerged(sheet, base, 16, base, signature_end)
            self._merge_if_unmerged(sheet, base + 1, 10, base + 1, 15)
            self._merge_if_unmerged(sheet, base + 2, 10, base + 2, 15)
            self._merge_if_unmerged(sheet, base + 1, 16, base + 1, signature_end)
            self._merge_if_unmerged(sheet, base + 2, 16, base + 2, signature_end)
        notes_start = base + 10
        sheet.cell(notes_start, 1, "Ghi chú:")
        if report_code == "15a":
            notes = (
                "1. Cột 12 được quy đổi từ cột 11 theo tỷ giá TSC thông báo ngày 31/12",
                "2. Cột 11, cột 12: Dòng tổng cộng khớp với tổng số dư Nợ của các tài khoản 21,22,23,24,25,26,27,28, 29 trên cân đối nguyên tệ và cân đối quy đổi.",
                '3. Cột 18: Dòng tổng cộng khớp với dòng "Lãi cho vay nhận trước" trên Sao kê chi tiết tài khoản 488 - Mẫu số 22/QT',
                "4. Cột 19: Dòng tổng cộng + lãi dự thu cho vay các TCTD khác (cột 12, Mẫu 11/QT) + Lãi dự thu cho vay TSC (cột 13, mẫu 26/QT) phải khớp với tổng số dư Nợ của tài khoản 394 - Lãi phải thu từ hoạt động tín dụng.",
                "5. Dòng Tổng cộng tính bằng công thức, không nhập số liệu trực tiếp",
            )
        elif report_code == "15b":
            notes = (
                "1. Cột 12 được quy đổi từ cột 11 theo tỷ giá TSC thông báo ngày 31/12",
                "2. Cột 18: Dòng tổng cộng khớp với số dư TK 941",
                "3. Dòng Tổng cộng tính bằng công thức, không nhập số liệu trực tiếp",
            )
        else:
            notes = (
                " Cột 10 được quy đổi từ cột 9 theo tỷ giá TSC thông báo ngày 31/12",
            )
        for offset, note in enumerate(notes, 1):
            target_row = notes_start if report_code == "16" else notes_start + offset
            sheet.cell(target_row, 2, note)

    def _write_report_header(
        self,
        sheet,
        request: SettlementRequest,
        report_code: str,
        report_date: str,
    ) -> None:
        profile = request.profile
        last_column = self._last_column(report_code)
        title = {
            "15a": " SAO KÊ CHI TIẾT TÀI KHOẢN CHO VAY KHÁCH HÀNG LÀ TỔ CHỨC, HỘ KINH DOANH VÀ CÁ NHÂN ",
            "15b": " SAO KÊ CHI TIẾT TÀI KHOẢN CHO VAY KHÁCH HÀNG LÀ TỔ CHỨC, HỘ KINH DOANH VÀ CÁ NHÂN CÓ LÃI PHẢI THU HOẠCH TOÁN NGOẠI BẢNG (TÀI KHOẢN 941)",
            "16": " SAO KÊ CHI TIẾT KHÁCH HÀNG CÓ DƯ NỢ TỪ 10 TỶ VIỆT NAM ĐỒNG TRỞ LÊN ",
        }[report_code]
        entries = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": "----------------------------------",
            "A4": f"Mã chi nhánh: {profile.branch_code}",
            "A5": f"Tên {profile.branch_name}",
            "A6": title,
            "A7": report_date,
        }
        note_column = 13 if report_code == "16" else 15
        entries[f"{get_column_letter(note_column)}1"] = f"1. Mẫu số {report_code}/QT"
        entries[f"{get_column_letter(note_column)}2"] = "2. CN loại I gửi file về TSC"
        entries[f"{get_column_letter(note_column)}3"] = "3. Lưu tại Chi nhánh"
        for address, value in entries.items():
            sheet[address] = value
        sheet.cell(8, last_column, "Đơn vị:VNĐ")
        for row in range(1, 6):
            self._merge_if_unmerged(sheet, row, 1, row, 4)
        self._merge_if_unmerged(sheet, 6, 1, 6, last_column)
        self._merge_if_unmerged(sheet, 7, 1, 7, last_column)
        for column, label in enumerate(self._headers(report_code), 1):
            sheet.cell(9, column, label)
            sheet.cell(10, column, column)

    def _write_unused_source_headers(self, sheet, report_code: str) -> None:
        for offset, header in enumerate(
            getattr(self, "_source_headers", ()),
            self._last_column(report_code) + 1,
        ):
            sheet.cell(10, offset, header)

    def _write_control_sheet(
        self,
        workbook: Workbook,
        request: SettlementRequest,
        report_code: str,
        records: list[LoanScheduleRecord],
    ) -> None:
        sheet = workbook.create_sheet("SoLieuTongHop")
        sheet.append((f"SỐ LIỆU TỔNG HỢP MẪU {report_code}",))
        include_accrual_accounts = request.options.include_accrual_accounts
        cash_groups: dict[str, list[LoanScheduleRecord]] = {}
        for record in records:
            currency = str(record.currency or "").strip() or "VND"
            key = f"{record.account}_{currency}"
            cash_groups.setdefault(key, []).append(record)
        accrual_accounts = sorted(
            {
                record.interest_account
                for record in records
                if include_accrual_accounts and record.interest_account
            },
            key=lambda value: (str(value).strip(),),
        )
        headers = [
            "Tài Khoản_VND",
            "Số Dư Nguyên Tệ",
            "Số Dư Qui Đổi VNĐ",
            "Số Lãi Trả Trước Đến Hết  31/12",
            (
                "Số Lãi Dự Thu Đến Hết 31/12"
                if report_code == "15a"
                else "SỐ LÃI PHẢI THU ĐẾN HẾT 31/12"
            ),
        ]
        headers.extend(accrual_accounts)
        sheet.append(headers)
        first_cash_row = 3
        for key in sorted(cash_groups):
            selected = cash_groups[key]
            row_values: list[object] = [
                key if key.upper().endswith("VND") else f"{key}_VND",
                sum((item.original_balance for item in selected), Decimal(0)),
                sum((item.converted_balance for item in selected), Decimal(0)),
                sum((item.prepaid_interest for item in selected), Decimal(0)),
                sum((item.accrued_interest for item in selected), Decimal(0)),
            ]
            for account in accrual_accounts:
                if report_code == "15a" and str(account).startswith("488"):
                    value = sum(
                        (
                            item.prepaid_interest
                            for item in selected
                            if item.interest_account == account
                        ),
                        Decimal(0),
                    )
                else:
                    value = sum(
                        (
                            item.accrued_interest
                            for item in selected
                            if item.interest_account == account
                        ),
                        Decimal(0),
                    )
                row_values.append(value)
            sheet.append(tuple(row_values))
        last_cash_row = sheet.max_row
        first_accrual_row = sheet.max_row + 1
        for account in accrual_accounts:
            selected = [item for item in records if item.interest_account == account]
            sheet.append(
                (
                    f"{account}_VND",
                    None,
                    None,
                    sum((item.prepaid_interest for item in selected), Decimal(0)),
                    sum((item.accrued_interest for item in selected), Decimal(0)),
                    *([Decimal(0)] * len(accrual_accounts)),
                )
            )
        last_accrual_row = sheet.max_row
        total_row = sheet.max_row + 1
        sheet.cell(total_row, 1, "Tổng cộng:")
        if accrual_accounts:
            sheet.cell(total_row, 2, f"=SUM(B{first_cash_row}:B{last_cash_row})")
            sheet.cell(total_row, 3, f"=SUM(C{first_cash_row}:C{last_cash_row})")
            sheet.cell(total_row, 4, f"=SUM(D{first_accrual_row}:D{last_accrual_row})")
            sheet.cell(total_row, 5, f"=SUM(E{first_accrual_row}:E{last_accrual_row})")
            for column in range(6, 6 + len(accrual_accounts)):
                letter = get_column_letter(column)
                sheet.cell(total_row, column, f"=SUM({letter}{first_cash_row}:{letter}{last_cash_row})")
        else:
            for column in range(2, 6):
                letter = get_column_letter(column)
                sheet.cell(total_row, column, f"=SUM({letter}3:{letter}{total_row - 1})")
        self._format_control_sheet(sheet, total_row)

    @staticmethod
    def _format_control_sheet(sheet, total_row: int) -> None:
        thin = Side(style="thin", color="000000")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=12)
                cell.alignment = Alignment(vertical="center")
        sheet["A1"].font = Font(name="Times New Roman", size=12, bold=True, color="000080")
        for cell in sheet[2]:
            cell.fill = PatternFill("solid", fgColor="FFFF00")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        max_column = sheet.max_column
        for row in sheet.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=max_column):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(3, total_row + 1):
            for column in range(2, max_column + 1):
                sheet.cell(row, column).number_format = "#,##0"
        sheet.row_dimensions[2].height = 50
        sheet.column_dimensions["A"].width = 18
        for column in range(2, max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 20

    def _format_report(
        self,
        sheet,
        report_code: str,
        last_detail_row: int,
        options: SettlementOptions,
    ) -> None:
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        last_column = self._last_column(report_code)
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=10)
                cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(min_row=9, max_row=last_detail_row, min_col=1, max_col=last_column):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(9, 11):
            for column in range(1, last_column + 1):
                cell = sheet.cell(row, column)
                cell.font = Font(name="Times New Roman", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet["A6"].font = Font(name="Times New Roman", size=13 if report_code == "15b" else 15, bold=True)
        sheet["A7"].font = Font(name="Times New Roman", size=16, bold=True, italic=True)
        for address in ("A1", "A2", "A4", "A5", "A6", "A7"):
            sheet[address].alignment = Alignment(horizontal="center", vertical="center")
        style_agency_header(sheet, start_column=1, end_column=4)
        style_currency_unit(sheet.cell(8, last_column))
        widths = self._column_widths(report_code, options)
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = width
        if not options.remove_unused_columns:
            for column in range(last_column + 1, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(column)].width = 15
        for row in range(11, last_detail_row + 1):
            for column in self._amount_columns(report_code):
                sheet.cell(row, column).number_format = "#,##0"
            for column in self._date_columns(report_code):
                sheet.cell(row, column).number_format = (
                    "dd/mm/yyyy" if options.four_digit_year else "dd/mm/yy"
                )
            if not self._is_total_row(sheet, row):
                for column in self._text_wrap_columns(report_code):
                    sheet.cell(row, column).alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )
                if self._row_needs_wrap(sheet, row, report_code):
                    sheet.row_dimensions[row].height = self._wrapped_row_height(
                        sheet,
                        row,
                        report_code,
                    )
            if self._is_total_row(sheet, row):
                for column in range(1, last_column + 1):
                    sheet.cell(row, column).font = Font(name="Times New Roman", size=9, bold=True)
                value_a = str(sheet.cell(row, 1).value or "")
                if value_a.startswith("TỔNG CỘNG:"):
                    for column in range(1, last_column + 1):
                        sheet.cell(row, column).border = Border(
                            left=medium, right=medium, top=medium, bottom=medium
                        )
        for row in range(last_detail_row + 1, sheet.max_row + 1):
            for column in range(1, last_column + 1):
                sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
        for row in (last_detail_row + 2,):
            for column in range(1, last_column + 1):
                sheet.cell(row, column).font = Font(name="Times New Roman", bold=True)
        for row in (last_detail_row + 3, last_detail_row + 9):
            for column in range(1, last_column + 1):
                sheet.cell(row, column).font = Font(name="Times New Roman", italic=True)
        for row in range(last_detail_row + 11, sheet.max_row + 1):
            sheet.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[6].height = 30 if report_code == "15b" else 24
        sheet.row_dimensions[9].height = 72
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "B11"
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:{get_column_letter(last_column)}{sheet.max_row}",
            title_rows="9:10",
        )

    @staticmethod
    def _required_headers(report_code: str) -> set[str]:
        common = {
            "NGAY", "NHOM_NO", "SO_HIEU_TK", "MA_KH", "TEN_KH",
            "SO_HOP_DONG", "CCY_THEOHOPDONG", "SO_GIAI_NGAN", "CCY_THEO_LDS",
            "SODU_NGUYENTE_TAI_3112", "SODU_QUYDOI_VND_TAI_3112",
            "NGAY_GIAI_NGAN", "NGAY_DAO_HAN", "LAISUAT",
            "NGAY_THU_LAI_CUOICUNG", "NGUON_VON",
        }
        if report_code == "15a":
            return common | {
                "KYHAN_THANG", "LAI_TRA_TRUOC_DEN_3112",
                "LAI_DU_THU_HET_3112", "MA_NGANH", "TEN_NGANH_KINH_TE",
                "NGHIEP_VU", "TK_LAI_DU_THU",
            }
        if report_code == "15b":
            return common | {
                "KYHAN", "SO_LAI_PHAI_THU_DEN_3112", "MA_NGANH",
                "TEN_NGANH_KINH_TE", "NGHIEP_VU", "TK_LAI_PHAI_THU",
            }
        return common | {
            "KYHAN", "LAI_TRA_TRUOC_DEN_3112", "LAI_DU_THU_HET_3112",
            "MUC_DICH_SU_DUNG",
        }

    @staticmethod
    def _headers(report_code: str) -> tuple[str, ...]:
        common = (
            "SỐ HIỆU TÀI KHOẢN",
            "MÃ KHÁCH HÀNG",
            "TÊN KHÁCH HÀNG",
            "SỐ KHẾ ƯỚC/HỢP ĐỒNG",
            "LOẠI TIỀN TỆ",
            "NHÓM NỢ",
        )
        if report_code == "16":
            middle = ("NGUỒN VỐN", "MỤC ĐÍCH VAY")
        else:
            middle = ("MÃ NGÀNH", "TÊN NGÀNH", "NGUỒN VỐN", "MỤC ĐÍCH VAY")
        tail = (
            "SỐ DƯ NGUYÊN TỆ TẠI THỜI ĐIỂM 31/12",
            "SỐ DƯ QUY ĐỔI VNĐ TẠI THỜI ĐIỂM 31/12",
            "KỲ HẠN (tháng)",
            "NGÀY GIẢI NGÂN",
            "NGÀY ĐÁO HẠN",
            "LÃI SUẤT (%/năm)",
            "NGÀY THU LÃI CUỐI CÙNG",
        )
        if report_code == "15b":
            return common + middle + tail + (
                "SỐ LÃI PHẢI THU ĐẾN HẾT 31/12",
                "TÀI KHOẢN LÃI PHẢI THU ",
            )
        if report_code == "16":
            return common + middle + tail + (
                "SỐ LÃI NHẬN TRƯỚC ĐẾN 31/12",
                "SỐ LÃI DỰ THU ĐẾN HẾT 31/12",
            )
        return common + middle + tail + (
            "SỐ LÃI NHẬN TRƯỚC ĐẾN 31/12",
            "SỐ LÃI DỰ THU ĐẾN HẾT 31/12",
            "TÀI KHOẢN LÃI DỰ THU ",
        )

    @staticmethod
    def _last_column(report_code: str) -> int:
        if report_code == "16":
            return 17
        if report_code == "15b":
            return 19
        return 20

    @staticmethod
    def _amount_columns(report_code: str) -> tuple[int, ...]:
        if report_code == "15b":
            return (11, 12, 18)
        if report_code == "16":
            return (9, 10, 16, 17)
        return (11, 12, 18, 19)

    @staticmethod
    def _subtotal_columns(report_code: str) -> tuple[int, ...]:
        if report_code == "15b":
            return (11, 12, 18)
        if report_code == "16":
            return (9, 10, 16, 17)
        return (11, 12, 18, 19)

    @staticmethod
    def _date_columns(report_code: str) -> tuple[int, ...]:
        return (14, 15, 17) if report_code != "16" else (12, 13, 15)

    @staticmethod
    def _text_wrap_columns(report_code: str) -> tuple[int, ...]:
        return (3, 8) if report_code == "16" else (3, 8, 10)

    @staticmethod
    def _is_total_row(sheet, row: int) -> bool:
        value_a = str(sheet.cell(row, 1).value or "")
        value_b = str(sheet.cell(row, 2).value or "")
        return value_a.startswith(("Cộng TK:", "TỔNG CỘNG:")) or value_b.startswith("Cộng KH:")

    @classmethod
    def _row_needs_wrap(cls, sheet, row: int, report_code: str) -> bool:
        thresholds = {3: 22, 8: 20, 10: 20}
        if report_code == "16":
            thresholds = {3: 22, 8: 22}
        for column in cls._text_wrap_columns(report_code):
            value = sheet.cell(row, column).value
            if len(str(value or "")) > thresholds[column]:
                return True
        return False

    @classmethod
    def _wrapped_row_height(cls, sheet, row: int, report_code: str) -> float:
        estimated_lines = 1
        for column in cls._text_wrap_columns(report_code):
            value = str(sheet.cell(row, column).value or "")
            if not value:
                continue
            width = sheet.column_dimensions[get_column_letter(column)].width or 12
            chars_per_line = max(8, int(width * 1.05))
            estimated_lines = max(
                estimated_lines,
                (len(value) + chars_per_line - 1) // chars_per_line,
            )
        return max(22, min(78, estimated_lines * 13.5))

    @staticmethod
    def _column_widths(report_code: str, options: SettlementOptions) -> dict[int, float]:
        widths = {
            1: 8.9,
            2: 13 if options.include_branch_in_customer_id else 9,
            3: 24,
            4: 18,
            5: 6.2,
            6: 4.9,
            7: 7.5,
            8: 13.7,
            9: 13.7,
            10: 13.7,
            11: 15,
            12: 15,
            13: 4.57,
            14: 9.2 if options.four_digit_year else 7.2,
            15: 9.2 if options.four_digit_year else 7.2,
            16: 5.8,
            17: 9.2 if options.four_digit_year else 7.2,
            18: 12,
            19: 12,
            20: 12,
        }
        if report_code == "16":
            return {
                1: 8.9,
                2: 13 if options.include_branch_in_customer_id else 9,
                3: 24,
                4: 18,
                5: 6.2,
                6: 4.9,
                7: 20,
                8: 20,
                9: 15,
                10: 15,
                11: 4.57,
                12: 9.2 if options.four_digit_year else 7.2,
                13: 9.2 if options.four_digit_year else 7.2,
                14: 5.8,
                15: 9.2 if options.four_digit_year else 7.2,
                16: 12,
                17: 17,
            }
        return widths

    @staticmethod
    def _group_key(report_code: str, record: LoanScheduleRecord) -> str:
        return record.customer_id if report_code == "16" else record.account

    @staticmethod
    def _sort_debt_group(value: str) -> tuple[int, str]:
        text = str(value or "").strip().lstrip("'")
        if not text:
            return (-1, "")
        return (int(text), text) if text.isdigit() else (999, text)

    @staticmethod
    def _numeric_code(value: Any) -> object:
        text = str(value or "").strip()
        if not text:
            return None
        return int(text) if text.isdigit() else value

    @staticmethod
    def _number_or_text(value: Any) -> object:
        text = str(value or "").strip()
        if not text:
            return value
        try:
            number = Decimal(text)
        except Exception:
            return value
        return int(number) if number == number.to_integral_value() else float(number)

    @staticmethod
    def _blank_preserving_display(raw_value: Any, parsed_value: object) -> object:
        text = str(raw_value or "")
        return raw_value if text and not text.strip() else parsed_value

    @staticmethod
    def _text_or_none(raw_value: Any) -> object:
        return None if raw_value == "" else raw_value

    @staticmethod
    def _report_code(request: SettlementRequest) -> str:
        code = request.spec.report_code.casefold()
        if code not in {"15a", "15b", "16"}:
            raise SettlementError(
                f"Processor mau15_16 không hỗ trợ Mẫu {request.spec.report_code}/QT.",
                code="unsupported_report",
            )
        return code

    @staticmethod
    def _output_prefix(options: SettlementOptions) -> str:
        prefix = (options.output_prefix or "QT").strip().upper()
        return "BN" if prefix == "BN" else "QT"

    @staticmethod
    def _merge_if_unmerged(sheet, start_row: int, start_column: int, end_row: int, end_column: int) -> None:
        cell_range = (
            f"{get_column_letter(start_column)}{start_row}:"
            f"{get_column_letter(end_column)}{end_row}"
        )
        if cell_range not in {str(merged) for merged in sheet.merged_cells.ranges}:
            sheet.merge_cells(cell_range)
