from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import StringIO
from pathlib import Path
from copy import copy
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.theme import theme_xml

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementRequest, SettlementResult
from agribank_v3.settlement.processors.formatting import setup_a4_print_layout, style_agency_header


@dataclass(frozen=True, slots=True)
class AllocationRecord:
    customer_id: str
    customer_name: str
    deferred_account: str
    income_expense_account: str
    income_expense_name: str
    reference: str
    currency: str
    open_date: datetime | None
    maturity_date: datetime | None
    allocation_date: datetime | None
    rate: Decimal
    principal: Decimal
    remaining_original: Decimal
    remaining_vnd: Decimal
    main_group: str
    sub_group: str


class Mau22Processor:
    """Processor for Mẫu 22/QT from glst34 deferred income/expense files."""

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if not request.source_paths:
            raise SettlementError("Mẫu 22/QT cần ít nhất một file nguồn.")
        records = self.read_sources(request)
        workbook = self.build_workbook(request, records)
        output_path = request.source_paths[0].with_name(
            f"{request.profile.branch_code.strip()}{request.options.output_prefix}22.xlsx"
        )
        workbook.save(output_path)
        workbook.close()
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name="22",
            processed_rows=len(records),
            warnings=(
                'Yêu cầu: Tại cột B nếu chưa có "Họ và Tên" khách hàng thì tự bổ sung, vì dữ liệu xuất ra không có.',
            ),
        )

    def read_sources(self, request: SettlementRequest) -> list[AllocationRecord]:
        records: list[AllocationRecord] = []
        for source_path in request.source_paths:
            records.extend(self._read_source(request, source_path))
        return sorted(records, key=lambda item: (item.main_group, item.sub_group, item.deferred_account, item.income_expense_account, item.reference))

    def build_workbook(self, request: SettlementRequest, records: list[AllocationRecord]) -> Workbook:
        workbook = Workbook()
        workbook.loaded_theme = self._times_new_roman_theme()
        workbook._named_styles["Normal"].font = Font(name="Times New Roman", size=10)
        sheet = workbook.active
        sheet.title = "22"
        self._write_header(sheet, request)
        current_row = self._write_table(sheet, records)
        self._write_signatures_and_notes(sheet, request, current_row + 1)
        self._format_report(sheet, current_row)
        return workbook

    def _read_source(self, request: SettlementRequest, source_path: Path) -> list[AllocationRecord]:
        if source_path.suffix.casefold() not in {".xls", ".xlsx", ".xlsm"}:
            raise SettlementError("Nguồn mẫu 22 phải là file Excel xuất từ glst34.")
        try:
            import xlrd
            workbook = xlrd.open_workbook(str(source_path), logfile=StringIO())
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            raise SettlementError(f"Không thể đọc file {source_path.name}: {exc}") from exc

        headers = [str(sheet.cell_value(0, col)).strip().casefold() for col in range(sheet.ncols)]
        required = {
            "custseq", "custnm", "acctcd", "adjacctcd", "refno", "ccy",
            "opndt", "matdt", "acrdt", "intrt", "acrbamt", "nextacramt", "nextbceqa",
        }
        missing = sorted(required - set(headers))
        if missing:
            raise SettlementError(f"File {source_path.name} thiếu cột: {', '.join(missing)}")
        indexes = {name: headers.index(name) for name in headers}
        records: list[AllocationRecord] = []
        for row_index in range(1, sheet.nrows):
            row = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
            remaining_original = self._decimal(row[indexes["nextacramt"]])
            remaining_vnd = self._decimal(row[indexes["nextbceqa"]])
            if remaining_original == 0 and remaining_vnd == 0:
                continue
            acct = self._text(row[indexes["acctcd"]])
            deferred = self._text(row[indexes["adjacctcd"]])
            account_name = self._text(row[indexes["acctnm"]])
            group = self._group_for(deferred, acct, account_name)
            if group is None:
                continue
            records.append(
                AllocationRecord(
                    customer_id=f"{request.profile.branch_code.strip()}{self._text(row[indexes['custseq']]).zfill(9)}",
                    customer_name=self._text(row[indexes["custnm"]]),
                    deferred_account=deferred,
                    income_expense_account=self._output_income_expense_account(acct),
                    income_expense_name=account_name,
                    reference=self._text(row[indexes["refno"]]).replace("-", ""),
                    currency=self._text(row[indexes["ccy"]]),
                    open_date=self._excel_date(row[indexes["opndt"]], workbook.datemode),
                    maturity_date=self._excel_date(row[indexes["matdt"]], workbook.datemode),
                    allocation_date=self._excel_date(row[indexes["acrdt"]], workbook.datemode),
                    rate=self._decimal(row[indexes["intrt"]]),
                    principal=self._decimal(row[indexes["acrbamt"]]),
                    remaining_original=remaining_original,
                    remaining_vnd=remaining_vnd,
                    main_group=group[0],
                    sub_group=group[1],
                )
            )
        return records

    def _write_header(self, sheet, request: SettlementRequest) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        values = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A4": f"Mã chi nhánh: {profile.branch_code.strip()}",
            "A5": f"Tên chi nhánh: {branch_name.removeprefix('Chi nhánh ').strip()}",
            "A6": "SAO KÊ CHI TIẾT SỐ DƯ TÀI KHOẢN DOANH THU VÀ CHI PHÍ CHỜ PHÂN BỔ",
            "A7": self._report_date_text(request),
            "L1": "1. Mẫu số 22/QT",
            "L2": "2. CN loại I gửi file về TSC",
            "L3": "3. Lưu tại chi nhánh",
            "M8": "Đơn vị : VNĐ",
        }
        for cell, value in values.items():
            sheet[cell] = value
        self._apply_times_new_roman(sheet, min_row=1, min_col=1, max_row=11, max_col=13)
        for ref in ("A1:D1", "A2:D2", "A3:D3", "A4:D4", "A5:D5", "A6:M6", "A7:M7"):
            sheet.merge_cells(ref)
        style_agency_header(sheet, start_column=1, end_column=4)
        sheet["A6"].font = Font(name="Times New Roman", size=14, bold=True)
        sheet["A7"].font = Font(name="Times New Roman", size=12, bold=True, italic=True)
        for address in ("A6", "A7"):
            sheet[address].alignment = Alignment(horizontal="center", vertical="center")
        for row in range(1, 4):
            sheet.cell(row, 12).font = Font(name="Times New Roman", size=8)
            sheet.cell(row, 12).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sheet["M8"].font = Font(name="Times New Roman", size=10, italic=True)
        sheet["M8"].alignment = Alignment(horizontal="right", vertical="center")
        self._write_table_headers(sheet)

    def _write_table_headers(self, sheet) -> None:
        headers = [
            "MÃ KH",
            "TÊN KHÁCH HÀNG",
            "TK DOANH THU (HOẶC CHI PHÍ) CHỜ PHÂN BỔ",
            "TK THU NHẬP (HOẶC CHI PHÍ)",
            "SỐ THAM CHIẾU",
            "LOẠI TIỀN TỆ",
            "NGÀY MỞ",
            "NGÀY ĐẾN HẠN",
            "NGÀY PHÂN BỔ",
            "LÃI SUẤT",
            "SỐ TIỀN GỐC",
            "SỐ TIỀN CÒN PHẢI PHÂN BỔ CHO KỲ SAU (NGUYÊN TỆ)",
            "SỐ TIỀN CÒN PHẢI PHÂN BỔ CHO KỲ SAU (QUY ĐỔI VND)",
        ]
        for column, header in enumerate(headers, start=1):
            sheet.cell(9, column, header)
            sheet.cell(11, column, column)

    def _write_table(self, sheet, records: list[AllocationRecord]) -> int:
        row = 12
        groups = (
            ("A", "DOANH THU CHỜ PHÂN BỔ", (("I", "Lãi cho vay nhận trước", "A1"), ("II", "Lãi chứng khoán nhận trước", "A2"), ("III", "Doanh thu chờ phân bổ khác", "A3"))),
            ("B", "CHI PHÍ CHỜ PHÂN BỔ", (("I", "Lãi huy động trả trước", "B1"), ("II", "Chi phí chờ phân bổ về CCDC", "B2"), ("III", "Chi phí chờ phân bổ khác", "B3"))),
        )
        first_data_row: int | None = None
        last_data_row: int | None = None
        for main_code, main_label, subgroups in groups:
            self._write_group_row(sheet, row, main_code, main_label)
            row += 1
            for subgroup_code, subgroup_label, key in subgroups:
                self._write_group_row(sheet, row, subgroup_code, subgroup_label)
                row += 1
                subgroup_records = [record for record in records if record.sub_group == key]
                for record in subgroup_records:
                    if first_data_row is None:
                        first_data_row = row
                    last_data_row = row
                    self._write_record(sheet, row, record)
                    row += 1
        total_row = row
        sheet.cell(total_row, 2, "Cộng")
        if first_data_row is not None and last_data_row is not None:
            for column in (11, 12, 13):
                letter = get_column_letter(column)
                sheet.cell(total_row, column, f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
        return total_row

    def _write_group_row(self, sheet, row: int, code: str, label: str) -> None:
        sheet.cell(row, 1, code)
        sheet.cell(row, 2, label)
        for column in range(1, 14):
            sheet.cell(row, column).font = Font(name="Times New Roman", size=10, bold=True)

    def _write_record(self, sheet, row: int, record: AllocationRecord) -> None:
        values = (
            record.customer_id,
            record.customer_name,
            record.deferred_account,
            record.income_expense_account,
            record.reference,
            record.currency,
            record.open_date,
            record.maturity_date,
            record.allocation_date,
            self._number(record.rate),
            self._number(record.principal),
            self._number(record.remaining_original),
            self._number(record.remaining_vnd),
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)

    def _write_signatures_and_notes(self, sheet, request: SettlementRequest, start_row: int) -> None:
        location = request.profile.report_location.strip()
        report_date = self._report_date_text(request)
        date_text = f"{location}, {report_date}" if location else report_date
        self._apply_times_new_roman(sheet, min_row=start_row, min_col=1, max_row=start_row + 9, max_col=13)
        sheet.merge_cells(start_row=start_row, start_column=11, end_row=start_row, end_column=13)
        sheet.cell(start_row, 11, date_text)
        sheet.cell(start_row, 11).font = Font(name="Times New Roman", size=10, italic=True)
        signature_row = start_row + 1
        note_row = start_row + 2
        blocks = ((1, 2, "LẬP BIỂU", "(Ký, ghi rõ họ tên, số ĐT liên hệ)"),
                  (3, 6, "TRƯỞNG PHÒNG KẾ TOÁN", "(Ký, ghi rõ họ tên)"),
                  (7, 10, "TRƯỞNG PHÒNG KTKSNB", "(Ký, ghi rõ họ tên)"),
                  (11, 13, "GIÁM ĐỐC", "(Ký, đóng dấu, ghi rõ họ tên)"))
        for start_col, end_col, title, note in blocks:
            sheet.merge_cells(start_row=signature_row, start_column=start_col, end_row=signature_row, end_column=end_col)
            sheet.merge_cells(start_row=note_row, start_column=start_col, end_row=note_row, end_column=end_col)
            sheet.cell(signature_row, start_col, title)
            sheet.cell(note_row, start_col, note)
            sheet.cell(signature_row, start_col).font = Font(name="Times New Roman", size=10, bold=True)
            sheet.cell(note_row, start_col).font = Font(name="Times New Roman", size=10, italic=True)
        for row in range(start_row, note_row + 1):
            for column in range(1, 14):
                sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
        preparer = request.profile.report_preparer.strip()
        phone = request.profile.phone.strip()
        if preparer or phone:
            sheet.merge_cells(start_row=start_row + 6, start_column=1, end_row=start_row + 6, end_column=2)
            sheet.cell(start_row + 6, 1, f"{preparer} ( SĐT {phone})" if phone else preparer)
            sheet.cell(start_row + 6, 1).font = Font(name="Times New Roman", size=10, italic=True)
        notes = [
            "Ghi chú: 1. Cột 13: Dòng cộng tài khoản phải khớp với số dư TK 488 và 388 trên cân đối ngày 31/12.",
            "                 2. Cột 7,8,9: Ngày ghi theo thứ tự ngày/tháng/năm (không ghi tháng trước ngày sau).",
            "                 3. Số tham chiếu: Số tài khoản tiền gửi của khách hàng, số giải ngân, …",
        ]
        for offset, text in enumerate(notes, start=7):
            sheet.cell(start_row + offset, 1, text)

    def _format_report(self, sheet, total_row: int) -> None:
        self._apply_times_new_roman(sheet)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(9, total_row + 1):
            for column in range(1, 14):
                cell = sheet.cell(row, column)
                cell.border = border
                font = copy(cell.font)
                font.name = "Times New Roman"
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=row == 9)
                if row >= 12 and column in (2,):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                if row >= 12 and column in (11, 12, 13):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
        for column in (7, 8, 9):
            for row in range(12, total_row):
                sheet.cell(row, column).number_format = "dd/mm/yyyy"
        sheet.row_dimensions[9].height = 76
        widths = {"A": 14, "B": 28, "C": 12, "D": 12, "E": 18, "F": 8, "G": 12, "H": 12, "I": 12, "J": 8, "K": 14, "L": 16, "M": 16}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        fill = PatternFill("solid", fgColor="FFFF99")
        for column in range(1, 14):
            sheet.cell(total_row, column).font = Font(name="Times New Roman", size=10, bold=True)
            sheet.cell(total_row, column).fill = fill
        setup_a4_print_layout(sheet, print_area=f"A1:M{sheet.max_row}", title_rows="$9:$11")

    def _group_for(self, deferred_account: str, account: str, account_name: str) -> tuple[str, str] | None:
        account_prefix = account[:3]
        if deferred_account.startswith("488"):
            if account_prefix == "702":
                return ("A", "A1")
            if account_prefix == "703":
                return ("A", "A2")
            return ("A", "A3")
        if deferred_account.startswith("388"):
            if account_prefix == "801":
                return ("B", "B1")
            if "công cụ lao động" in account_name.casefold():
                return ("B", "B2")
            return ("B", "B3")
        return None

    def _apply_times_new_roman(
        self,
        sheet,
        *,
        min_row: int | None = None,
        min_col: int | None = None,
        max_row: int | None = None,
        max_col: int | None = None,
    ) -> None:
        for row in sheet.iter_rows(
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col,
        ):
            for cell in row:
                font = copy(cell.font)
                font.name = "Times New Roman"
                cell.font = font

    def _output_income_expense_account(self, account: str) -> str:
        if account.startswith("875"):
            return "874" + account[3:]
        return account

    def _report_date_text(self, request: SettlementRequest) -> str:
        today = date.today()
        if request.options.output_prefix == "BN":
            return f"ngày 30 tháng 6 năm {today.year}"
        return f"ngày 31 tháng 12 năm {today.year}"

    def _excel_date(self, value: Any, datemode: int) -> datetime | None:
        try:
            serial = float(value)
        except (TypeError, ValueError):
            return None
        if serial < 1000:
            return None
        try:
            import xlrd
            return xlrd.xldate_as_datetime(serial, datemode)
        except Exception:
            return None

    def _decimal(self, value: Any) -> Decimal:
        if value in (None, ""):
            return Decimal(0)
        try:
            return Decimal(str(value).replace(",", ""))
        except (InvalidOperation, ValueError):
            return Decimal(0)

    def _number(self, value: Decimal) -> int | float:
        return int(value) if value == value.to_integral_value() else float(value)

    def _text(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return text[:-2] if text.endswith(".0") else text

    def _times_new_roman_theme(self) -> str | bytes:
        if isinstance(theme_xml, bytes):
            return theme_xml.replace(b"Calibri", b"Times New Roman")
        return theme_xml.replace("Calibri", "Times New Roman")
