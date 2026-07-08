from __future__ import annotations

from collections import OrderedDict
import csv
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
import gc
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import (
    SettlementOptions,
    SettlementRequest,
    SettlementResult,
)
from agribank_v3.settlement.processors.formatting import (
    setup_a4_print_layout,
    style_agency_header,
    style_currency_unit,
)
from agribank_v3.settlement.transforms import parse_yyyymmdd, vietnamese_report_date


@dataclass(frozen=True, slots=True)
class ReceivablePayableRecord:
    transaction_date_text: str
    branch_code: str
    account: str
    account_name: str
    counterparty_code: str
    counterparty_name: str
    reason: str
    arising_year: date | None
    currency: str
    original_balance: Decimal
    converted_balance: Decimal
    extra_values: tuple[object, ...]

    @property
    def group_account(self) -> str:
        return self.account

    @property
    def control_key(self) -> str:
        return f"{self.account}_{self.currency}"


class Mau24Processor:
    """Processor for Mẫu 24/QT - phải thu, phải trả."""

    REQUIRED_HEADERS = {
        "NGAY_GIAO_DICH",
        "MA_CN",
        "TAI_KHOAN",
        "TEN_TAI_KHOAN",
        "MA_DOI_TUONG_PHAI_THU_PHAI_TRA",
        "TEN_DOI_TUONG",
        "LY_DO_HACH_TOAN",
        "NAM_PHAT_SINH",
        "LOAI_TIEN",
        "SO_DU_NGUYENTE_TAI_NGAY_3112",
        "SO_DU_QUYDOI_VND_TAI_NGAY_3112",
    }
    EXTRA_HEADERS = ()
    def execute(self, request: SettlementRequest) -> SettlementResult:
        if len(request.source_paths) != 1:
            raise SettlementError(
                "Mẫu 24/QT cần đúng một file CSV nguồn.",
                code="invalid_source_count",
            )
        source_path = request.source_paths[0]
        records, report_date = self.read_source(request, source_path)
        processed_rows = len(records)
        workbook = self.build_workbook(request, records, report_date)
        output_path = source_path.with_name(
            f"{request.profile.branch_code.strip()}{self._output_prefix(request.options)}24a.xlsx"
        )
        try:
            workbook.save(output_path)
        finally:
            workbook.close()
        del workbook
        del records
        gc.collect()
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name="24",
            processed_rows=processed_rows,
            warnings=(
                "Yêu cầu: tại cột 4, người dùng phải nhập lý do đã hạch toán vào.",
            ),
        )

    def read_source(
        self,
        request: SettlementRequest,
        source_path: Path,
    ) -> tuple[list[ReceivablePayableRecord], str]:
        if source_path.suffix.casefold() not in {".csv", ".txt"}:
            raise SettlementError(
                "Nguồn Mẫu 24/QT phải là file CSV.",
                code="invalid_source_type",
            )
        records: list[ReceivablePayableRecord] = []
        report_date = ""
        try:
            with source_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                headers = set(reader.fieldnames or ())
                missing = sorted(self.REQUIRED_HEADERS - headers)
                if missing:
                    raise SettlementError(
                        f"File nguồn thiếu cột: {', '.join(missing)}",
                        code="invalid_mau24_headers",
                    )
                extra_headers = tuple(
                    header
                    for header in (reader.fieldnames or ())
                    if header not in self.REQUIRED_HEADERS
                )
                for row in reader:
                    if not report_date:
                        report_date = vietnamese_report_date(row["NGAY_GIAO_DICH"])
                    records.append(self._row_to_record(row, extra_headers))
        except SettlementError:
            raise
        except (OSError, UnicodeError, csv.Error) as exc:
            raise SettlementError(f"Không thể đọc file nguồn: {exc}") from exc

        records.sort(
            key=lambda record: (
                record.group_account,
                record.arising_year or date.min,
            )
        )
        return records, report_date

    def build_workbook(
        self,
        request: SettlementRequest,
        records: list[ReceivablePayableRecord],
        report_date: str,
    ) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "24"
        self._write_header(sheet, request, report_date)
        final_row, summary_rows = self._write_rows(
            sheet,
            records,
            keep_unused=not request.options.remove_unused_columns,
        )
        self._write_signatures_and_notes(sheet, request, report_date, final_row)
        self._format_report(
            sheet,
            request.options,
            final_row,
            summary_rows,
        )
        if request.options.create_control_sheet:
            self._write_control_sheet(workbook, records)
        workbook.calculation.fullCalcOnLoad = False
        workbook.calculation.forceFullCalc = False
        workbook.calculation.calcMode = "auto"
        return workbook

    def _write_header(
        self,
        sheet,
        request: SettlementRequest,
        report_date: str,
    ) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        for address in ("A1:B1", "A2:B2", "A3:B3", "A4:B4", "A6:H6", "A7:H7"):
            sheet.merge_cells(address)
        sheet["A1"] = "NGÂN HÀNG NÔNG NGHIỆP"
        sheet["A2"] = "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM"
        sheet["A3"] = f"Mã chi nhánh: {profile.branch_code.strip()}"
        sheet["A4"] = f"Tên {branch_name}"
        style_agency_header(sheet, start_column=1, end_column=2)
        sheet["A6"] = "SAO KÊ CHI TIẾT TÀI KHOẢN PHẢI THU, PHẢI TRẢ"
        sheet["A7"] = report_date
        sheet["G1"] = "1. Mẫu số 24a/QT"
        sheet["G2"] = "2. CN loại I gửi file về TSC"
        sheet["G3"] = "3. Lưu tại Chi nhánh"
        for row in range(1, 4):
            cell = sheet.cell(row, 7)
            cell.font = Font(name="Times New Roman", size=8)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet["H8"] = " Đơn vị: VNĐ "
        style_currency_unit(sheet["H8"])

        headers = (
            "TÀI KHOẢN",
            "TÊN TÀI KHOẢN",
            "TÊN ĐỐI TƯỢNG PHẢI THU/PHẢI TRẢ",
            "LÝ DO HẠCH TOÁN",
            "NĂM PHÁT SINH",
            "LOẠI TIỀN",
            "SỐ DƯ TẠI NGÀY 31/12 (nguyên tệ)",
            "SỐ DƯ TẠI NGÀY 31/12 (quy đổi VNĐ)",
        )
        for column, value in enumerate(headers, start=1):
            sheet.cell(9, column).value = value
            sheet.cell(10, column).value = column
            sheet.cell(10, column).number_format = "(0)"

    def _write_rows(
        self,
        sheet,
        records: list[ReceivablePayableRecord],
        *,
        keep_unused: bool,
    ) -> tuple[int, set[int]]:
        grouped: OrderedDict[str, list[ReceivablePayableRecord]] = OrderedDict()
        for record in records:
            grouped.setdefault(record.group_account, []).append(record)
        row = 11
        summary_rows: set[int] = set()
        for group_account, group_records in grouped.items():
            start_row = row
            for record in group_records:
                self._write_record(sheet, row, record, keep_unused=keep_unused)
                row += 1
            sheet.cell(row, 1).value = f"Cộng TK {group_account}"
            sheet.cell(row, 8).value = f"=SUM(H{start_row}:H{row - 1})"
            summary_rows.add(row)
            row += 1
        return max(10, row - 1), summary_rows

    def _write_record(
        self,
        sheet,
        row: int,
        record: ReceivablePayableRecord,
        *,
        keep_unused: bool,
    ) -> None:
        values = (
            self._numeric_code(record.account),
            record.account_name,
            record.counterparty_name,
            record.reason or None,
            record.arising_year,
            record.currency,
            self._number_to_cell(record.original_balance),
            self._number_to_cell(record.converted_balance),
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column).value = value
        if keep_unused:
            for column, value in enumerate(record.extra_values, start=9):
                sheet.cell(row, column).value = value or None

    def _write_signatures_and_notes(
        self,
        sheet,
        request: SettlementRequest,
        report_date: str,
        final_row: int,
    ) -> None:
        profile = request.profile
        date_row = final_row + 1
        title_row = final_row + 2
        note_row = final_row + 3
        preparer_row = final_row + 9
        notes_row = final_row + 11
        location = profile.report_location.strip()
        sheet.merge_cells(
            start_row=date_row,
            start_column=5,
            end_row=date_row,
            end_column=8,
        )
        sheet.cell(date_row, 5).value = (
            f"{location}, {report_date}" if location else report_date
        )
        for start_column, end_column, title, note in (
            (1, 2, "LẬP BIỂU", "(Ký, ghi rõ họ tên, số ĐT liên hệ)"),
            (3, 4, "TRƯỞNG PHÒNG KẾ TOÁN", "(Ký, ghi rõ họ tên)"),
            (5, 8, "GIÁM ĐỐC", "(Ký, đóng dấu, ghi rõ họ tên)"),
        ):
            sheet.merge_cells(
                start_row=title_row,
                start_column=start_column,
                end_row=title_row,
                end_column=end_column,
            )
            sheet.merge_cells(
                start_row=note_row,
                start_column=start_column,
                end_row=note_row,
                end_column=end_column,
            )
            sheet.cell(title_row, start_column).value = title
            sheet.cell(note_row, start_column).value = note
        sheet.merge_cells(
            start_row=preparer_row,
            start_column=1,
            end_row=preparer_row,
            end_column=2,
        )
        sheet.cell(preparer_row, 1).value = (
            f"({profile.report_preparer} - SĐT: {profile.phone})"
        )

        sheet.cell(notes_row, 1).value = "Ghi chú:"
        notes = (
            "1. Các  khoản phải thu cần sao kê bao gồm: 321, 322, 323, 35, 36, 387",
            "2. Các tài khoản phải trả cần sao kê bao gồm: 45, 46, 48 (trừ các TK 483,484, 485, 486, 488).",
            "3. Tổng số dư của mỗi tài khoản (Số Cộng TK) phải được đối chiếu khớp đúng với bảng cân đối tài khoản.",
            "4. Yêu cầu mỗi khoản phải thu, phải trả sao kê phải điền đầy đủ, rõ ràng thông tin tại cột 3, 4 và 5.",
            "5. Riêng cột 5 yêu cầu ghi rõ theo thứ tự ngày tháng năm phát sinh, không báo cáo năm phát sinh hoặc để trống. Đây là cơ sở quan trọng để Kiểm toán độc lập",
            "làm căn cứ tính trích lập dự phòng rủi ro khó đòi.",
            "6. Cột 10: Ghi nguyên nhân chưa tất toán đối với các khoản phải thu, phải trả phát sinh từ 24 tháng trở lên.",
            "7. Lưu ý:  Đối với các khoản phải thu, phải trả hạch toán tự động và có tính chất đặc thù như:  Thuế, hỗ trợ lãi suất không thực hiện sao kê chi tiết, nhập",
            "tổng hợp theo từng nội dung: tại cột 5 ghi tổng số món, cột 7, 8  ghi tổng số tiền.",
        )
        for offset, text in enumerate(notes):
            sheet.cell(notes_row + offset, 2).value = text

    def _write_control_sheet(
        self,
        workbook: Workbook,
        records: list[ReceivablePayableRecord],
    ) -> None:
        sheet = workbook.create_sheet("SoLieuTongHop")
        sheet["A1"] = "SỐ LIỆU TỔNG HỢP MẪU 24"
        headers = ("Tài Khoản", "Số Dư Nguyên Tệ", "Số Dư Qui Đổi VNĐ")
        for column, header in enumerate(headers, start=1):
            sheet.cell(2, column).value = header
        totals: OrderedDict[str, list[Decimal]] = OrderedDict()
        for record in records:
            bucket = totals.setdefault(record.control_key, [Decimal(0), Decimal(0)])
            bucket[0] += record.original_balance
            bucket[1] += record.converted_balance
        row = 3
        for key, amounts in totals.items():
            sheet.cell(row, 1).value = key
            sheet.cell(row, 2).value = self._number_to_cell(amounts[0])
            sheet.cell(row, 3).value = self._number_to_cell(amounts[1])
            row += 1
        total_row = row
        sheet.cell(total_row, 1).value = "Tổng cộng:"
        sheet.cell(total_row, 2).value = f"=SUM(B3:B{total_row - 1})"
        sheet.cell(total_row, 3).value = f"=SUM(C3:C{total_row - 1})"
        self._format_control_sheet(sheet, total_row)

    @staticmethod
    def _format_control_sheet(sheet, total_row: int) -> None:
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in sheet.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = border
                cell.font = Font(name="Times New Roman", size=12)
                cell.alignment = Alignment(vertical="center")
        sheet["A1"].font = Font(
            name="Times New Roman",
            size=12,
            bold=True,
            color="000080",
        )
        for cell in sheet[2][:3]:
            cell.font = Font(name="Times New Roman", size=12, bold=True)
            cell.fill = PatternFill("solid", fgColor="FFFF00")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in sheet[total_row][:3]:
            cell.font = Font(name="Times New Roman", size=12, bold=True)
        for row in sheet.iter_rows(
            min_row=3,
            max_row=total_row,
            min_col=2,
            max_col=3,
        ):
            for cell in row:
                cell.number_format = "#,##0"
        sheet.column_dimensions["A"].width = 16
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 20
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:C{total_row}",
            orientation="portrait",
        )

    def _format_report(
        self,
        sheet,
        options: SettlementOptions,
        final_row: int,
        summary_rows: set[int],
    ) -> None:
        notes_end = final_row + 19
        final_column = 8 if options.remove_unused_columns else sheet.max_column
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=notes_end,
            min_col=1,
            max_col=final_column,
        ):
            for cell in row:
                cell.font = Font(name="Times New Roman", size=11)
                cell.alignment = Alignment(vertical="center")

        style_agency_header(sheet, start_column=1, end_column=2)
        sheet["A6"].font = Font(name="Times New Roman", size=18, bold=True)
        sheet["A7"].font = Font(name="Times New Roman", size=14, bold=True, italic=True)
        for address in ("A6", "A7"):
            sheet[address].alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[5].height = 4

        for row in range(9, 11):
            for column in range(1, 9):
                cell = sheet.cell(row, column)
                cell.font = Font(name="Times New Roman", size=11, bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = body_border
        for row in range(11, final_row + 1):
            is_summary = row in summary_rows
            for column in range(1, 9):
                cell = sheet.cell(row, column)
                cell.border = body_border
                if is_summary:
                    cell.font = Font(name="Times New Roman", size=11, bold=True)
            for column in (5, 6):
                sheet.cell(row, column).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
            for column in (7, 8):
                sheet.cell(row, column).number_format = "#,##0"
            sheet.cell(row, 1).number_format = "@"
            sheet.cell(row, 5).number_format = (
                "dd/mm/yyyy" if options.four_digit_year else "dd/mm/yy"
            )
            if not is_summary:
                for column in (2, 3):
                    sheet.cell(row, column).alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )
                text_length = max(
                    len(str(sheet.cell(row, column).value or ""))
                    for column in (2, 3)
                )
                if text_length > 28:
                    sheet.row_dimensions[row].height = min(
                        60,
                        15 * ((text_length // 28) + 1),
                    )

        for row in summary_rows:
            for column in range(1, 9):
                cell = sheet.cell(row, column)
                cell.border = Border(
                    left=medium if column == 1 else thin,
                    right=medium if column == 8 else thin,
                    top=medium,
                    bottom=medium,
                )

        date_row = final_row + 1
        title_row = final_row + 2
        note_row = final_row + 3
        preparer_row = final_row + 9
        notes_row = final_row + 11
        for row in range(date_row, preparer_row + 1):
            for column in range(1, 9):
                sheet.cell(row, column).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )
        sheet.cell(date_row, 5).font = Font(
            name="Times New Roman",
            size=11,
            italic=True,
        )
        for column in (1, 3, 5):
            sheet.cell(title_row, column).font = Font(
                name="Times New Roman",
                size=11,
                bold=True,
            )
            sheet.cell(note_row, column).font = Font(
                name="Times New Roman",
                size=11,
                italic=True,
            )
        sheet.cell(preparer_row, 1).font = Font(
            name="Times New Roman",
            size=11,
            italic=True,
        )
        sheet.cell(notes_row, 1).font = Font(
            name="Times New Roman",
            size=11,
            bold=True,
        )
        sheet.cell(notes_row, 1).alignment = Alignment(horizontal="right")
        for row in range(notes_row, notes_end + 1):
            sheet.cell(row, 2).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False,
            )

        widths = {
            "A": 9.5,
            "B": 34,
            "C": 36,
            "D": 22,
            "E": 10,
            "F": 6,
            "G": 17,
            "H": 17,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        setup_a4_print_layout(
            sheet,
            print_area=f"A1:{get_column_letter(final_column)}{notes_end}",
            orientation="landscape",
            title_rows="$9:$9",
        )
        sheet.print_title_cols = f"$A:${get_column_letter(min(final_column, 8))}"
        sheet.page_setup.scale = 98
        sheet.page_setup.fitToWidth = 0
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = False
        sheet.sheet_view.showGridLines = False
        if options.remove_unused_columns and sheet.max_column > 8:
            sheet.delete_cols(9, sheet.max_column - 8)

    def _row_to_record(
        self,
        row: dict[str, str],
        extra_headers: tuple[str, ...],
    ) -> ReceivablePayableRecord:
        return ReceivablePayableRecord(
            transaction_date_text=str(row["NGAY_GIAO_DICH"] or "").strip(),
            branch_code=str(row["MA_CN"] or "").strip(),
            account=self._clean_code(row["TAI_KHOAN"]),
            account_name=str(row["TEN_TAI_KHOAN"] or ""),
            counterparty_code=self._clean_code(row["MA_DOI_TUONG_PHAI_THU_PHAI_TRA"]),
            counterparty_name=str(row["TEN_DOI_TUONG"] or ""),
            reason=str(row["LY_DO_HACH_TOAN"] or "").strip(),
            arising_year=parse_yyyymmdd(row["NAM_PHAT_SINH"]),
            currency=str(row["LOAI_TIEN"] or "").strip(),
            original_balance=self._parse_source_amount(
                row["SO_DU_NGUYENTE_TAI_NGAY_3112"]
            ),
            converted_balance=self._parse_source_amount(
                row["SO_DU_QUYDOI_VND_TAI_NGAY_3112"]
            ),
            extra_values=tuple(row.get(header, "") for header in extra_headers),
        )

    @staticmethod
    def _output_prefix(options: SettlementOptions) -> str:
        prefix = (options.output_prefix or "QT").strip().upper()
        return "BN" if prefix == "BN" else "QT"

    @staticmethod
    def _clean_code(value: Any) -> str:
        return str(value or "").strip().lstrip("'").strip()

    @staticmethod
    def _numeric_code(value: str) -> int | str | None:
        text = str(value or "").strip()
        if not text:
            return None
        return int(text) if text.isdigit() else text

    @staticmethod
    def _parse_source_amount(value: Any) -> Decimal:
        text = str(value or "").replace("\u00a0", "").replace(" ", "").strip()
        if not text or text == "-":
            return Decimal(0)
        if "," in text and "." not in text:
            text = text.replace(",", ".")
        elif "," in text and "." in text:
            text = text.replace(",", "")
        try:
            return Decimal(text)
        except InvalidOperation:
            return Decimal(0)

    @staticmethod
    def _number_to_cell(value: Decimal | int | float) -> int | float:
        if not isinstance(value, Decimal):
            value = Decimal(str(value))
        if value == value.to_integral_value():
            return int(value)
        return float(value)
