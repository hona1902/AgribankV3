from __future__ import annotations

import csv
from collections import OrderedDict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementRequest
from agribank_v3.settlement.processors.mau05 import Mau05Processor


class Summary05Processor:
    """Create the final Tổng hợp Mẫu 05 workbook without Excel/VBA."""

    DETAIL_SHEET = "SoLieu_Mau05"
    SUMMARY_SHEET = "TongHop_Mau05"

    def execute(
        self,
        request: SettlementRequest,
        merged_csv_path: Path,
        output_path: Path,
    ) -> int:
        mau05 = Mau05Processor()
        records, report_date = mau05.read_source(request, merged_csv_path)
        workbook = mau05.build_workbook(request, records, report_date)
        detail = workbook["05"]
        detail.title = self.DETAIL_SHEET
        if "SoLieuTongHop" in workbook.sheetnames:
            del workbook["SoLieuTongHop"]
        processed_rows = self._build_summary(workbook, merged_csv_path)
        workbook.save(output_path)
        workbook.close()
        return processed_rows

    def _build_summary(self, workbook, merged_csv_path: Path) -> int:
        headers, rows = self._read_merged_csv(merged_csv_path)
        required = {"MA_CN", "LOAI_TIEN", "SO_TIEN_QUY_DOI_VND", "TAI_KHOAN"}
        missing = sorted(required - set(headers))
        if missing:
            raise SettlementError(f"File nối mẫu 05 thiếu cột: {', '.join(missing)}")

        branches: OrderedDict[str, None] = OrderedDict()
        account_keys: OrderedDict[str, None] = OrderedDict()
        values: dict[tuple[str, str], Decimal] = {}
        for row in rows:
            branch = self._text(row.get("MA_CN"))
            if not branch:
                continue
            account = self._text(row.get("TAI_KHOAN")) or "994009"
            currency = self._text(row.get("LOAI_TIEN"))
            key = f"{account}_{currency}" if currency else account
            amount = self._decimal(row.get("SO_TIEN_QUY_DOI_VND"))
            branches.setdefault(branch, None)
            account_keys.setdefault(key, None)
            values[(branch, key)] = values.get((branch, key), Decimal(0)) + amount

        if self.SUMMARY_SHEET in workbook.sheetnames:
            del workbook[self.SUMMARY_SHEET]
        sheet = workbook.create_sheet(self.SUMMARY_SHEET)
        sheet["A1"] = "Báo cáo tổng hợp số liệu quyết toán mẫu 05"
        sheet["A2"] = " Mã Chi nhánh"
        for offset, account_key in enumerate(account_keys, start=2):
            sheet.cell(2, offset, account_key)
        total_column = 2 + len(account_keys)
        sheet.cell(2, total_column, "Cộng chi nhánh")

        for row_offset, branch in enumerate(branches, start=3):
            sheet.cell(row_offset, 1, self._branch_to_cell(branch))
            row_total = Decimal(0)
            for column_offset, account_key in enumerate(account_keys, start=2):
                amount = values.get((branch, account_key), Decimal(0))
                row_total += amount
                sheet.cell(row_offset, column_offset, self._number_to_cell(amount))
            sheet.cell(row_offset, total_column, self._number_to_cell(row_total))

        total_row = 3 + len(branches)
        sheet.cell(total_row, 1, "Tổng cộng")
        for column in range(2, total_column + 1):
            total = sum(
                self._decimal(sheet.cell(row, column).value)
                for row in range(3, total_row)
            )
            sheet.cell(total_row, column, self._number_to_cell(total))

        self._format(sheet, total_row, total_column)
        return len(rows)

    def _format(self, sheet, total_row: int, total_column: int) -> None:
        thin = Side(style="thin", color="000000")
        medium = Side(style="medium", color="000000")
        for row in sheet.iter_rows(min_row=1, max_row=total_row, min_col=1, max_col=total_column):
            for cell in row:
                cell.font = Font(name="Times New Roman", size=11)
                cell.alignment = Alignment(vertical="center")
        sheet["A1"].font = Font(name="Times New Roman", size=12, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        for column in range(1, total_column + 1):
            cell = sheet.cell(2, column)
            cell.font = Font(name="Times New Roman", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=medium, right=medium, top=medium, bottom=medium)
        for row in range(3, total_row + 1):
            sheet.cell(row, 1).font = Font(name="Times New Roman", size=11, bold=True)
            sheet.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
            for column in range(1, total_column + 1):
                sheet.cell(row, column).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for column in range(2, total_column + 1):
                sheet.cell(row, column).number_format = "#,##0"
        for column in range(1, total_column + 1):
            sheet.cell(total_row, column).font = Font(name="Times New Roman", size=11, bold=True)
        sheet.column_dimensions["A"].width = 15
        for column in range(2, total_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 18
        sheet.freeze_panes = "B3"

    @staticmethod
    def _read_merged_csv(path: Path) -> tuple[set[str], list[dict[str, str]]]:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            rows = list(reader)
            return set(reader.fieldnames or ()), rows

    @staticmethod
    def _text(value: Any) -> str:
        return str(value or "").strip().lstrip("'").strip()

    @staticmethod
    def _decimal(value: Any) -> Decimal:
        if value in (None, ""):
            return Decimal(0)
        try:
            return Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, ValueError):
            return Decimal(0)

    @staticmethod
    def _number_to_cell(value: Decimal) -> int | float:
        return int(value) if value == value.to_integral_value() else float(value)

    @staticmethod
    def _branch_to_cell(value: str) -> int | str:
        return int(value) if value.isdigit() else value
