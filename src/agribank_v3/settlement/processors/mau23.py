from __future__ import annotations

from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import StringIO
from pathlib import Path
from typing import Any
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.writer.theme import theme_xml

from agribank_v3.settlement.engine import SettlementError
from agribank_v3.settlement.models import SettlementRequest, SettlementResult
from agribank_v3.settlement.processors.formatting import setup_a4_print_layout, style_agency_header


@dataclass(frozen=True, slots=True)
class UnusualRecord:
    account: str
    description: str
    posting_date: datetime | int
    amount: Decimal
    sort_date: datetime | int


class Mau23Processor:
    """Processor for Mẫu 23/QT from glcb06 account detail files."""

    TARGET_ACCOUNTS = ("790008", "790009", "899001", "899009")

    def execute(self, request: SettlementRequest) -> SettlementResult:
        if not request.source_paths:
            raise SettlementError("Mẫu 23/QT cần ít nhất một file nguồn.")
        records, skipped = self.read_sources(request.source_paths)
        if not records:
            raise SettlementError("Tất cả các file đã chọn đều chỉ có tiêu đề hoặc không có dữ liệu.")
        workbook = self.build_workbook(request, records)
        output_path = request.source_paths[0].with_name(
            f"{request.profile.branch_code.strip()}{request.options.output_prefix}23.xlsx"
        )
        workbook.save(output_path)
        workbook.close()
        warnings = ()
        if skipped:
            warnings = ("Các file không có dữ liệu đã được bỏ qua: " + ", ".join(skipped),)
        return SettlementResult(
            spec_key=request.spec.key,
            output_path=output_path,
            workbook_name=output_path.name,
            worksheet_name="23",
            processed_rows=len(records),
            warnings=warnings,
        )

    def read_sources(self, source_paths: tuple[Path, ...]) -> tuple[list[UnusualRecord], list[str]]:
        records: list[UnusualRecord] = []
        skipped: list[str] = []
        for source_path in source_paths:
            source_records = self._read_source(source_path)
            if source_records:
                records.extend(source_records)
            else:
                skipped.append(source_path.name)
        return self._summarize_automatic_records(records), skipped

    def build_workbook(self, request: SettlementRequest, records: list[UnusualRecord]) -> Workbook:
        workbook = Workbook()
        workbook.loaded_theme = self._times_new_roman_theme()
        workbook._named_styles["Normal"].font = Font(name="Times New Roman", size=10)
        sheet = workbook.active
        sheet.title = "23"
        self._write_header(sheet, request)
        total_row = self._write_table(sheet, records)
        self._write_signatures_and_notes(sheet, request, total_row + 3)
        self._format_report(sheet, total_row)
        return workbook

    def _read_source(self, source_path: Path) -> list[UnusualRecord]:
        if source_path.suffix.casefold() not in {".xls", ".xlsx", ".xlsm"}:
            raise SettlementError("Nguồn mẫu 23 phải là file Excel xuất từ glcb06.")
        try:
            import xlrd
            workbook = xlrd.open_workbook(str(source_path), logfile=StringIO())
            sheet = workbook.sheet_by_index(0)
        except Exception as exc:
            raise SettlementError(f"Không thể đọc file {source_path.name}: {exc}") from exc
        if sheet.nrows <= 1:
            return []
        headers = [str(sheet.cell_value(0, col)).strip().casefold() for col in range(sheet.ncols)]
        required = {"tran_time", "user", "journal", "acctnm", "remark", "dramt", "cramt", "trtp"}
        missing = sorted(required - set(headers))
        if missing:
            raise SettlementError(f"File {source_path.name} thiếu cột: {', '.join(missing)}")
        index = {header: headers.index(header) for header in headers}
        records: list[UnusualRecord] = []
        for row_index in range(1, sheet.nrows):
            row = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
            if self._text(row[index["trtp"]]).casefold() == "cancel":
                continue
            account = self._text(row[index["acctnm"]])[:6]
            if account not in self.TARGET_ACCOUNTS:
                continue
            amount = self._decimal(row[index["cramt"] if account.startswith("7") else index["dramt"]])
            if amount == 0:
                continue
            posting_date = self._excel_date(row[index["tran_time"]], workbook.datemode) or row[index["tran_time"]]
            remark = self._text(row[index["remark"]])
            description = self._description(
                remark=remark,
                user=self._text(row[index["user"]]),
                journal=self._text(row[index["journal"]]),
                amount=amount,
            )
            records.append(
                UnusualRecord(
                    account=account,
                    description=description,
                    posting_date=posting_date,
                    amount=amount,
                    sort_date=posting_date,
                )
            )
        return records

    def _write_header(self, sheet, request: SettlementRequest) -> None:
        profile = request.profile
        branch_name = profile.reporting_branch_name.strip() or profile.branch_name.strip()
        values = {
            "A1": "NGÂN HÀNG NÔNG NGHIỆP",
            "A2": "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM",
            "A3": f"Mã chi nhánh: {profile.branch_code.strip()}",
            "A4": f"Tên chi nhánh: {branch_name.removeprefix('Chi nhánh ').strip()}",
            "A6": "SAO KÊ CHI TIẾT TÀI KHOẢN THU NHẬP VÀ CHI PHÍ BẤT THƯỜNG",
            "A7": self._report_date_text(request),
            "D8": "Đơn vị : VNĐ",
            "D1": "1. Mẫu số 23/QT",
            "D2": "2. CN loại I gửi file về TSC",
            "D3": "3. Lưu tại chi nhánh",
            "A9": "TÀI KHOẢN",
            "B9": "DIỄN GIẢI NGHIỆP VỤ",
            "C9": "NGÀY HẠCH TOÁN",
            "D9": "SỐ TIỀN HẠCH TOÁN (quy đổi VNĐ)",
        }
        for cell, value in values.items():
            sheet[cell] = value
        self._apply_times_new_roman(sheet, min_row=1, min_col=1, max_row=11, max_col=4)
        for ref in ("A1:B1", "A2:B2", "A3:B3", "A4:B4", "A6:D6", "A7:D7", "A9:A10", "B9:B10", "C9:C10", "D9:D10"):
            sheet.merge_cells(ref)
        style_agency_header(sheet, start_column=1, end_column=3)
        sheet["A6"].font = Font(name="Times New Roman", size=14, bold=True)
        sheet["A7"].font = Font(name="Times New Roman", size=12, bold=True, italic=True)
        sheet["D8"].font = Font(name="Times New Roman", size=10, italic=True)
        sheet["D8"].alignment = Alignment(horizontal="right", vertical="center")
        for row in range(1, 4):
            sheet.cell(row, 4).font = Font(name="Times New Roman", size=8)
        for address in ("A6", "A7"):
            sheet[address].alignment = Alignment(horizontal="center", vertical="center")
        for column in range(1, 5):
            sheet.cell(11, column, column)

    def _write_table(self, sheet, records: list[UnusualRecord]) -> int:
        grouped: dict[str, list[UnusualRecord]] = {account: [] for account in self.TARGET_ACCOUNTS}
        for record in sorted(records, key=lambda item: (item.account, self._sort_key(item.sort_date), item.description)):
            grouped.setdefault(record.account, []).append(record)
        row = 12
        for account in self.TARGET_ACCOUNTS:
            start_row = row
            sheet.cell(row, 1, int(account))
            row += 1
            for record in grouped.get(account, []):
                sheet.cell(row, 1, None)
                sheet.cell(row, 2, record.description)
                sheet.cell(row, 3, record.posting_date)
                sheet.cell(row, 4, self._number(record.amount))
                row += 1
            total_row = row
            sheet.cell(total_row, 1, f"Cộng TK {account}")
            sheet.cell(total_row, 4, f"=SUM(D{start_row}:D{total_row - 1})")
            row += 1
        return row - 1

    def _write_signatures_and_notes(self, sheet, request: SettlementRequest, start_row: int) -> None:
        report_date = self._report_date_text(request)
        location = request.profile.report_location.strip()
        sheet.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=4)
        sheet.cell(start_row, 3, f"{location}, {report_date}" if location else report_date)
        sheet.cell(start_row, 3).font = Font(name="Times New Roman", size=10, italic=True)
        sheet.cell(start_row, 3).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(start_row + 1, 1, "LẬP BIỂU")
        sheet.cell(start_row + 1, 2, "TRƯỞNG PHÒNG KẾ TOÁN")
        sheet.merge_cells(start_row=start_row + 1, start_column=3, end_row=start_row + 1, end_column=4)
        sheet.cell(start_row + 1, 3, "GIÁM ĐỐC")
        sheet.cell(start_row + 2, 1, "(Ký, ghi rõ họ tên, số ĐT liên hệ)")
        sheet.cell(start_row + 2, 2, "(Ký, ghi rõ họ tên)")
        sheet.merge_cells(start_row=start_row + 2, start_column=3, end_row=start_row + 2, end_column=4)
        sheet.cell(start_row + 2, 3, "(Ký, đóng dấu, ghi rõ họ tên)")
        for row in (start_row + 1, start_row + 2):
            for col in range(1, 5):
                sheet.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        for col in (1, 2, 3):
            sheet.cell(start_row + 1, col).font = Font(name="Times New Roman", size=10, bold=True)
            sheet.cell(start_row + 2, col).font = Font(name="Times New Roman", size=10, italic=True)
        preparer = request.profile.report_preparer.strip()
        phone = request.profile.phone.strip()
        if preparer:
            sheet.cell(start_row + 7, 1, preparer)
        if phone:
            sheet.cell(start_row + 8, 1, f"SĐT {phone}")
        notes = [
            "Ghi chú: ",
            '1. Tổng số tiền hạch toán vào tài khoản (Số "Cộng TK") phải được đối chiếu khớp với số dư trên bảng cân đối của chi nhánh.',
            "2. Yêu cầu mỗi khoản sao kê phải có diễn giải nghiệp vụ.",
            "3. Lưu ý: Đối với các khoản thu nhập, chi phí bất thường hạch toán tự động và có tính trùng lắp như: Thay đổi trạng thái, hủy tài khoản không hoạt động,",
            "tài khoản ngủ… không thực hiện sao kê chi tiết, nhập tổng hợp theo từng nội dung: tại cột 3 ghi tổng số món, cột 4 ghi tổng số tiền.",
        ]
        for offset, text in enumerate(notes, start=9):
            sheet.cell(start_row + offset, 1, text)
        sheet.cell(start_row + 9, 1).font = Font(name="Times New Roman", size=10, bold=True)
        for row in range(start_row + 10, start_row + 14):
            sheet.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _format_report(self, sheet, total_row: int) -> None:
        self._apply_times_new_roman(sheet)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(9, total_row + 1):
            for column in range(1, 5):
                cell = sheet.cell(row, column)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=row == 9)
                if row >= 12 and column == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if row >= 12 and column == 4:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
        for row in range(12, total_row + 1):
            value = sheet.cell(row, 1).value
            if value is not None and sheet.cell(row, 2).value is None:
                for column in range(1, 5):
                    sheet.cell(row, column).font = Font(name="Times New Roman", size=10, bold=True)
        for row in range(12, total_row + 1):
            if isinstance(sheet.cell(row, 3).value, datetime):
                sheet.cell(row, 3).number_format = "dd/mm/yyyy"
            else:
                sheet.cell(row, 3).number_format = "0"
        sheet.column_dimensions["A"].width = 27
        sheet.column_dimensions["B"].width = 50
        sheet.column_dimensions["C"].width = 25
        sheet.column_dimensions["D"].width = 25
        setup_a4_print_layout(sheet, print_area=f"A1:D{sheet.max_row}", title_rows="$9:$11", orientation="portrait")

    def _summarize_automatic_records(self, records: list[UnusualRecord]) -> list[UnusualRecord]:
        records = self._remove_offsetting_pairs(records)
        result: list[UnusualRecord] = []
        coin_count = 0
        coin_amount = Decimal(0)
        dormant_count = 0
        dormant_amount = Decimal(0)
        for record in records:
            desc_key = record.description.casefold()
            if self._is_coin_rounding(desc_key):
                coin_count += 1
                coin_amount += record.amount
            elif "dormant account" in desc_key:
                dormant_count += 1
                dormant_amount += record.amount
            else:
                result.append(record)
        if coin_count:
            result.append(UnusualRecord("790009", "Thu tiền lẻ cuối ngày", coin_count, coin_amount, coin_count))
        if dormant_count:
            result.append(UnusualRecord("790009", "TK không hoạt động chuyển sang tài khoản ngủ", dormant_count, dormant_amount, dormant_count))
        return result

    def _remove_offsetting_pairs(self, records: list[UnusualRecord]) -> list[UnusualRecord]:
        removed: set[int] = set()
        for i, first in enumerate(records):
            if i in removed:
                continue
            for j in range(i + 1, len(records)):
                second = records[j]
                if j in removed or first.account != second.account:
                    continue
                if first.amount + second.amount != 0:
                    continue
                if self._similar_business_text(first.description, second.description):
                    removed.add(i)
                    removed.add(j)
                    break
        return [record for index, record in enumerate(records) if index not in removed]

    def _similar_business_text(self, left: str, right: str) -> bool:
        left_key = self._compact_text(left)
        right_key = self._compact_text(right)
        return any(token in left_key and token in right_key for token in ("atm05", "bt", "thuadequy", "thuaquy"))

    def _description(self, *, remark: str, user: str, journal: str, amount: Decimal) -> str:
        text = remark.strip()
        lower = text.casefold()
        if "quyết toán thuế gtgt" in lower:
            return text.replace("năm ", "")
        if amount == Decimal("36623"):
            return "Hoàn phí BH tai nạn chủ thẻ Quốc Tế đối với thẻ đóng năm 2024"
        if "thừa quỹ" in lower:
            text = text.replace("Xử lý thừa quỹ 5491A005", "Thu nghiệp vụ các khoản thừa quỹ ATM05")
            text = text.replace("Xử lý thừa quỹ 5491ATM05", "Thu nghiệp vụ các khoản thừa quỹ ATM05")
            text = text.replace("HT Xử lý thừa quỹ ATM05", "Thu nghiệp vụ các khoản thừa quỹ  ATM05")
            text = text.replace("chu kỳ ngày ", "chu kỳ ")
            text = text.rstrip(".")
            return text
        if not text:
            return f"{user}_BT{journal}"
        return text

    def _is_coin_rounding(self, text: str) -> bool:
        compact = self._compact_text(text)
        return (
            "thutienlecuoingay" in compact
            or "thutienlecuingay" in compact
            or "thutienlecuoi" in compact
            or "thutinlecuoingay" in compact
            or "thutienle" in compact and "ngay" in compact
            or "thutinle" in compact and "ngay" in compact
        )

    def _compact_text(self, text: str) -> str:
        normalized = "".join(
            ch
            for ch in unicodedata.normalize("NFD", text)
            if unicodedata.category(ch) != "Mn"
        )
        normalized = normalized.replace("đ", "d").replace("Đ", "D")
        return "".join(ch for ch in normalized.casefold() if ch.isalnum())

    def _sort_key(self, value: datetime | int) -> tuple[int, Any]:
        if isinstance(value, datetime):
            return (0, value)
        return (1, value)

    def _report_date_text(self, request: SettlementRequest) -> str:
        today = date.today()
        if request.options.output_prefix == "BN":
            return f"Ngày 30 tháng 6 năm {today.year}"
        return f"Ngày 31 tháng 12 năm {today.year}"

    def _excel_date(self, value: Any, datemode: int) -> datetime | None:
        try:
            serial = float(value)
        except (TypeError, ValueError):
            return None
        try:
            import xlrd
            return xlrd.xldate_as_datetime(serial, datemode)
        except Exception:
            return None

    def _decimal(self, value: Any) -> Decimal:
        if value in (None, ""):
            return Decimal(0)
        try:
            return Decimal(str(value).replace(",", ""))
        except (InvalidOperation, ValueError):
            return Decimal(0)

    def _number(self, value: Decimal) -> int | float:
        return int(value) if value == value.to_integral_value() else float(value)

    def _text(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return text[:-2] if text.endswith(".0") else text

    def _apply_times_new_roman(
        self,
        sheet,
        *,
        min_row: int | None = None,
        min_col: int | None = None,
        max_row: int | None = None,
        max_col: int | None = None,
    ) -> None:
        for row in sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in row:
                font = copy(cell.font)
                font.name = "Times New Roman"
                cell.font = font

    def _times_new_roman_theme(self) -> str | bytes:
        if isinstance(theme_xml, bytes):
            return theme_xml.replace(b"Calibri", b"Times New Roman")
        return theme_xml.replace("Calibri", "Times New Roman")
