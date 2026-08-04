from __future__ import annotations

from collections.abc import Mapping, Sequence
import hashlib
import json
import os
import re
import shutil
import uuid
import zipfile
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from agribank_v3.runtime_paths import application_root

from ..models import CreditLimitRow, DashboardData, DashboardMetric, PageResult, SummaryError
from .models import (
    CREDIT_LIMIT_BATCH_SCHEMA_VERSION,
    DATA_SHEET_NAME,
    META_SHEET_NAME,
    STORAGE_FOLDER_NAME,
    CreditLimitBatchMetadata,
    CreditLimitStorageStatus,
)


DATA_HEADERS: tuple[str, ...] = (
    "batch_id",
    "branch_code",
    "customer_code",
    "customer_name",
    "account_number",
    "credit_line_type",
    "approval_sequence",
    "approval_date",
    "approved_limit",
    "maturity_date",
    "outstanding_balance",
    "officer_code",
    "officer_name",
    "address",
    "source_row_count",
)
REQUIRED_DATA_HEADERS_FOR_READ: tuple[str, ...] = tuple(
    header for header in DATA_HEADERS if header != "outstanding_balance"
)
TEXT_COLUMNS = {
    "batch_id",
    "branch_code",
    "customer_code",
    "customer_name",
    "account_number",
    "credit_line_type",
    "approval_sequence",
    "officer_code",
    "officer_name",
    "address",
}
MONEY_COLUMNS = {"approved_limit", "outstanding_balance"}
DATE_COLUMNS = {"approval_date", "maturity_date"}


@dataclass(frozen=True, slots=True)
class CreditLimitRowContext:
    metadata: CreditLimitBatchMetadata
    row: CreditLimitRow
    reference_date: date
    warn_days: int
    min_limit: float
    dynamic_status: bool
    outstanding_balance_available: bool = True


def credit_limit_storage_directory(main_database_path: str | Path | None = None) -> Path:
    if main_database_path:
        return Path(main_database_path).expanduser().resolve().parent / STORAGE_FOLDER_NAME
    return application_root() / "data" / STORAGE_FOLDER_NAME


class CreditLimitExcelBatchStore:
    """Stores each imported LN01 credit-limit batch as one atomic Excel workbook."""

    def __init__(self, main_database_path: str | Path | None = None) -> None:
        self.storage_path = credit_limit_storage_directory(main_database_path)
        self.temp_path = self.storage_path / "Temp"
        self.backup_path = self.storage_path / "Backup"
        self.trash_path = self.storage_path / "Trash"
        self._metadata_cache: dict[Path, tuple[int, int, CreditLimitBatchMetadata]] = {}
        self._row_cache: dict[Path, tuple[int, int, tuple[CreditLimitRow, ...], bool]] = {}

    def ensure_storage(self) -> Path:
        self.storage_path.mkdir(parents=True, exist_ok=True)
        self.temp_path.mkdir(parents=True, exist_ok=True)
        self.backup_path.mkdir(parents=True, exist_ok=True)
        self.trash_path.mkdir(parents=True, exist_ok=True)
        return self.storage_path

    def create_batch(
        self,
        *,
        source_path: Path,
        rows: Sequence[CreditLimitRow] | Sequence[Mapping[str, Any]],
        accepted_rows: Sequence[CreditLimitRow] | Sequence[Mapping[str, Any]],
        reference_date: date,
        min_limit: float,
        warn_days: int,
        source_file_sha256: str,
        source_file_size: int,
        imported_by: str = "",
        source_row_count: int = 0,
        duplicate_policy: str = "error",
        progress: Callable[[str], None] | None = None,
    ) -> CreditLimitBatchMetadata:
        self.ensure_storage()
        duplicate = self.find_duplicate_by_sha(source_file_sha256)
        if duplicate and duplicate_policy == "error":
            first = duplicate[0]
            raise SummaryError(
                f"File LN01 này đã được nhập trước đó: {first.batch_name} ({first.file_name})."
            )
        if progress:
            progress("Đang tạo workbook batch HMHETHAN...")
        now = datetime.now()
        source_path = Path(source_path)
        materialized_rows = tuple(normalize_credit_limit_row(row) for row in rows)
        materialized_accepted = tuple(normalize_credit_limit_row(row) for row in accepted_rows)
        expired = sum(1 for row in materialized_accepted if row.status == "Đã hết hạn")
        expiring = sum(1 for row in materialized_accepted if row.status == "Sắp hết hạn")
        batch_id = self._new_batch_id(now, source_file_sha256)
        batch_name = f"{source_path.stem} - {now:%d/%m/%Y %H:%M:%S}"
        file_name = self._new_file_name(now, source_path.stem, source_file_sha256)
        target = self.storage_path / file_name
        temp_file = self.temp_path / f".{file_name}.{uuid.uuid4().hex}.tmp.xlsx"
        metadata = CreditLimitBatchMetadata(
            batch_id=batch_id,
            batch_name=batch_name,
            file_path=target,
            file_name=file_name,
            source_file_name=source_path.name,
            source_file_sha256=source_file_sha256,
            source_file_size=source_file_size,
            imported_at=now,
            imported_by=str(imported_by or ""),
            app_version="AgribankV3",
            source_row_count=int(source_row_count or len(materialized_rows)),
            accepted_row_count=len(materialized_accepted),
            rejected_row_count=max(0, int(source_row_count or len(materialized_rows)) - len(materialized_accepted)),
            warning_count=len(materialized_accepted),
            reference_date_at_import=reference_date,
            minimum_limit_at_import=float(min_limit or 0),
            warning_days_at_import=int(warn_days),
            expired_count_at_import=expired,
            expiring_count_at_import=expiring,
            status="OK",
            notes="",
        )
        workbook: Workbook | None = None
        try:
            workbook = Workbook()
            meta_sheet = workbook.active
            meta_sheet.title = META_SHEET_NAME
            data_sheet = workbook.create_sheet(DATA_SHEET_NAME)
            self._write_metadata_sheet(meta_sheet, metadata)
            self._write_data_sheet(data_sheet, metadata.batch_id, materialized_rows)
            workbook.save(temp_file)
            self._validate_workbook(temp_file)
            os.replace(temp_file, target)
        except Exception:
            if temp_file.exists():
                temp_file.unlink(missing_ok=True)
            raise
        finally:
            if workbook is not None:
                workbook.close()
        self.invalidate_cache()
        return self._metadata_from_file(target)

    def list_batches(self) -> list[CreditLimitBatchMetadata]:
        self.ensure_storage()
        batches: list[CreditLimitBatchMetadata] = []
        for path in self._iter_candidate_files():
            try:
                batches.append(self._metadata_from_file(path))
            except Exception:
                continue
        batches.sort(
            key=lambda item: (
                item.imported_at or datetime.min,
                item.modified_at or datetime.min,
                item.file_name,
            ),
            reverse=True,
        )
        return batches

    def find_duplicate_by_sha(self, source_file_sha256: str) -> list[CreditLimitBatchMetadata]:
        needle = str(source_file_sha256 or "").strip().lower()
        if not needle:
            return []
        return [
            item
            for item in self.list_batches()
            if str(item.source_file_sha256 or "").strip().lower() == needle
        ]

    def query_credit_limits(
        self,
        *,
        batch_id: str | int | None = None,
        search: str = "",
        status: str = "",
        officer: str = "",
        page: int = 1,
        page_size: int = 200,
        min_limit: float | None = None,
        warn_days: int | None = None,
        reference_date: date | None = None,
    ) -> PageResult:
        contexts = self.build_page_rows(
            batch_id=batch_id,
            min_limit=min_limit,
            warn_days=warn_days,
            reference_date=reference_date,
        )
        contexts = self._apply_view_filters(contexts, search=search, status=status, officer=officer)
        contexts.sort(
            key=lambda context: (
                _date_sort_value(context.row.expiry_date),
                -float(context.row.approved_amount or 0),
                context.row.customer_code,
            )
        )
        total = len(contexts)
        page_size = max(1, int(page_size))
        page = max(1, int(page))
        start = (page - 1) * page_size
        end = start + page_size
        return PageResult(
            rows=[credit_limit_context_to_page_dict(context) for context in contexts[start:end]],
            total_rows=total,
            page=page,
            page_size=page_size,
        )

    def dashboard_credit_limits(
        self,
        *,
        batch_id: str | int | None = None,
        search: str = "",
        status: str = "",
        officer: str = "",
        min_limit: float | None = None,
        warn_days: int | None = None,
        reference_date: date | None = None,
    ) -> DashboardData:
        contexts = self.calculate_kpis(
            batch_id=batch_id,
            min_limit=min_limit,
            warn_days=warn_days,
            reference_date=reference_date,
        )
        contexts = self._apply_view_filters(contexts, search=search, status=status, officer=officer)
        status_counts: dict[str, int] = {}
        officer_stats: dict[str, dict[str, object]] = {}
        month_counts: dict[str, int] = {}
        total_limit = 0.0
        total_balance = 0.0
        outstanding_available = not contexts or any(context.outstanding_balance_available for context in contexts)
        for context in contexts:
            row = context.row
            row_status = row.status or "Không rõ"
            status_counts[row_status] = status_counts.get(row_status, 0) + 1
            row_officer = row.officer or "Không xác định CBTD"
            officer_key = _officer_identity(row)
            stats = officer_stats.setdefault(
                officer_key,
                {
                    "officer_code": row.officer_code,
                    "officer_name": row_officer,
                    "expired_count": 0,
                    "expiring_count": 0,
                    "total_approved_limit": 0.0,
                    "total_outstanding_balance": 0.0,
                },
            )
            if row_status == "Đã hết hạn":
                stats["expired_count"] = int(stats["expired_count"]) + 1
            elif row_status == "Sắp hết hạn":
                stats["expiring_count"] = int(stats["expiring_count"]) + 1
            stats["total_approved_limit"] = float(stats["total_approved_limit"]) + float(row.approved_amount or 0)
            stats["total_outstanding_balance"] = float(stats["total_outstanding_balance"]) + float(row.outstanding_balance or 0)
            expiry = row.expiry_date
            month = expiry.isoformat()[:7] if isinstance(expiry, date) else "Không rõ"
            month_counts[month] = month_counts.get(month, 0) + 1
            total_limit += float(row.approved_amount or 0)
            total_balance += float(row.outstanding_balance or 0)
        metrics = (
            DashboardMetric(
                "HĐTD đã hết hạn",
                str(status_counts.get("Đã hết hạn", 0)),
                "Số HĐTD có ngày hết hạn nhỏ hơn ngày tham chiếu.",
            ),
            DashboardMetric(
                "HĐTD sắp hết hạn",
                str(status_counts.get("Sắp hết hạn", 0)),
                "Số HĐTD còn từ 0 đến số ngày cảnh báo.",
            ),
            DashboardMetric(
                "Tổng HĐTD cảnh báo",
                str(sum(status_counts.values())),
                "HĐTD đã hết hạn + HĐTD sắp hết hạn.",
            ),
            DashboardMetric(
                "Tổng hạn mức",
                _format_integer_vn(total_limit),
                f"Tổng hạn mức: {_format_integer_vn(total_limit)} đồng",
            ),
            DashboardMetric(
                "Tổng dư nợ",
                _format_integer_vn(total_balance) if outstanding_available else "—",
                (
                    f"Tổng dư nợ: {_format_integer_vn(total_balance)} đồng"
                    if outstanding_available
                    else "Batch này không có dữ liệu dư nợ HĐTD."
                ),
            ),
        )
        officer_payload = tuple(
            (
                str(stats["officer_code"]),
                str(stats["officer_name"] or "Không xác định CBTD"),
                int(stats["expired_count"]),
                int(stats["expiring_count"]),
                int(stats["expired_count"]) + int(stats["expiring_count"]),
                float(stats["total_approved_limit"]),
                float(stats["total_outstanding_balance"]),
            )
            for stats in sorted(
                officer_stats.values(),
                key=lambda item: (-(int(item["expired_count"]) + int(item["expiring_count"])), str(item["officer_name"])),
            )[:10]
        )
        return DashboardData(
            metrics=metrics,
            bars=tuple((item[1], item[4]) for item in officer_payload),
            lines=tuple(sorted(month_counts.items())),
            pies=officer_payload,
        )

    def distinct_values(
        self,
        column_name: str,
        *,
        batch_id: str | int | None = None,
        min_limit: float | None = None,
        warn_days: int | None = None,
        reference_date: date | None = None,
    ) -> list[str]:
        if column_name == "status":
            values = {
                context.row.status.strip()
                for context in self.filter_rows(
                    batch_id=batch_id,
                    min_limit=min_limit,
                    warn_days=warn_days,
                    reference_date=reference_date,
                )
            }
        elif column_name == "officer":
            values = {
                row.officer.strip()
                for _meta, rows in self._iter_batch_rows(batch_id)
                for row in rows
            }
        else:
            raise SummaryError("Trường lọc hạn mức không hợp lệ.")
        return sorted((value for value in values if value), key=str.casefold)

    def _apply_view_filters(
        self,
        contexts: list[CreditLimitRowContext],
        *,
        search: str,
        status: str,
        officer: str,
    ) -> list[CreditLimitRowContext]:
        search_text = str(search or "").strip().casefold()
        status_text = str(status or "").strip()
        officer_text = str(officer or "").strip()
        if status_text:
            contexts = [context for context in contexts if context.row.status == status_text]
        if officer_text:
            contexts = [context for context in contexts if context.row.officer == officer_text]
        if search_text:
            contexts = [
                context
                for context in contexts
                if search_text
                in " ".join(
                    (
                        context.row.customer_code,
                        context.row.customer_name,
                        context.row.contract_number,
                        context.row.officer_code,
                        context.row.officer,
                        context.row.address,
                    )
                ).casefold()
            ]
        return contexts

    def delete_batch(self, batch_id: str | int) -> Path:
        metadata = self.get_batch(batch_id)
        if metadata is None:
            raise SummaryError("Không tìm thấy batch hạn mức cần xóa.")
        self.ensure_storage()
        suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
        trash_file = self.trash_path / f"{metadata.file_path.stem}_deleted_{suffix}{metadata.file_path.suffix}"
        shutil.move(str(metadata.file_path), str(trash_file))
        self.invalidate_cache()
        return trash_file

    def get_batch(self, batch_id: str | int | None) -> CreditLimitBatchMetadata | None:
        if batch_id is None or str(batch_id).strip() in {"", "0"}:
            return None
        clean = str(batch_id).strip()
        for metadata in self.list_batches():
            if metadata.batch_id == clean or metadata.file_name == clean:
                return metadata
        return None

    def backup_storage(
        self,
        destination: Path | None = None,
        *,
        progress: Callable[[str], None] | None = None,
    ) -> Path:
        self.ensure_storage()
        if destination is None:
            destination = self.backup_path / f"HMHETHAN_Backup_{datetime.now():%Y%m%d_%H%M%S}.zip"
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        batches = self.list_batches()
        manifest = {
            "backup_version": 1,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "app_version": "AgribankV3",
            "file_count": len(batches),
            "total_size_bytes": sum(batch.file_path.stat().st_size for batch in batches if batch.file_path.exists()),
            "files": [],
        }
        temp_zip = destination.with_name(f".{destination.name}.{uuid.uuid4().hex}.tmp")
        try:
            with zipfile.ZipFile(temp_zip, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for batch in batches:
                    if progress:
                        progress(f"Đang sao lưu {batch.file_name}...")
                    archive.write(batch.file_path, arcname=batch.file_name)
                    manifest["files"].append(
                        {
                            "name": batch.file_name,
                            "sha256": _file_sha256(batch.file_path),
                            "size": batch.file_path.stat().st_size,
                        }
                    )
                archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
            os.replace(temp_zip, destination)
        except Exception:
            if temp_zip.exists():
                temp_zip.unlink(missing_ok=True)
            raise
        return destination

    def restore_storage(
        self,
        archive_path: Path,
        *,
        conflict_policy: str = "skip",
        progress: Callable[[str], None] | None = None,
    ) -> dict[str, int]:
        self.ensure_storage()
        archive_path = Path(archive_path)
        if not archive_path.exists():
            raise SummaryError("File sao lưu hạn mức không tồn tại.")
        if conflict_policy not in {"skip", "overwrite", "keep_both"}:
            raise SummaryError("Chính sách khôi phục không hợp lệ.")
        restored = skipped = overwritten = invalid = 0
        temp_dir = self.temp_path / f"restore_{uuid.uuid4().hex}"
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            with zipfile.ZipFile(archive_path, "r") as archive:
                manifest = self._load_backup_manifest(archive)
                expected_hashes = {
                    str(item.get("name")): str(item.get("sha256") or "")
                    for item in manifest.get("files", [])
                    if isinstance(item, dict)
                }
                for member in archive.namelist():
                    if member == "manifest.json" or "/" in member or "\\" in member:
                        continue
                    if not member.lower().endswith(".xlsx"):
                        continue
                    if progress:
                        progress(f"Đang khôi phục {member}...")
                    extracted = temp_dir / Path(member).name
                    with archive.open(member) as source, extracted.open("wb") as target:
                        shutil.copyfileobj(source, target)
                    try:
                        self._validate_workbook(extracted)
                    except Exception:
                        invalid += 1
                        continue
                    expected_hash = expected_hashes.get(member)
                    if expected_hash and _file_sha256(extracted).lower() != expected_hash.lower():
                        invalid += 1
                        continue
                    target = self.storage_path / Path(member).name
                    if target.exists():
                        if conflict_policy == "skip":
                            skipped += 1
                            continue
                        if conflict_policy == "keep_both":
                            target = self._unique_restore_target(target)
                        else:
                            overwritten += 1
                    os.replace(extracted, target)
                    restored += 1
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
        self.invalidate_cache()
        return {"restored": restored, "skipped": skipped, "overwritten": overwritten, "invalid": invalid}

    def maintenance_status(self) -> CreditLimitStorageStatus:
        self.ensure_storage()
        valid: list[CreditLimitBatchMetadata] = []
        invalid: list[tuple[str, str]] = []
        temporary = 0
        total_size = 0
        for path in self.storage_path.iterdir():
            if not path.is_file():
                continue
            if path.name.startswith("~$") or path.suffix.lower() != ".xlsx" or path.name.startswith("."):
                temporary += 1
                continue
            total_size += path.stat().st_size
            try:
                valid.append(self._metadata_from_file(path))
            except Exception as exc:
                invalid.append((path.name, str(exc)))
        valid.sort(key=lambda item: item.imported_at or datetime.min, reverse=True)
        return CreditLimitStorageStatus(
            storage_path=self.storage_path,
            valid_files=len(valid),
            invalid_files=len(invalid),
            temporary_files=temporary,
            total_size_bytes=total_size,
            last_checked_at=datetime.now(),
            batches=tuple(valid),
            invalid_entries=tuple(invalid),
        )

    def invalidate_cache(self) -> None:
        self._metadata_cache.clear()
        self._row_cache.clear()

    def build_page_rows(
        self,
        *,
        batch_id: str | int | None,
        min_limit: float | None,
        warn_days: int | None,
        reference_date: date | None,
    ) -> list[CreditLimitRowContext]:
        return self.filter_rows(
            batch_id=batch_id,
            min_limit=min_limit,
            warn_days=warn_days,
            reference_date=reference_date,
        )

    def calculate_kpis(
        self,
        *,
        batch_id: str | int | None,
        min_limit: float | None,
        warn_days: int | None,
        reference_date: date | None,
    ) -> list[CreditLimitRowContext]:
        return self.filter_rows(
            batch_id=batch_id,
            min_limit=min_limit,
            warn_days=warn_days,
            reference_date=reference_date,
        )

    def filter_rows(
        self,
        *,
        batch_id: str | int | None,
        min_limit: float | None,
        warn_days: int | None,
        reference_date: date | None,
    ) -> list[CreditLimitRowContext]:
        from ..services import filter_credit_limit_rows

        output: list[CreditLimitRowContext] = []
        explicit_filter = min_limit is not None or warn_days is not None or reference_date is not None
        for metadata, rows, has_outstanding_balance in self._iter_batch_rows_with_profile(batch_id):
            effective_reference = reference_date or metadata.reference_date_at_import or date.today()
            effective_min = float(min_limit) if min_limit is not None else float(metadata.minimum_limit_at_import or 0)
            effective_warn = int(warn_days) if warn_days is not None else int(metadata.warning_days_at_import or 30)
            filtered = filter_credit_limit_rows(
                rows,
                min_limit=effective_min,
                warn_days=effective_warn,
                reference_date=effective_reference,
            )
            for row in filtered:
                output.append(
                    CreditLimitRowContext(
                        metadata=metadata,
                        row=row,
                        reference_date=effective_reference,
                        warn_days=effective_warn,
                        min_limit=effective_min,
                        dynamic_status=explicit_filter,
                        outstanding_balance_available=has_outstanding_balance,
                    )
                )
        return output

    def _iter_batch_rows(
        self,
        batch_id: str | int | None = None,
    ) -> Iterable[tuple[CreditLimitBatchMetadata, tuple[CreditLimitRow, ...]]]:
        for metadata, rows, _has_outstanding_balance in self._iter_batch_rows_with_profile(batch_id):
            yield metadata, rows

    def _iter_batch_rows_with_profile(
        self,
        batch_id: str | int | None = None,
    ) -> Iterable[tuple[CreditLimitBatchMetadata, tuple[CreditLimitRow, ...], bool]]:
        metadata = self.get_batch(batch_id) if batch_id else None
        batches = [metadata] if metadata else self.list_batches()
        for batch in batches:
            if batch is None:
                continue
            rows, has_outstanding_balance = self._rows_from_file_with_profile(batch.file_path)
            yield batch, rows, has_outstanding_balance

    def _iter_candidate_files(self) -> Iterable[Path]:
        self.ensure_storage()
        for path in self.storage_path.iterdir():
            if not path.is_file():
                continue
            if path.name.startswith("~$") or path.name.startswith("."):
                continue
            if path.suffix.lower() != ".xlsx":
                continue
            yield path

    def _metadata_from_file(self, path: Path) -> CreditLimitBatchMetadata:
        path = Path(path)
        stat = path.stat()
        cached = self._metadata_cache.get(path)
        if cached and cached[0] == stat.st_mtime_ns and cached[1] == stat.st_size:
            return cached[2]
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if META_SHEET_NAME not in workbook.sheetnames:
                raise SummaryError(f"Workbook thiếu sheet {META_SHEET_NAME}.")
            sheet = workbook[META_SHEET_NAME]
            values = _read_key_value_sheet(sheet)
            metadata = _metadata_from_values(values, path)
        finally:
            workbook.close()
        self._metadata_cache[path] = (stat.st_mtime_ns, stat.st_size, metadata)
        return metadata

    def _rows_from_file(self, path: Path) -> tuple[CreditLimitRow, ...]:
        rows, _has_outstanding_balance = self._rows_from_file_with_profile(path)
        return rows

    def _rows_from_file_with_profile(self, path: Path) -> tuple[tuple[CreditLimitRow, ...], bool]:
        path = Path(path)
        stat = path.stat()
        cached = self._row_cache.get(path)
        if cached and cached[0] == stat.st_mtime_ns and cached[1] == stat.st_size:
            return cached[2], cached[3]
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if DATA_SHEET_NAME not in workbook.sheetnames:
                raise SummaryError(f"Workbook thiếu sheet {DATA_SHEET_NAME}.")
            sheet = workbook[DATA_SHEET_NAME]
            rows, has_outstanding_balance = _read_data_sheet(sheet)
        finally:
            workbook.close()
        self._row_cache[path] = (stat.st_mtime_ns, stat.st_size, rows, has_outstanding_balance)
        return rows, has_outstanding_balance

    def _write_metadata_sheet(self, sheet: Worksheet, metadata: CreditLimitBatchMetadata) -> None:
        sheet.append(("key", "value"))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7A003C")
        for key, value in _metadata_values(metadata).items():
            sheet.append((key, _excel_value(value)))
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 72
        for row in sheet.iter_rows(min_row=2, max_col=2):
            row[0].alignment = Alignment(vertical="top")
            row[1].alignment = Alignment(vertical="top", wrap_text=True)

    def _write_data_sheet(
        self,
        sheet: Worksheet,
        batch_id: str,
        rows: Sequence[CreditLimitRow],
    ) -> None:
        sheet.append(DATA_HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7A003C")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            sheet.append(credit_limit_row_to_excel_values(row=row, batch_id=batch_id))
        for index, header in enumerate(DATA_HEADERS, start=1):
            column = sheet.cell(row=1, column=index).column_letter
            if header in TEXT_COLUMNS:
                sheet.column_dimensions[column].width = 18 if header != "address" else 42
                for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                    for item in cell:
                        item.number_format = "@"
                        item.alignment = Alignment(vertical="top")
            elif header in MONEY_COLUMNS:
                sheet.column_dimensions[column].width = 18
                for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                    for item in cell:
                        item.number_format = "#,##0"
                        item.alignment = Alignment(horizontal="right")
            elif header in DATE_COLUMNS:
                sheet.column_dimensions[column].width = 14
                for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                    for item in cell:
                        item.number_format = "dd/mm/yyyy"
                        item.alignment = Alignment(horizontal="center")
            else:
                sheet.column_dimensions[column].width = 14
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    def _validate_workbook(self, path: Path) -> None:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            missing = {META_SHEET_NAME, DATA_SHEET_NAME}.difference(workbook.sheetnames)
            if missing:
                raise SummaryError(f"Workbook thiếu sheet: {', '.join(sorted(missing))}.")
            sheet = workbook[DATA_SHEET_NAME]
            headers = tuple(str(cell.value or "").strip() for cell in next(sheet.iter_rows(max_row=1)))
            if headers[: len(DATA_HEADERS)] != DATA_HEADERS:
                raise SummaryError("Workbook HMHETHAN không đúng cấu trúc cột.")
        finally:
            workbook.close()

    def _new_batch_id(self, imported_at: datetime, source_hash: str) -> str:
        prefix = f"{imported_at:%Y%m%d_%H%M%S}_{source_hash[:8].lower()}"
        candidate = prefix
        existing = {batch.batch_id for batch in self.list_batches()}
        while candidate in existing:
            candidate = f"{prefix}_{uuid.uuid4().hex[:6]}"
        return candidate

    def _new_file_name(self, imported_at: datetime, source_stem: str, source_hash: str) -> str:
        safe_stem = _safe_file_part(source_stem)[:60] or "LN01"
        prefix = f"HMHETHAN_{imported_at:%Y%m%d_%H%M%S}_{safe_stem}_{source_hash[:8].lower()}"
        candidate = f"{prefix}.xlsx"
        counter = 1
        while (self.storage_path / candidate).exists():
            counter += 1
            candidate = f"{prefix}_{counter}.xlsx"
        return candidate

    def _unique_restore_target(self, target: Path) -> Path:
        stem = target.stem
        suffix = target.suffix
        counter = 2
        candidate = target.with_name(f"{stem}_{counter}{suffix}")
        while candidate.exists():
            counter += 1
            candidate = target.with_name(f"{stem}_{counter}{suffix}")
        return candidate

    def _load_backup_manifest(self, archive: zipfile.ZipFile) -> dict[str, object]:
        if "manifest.json" not in archive.namelist():
            raise SummaryError("File sao lưu HMHETHAN thiếu manifest.json.")
        with archive.open("manifest.json") as handle:
            return json.loads(handle.read().decode("utf-8"))


def _read_key_value_sheet(sheet: Worksheet) -> dict[str, object]:
    output: dict[str, object] = {}
    for key, value in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        clean_key = str(key or "").strip()
        if clean_key:
            output[clean_key] = value
    return output


def _read_data_sheet(sheet: Worksheet) -> tuple[tuple[CreditLimitRow, ...], bool]:
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = tuple(str(value or "").strip() for value in next(rows))
    except StopIteration:
        return (), True
    missing = [header for header in REQUIRED_DATA_HEADERS_FOR_READ if header not in headers]
    if missing:
        raise SummaryError(f"Workbook HMHETHAN thiếu cột: {', '.join(missing)}.")
    has_outstanding_balance = "outstanding_balance" in headers
    output: list[CreditLimitRow] = []
    for raw in rows:
        record = {header: raw[index] if index < len(raw) else None for index, header in enumerate(headers)}
        output.append(excel_record_to_credit_limit_row(record))
    return tuple(output), has_outstanding_balance


def normalize_credit_limit_row(value: CreditLimitRow | Mapping[str, Any]) -> CreditLimitRow:
    if isinstance(value, CreditLimitRow):
        return value
    if isinstance(value, Mapping):
        return excel_record_to_credit_limit_row(value)
    raise TypeError(f"Unsupported credit limit row type: {type(value).__name__}")


def credit_limit_row_to_excel_values(row: CreditLimitRow, batch_id: str) -> list[Any]:
    if not isinstance(row, CreditLimitRow):
        raise TypeError(f"Expected CreditLimitRow, got {type(row).__name__}")
    values = (
        batch_id,
        row.branch_code,
        row.customer_code,
        row.customer_name,
        row.account_number,
        row.credit_line_type,
        row.contract_number,
        row.approved_date,
        row.approved_amount,
        row.expiry_date,
        row.outstanding_balance,
        row.officer_code,
        row.officer,
        row.address,
        row.source_row_count,
    )
    return [_excel_data_value(value) for value in values]


def excel_record_to_credit_limit_row(record: Mapping[str, Any]) -> CreditLimitRow:
    if not isinstance(record, Mapping):
        raise TypeError(f"Expected Mapping, got {type(record).__name__}")
    raw_officer = _clean_mapping_text(record, "officer_name", "officer")
    inline_code, inline_name = _split_inline_officer(raw_officer)
    officer_code = _clean_mapping_text(record, "officer_code") or inline_code
    return CreditLimitRow(
        customer_code=_clean_mapping_text(record, "customer_code"),
        customer_name=_clean_mapping_text(record, "customer_name"),
        contract_number=_clean_mapping_text(record, "approval_sequence", "contract_number"),
        approved_date=_to_date(_mapping_value(record, "approval_date", "approved_date")),
        approved_amount=_to_float(_mapping_value(record, "approved_limit", "approved_amount")),
        outstanding_balance=_to_float(_mapping_value(record, "outstanding_balance")),
        expiry_date=_to_date(_mapping_value(record, "maturity_date", "expiry_date")),
        address=_clean_mapping_text(record, "address"),
        officer=inline_name or raw_officer,
        officer_code=officer_code,
        note=_clean_mapping_text(record, "note"),
        days_to_expiry=_mapping_int_or_none(record, "days_to_expiry"),
        status=_clean_mapping_text(record, "status"),
        branch_code=_clean_mapping_text(record, "branch_code"),
        account_number=_clean_mapping_text(record, "account_number"),
        credit_line_type=_clean_mapping_text(record, "credit_line_type") or "Line of Credit",
        source_row_count=max(1, _to_int(_mapping_value(record, "source_row_count"))),
    )


def credit_limit_context_to_page_dict(context: CreditLimitRowContext) -> dict[str, object]:
    row = context.row
    metadata = context.metadata
    return {
        "batch_id": metadata.batch_id,
        "batch_name": metadata.batch_name,
        "customer_code": row.customer_code,
        "customer_name": row.customer_name,
        "contract_number": row.contract_number,
        "approved_date": row.approved_date,
        "approved_amount": row.approved_amount,
        "outstanding_balance": row.outstanding_balance,
        "expiry_date": row.expiry_date,
        "address": row.address,
        "officer": row.officer,
        "officer_code": row.officer_code,
        "note": row.note,
        "days_to_expiry": row.days_to_expiry,
        "status": row.status,
        "branch_code": row.branch_code,
        "account_number": row.account_number,
        "credit_line_type": row.credit_line_type,
        "source_row_count": row.source_row_count,
        "reference_date": context.reference_date,
        "warn_days": context.warn_days,
        "min_limit": context.min_limit,
        "dynamic_status": context.dynamic_status,
    }


def _mapping_value(record: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in record and record.get(key) not in (None, ""):
            return record.get(key)
    return None


def _clean_mapping_text(record: Mapping[str, Any], *keys: str) -> str:
    return _clean_text(_mapping_value(record, *keys))


def _split_inline_officer(value: object) -> tuple[str, str]:
    text = _clean_text(value)
    if text.startswith("[") and "]" in text:
        code, name = text[1:].split("]", 1)
        return code.strip(), name.strip()
    return "", text


def _mapping_int_or_none(record: Mapping[str, Any], *keys: str) -> int | None:
    value = _mapping_value(record, *keys)
    if value in (None, ""):
        return None
    return _to_int(value)


def _clean_text(value: object) -> str:
    text = str(value or "").strip()
    if text.startswith("'") and len(text) > 1 and text[1] in ("=", "+", "-", "@"):
        return text[1:]
    return text


def _metadata_values(metadata: CreditLimitBatchMetadata) -> dict[str, object]:
    values = asdict(metadata)
    values.pop("file_path", None)
    values.pop("file_size", None)
    values.pop("modified_at", None)
    values["schema_version"] = CREDIT_LIMIT_BATCH_SCHEMA_VERSION
    return values


def _metadata_from_values(values: dict[str, object], path: Path) -> CreditLimitBatchMetadata:
    stat = path.stat()
    return CreditLimitBatchMetadata(
        batch_id=str(values.get("batch_id") or path.stem),
        batch_name=str(values.get("batch_name") or path.stem),
        file_path=path,
        file_name=path.name,
        source_file_name=str(values.get("source_file_name") or ""),
        source_file_sha256=str(values.get("source_file_sha256") or ""),
        source_file_size=_to_int(values.get("source_file_size")),
        imported_at=_to_datetime(values.get("imported_at")),
        imported_by=str(values.get("imported_by") or ""),
        app_version=str(values.get("app_version") or ""),
        source_row_count=_to_int(values.get("source_row_count")),
        accepted_row_count=_to_int(values.get("accepted_row_count")),
        rejected_row_count=_to_int(values.get("rejected_row_count")),
        warning_count=_to_int(values.get("warning_count")),
        reference_date_at_import=_to_date(values.get("reference_date_at_import")),
        minimum_limit_at_import=_to_float(values.get("minimum_limit_at_import")),
        warning_days_at_import=_to_int(values.get("warning_days_at_import")) or 30,
        expired_count_at_import=_to_int(values.get("expired_count_at_import")),
        expiring_count_at_import=_to_int(values.get("expiring_count_at_import")),
        status=str(values.get("status") or "OK"),
        notes=str(values.get("notes") or ""),
        schema_version=str(values.get("schema_version") or CREDIT_LIMIT_BATCH_SCHEMA_VERSION),
        file_size=stat.st_size,
        modified_at=datetime.fromtimestamp(stat.st_mtime),
    )


def _read_text(raw: tuple[object, ...], index: dict[str, int], key: str) -> str:
    value = raw[index[key]] if index[key] < len(raw) else ""
    return str(value or "").strip()


def _read_float(raw: tuple[object, ...], index: dict[str, int], key: str) -> float:
    value = raw[index[key]] if index[key] < len(raw) else 0
    return _to_float(value)


def _read_date(raw: tuple[object, ...], index: dict[str, int], key: str) -> date | None:
    value = raw[index[key]] if index[key] < len(raw) else None
    return _to_date(value)


def _to_int(value: object) -> int:
    try:
        return int(float(str(value or "0").replace(",", "")))
    except (TypeError, ValueError):
        return 0


def _to_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "." in text:
        parts = text.split(".")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]) and len(parts[0]) <= 3:
            text = "".join(parts)
    elif "," in text:
        parts = text.split(",")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]) and len(parts[0]) <= 3:
            text = "".join(parts)
        else:
            text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _to_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _to_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _excel_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return _escape_excel_text(value)
    return value


def _excel_data_value(value: object) -> object:
    if isinstance(value, str):
        return _escape_excel_text(value)
    return value


def _escape_excel_text(value: str) -> str:
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _format_integer_vn(value: object) -> str:
    try:
        number = int(round(float(value or 0)))
    except (TypeError, ValueError):
        number = 0
    return f"{number:,}".replace(",", ".")


def _officer_identity(row: CreditLimitRow) -> str:
    code = str(row.officer_code or "").strip()
    if code:
        return f"CODE:{code.casefold()}"
    name = row.officer.strip() or "Không xác định CBTD"
    return name.casefold()


def _safe_file_part(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    text = re.sub(r"_+", "_", text).strip("._-")
    return text


def _date_sort_value(value: object) -> str:
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "")
