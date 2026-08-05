from __future__ import annotations

from collections.abc import Callable, Sequence
from contextlib import closing, contextmanager
from dataclasses import dataclass, replace
from datetime import date, datetime
import csv
import getpass
import hashlib
import logging
from io import StringIO
import os
from pathlib import Path
import re
import shutil
import sqlite3
import tempfile
from time import perf_counter
from typing import Iterable
from enum import StrEnum

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from agribank_v3.features.credit.summary.models import (
    CreditLimitRow,
    ImportResult,
    LoanSnapshotRow,
    NIM_DN_CONFIG,
    NIM_NV_CONFIG,
    NormalizedLn01Row,
    NormalizedLoanRow,
    OfficerHistory,
    OfficerHistoryPoint,
    NimConfig,
    NimRow,
    SummaryDataType,
    SummaryError,
)
from agribank_v3.features.credit.summary.reports import (
    export_credit_limit_vba,
    export_loan_compare_vba,
    export_nim_vba,
)
from agribank_v3.features.credit.summary.repository import SummaryRepository
from agribank_v3.features.credit.summary.customer.repository import CustomerRepository
from agribank_v3.features.credit.summary.customer.services import (
    CustomerAggregationService,
    build_customer_code,
    build_office_code,
    classify_office_type,
    map_customer_type_code,
    normalize_debt_group,
    normalize_trctcd,
    normalize_customer_sequence,
    split_officer as split_customer_officer,
)
from agribank_v3.features.credit.summary.ln01 import (
    ln01_has_group_code_header,
    parse_ln01_bytes,
    project_credit_limit_rows,
)
from agribank_v3.features.credit.summary.credit_limit.excel_batch_store import PendingCreditLimitBatch
from agribank_v3.features.credit.summary.credit_limit.models import (
    CreditLimitBatchLookup,
    CreditLimitBatchLookupState,
    CreditLimitBatchMetadata,
)
from agribank_v3.features.credit.summary.credit_report import CreditReportRepository
from agribank_v3.features.settings.unit_directory.service import (
    UnitDirectoryService,
    get_unit_directory_service,
)
from agribank_v3.runtime_paths import application_root


ProgressCallback = Callable[[str], None]


class Ln01DuplicateDecision(StrEnum):
    CREATE_BOTH = "CREATE_BOTH"
    CREATE_CREDIT_ONLY = "CREATE_CREDIT_ONLY"
    CREATE_HMHE_THAN_ONLY = "CREATE_HMHEthan_ONLY"
    OVERWRITE_CREDIT = "OVERWRITE_CREDIT"
    OVERWRITE_BOTH = "OVERWRITE_BOTH"
    CANCEL = "CANCEL"


@dataclass(frozen=True, slots=True)
class CreditPeriodStatus:
    period: str
    exists: bool
    same_sha: bool
    different_sha: bool
    source_file_name: str = ""
    source_sha256: str = ""
    imported_at: str = ""
    imported_by: str = ""


@dataclass(frozen=True, slots=True)
class Ln01ImportDecisionResolution:
    default_decision: Ln01DuplicateDecision
    allowed_decisions: tuple[Ln01DuplicateDecision, ...]
    requires_confirmation: bool
    reason: str = ""


@dataclass(frozen=True, slots=True)
class Ln01ImportContext:
    period: str
    branch_code: str
    source_file_name: str
    source_sha256: str
    source_file_size: int
    source_row_count: int
    hmhethan_status: CreditLimitBatchLookup
    credit_status: CreditPeriodStatus
    resolution: Ln01ImportDecisionResolution


Ln01DecisionProvider = Callable[[Ln01ImportContext], Ln01DuplicateDecision | str | None]


@dataclass(frozen=True, slots=True)
class Ln01PreparedImport:
    file_path: Path
    period: str
    branch_code: str
    source_sha256: str
    source_file_size: int
    source_row_count: int
    normalized_rows: tuple[NormalizedLn01Row, ...]
    agreements: tuple[CreditLimitRow, ...]
    accepted_rows: tuple[CreditLimitRow, ...]
    reference_date: date
    min_limit: float
    warn_days: int
    has_group_code_column: bool
    context: Ln01ImportContext


NIM_COMMON_HEADERS = (
    "BRCD",
    "FTP",
    "INTRT",
    "MUCFTPDC",
    "LDRBAL",
)

CUSTOMER_FTPLN_REQUIRED_HEADERS = (
    "BRCD",
    "CUSTSEQ",
    "CUSTNM",
    "CUSTTP",
    "FTPCD",
    "LDRBAL",
    "FTP",
    "INTRT",
    "MUCFTPDC",
    "CBTD",
)

CUSTTP_TOTAL = "[Tất cả KH]"
BRANCH_TOTAL = "[Tổng Chi Nhánh]"
CREDIT_LIMIT_EXPIRED_NOTE = "Hợp đồng hạn mức tín dụng ã quá hạn đến thời đểm hiện tại"
CREDIT_LIMIT_SOON_NOTE_TEMPLATE = "Hợp đồng tín dụng dến hạn trong vòng {warn_days} ngày tới theo thời đểm hiện tại"
OFFICER_HISTORY_HEADERS = (
    "Kỳ",
    "Dư nợ",
    "Lãi suất bình quân",
    "NIM trước ĐC",
    "NIM sau ĐC",
)


@dataclass(frozen=True, slots=True)
class ParsedNimFile:
    rows: list[NimRow]
    period: str
    source_hash: str
    customer_rows: list[NormalizedLoanRow]
    customer_missing_headers: tuple[str, ...]
    customer_source_total_balance: float
    customer_invalid_row_count: int
    customer_warning_count: int
    customer_warnings: tuple[str, ...]
    unit_warnings: tuple[str, ...]
    debt_group_header_present: bool = False
    debt_group_valid_row_count: int = 0
    debt_group_1_row_count: int = 0
    debt_group_2_row_count: int = 0
    debt_group_3_row_count: int = 0
    debt_group_4_row_count: int = 0
    debt_group_5_row_count: int = 0
    debt_group_unknown_row_count: int = 0
    debt_group_invalid_samples: tuple[str, ...] = ()


def import_nim_folder(
    repository: SummaryRepository,
    folder_path: Path,
    config: NimConfig,
    *,
    credit_card_rate: float = 0.0,
    export_path: Path | None = None,
    imported_by: str | None = None,
    progress: ProgressCallback | None = None,
    replace_existing_periods: bool = False,
) -> ImportResult:
    folder_path = Path(folder_path)
    if not folder_path.is_dir():
        raise SummaryError("Thư mục import NIM không tồn tại.")
    files = sorted(
        path for path in folder_path.iterdir()
        if path.is_file()
        and path.suffix.casefold() == ".csv"
        and config.file_pattern_token in path.name.casefold()
    )
    if not files:
        raise SummaryError(f"Không tìm thấy file CSV phù hợp {config.file_pattern_token} trong thư mục đã chọn.")

    parsed_files: list[tuple[Path, ParsedNimFile, int]] = []
    user = imported_by or _current_user()
    customer_aggregation = (
        CustomerAggregationService(folder_path)
        if config.data_type == SummaryDataType.NIM_DN
        else None
    )
    unit_directory = get_unit_directory_service(repository.main_database_path)
    customer_missing_messages: list[str] = []
    unit_warning_messages: list[str] = []
    parse_started = perf_counter()
    for file_index, file_path in enumerate(files, start=1):
        if progress:
            progress(f"Đang đọc file {file_index}/{len(files)}: {file_path.name}")
        started = perf_counter()
        parsed = _parse_nim_file(
            file_path,
            config,
            credit_card_rate=credit_card_rate,
            unit_directory=unit_directory,
        )
        duration_ms = int((perf_counter() - started) * 1000)
        parsed_files.append((file_path, parsed, duration_ms))
        unit_warning_messages.extend(parsed.unit_warnings)
        if customer_aggregation is not None:
            if parsed.customer_missing_headers:
                customer_missing_messages.append(
                    f"{file_path.name} thiếu cột Customer.db: {', '.join(parsed.customer_missing_headers)}"
                )
            else:
                customer_aggregation.add_file(
                    parsed.customer_rows,
                    period=parsed.period,
                    file_path=file_path,
                    file_hash=parsed.source_hash,
                    source_row_count=len(parsed.rows),
                    source_total_balance=parsed.customer_source_total_balance,
                    invalid_row_count=parsed.customer_invalid_row_count,
                    warning_count=parsed.customer_warning_count,
                    warnings=parsed.customer_warnings,
                    debt_group_invalid_samples=parsed.debt_group_invalid_samples,
                    debt_group_header_present=parsed.debt_group_header_present,
                )
        if progress:
            customer_count = len({row.customer_code for row in parsed.customer_rows})
            progress(
                f"Đã đọc {len(parsed.rows):,} dòng nguồn, "
                f"{customer_count:,} khách hàng từ {file_path.name}"
            )

    customer_result = None
    if customer_aggregation is not None and not customer_missing_messages:
        customer_result = customer_aggregation.build_result()
        if customer_result.source_row_count <= 0:
            customer_result = None

    customer_repository = CustomerRepository(repository.main_database_path) if customer_result is not None else None
    credit_report_repository = (
        CreditReportRepository(repository.main_database_path)
        if config.data_type == SummaryDataType.NIM_DN
        else None
    )
    snapshot_paths = [repository.database_path]
    if customer_repository is not None:
        snapshot_paths.append(customer_repository.database_path)
    if credit_report_repository is not None:
        snapshot_paths.append(credit_report_repository.database_path)

    total_rows = 0
    first_batch_id = 0
    export_rows_buffer: list[NimRow] = []
    total_summary_rows = 0
    customer_run_id = 0
    write_started = perf_counter()
    with _sqlite_snapshots(snapshot_paths):
        if replace_existing_periods and config.data_type == SummaryDataType.NIM_DN:
            for period in sorted({parsed.period for _file_path, parsed, _duration_ms in parsed_files}):
                try:
                    repository.delete_nim_period(config.data_type, period, created_by=user)
                except SummaryError as exc:
                    if "Không có dữ liệu" not in str(exc):
                        raise
                if credit_report_repository is not None:
                    credit_report_repository.delete_credit_card_period(period)
        for file_path, parsed, duration_ms in parsed_files:
            if progress:
                progress(f"Đang ghi CreditSummary.db: {file_path.name}")
            batch_id = repository.create_batch(
                config.data_type,
                period=parsed.period,
                source_path=file_path,
                imported_by=user,
                row_count=len(parsed.rows),
                duration_ms=duration_ms,
                message=f"Import {file_path.name}",
                source_hash=parsed.source_hash,
            )
            rows = [replace(row, batch_id=batch_id) for row in parsed.rows]
            summary_rows = aggregate_nim_rows(rows)
            repository.save_nim_rows(summary_rows)
            export_rows_buffer.extend(rows)
            if not first_batch_id:
                first_batch_id = batch_id
            total_rows += len(rows)
            total_summary_rows += len(summary_rows)
            if (
                credit_report_repository is not None
                and not parsed.customer_missing_headers
            ):
                if progress:
                    progress(f"Đang ghi dư nợ thẻ DN15 sang Credit.db: {file_path.name}")
                credit_report_repository.save_credit_card_projection(
                    period=parsed.period,
                    file_name=file_path.name,
                    source_sha256=parsed.source_hash,
                    source_file_size=file_path.stat().st_size if file_path.is_file() else 0,
                    source_row_count=len(parsed.rows),
                    rows=parsed.customer_rows,
                    imported_by=user,
                    replace_period=False,
                    duration_ms=duration_ms,
                )
            if progress:
                progress(f"Đã lưu {len(summary_rows):,} dòng tổng hợp NIM từ {len(rows):,} dòng nguồn")
        if customer_repository is not None and customer_result is not None:
            if progress:
                progress("Đang ghi Customer.db")
            customer_run_id = customer_repository.save_aggregation(
                customer_result,
                replace_periods=replace_existing_periods,
                created_by=user,
                duration_ms=int((perf_counter() - write_started) * 1000),
            )
    output_path = None
    if export_path is not None and export_rows_buffer:
        detail_rows, branch_rows = build_nim_vba_tables(export_rows_buffer, config.data_type)
        output_path = export_nim_vba(detail_rows, branch_rows, Path(export_path), data_type=config.data_type)
    customer_message = ""
    if customer_result is not None:
        unknown_count = len(customer_result.unknown_ftp_codes)
        customer_message = (
            f" Customer.db: run {customer_run_id}, {customer_result.customer_count:,} khách hàng, "
            f"tổng dư nợ {customer_result.total_balance:,.0f}, "
            f"ngắn hạn {customer_result.short_term_balance:,.0f}, "
            f"trung/dài hạn {customer_result.medium_long_term_balance:,.0f}, "
            f"chưa phân loại {customer_result.other_balance:,.0f}, "
            f"KH nhiều cán bộ {customer_result.multiple_officer_customer_count:,}, "
            f"FTPCD lạ {unknown_count:,}, lỗi dòng {customer_result.invalid_row_count:,}, "
            f"nhóm nợ hợp lệ {customer_result.debt_group_valid_row_count:,}, "
            f"UNKNOWN {customer_result.debt_group_unknown_row_count:,}."
        )
    elif customer_missing_messages:
        customer_message = " Customer.db chưa cập nhật vì " + "; ".join(customer_missing_messages[:3]) + "."
    if progress:
        progress("Hoàn thành")
    unit_message = ""
    if unit_warning_messages:
        unique_warnings = list(dict.fromkeys(unit_warning_messages))
        unit_message = " Cài đặt đơn vị: " + "; ".join(unique_warnings[:5])
        if len(unique_warnings) > 5:
            unit_message += f"; còn {len(unique_warnings) - 5} cảnh báo khác"
        unit_message += "."
    return ImportResult(
        batch_id=first_batch_id,
        row_count=total_rows,
        message=(
            f"Đã import {len(files)} file, {total_rows:,} dòng nguồn, "
            f"{total_summary_rows:,} dòng tổng hợp trong {int((perf_counter() - parse_started) * 1000):,} ms."
            f"{customer_message}"
            f"{unit_message}"
        ),
        output_path=output_path,
    )


def import_nim_dn(
    repository: SummaryRepository,
    folder_path: Path,
    *,
    credit_card_rate: float = 0.0,
    export_path: Path | None = None,
    imported_by: str | None = None,
    progress: ProgressCallback | None = None,
    replace_existing_periods: bool = False,
) -> ImportResult:
    return import_nim_folder(
        repository,
        folder_path,
        NIM_DN_CONFIG,
        credit_card_rate=credit_card_rate,
        export_path=export_path,
        imported_by=imported_by,
        progress=progress,
        replace_existing_periods=replace_existing_periods,
    )


def import_nim_nv(
    repository: SummaryRepository,
    folder_path: Path,
    *,
    credit_card_rate: float = 0.0,
    export_path: Path | None = None,
    imported_by: str | None = None,
    progress: ProgressCallback | None = None,
    replace_existing_periods: bool = False,
) -> ImportResult:
    _ = replace_existing_periods
    return import_nim_folder(
        repository,
        folder_path,
        NIM_NV_CONFIG,
        credit_card_rate=credit_card_rate,
        export_path=export_path,
        imported_by=imported_by,
        progress=progress,
    )


def build_officer_history(
    repository: SummaryRepository,
    data_type: SummaryDataType,
    *,
    officer: str,
    branch: str = "",
    transaction_office: str = "",
    customer_type: str = "",
) -> OfficerHistory:
    rows = repository.get_officer_history(
        data_type,
        officer=officer,
        branch=branch,
        transaction_office=transaction_office,
        customer_type=customer_type,
    )
    points = tuple(
        OfficerHistoryPoint(
            period=str(row.get("period") or ""),
            balance=float(row.get("balance") or 0),
            average_rate=float(row.get("average_rate") or 0),
            nim_before=float(row.get("nim_before") or 0),
            nim_after=float(row.get("nim_after") or 0),
        )
        for row in rows
    )
    current = points[-1] if points else None
    return OfficerHistory(
        data_type=data_type,
        officer=officer,
        branch=branch,
        transaction_office=transaction_office,
        customer_type=customer_type,
        current_period=current.period if current else "",
        current_balance=current.balance if current else 0.0,
        current_average_rate=current.average_rate if current else 0.0,
        current_nim_before=current.nim_before if current else 0.0,
        current_nim_after=current.nim_after if current else 0.0,
        points=points,
    )


def export_officer_history_excel(history: OfficerHistory, destination: Path) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "LichSu_NIM_CBTD"
    worksheet["A1"] = "Lịch sử NIM CBTD"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A3"] = "CBTD"
    worksheet["B3"] = history.officer
    worksheet["A4"] = "Chi nhánh"
    worksheet["B4"] = history.branch
    worksheet["A5"] = "Phòng GD"
    worksheet["B5"] = history.transaction_office
    worksheet["A6"] = "Loại khách hàng"
    worksheet["B6"] = history.customer_type or "Tất cả"
    for column_index, value in enumerate(OFFICER_HISTORY_HEADERS, start=1):
        cell = worksheet.cell(8, column_index, value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row_index, point in enumerate(history.points, start=9):
        worksheet.cell(row_index, 1, point.period)
        worksheet.cell(row_index, 2, point.balance)
        worksheet.cell(row_index, 3, point.average_rate)
        worksheet.cell(row_index, 4, point.nim_before)
        worksheet.cell(row_index, 5, point.nim_after)
    for cell in worksheet["B"]:
        if cell.row >= 9:
            cell.number_format = "#,##0"
    for column_letter in ("C", "D", "E"):
        for cell in worksheet[column_letter]:
            if cell.row >= 9:
                cell.number_format = '0.00"%"'
    widths = {
        "A": 14,
        "B": 18,
        "C": 20,
        "D": 16,
        "E": 16,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width
    workbook.save(destination)
    return destination


def build_nim_vba_tables(
    rows: Iterable[NimRow],
    data_type: SummaryDataType,
) -> tuple[list[list[object]], list[list[object]]]:
    detail: dict[str, list[float]] = {}
    branch: dict[str, list[float]] = {}
    branch_has_pgd: dict[str, bool] = {}

    for row in rows:
        detail_key = "|".join(
            [
                row.period,
                row.branch_name,
                row.transaction_office,
                row.customer_type,
                row.officer,
            ]
        )
        detail_total_key = "|".join(
            [
                row.period,
                row.branch_name,
                row.transaction_office,
                CUSTTP_TOTAL,
                row.officer,
            ]
        )
        old_trctcd = row.trctcd.replace("'", "")
        if old_trctcd != "00":
            branch_has_pgd[row.branch_name] = True
        elif row.branch_name not in branch_has_pgd:
            branch_has_pgd[row.branch_name] = False

        _add_nim_values(detail, detail_key, row, data_type)
        _add_nim_values(detail, detail_total_key, row, data_type)

        branch_key = "|".join([row.period, row.branch_name, BRANCH_TOTAL, row.customer_type])
        branch_total_key = "|".join([row.period, row.branch_name, BRANCH_TOTAL, CUSTTP_TOTAL])
        branch_pgd_key = "|".join([row.period, row.branch_name, row.transaction_office, row.customer_type])
        branch_pgd_total_key = "|".join([row.period, row.branch_name, row.transaction_office, CUSTTP_TOTAL])
        _add_nim_values(branch, branch_key, row, data_type)
        _add_nim_values(branch, branch_total_key, row, data_type)
        _add_nim_values(branch, branch_pgd_key, row, data_type)
        _add_nim_values(branch, branch_pgd_total_key, row, data_type)

    detail_rows = [_nim_detail_output(key, values, data_type) for key, values in detail.items()]
    branch_rows: list[list[object]] = []
    for key, values in branch.items():
        period, branch_name, transaction_office, _customer_type = key.split("|")
        if transaction_office == "Hội sở" and branch_has_pgd.get(branch_name) is False:
            continue
        branch_rows.append(_nim_branch_output(key, values, data_type))
    return detail_rows, branch_rows


def aggregate_nim_rows(rows: Iterable[NimRow]) -> list[NimRow]:
    grouped: dict[tuple[object, ...], dict[str, object]] = {}
    for row in rows:
        officer_code, officer_name = _split_officer(row.officer)
        key = (
            row.batch_id,
            row.data_type,
            row.period,
            row.branch_code,
            row.branch_name,
            row.trctcd,
            row.transaction_office,
            row.customer_type,
            officer_code,
            officer_name,
        )
        item = grouped.get(key)
        if item is None:
            item = {
                "row": row,
                "officer_code": officer_code,
                "officer_name": officer_name,
                "balance": 0.0,
                "average_rate_numerator": 0.0,
                "numerator_before": 0.0,
                "numerator_after": 0.0,
                "source_row_count": 0,
            }
            grouped[key] = item
        item["balance"] = float(item["balance"]) + float(row.balance or 0)
        item["average_rate_numerator"] = float(item["average_rate_numerator"]) + float(row.average_rate_numerator or 0)
        item["numerator_before"] = float(item["numerator_before"]) + float(row.numerator_before or 0)
        item["numerator_after"] = float(item["numerator_after"]) + float(row.numerator_after or 0)
        item["source_row_count"] = int(item["source_row_count"]) + int(row.source_row_count or 1)

    summary_rows: list[NimRow] = []
    for item in grouped.values():
        base = item["row"]
        if not isinstance(base, NimRow):
            continue
        balance = float(item["balance"] or 0)
        average_rate_numerator = float(item["average_rate_numerator"] or 0)
        summary_rows.append(
            replace(
                base,
                officer=_format_officer(str(item["officer_code"] or ""), str(item["officer_name"] or "")),
                balance=balance,
                interest_rate=average_rate_numerator / balance if balance else 0.0,
                ftp_rate=0.0,
                adjustment_rate=0.0,
                numerator_before=float(item["numerator_before"] or 0),
                numerator_after=float(item["numerator_after"] or 0),
                average_rate_numerator=average_rate_numerator,
                source_row_count=int(item["source_row_count"] or 0),
            )
        )
    return summary_rows


def compare_loan_balances(
    repository: SummaryRepository,
    previous_file: Path,
    current_file: Path,
    *,
    export_path: Path | None = None,
    imported_by: str | None = None,
    progress: ProgressCallback | None = None,
) -> ImportResult:
    previous_file = Path(previous_file)
    current_file = Path(current_file)
    if progress:
        progress(f"Đang đọc kỳ trước: {previous_file.name}")
    previous, info = _load_loan_file(previous_file)
    if progress:
        progress(f"Đang đọc kỳ này: {current_file.name}")
    current, info_current = _load_loan_file(current_file)
    for key, value in info_current.items():
        info.setdefault(key, value)
    rows = build_loan_compare_rows(previous, current, info)
    started = perf_counter()
    period = f"{previous_file.stem} -> {current_file.stem}"
    batch_id = repository.create_batch(
        SummaryDataType.LOAN_COMPARE,
        period=period,
        source_path=current_file,
        imported_by=imported_by or _current_user(),
        row_count=len(rows),
        duration_ms=int((perf_counter() - started) * 1000),
        message=f"So sánh {previous_file.name} và {current_file.name}",
        source_hash=_combined_file_hash((previous_file, current_file)),
    )
    repository.save_loan_compare_rows(batch_id, rows)
    output_path = export_loan_compare_vba(rows, Path(export_path)) if export_path is not None and rows else None
    if progress:
        progress(f"Đã lưu {len(rows):,} dòng đối chiếu")
    return ImportResult(
        batch_id=batch_id,
        row_count=len(rows),
        message=f"Đã đối chiếu {len(rows):,} khách hàng.",
        output_path=output_path,
    )


def compare_loan_folder(
    repository: SummaryRepository,
    folder_path: Path,
    *,
    export_path: Path | None = None,
    imported_by: str | None = None,
    progress: ProgressCallback | None = None,
) -> ImportResult:
    files = sorted(
        path for path in Path(folder_path).iterdir()
        if path.is_file() and path.suffix.casefold() in {".csv", ".txt", ".xlsx"}
    )
    if len(files) < 2:
        raise SummaryError("Không tìm thấy đủ 2 file dữ liệu trong thư mục đã chọn.")
    return compare_loan_balances(
        repository,
        files[0],
        files[1],
        export_path=export_path,
        imported_by=imported_by,
        progress=progress,
    )


def build_loan_compare_rows(
    previous: dict[str, float],
    current: dict[str, float],
    info: dict[str, tuple[str, str, str]],
) -> list[LoanSnapshotRow]:
    all_keys = list(previous)
    for key in current:
        if key not in previous:
            all_keys.append(key)
    rows: list[LoanSnapshotRow] = []
    for key in all_keys:
        old_value = previous.get(key, 0.0)
        new_value = current.get(key, 0.0)
        if old_value <= 0 and new_value > 0:
            category = "Khach hang vay moi"
        elif old_value > 0 and new_value <= 0:
            category = "Khach hang tat toan"
        elif new_value > old_value:
            category = "Khach hang vay tang"
        elif new_value < old_value:
            category = "Khach hang vay giam"
        else:
            category = "Khong thay doi"
        customer_name, address, officer = info.get(key, ("", "", ""))
        rows.append(
            LoanSnapshotRow(
                customer_code=key,
                customer_name=customer_name,
                address=address,
                officer=officer,
                previous_balance=old_value,
                current_balance=new_value,
                category=category,
            )
        )
    return rows


def import_credit_limit_file(
    repository: SummaryRepository,
    file_path: Path,
    *,
    period: str | None = None,
    min_limit: float = 0.0,
    warn_days: int = 30,
    reference_date: date | None = None,
    export_path: Path | None = None,
    imported_by: str | None = None,
    duplicate_policy: str = "error",
    overwrite_report_period: bool | None = None,
    duplicate_decision: Ln01DuplicateDecision | str | None = None,
    decision_provider: Ln01DecisionProvider | None = None,
    progress: ProgressCallback | None = None,
) -> ImportResult:
    coordinator = Ln01ImportCoordinator(repository)
    return coordinator.import_file(
        file_path,
        period=period,
        min_limit=min_limit,
        warn_days=warn_days,
        reference_date=reference_date,
        export_path=export_path,
        imported_by=imported_by,
        duplicate_policy=duplicate_policy,
        overwrite_report_period=overwrite_report_period,
        duplicate_decision=duplicate_decision,
        decision_provider=decision_provider,
        progress=progress,
    )


class Ln01ImportCoordinator:
    def __init__(
        self,
        repository: SummaryRepository,
        credit_repository: CreditReportRepository | None = None,
    ) -> None:
        self.repository = repository
        self.credit_repository = credit_repository or CreditReportRepository(repository.main_database_path)

    def import_file(
        self,
        file_path: Path,
        *,
        period: str | None = None,
        min_limit: float = 0.0,
        warn_days: int = 30,
        reference_date: date | None = None,
        export_path: Path | None = None,
        imported_by: str | None = None,
        duplicate_policy: str = "error",
        overwrite_report_period: bool | None = None,
        duplicate_decision: Ln01DuplicateDecision | str | None = None,
        decision_provider: Ln01DecisionProvider | None = None,
        progress: ProgressCallback | None = None,
    ) -> ImportResult:
        prepared = self.prepare_import(
            file_path,
            period=period,
            min_limit=min_limit,
            warn_days=warn_days,
            reference_date=reference_date,
            duplicate_policy=duplicate_policy,
            overwrite_report_period=overwrite_report_period,
            progress=progress,
        )
        decision = self._choose_decision(
            prepared.context,
            duplicate_decision=duplicate_decision,
            decision_provider=decision_provider,
        )
        return self.execute_prepared_import(
            prepared,
            duplicate_decision=decision,
            export_path=export_path,
            imported_by=imported_by,
            progress=progress,
        )

    def prepare_import(
        self,
        file_path: Path,
        *,
        period: str | None = None,
        min_limit: float = 0.0,
        warn_days: int = 30,
        reference_date: date | None = None,
        duplicate_policy: str = "error",
        overwrite_report_period: bool | None = None,
        progress: ProgressCallback | None = None,
    ) -> Ln01PreparedImport:
        file_path = Path(file_path)
        if file_path.suffix.casefold() != ".csv":
            raise SummaryError("Hạn mức tín dụng chỉ hỗ trợ file CSV LN01.")
        reference_date = reference_date or date.today()
        report_period = _credit_report_period(period, file_path, reference_date)
        if progress:
            progress(f"Đang đọc file LN01: {file_path.name}")
        raw_data = file_path.read_bytes()
        source_hash = hashlib.sha256(raw_data).hexdigest()
        has_group_code_column = ln01_has_group_code_header(raw_data)
        normalized_rows, source_row_count = parse_ln01_bytes(file_path, raw_data, period=report_period)
        agreements = tuple(project_credit_limit_rows(normalized_rows))
        branch_code = _ln01_branch_code(agreements, file_path)
        if progress:
            progress("Đang lọc hợp đồng hạn mức cảnh báo")
        accepted_rows = tuple(
            filter_credit_limit_rows(
                agreements,
                min_limit=float(min_limit),
                warn_days=int(warn_days),
                reference_date=reference_date,
            )
        )
        hmhethan_status = self.find_hmhethan_batch_by_sha(
            source_hash,
            period=report_period,
            branch_code=branch_code,
        )
        credit_status = self.get_credit_period_status(report_period, source_hash)
        resolution = self.resolve_ln01_import_decision(
            hmhethan_status,
            credit_status,
            duplicate_policy=duplicate_policy,
            overwrite_report_period=overwrite_report_period,
        )
        context = Ln01ImportContext(
            period=report_period,
            branch_code=branch_code,
            source_file_name=file_path.name,
            source_sha256=source_hash,
            source_file_size=len(raw_data),
            source_row_count=source_row_count,
            hmhethan_status=hmhethan_status,
            credit_status=credit_status,
            resolution=resolution,
        )
        return Ln01PreparedImport(
            file_path=file_path,
            period=report_period,
            branch_code=branch_code,
            source_sha256=source_hash,
            source_file_size=len(raw_data),
            source_row_count=source_row_count,
            normalized_rows=tuple(normalized_rows),
            agreements=agreements,
            accepted_rows=accepted_rows,
            reference_date=reference_date,
            min_limit=float(min_limit),
            warn_days=int(warn_days),
            has_group_code_column=has_group_code_column,
            context=context,
        )

    def execute_prepared_import(
        self,
        prepared: Ln01PreparedImport,
        *,
        duplicate_decision: Ln01DuplicateDecision | str | None = None,
        export_path: Path | None = None,
        imported_by: str | None = None,
        progress: ProgressCallback | None = None,
    ) -> ImportResult:
        decision = self._normalize_decision(duplicate_decision) or prepared.context.resolution.default_decision
        self._validate_decision(prepared.context, decision)
        if decision == Ln01DuplicateDecision.CANCEL:
            return ImportResult(batch_id="", row_count=0, message="Đã hủy import LN01.")

        user = imported_by or _current_user()
        pending_batch: PendingCreditLimitBatch | None = None
        metadata = prepared.context.hmhethan_status.metadata
        backup_file: Path | None = None
        target_existed = False
        batch_finalized = False
        stage = "resolve_ln01_import"
        report_result: dict[str, object] = {
            "normalized_loan_count": 0,
            "customer_count": 0,
            "ln01_total_balance": 0.0,
        }
        started = perf_counter()
        try:
            if self._decision_creates_hmhethan(decision):
                stage = "prepare_excel_batch"
                pending_batch = self._prepare_hmhethan_batch(prepared, decision, imported_by=user, progress=progress)
                target_existed = pending_batch.target_file.exists()
                if target_existed:
                    backup_file = self._backup_batch_target(pending_batch.target_file)

            if progress and self._decision_writes_credit(decision):
                progress(f"Đang ghi dữ liệu kỳ {prepared.period} vào Credit.db")
            stage = "write_credit_db"
            with closing(self.credit_repository.connect()) as credit_db:
                credit_db.execute("BEGIN IMMEDIATE")
                try:
                    if self._decision_writes_credit(decision):
                        report_result = self.credit_repository.write_ln01_period(
                            credit_db,
                            period=prepared.period,
                            source_file_name=prepared.file_path.name,
                            source_sha256=prepared.source_sha256,
                            source_file_size=prepared.source_file_size,
                            source_row_count=prepared.source_row_count,
                            rows=list(prepared.normalized_rows),
                            imported_by=user,
                            overwrite_period=self._decision_overwrites_credit(prepared.context, decision),
                            duration_ms=int((perf_counter() - started) * 1000),
                            has_group_code_column=prepared.has_group_code_column,
                            message=self._credit_history_message(prepared.context, decision),
                        )
                    if pending_batch is not None:
                        stage = "finalize_excel_batch"
                        metadata = self.repository.credit_limit_store.finalize_prepared_batch(pending_batch)
                        batch_finalized = True
                    if metadata is None:
                        raise SummaryError("Không tìm thấy batch HMHETHAN để ghi nhận import LN01.")
                    self.credit_repository.log_ln01_import_action(
                        credit_db,
                        action=self._action_name_for_decision(decision),
                        period=prepared.period,
                        source_sha256=prepared.source_sha256,
                        source_file_name=prepared.file_path.name,
                        batch_id=str(metadata.batch_id),
                        created_by=user,
                    )
                    credit_db.commit()
                except Exception:
                    credit_db.rollback()
                    self._rollback_batch_file(
                        pending_batch=pending_batch,
                        backup_file=backup_file,
                        target_existed=target_existed,
                        batch_finalized=batch_finalized,
                    )
                    raise
            if backup_file is not None:
                self._archive_batch_backup(backup_file, pending_batch.target_file if pending_batch else prepared.file_path)
            stage = "export_report"
            output_path = (
                export_credit_limit_vba(list(prepared.accepted_rows), Path(export_path))
                if export_path is not None and prepared.accepted_rows
                else None
            )
            if progress:
                progress(self._progress_done_message(prepared, decision))
            return ImportResult(
                batch_id=str(metadata.batch_id) if metadata is not None else "",
                row_count=len(prepared.accepted_rows),
                message=self._success_message(prepared, decision, metadata, report_result),
                output_path=output_path,
            )
        except SummaryError as exc:
            if pending_batch is not None and not batch_finalized:
                self.repository.credit_limit_store.cleanup_prepared_batch(pending_batch)
            if backup_file is not None and backup_file.exists() and not batch_finalized:
                backup_file.unlink(missing_ok=True)
            _log_credit_limit_import_failure(
                repository=self.repository,
                source_file=prepared.file_path,
                stage=stage,
                parsed_rows=len(prepared.agreements),
                first_row_type=type(prepared.agreements[0]).__name__ if prepared.agreements else "",
                batch_id=str(getattr(getattr(pending_batch, "metadata", None), "batch_id", "")),
                exc=exc,
            )
            raise
        except Exception as exc:
            if pending_batch is not None and not batch_finalized:
                self.repository.credit_limit_store.cleanup_prepared_batch(pending_batch)
            if backup_file is not None and backup_file.exists() and not batch_finalized:
                backup_file.unlink(missing_ok=True)
            _log_credit_limit_import_failure(
                repository=self.repository,
                source_file=prepared.file_path,
                stage=stage,
                parsed_rows=len(prepared.agreements),
                first_row_type=type(prepared.agreements[0]).__name__ if prepared.agreements else "",
                batch_id=str(getattr(getattr(pending_batch, "metadata", None), "batch_id", "")),
                exc=exc,
            )
            raise SummaryError("Không thể tạo batch Hạn mức tín dụng từ file LN01.") from exc

    def find_hmhethan_batch_by_sha(
        self,
        source_sha256: str,
        *,
        period: str = "",
        branch_code: str = "",
    ) -> CreditLimitBatchLookup:
        return self.repository.credit_limit_store.find_batch_by_sha_status(
            source_sha256,
            period=period,
            branch_code=branch_code,
        )

    def get_credit_period_status(self, period: str, source_sha256: str) -> CreditPeriodStatus:
        clean_hash = str(source_sha256 or "").strip().lower()
        metadata = self.credit_repository.ln01_period_import_metadata(period)
        if not metadata:
            return CreditPeriodStatus(period=period, exists=False, same_sha=False, different_sha=False)
        existing_hash = str(metadata.get("source_sha256") or "").strip().lower()
        same_sha = bool(clean_hash and existing_hash == clean_hash)
        return CreditPeriodStatus(
            period=period,
            exists=True,
            same_sha=same_sha,
            different_sha=bool(clean_hash and existing_hash and existing_hash != clean_hash),
            source_file_name=str(metadata.get("source_file_name") or ""),
            source_sha256=existing_hash,
            imported_at=str(metadata.get("created_at") or ""),
            imported_by=str(metadata.get("imported_by") or ""),
        )

    def resolve_ln01_import_decision(
        self,
        hmhethan_status: CreditLimitBatchLookup,
        credit_status: CreditPeriodStatus,
        *,
        duplicate_policy: str = "error",
        overwrite_report_period: bool | None = None,
    ) -> Ln01ImportDecisionResolution:
        overwrite_requested = (
            bool(duplicate_policy == "new")
            if overwrite_report_period is None
            else bool(overwrite_report_period)
        )
        hm_valid = hmhethan_status.state == CreditLimitBatchLookupState.FOUND_VALID
        if not credit_status.exists:
            if hm_valid:
                return Ln01ImportDecisionResolution(
                    default_decision=Ln01DuplicateDecision.CREATE_CREDIT_ONLY,
                    allowed_decisions=(
                        Ln01DuplicateDecision.CREATE_CREDIT_ONLY,
                        Ln01DuplicateDecision.CANCEL,
                    ),
                    requires_confirmation=True,
                    reason="HMHETHAN_EXISTS_CREDIT_MISSING",
                )
            return Ln01ImportDecisionResolution(
                default_decision=Ln01DuplicateDecision.CREATE_BOTH,
                allowed_decisions=(Ln01DuplicateDecision.CREATE_BOTH, Ln01DuplicateDecision.CANCEL),
                requires_confirmation=hmhethan_status.state == CreditLimitBatchLookupState.FOUND_INVALID,
                reason="NO_CREDIT_PERIOD",
            )
        if credit_status.different_sha:
            hm_hash = ""
            if hmhethan_status.metadata is not None:
                hm_hash = str(hmhethan_status.metadata.source_file_sha256 or "").strip().lower()
            credit_hash = str(credit_status.source_sha256 or "").strip().lower()
            default = (
                Ln01DuplicateDecision.OVERWRITE_BOTH
                if hm_valid and hm_hash == credit_hash
                else Ln01DuplicateDecision.OVERWRITE_CREDIT
                if hm_valid
                else Ln01DuplicateDecision.CREATE_BOTH
            )
            allowed: tuple[Ln01DuplicateDecision, ...]
            if hm_valid:
                allowed = (
                    Ln01DuplicateDecision.OVERWRITE_BOTH,
                    Ln01DuplicateDecision.OVERWRITE_CREDIT,
                    Ln01DuplicateDecision.CREATE_BOTH,
                    Ln01DuplicateDecision.CANCEL,
                )
            else:
                allowed = (Ln01DuplicateDecision.CREATE_BOTH, Ln01DuplicateDecision.CANCEL)
            return Ln01ImportDecisionResolution(
                default_decision=default,
                allowed_decisions=allowed,
                requires_confirmation=not overwrite_requested,
                reason="CREDIT_PERIOD_DIFFERENT_SHA",
            )
        if hm_valid:
            return Ln01ImportDecisionResolution(
                default_decision=Ln01DuplicateDecision.OVERWRITE_CREDIT,
                allowed_decisions=(
                    Ln01DuplicateDecision.OVERWRITE_CREDIT,
                    Ln01DuplicateDecision.OVERWRITE_BOTH,
                    Ln01DuplicateDecision.CANCEL,
                ),
                requires_confirmation=not overwrite_requested,
                reason="BOTH_EXIST_SAME_SHA",
            )
        default = Ln01DuplicateDecision.CREATE_BOTH if overwrite_requested else Ln01DuplicateDecision.CREATE_HMHE_THAN_ONLY
        return Ln01ImportDecisionResolution(
            default_decision=default,
            allowed_decisions=(
                Ln01DuplicateDecision.CREATE_BOTH,
                Ln01DuplicateDecision.CREATE_HMHE_THAN_ONLY,
                Ln01DuplicateDecision.CANCEL,
            ),
            requires_confirmation=not overwrite_requested,
            reason="CREDIT_EXISTS_HMHEthan_MISSING",
        )

    def _choose_decision(
        self,
        context: Ln01ImportContext,
        *,
        duplicate_decision: Ln01DuplicateDecision | str | None,
        decision_provider: Ln01DecisionProvider | None,
    ) -> Ln01DuplicateDecision:
        decision = self._normalize_decision(duplicate_decision)
        if decision is None and decision_provider is not None and context.resolution.requires_confirmation:
            decision = self._normalize_decision(decision_provider(context)) or Ln01DuplicateDecision.CANCEL
        if decision is None:
            if context.resolution.requires_confirmation and self._decision_requires_explicit_confirmation(context):
                raise SummaryError(self._confirmation_required_message(context))
            decision = context.resolution.default_decision
        self._validate_decision(context, decision)
        return decision

    def _normalize_decision(
        self,
        decision: Ln01DuplicateDecision | str | None,
    ) -> Ln01DuplicateDecision | None:
        if decision is None:
            return None
        if isinstance(decision, Ln01DuplicateDecision):
            return decision
        text = str(decision or "").strip()
        if not text:
            return None
        if text in Ln01DuplicateDecision.__members__:
            return Ln01DuplicateDecision[text]
        try:
            return Ln01DuplicateDecision(text)
        except ValueError as exc:
            raise SummaryError(f"Lựa chọn import LN01 không hợp lệ: {text}") from exc

    def _validate_decision(self, context: Ln01ImportContext, decision: Ln01DuplicateDecision) -> None:
        if decision == Ln01DuplicateDecision.CANCEL:
            return
        if decision not in context.resolution.allowed_decisions:
            raise SummaryError("Lựa chọn import LN01 không hợp lệ cho trạng thái hiện tại.")
        hm_valid = context.hmhethan_status.state == CreditLimitBatchLookupState.FOUND_VALID
        if decision in {Ln01DuplicateDecision.CREATE_CREDIT_ONLY, Ln01DuplicateDecision.OVERWRITE_CREDIT} and not hm_valid:
            raise SummaryError("Không có batch HMHETHAN hợp lệ để tái sử dụng.")
        if decision == Ln01DuplicateDecision.OVERWRITE_BOTH and context.hmhethan_status.metadata is None:
            raise SummaryError("Không có batch HMHETHAN hợp lệ để ghi đè.")

    def _decision_requires_explicit_confirmation(self, context: Ln01ImportContext) -> bool:
        decision = context.resolution.default_decision
        if decision in {Ln01DuplicateDecision.OVERWRITE_CREDIT, Ln01DuplicateDecision.OVERWRITE_BOTH}:
            return True
        return decision == Ln01DuplicateDecision.CREATE_BOTH and context.credit_status.exists

    def _confirmation_required_message(self, context: Ln01ImportContext) -> str:
        if context.credit_status.different_sha:
            return (
                f"Kỳ {context.period} đã được tạo từ file LN01 khác. "
                "Cần xác nhận trước khi ghi đè dữ liệu báo cáo."
            )
        return f"Kỳ {context.period} đã có dữ liệu báo cáo. Cần xác nhận phạm vi ghi đè."

    def _prepare_hmhethan_batch(
        self,
        prepared: Ln01PreparedImport,
        decision: Ln01DuplicateDecision,
        *,
        imported_by: str,
        progress: ProgressCallback | None,
    ) -> PendingCreditLimitBatch:
        replace_metadata = (
            prepared.context.hmhethan_status.metadata
            if decision == Ln01DuplicateDecision.OVERWRITE_BOTH
            or (decision == Ln01DuplicateDecision.CREATE_BOTH and prepared.context.hmhethan_status.metadata is not None)
            else None
        )
        duplicate_policy = "replace" if replace_metadata is not None else "new"
        return self.repository.credit_limit_store.prepare_batch(
            source_path=prepared.file_path,
            rows=prepared.agreements,
            accepted_rows=prepared.accepted_rows,
            reference_date=prepared.reference_date,
            warn_days=prepared.warn_days,
            min_limit=prepared.min_limit,
            source_file_sha256=prepared.source_sha256,
            source_file_size=prepared.source_file_size,
            imported_by=imported_by,
            source_row_count=prepared.source_row_count,
            duplicate_policy=duplicate_policy,
            replace_metadata=replace_metadata,
            period=prepared.period,
            branch_code=prepared.branch_code,
            progress=progress,
        )

    def _decision_creates_hmhethan(self, decision: Ln01DuplicateDecision) -> bool:
        return decision in {
            Ln01DuplicateDecision.CREATE_BOTH,
            Ln01DuplicateDecision.CREATE_HMHE_THAN_ONLY,
            Ln01DuplicateDecision.OVERWRITE_BOTH,
        }

    def _decision_writes_credit(self, decision: Ln01DuplicateDecision) -> bool:
        return decision in {
            Ln01DuplicateDecision.CREATE_BOTH,
            Ln01DuplicateDecision.CREATE_CREDIT_ONLY,
            Ln01DuplicateDecision.OVERWRITE_CREDIT,
            Ln01DuplicateDecision.OVERWRITE_BOTH,
        }

    def _decision_overwrites_credit(self, context: Ln01ImportContext, decision: Ln01DuplicateDecision) -> bool:
        return decision in {Ln01DuplicateDecision.OVERWRITE_CREDIT, Ln01DuplicateDecision.OVERWRITE_BOTH} or (
            decision == Ln01DuplicateDecision.CREATE_BOTH and context.credit_status.exists
        )

    def _backup_batch_target(self, target_file: Path) -> Path:
        target_file = Path(target_file)
        backup_file = target_file.with_name(f".{target_file.name}.{os.getpid()}.bak")
        counter = 1
        while backup_file.exists():
            counter += 1
            backup_file = target_file.with_name(f".{target_file.name}.{os.getpid()}.{counter}.bak")
        shutil.copy2(target_file, backup_file)
        return backup_file

    def _archive_batch_backup(self, backup_file: Path, target_file: Path) -> Path:
        history_path = self.repository.credit_limit_store.history_path
        history_path.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = Path(target_file)
        archive = history_path / f"{target.stem}_replaced_{timestamp}{target.suffix}"
        counter = 1
        while archive.exists():
            counter += 1
            archive = history_path / f"{target.stem}_replaced_{timestamp}_{counter}{target.suffix}"
        os.replace(backup_file, archive)
        return archive

    def _rollback_batch_file(
        self,
        *,
        pending_batch: PendingCreditLimitBatch | None,
        backup_file: Path | None,
        target_existed: bool,
        batch_finalized: bool,
    ) -> None:
        if pending_batch is None:
            return
        if not batch_finalized:
            self.repository.credit_limit_store.cleanup_prepared_batch(pending_batch)
            if backup_file is not None:
                backup_file.unlink(missing_ok=True)
            return
        if target_existed and backup_file is not None and backup_file.is_file():
            os.replace(backup_file, pending_batch.target_file)
        else:
            pending_batch.target_file.unlink(missing_ok=True)
        self.repository.credit_limit_store.invalidate_cache()

    def _action_name_for_decision(self, decision: Ln01DuplicateDecision) -> str:
        return {
            Ln01DuplicateDecision.CREATE_BOTH: "CREATE_BOTH",
            Ln01DuplicateDecision.CREATE_CREDIT_ONLY: "CREATE_CREDIT_FROM_EXISTING_HMHEthan",
            Ln01DuplicateDecision.CREATE_HMHE_THAN_ONLY: "CREATE_HMHEthan_ONLY",
            Ln01DuplicateDecision.OVERWRITE_CREDIT: "OVERWRITE_CREDIT_KEEP_HMHEthan",
            Ln01DuplicateDecision.OVERWRITE_BOTH: "OVERWRITE_BOTH",
        }[decision]

    def _credit_history_message(self, context: Ln01ImportContext, decision: Ln01DuplicateDecision) -> str:
        action = self._action_name_for_decision(decision)
        reused = decision in {Ln01DuplicateDecision.CREATE_CREDIT_ONLY, Ln01DuplicateDecision.OVERWRITE_CREDIT}
        suffix = "; hmhethan_batch_reused=1" if reused else ""
        return f"Import LN01 {context.source_file_name}; action={action}{suffix}"

    def _progress_done_message(self, prepared: Ln01PreparedImport, decision: Ln01DuplicateDecision) -> str:
        if decision == Ln01DuplicateDecision.CREATE_HMHE_THAN_ONLY:
            return f"Đã tạo batch HMHETHAN từ {prepared.file_path.name}"
        return f"Đã lưu kỳ {prepared.period} trong Credit.db"

    def _success_message(
        self,
        prepared: Ln01PreparedImport,
        decision: Ln01DuplicateDecision,
        metadata: CreditLimitBatchMetadata | None,
        report_result: dict[str, object],
    ) -> str:
        batch_file = metadata.file_name if metadata is not None else ""
        group_warning = self._group_code_warning(prepared)
        if decision == Ln01DuplicateDecision.CREATE_CREDIT_ONLY:
            return (
                f"Đã tạo lại dữ liệu báo cáo kỳ {prepared.period}.\n\n"
                f"Batch Hạn mức đã có được giữ nguyên: {batch_file}"
                f"{group_warning}"
            )
        if decision == Ln01DuplicateDecision.OVERWRITE_CREDIT:
            return (
                f"Đã ghi đè dữ liệu báo cáo kỳ {prepared.period}.\n\n"
                "Batch Hạn mức hiện có không bị thay đổi."
                f"{group_warning}"
            )
        if decision == Ln01DuplicateDecision.OVERWRITE_BOTH:
            return f"Đã ghi đè dữ liệu báo cáo và batch Hạn mức kỳ {prepared.period}.{group_warning}"
        if decision == Ln01DuplicateDecision.CREATE_HMHE_THAN_ONLY:
            return f"Đã tạo batch Hạn mức {batch_file}. Dữ liệu Credit.db không thay đổi.{group_warning}"
        return (
            f"Đã lưu {len(prepared.accepted_rows):,} hợp đồng hạn mức. "
            f"Credit.db kỳ {prepared.period}: {int(report_result['normalized_loan_count']):,} món, "
            f"{int(report_result['customer_count']):,} khách hàng, "
            f"tổng dư nợ LN01 {float(report_result['ln01_total_balance']):,.0f}."
            f"{group_warning}"
        )

    @staticmethod
    def _group_code_warning(prepared: Ln01PreparedImport) -> str:
        if prepared.has_group_code_column:
            return ""
        return (
            "\n\nLưu ý: File LN01 không có header GRPNO/MaToVayVon/MaTo. "
            "Các báo cáo khác vẫn được import, nhưng kỳ này chưa có dữ liệu mã tổ vay vốn."
        )


def _credit_report_period(period: str | None, file_path: Path, reference_date: date) -> str:
    clean_period = str(period or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}", clean_period):
        return clean_period
    parsed = parse_period_from_filename(file_path.name)
    if re.fullmatch(r"\d{4}-\d{2}", parsed):
        return parsed
    raise SummaryError(
        "Không xác định được kỳ dữ liệu LN01 từ tên file. "
        "Vui lòng nhập kỳ theo dạng YYYY-MM."
    )


def _ln01_branch_code(rows: Sequence[CreditLimitRow], file_path: Path) -> str:
    for row in rows:
        branch_code = str(row.branch_code or "").strip()
        if branch_code:
            return branch_code
    match = re.match(r"^\D*(\d{4})(?:\D|$)", Path(file_path).stem)
    return match.group(1) if match else ""


def _legacy_import_credit_limit_file(
    repository: SummaryRepository,
    file_path: Path,
    *,
    min_limit: float = 0.0,
    warn_days: int = 30,
    reference_date: date | None = None,
    export_path: Path | None = None,
    imported_by: str | None = None,
    duplicate_policy: str = "error",
    progress: ProgressCallback | None = None,
) -> ImportResult:
    file_path = Path(file_path)
    if file_path.suffix.casefold() != ".csv":
        raise SummaryError("Hạn mức tín dụng chỉ hỗ trợ file CSV LN01.")
    reference_date = reference_date or date.today()
    stage = "read_ln01"
    agreements: list[CreditLimitRow] = []
    rows: list[CreditLimitRow] = []
    batch_id = ""
    try:
        if progress:
            progress(f"Đang đọc file LN01: {file_path.name}")
        started = perf_counter()
        raw_data = file_path.read_bytes()
        source_hash = hashlib.sha256(raw_data).hexdigest()
        agreements, source_row_count = _load_credit_limit_agreements_from_bytes(file_path, raw_data)
        stage = "filter_rows"
        rows = filter_credit_limit_rows(
            agreements,
            min_limit=float(min_limit),
            warn_days=int(warn_days),
            reference_date=reference_date,
        )
        _ = int((perf_counter() - started) * 1000)
        stage = "write_excel_batch"
        metadata = repository.save_credit_limit_excel_batch(
            source_path=file_path,
            rows=agreements,
            accepted_rows=rows,
            reference_date=reference_date.isoformat(),
            warn_days=warn_days,
            min_limit=min_limit,
            source_file_sha256=source_hash,
            source_file_size=len(raw_data),
            imported_by=imported_by or _current_user(),
            source_row_count=source_row_count,
            duplicate_policy=duplicate_policy,
            progress=progress,
        )
        batch_id = str(metadata.batch_id)
        stage = "export_report"
        output_path = export_credit_limit_vba(rows, Path(export_path)) if export_path is not None and rows else None
        if progress:
            progress(f"Đã lưu {len(rows):,} hợp đồng cảnh báo")
        return ImportResult(
            batch_id=metadata.batch_id,
            row_count=len(rows),
            message=f"Đã lưu {len(rows):,} hợp đồng hạn mức.",
            output_path=output_path,
        )
    except SummaryError as exc:
        _log_credit_limit_import_failure(
            repository=repository,
            source_file=file_path,
            stage=stage,
            parsed_rows=len(agreements),
            first_row_type=type(agreements[0]).__name__ if agreements else "",
            batch_id=batch_id,
            exc=exc,
        )
        raise
    except Exception as exc:
        _log_credit_limit_import_failure(
            repository=repository,
            source_file=file_path,
            stage=stage,
            parsed_rows=len(agreements),
            first_row_type=type(agreements[0]).__name__ if agreements else "",
            batch_id=batch_id,
            exc=exc,
        )
        raise SummaryError("Không thể tạo batch Hạn mức tín dụng từ file LN01.") from exc


def filter_credit_limit_rows(
    agreements: Iterable[CreditLimitRow],
    *,
    min_limit: float,
    warn_days: int,
    reference_date: date,
) -> list[CreditLimitRow]:
    selected: list[CreditLimitRow] = []
    for row in agreements:
        if min_limit != 0 and not (row.approved_amount > min_limit):
            continue
        if row.expiry_date is None:
            continue
        days = (row.expiry_date - reference_date).days
        if days < 0:
            selected.append(
                replace(
                    row,
                    days_to_expiry=days,
                    status="Đã hết hạn",
                    note=CREDIT_LIMIT_EXPIRED_NOTE,
                )
            )
        elif 0 <= days <= warn_days:
            selected.append(
                replace(
                    row,
                    days_to_expiry=days,
                    status="Sắp hết hạn",
                    note=CREDIT_LIMIT_SOON_NOTE_TEMPLATE.format(warn_days=warn_days),
                )
            )
    return selected


def _parse_nim_file(
    file_path: Path,
    config: NimConfig,
    *,
    credit_card_rate: float,
    unit_directory: UnitDirectoryService | None = None,
) -> ParsedNimFile:
    unit_directory = unit_directory or get_unit_directory_service()
    data = Path(file_path).read_bytes()
    source_hash = hashlib.sha256(data).hexdigest()
    text = data.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    try:
        headers = next(reader)
    except StopIteration as exc:
        raise SummaryError(f"File {file_path.name} không có dữ liệu.") from exc
    header_map = {clean_header(header): index for index, header in enumerate(headers)}
    required = set(NIM_COMMON_HEADERS)
    missing = [header for header in required if header not in header_map]
    if missing:
        raise SummaryError(f"File {file_path.name} thiếu cột: {', '.join(missing)}")
    period = parse_period_from_filename(file_path.name)
    rows: list[NimRow] = []
    customer_rows: list[NormalizedLoanRow] = []
    collect_customer = config.data_type == SummaryDataType.NIM_DN
    customer_missing_headers: tuple[str, ...] = ()
    if collect_customer:
        customer_missing_headers = tuple(
            header for header in CUSTOMER_FTPLN_REQUIRED_HEADERS if header not in header_map
        )
    customer_source_total_balance = 0.0
    customer_invalid_row_count = 0
    customer_warning_count = 0
    customer_warnings: list[str] = []
    debt_group_header_present = "AQCCDFIN" in header_map
    debt_group_counts = {"01": 0, "02": 0, "03": 0, "04": 0, "05": 0, "UNKNOWN": 0}
    debt_group_invalid_samples: list[str] = []
    unit_warnings: list[str] = []
    checked_units: set[tuple[str, str]] = set()
    missing_trctcd = "TRCTCD" not in header_map
    if collect_customer and missing_trctcd:
        customer_warning_count += 1
        customer_warnings.append(
            f"{file_path.name} thieu cot TRCTCD; ghi chi tiet don vi UNKNOWN, khong tu gan Hoi so."
        )
    for source_row_number, raw in enumerate(reader, start=2):
        if not raw or not any(str(item).strip() for item in raw):
            continue
        value = lambda header: _field(raw, header_map[header])
        optional_value = lambda header: _field(raw, header_map[header]) if header in header_map else ""
        trref = optional_value("TRREF")
        intrt_text = value("INTRT")
        is_glx = trref.upper() == "GLX" or intrt_text == ""
        intrt = float(credit_card_rate) if is_glx else safe_rate(intrt_text)
        ftp = safe_rate(value("FTP"))
        adjustment = safe_rate(value("MUCFTPDC"))
        balance = safe_rate(value("LDRBAL"))
        row_has_numeric_warning = False
        if collect_customer and not customer_missing_headers:
            customer_source_total_balance += balance
            for header in ("LDRBAL", "FTP", "INTRT", "MUCFTPDC"):
                text_value = value(header)
                if text_value and not _has_vba_numeric_prefix(str(text_value).replace(",", ".")):
                    row_has_numeric_warning = True
                    customer_warning_count += 1
                    if len(customer_warnings) < 50:
                        customer_warnings.append(
                            f"{file_path.name}:{source_row_number} gia tri so khong hop le o cot {header}: {text_value}"
                        )
            if row_has_numeric_warning:
                customer_invalid_row_count += 1
        if config.data_type == SummaryDataType.NIM_DN:
            numerator_after = (intrt - ftp - adjustment) * balance
            numerator_before = (intrt - ftp) * balance
            average_rate_numerator = intrt * balance
        else:
            numerator_after = (ftp - intrt + adjustment) * balance
            numerator_before = (ftp - intrt) * balance
            average_rate_numerator = 0.0
        branch_code = value("BRCD")
        trctcd = normalize_trctcd(optional_value("TRCTCD"))
        unit_key = (branch_code, trctcd)
        if unit_key not in checked_units:
            checked_units.add(unit_key)
            unit_warnings.extend(
                unit_directory.ensure_known_unit(
                    branch_code,
                    trctcd,
                    updated_by="import_nim",
                )
            )
        office_code = build_office_code(branch_code, trctcd)
        office_type = classify_office_type(trctcd).value
        transaction_office = unit_directory.get_office_name(branch_code, trctcd)
        officer = optional_value(config.officer_header)
        if config.data_type == SummaryDataType.NIM_DN:
            if not officer or officer == "Không xác định":
                officer = "Thẻ tín dụng"
        elif not officer or officer == "Không xác định":
            officer = "Không gắn mã CB"
        rows.append(
            NimRow(
                batch_id=0,
                data_type=config.data_type,
                period=period,
                branch_code=branch_code,
                branch_name=unit_directory.get_branch_display_name(branch_code),
                trctcd=trctcd,
                transaction_office=transaction_office,
                customer_type=map_customer_type(optional_value("CUSTTP")),
                officer=officer,
                balance=balance,
                interest_rate=intrt,
                ftp_rate=ftp,
                adjustment_rate=adjustment,
                numerator_before=numerator_before,
                numerator_after=numerator_after,
                average_rate_numerator=average_rate_numerator,
                source_file=file_path.name,
            )
        )
        if collect_customer and not customer_missing_headers:
            branch_code = value("BRCD")
            customer_sequence = normalize_customer_sequence(value("CUSTSEQ"))
            customer_code = build_customer_code(branch_code, customer_sequence) if branch_code and customer_sequence else ""
            if not branch_code or not customer_sequence or not customer_code:
                customer_invalid_row_count += 1
                customer_warning_count += 1
                if len(customer_warnings) < 50:
                    customer_warnings.append(
                        f"{file_path.name}:{source_row_number} thieu BRCD/CUSTSEQ, khong tao customer_code."
                    )
                continue
            officer_identity = split_customer_officer(officer)
            raw_debt_group = optional_value("AQCCDFIN") if debt_group_header_present else ""
            debt_group_code, debt_group_number, debt_group_category, has_valid_debt_group = normalize_debt_group(raw_debt_group)
            if has_valid_debt_group:
                debt_group_counts[debt_group_code] += 1
            else:
                debt_group_counts["UNKNOWN"] += 1
                sample = str(raw_debt_group or "").strip()
                if debt_group_header_present and sample and sample not in debt_group_invalid_samples and len(debt_group_invalid_samples) < 20:
                    debt_group_invalid_samples.append(sample)
            customer_rows.append(
                NormalizedLoanRow(
                    period=period,
                    source_file=file_path.name,
                    source_row_number=source_row_number,
                    branch_code=branch_code,
                    trctcd=trctcd,
                    transaction_office=transaction_office,
                    customer_sequence=customer_sequence,
                    customer_code=customer_code,
                    customer_name=clean_cell(value("CUSTNM")),
                    customer_type=map_customer_type_code(value("CUSTTP")).value,
                    ftp_code=clean_cell(value("FTPCD")).upper(),
                    balance=balance,
                    ftp=ftp,
                    interest_rate=intrt,
                    ftp_adjustment=adjustment,
                    officer_code=officer_identity.officer_code,
                    officer_name=officer_identity.officer_name,
                    office_code=office_code,
                    office_name=transaction_office,
                    office_type=office_type,
                    debt_group_code=debt_group_code,
                    debt_group_number=debt_group_number,
                    debt_group_category=debt_group_category,
                    has_valid_debt_group=has_valid_debt_group,
                )
            )
    return ParsedNimFile(
        rows=rows,
        period=period,
        source_hash=source_hash,
        customer_rows=customer_rows,
        customer_missing_headers=customer_missing_headers,
        customer_source_total_balance=customer_source_total_balance,
        customer_invalid_row_count=customer_invalid_row_count,
        customer_warning_count=customer_warning_count,
        customer_warnings=tuple(customer_warnings),
        unit_warnings=tuple(unit_warnings),
        debt_group_header_present=debt_group_header_present,
        debt_group_valid_row_count=sum(debt_group_counts[code] for code in ("01", "02", "03", "04", "05")),
        debt_group_1_row_count=debt_group_counts["01"],
        debt_group_2_row_count=debt_group_counts["02"],
        debt_group_3_row_count=debt_group_counts["03"],
        debt_group_4_row_count=debt_group_counts["04"],
        debt_group_5_row_count=debt_group_counts["05"],
        debt_group_unknown_row_count=debt_group_counts["UNKNOWN"],
        debt_group_invalid_samples=tuple(debt_group_invalid_samples),
    )


def _load_loan_file(file_path: Path) -> tuple[dict[str, float], dict[str, tuple[str, str, str]]]:
    suffix = file_path.suffix.casefold()
    if suffix in {".csv", ".txt"}:
        rows = _read_vba_simple_split_rows(file_path)
        use_vba_val = True
    elif suffix == ".xlsx":
        rows = _read_excel_rows(file_path)
        use_vba_val = False
    else:
        raise SummaryError(f"Không hỗ trợ định dạng file so sánh tăng giảm khách hàng: {file_path.name}")
    if not rows:
        return {}, {}
    headers = [clean_header(item) for item in rows[0]]
    try:
        col_customer = headers.index("CUSTSEQ")
        col_balance = headers.index("DU_NO")
    except ValueError as exc:
        raise SummaryError(f"File {file_path.name} thiếu cột CUSTSEQ hoặc DU_NO.") from exc
    col_name = _optional_index(headers, "CUSTNM")
    col_address = _optional_index(headers, "ADDR1")
    col_officer = _optional_index(headers, "OFFICER_NAME")
    balances: dict[str, float] = {}
    info: dict[str, tuple[str, str, str]] = {}
    for raw in rows[1:]:
        if len(raw) <= max(col_customer, col_balance):
            continue
        customer_code = clean_cell(raw[col_customer])
        if not customer_code:
            continue
        balance = _vba_loan_amount(raw[col_balance]) if use_vba_val else safe_amount(raw[col_balance])
        balances[customer_code] = balances.get(customer_code, 0.0) + balance
        if customer_code not in info:
            info[customer_code] = (
                clean_cell(raw[col_name]) if col_name is not None and len(raw) > col_name else "",
                clean_cell(raw[col_address]) if col_address is not None and len(raw) > col_address else "",
                clean_cell(raw[col_officer]) if col_officer is not None and len(raw) > col_officer else "",
            )
    return balances, info


def _load_credit_limit_agreements(file_path: Path) -> list[CreditLimitRow]:
    rows = _read_delimited_rows(file_path)
    agreements, _source_row_count = _load_credit_limit_agreements_from_rows(rows)
    return agreements


def _load_credit_limit_agreements_from_bytes(file_path: Path, raw_data: bytes) -> tuple[list[CreditLimitRow], int]:
    try:
        normalized_rows, source_row_count = parse_ln01_bytes(file_path, raw_data)
        return project_credit_limit_rows(normalized_rows), source_row_count
    except SummaryError as exc:
        raise SummaryError(str(exc)) from exc


def _load_credit_limit_agreements_from_rows(rows: list[list[str]]) -> tuple[list[CreditLimitRow], int]:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerows(rows)
    return _load_credit_limit_agreements_from_bytes(Path("LN01.csv"), output.getvalue().encode("utf-8"))


def _read_delimited_rows(file_path: Path) -> list[list[str]]:
    text = file_path.read_text(encoding="utf-8-sig")
    return _read_delimited_rows_from_text(text)


def _read_delimited_rows_from_text(text: str) -> list[list[str]]:
    first_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = ","
    if "\t" in first_line:
        delimiter = "\t"
    elif ";" in first_line:
        delimiter = ";"
    return [row for row in csv.reader(StringIO(text), delimiter=delimiter)]


def _read_vba_simple_split_rows(file_path: Path) -> list[list[str]]:
    text = file_path.read_text(encoding="utf-8-sig")
    text = text.replace("\r\n", "\n")
    lines = text.split("\n")
    first_line = lines[0] if lines else ""
    delimiter = ","
    if "\t" in first_line:
        delimiter = "\t"
    elif ";" in first_line:
        delimiter = ";"
    return [line.split(delimiter) for line in lines]


def _read_excel_rows(file_path: Path) -> list[list[object]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        header_row = None
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = list(row)
            if any(clean_header(value) == "CUSTSEQ" for value in values):
                header_row = row_index
                break
        if header_row is None:
            raise SummaryError(f"File {file_path.name} không có cột CUSTSEQ.")
        rows: list[list[object]] = []
        for row in worksheet.iter_rows(min_row=header_row, values_only=True):
            rows.append(list(row))
        return rows
    finally:
        workbook.close()


def parse_period_from_filename(file_name: str) -> str:
    stem = Path(file_name).stem
    date_part = stem[-8:]
    if len(date_part) == 8 and date_part.isdigit():
        return f"{date_part[:4]}-{date_part[4:6]}"
    return "Không Rõ"


def get_branch_name(brcd: str) -> str:
    return get_unit_directory_service().get_branch_display_name(brcd)


def get_trctcd_name(brcd: str, trctcd: str) -> str:
    return get_unit_directory_service().get_office_name(brcd, trctcd)


def map_customer_type(value: str) -> str:
    clean = str(value).strip()
    if clean == "CN":
        return "Cá nhân (CN)"
    if clean == "TC":
        return "Tổ chức (TC)"
    if clean:
        return "Khách hàng khác"
    return "Không rõ"


def safe_rate(value: object) -> float:
    text = clean_cell(value)
    if not text:
        return 0.0
    text = text.replace(",", ".")
    return _vba_val(text)


def safe_amount(value: object) -> float:
    text = clean_cell(value)
    if not text:
        return 0.0
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value: object) -> date | None:
    text = clean_cell(value)
    if not text:
        return None
    if " " in text:
        text = text.split(" ", 1)[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def clean_header(value: object) -> str:
    text = str(value or "").replace("\ufeff", "").strip().replace('"', "")
    upper = text.upper()
    if upper.endswith("CUSTSEQ"):
        return "CUSTSEQ"
    return upper


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().replace('"', "")


def _split_officer(raw_name: object) -> tuple[str, str]:
    text = str(raw_name or "").strip()
    match = re.match(r"^\s*\[([^\]]+)\]\s*(.*)$", text)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return "", text


def _format_officer(officer_code: str, officer_name: str) -> str:
    code = str(officer_code or "").strip()
    name = str(officer_name or "").strip()
    if code:
        return f"[{code}] {name}".strip()
    return name


def _field(row: list[str], index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    return clean_cell(row[index])


def _optional_index(headers: list[str], header: str) -> int | None:
    try:
        return headers.index(header)
    except ValueError:
        return None


def _add_nim_values(
    target: dict[str, list[float]],
    key: str,
    row: NimRow,
    data_type: SummaryDataType,
) -> None:
    if data_type == SummaryDataType.NIM_DN:
        values = [
            row.numerator_after,
            row.balance,
            row.numerator_before,
            row.average_rate_numerator,
        ]
    else:
        values = [
            row.numerator_after,
            row.balance,
            row.numerator_before,
        ]
    if key not in target:
        target[key] = values
        return
    current = target[key]
    for index, value in enumerate(values):
        current[index] += value


def _nim_detail_output(
    key: str,
    values: list[float],
    data_type: SummaryDataType,
) -> list[object]:
    period, branch_name, transaction_office, customer_type, officer = key.split("|")
    if data_type == SummaryDataType.NIM_DN:
        nim_after, balance, nim_before, average_rate = _nim_rates(values, include_average=True)
        return [
            period,
            branch_name,
            transaction_office,
            customer_type,
            officer,
            average_rate,
            nim_before,
            nim_after,
        ]
    nim_after, _balance, nim_before = _nim_rates(values, include_average=False)
    return [
        period,
        branch_name,
        transaction_office,
        customer_type,
        officer,
        nim_before,
        nim_after,
    ]


def _nim_branch_output(
    key: str,
    values: list[float],
    data_type: SummaryDataType,
) -> list[object]:
    period, branch_name, transaction_office, customer_type = key.split("|")
    if data_type == SummaryDataType.NIM_DN:
        nim_after, balance, nim_before, average_rate = _nim_rates(values, include_average=True)
        _ = balance
        return [
            period,
            branch_name,
            transaction_office,
            customer_type,
            average_rate,
            nim_before,
            nim_after,
        ]
    nim_after, _balance, nim_before = _nim_rates(values, include_average=False)
    return [
        period,
        branch_name,
        transaction_office,
        customer_type,
        nim_before,
        nim_after,
    ]


def _nim_rates(values: list[float], *, include_average: bool) -> tuple[float, ...]:
    numerator_after = values[0]
    balance = values[1]
    numerator_before = values[2]
    nim_after = numerator_after / balance if balance else 0.0
    nim_before = numerator_before / balance if balance else 0.0
    if include_average:
        average_rate_numerator = values[3]
        average_rate = average_rate_numerator / balance if balance else 0.0
        return nim_after, balance, nim_before, average_rate
    return nim_after, balance, nim_before


def _vba_val(value: object) -> float:
    text = clean_cell(value)
    if not text:
        return 0.0
    match = re.match(r"^[\s]*([+-]?(?:\d+(?:\.\d*)?|\.\d+))", text)
    if match is None:
        return 0.0
    try:
        return float(match.group(1))
    except ValueError:
        return 0.0


def _has_vba_numeric_prefix(value: object) -> bool:
    text = clean_cell(value)
    return bool(re.match(r"^[\s]*([+-]?(?:\d+(?:\.\d*)?|\.\d+))", text))


def _vba_loan_amount(value: object) -> float:
    return _vba_val(clean_cell(value).replace(",", ""))


@contextmanager
def _sqlite_snapshots(database_paths: list[Path]) -> Iterable[None]:
    paths = [Path(path) for path in database_paths]
    with tempfile.TemporaryDirectory(prefix="agribank-import-snapshot-") as temporary_directory:
        temporary_root = Path(temporary_directory)
        snapshots: list[tuple[Path, Path | None]] = []
        for index, path in enumerate(paths):
            snapshot_path = temporary_root / f"{index}-{path.name}"
            if path.is_file():
                _checkpoint_sqlite(path)
                with closing(sqlite3.connect(path, timeout=30)) as source:
                    with closing(sqlite3.connect(snapshot_path, timeout=30)) as target:
                        source.backup(target)
                snapshots.append((path, snapshot_path))
            else:
                snapshots.append((path, None))
        try:
            yield
        except Exception:
            for path, snapshot_path in reversed(snapshots):
                _restore_sqlite_snapshot(path, snapshot_path)
            raise


def _checkpoint_sqlite(path: Path) -> None:
    if not path.is_file():
        return
    try:
        with closing(sqlite3.connect(path, timeout=30)) as connection:
            connection.execute("PRAGMA wal_checkpoint(FULL)")
    except sqlite3.Error:
        pass


def _restore_sqlite_snapshot(path: Path, snapshot_path: Path | None) -> None:
    for suffix in ("-wal", "-shm"):
        Path(f"{path}{suffix}").unlink(missing_ok=True)
    if snapshot_path is None:
        path.unlink(missing_ok=True)
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    replacement = path.with_name(f".restore-{os.getpid()}-{path.name}")
    if replacement.exists():
        replacement.unlink()
    with closing(sqlite3.connect(snapshot_path, timeout=30)) as source:
        with closing(sqlite3.connect(replacement, timeout=30)) as target:
            source.backup(target)
    os.replace(replacement, path)


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _combined_file_hash(paths: Iterable[Path]) -> str:
    digest = hashlib.sha256()
    for path in paths:
        file_path = Path(path)
        digest.update(file_path.name.encode("utf-8", errors="ignore"))
        digest.update(b"\0")
        digest.update(_file_hash(file_path).encode("ascii"))
        digest.update(b"\0")
    return digest.hexdigest()


def _current_user() -> str:
    try:
        return getpass.getuser()
    except Exception:
        return ""


def _credit_limit_import_logger() -> logging.Logger:
    logger = logging.getLogger("agribank_v3.credit_limit_import")
    if not any(getattr(handler, "_agribank_credit_limit_handler", False) for handler in logger.handlers):
        log_path = application_root() / "logs" / "credit_limit_import.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        handler = logging.FileHandler(log_path, encoding="utf-8")
        handler._agribank_credit_limit_handler = True
        handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)
    return logger


def _log_credit_limit_import_failure(
    *,
    repository: SummaryRepository,
    source_file: Path,
    stage: str,
    parsed_rows: int,
    first_row_type: str,
    batch_id: str,
    exc: Exception,
) -> None:
    store = getattr(repository, "credit_limit_store", None)
    temp_path = getattr(store, "temp_path", "")
    _credit_limit_import_logger().error(
        "Credit limit import failed:\n"
        "stage=%s\n"
        "source_file=%s\n"
        "parsed_rows=%s\n"
        "row_type=%s\n"
        "temp_path=%s\n"
        "batch_id=%s\n"
        "exception=%s",
        stage,
        source_file,
        parsed_rows,
        first_row_type or "",
        temp_path,
        batch_id or "",
        type(exc).__name__,
        exc_info=(type(exc), exc, exc.__traceback__),
    )
