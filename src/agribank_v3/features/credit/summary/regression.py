from __future__ import annotations

from dataclasses import dataclass
from itertools import zip_longest
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from agribank_v3.features.credit.summary.models import SummaryError


@dataclass(frozen=True, slots=True)
class WorkbookDiff:
    sheet_name: str
    cell: str
    issue: str
    vba_value: object
    python_value: object


@dataclass(frozen=True, slots=True)
class WorkbookCompareResult:
    vba_path: Path
    python_path: Path
    same: bool
    sheet_count_vba: int
    sheet_count_python: int
    diff_count: int
    report_path: Path


def compare_workbooks(
    vba_path: Path,
    python_path: Path,
    report_path: Path,
    *,
    max_diffs: int = 100_000,
) -> WorkbookCompareResult:
    vba_path = Path(vba_path)
    python_path = Path(python_path)
    report_path = Path(report_path)
    if not vba_path.is_file():
        raise SummaryError(f"Không tìm thấy file VBA: {vba_path}")
    if not python_path.is_file():
        raise SummaryError(f"Không tìm thấy file Python: {python_path}")
    if vba_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise SummaryError("Công cụ so sánh hiện hỗ trợ file VBA .xlsx/.xlsm.")
    if python_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise SummaryError("Công cụ so sánh hiện hỗ trợ file Python .xlsx/.xlsm.")

    vba_wb = load_workbook(vba_path, data_only=True, read_only=True)
    py_wb = load_workbook(python_path, data_only=True, read_only=True)
    try:
        diffs = list(_compare_open_workbooks(vba_wb, py_wb, max_diffs=max_diffs))
        _write_diff_report(
            report_path,
            vba_path=vba_path,
            python_path=python_path,
            vba_sheet_count=len(vba_wb.sheetnames),
            python_sheet_count=len(py_wb.sheetnames),
            diffs=diffs,
        )
        return WorkbookCompareResult(
            vba_path=vba_path,
            python_path=python_path,
            same=not diffs and vba_wb.sheetnames == py_wb.sheetnames,
            sheet_count_vba=len(vba_wb.sheetnames),
            sheet_count_python=len(py_wb.sheetnames),
            diff_count=len(diffs),
            report_path=report_path,
        )
    finally:
        vba_wb.close()
        py_wb.close()


def _compare_open_workbooks(vba_wb, py_wb, *, max_diffs: int) -> Iterable[WorkbookDiff]:
    vba_sheets = list(vba_wb.sheetnames)
    py_sheets = list(py_wb.sheetnames)
    if vba_sheets != py_sheets:
        yield WorkbookDiff(
            sheet_name="[Workbook]",
            cell="",
            issue="Danh sách sheet khác nhau",
            vba_value=", ".join(vba_sheets),
            python_value=", ".join(py_sheets),
        )
    for sheet_name in vba_sheets:
        if sheet_name not in py_wb.sheetnames:
            yield WorkbookDiff(sheet_name, "", "Thiếu sheet trong file Python", "Có", "Không")
            continue
        vba_ws = vba_wb[sheet_name]
        py_ws = py_wb[sheet_name]
        if vba_ws.max_row != py_ws.max_row:
            yield WorkbookDiff(sheet_name, "", "Số dòng khác nhau", vba_ws.max_row, py_ws.max_row)
        if vba_ws.max_column != py_ws.max_column:
            yield WorkbookDiff(sheet_name, "", "Số cột khác nhau", vba_ws.max_column, py_ws.max_column)
        max_row = max(vba_ws.max_row, py_ws.max_row)
        max_column = max(vba_ws.max_column, py_ws.max_column)
        count = 0
        vba_rows = vba_ws.iter_rows(max_row=max_row, max_col=max_column, values_only=True)
        py_rows = py_ws.iter_rows(max_row=max_row, max_col=max_column, values_only=True)
        empty_row = tuple("" for _ in range(max_column))
        for row_index, (vba_row, py_row) in enumerate(zip_longest(vba_rows, py_rows, fillvalue=empty_row), start=1):
            for column_index, (vba_value, py_value) in enumerate(zip_longest(vba_row, py_row, fillvalue=""), start=1):
                if _normalize_cell(vba_value) != _normalize_cell(py_value):
                    count += 1
                    coordinate = f"{get_column_letter(column_index)}{row_index}"
                    yield WorkbookDiff(
                        sheet_name=sheet_name,
                        cell=coordinate,
                        issue="Giá trị ô khác nhau",
                        vba_value=vba_value,
                        python_value=py_value,
                    )
                    if count >= max_diffs:
                        yield WorkbookDiff(
                            sheet_name=sheet_name,
                            cell="",
                            issue="Đã đạt giới hạn số khác biệt",
                            vba_value=max_diffs,
                            python_value=max_diffs,
                        )
                        return


def _write_diff_report(
    report_path: Path,
    *,
    vba_path: Path,
    python_path: Path,
    vba_sheet_count: int,
    python_sheet_count: int,
    diffs: list[WorkbookDiff],
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "TongHop"
    summary.append(["Chỉ tiêu", "VBA", "Python"])
    summary.append(["File", str(vba_path), str(python_path)])
    summary.append(["Số sheet", vba_sheet_count, python_sheet_count])
    summary.append(["Số khác biệt", len(diffs), len(diffs)])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary["C1"].font = Font(bold=True)

    details = workbook.create_sheet("ChiTiet")
    details.append(["Sheet", "Ô", "Lỗi", "Giá trị VBA", "Giá trị Python"])
    fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in details[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for diff in diffs:
        details.append(
            [
                diff.sheet_name,
                diff.cell,
                diff.issue,
                diff.vba_value,
                diff.python_value,
            ]
        )
    for worksheet in (summary, details):
        for column_cells in worksheet.columns:
            width = min(60, max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2))
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(report_path)


def _normalize_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 10)
    return str(value).strip() if isinstance(value, str) else value
