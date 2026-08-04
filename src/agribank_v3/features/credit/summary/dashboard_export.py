from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agribank_v3.features.credit.summary.dashboard_service import (
    METRIC_LABELS,
    DashboardNimData,
    latest_branch_rows,
    metric_value,
)
from agribank_v3.features.credit.summary.nim_ui_config import NIM_DN_UI_CONFIG, NIM_NV_UI_CONFIG, NimUiConfig
from agribank_v3.features.credit.summary.officer_history.models import (
    METRIC_AVERAGE_RATE,
    METRIC_BALANCE,
    METRIC_BALANCE_GROWTH,
    METRIC_NIM_AFTER,
    METRIC_NIM_BEFORE,
)


SHEET_OVERVIEW = "TongQuanTheoKy"
SHEET_BRANCH = "SoSanhChiNhanh"
SHEET_GROWTH = "TangTruong"
SHEET_DETAIL = "BangDuLieuChiTiet"
SHEET_BY_TAB = {
    "overview": SHEET_OVERVIEW,
    "branch": SHEET_BRANCH,
    "growth": SHEET_GROWTH,
    "detail": SHEET_DETAIL,
}
MONEY_HEADERS = {
    "Tổng dư nợ",
    "Dư nợ",
    "Tăng/giảm dư nợ tuyệt đối",
    "Tăng/giảm dư nợ",
    "Tổng nguồn vốn",
    "Số dư nguồn vốn",
    "Tăng/giảm nguồn vốn tuyệt đối",
    "Tăng/giảm nguồn vốn",
}
PERCENT_HEADERS = {
    "Lãi suất bình quân",
    "NIM trước ĐC",
    "NIM sau ĐC",
    "Tăng trưởng dư nợ (%)",
    "Tăng trưởng nguồn vốn (%)",
    "Biến động NIM trước ĐC",
    "Biến động NIM sau ĐC",
    "Tăng trưởng (%)",
    "Thay đổi (điểm %)",
}


class DashboardNimExportService:
    def __init__(self, data: DashboardNimData, *, metric: str = METRIC_BALANCE) -> None:
        self.data = data
        self.metric = metric
        self.ui_config = data.ui_config

    def rows_for_tab(self, tab_key: str) -> list[dict[str, object]]:
        if tab_key == "branch":
            return self.branch_comparison_rows()
        if tab_key == "growth":
            return self.growth_rows()
        if tab_key == "detail":
            return self.detail_rows()
        return self.overview_by_period_rows()

    def overview_by_period_rows(self) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for row in self.data.period_rows:
            rows.append(self._overview_row(row, row.balance_delta, row.balance_growth_percent))
        return rows

    def overview_endpoint_rows(self) -> list[dict[str, object]]:
        period_rows = tuple(self.data.period_rows)
        if not period_rows:
            return []
        from_row, to_row = _period_endpoint_rows(period_rows, self.data.filters.period_from, self.data.filters.period_to)
        if from_row.period == to_row.period:
            return [self._overview_row(from_row, None, None)]
        delta = to_row.balance - from_row.balance
        growth = None if from_row.balance == 0 else (delta / from_row.balance) * 100
        return [
            self._overview_row(from_row, None, None),
            self._overview_row(to_row, delta, growth),
        ]

    def branch_comparison_rows(self) -> list[dict[str, object]]:
        rows = latest_branch_rows(self.data.branch_rows)
        rows = tuple(
            sorted(
                rows,
                key=lambda row: float("-inf") if metric_value(row, self.metric) is None else float(metric_value(row, self.metric) or 0),
                reverse=True,
            )
        )
        output: list[dict[str, object]] = []
        for row in rows:
            payload = {
                "Kỳ": row.period,
                "Tên chi nhánh": row.branch,
                self.ui_config.balance_label: row.balance,
                "NIM trước ĐC": row.nim_before,
                "NIM sau ĐC": row.nim_after,
                "Chỉ tiêu đang chọn": self.ui_config.metric_labels().get(self.metric, self.metric),
                "Giá trị chỉ tiêu": metric_value(row, self.metric),
            }
            if self.ui_config.include_average_rate:
                payload = _insert_after(payload, self.ui_config.balance_label, "Lãi suất bình quân", row.average_rate)
            output.append(payload)
        return output

    def branch_current_rows(self) -> list[dict[str, object]]:
        rows = latest_branch_rows(self.data.branch_rows)
        rows = tuple(
            sorted(
                rows,
                key=lambda row: float("-inf") if metric_value(row, self.metric) is None else float(metric_value(row, self.metric) or 0),
                reverse=True,
            )
        )
        output: list[dict[str, object]] = []
        for row in rows:
            payload = {
                "Kỳ": row.period,
                "Mã chi nhánh": row.branch_code,
                "Tên chi nhánh": row.branch,
                self.ui_config.balance_label: row.balance,
                "NIM trước ĐC": row.nim_before,
                "NIM sau ĐC": row.nim_after,
                "Chỉ tiêu đang chọn": self.ui_config.metric_labels().get(self.metric, self.metric),
                "Giá trị chỉ tiêu": metric_value(row, self.metric),
            }
            if self.ui_config.include_average_rate:
                payload = _insert_after(payload, self.ui_config.balance_label, "Lãi suất bình quân", row.average_rate)
            output.append(payload)
        return output

    def branch_period_comparison_rows(self) -> list[dict[str, object]]:
        branch_rows = tuple(self.data.branch_rows)
        if not branch_rows:
            return []
        from_period, to_period = _period_bounds(branch_rows, self.data.filters.period_from, self.data.filters.period_to)
        from_rows = {row.branch_code or row.branch: row for row in branch_rows if row.period == from_period}
        to_rows = {row.branch_code or row.branch: row for row in branch_rows if row.period == to_period}
        keys = sorted(
            set(from_rows) | set(to_rows),
            key=lambda key: ((to_rows.get(key) or from_rows.get(key)).branch if (to_rows.get(key) or from_rows.get(key)) else "").casefold(),
        )
        output: list[dict[str, object]] = []
        metric_label = self.ui_config.metric_labels().get(self.metric, self.metric)
        for key in keys:
            from_row = from_rows.get(key)
            to_row = to_rows.get(key)
            display_row = to_row or from_row
            if display_row is None:
                continue
            if self.metric == METRIC_BALANCE:
                from_value = from_row.balance if from_row else None
                to_value = to_row.balance if to_row else None
                delta = None if from_value is None or to_value is None else to_value - from_value
                growth = None if delta is None or from_value in (None, 0) else (delta / from_value) * 100
                output.append(
                    {
                        "Mã chi nhánh": display_row.branch_code,
                        "Tên chi nhánh": display_row.branch,
                        f"{self.ui_config.balance_label} Từ kỳ": from_value,
                        f"{self.ui_config.balance_label} Đến kỳ": to_value,
                        "Tăng/giảm tuyệt đối": delta,
                        "Tăng trưởng (%)": growth,
                    }
                )
            else:
                from_value = metric_value(from_row, self.metric) if from_row else None
                to_value = metric_value(to_row, self.metric) if to_row else None
                delta = None if from_value is None or to_value is None else to_value - from_value
                output.append(
                    {
                        "Mã chi nhánh": display_row.branch_code,
                        "Tên chi nhánh": display_row.branch,
                        f"{metric_label} Từ kỳ": from_value,
                        f"{metric_label} Đến kỳ": to_value,
                        "Thay đổi (điểm %)": delta,
                    }
                )
        return output

    def growth_rows(self) -> list[dict[str, object]]:
        return [
            {
                "Kỳ": row.period,
                "Tên chi nhánh": row.branch,
                "Phòng GD": row.transaction_office,
                "Loại KH": row.customer_type,
                self.ui_config.balance_label: row.balance,
                self.ui_config.balance_delta_label: row.balance_delta,
                self.ui_config.growth_percent_label: row.balance_growth_percent,
                "Biến động NIM trước ĐC": row.nim_before_delta,
                "Biến động NIM sau ĐC": row.nim_after_delta,
            }
            for row in self.data.detail_rows
        ]

    def detail_rows(self) -> list[dict[str, object]]:
        output: list[dict[str, object]] = []
        for row in self.data.detail_rows:
            payload = {
                "Kỳ": row.period,
                "Tên chi nhánh": row.branch,
                "Phòng GD": row.transaction_office,
                "Loại KH": row.customer_type,
                self.ui_config.balance_label: row.balance,
                "NIM trước ĐC": row.nim_before,
                "NIM sau ĐC": row.nim_after,
                self.ui_config.growth_percent_label: row.balance_growth_percent,
                self.ui_config.balance_delta_label: row.balance_delta,
            }
            if self.ui_config.include_average_rate:
                payload = _insert_after(payload, self.ui_config.balance_label, "Lãi suất bình quân", row.average_rate)
            output.append(payload)
        return output

    def detail_branch_rows(self) -> list[dict[str, object]]:
        output: list[dict[str, object]] = []
        customer_type = self.data.filters.customer_type or "Tất cả"
        for row in self.data.branch_rows:
            payload = {
                "Kỳ": row.period,
                "Mã chi nhánh": row.branch_code,
                "Tên chi nhánh": row.branch,
                "Loại KH": customer_type,
                self.ui_config.balance_label: row.balance,
                "NIM trước ĐC": row.nim_before,
                "NIM sau ĐC": row.nim_after,
                self.ui_config.growth_percent_label: row.balance_growth_percent,
                self.ui_config.balance_delta_label: row.balance_delta,
            }
            if self.ui_config.include_average_rate:
                payload = _insert_after(payload, self.ui_config.balance_label, "Lãi suất bình quân", row.average_rate)
            output.append(payload)
        return output

    def detail_office_rows(self) -> list[dict[str, object]]:
        output: list[dict[str, object]] = []
        for row in self.data.detail_rows:
            payload = {
                "Kỳ": row.period,
                "Mã chi nhánh": row.branch_code,
                "Tên chi nhánh": row.branch,
                "Mã đơn vị": row.office_code,
                "Hội sở/Phòng GD": row.transaction_office,
                "Loại đơn vị": row.office_type,
                "Loại KH": row.customer_type,
                self.ui_config.balance_label: row.balance,
                "NIM trước ĐC": row.nim_before,
                "NIM sau ĐC": row.nim_after,
                self.ui_config.growth_percent_label: row.balance_growth_percent,
                self.ui_config.balance_delta_label: row.balance_delta,
            }
            if self.ui_config.include_average_rate:
                payload = _insert_after(payload, self.ui_config.balance_label, "Lãi suất bình quân", row.average_rate)
            output.append(payload)
        return output

    def _overview_row(self, row, balance_delta: float | None, balance_growth_percent: float | None) -> dict[str, object]:
        payload = {
            "Kỳ": row.period,
            self.ui_config.total_balance_label: row.balance,
            "NIM trước ĐC": row.nim_before,
            "NIM sau ĐC": row.nim_after,
            self.ui_config.balance_delta_label: balance_delta,
            self.ui_config.growth_percent_label: balance_growth_percent,
        }
        if self.ui_config.include_average_rate:
            payload = _insert_after(payload, self.ui_config.total_balance_label, "Lãi suất bình quân", row.average_rate)
        return payload

    def export_overview_by_period(self, destination: Path) -> Path:
        return self._export_single(destination, "overview")

    def export_branch_comparison(self, destination: Path) -> Path:
        return self._export_single(destination, "branch")

    def export_growth(self, destination: Path) -> Path:
        return self._export_single(destination, "growth")

    def export_detail(self, destination: Path) -> Path:
        return self._export_single(destination, "detail")

    def export_all_tabs(self, destination: Path) -> Path:
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        for tab_key in ("overview", "branch", "growth", "detail"):
            _write_sheet(workbook, self.ui_config.dashboard_sheets[tab_key], self.rows_for_tab(tab_key))
        workbook.save(destination)
        return destination

    def _export_single(self, destination: Path, tab_key: str) -> Path:
        return export_dashboard_rows(self.rows_for_tab(tab_key), destination, sheet_name=self.ui_config.dashboard_sheets[tab_key])


def export_dashboard_rows(rows: list[dict[str, object]], destination: Path, *, sheet_name: str, metadata: list[tuple[str, object]] | None = None) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or "DashboardNIM"
    _write_rows(worksheet, rows)
    if metadata:
        _write_metadata(workbook, metadata)
    workbook.save(destination)
    return destination


def export_dashboard_workbook(
    sheets: list[tuple[str, list[dict[str, object]]]],
    destination: Path,
    *,
    metadata: list[tuple[str, object]] | None = None,
) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_name, rows in sheets:
        _write_sheet(workbook, sheet_name[:31] or "DashboardNIM", rows)
    if metadata:
        _write_metadata(workbook, metadata)
    workbook.save(destination)
    return destination


def _write_sheet(workbook: Workbook, sheet_name: str, rows: list[dict[str, object]]) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    _write_rows(worksheet, rows)


def _write_rows(worksheet, rows: list[dict[str, object]]) -> None:
    headers = list(rows[0].keys()) if rows else ["Thông báo"]
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if rows:
        for row in rows:
            worksheet.append([_cell_value(row.get(header)) for header in headers])
    else:
        worksheet.append(["Không có dữ liệu"])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _format_columns(worksheet, headers)
    _fit_columns(worksheet)


def _write_metadata(workbook: Workbook, metadata: list[tuple[str, object]]) -> None:
    worksheet = workbook.create_sheet("ThongTin")
    worksheet.append(["Thông tin", "Giá trị"])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for key, value in metadata:
        worksheet.append([key, _cell_value(value)])
    _fit_columns(worksheet)


def _format_columns(worksheet, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=worksheet.max_row):
            for item in cell:
                if _is_money_header(header):
                    item.number_format = "#,##0"
                elif _is_percent_header(header) or (header == "Giá trị chỉ tiêu" and _selected_metric_is_percent(worksheet)):
                    item.number_format = '0.00"%"'
                elif header == "Giá trị chỉ tiêu" and _selected_metric_is_money(worksheet):
                    item.number_format = "#,##0"
                if isinstance(item.value, (int, float)):
                    item.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    item.alignment = Alignment(horizontal="left", vertical="center")


def _fit_columns(worksheet) -> None:
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(48, max(10, max_length + 2))


def _cell_value(value: object) -> object:
    if value is None:
        return ""
    return value


def _period_endpoint_rows(rows, from_period: str, to_period: str):
    row_by_period = {row.period: row for row in rows}
    periods = sorted(row_by_period)
    clean_from = from_period if from_period in row_by_period else periods[0]
    clean_to = to_period if to_period in row_by_period else periods[-1]
    return row_by_period[clean_from], row_by_period[clean_to]


def _period_bounds(rows, from_period: str, to_period: str) -> tuple[str, str]:
    periods = sorted({row.period for row in rows})
    if not periods:
        return "", ""
    clean_from = from_period if from_period in periods else periods[0]
    clean_to = to_period if to_period in periods else periods[-1]
    return clean_from, clean_to


def _is_money_header(header: str) -> bool:
    if header in MONEY_HEADERS:
        return True
    lowered = header.casefold()
    return (
        "dư nợ" in lowered
        or "nguồn vốn" in lowered
        or "số dư" in lowered
        or "tăng/giảm tuyệt đối" in lowered
    ) and "tăng trưởng" not in lowered


def _is_percent_header(header: str) -> bool:
    if header in PERCENT_HEADERS:
        return True
    lowered = header.casefold()
    return "nim" in lowered or "lãi suất" in lowered or "tăng trưởng" in lowered or "điểm %" in lowered


def _selected_metric_is_percent(worksheet) -> bool:
    label = _selected_metric_label(worksheet)
    return label in _metric_label_values(METRIC_NIM_BEFORE, METRIC_NIM_AFTER, METRIC_AVERAGE_RATE, METRIC_BALANCE_GROWTH)


def _selected_metric_is_money(worksheet) -> bool:
    return _selected_metric_label(worksheet) in {config.metric_labels()[METRIC_BALANCE] for config in (NIM_DN_UI_CONFIG, NIM_NV_UI_CONFIG)}


def _selected_metric_label(worksheet) -> str:
    headers = [cell.value for cell in worksheet[1]]
    if "Chỉ tiêu đang chọn" not in headers:
        return ""
    column_index = headers.index("Chỉ tiêu đang chọn") + 1
    return str(worksheet.cell(2, column_index).value or "")


def _insert_after(source: dict[str, object], after_key: str, key: str, value: object) -> dict[str, object]:
    output: dict[str, object] = {}
    for existing_key, existing_value in source.items():
        output[existing_key] = existing_value
        if existing_key == after_key:
            output[key] = value
    if key not in output:
        output[key] = value
    return output


def _metric_label_values(*metrics: str) -> set[str]:
    labels = set()
    default_labels = dict(METRIC_LABELS)
    for metric in metrics:
        if metric in default_labels:
            labels.add(default_labels[metric])
    for config in (NIM_DN_UI_CONFIG, NIM_NV_UI_CONFIG):
        config_labels = config.metric_labels()
        for metric in metrics:
            if metric in config_labels:
                labels.add(config_labels[metric])
    return labels
