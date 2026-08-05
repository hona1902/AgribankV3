from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from agribank_v3.features.credit.summary.credit_report import VIEW_COMPARE_PERIODS, VIEW_CURRENT_PERIOD

from .models import DETAIL_BY_GROUP, SUMMARY_BY_ASSOCIATION, GroupLendingFilters
from .service import GroupLendingService


class GroupLendingExportService:
    def __init__(self, service: GroupLendingService) -> None:
        self.service = service

    def export(
        self,
        destination: Path,
        *,
        filters: GroupLendingFilters,
        view_mode: str,
        display_mode: str,
    ) -> Path:
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        try:
            if view_mode == VIEW_COMPARE_PERIODS:
                from_period = filters.from_period
                to_period = filters.to_period
                if display_mode == SUMMARY_BY_ASSOCIATION:
                    result = self.service.compare_associations(from_period, to_period, filters)
                    sheet_name = "SoSanhTheoHoi"
                else:
                    result = self.service.compare_groups(from_period, to_period, filters)
                    sheet_name = "SoSanhTheoTo"
            else:
                period = filters.period
                if display_mode == SUMMARY_BY_ASSOCIATION:
                    result = self.service.get_association_summary(period, filters)
                    sheet_name = "TongHopTheoHoi"
                else:
                    result = self.service.get_group_lending_snapshot(period, filters, page=1, page_size=50000)
                    sheet_name = "ChoVayQuaTo"
            sheet = workbook.active
            sheet.title = sheet_name
            _write_rows(sheet, [_row_to_dict(row, index) for index, row in enumerate(result.rows, start=1)])
            info = workbook.create_sheet("ThongTin")
            _write_rows(info, _metadata_rows(filters, view_mode, display_mode, result.diagnostics or {}))
            workbook.save(destination)
        finally:
            workbook.close()
        return destination


def _row_to_dict(row: object, index: int) -> dict[str, object]:
    if hasattr(row, "to_dict"):
        data = row.to_dict(index) if row.__class__.__name__ == "GroupLendingRow" else row.to_dict()
        return dict(data)
    return dict(row) if isinstance(row, dict) else {"Giá trị": row}


def _metadata_rows(
    filters: GroupLendingFilters,
    view_mode: str,
    display_mode: str,
    diagnostics: dict[str, object],
) -> list[dict[str, object]]:
    return [
        {"Chỉ tiêu": "Chế độ", "Giá trị": "So sánh các kỳ" if view_mode == VIEW_COMPARE_PERIODS else "Kỳ hiện tại"},
        {"Chỉ tiêu": "Kiểu hiển thị", "Giá trị": "Tổng hợp theo Hội" if display_mode == SUMMARY_BY_ASSOCIATION else "Chi tiết theo tổ"},
        {"Chỉ tiêu": "Kỳ báo cáo", "Giá trị": filters.period},
        {"Chỉ tiêu": "Từ kỳ", "Giá trị": filters.from_period},
        {"Chỉ tiêu": "Đến kỳ", "Giá trị": filters.to_period},
        {"Chỉ tiêu": "Chi nhánh", "Giá trị": filters.branch_code or "Tất cả"},
        {"Chỉ tiêu": "Phòng giao dịch", "Giá trị": filters.office_code or "Tất cả"},
        {"Chỉ tiêu": "Loại Hội", "Giá trị": filters.association_type or "Tất cả"},
        {"Chỉ tiêu": "Trạng thái tổ", "Giá trị": filters.group_status or "Tất cả"},
        {"Chỉ tiêu": "CBTD", "Giá trị": filters.officer or "Tất cả"},
        {"Chỉ tiêu": "Tìm kiếm", "Giá trị": filters.search},
        {"Chỉ tiêu": "Nguồn GRPNO", "Giá trị": "LN01.GRPNO -> credit_groups.ma_to"},
        {"Chỉ tiêu": "Dư nợ không có GRPNO", "Giá trị": diagnostics.get("no_group_balance", 0)},
        {"Chỉ tiêu": "Khách hàng ở nhiều tổ", "Giá trị": diagnostics.get("multi_group_customer_count", 0)},
        {"Chỉ tiêu": "Dư nợ khách hàng ở nhiều tổ", "Giá trị": diagnostics.get("multi_group_customer_balance", 0)},
        {"Chỉ tiêu": "Dư nợ thẻ", "Giá trị": "Không cộng DN15 khi nguồn thẻ không có mã tổ"},
        {"Chỉ tiêu": "Thời gian xuất", "Giá trị": datetime.now().astimezone().isoformat(timespec="seconds")},
    ]


def _write_rows(sheet, rows: list[dict[str, object]]) -> None:
    headers = list(rows[0].keys()) if rows else ["Thông tin"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row_index, row in enumerate(rows, start=2):
        for column, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = sheet.cell(row_index, column, value)
            if header in _MONEY_HEADERS:
                cell.number_format = "#,##0"
            elif header in _PERCENT_HEADERS:
                cell.number_format = '0.00"%"'
            elif header in {"Mã tổ"}:
                cell.number_format = "@"
    for column, header in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = min(max(12, len(header) + 4), 28)


_MONEY_HEADERS = {
    "Tổng dư nợ",
    "Dư nợ bình quân/tổ viên",
    "Dư nợ bình quân/tổ",
    "Dư nợ Từ kỳ",
    "Dư nợ Đến kỳ",
    "Tăng/giảm dư nợ",
    "Dư nợ không có GRPNO",
    "Dư nợ khách hàng ở nhiều tổ",
}
_PERCENT_HEADERS = {
    "Tỷ trọng",
    "Tăng trưởng (%)",
    "Tăng trưởng dư nợ (%)",
    "Tỷ trọng Từ kỳ",
    "Tỷ trọng Đến kỳ",
    "Thay đổi tỷ trọng (điểm %)",
}
