from __future__ import annotations

from collections import defaultdict
from contextlib import closing
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
import hashlib
import json
import os
import shutil
import sqlite3
import tempfile
from time import perf_counter
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from agribank_v3.features.credit.summary.models import (
    DEBT_GROUP_UNKNOWN,
    NormalizedLn01Row,
    NormalizedLoanRow,
    SummaryError,
    now_text,
)
from agribank_v3.runtime_paths import application_root


CREDIT_DATABASE_NAME = "Credit.db"
CREDIT_SCHEMA_VERSION = "0.2.0"
MAIN_DATABASE_NAME = "DuLieuV3.db"
CREDIT_SOURCE_LN01 = "LN01"
CREDIT_SOURCE_NIM_DN_CARD = "NIM_DN_CARD"
CUSTOMER_TYPE_PERSONAL = "PERSONAL"
CUSTOMER_TYPE_LEGAL = "LEGAL"
CUSTOMER_TYPE_UNKNOWN = "UNKNOWN"
TERM_SHORT = "SHORT"
TERM_MEDIUM = "MEDIUM"
TERM_LONG = "LONG"
TERM_UNKNOWN = "UNKNOWN"
INDUSTRY_HIGH_TECH = "HIGH_TECH"
INDUSTRY_WHOLESALE_RETAIL = "WHOLESALE_RETAIL"
INDUSTRY_REAL_ESTATE = "REAL_ESTATE"
INDUSTRY_OTHER = "OTHER"
VIEW_CURRENT_PERIOD = "CURRENT_PERIOD"
VIEW_COMPARE_PERIODS = "COMPARE_PERIODS"
GROUP_SUMMARY = "SUMMARY"
GROUP_TERM_STRUCTURE = "TERM_STRUCTURE"
GROUP_CUSTOMER_TYPE = "CUSTOMER_TYPE"
GROUP_CREDIT_QUALITY = "CREDIT_QUALITY"
GROUP_DECREE55 = "DECREE_55"
GROUP_INDUSTRY = "INDUSTRY"
CREDIT_QUALITY_DISPLAY_NAME = "Chất lượng tín dụng"

CUSTOMER_TYPE_LABELS = {
    CUSTOMER_TYPE_PERSONAL: "Cá nhân",
    CUSTOMER_TYPE_LEGAL: "Pháp nhân",
    CUSTOMER_TYPE_UNKNOWN: "Chưa xác định",
}
TERM_LABELS = {
    TERM_SHORT: "Ngắn hạn",
    TERM_MEDIUM: "Trung hạn",
    TERM_LONG: "Dài hạn",
    TERM_UNKNOWN: "Chưa phân loại",
}
DEBT_BUCKET_LABELS = {
    "GROUP_1": "Nợ nhóm 1",
    "GROUP_2": "Nợ nhóm 2 - Nợ cần chú ý",
    "BAD_DEBT": "Nợ nhóm 3-5 - Nợ xấu",
    "UNKNOWN": "Chưa xác định nhóm nợ",
}
INDUSTRY_LABELS = {
    INDUSTRY_HIGH_TECH: "Cho vay ứng dụng công nghệ cao",
    INDUSTRY_WHOLESALE_RETAIL: "Cho vay bán buôn, bán lẻ",
    INDUSTRY_REAL_ESTATE: "Cho vay bất động sản",
    INDUSTRY_OTHER: "Ngành khác",
}
HIGH_TECH_CODES = {"10104", "10105", "10106"}
REAL_ESTATE_CODES = {
    "240101",
    "240106",
    "240108",
    "240116",
    "240117",
    "240118",
}


@dataclass(frozen=True, slots=True)
class CreditReportFilters:
    period: str = ""
    from_period: str = ""
    to_period: str = ""
    branch_code: str = ""
    transaction_office: str = ""
    customer_type: str = ""
    debt_group: str = ""
    term_category: str = ""
    officer: str = ""
    search: str = ""


@dataclass(frozen=True, slots=True)
class ReportMetric:
    label: str
    value: object = None
    value_kind: str = "number"
    tooltip: str = ""
    from_value: object = None
    to_value: object = None
    difference: object = None
    growth_rate: float | None = None
    note: str = ""


@dataclass(frozen=True, slots=True)
class ReportTableRow:
    values: tuple[tuple[str, object], ...]

    @classmethod
    def from_mapping(cls, row: dict[str, object]) -> "ReportTableRow":
        return cls(tuple(row.items()))

    def to_dict(self) -> dict[str, object]:
        return dict(self.values)

    def get(self, key: str, default: object = None) -> object:
        return self.to_dict().get(key, default)


@dataclass(frozen=True, slots=True)
class ReportSnapshot:
    period: str
    filters: CreditReportFilters
    exists: bool
    summary: dict[str, object]
    groups: dict[str, tuple[ReportTableRow, ...]]
    card_status: str
    card_message: str
    note: str = ""


@dataclass(frozen=True, slots=True)
class ReportComparisonResult:
    from_snapshot: ReportSnapshot
    to_snapshot: ReportSnapshot
    group_key: str
    rows: tuple[ReportTableRow, ...]
    kpis: tuple[ReportMetric, ...]
    notes: tuple[str, ...] = ()


def credit_database_path(main_database_path: Path | None = None) -> Path:
    if main_database_path is None:
        return application_root() / "data" / CREDIT_DATABASE_NAME
    path = Path(main_database_path)
    if path.name.casefold() == CREDIT_DATABASE_NAME.casefold():
        return path
    return path.parent / CREDIT_DATABASE_NAME


def get_credit_connection(database_path: Path | None = None) -> sqlite3.Connection:
    path = credit_database_path(database_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(path, timeout=30)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute("PRAGMA journal_mode = WAL")
    connection.execute("PRAGMA busy_timeout = 30000")
    return connection


class CreditReportRepository:
    def __init__(self, main_database_path: Path | None = None) -> None:
        self.main_database_path = Path(main_database_path) if main_database_path is not None else None
        self.database_path = credit_database_path(main_database_path)
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self.ensure_schema()

    def connect(self) -> sqlite3.Connection:
        return get_credit_connection(self.database_path)

    def ensure_schema(self) -> None:
        with closing(self.connect()) as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS credit_schema_migrations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    version TEXT NOT NULL,
                    migration_name TEXT NOT NULL DEFAULT '',
                    applied_at TEXT NOT NULL,
                    checksum TEXT NOT NULL DEFAULT '',
                    success INTEGER NOT NULL DEFAULT 1 CHECK(success IN (0, 1))
                );
                CREATE UNIQUE INDEX IF NOT EXISTS idx_credit_schema_migrations_version
                    ON credit_schema_migrations(version, migration_name)
                    WHERE success = 1;

                CREATE TABLE IF NOT EXISTS credit_import_runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    period TEXT NOT NULL,
                    source_type TEXT NOT NULL,
                    source_file_name TEXT NOT NULL DEFAULT '',
                    source_sha256 TEXT NOT NULL DEFAULT '',
                    source_file_size INTEGER NOT NULL DEFAULT 0,
                    source_row_count INTEGER NOT NULL DEFAULT 0,
                    normalized_loan_count INTEGER NOT NULL DEFAULT 0,
                    customer_count INTEGER NOT NULL DEFAULT 0,
                    ln01_total_balance REAL NOT NULL DEFAULT 0,
                    card_balance REAL NOT NULL DEFAULT 0,
                    status TEXT NOT NULL DEFAULT 'success',
                    warning_count INTEGER NOT NULL DEFAULT 0,
                    message TEXT NOT NULL DEFAULT '',
                    imported_by TEXT NOT NULL DEFAULT '',
                    group_code_stats_json TEXT NOT NULL DEFAULT '{}',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    duration_ms INTEGER NOT NULL DEFAULT 0
                );
                CREATE INDEX IF NOT EXISTS idx_credit_import_runs_period
                    ON credit_import_runs(period, source_type, created_at DESC);
                CREATE UNIQUE INDEX IF NOT EXISTS idx_credit_import_runs_source_sha
                    ON credit_import_runs(source_type, period, source_sha256)
                    WHERE source_sha256 <> '' AND status = 'success';

                CREATE TABLE IF NOT EXISTS credit_import_files (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    run_id INTEGER NOT NULL,
                    period TEXT NOT NULL,
                    source_type TEXT NOT NULL,
                    file_name TEXT NOT NULL DEFAULT '',
                    sha256 TEXT NOT NULL DEFAULT '',
                    file_size INTEGER NOT NULL DEFAULT 0,
                    row_count INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(run_id) REFERENCES credit_import_runs(id)
                        ON DELETE CASCADE
                );

                CREATE TABLE IF NOT EXISTS credit_customer_master (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    branch_code TEXT NOT NULL,
                    customer_sequence TEXT NOT NULL,
                    customer_code TEXT NOT NULL DEFAULT '',
                    customer_name TEXT NOT NULL DEFAULT '',
                    latest_customer_type_code TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(branch_code, customer_sequence)
                );
                CREATE INDEX IF NOT EXISTS idx_credit_customer_master_code
                    ON credit_customer_master(customer_code);

                CREATE TABLE IF NOT EXISTS credit_loan_period (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    period TEXT NOT NULL,
                    customer_id INTEGER NOT NULL,
                    branch_code TEXT NOT NULL,
                    loan_key TEXT NOT NULL,
                    account_number TEXT NOT NULL DEFAULT '',
                    approval_sequence TEXT NOT NULL DEFAULT '',
                    outstanding_balance REAL NOT NULL DEFAULT 0,
                    customer_type_code TEXT NOT NULL DEFAULT '',
                    debt_group_code TEXT NOT NULL DEFAULT 'UNKNOWN',
                    secured_percent REAL,
                    industry_code TEXT NOT NULL DEFAULT '',
                    officer_code TEXT NOT NULL DEFAULT '',
                    officer_name TEXT NOT NULL DEFAULT '',
                    term_category TEXT NOT NULL DEFAULT 'UNKNOWN',
                    group_code TEXT NULL,
                    source_run_id INTEGER NOT NULL,
                    source_row_count INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(customer_id) REFERENCES credit_customer_master(id)
                        ON DELETE RESTRICT,
                    FOREIGN KEY(source_run_id) REFERENCES credit_import_runs(id)
                        ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_branch
                    ON credit_loan_period(period, branch_code);
                CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_customer
                    ON credit_loan_period(period, customer_id);
                CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_debt
                    ON credit_loan_period(period, debt_group_code);
                CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_term
                    ON credit_loan_period(period, term_category);
                CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_type
                    ON credit_loan_period(period, customer_type_code);
                CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_industry
                    ON credit_loan_period(period, industry_code);

                CREATE TABLE IF NOT EXISTS credit_card_period (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    period TEXT NOT NULL,
                    customer_id INTEGER NOT NULL,
                    branch_code TEXT NOT NULL,
                    customer_code TEXT NOT NULL DEFAULT '',
                    customer_name TEXT NOT NULL DEFAULT '',
                    customer_type TEXT NOT NULL DEFAULT 'UNKNOWN',
                    debt_group_code TEXT NOT NULL DEFAULT 'UNKNOWN',
                    officer_code TEXT NOT NULL DEFAULT '',
                    officer_name TEXT NOT NULL DEFAULT '',
                    balance REAL NOT NULL DEFAULT 0,
                    source_run_id INTEGER NOT NULL,
                    source_row_count INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(customer_id) REFERENCES credit_customer_master(id)
                        ON DELETE RESTRICT,
                    FOREIGN KEY(source_run_id) REFERENCES credit_import_runs(id)
                        ON DELETE CASCADE,
                    UNIQUE(
                        period,
                        branch_code,
                        customer_id,
                        customer_type,
                        debt_group_code,
                        officer_code,
                        officer_name
                    )
                );
                CREATE INDEX IF NOT EXISTS idx_credit_card_period_period_branch_customer
                    ON credit_card_period(period, branch_code, customer_id);

                CREATE TABLE IF NOT EXISTS credit_customer_type_rules (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code TEXT NOT NULL,
                    label TEXT NOT NULL DEFAULT '',
                    is_personal INTEGER NOT NULL DEFAULT 1 CHECK(is_personal IN (0, 1)),
                    is_enabled INTEGER NOT NULL DEFAULT 1 CHECK(is_enabled IN (0, 1)),
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(code)
                );

                CREATE TABLE IF NOT EXISTS credit_action_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    action TEXT NOT NULL,
                    target_id TEXT NOT NULL DEFAULT '',
                    detail TEXT NOT NULL DEFAULT '',
                    created_by TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_credit_action_log_created
                    ON credit_action_log(created_at DESC);
                """
            )
            self._migrate_credit_loan_group_code(connection)
            self._ensure_group_code_metadata_column(connection)
            self._ensure_credit_loan_indexes(connection)
            self._seed_default_customer_type_rules(connection)
            self._mark_migration(connection)
            connection.commit()

    def _ensure_group_code_metadata_column(self, connection: sqlite3.Connection) -> None:
        if not _table_has_column(connection, "credit_import_runs", "group_code_stats_json"):
            self._backup_database_before_schema_change("group-code-metadata")
            connection.execute("ALTER TABLE credit_import_runs ADD COLUMN group_code_stats_json TEXT NOT NULL DEFAULT '{}'")

    def _migrate_credit_loan_group_code(self, connection: sqlite3.Connection) -> None:
        if not _table_has_column(connection, "credit_loan_period", "group_code"):
            self._backup_database_before_schema_change("loan-group-code")
            self._rebuild_credit_loan_period(connection, has_group_code=False)
            return
        if _has_old_credit_loan_unique_index(connection):
            self._backup_database_before_schema_change("loan-group-unique")
            self._rebuild_credit_loan_period(connection, has_group_code=True)

    def _rebuild_credit_loan_period(self, connection: sqlite3.Connection, *, has_group_code: bool) -> None:
        temp_table = "credit_loan_period_before_group_code"
        connection.execute(f"DROP TABLE IF EXISTS {temp_table}")
        connection.execute(f"ALTER TABLE credit_loan_period RENAME TO {temp_table}")
        connection.execute(
            """
            CREATE TABLE credit_loan_period (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period TEXT NOT NULL,
                customer_id INTEGER NOT NULL,
                branch_code TEXT NOT NULL,
                loan_key TEXT NOT NULL,
                account_number TEXT NOT NULL DEFAULT '',
                approval_sequence TEXT NOT NULL DEFAULT '',
                outstanding_balance REAL NOT NULL DEFAULT 0,
                customer_type_code TEXT NOT NULL DEFAULT '',
                debt_group_code TEXT NOT NULL DEFAULT 'UNKNOWN',
                secured_percent REAL,
                industry_code TEXT NOT NULL DEFAULT '',
                officer_code TEXT NOT NULL DEFAULT '',
                officer_name TEXT NOT NULL DEFAULT '',
                term_category TEXT NOT NULL DEFAULT 'UNKNOWN',
                group_code TEXT NULL,
                source_run_id INTEGER NOT NULL,
                source_row_count INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES credit_customer_master(id)
                    ON DELETE RESTRICT,
                FOREIGN KEY(source_run_id) REFERENCES credit_import_runs(id)
                    ON DELETE CASCADE
            )
            """
        )
        group_expression = "group_code" if has_group_code else "NULL"
        connection.execute(
            f"""
            INSERT INTO credit_loan_period(
                id, period, customer_id, branch_code, loan_key, account_number,
                approval_sequence, outstanding_balance, customer_type_code,
                debt_group_code, secured_percent, industry_code, officer_code,
                officer_name, term_category, group_code, source_run_id,
                source_row_count, created_at
            )
            SELECT
                id, period, customer_id, branch_code, loan_key, account_number,
                approval_sequence, outstanding_balance, customer_type_code,
                debt_group_code, secured_percent, industry_code, officer_code,
                officer_name, term_category, {group_expression}, source_run_id,
                source_row_count, created_at
            FROM {temp_table}
            """
        )
        connection.execute(f"DROP TABLE {temp_table}")

    def _ensure_credit_loan_indexes(self, connection: sqlite3.Connection) -> None:
        connection.executescript(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_credit_loan_period_unique_period_branch_loan_group
                ON credit_loan_period(period, branch_code, loan_key, COALESCE(group_code, ''));
            CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_branch
                ON credit_loan_period(period, branch_code);
            CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_customer
                ON credit_loan_period(period, customer_id);
            CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_debt
                ON credit_loan_period(period, debt_group_code);
            CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_term
                ON credit_loan_period(period, term_category);
            CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_type
                ON credit_loan_period(period, customer_type_code);
            CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_industry
                ON credit_loan_period(period, industry_code);
            CREATE INDEX IF NOT EXISTS idx_credit_loan_period_period_group_customer
                ON credit_loan_period(period, group_code, customer_id);
            """
        )

    def _backup_database_before_schema_change(self, suffix: str) -> Path | None:
        if not self.database_path.is_file():
            return None
        backup_dir = self.database_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{self.database_path.stem}-before-{suffix}-{stamp}.db"
        counter = 1
        while backup_path.exists():
            counter += 1
            backup_path = backup_dir / f"{self.database_path.stem}-before-{suffix}-{stamp}_{counter}.db"
        shutil.copy2(self.database_path, backup_path)
        return backup_path

    def ln01_period_exists(self, period: str) -> bool:
        clean_period = _require_period(period)
        with closing(self.connect()) as connection:
            return self._ln01_period_exists(connection, clean_period)

    def ln01_period_import_metadata(self, period: str) -> dict[str, object]:
        clean_period = _require_period(period)
        with closing(self.connect()) as connection:
            row = connection.execute(
                """
                SELECT id, period, source_file_name, source_sha256, source_file_size,
                       source_row_count, created_at, imported_by, message
                FROM credit_import_runs
                WHERE period = ? AND source_type = ? AND status = 'success'
                ORDER BY created_at DESC, id DESC
                LIMIT 1
                """,
                (clean_period, CREDIT_SOURCE_LN01),
            ).fetchone()
        if row is None:
            return {}
        return {
            "id": row["id"],
            "period": row["period"],
            "source_file_name": row["source_file_name"],
            "source_sha256": row["source_sha256"],
            "source_file_size": row["source_file_size"],
            "source_row_count": row["source_row_count"],
            "created_at": row["created_at"],
            "imported_by": row["imported_by"],
            "message": row["message"],
        }

    def source_sha_exists(
        self,
        source_sha256: str,
        *,
        source_type: str = CREDIT_SOURCE_LN01,
    ) -> bool:
        needle = str(source_sha256 or "").strip().lower()
        if not needle:
            return False
        with closing(self.connect()) as connection:
            return self._source_sha_exists(connection, source_type, needle)

    def write_ln01_period(
        self,
        connection: sqlite3.Connection,
        *,
        period: str,
        source_file_name: str,
        source_sha256: str,
        source_file_size: int,
        source_row_count: int,
        rows: list[NormalizedLn01Row],
        imported_by: str,
        overwrite_period: bool = False,
        duration_ms: int = 0,
        has_group_code_column: bool | None = None,
        message: str = "",
    ) -> dict[str, object]:
        clean_period = _require_period(period)
        source_hash = str(source_sha256 or "").strip().lower()
        if self._ln01_period_exists(connection, clean_period) and not overwrite_period:
            raise SummaryError(f"Kỳ {clean_period} đã có dữ liệu báo cáo.")
        if self._source_sha_exists(connection, CREDIT_SOURCE_LN01, source_hash) and not overwrite_period:
            raise SummaryError("File LN01 này đã được nhập vào dữ liệu báo cáo trước đó.")
        if overwrite_period:
            self._delete_ln01_period(connection, clean_period)
        now = now_text()
        run_id = self._insert_import_run(
            connection,
            period=clean_period,
            source_type=CREDIT_SOURCE_LN01,
            source_file_name=source_file_name,
            source_sha256=source_hash,
            source_file_size=source_file_size,
            source_row_count=source_row_count,
            imported_by=imported_by,
            message=message or f"Import LN01 {source_file_name}",
            duration_ms=duration_ms,
        )
        self._insert_import_file(
            connection,
            run_id=run_id,
            period=clean_period,
            source_type=CREDIT_SOURCE_LN01,
            file_name=source_file_name,
            sha256=source_hash,
            file_size=source_file_size,
            row_count=source_row_count,
        )
        grouped: dict[tuple[str, str], dict[str, object]] = {}
        skipped = 0
        for row in rows:
            branch_code = str(row.branch_code or "").strip()
            customer_sequence = str(row.customer_sequence or "").strip()
            if not branch_code or not customer_sequence:
                skipped += 1
                continue
            customer_id = self._upsert_customer(connection, row)
            loan_key = _loan_key(customer_id, row.account_number, row.approval_sequence)
            group_code = str(row.group_code or "").strip() or None
            item = grouped.setdefault(
                (loan_key, group_code or ""),
                {
                    "period": clean_period,
                    "customer_id": customer_id,
                    "branch_code": branch_code,
                    "loan_key": loan_key,
                    "account_number": row.account_number,
                    "approval_sequence": row.approval_sequence,
                    "outstanding_balance": 0.0,
                    "customer_type_code": row.customer_type_code,
                    "debt_group_code": _normalize_debt_code(row.debt_group_code),
                    "secured_percent": row.secured_percent,
                    "industry_code": row.industry_code,
                    "officer_code": row.officer_code,
                    "officer_name": row.officer_name,
                    "term_category": classify_term_from_account(row.account_number),
                    "group_code": group_code,
                    "source_run_id": run_id,
                    "source_row_count": 0,
                    "created_at": now,
                },
            )
            item["outstanding_balance"] = float(item["outstanding_balance"] or 0) + float(row.outstanding_balance or 0)
            item["source_row_count"] = int(item["source_row_count"] or 0) + int(
                getattr(row, "source_row_count", 1) or 1
            )
        connection.executemany(
            """
            INSERT INTO credit_loan_period(
                period, customer_id, branch_code, loan_key, account_number,
                approval_sequence, outstanding_balance, customer_type_code,
                debt_group_code, secured_percent, industry_code, officer_code,
                officer_name, term_category, group_code, source_run_id, source_row_count, created_at
            )
            VALUES (
                :period, :customer_id, :branch_code, :loan_key, :account_number,
                :approval_sequence, :outstanding_balance, :customer_type_code,
                :debt_group_code, :secured_percent, :industry_code, :officer_code,
                :officer_name, :term_category, :group_code, :source_run_id, :source_row_count, :created_at
            )
            """,
            list(grouped.values()),
        )
        ln01_total = sum(float(item["outstanding_balance"] or 0) for item in grouped.values())
        customer_count = len(
            {
                int(item["customer_id"])
                for item in grouped.values()
                if float(item["outstanding_balance"] or 0) > 0
            }
        )
        group_stats = self._ln01_group_code_stats(
            connection,
            list(grouped.values()),
            source_rows=rows,
            has_group_code_column=has_group_code_column,
        )
        connection.execute(
            """
            UPDATE credit_import_runs
            SET normalized_loan_count = ?,
                customer_count = ?,
                ln01_total_balance = ?,
                warning_count = ?,
                group_code_stats_json = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (len(grouped), customer_count, ln01_total, skipped + int(group_stats["unknown_group_code_count"]), json.dumps(group_stats, ensure_ascii=False, sort_keys=True), now, run_id),
        )
        self._validate_ln01_period(connection, clean_period, expected_balance=ln01_total)
        return {
            "run_id": run_id,
            "normalized_loan_count": len(grouped),
            "customer_count": customer_count,
            "ln01_total_balance": ln01_total,
            "skipped_row_count": skipped,
            **group_stats,
        }

    def replace_ln01_period(
        self,
        connection: sqlite3.Connection,
        **values: object,
    ) -> dict[str, object]:
        return self.write_ln01_period(connection, overwrite_period=True, **values)

    def _ln01_group_code_stats(
        self,
        connection: sqlite3.Connection,
        grouped_rows: list[dict[str, object]],
        *,
        source_rows: list[NormalizedLn01Row],
        has_group_code_column: bool | None = None,
    ) -> dict[str, object]:
        _ = connection
        declared_codes = self._declared_credit_group_codes()
        source_rows_with_group = sum(1 for row in source_rows if str(row.group_code or "").strip())
        source_rows_without_group = max(0, len(source_rows) - source_rows_with_group)
        loans_with_group = [row for row in grouped_rows if str(row.get("group_code") or "").strip()]
        loans_without_group = [row for row in grouped_rows if not str(row.get("group_code") or "").strip()]
        unknown_codes = sorted(
            {
                str(row.get("group_code") or "").strip()
                for row in loans_with_group
                if str(row.get("group_code") or "").strip() not in declared_codes
            }
        )
        unknown_balance = sum(
            float(row.get("outstanding_balance") or 0)
            for row in loans_with_group
            if str(row.get("group_code") or "").strip() in set(unknown_codes)
        )
        loan_group_pairs: dict[tuple[object, object, object, object], set[str]] = defaultdict(set)
        for row in grouped_rows:
            group_code = str(row.get("group_code") or "").strip()
            if not group_code:
                continue
            loan_group_pairs[
                (
                    row.get("period"),
                    row.get("branch_code"),
                    row.get("customer_id"),
                    row.get("loan_key"),
                )
            ].add(group_code)
        conflict_pairs = {key: codes for key, codes in loan_group_pairs.items() if len(codes) > 1}
        conflict_balance = sum(
            float(row.get("outstanding_balance") or 0)
            for row in grouped_rows
            if (
                row.get("period"),
                row.get("branch_code"),
                row.get("customer_id"),
                row.get("loan_key"),
            )
            in conflict_pairs
        )
        return {
            "version": 1,
            "has_group_code_column": bool(has_group_code_column) if has_group_code_column is not None else source_rows_with_group > 0,
            "source_rows_with_group_code": source_rows_with_group,
            "source_rows_without_group_code": source_rows_without_group,
            "normalized_loans_with_group_code": len(loans_with_group),
            "normalized_loans_without_group_code": len(loans_without_group),
            "balance_with_group_code": sum(float(row.get("outstanding_balance") or 0) for row in loans_with_group),
            "balance_without_group_code": sum(float(row.get("outstanding_balance") or 0) for row in loans_without_group),
            "matched_group_code_count": len({str(row.get("group_code") or "").strip() for row in loans_with_group if str(row.get("group_code") or "").strip() in declared_codes}),
            "unknown_group_code_count": len(unknown_codes),
            "unknown_group_codes": unknown_codes[:50],
            "unknown_group_balance": unknown_balance,
            "loan_group_conflict_count": len(conflict_pairs),
            "loan_group_conflict_balance": conflict_balance,
            "credit_card_group_code_scope": "excluded_without_group_code",
        }

    def _declared_credit_group_codes(self) -> set[str]:
        path = self.group_directory_database_path()
        if not path.is_file():
            return set()
        try:
            with closing(sqlite3.connect(path, timeout=15)) as connection:
                table = connection.execute(
                    """
                    SELECT name
                    FROM sqlite_master
                    WHERE type = 'table' AND name = 'credit_groups'
                    """
                ).fetchone()
                if table is None:
                    return set()
                return {
                    str(row[0] or "").strip()
                    for row in connection.execute("SELECT ma_to FROM credit_groups").fetchall()
                    if str(row[0] or "").strip()
                }
        except sqlite3.Error:
            return set()

    def group_directory_database_path(self) -> Path:
        if self.main_database_path is not None:
            path = Path(self.main_database_path)
            if path.name.casefold() == CREDIT_DATABASE_NAME.casefold():
                return path.parent / MAIN_DATABASE_NAME
            return path
        return application_root() / "data" / MAIN_DATABASE_NAME

    def log_ln01_import_action(
        self,
        connection: sqlite3.Connection,
        *,
        action: str,
        period: str,
        source_sha256: str,
        source_file_name: str,
        batch_id: str,
        created_by: str = "",
        result: str = "success",
    ) -> None:
        detail = json.dumps(
            {
                "period": period,
                "source_sha256": source_sha256,
                "source_file_name": source_file_name,
                "batch_id": batch_id,
                "result": result,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        self._log_action(
            connection,
            action=action,
            target_id=period,
            detail=detail,
            created_by=created_by,
        )

    def save_credit_card_projection(
        self,
        *,
        period: str,
        file_name: str,
        source_sha256: str,
        source_file_size: int,
        source_row_count: int,
        rows: list[NormalizedLoanRow],
        imported_by: str = "",
        replace_period: bool = False,
        duration_ms: int = 0,
    ) -> dict[str, object]:
        clean_period = _require_period(period)
        source_hash = str(source_sha256 or "").strip().lower()
        with closing(self.connect()) as connection:
            connection.execute("BEGIN IMMEDIATE")
            try:
                if replace_period:
                    self._delete_card_period(connection, clean_period)
                elif self._source_sha_exists(connection, CREDIT_SOURCE_NIM_DN_CARD, source_hash):
                    raise SummaryError("File NIM Dư nợ này đã được chiếu sang Credit.db trước đó.")
                run_id = self._insert_import_run(
                    connection,
                    period=clean_period,
                    source_type=CREDIT_SOURCE_NIM_DN_CARD,
                    source_file_name=file_name,
                    source_sha256=source_hash,
                    source_file_size=source_file_size,
                    source_row_count=source_row_count,
                    imported_by=imported_by,
                    message=f"Projection DN15 từ {file_name}",
                    duration_ms=duration_ms,
                )
                self._insert_import_file(
                    connection,
                    run_id=run_id,
                    period=clean_period,
                    source_type=CREDIT_SOURCE_NIM_DN_CARD,
                    file_name=file_name,
                    sha256=source_hash,
                    file_size=source_file_size,
                    row_count=source_row_count,
                )
                card_rows = [row for row in rows if str(row.ftp_code or "").strip().upper() == "DN15"]
                grouped = self._aggregate_card_rows(connection, run_id, clean_period, card_rows)
                connection.executemany(
                    """
                    INSERT INTO credit_card_period(
                        period, customer_id, branch_code, customer_code, customer_name,
                        customer_type, debt_group_code, officer_code, officer_name,
                        balance, source_run_id, source_row_count, created_at
                    )
                    VALUES (
                        :period, :customer_id, :branch_code, :customer_code, :customer_name,
                        :customer_type, :debt_group_code, :officer_code, :officer_name,
                        :balance, :source_run_id, :source_row_count, :created_at
                    )
                    ON CONFLICT(
                        period, branch_code, customer_id, customer_type,
                        debt_group_code, officer_code, officer_name
                    ) DO UPDATE SET
                        balance = credit_card_period.balance + excluded.balance,
                        source_row_count = credit_card_period.source_row_count + excluded.source_row_count
                    """,
                    grouped,
                )
                card_balance = sum(float(item["balance"] or 0) for item in grouped)
                customer_count = len({int(item["customer_id"]) for item in grouped if float(item["balance"] or 0) > 0})
                connection.execute(
                    """
                    UPDATE credit_import_runs
                    SET normalized_loan_count = ?,
                        customer_count = ?,
                        card_balance = ?,
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (len(grouped), customer_count, card_balance, now_text(), run_id),
                )
                connection.commit()
            except Exception:
                connection.rollback()
                raise
        return {
            "run_id": run_id,
            "card_row_count": len(grouped),
            "card_balance": card_balance,
            "customer_count": customer_count,
        }

    def delete_credit_card_period(self, period: str) -> None:
        clean_period = _require_period(period)
        with closing(self.connect()) as connection:
            with connection:
                self._delete_card_period(connection, clean_period)

    def available_periods(self) -> list[str]:
        with closing(self.connect()) as connection:
            rows = connection.execute(
                """
                SELECT period FROM credit_loan_period
                UNION
                SELECT period FROM credit_card_period
                ORDER BY period
                """
            ).fetchall()
        return [str(row[0]) for row in rows]

    def filter_values(self) -> dict[str, list[tuple[str, str]]]:
        periods = self.available_periods()
        rows = self._combined_rows(CreditReportFilters())
        branches = sorted({str(row["branch_code"]) for row in rows if row["branch_code"]})
        officers = sorted(
            {
                _officer_display(str(row["officer_code"]), str(row["officer_name"]))
                for row in rows
                if row["officer_code"] or row["officer_name"]
            }
        )
        return {
            "periods": [(period, period) for period in periods],
            "branches": [(branch, branch) for branch in branches],
            "customer_types": [(CUSTOMER_TYPE_LABELS[key], key) for key in (CUSTOMER_TYPE_PERSONAL, CUSTOMER_TYPE_LEGAL, CUSTOMER_TYPE_UNKNOWN)],
            "debt_groups": [(DEBT_BUCKET_LABELS[key], key) for key in ("GROUP_1", "GROUP_2", "BAD_DEBT", "UNKNOWN")],
            "terms": [(TERM_LABELS[key], key) for key in (TERM_SHORT, TERM_MEDIUM, TERM_LONG, TERM_UNKNOWN)],
            "officers": [(officer, officer) for officer in officers],
        }

    def overall_summary(self, filters: CreditReportFilters) -> dict[str, object]:
        rows = self._combined_rows(filters)
        totals = _total_by_source(rows)
        customer_totals = _customer_totals(rows)
        customer_count = sum(1 for value in customer_totals.values() if value > 0)
        by_type = _balance_by(rows, "customer_type")
        by_debt = _balance_by_debt_bucket(rows)
        by_term = _term_rows_from_combined(rows)
        decree55_balance = sum(
            float(row["balance"] or 0)
            for row in rows
            if row["source"] == CREDIT_SOURCE_LN01 and _is_secured_zero(row.get("secured_percent"))
        )
        status = self.card_data_status(filters.period)
        total_balance = totals["ln01"] + totals["card"]
        return {
            "period": filters.period,
            "ln01_total_balance": totals["ln01"],
            "credit_card_balance": totals["card"],
            "total_balance": total_balance,
            "customer_count": customer_count,
            "personal_balance": by_type[CUSTOMER_TYPE_PERSONAL],
            "legal_balance": by_type[CUSTOMER_TYPE_LEGAL],
            "unknown_customer_type_balance": by_type[CUSTOMER_TYPE_UNKNOWN],
            "short_term_balance": by_term[TERM_SHORT]["total_balance"],
            "medium_term_balance": by_term[TERM_MEDIUM]["total_balance"],
            "long_term_balance": by_term[TERM_LONG]["total_balance"],
            "unknown_term_balance": by_term[TERM_UNKNOWN]["total_balance"],
            "debt_group_1_balance": by_debt["GROUP_1"],
            "debt_group_2_balance": by_debt["GROUP_2"],
            "bad_debt_balance": by_debt["BAD_DEBT"],
            "unknown_debt_group_balance": by_debt["UNKNOWN"],
            "decree55_balance": decree55_balance,
            "card_data_status": status["status"],
            "card_data_message": status["message"],
        }

    def overview_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        summary = self.overall_summary(filters)
        return [
            {"Chỉ tiêu": "Tổng dư nợ", "Giá trị": summary["total_balance"], "Ghi chú": "LN01 + thẻ DN15 cùng kỳ"},
            {"Chỉ tiêu": "Dư nợ LN01", "Giá trị": summary["ln01_total_balance"], "Ghi chú": ""},
            {"Chỉ tiêu": "Dư nợ thẻ tín dụng", "Giá trị": summary["credit_card_balance"], "Ghi chú": summary["card_data_message"]},
            {"Chỉ tiêu": "Tổng khách hàng còn dư nợ", "Giá trị": summary["customer_count"], "Ghi chú": "Distinct khách hàng có tổng dư nợ > 0"},
            {"Chỉ tiêu": "Dư nợ ngắn hạn", "Giá trị": summary["short_term_balance"], "Ghi chú": "Bao gồm thẻ DN15"},
            {"Chỉ tiêu": "Dư nợ trung hạn", "Giá trị": summary["medium_term_balance"], "Ghi chú": ""},
            {"Chỉ tiêu": "Dư nợ dài hạn", "Giá trị": summary["long_term_balance"], "Ghi chú": ""},
            {"Chỉ tiêu": "Dư nợ cá nhân", "Giá trị": summary["personal_balance"], "Ghi chú": ""},
            {"Chỉ tiêu": "Dư nợ pháp nhân", "Giá trị": summary["legal_balance"], "Ghi chú": ""},
            {"Chỉ tiêu": "Nợ cần chú ý", "Giá trị": summary["debt_group_2_balance"], "Ghi chú": "Nhóm 02"},
            {"Chỉ tiêu": "Nợ xấu", "Giá trị": summary["bad_debt_balance"], "Ghi chú": "Nhóm 03-05"},
            {"Chỉ tiêu": "Dư nợ Nghị định 55", "Giá trị": summary["decree55_balance"], "Ghi chú": "LN01 SECURED_PERCENT = 0"},
        ]

    def term_structure_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        rows = self._combined_rows(filters)
        term_data = _term_rows_from_combined(rows)
        total = sum(item["total_balance"] for item in term_data.values())
        output: list[dict[str, object]] = []
        for key in (TERM_SHORT, TERM_MEDIUM, TERM_LONG, TERM_UNKNOWN):
            item = term_data[key]
            output.append(
                {
                    "Loại thời hạn": TERM_LABELS[key],
                    "Dư nợ LN01": item["ln01_balance"],
                    "Dư nợ thẻ": item["card_balance"],
                    "Tổng dư nợ": item["total_balance"],
                    "Tỷ trọng": _ratio(item["total_balance"], total),
                    "Số khách hàng": item["customer_count"],
                }
            )
        return output

    def customer_type_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        rows = self._combined_rows(filters)
        total = sum(float(row["balance"] or 0) for row in rows)
        output = []
        for key in (CUSTOMER_TYPE_PERSONAL, CUSTOMER_TYPE_LEGAL, CUSTOMER_TYPE_UNKNOWN):
            scoped = [row for row in rows if row["customer_type"] == key]
            balance = sum(float(row["balance"] or 0) for row in scoped)
            if key == CUSTOMER_TYPE_UNKNOWN and balance == 0:
                continue
            customer_count = sum(1 for value in _customer_totals(scoped).values() if value > 0)
            output.append(
                {
                    "Loại khách hàng": CUSTOMER_TYPE_LABELS[key],
                    "Dư nợ LN01": sum(float(row["balance"] or 0) for row in scoped if row["source"] == CREDIT_SOURCE_LN01),
                    "Dư nợ thẻ": sum(float(row["balance"] or 0) for row in scoped if row["source"] == CREDIT_SOURCE_NIM_DN_CARD),
                    "Tổng dư nợ": balance,
                    "Tỷ trọng": _ratio(balance, total),
                    "Số khách hàng": customer_count,
                    "Số món": len({row["loan_key"] for row in scoped if row["source"] == CREDIT_SOURCE_LN01}),
                }
            )
        return output

    def debt_group_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        rows = self._combined_rows(filters)
        total = sum(float(row["balance"] or 0) for row in rows)
        output = []
        for key in ("GROUP_1", "GROUP_2", "BAD_DEBT", "UNKNOWN"):
            scoped = [row for row in rows if _debt_bucket(row["debt_group_code"]) == key]
            balance = sum(float(row["balance"] or 0) for row in scoped)
            if key == "UNKNOWN" and balance == 0:
                continue
            customer_totals = _customer_totals(scoped)
            output.append(
                {
                    "Nhóm nợ": DEBT_BUCKET_LABELS[key],
                    "Tổng dư nợ": balance,
                    "Tỷ lệ": _ratio(balance, total),
                    "Số món": len({row["loan_key"] for row in scoped if row["source"] == CREDIT_SOURCE_LN01}),
                    "Số lượng khách hàng": sum(1 for value in customer_totals.values() if value > 0),
                    "Số khách hàng cá nhân": _customer_count_by_type(scoped, CUSTOMER_TYPE_PERSONAL),
                    "Số khách hàng pháp nhân": _customer_count_by_type(scoped, CUSTOMER_TYPE_LEGAL),
                    "Ghi chú": "Số món thẻ chưa xác định" if any(row["source"] == CREDIT_SOURCE_NIM_DN_CARD for row in scoped) else "",
                }
            )
        return output

    def decree55_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        ln01_rows = [row for row in self._combined_rows(filters) if row["source"] == CREDIT_SOURCE_LN01]
        rows = [
            row
            for row in ln01_rows
            if _is_secured_zero(row.get("secured_percent"))
        ]
        balance = sum(float(row["balance"] or 0) for row in rows)
        ln01_total = sum(float(row["balance"] or 0) for row in ln01_rows)
        return [
            {
                "Chỉ tiêu": "Cho vay theo Nghị định 55 không TSĐB",
                "Tổng dư nợ": balance,
                "Tỷ trọng trên dư nợ LN01": _ratio(balance, ln01_total),
                "Số món": len({row["loan_key"] for row in rows}),
                "Số lượng khách hàng": sum(1 for value in _customer_totals(rows).values() if value > 0),
                "Số khách hàng cá nhân": _customer_count_by_type(rows, CUSTOMER_TYPE_PERSONAL),
                "Số khách hàng pháp nhân": _customer_count_by_type(rows, CUSTOMER_TYPE_LEGAL),
                "Ghi chú": "Chỉ tiêu Nghị định 55 được xác định từ dữ liệu LN01 có SECURED_PERCENT = 0.",
            }
        ]

    def industry_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        rows = [row for row in self._combined_rows(filters) if row["source"] == CREDIT_SOURCE_LN01]
        total = sum(float(row["balance"] or 0) for row in rows)
        output = []
        for key in (INDUSTRY_HIGH_TECH, INDUSTRY_WHOLESALE_RETAIL, INDUSTRY_REAL_ESTATE, INDUSTRY_OTHER):
            scoped = [row for row in rows if classify_industry(row["industry_code"]) == key]
            balance = sum(float(row["balance"] or 0) for row in scoped)
            output.append(
                {
                    "Nhóm ngành": INDUSTRY_LABELS[key],
                    "Tổng dư nợ": balance,
                    "Tỷ trọng LN01": _ratio(balance, total),
                    "Số món": len({row["loan_key"] for row in scoped}),
                    "Số lượng khách hàng": sum(1 for value in _customer_totals(scoped).values() if value > 0),
                    "Số khách hàng cá nhân": _customer_count_by_type(scoped, CUSTOMER_TYPE_PERSONAL),
                    "Số khách hàng pháp nhân": _customer_count_by_type(scoped, CUSTOMER_TYPE_LEGAL),
                    "Ghi chú": "Phân loại ngành kinh tế sử dụng dữ liệu LN01.",
                }
            )
        return output

    def period_comparison_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        from_period = filters.from_period or filters.period
        to_period = filters.to_period or filters.period
        if not from_period or not to_period:
            return []
        from_filters = replace(filters, period=from_period)
        to_filters = replace(filters, period=to_period)
        from_summary = self.overall_summary(from_filters)
        to_summary = self.overall_summary(to_filters)
        metrics = (
            ("Tổng dư nợ", "total_balance"),
            ("Tổng khách hàng", "customer_count"),
            ("Dư nợ ngắn hạn", "short_term_balance"),
            ("Dư nợ trung hạn", "medium_term_balance"),
            ("Dư nợ dài hạn", "long_term_balance"),
            ("Dư nợ cá nhân", "personal_balance"),
            ("Dư nợ pháp nhân", "legal_balance"),
            ("Nợ nhóm 1", "debt_group_1_balance"),
            ("Nợ nhóm 2", "debt_group_2_balance"),
            ("Nợ xấu", "bad_debt_balance"),
            ("Cho vay Nghị định 55", "decree55_balance"),
        )
        industry_from = {row["Nhóm ngành"]: row["Tổng dư nợ"] for row in self.industry_rows(from_filters)}
        industry_to = {row["Nhóm ngành"]: row["Tổng dư nợ"] for row in self.industry_rows(to_filters)}
        output = [_comparison_row(label, from_summary[key], to_summary[key]) for label, key in metrics]
        for label in (
            INDUSTRY_LABELS[INDUSTRY_HIGH_TECH],
            INDUSTRY_LABELS[INDUSTRY_WHOLESALE_RETAIL],
            INDUSTRY_LABELS[INDUSTRY_REAL_ESTATE],
        ):
            output.append(_comparison_row(label, industry_from.get(label, 0), industry_to.get(label, 0)))
        return output

    def import_history_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        clauses: list[str] = []
        params: list[object] = []
        if filters.period:
            clauses.append("period = ?")
            params.append(filters.period)
        where = " WHERE " + " AND ".join(clauses) if clauses else ""
        with closing(self.connect()) as connection:
            rows = connection.execute(
                f"""
                SELECT *
                FROM credit_import_runs
                {where}
                ORDER BY created_at DESC, id DESC
                LIMIT 500
                """,
                params,
            ).fetchall()
        return [
            {
                "Kỳ": row["period"],
                "Tên file nguồn": row["source_file_name"],
                "SHA-256": row["source_sha256"],
                "Thời gian import": row["created_at"],
                "Số dòng nguồn": row["source_row_count"],
                "Số món sau chuẩn hóa": row["normalized_loan_count"],
                "Số món có GRPNO": _import_group_stat(row, "normalized_loans_with_group_code"),
                "Số khách hàng": row["customer_count"],
                "Tổng dư nợ LN01": row["ln01_total_balance"],
                "Dư nợ có GRPNO": _import_group_stat(row, "balance_with_group_code"),
                "Mã tổ chưa khai báo": _import_group_stat(row, "unknown_group_code_count"),
                "Dư nợ thẻ": row["card_balance"],
                "Trạng thái": row["status"],
                "Số cảnh báo": row["warning_count"],
                "Người thực hiện": row["imported_by"],
            }
            for row in rows
        ]

    def customer_type_rule_rows(self) -> list[dict[str, object]]:
        with closing(self.connect()) as connection:
            rows = connection.execute(
                """
                SELECT code, label, is_personal, is_enabled, updated_at
                FROM credit_customer_type_rules
                ORDER BY code COLLATE NOCASE
                """
            ).fetchall()
        return [
            {
                "Mã": row["code"],
                "Tên": row["label"],
                "Loại": "Cá nhân" if int(row["is_personal"] or 0) else "Pháp nhân",
                "Đang dùng": "Có" if int(row["is_enabled"] or 0) else "Không",
                "Cập nhật": row["updated_at"],
            }
            for row in rows
        ]

    def save_personal_type_rule(self, code: str, *, label: str = "", enabled: bool = True) -> None:
        clean_code = _normalize_code(code)
        if not clean_code:
            raise SummaryError("Mã khách hàng cá nhân không được để trống.")
        now = now_text()
        with closing(self.connect()) as connection:
            with connection:
                connection.execute(
                    """
                    INSERT INTO credit_customer_type_rules(
                        code, label, is_personal, is_enabled, created_at, updated_at
                    )
                    VALUES (?, ?, 1, ?, ?, ?)
                    ON CONFLICT(code) DO UPDATE SET
                        label = excluded.label,
                        is_personal = 1,
                        is_enabled = excluded.is_enabled,
                        updated_at = excluded.updated_at
                    """,
                    (clean_code, label or clean_code, 1 if enabled else 0, now, now),
                )

    def delete_personal_type_rule(self, code: str) -> None:
        clean_code = _normalize_code(code)
        if clean_code in {"100", "570"}:
            raise SummaryError("Mã mặc định chỉ có thể tắt hoặc khôi phục mặc định.")
        with closing(self.connect()) as connection:
            with connection:
                connection.execute("DELETE FROM credit_customer_type_rules WHERE code = ?", (clean_code,))

    def restore_default_type_rules(self) -> None:
        with closing(self.connect()) as connection:
            with connection:
                connection.execute("DELETE FROM credit_customer_type_rules")
                self._seed_default_customer_type_rules(connection)

    def delete_report_period(self, period: str, *, created_by: str = "") -> dict[str, object]:
        clean_period = _require_period(period)
        with closing(self.connect()) as connection:
            connection.execute("BEGIN IMMEDIATE")
            try:
                loan_rows = int(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = ?", (clean_period,)).fetchone()[0] or 0)
                card_rows = int(connection.execute("SELECT COUNT(*) FROM credit_card_period WHERE period = ?", (clean_period,)).fetchone()[0] or 0)
                run_rows = int(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = ?", (clean_period,)).fetchone()[0] or 0)
                file_rows = int(connection.execute("SELECT COUNT(*) FROM credit_import_files WHERE period = ?", (clean_period,)).fetchone()[0] or 0)
                before_period_count = len(
                    {
                        str(row[0])
                        for row in connection.execute(
                            """
                            SELECT period FROM credit_loan_period
                            UNION
                            SELECT period FROM credit_card_period
                            """
                        ).fetchall()
                    }
                )
                connection.execute("DELETE FROM credit_loan_period WHERE period = ?", (clean_period,))
                connection.execute("DELETE FROM credit_card_period WHERE period = ?", (clean_period,))
                connection.execute("DELETE FROM credit_import_files WHERE period = ?", (clean_period,))
                connection.execute("DELETE FROM credit_import_runs WHERE period = ?", (clean_period,))
                orphan_customers = self._delete_orphan_customers(connection)
                page_size = int(connection.execute("PRAGMA page_size").fetchone()[0] or 0)
                freelist_count = int(connection.execute("PRAGMA freelist_count").fetchone()[0] or 0)
                self._log_action(
                    connection,
                    action="delete_period",
                    target_id=clean_period,
                    detail=json.dumps(
                        {
                            "loan_rows": loan_rows,
                            "card_rows": card_rows,
                            "runs": run_rows,
                            "files": file_rows,
                            "orphan_customers": orphan_customers,
                            "freelist_count": freelist_count,
                            "reclaimable_bytes": freelist_count * page_size,
                        },
                        ensure_ascii=False,
                    ),
                    created_by=created_by,
                )
                connection.commit()
            except Exception:
                connection.rollback()
                raise
        after = self.diagnostics()
        return {
            "period": clean_period,
            "loan_rows": loan_rows,
            "card_rows": card_rows,
            "runs": run_rows,
            "files": file_rows,
            "orphan_customers": orphan_customers,
            "last_period_deleted": before_period_count == 1,
            "size_bytes": after["db_size_bytes"],
            "freelist_count": after["freelist_count"],
            "reclaimable_bytes": after["reclaimable_bytes"],
        }

    def card_data_status(self, period: str) -> dict[str, str]:
        clean_period = str(period or "").strip()
        if not clean_period:
            return {"status": "missing", "message": "Chưa chọn kỳ báo cáo."}
        with closing(self.connect()) as connection:
            run = connection.execute(
                """
                SELECT COUNT(*) FROM credit_import_runs
                WHERE period = ? AND source_type = ? AND status = 'success'
                """,
                (clean_period, CREDIT_SOURCE_NIM_DN_CARD),
            ).fetchone()[0]
            balance = connection.execute(
                "SELECT COALESCE(SUM(balance), 0) FROM credit_card_period WHERE period = ?",
                (clean_period,),
            ).fetchone()[0]
        if int(run or 0) <= 0:
            return {"status": "missing", "message": f"Chưa có dữ liệu thẻ tín dụng kỳ {clean_period}."}
        return {"status": "ok", "message": "Dư nợ thẻ tín dụng: 0" if float(balance or 0) == 0 else "Đã cộng dư nợ thẻ DN15 từ NIM Dư nợ."}

    def status(self) -> dict[str, object]:
        diagnostics = self.diagnostics(include_dbstat=False)
        return {
            "database_path": self.database_path,
            "size_bytes": diagnostics["db_size_bytes"],
            "integrity": diagnostics["integrity_check"],
            "period_count": diagnostics["period_count"],
            "loan_rows": diagnostics["tables"].get("credit_loan_period", 0),
            "card_rows": diagnostics["tables"].get("credit_card_period", 0),
            "customer_rows": diagnostics["tables"].get("credit_customer_master", 0),
            "import_rows": diagnostics["tables"].get("credit_import_runs", 0),
            "import_file_rows": diagnostics["tables"].get("credit_import_files", 0),
            "freelist_count": diagnostics["freelist_count"],
            "reclaimable_bytes": diagnostics["reclaimable_bytes"],
        }

    def optimize_database(self, *, vacuum: bool = False) -> dict[str, object]:
        if vacuum:
            return self.compact_database()
        before = self.diagnostics(include_dbstat=False)
        started = perf_counter()
        with closing(self.connect()) as connection:
            checkpoint = connection.execute("PRAGMA wal_checkpoint(TRUNCATE)").fetchone()
            connection.execute("ANALYZE")
            connection.execute("PRAGMA optimize")
            connection.commit()
        after = self.diagnostics(include_dbstat=False)
        return {
            "before_size_bytes": before["db_size_bytes"],
            "after_size_bytes": after["db_size_bytes"],
            "before_wal_size_bytes": before["wal_size_bytes"],
            "after_wal_size_bytes": after["wal_size_bytes"],
            "duration_ms": int((perf_counter() - started) * 1000),
            "vacuum": False,
            "wal_checkpoint": tuple(checkpoint) if checkpoint is not None else (),
        }

    def diagnostics(self, *, include_dbstat: bool = True) -> dict[str, object]:
        db_path = self.database_path
        wal_path = Path(f"{db_path}-wal")
        shm_path = Path(f"{db_path}-shm")
        with closing(self.connect()) as connection:
            tables = [
                str(row[0])
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
                ).fetchall()
            ]
            table_counts = {
                name: int(connection.execute(f'SELECT COUNT(*) FROM "{name}"').fetchone()[0] or 0)
                for name in tables
            }
            indexes = [
                str(row[0])
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'index' AND name NOT LIKE 'sqlite_%' ORDER BY name"
                ).fetchall()
            ]
            periods = {
                str(row[0])
                for row in connection.execute(
                    """
                    SELECT period FROM credit_loan_period
                    UNION
                    SELECT period FROM credit_card_period
                    UNION
                    SELECT period FROM credit_import_runs
                    UNION
                    SELECT period FROM credit_import_files
                    """
                ).fetchall()
                if str(row[0] or "").strip()
            }
            page_size = int(connection.execute("PRAGMA page_size").fetchone()[0] or 0)
            page_count = int(connection.execute("PRAGMA page_count").fetchone()[0] or 0)
            freelist_count = int(connection.execute("PRAGMA freelist_count").fetchone()[0] or 0)
            diagnostics: dict[str, object] = {
                "database_path": db_path,
                "db_size_bytes": db_path.stat().st_size if db_path.is_file() else 0,
                "wal_size_bytes": wal_path.stat().st_size if wal_path.is_file() else 0,
                "shm_size_bytes": shm_path.stat().st_size if shm_path.is_file() else 0,
                "page_size": page_size,
                "page_count": page_count,
                "freelist_count": freelist_count,
                "reclaimable_bytes": page_size * freelist_count,
                "auto_vacuum": int(connection.execute("PRAGMA auto_vacuum").fetchone()[0] or 0),
                "journal_mode": str(connection.execute("PRAGMA journal_mode").fetchone()[0] or ""),
                "integrity_check": str(connection.execute("PRAGMA integrity_check").fetchone()[0] or ""),
                "schema_version": int(connection.execute("PRAGMA schema_version").fetchone()[0] or 0),
                "period_count": len(periods),
                "periods": sorted(periods),
                "tables": table_counts,
                "indexes": indexes,
                "table_bytes": {},
                "index_bytes": {},
                "dbstat_supported": False,
                "dbstat_error": "",
            }
            if include_dbstat:
                try:
                    dbstat_rows = connection.execute(
                        """
                        SELECT name, SUM(pgsize) AS bytes
                        FROM dbstat
                        GROUP BY name
                        ORDER BY bytes DESC
                        """
                    ).fetchall()
                    index_names = set(indexes)
                    diagnostics["dbstat_supported"] = True
                    diagnostics["table_bytes"] = {
                        str(row["name"]): int(row["bytes"] or 0)
                        for row in dbstat_rows
                        if str(row["name"]) not in index_names and str(row["name"]) != "sqlite_schema"
                    }
                    diagnostics["index_bytes"] = {
                        str(row["name"]): int(row["bytes"] or 0)
                        for row in dbstat_rows
                        if str(row["name"]) in index_names
                    }
                except Exception as exc:
                    diagnostics["dbstat_error"] = str(exc)
        return diagnostics

    def cleanup_orphan_customers(self, *, created_by: str = "") -> dict[str, object]:
        with closing(self.connect()) as connection:
            connection.execute("BEGIN IMMEDIATE")
            try:
                deleted = self._delete_orphan_customers(connection)
                self._log_action(
                    connection,
                    action="cleanup_orphan_customers",
                    target_id="credit_customer_master",
                    detail=json.dumps({"deleted_customers": deleted}, ensure_ascii=False),
                    created_by=created_by,
                )
                connection.commit()
            except Exception:
                connection.rollback()
                raise
        return {"deleted_customers": deleted, **self.diagnostics(include_dbstat=False)}

    def compact_database(self, *, backup_directory: Path | None = None, created_by: str = "") -> dict[str, object]:
        backup_root = Path(backup_directory) if backup_directory is not None else self.database_path.parent / "backup"
        backup_root.mkdir(parents=True, exist_ok=True)
        backup_path = backup_root / f"Credit-before-vacuum-{datetime.now():%Y%m%d_%H%M%S}.zip"
        before = self.diagnostics(include_dbstat=False)
        started = perf_counter()
        self.backup_database(backup_path)
        with closing(self.connect()) as connection:
            checkpoint = connection.execute("PRAGMA wal_checkpoint(TRUNCATE)").fetchone()
            connection.execute("VACUUM")
            connection.execute("ANALYZE")
            connection.execute("PRAGMA optimize")
            self._log_action(
                connection,
                action="compact_database",
                target_id=CREDIT_DATABASE_NAME,
                detail=json.dumps({"backup_path": str(backup_path)}, ensure_ascii=False),
                created_by=created_by,
            )
            connection.commit()
        after = self.diagnostics(include_dbstat=False)
        return {
            "backup_path": backup_path,
            "before_size_bytes": before["db_size_bytes"],
            "after_size_bytes": after["db_size_bytes"],
            "before_wal_size_bytes": before["wal_size_bytes"],
            "after_wal_size_bytes": after["wal_size_bytes"],
            "before_shm_size_bytes": before["shm_size_bytes"],
            "after_shm_size_bytes": after["shm_size_bytes"],
            "reduced_bytes": int(before["db_size_bytes"]) - int(after["db_size_bytes"]),
            "duration_ms": int((perf_counter() - started) * 1000),
            "vacuum": True,
            "wal_checkpoint": tuple(checkpoint) if checkpoint is not None else (),
            "integrity_check": after["integrity_check"],
        }

    def backup_database(self, destination: Path) -> Path:
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        if destination.suffix.casefold() == ".zip":
            with tempfile.TemporaryDirectory(prefix="credit-db-backup-") as temporary:
                snapshot = Path(temporary) / CREDIT_DATABASE_NAME
                with closing(self.connect()) as source:
                    with closing(sqlite3.connect(snapshot, timeout=30)) as target:
                        source.backup(target)
                with zipfile.ZipFile(destination, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                    archive.write(snapshot, CREDIT_DATABASE_NAME)
                    archive.writestr(
                        "manifest.json",
                        json.dumps({"format": "agribank-v3-credit-db", "created_at": now_text()}, ensure_ascii=False),
                    )
            return destination
        with closing(self.connect()) as source:
            with closing(sqlite3.connect(destination, timeout=30)) as target:
                source.backup(target)
        return destination

    def restore_database(self, source_path: Path) -> Path:
        source_path = Path(source_path)
        if not source_path.is_file():
            raise SummaryError("Không tìm thấy file khôi phục Credit.db.")
        with tempfile.TemporaryDirectory(prefix="credit-db-restore-") as temporary:
            temporary_root = Path(temporary)
            staged = temporary_root / CREDIT_DATABASE_NAME
            if source_path.suffix.casefold() == ".zip":
                with zipfile.ZipFile(source_path) as archive:
                    if CREDIT_DATABASE_NAME not in archive.namelist():
                        raise SummaryError("Gói sao lưu không chứa Credit.db.")
                    archive.extract(CREDIT_DATABASE_NAME, temporary_root)
            else:
                shutil.copy2(source_path, staged)
            with closing(sqlite3.connect(staged, timeout=30)) as connection:
                result = connection.execute("PRAGMA integrity_check").fetchone()[0]
            if result != "ok":
                raise SummaryError(f"Credit.db khôi phục không hợp lệ: {result}")
            backup = self.database_path.parent / f"Credit-before-restore-{datetime.now():%Y%m%d_%H%M%S}.db"
            if self.database_path.is_file():
                shutil.copy2(self.database_path, backup)
            for suffix in ("-wal", "-shm"):
                Path(f"{self.database_path}{suffix}").unlink(missing_ok=True)
            os.replace(staged, self.database_path)
        self.ensure_schema()
        return self.database_path

    def export_workbook(
        self,
        destination: Path,
        *,
        filters: CreditReportFilters,
        view_mode: str = VIEW_CURRENT_PERIOD,
    ) -> Path:
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        try:
            if view_mode == VIEW_COMPARE_PERIODS:
                from_period = filters.from_period or filters.period
                to_period = filters.to_period or filters.period
                from_snapshot = get_report_snapshot(self, from_period, filters)
                to_snapshot = get_report_snapshot(self, to_period, filters if to_period != from_period else replace(filters, to_period=from_period))
                sheets = (
                    ("SoSanhTongHop", _row_dicts(compare_report_snapshots(from_snapshot, to_snapshot, GROUP_SUMMARY).rows)),
                    ("SoSanhCoCauThoiHan", _row_dicts(compare_report_snapshots(from_snapshot, to_snapshot, GROUP_TERM_STRUCTURE).rows)),
                    ("SoSanhLoaiHinhKH", _row_dicts(compare_report_snapshots(from_snapshot, to_snapshot, GROUP_CUSTOMER_TYPE).rows)),
                    ("SoSanhChatLuongTinDung", _row_dicts(compare_report_snapshots(from_snapshot, to_snapshot, GROUP_CREDIT_QUALITY).rows)),
                    ("SoSanhNghiDinh55", _row_dicts(compare_report_snapshots(from_snapshot, to_snapshot, GROUP_DECREE55).rows)),
                    ("SoSanhNganhKinhTe", _row_dicts(compare_report_snapshots(from_snapshot, to_snapshot, GROUP_INDUSTRY).rows)),
                    ("ThongTin", _metadata_rows(filters, VIEW_COMPARE_PERIODS, from_snapshot, to_snapshot, self.status())),
                )
            else:
                snapshot = get_report_snapshot(self, filters.period, filters)
                sheets = (
                    ("TongHop", _row_dicts(snapshot.groups.get(GROUP_SUMMARY, ()))),
                    ("CoCauThoiHan", _row_dicts(snapshot.groups.get(GROUP_TERM_STRUCTURE, ()))),
                    ("LoaiHinhKhachHang", _row_dicts(snapshot.groups.get(GROUP_CUSTOMER_TYPE, ()))),
                    ("ChatLuongTinDung", _row_dicts(snapshot.groups.get(GROUP_CREDIT_QUALITY, ()))),
                    ("NghiDinh55", _row_dicts(snapshot.groups.get(GROUP_DECREE55, ()))),
                    ("NganhKinhTe", _row_dicts(snapshot.groups.get(GROUP_INDUSTRY, ()))),
                    ("ThongTin", _metadata_rows(filters, VIEW_CURRENT_PERIOD, snapshot, None, self.status())),
                )
            first = True
            for sheet_name, rows in sheets:
                sheet = workbook.active if first else workbook.create_sheet(sheet_name)
                first = False
                sheet.title = sheet_name
                _write_sheet(sheet, rows)
            workbook.save(destination)
        finally:
            workbook.close()
        return destination

    def _combined_rows(self, filters: CreditReportFilters) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        with closing(self.connect()) as connection:
            rule_codes = self._personal_rule_codes(connection)
            loan_where = "WHERE l.period = ?" if filters.period else ""
            card_where = "WHERE p.period = ?" if filters.period else ""
            period_params: tuple[object, ...] = (filters.period,) if filters.period else ()
            loans = connection.execute(
                f"""
                SELECT
                    l.*,
                    c.customer_sequence,
                    c.customer_code,
                    c.customer_name
                FROM credit_loan_period l
                JOIN credit_customer_master c ON c.id = l.customer_id
                {loan_where}
                """,
                period_params,
            ).fetchall()
            cards = connection.execute(
                f"""
                SELECT
                    p.*,
                    c.customer_sequence
                FROM credit_card_period p
                JOIN credit_customer_master c ON c.id = p.customer_id
                {card_where}
                """,
                period_params,
            ).fetchall()
        for row in loans:
            payload = {
                "source": CREDIT_SOURCE_LN01,
                "period": row["period"],
                "customer_id": row["customer_id"],
                "customer_key": f"{row['branch_code']}|{row['customer_sequence']}",
                "branch_code": row["branch_code"],
                "customer_sequence": row["customer_sequence"],
                "customer_code": row["customer_code"],
                "customer_name": row["customer_name"],
                "account_number": row["account_number"],
                "approval_sequence": row["approval_sequence"],
                "loan_key": row["loan_key"],
                "balance": float(row["outstanding_balance"] or 0),
                "customer_type": _classify_ln01_customer_type(row["customer_type_code"], rule_codes),
                "customer_type_code": row["customer_type_code"],
                "debt_group_code": _normalize_debt_code(row["debt_group_code"]),
                "secured_percent": row["secured_percent"],
                "industry_code": row["industry_code"],
                "officer_code": row["officer_code"],
                "officer_name": row["officer_name"],
                "term_category": row["term_category"],
                "group_code": row["group_code"],
            }
            if _passes_filters(payload, filters):
                rows.append(payload)
        for row in cards:
            payload = {
                "source": CREDIT_SOURCE_NIM_DN_CARD,
                "period": row["period"],
                "customer_id": row["customer_id"],
                "customer_key": f"{row['branch_code']}|{row['customer_sequence']}",
                "branch_code": row["branch_code"],
                "customer_sequence": row["customer_sequence"],
                "customer_code": row["customer_code"],
                "customer_name": row["customer_name"],
                "account_number": "",
                "approval_sequence": "",
                "loan_key": f"CARD|{row['id']}",
                "balance": float(row["balance"] or 0),
                "customer_type": row["customer_type"],
                "customer_type_code": row["customer_type"],
                "debt_group_code": _normalize_debt_code(row["debt_group_code"]),
                "secured_percent": None,
                "industry_code": "",
                "officer_code": row["officer_code"],
                "officer_name": row["officer_name"],
                "term_category": TERM_SHORT,
                "group_code": None,
            }
            if _passes_filters(payload, filters):
                rows.append(payload)
        return rows

    def _aggregate_card_rows(
        self,
        connection: sqlite3.Connection,
        run_id: int,
        period: str,
        rows: list[NormalizedLoanRow],
    ) -> list[dict[str, object]]:
        grouped: dict[tuple[object, ...], dict[str, object]] = {}
        now = now_text()
        for row in rows:
            customer_id = self._upsert_customer_from_card(connection, row)
            if not customer_id:
                continue
            customer_type = _classify_card_customer_type(row.customer_type)
            debt_group = _normalize_debt_code(row.debt_group_code)
            key = (
                period,
                row.branch_code,
                customer_id,
                customer_type,
                debt_group,
                row.officer_code,
                row.officer_name,
            )
            item = grouped.setdefault(
                key,
                {
                    "period": period,
                    "customer_id": customer_id,
                    "branch_code": row.branch_code,
                    "customer_code": row.customer_code,
                    "customer_name": row.customer_name,
                    "customer_type": customer_type,
                    "debt_group_code": debt_group,
                    "officer_code": row.officer_code,
                    "officer_name": row.officer_name,
                    "balance": 0.0,
                    "source_run_id": run_id,
                    "source_row_count": 0,
                    "created_at": now,
                },
            )
            item["balance"] = float(item["balance"] or 0) + float(row.balance or 0)
            item["source_row_count"] = int(item["source_row_count"] or 0) + int(
                getattr(row, "source_row_count", 1) or 1
            )
        return list(grouped.values())

    def _upsert_customer(self, connection: sqlite3.Connection, row: NormalizedLn01Row) -> int:
        now = now_text()
        connection.execute(
            """
            INSERT INTO credit_customer_master(
                branch_code, customer_sequence, customer_code, customer_name,
                latest_customer_type_code, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(branch_code, customer_sequence) DO UPDATE SET
                customer_code = excluded.customer_code,
                customer_name = CASE
                    WHEN excluded.customer_name <> '' THEN excluded.customer_name
                    ELSE credit_customer_master.customer_name
                END,
                latest_customer_type_code = CASE
                    WHEN excluded.latest_customer_type_code <> '' THEN excluded.latest_customer_type_code
                    ELSE credit_customer_master.latest_customer_type_code
                END,
                updated_at = excluded.updated_at
            """,
            (
                row.branch_code,
                row.customer_sequence,
                row.customer_code,
                row.customer_name,
                row.customer_type_code,
                now,
                now,
            ),
        )
        return int(
            connection.execute(
                """
                SELECT id FROM credit_customer_master
                WHERE branch_code = ? AND customer_sequence = ?
                """,
                (row.branch_code, row.customer_sequence),
            ).fetchone()["id"]
        )

    def _upsert_customer_from_card(self, connection: sqlite3.Connection, row: NormalizedLoanRow) -> int:
        if not row.branch_code or not row.customer_sequence:
            return 0
        now = now_text()
        connection.execute(
            """
            INSERT INTO credit_customer_master(
                branch_code, customer_sequence, customer_code, customer_name,
                latest_customer_type_code, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(branch_code, customer_sequence) DO UPDATE SET
                customer_code = CASE
                    WHEN excluded.customer_code <> '' THEN excluded.customer_code
                    ELSE credit_customer_master.customer_code
                END,
                customer_name = CASE
                    WHEN excluded.customer_name <> '' THEN excluded.customer_name
                    ELSE credit_customer_master.customer_name
                END,
                updated_at = excluded.updated_at
            """,
            (
                row.branch_code,
                row.customer_sequence,
                row.customer_code,
                row.customer_name,
                row.customer_type,
                now,
                now,
            ),
        )
        return int(
            connection.execute(
                """
                SELECT id FROM credit_customer_master
                WHERE branch_code = ? AND customer_sequence = ?
                """,
                (row.branch_code, row.customer_sequence),
            ).fetchone()["id"]
        )

    def _insert_import_run(self, connection: sqlite3.Connection, **values: object) -> int:
        now = now_text()
        cursor = connection.execute(
            """
            INSERT INTO credit_import_runs(
                period, source_type, source_file_name, source_sha256, source_file_size,
                source_row_count, status, message, imported_by, created_at, updated_at,
                duration_ms
            )
            VALUES (?, ?, ?, ?, ?, ?, 'success', ?, ?, ?, ?, ?)
            """,
            (
                values["period"],
                values["source_type"],
                values["source_file_name"],
                values["source_sha256"],
                values["source_file_size"],
                values["source_row_count"],
                values.get("message", ""),
                values.get("imported_by", ""),
                now,
                now,
                values.get("duration_ms", 0),
            ),
        )
        return int(cursor.lastrowid)

    def _insert_import_file(self, connection: sqlite3.Connection, **values: object) -> None:
        connection.execute(
            """
            INSERT INTO credit_import_files(
                run_id, period, source_type, file_name, sha256, file_size,
                row_count, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                values["run_id"],
                values["period"],
                values["source_type"],
                values["file_name"],
                values["sha256"],
                values["file_size"],
                values["row_count"],
                now_text(),
            ),
        )

    def _ln01_period_exists(self, connection: sqlite3.Connection, period: str) -> bool:
        return (
            connection.execute(
                """
                SELECT 1
                FROM credit_import_runs
                WHERE period = ? AND source_type = ? AND status = 'success'
                LIMIT 1
                """,
                (period, CREDIT_SOURCE_LN01),
            ).fetchone()
            is not None
        )

    def _source_sha_exists(self, connection: sqlite3.Connection, source_type: str, source_sha256: str) -> bool:
        if not source_sha256:
            return False
        return (
            connection.execute(
                """
                SELECT 1
                FROM credit_import_runs
                WHERE source_type = ? AND source_sha256 = ? AND status = 'success'
                LIMIT 1
                """,
                (source_type, source_sha256),
            ).fetchone()
            is not None
        )

    def _delete_ln01_period(self, connection: sqlite3.Connection, period: str) -> None:
        connection.execute("DELETE FROM credit_loan_period WHERE period = ?", (period,))
        connection.execute(
            "DELETE FROM credit_import_files WHERE period = ? AND source_type = ?",
            (period, CREDIT_SOURCE_LN01),
        )
        connection.execute(
            "DELETE FROM credit_import_runs WHERE period = ? AND source_type = ?",
            (period, CREDIT_SOURCE_LN01),
        )
        self._delete_orphan_customers(connection)

    def _delete_card_period(self, connection: sqlite3.Connection, period: str) -> None:
        connection.execute("DELETE FROM credit_card_period WHERE period = ?", (period,))
        connection.execute(
            "DELETE FROM credit_import_files WHERE period = ? AND source_type = ?",
            (period, CREDIT_SOURCE_NIM_DN_CARD),
        )
        connection.execute(
            "DELETE FROM credit_import_runs WHERE period = ? AND source_type = ?",
            (period, CREDIT_SOURCE_NIM_DN_CARD),
        )
        self._delete_orphan_customers(connection)

    def _delete_orphan_customers(self, connection: sqlite3.Connection) -> int:
        cursor = connection.execute(
            """
            DELETE FROM credit_customer_master
            WHERE NOT EXISTS (
                SELECT 1
                FROM credit_loan_period l
                WHERE l.customer_id = credit_customer_master.id
            )
            AND NOT EXISTS (
                SELECT 1
                FROM credit_card_period c
                WHERE c.customer_id = credit_customer_master.id
            )
            """
        )
        return int(cursor.rowcount or 0)

    def _validate_ln01_period(
        self,
        connection: sqlite3.Connection,
        period: str,
        *,
        expected_balance: float,
    ) -> None:
        actual = float(
            connection.execute(
                "SELECT COALESCE(SUM(outstanding_balance), 0) FROM credit_loan_period WHERE period = ?",
                (period,),
            ).fetchone()[0]
            or 0
        )
        if abs(actual - expected_balance) > 0.01:
            raise SummaryError(
                f"Tổng dư nợ LN01 không khớp sau ghi Credit.db: nguồn={expected_balance}, db={actual}."
            )

    def _personal_rule_codes(self, connection: sqlite3.Connection) -> set[str]:
        return {
            str(row["code"] or "").strip()
            for row in connection.execute(
                """
                SELECT code
                FROM credit_customer_type_rules
                WHERE is_personal = 1 AND is_enabled = 1
                """
            ).fetchall()
            if str(row["code"] or "").strip()
        }

    def _seed_default_customer_type_rules(self, connection: sqlite3.Connection) -> None:
        now = now_text()
        for code in ("100", "570"):
            connection.execute(
                """
                INSERT OR IGNORE INTO credit_customer_type_rules(
                    code, label, is_personal, is_enabled, created_at, updated_at
                )
                VALUES (?, ?, 1, 1, ?, ?)
                """,
                (code, f"Mã cá nhân {code}", now, now),
            )

    def _mark_migration(self, connection: sqlite3.Connection) -> None:
        checksum = hashlib.sha256(b"credit-report-schema-v2-group-code").hexdigest()
        existing = connection.execute(
            """
            SELECT 1
            FROM credit_schema_migrations
            WHERE version = ? AND migration_name = 'credit-report-schema' AND success = 1
            LIMIT 1
            """,
            (CREDIT_SCHEMA_VERSION,),
        ).fetchone()
        if existing is None:
            connection.execute(
                """
                INSERT INTO credit_schema_migrations(version, migration_name, applied_at, checksum, success)
                VALUES (?, 'credit-report-schema', ?, ?, 1)
                """,
                (CREDIT_SCHEMA_VERSION, now_text(), checksum),
            )

    def _log_action(
        self,
        connection: sqlite3.Connection,
        *,
        action: str,
        target_id: str,
        detail: str,
        created_by: str,
    ) -> None:
        connection.execute(
            """
            INSERT INTO credit_action_log(action, target_id, detail, created_by, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (action, target_id, detail, created_by, now_text()),
        )


REPORT_GROUP_LABELS = {
    GROUP_SUMMARY: "Tổng hợp",
    GROUP_TERM_STRUCTURE: "Cơ cấu thời hạn",
    GROUP_CUSTOMER_TYPE: "Loại hình khách hàng",
    GROUP_CREDIT_QUALITY: CREDIT_QUALITY_DISPLAY_NAME,
    GROUP_DECREE55: "Cho vay Nghị định 55",
    GROUP_INDUSTRY: "Ngành kinh tế",
}


def get_report_snapshot(
    repository: CreditReportRepository,
    period: str,
    filters: CreditReportFilters,
) -> ReportSnapshot:
    clean_period = str(period or "").strip()
    scoped_filters = replace(filters, period=clean_period, from_period="", to_period="")
    exists = bool(clean_period and clean_period in repository.available_periods())
    if not exists:
        message = f"Chưa có dữ liệu báo cáo kỳ {clean_period}." if clean_period else "Chưa chọn kỳ báo cáo."
        return ReportSnapshot(
            period=clean_period,
            filters=scoped_filters,
            exists=False,
            summary={"period": clean_period, "card_data_status": "missing", "card_data_message": message},
            groups={
                GROUP_SUMMARY: (),
                GROUP_TERM_STRUCTURE: (),
                GROUP_CUSTOMER_TYPE: (),
                GROUP_CREDIT_QUALITY: (),
                GROUP_DECREE55: (),
                GROUP_INDUSTRY: (),
            },
            card_status="missing",
            card_message=message,
            note=message,
        )
    summary = repository.overall_summary(scoped_filters)
    return ReportSnapshot(
        period=clean_period,
        filters=scoped_filters,
        exists=True,
        summary=summary,
        groups={
            GROUP_SUMMARY: _table_rows(repository.overview_rows(scoped_filters)),
            GROUP_TERM_STRUCTURE: _table_rows(repository.term_structure_rows(scoped_filters)),
            GROUP_CUSTOMER_TYPE: _table_rows(repository.customer_type_rows(scoped_filters)),
            GROUP_CREDIT_QUALITY: _table_rows(repository.debt_group_rows(scoped_filters)),
            GROUP_DECREE55: _table_rows(repository.decree55_rows(scoped_filters)),
            GROUP_INDUSTRY: _table_rows(repository.industry_rows(scoped_filters)),
        },
        card_status=str(summary.get("card_data_status") or ""),
        card_message=str(summary.get("card_data_message") or ""),
    )


def compare_report_snapshots(
    from_snapshot: ReportSnapshot,
    to_snapshot: ReportSnapshot,
    group_key: str,
) -> ReportComparisonResult:
    group = group_key if group_key in REPORT_GROUP_LABELS else GROUP_SUMMARY
    if group == GROUP_TERM_STRUCTURE:
        rows = _compare_named_rows(
            from_snapshot,
            to_snapshot,
            group,
            label_header="Loại thời hạn",
            value_header="Tổng dư nợ",
            share_header="Tỷ trọng",
            count_headers=(("Số KH Từ kỳ", "Số KH Đến kỳ", "Số khách hàng"),),
        )
    elif group == GROUP_CUSTOMER_TYPE:
        rows = _compare_named_rows(
            from_snapshot,
            to_snapshot,
            group,
            label_header="Loại khách hàng",
            value_header="Tổng dư nợ",
            share_header="Tỷ trọng",
            count_headers=(
                ("Số KH Từ kỳ", "Số KH Đến kỳ", "Số khách hàng"),
                ("Số món Từ kỳ", "Số món Đến kỳ", "Số món"),
            ),
            extra_difference=("Thay đổi số KH", "Số khách hàng"),
        )
    elif group == GROUP_CREDIT_QUALITY:
        rows = _compare_named_rows(
            from_snapshot,
            to_snapshot,
            group,
            label_header="Nhóm nợ",
            value_header="Tổng dư nợ",
            share_header="Tỷ lệ",
            count_headers=(
                ("Số món Từ kỳ", "Số món Đến kỳ", "Số món"),
                ("Số KH Từ kỳ", "Số KH Đến kỳ", "Số lượng khách hàng"),
                ("KH cá nhân Từ kỳ", "KH cá nhân Đến kỳ", "Số khách hàng cá nhân"),
                ("KH pháp nhân Từ kỳ", "KH pháp nhân Đến kỳ", "Số khách hàng pháp nhân"),
            ),
        )
    elif group == GROUP_DECREE55:
        rows = _compare_decree55_rows(from_snapshot, to_snapshot)
    elif group == GROUP_INDUSTRY:
        rows = _compare_named_rows(
            from_snapshot,
            to_snapshot,
            group,
            label_header="Nhóm ngành",
            value_header="Tổng dư nợ",
            share_header="Tỷ trọng LN01",
            count_headers=(
                ("Số KH Từ kỳ", "Số KH Đến kỳ", "Số lượng khách hàng"),
                ("KH cá nhân Từ kỳ", "KH cá nhân Đến kỳ", "Số khách hàng cá nhân"),
                ("KH pháp nhân Từ kỳ", "KH pháp nhân Đến kỳ", "Số khách hàng pháp nhân"),
            ),
        )
    else:
        rows = _compare_summary_rows(from_snapshot, to_snapshot)
    return ReportComparisonResult(
        from_snapshot=from_snapshot,
        to_snapshot=to_snapshot,
        group_key=group,
        rows=rows,
        kpis=_comparison_kpis(from_snapshot, to_snapshot),
        notes=_comparison_notes(from_snapshot, to_snapshot),
    )


def _table_rows(rows: list[dict[str, object]]) -> tuple[ReportTableRow, ...]:
    return tuple(ReportTableRow.from_mapping(row) for row in rows)


def _compare_summary_rows(from_snapshot: ReportSnapshot, to_snapshot: ReportSnapshot) -> tuple[ReportTableRow, ...]:
    specs = (
        ("Tổng dư nợ", "total_balance"),
        ("Dư nợ LN01", "ln01_total_balance"),
        ("Dư nợ thẻ tín dụng", "credit_card_balance"),
        ("Tổng khách hàng còn dư nợ", "customer_count"),
        ("Dư nợ ngắn hạn", "short_term_balance"),
        ("Dư nợ trung hạn", "medium_term_balance"),
        ("Dư nợ dài hạn", "long_term_balance"),
        ("Dư nợ cá nhân", "personal_balance"),
        ("Dư nợ pháp nhân", "legal_balance"),
        ("Nợ cần chú ý", "debt_group_2_balance"),
        ("Nợ xấu", "bad_debt_balance"),
        ("Dư nợ Nghị định 55", "decree55_balance"),
    )
    return tuple(
        _comparison_table_row(
            "Chỉ tiêu",
            label,
            _summary_value(from_snapshot, key),
            _summary_value(to_snapshot, key),
            diff_header="Tăng/giảm tuyệt đối",
            note=_summary_compare_note(from_snapshot, to_snapshot, key),
        )
        for label, key in specs
    )


def _comparison_kpis(from_snapshot: ReportSnapshot, to_snapshot: ReportSnapshot) -> tuple[ReportMetric, ...]:
    specs = (
        ("Tổng dư nợ", "total_balance", "money"),
        ("Tổng khách hàng", "customer_count", "count"),
        ("Dư nợ ngắn hạn", "short_term_balance", "money"),
        ("Dư nợ trung/dài hạn", ("medium_term_balance", "long_term_balance"), "money"),
        ("Nợ cần chú ý", "debt_group_2_balance", "money"),
        ("Nợ xấu", "bad_debt_balance", "money"),
    )
    metrics: list[ReportMetric] = []
    for label, key, value_kind in specs:
        from_value = _summary_value(from_snapshot, key)
        to_value = _summary_value(to_snapshot, key)
        difference, growth = _difference_growth(from_value, to_value)
        metrics.append(
            ReportMetric(
                label=label,
                value=to_value,
                value_kind=value_kind,
                from_value=from_value,
                to_value=to_value,
                difference=difference,
                growth_rate=growth,
                note="N/A do giá trị Từ kỳ = 0" if growth is None and _numeric(from_value) == 0 else "",
            )
        )
    return tuple(metrics)


def _compare_named_rows(
    from_snapshot: ReportSnapshot,
    to_snapshot: ReportSnapshot,
    group_key: str,
    *,
    label_header: str,
    value_header: str,
    share_header: str,
    count_headers: tuple[tuple[str, str, str], ...] = (),
    extra_difference: tuple[str, str] | None = None,
) -> tuple[ReportTableRow, ...]:
    from_rows = {str(row.get(label_header, "")): row for row in from_snapshot.groups.get(group_key, ())}
    to_rows = {str(row.get(label_header, "")): row for row in to_snapshot.groups.get(group_key, ())}
    labels = _ordered_labels(group_key, tuple(from_rows.keys()) + tuple(to_rows.keys()))
    output: list[ReportTableRow] = []
    for label in labels:
        before = _row_value(from_rows.get(label), value_header, from_snapshot.exists)
        after = _row_value(to_rows.get(label), value_header, to_snapshot.exists)
        difference, growth = _difference_growth(before, after)
        before_share = _row_value(from_rows.get(label), share_header, from_snapshot.exists)
        after_share = _row_value(to_rows.get(label), share_header, to_snapshot.exists)
        values: list[tuple[str, object]] = [
            (label_header, label),
            ("Dư nợ Từ kỳ", before),
            ("Dư nợ Đến kỳ", after),
            ("Tăng/giảm", difference),
            ("Tăng trưởng (%)", growth),
            ("Tỷ trọng Từ kỳ", before_share),
            ("Tỷ trọng Đến kỳ", after_share),
            ("Thay đổi tỷ trọng (điểm %)", _difference(before_share, after_share)),
        ]
        for from_header, to_header, source_header in count_headers:
            values.append((from_header, _row_value(from_rows.get(label), source_header, from_snapshot.exists)))
            values.append((to_header, _row_value(to_rows.get(label), source_header, to_snapshot.exists)))
        if extra_difference is not None:
            output_header, source_header = extra_difference
            values.append(
                (
                    output_header,
                    _difference(
                        _row_value(from_rows.get(label), source_header, from_snapshot.exists),
                        _row_value(to_rows.get(label), source_header, to_snapshot.exists),
                    ),
                )
            )
        output.append(ReportTableRow(tuple(values)))
    return tuple(output)


def _compare_decree55_rows(from_snapshot: ReportSnapshot, to_snapshot: ReportSnapshot) -> tuple[ReportTableRow, ...]:
    from_rows = from_snapshot.groups.get(GROUP_DECREE55, ())
    to_rows = to_snapshot.groups.get(GROUP_DECREE55, ())
    before = from_rows[0] if from_rows else None
    after = to_rows[0] if to_rows else None
    specs = (
        ("Tổng dư nợ", "Tổng dư nợ", False),
        ("Tỷ trọng", "Tỷ trọng trên dư nợ LN01", True),
        ("Số món", "Số món", False),
        ("Số khách hàng", "Số lượng khách hàng", False),
        ("Khách hàng cá nhân", "Số khách hàng cá nhân", False),
        ("Khách hàng pháp nhân", "Số khách hàng pháp nhân", False),
    )
    rows: list[ReportTableRow] = []
    for label, source_header, is_share in specs:
        from_value = _row_value(before, source_header, from_snapshot.exists)
        to_value = _row_value(after, source_header, to_snapshot.exists)
        difference, growth = _difference_growth(from_value, to_value)
        rows.append(
            ReportTableRow(
                (
                    ("Chỉ tiêu", label),
                    ("Giá trị Từ kỳ", from_value),
                    ("Giá trị Đến kỳ", to_value),
                    ("Tăng/giảm", difference),
                    ("Tăng trưởng (%)", difference if is_share else growth),
                    ("Ghi chú", "Chênh lệch điểm %" if is_share else "Xác định từ LN01 có SECURED_PERCENT = 0."),
                )
            )
        )
    return tuple(rows)


def _comparison_table_row(
    label_header: str,
    label: str,
    from_value: object,
    to_value: object,
    *,
    diff_header: str = "Tăng/giảm",
    note: str = "",
) -> ReportTableRow:
    difference, growth = _difference_growth(from_value, to_value)
    if not note and growth is None and _numeric(from_value) == 0:
        note = "N/A do giá trị Từ kỳ = 0"
    return ReportTableRow(
        (
            (label_header, label),
            ("Giá trị Từ kỳ", from_value),
            ("Giá trị Đến kỳ", to_value),
            (diff_header, difference),
            ("Tăng trưởng (%)", growth),
            ("Ghi chú", note),
        )
    )


def _ordered_labels(group_key: str, labels: tuple[str, ...]) -> tuple[str, ...]:
    preferred = {
        GROUP_TERM_STRUCTURE: tuple(TERM_LABELS[key] for key in (TERM_SHORT, TERM_MEDIUM, TERM_LONG, TERM_UNKNOWN)),
        GROUP_CUSTOMER_TYPE: tuple(CUSTOMER_TYPE_LABELS[key] for key in (CUSTOMER_TYPE_PERSONAL, CUSTOMER_TYPE_LEGAL, CUSTOMER_TYPE_UNKNOWN)),
        GROUP_CREDIT_QUALITY: tuple(DEBT_BUCKET_LABELS[key] for key in ("GROUP_1", "GROUP_2", "BAD_DEBT", "UNKNOWN")),
        GROUP_INDUSTRY: tuple(
            INDUSTRY_LABELS[key]
            for key in (INDUSTRY_HIGH_TECH, INDUSTRY_WHOLESALE_RETAIL, INDUSTRY_REAL_ESTATE, INDUSTRY_OTHER)
        ),
    }.get(group_key, ())
    seen = {label for label in labels if label}
    ordered = [label for label in preferred if label in seen]
    ordered.extend(sorted(seen.difference(ordered)))
    return tuple(ordered)


def _summary_value(snapshot: ReportSnapshot, key: str | tuple[str, ...]) -> object:
    if not snapshot.exists:
        return None
    if isinstance(key, tuple):
        values = [_summary_value(snapshot, item) for item in key]
        if any(value is None for value in values):
            return None
        return sum(float(value or 0) for value in values)
    if key == "credit_card_balance" and snapshot.card_status == "missing":
        return None
    return snapshot.summary.get(key)


def _summary_compare_note(from_snapshot: ReportSnapshot, to_snapshot: ReportSnapshot, key: str) -> str:
    notes: list[str] = []
    if key in {"credit_card_balance", "total_balance", "short_term_balance"}:
        if from_snapshot.exists and from_snapshot.card_status == "missing":
            notes.append(f"Từ kỳ {from_snapshot.period}: chưa có dữ liệu thẻ; tổng chỉ gồm LN01.")
        if to_snapshot.exists and to_snapshot.card_status == "missing":
            notes.append(f"Đến kỳ {to_snapshot.period}: chưa có dữ liệu thẻ; tổng chỉ gồm LN01.")
    return " ".join(notes)


def _comparison_notes(from_snapshot: ReportSnapshot, to_snapshot: ReportSnapshot) -> tuple[str, ...]:
    notes: list[str] = []
    if not from_snapshot.exists:
        notes.append(from_snapshot.note)
    if not to_snapshot.exists:
        notes.append(to_snapshot.note)
    if from_snapshot.period and from_snapshot.period == to_snapshot.period:
        notes.append("Từ kỳ và Đến kỳ đang giống nhau.")
    for label, snapshot in (("Từ kỳ", from_snapshot), ("Đến kỳ", to_snapshot)):
        if snapshot.exists and snapshot.card_status == "missing":
            notes.append(f"{label} {snapshot.period}: {snapshot.card_message}")
    return tuple(note for note in notes if note)


def _row_value(row: ReportTableRow | None, header: str, exists: bool) -> object:
    if not exists:
        return None
    if row is None:
        return 0
    return row.get(header)


def _difference_growth(from_value: object, to_value: object) -> tuple[float | None, float | None]:
    before = _numeric(from_value)
    after = _numeric(to_value)
    if before is None or after is None:
        return None, None
    difference = after - before
    growth = None if before == 0 else difference / before * 100
    return difference, growth


def _difference(from_value: object, to_value: object) -> float | None:
    before = _numeric(from_value)
    after = _numeric(to_value)
    if before is None or after is None:
        return None
    return after - before


def _numeric(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def classify_term_from_account(account_number: object) -> str:
    text = _normalize_code(account_number)
    if text.startswith("211"):
        return TERM_SHORT
    if text.startswith("212") or text.startswith("251"):
        return TERM_MEDIUM
    if text.startswith("213") or text.startswith("252"):
        return TERM_LONG
    return TERM_UNKNOWN


def classify_industry(industry_code: object) -> str:
    code = _normalize_code(industry_code)
    if code in HIGH_TECH_CODES:
        return INDUSTRY_HIGH_TECH
    if len(code) == 6 and code[:4].isdigit() and 1001 <= int(code[:4]) <= 1013:
        return INDUSTRY_WHOLESALE_RETAIL
    if code.startswith("8") or code in REAL_ESTATE_CODES:
        return INDUSTRY_REAL_ESTATE
    return INDUSTRY_OTHER


def _classify_ln01_customer_type(value: object, personal_codes: set[str]) -> str:
    code = _normalize_code(value)
    if not code:
        return CUSTOMER_TYPE_UNKNOWN
    if code in personal_codes:
        return CUSTOMER_TYPE_PERSONAL
    return CUSTOMER_TYPE_LEGAL


def _classify_card_customer_type(value: object) -> str:
    text = str(value or "").strip().upper()
    if text == "CN":
        return CUSTOMER_TYPE_PERSONAL
    if text == "TC":
        return CUSTOMER_TYPE_LEGAL
    return CUSTOMER_TYPE_UNKNOWN


def _normalize_code(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if text.startswith("'"):
        text = text[1:].strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _normalize_debt_code(value: object) -> str:
    text = _normalize_code(value)
    if text in {"1", "2", "3", "4", "5"}:
        return f"0{text}"
    if text in {"01", "02", "03", "04", "05"}:
        return text
    return DEBT_GROUP_UNKNOWN


def _debt_bucket(debt_group_code: object) -> str:
    code = _normalize_debt_code(debt_group_code)
    if code == "01":
        return "GROUP_1"
    if code == "02":
        return "GROUP_2"
    if code in {"03", "04", "05"}:
        return "BAD_DEBT"
    return "UNKNOWN"


def _loan_key(customer_id: int, account_number: object, approval_sequence: object) -> str:
    account = _normalize_code(account_number) or "NO_ACCOUNT"
    approval = _normalize_code(approval_sequence) or "NO_APPRSEQ"
    return f"{customer_id}|{account}|{approval}"


def _table_has_column(connection: sqlite3.Connection, table_name: str, column_name: str) -> bool:
    return any(str(row["name"]).casefold() == column_name.casefold() for row in connection.execute(f"PRAGMA table_info({table_name})"))


def _has_old_credit_loan_unique_index(connection: sqlite3.Connection) -> bool:
    for index_row in connection.execute("PRAGMA index_list(credit_loan_period)").fetchall():
        if int(index_row["unique"] or 0) != 1:
            continue
        index_name = str(index_row["name"] or "")
        if index_name == "idx_credit_loan_period_unique_period_branch_loan_group":
            continue
        columns = [
            str(info["name"] or "")
            for info in connection.execute(f"PRAGMA index_info({index_name})").fetchall()
            if str(info["name"] or "")
        ]
        if columns == ["period", "branch_code", "loan_key"]:
            return True
    return False


def _import_group_stat(row: sqlite3.Row, key: str) -> object:
    try:
        stats = json.loads(str(row["group_code_stats_json"] or "{}"))
    except (KeyError, IndexError, TypeError, ValueError, json.JSONDecodeError):
        stats = {}
    return stats.get(key, 0)


def _require_period(value: str) -> str:
    text = str(value or "").strip()
    if not text or len(text) != 7 or text[4] != "-" or not text[:4].isdigit() or not text[5:].isdigit():
        raise SummaryError("Kỳ dữ liệu báo cáo phải có dạng YYYY-MM.")
    month = int(text[5:])
    if month < 1 or month > 12:
        raise SummaryError("Tháng trong kỳ dữ liệu báo cáo không hợp lệ.")
    return text


def _passes_filters(row: dict[str, object], filters: CreditReportFilters) -> bool:
    if filters.branch_code and row["branch_code"] != filters.branch_code:
        return False
    if filters.customer_type and row["customer_type"] != filters.customer_type:
        return False
    if filters.debt_group and _debt_bucket(row["debt_group_code"]) != filters.debt_group:
        return False
    if filters.term_category and row["term_category"] != filters.term_category:
        return False
    if filters.officer:
        display = _officer_display(str(row["officer_code"]), str(row["officer_name"]))
        if filters.officer not in {display, str(row["officer_code"]), str(row["officer_name"])}:
            return False
    needle = str(filters.search or "").strip().casefold()
    if needle:
        haystack = " ".join(
            str(row.get(key) or "")
            for key in (
                "customer_sequence",
                "customer_code",
                "customer_name",
                "account_number",
                "approval_sequence",
            )
        ).casefold()
        if needle not in haystack:
            return False
    return True


def _officer_display(code: str, name: str) -> str:
    clean_code = str(code or "").strip()
    clean_name = str(name or "").strip()
    if clean_code:
        return f"[{clean_code}] {clean_name or clean_code}"
    return clean_name


def _total_by_source(rows: list[dict[str, object]]) -> dict[str, float]:
    return {
        "ln01": sum(float(row["balance"] or 0) for row in rows if row["source"] == CREDIT_SOURCE_LN01),
        "card": sum(float(row["balance"] or 0) for row in rows if row["source"] == CREDIT_SOURCE_NIM_DN_CARD),
    }


def _customer_totals(rows: list[dict[str, object]]) -> dict[object, float]:
    totals: dict[object, float] = defaultdict(float)
    for row in rows:
        totals[row["customer_key"]] += float(row["balance"] or 0)
    return totals


def _balance_by(rows: list[dict[str, object]], key: str) -> dict[str, float]:
    totals = defaultdict(float)
    for row in rows:
        totals[str(row[key])] += float(row["balance"] or 0)
    for required in (CUSTOMER_TYPE_PERSONAL, CUSTOMER_TYPE_LEGAL, CUSTOMER_TYPE_UNKNOWN):
        totals.setdefault(required, 0.0)
    return totals


def _balance_by_debt_bucket(rows: list[dict[str, object]]) -> dict[str, float]:
    totals = defaultdict(float)
    for row in rows:
        totals[_debt_bucket(row["debt_group_code"])] += float(row["balance"] or 0)
    for required in ("GROUP_1", "GROUP_2", "BAD_DEBT", "UNKNOWN"):
        totals.setdefault(required, 0.0)
    return totals


def _term_rows_from_combined(rows: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    output = {
        key: {
            "ln01_balance": 0.0,
            "card_balance": 0.0,
            "total_balance": 0.0,
            "customer_totals": defaultdict(float),
            "customer_count": 0,
        }
        for key in (TERM_SHORT, TERM_MEDIUM, TERM_LONG, TERM_UNKNOWN)
    }
    for row in rows:
        term = str(row["term_category"] or TERM_UNKNOWN)
        if term not in output:
            term = TERM_UNKNOWN
        balance = float(row["balance"] or 0)
        if row["source"] == CREDIT_SOURCE_NIM_DN_CARD:
            output[TERM_SHORT]["card_balance"] += balance
            output[TERM_SHORT]["total_balance"] += balance
            output[TERM_SHORT]["customer_totals"][row["customer_key"]] += balance
        else:
            output[term]["ln01_balance"] += balance
            output[term]["total_balance"] += balance
            output[term]["customer_totals"][row["customer_key"]] += balance
    for item in output.values():
        item["customer_count"] = sum(1 for value in item["customer_totals"].values() if value > 0)
    return output


def _customer_count_by_type(rows: list[dict[str, object]], customer_type: str) -> int:
    scoped = [row for row in rows if row["customer_type"] == customer_type]
    return sum(1 for value in _customer_totals(scoped).values() if value > 0)


def _is_secured_zero(value: object) -> bool:
    try:
        return abs(float(value)) < 0.0000001
    except (TypeError, ValueError):
        return False


def _ratio(value: object, total: object) -> float | None:
    denominator = float(total or 0)
    if denominator == 0:
        return None
    return float(value or 0) / denominator * 100


def _comparison_row(label: str, from_value: object, to_value: object) -> dict[str, object]:
    before = float(from_value or 0)
    after = float(to_value or 0)
    difference = after - before
    growth = None if before == 0 else difference / before * 100
    return {
        "Chỉ tiêu": label,
        "Giá trị Từ kỳ": before,
        "Giá trị Đến kỳ": after,
        "Tăng/giảm tuyệt đối": difference,
        "Tăng trưởng (%)": growth,
        "Ghi chú": "N/A do giá trị Từ kỳ = 0" if growth is None else "",
    }


def _info_row(key: object, value: object) -> dict[str, object]:
    return {"Thông tin": str(key), "Giá trị": str(value)}


def _row_dicts(rows: tuple[ReportTableRow, ...]) -> list[dict[str, object]]:
    return [row.to_dict() for row in rows]


def _metadata_rows(
    filters: CreditReportFilters,
    view_mode: str,
    snapshot: ReportSnapshot,
    to_snapshot: ReportSnapshot | None,
    status: dict[str, object],
) -> list[dict[str, object]]:
    rows = [
        {"Thông tin": "Chế độ xem", "Giá trị": "So sánh các kỳ" if view_mode == VIEW_COMPARE_PERIODS else "Kỳ hiện tại"},
        {"Thông tin": "Kỳ báo cáo", "Giá trị": filters.period},
        {"Thông tin": "Từ kỳ", "Giá trị": filters.from_period},
        {"Thông tin": "Đến kỳ", "Giá trị": filters.to_period},
        {"Thông tin": "Chi nhánh", "Giá trị": filters.branch_code},
        {"Thông tin": "Phòng giao dịch", "Giá trị": filters.transaction_office},
        {"Thông tin": "Loại khách hàng", "Giá trị": filters.customer_type},
        {"Thông tin": "Nhóm nợ", "Giá trị": filters.debt_group},
        {"Thông tin": "Loại thời hạn", "Giá trị": filters.term_category},
        {"Thông tin": "CBTD", "Giá trị": filters.officer},
        {"Thông tin": f"Trạng thái thẻ {snapshot.period}", "Giá trị": snapshot.card_message},
    ]
    if to_snapshot is not None:
        rows.append({"Thông tin": f"Trạng thái thẻ {to_snapshot.period}", "Giá trị": to_snapshot.card_message})
    rows.extend(_info_row(key, value) for key, value in status.items())
    rows.append({"Thông tin": "Thời gian xuất", "Giá trị": now_text()})
    return rows


def _write_sheet(sheet, rows: list[dict[str, object]]) -> None:
    headers = list(rows[0].keys()) if rows else ["Thông tin"]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="7A003C")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    for index, header in enumerate(headers, start=1):
        column = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column].width = min(max(12, len(str(header)) + 4), 36)
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for item in cell:
                if isinstance(item.value, (int, float)):
                    item.number_format = "#,##0.00" if "Tỷ" in str(header) or "%" in str(header) else "#,##0"
                    item.alignment = Alignment(horizontal="right", vertical="top")
                else:
                    item.number_format = "@"
                    item.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
