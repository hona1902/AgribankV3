from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agribank_v3.features.credit.summary.customer.filters import CustomerFilters
from agribank_v3.features.credit.summary.customer.formatters import format_customer_type
from agribank_v3.features.credit.summary.customer.repository import CustomerRepository


ColumnSpec = tuple[str, str, str]

MONEY_FORMAT = "#,##0"
PERCENT_FORMAT = '0.00"%"'
TEXT_FORMAT = "@"
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")

CUSTOMER_LIST_COLUMNS: tuple[ColumnSpec, ...] = (
    ("period", "Kỳ", "center"),
    ("customer_code", "Mã khách hàng", "text"),
    ("customer_name", "Tên khách hàng", "text"),
    ("customer_type", "Loại khách hàng", "customer_type"),
    ("branch_code", "Mã chi nhánh", "text"),
    ("effective_officer_name", "Cán bộ quản lý hiệu lực", "text"),
    ("imported_officer_name", "Cán bộ gốc từ file", "text"),
    ("officer_count", "Số cán bộ quản lý", "integer"),
    ("total_balance", "Tổng dư nợ", "money"),
    ("short_term_balance", "Dư nợ ngắn hạn", "money"),
    ("medium_long_term_balance", "Dư nợ trung/dài hạn", "money"),
    ("other_balance", "Dư nợ chưa phân loại", "money"),
    ("medium_long_ratio", "Tỷ lệ trung/dài hạn", "percent"),
    ("average_rate", "Lãi suất bình quân", "percent"),
    ("nim_before", "NIM trước ĐC", "percent"),
    ("nim_after", "NIM sau ĐC", "percent"),
    ("source_loan_count", "Số bản ghi nguồn", "integer"),
    ("override_status", "Trạng thái override", "text"),
)

MOVEMENT_COLUMNS: tuple[ColumnSpec, ...] = (
    ("customer_code", "Mã KH", "text"),
    ("customer_name", "Tên khách hàng", "text"),
    ("customer_type", "Loại KH", "customer_type"),
    ("branch_display", "Chi nhánh", "text"),
    ("effective_officer_name", "Cán bộ quản lý hiệu lực", "text"),
    ("previous_balance", "Dư nợ kỳ trước", "money"),
    ("current_balance", "Dư nợ kỳ hiện tại", "money"),
    ("difference", "Tăng/giảm", "money"),
    ("growth_rate", "Tăng trưởng (%)", "percent_or_blank"),
    ("movement_status", "Phân loại", "text"),
)

TOP_BALANCE_COLUMNS: tuple[ColumnSpec, ...] = (
    ("rank", "STT", "integer"),
    ("customer_code", "Mã KH", "text"),
    ("customer_name", "Tên khách hàng", "text"),
    ("customer_type", "Loại KH", "customer_type"),
    ("branch_code", "Chi nhánh", "branch_display"),
    ("effective_officer_name", "Cán bộ quản lý hiệu lực", "text"),
    ("total_balance", "Tổng dư nợ", "money"),
    ("short_term_balance", "Dư nợ ngắn hạn", "money"),
    ("medium_long_term_balance", "Dư nợ trung/dài hạn", "money"),
    ("medium_long_ratio", "Tỷ lệ trung/dài hạn", "percent"),
    ("nim_after", "NIM sau ĐC", "percent"),
)

TOP_MOVEMENT_COLUMNS: tuple[ColumnSpec, ...] = (
    ("rank", "STT", "integer"),
    ("customer_code", "Mã KH", "text"),
    ("customer_name", "Tên khách hàng", "text"),
    ("customer_type", "Loại KH", "customer_type"),
    ("branch_code", "Chi nhánh", "branch_display"),
    ("effective_officer_name", "Cán bộ quản lý hiệu lực", "text"),
    ("previous_balance", "Dư nợ kỳ trước", "money"),
    ("current_balance", "Dư nợ kỳ hiện tại", "money"),
    ("difference", "Tăng/giảm", "money"),
    ("growth_rate", "Tăng trưởng (%)", "percent_or_blank"),
    ("movement_status", "Phân loại", "text"),
)

MULTIPLE_OFFICER_COLUMNS: tuple[ColumnSpec, ...] = (
    ("period", "Kỳ", "center"),
    ("customer_code", "Mã KH", "text"),
    ("customer_name", "Tên khách hàng", "text"),
    ("customer_type", "Loại KH", "customer_type"),
    ("branch_code", "Chi nhánh", "branch_display"),
    ("total_balance", "Tổng dư nợ", "money"),
    ("officer_count", "Số cán bộ", "integer"),
    ("imported_officer_name", "Cán bộ chính", "text"),
    ("effective_officer_name", "Cán bộ hiệu lực", "text"),
    ("officer_list", "Danh sách cán bộ", "text"),
    ("override_status", "Trạng thái override", "text"),
)

CROSS_BRANCH_COLUMNS: tuple[ColumnSpec, ...] = (
    ("rank", "STT", "integer"),
    ("period", "Kỳ", "center"),
    ("customer_sequence", "Mã khách hàng gốc", "text"),
    ("customer_name", "Tên khách hàng", "text"),
    ("customer_type_display", "Loại khách hàng", "text"),
    ("branch_count", "Số chi nhánh vay", "integer"),
    ("office_count", "Số đơn vị vay", "integer"),
    ("head_office_count", "Số Hội sở", "integer"),
    ("pgd_count", "Số PGD", "integer"),
    ("has_head_and_pgd_text", "Có Hội sở và PGD", "text"),
    ("has_multi_pgd_text", "Có nhiều PGD", "text"),
    ("scope_status", "Loại phạm vi vay", "text"),
    ("representative_office_list", "Đơn vị đại diện", "text"),
    ("representative_office_type_list", "Loại đơn vị đại diện", "text"),
    ("representative_reason_list", "Lý do phân bổ", "text"),
    ("branch_list", "Danh sách chi nhánh", "text"),
    ("office_list", "Danh sách đơn vị thực tế", "text"),
    ("total_balance", "Tổng dư nợ", "money"),
    ("head_office_balance", "Dư nợ Hội sở", "money"),
    ("pgd_balance", "Dư nợ PGD", "money"),
    ("short_term_balance", "Dư nợ ngắn hạn", "money"),
    ("medium_long_term_balance", "Dư nợ trung/dài hạn", "money"),
    ("medium_long_ratio", "Tỷ lệ trung/dài hạn", "percent"),
    ("average_rate", "Lãi suất bình quân", "percent"),
    ("nim_before", "NIM trước ĐC", "percent"),
    ("nim_after", "NIM sau ĐC", "percent"),
    ("officer_count", "Số cán bộ quản lý", "integer"),
    ("officer_list", "Danh sách cán bộ", "text"),
    ("has_override", "Có override", "yes_no"),
    ("conflict_status", "Có xung đột tên/loại KH", "text"),
)

CROSS_BRANCH_DETAIL_COLUMNS: tuple[ColumnSpec, ...] = (
    ("period", "Kỳ", "center"),
    ("customer_code", "Mã khách hàng đầy đủ", "text"),
    ("branch_code", "Mã chi nhánh", "text"),
    ("branch_name", "Tên chi nhánh", "text"),
    ("office_code", "Mã đơn vị", "text"),
    ("office_type_display", "Loại đơn vị", "text"),
    ("office_name", "Tên đơn vị", "text"),
    ("customer_name", "Tên khách hàng tại đơn vị", "text"),
    ("customer_type_display", "Loại khách hàng", "text"),
    ("imported_officer_display", "Cán bộ gốc", "text"),
    ("effective_officer_display", "Cán bộ hiệu lực", "text"),
    ("total_balance", "Tổng dư nợ", "money"),
    ("short_term_balance", "Dư nợ ngắn hạn", "money"),
    ("medium_long_term_balance", "Dư nợ trung/dài hạn", "money"),
    ("medium_long_ratio", "Tỷ lệ trung/dài hạn", "percent"),
    ("average_rate", "Lãi suất bình quân", "percent"),
    ("nim_before", "NIM trước ĐC", "percent"),
    ("nim_after", "NIM sau ĐC", "percent"),
    ("source_loan_count", "Số bản ghi nguồn", "integer"),
    ("override_status", "Trạng thái override", "text"),
)

CROSS_BRANCH_HISTORY_COLUMNS: tuple[ColumnSpec, ...] = (
    ("period", "Kỳ", "center"),
    ("branch_count", "Số chi nhánh vay", "integer"),
    ("office_count", "Số đơn vị vay", "integer"),
    ("head_office_count", "Số Hội sở", "integer"),
    ("pgd_count", "Số PGD", "integer"),
    ("has_head_and_pgd", "Có Hội sở và PGD cùng chi nhánh", "yes_no"),
    ("has_multi_pgd", "Có nhiều PGD cùng chi nhánh", "yes_no"),
    ("branch_list", "Danh sách chi nhánh", "text"),
    ("office_list", "Danh sách đơn vị", "text"),
    ("total_balance", "Tổng dư nợ", "money"),
    ("head_office_balance", "Dư nợ Hội sở", "money"),
    ("pgd_balance", "Dư nợ PGD", "money"),
    ("difference", "Tăng/giảm dư nợ", "money_or_blank"),
    ("average_rate", "Lãi suất bình quân", "percent"),
    ("nim_before", "NIM trước ĐC", "percent"),
    ("nim_after", "NIM sau ĐC", "percent"),
)

IMPORT_RUN_COLUMNS: tuple[ColumnSpec, ...] = (
    ("id", "Run ID", "integer"),
    ("period", "Kỳ", "center"),
    ("source_folder", "Thư mục nguồn", "text"),
    ("file_count", "Số file", "integer"),
    ("source_row_count", "Tổng dòng nguồn", "integer"),
    ("customer_count", "Số khách hàng", "integer"),
    ("status", "Trạng thái", "text"),
    ("started_at", "Thời gian bắt đầu", "text"),
    ("completed_at", "Thời gian kết thúc", "text"),
    ("created_by", "Người thực hiện", "text"),
    ("computer_name", "Máy thực hiện", "text"),
    ("error_message", "Thông báo lỗi", "text"),
)

IMPORT_FILE_COLUMNS: tuple[ColumnSpec, ...] = (
    ("file_name", "Tên file", "text"),
    ("file_path", "Đường dẫn", "text"),
    ("file_hash", "SHA-256", "text"),
    ("branch_code", "Chi nhánh", "branch_display"),
    ("period", "Kỳ", "center"),
    ("source_row_count", "Số dòng nguồn", "integer"),
    ("customer_count", "Số khách hàng", "integer"),
    ("status", "Trạng thái", "text"),
    ("error_message", "Lỗi nếu có", "text"),
)

OFFICER_DIRECTORY_COLUMNS: tuple[ColumnSpec, ...] = (
    ("officer_code", "Mã cán bộ", "text"),
    ("officer_name", "Tên cán bộ", "text"),
    ("branch_code", "Chi nhánh", "branch_display"),
    ("transaction_office", "Phòng GD", "text"),
    ("is_active", "Trạng thái", "active_status"),
    ("updated_at", "Ngày cập nhật", "text"),
)

DETAIL_BALANCE_COLUMNS: tuple[ColumnSpec, ...] = (
    ("period", "Kỳ", "center"),
    ("total_balance", "Tổng dư nợ", "money"),
    ("short_term_balance", "Dư nợ ngắn hạn", "money"),
    ("medium_long_term_balance", "Dư nợ trung/dài hạn", "money"),
    ("other_balance", "Dư nợ khác", "money"),
    ("difference", "Tăng/giảm tuyệt đối", "money_or_blank"),
    ("growth_rate", "Tăng trưởng (%)", "percent_or_blank"),
)

DETAIL_NIM_COLUMNS: tuple[ColumnSpec, ...] = (
    ("period", "Kỳ", "center"),
    ("average_rate", "Lãi suất bình quân", "percent"),
    ("nim_before", "NIM trước ĐC", "percent"),
    ("nim_after", "NIM sau ĐC", "percent"),
)

DETAIL_OFFICER_COLUMNS: tuple[ColumnSpec, ...] = (
    ("period", "Kỳ", "center"),
    ("imported_officer_code", "Mã cán bộ gốc", "text"),
    ("imported_officer_name", "Tên cán bộ gốc", "text"),
    ("balance_managed", "Dư nợ cán bộ quản lý", "money"),
    ("source_loan_count", "Số bản ghi nguồn", "integer"),
    ("is_primary", "Là cán bộ chính", "yes_no"),
    ("override_officer_name", "Cán bộ override", "text"),
    ("override_scope", "Phạm vi override", "text"),
    ("override_reason", "Lý do", "text"),
    ("override_created_by", "Người cập nhật", "text"),
    ("override_updated_at", "Thời gian cập nhật", "text"),
)


def suggested_customer_export_name(tab_name: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"Customer_{tab_name}_{stamp}.xlsx"


def export_customer_dashboard(
    repository: CustomerRepository,
    filters: CustomerFilters,
    destination: Path,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TongQuanKhachHang"
    _write_dashboard_sheet(worksheet, repository, filters)
    workbook.save(destination)
    return Path(destination)


def export_customer_list(
    repository: CustomerRepository,
    filters: CustomerFilters,
    destination: Path,
    *,
    sort_by: str = "period",
    sort_desc: bool = True,
) -> Path:
    rows = repository.all_customer_rows(filters, sort_by=sort_by, sort_desc=sort_desc)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DanhSachKhachHang"
    _write_rows(worksheet, rows, CUSTOMER_LIST_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def export_customer_growth(
    repository: CustomerRepository,
    previous_period: str,
    current_period: str,
    filters: CustomerFilters,
    destination: Path,
    *,
    sort_by: str = "difference",
    sort_desc: bool = True,
) -> Path:
    rows = repository.all_movement_rows(
        previous_period,
        current_period,
        filters,
        sort_by=sort_by,
        sort_desc=sort_desc,
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "BienDongDuNo"
    _write_rows(worksheet, rows, MOVEMENT_COLUMNS, repository=repository)
    _write_movement_metadata_sheet(
        workbook.create_sheet("ThongTinLoc"),
        previous_period,
        current_period,
        filters,
        repository,
    )
    workbook.save(destination)
    return Path(destination)


def export_top_customer_balance(
    repository: CustomerRepository,
    filters: CustomerFilters,
    period: str,
    limit: int,
    destination: Path,
) -> Path:
    rows = _rank_rows(repository.get_top_customers_by_balance(filters, period, limit))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TopKhachHangDuNo"
    _write_rows(worksheet, rows, TOP_BALANCE_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def export_top_customer_balance_rows(rows: Iterable[dict[str, object]], destination: Path) -> Path:
    output_rows = _ensure_rank_rows(rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TopKhachHangDuNo"
    _write_rows(worksheet, output_rows, TOP_BALANCE_COLUMNS)
    workbook.save(destination)
    return Path(destination)


def export_top_customer_movement(
    repository: CustomerRepository,
    filters: CustomerFilters,
    previous_period: str,
    current_period: str,
    direction: str,
    limit: int,
    destination: Path,
) -> Path:
    direction = str(direction or "increase").casefold()
    rows = _rank_rows(
        repository.get_top_customer_movements(
            filters,
            previous_period,
            current_period,
            direction=direction,
            limit=limit,
        )
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TopGiamDuNo" if direction == "decrease" else "TopTangDuNo"
    _write_rows(worksheet, rows, TOP_MOVEMENT_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def export_top_customer_movement_rows(
    rows: Iterable[dict[str, object]],
    direction: str,
    destination: Path,
) -> Path:
    output_rows = _ensure_rank_rows(rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TopGiamDuNo" if str(direction or "").casefold() == "decrease" else "TopTangDuNo"
    _write_rows(worksheet, output_rows, TOP_MOVEMENT_COLUMNS)
    workbook.save(destination)
    return Path(destination)


def export_multiple_officers(
    repository: CustomerRepository,
    filters: CustomerFilters,
    destination: Path,
    *,
    sort_by: str = "total_balance",
    sort_desc: bool = True,
) -> Path:
    rows: list[dict[str, object]] = []
    page = 1
    while True:
        result = repository.multiple_officer_rows(
            filters,
            page=page,
            page_size=1000,
            sort_by=sort_by,
            sort_desc=sort_desc,
        )
        rows.extend(result.rows)
        if page * result.page_size >= result.total_rows:
            break
        page += 1
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "NhieuCanBoQuanLy"
    _write_rows(worksheet, rows, MULTIPLE_OFFICER_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def export_cross_branch_customers(
    repository: CustomerRepository,
    period: str,
    filters: CustomerFilters,
    minimum_branch_count: int,
    destination: Path,
    *,
    scope_type: object = "cross_branch",
    office_code: str = "",
    office_filter_mode: str = "actual",
    sort_by: str = "branch_count",
    sort_desc: bool = True,
) -> Path:
    rows = repository.all_cross_branch_customers(
        period,
        filters,
        minimum_branch_count=minimum_branch_count,
        scope_type=scope_type,
        office_code=office_code,
        office_filter_mode=office_filter_mode,
        sort_by=sort_by,
        sort_desc=sort_desc,
    )
    detail_rows: list[dict[str, object]] = []
    for row in rows:
        detail_rows.extend(
            repository.get_cross_branch_customer_detail(
                str(row.get("period") or period),
                str(row.get("customer_sequence") or ""),
            )
        )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "KhachHangLienChiNhanh"
    _write_rows(worksheet, rows, CROSS_BRANCH_COLUMNS, repository=repository)
    _write_rows(workbook.create_sheet("ChiTietTheoDonVi"), detail_rows, CROSS_BRANCH_DETAIL_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def export_cross_branch_customer_detail(
    repository: CustomerRepository,
    customer_sequence: str,
    destination: Path,
    *,
    period_from: str = "",
    period_to: str = "",
    report_period: str = "",
    branch_code: str = "",
    office_code: str = "",
    scope_type: str = "all",
) -> Path:
    if not report_period:
        periods = repository.customer_sequence_periods(customer_sequence)
        report_period = periods[-1] if periods else ""
    kpis = repository.get_cross_branch_customer_filtered_kpis(
        customer_sequence,
        report_period,
        branch_code=branch_code,
        office_code=office_code,
        scope_type=scope_type,
    )
    detail_rows = repository.get_cross_branch_customer_offices(
        customer_sequence,
        report_period,
        branch_code=branch_code,
        office_code=office_code,
        scope_type=scope_type,
    )
    history_rows = repository.get_cross_branch_customer_unit_history(
        customer_sequence,
        period_from,
        period_to,
        branch_code=branch_code,
        office_code=office_code,
        scope_type=scope_type,
    )
    workbook = Workbook()
    overview = workbook.active
    overview.title = "TongQuanLienChiNhanh"
    metadata = [
        {"metric": "Mã khách hàng gốc", "value": customer_sequence},
        {"metric": "Từ kỳ", "value": period_from or "Tất cả"},
        {"metric": "Đến kỳ", "value": period_to or "Tất cả"},
        {"metric": "Kỳ báo cáo", "value": report_period or "Tất cả"},
        {"metric": "Chi nhánh lọc", "value": _branch_filter_label(repository, branch_code)},
        {"metric": "PGD/Đơn vị lọc", "value": _office_filter_label(repository, office_code)},
        {"metric": "Loại phạm vi vay", "value": scope_type or "Tất cả"},
        {"metric": "Tên khách hàng", "value": kpis.get("customer_name", "")},
        {"metric": "Loại khách hàng", "value": kpis.get("customer_type_display", "")},
        {"metric": "Số chi nhánh vay", "value": kpis.get("branch_count", 0)},
        {"metric": "Số đơn vị vay", "value": kpis.get("office_count", 0)},
        {"metric": "Số Hội sở", "value": kpis.get("head_office_count", 0)},
        {"metric": "Số PGD", "value": kpis.get("pgd_count", 0)},
        {"metric": "Tổng dư nợ", "value": kpis.get("total_balance", 0)},
        {"metric": "Dư nợ Hội sở", "value": kpis.get("head_office_balance", 0)},
        {"metric": "Dư nợ PGD", "value": kpis.get("pgd_balance", 0)},
        {"metric": "Lãi suất bình quân", "value": kpis.get("average_rate", 0)},
        {"metric": "NIM trước ĐC", "value": kpis.get("nim_before", 0)},
        {"metric": "NIM sau ĐC", "value": kpis.get("nim_after", 0)},
    ]
    _write_rows(overview, metadata, (("metric", "Chỉ tiêu", "text"), ("value", "Giá trị", "text")))
    _write_rows(workbook.create_sheet("ChiTietTheoDonVi"), detail_rows, CROSS_BRANCH_DETAIL_COLUMNS, repository=repository)
    _write_rows(workbook.create_sheet("LichSuTheoKy"), history_rows, CROSS_BRANCH_HISTORY_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def export_import_history(repository: CustomerRepository, destination: Path) -> Path:
    runs = repository.import_runs(page=1, page_size=5000).rows
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "LichSuImport"
    _write_rows(worksheet, runs, IMPORT_RUN_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def export_officer_directory(repository: CustomerRepository, destination: Path) -> Path:
    rows = repository.officer_directory(page=1, page_size=5000).rows
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DanhMucCanBo"
    _write_rows(worksheet, rows, OFFICER_DIRECTORY_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def export_customer_detail(repository: CustomerRepository, customer_code: str, destination: Path) -> Path:
    workbook = Workbook()
    balance_sheet = workbook.active
    balance_sheet.title = "LichSuDuNo"
    history = repository.customer_history(customer_code)
    _write_rows(balance_sheet, history, DETAIL_BALANCE_COLUMNS)
    term_sheet = workbook.create_sheet("CoCauKyHan")
    _write_rows(term_sheet, history, DETAIL_BALANCE_COLUMNS[:5] + (("medium_long_ratio", "Tỷ lệ trung/dài hạn", "percent"),))
    nim_sheet = workbook.create_sheet("NIMLaiSuat")
    _write_rows(nim_sheet, history, DETAIL_NIM_COLUMNS)
    officer_sheet = workbook.create_sheet("LichSuCanBo")
    _write_rows(officer_sheet, repository.customer_officer_history(customer_code), DETAIL_OFFICER_COLUMNS)
    compare_sheet = workbook.create_sheet("SoSanhCacKy")
    _write_rows(compare_sheet, _detail_compare_rows(history), DETAIL_BALANCE_COLUMNS)
    workbook.save(destination)
    return Path(destination)


def export_all_customer_sheets(
    repository: CustomerRepository,
    filters: CustomerFilters,
    destination: Path,
    *,
    previous_period: str = "",
    current_period: str = "",
) -> Path:
    workbook = Workbook()
    dashboard_sheet = workbook.active
    dashboard_sheet.title = "TongQuanKhachHang"
    _write_dashboard_sheet(dashboard_sheet, repository, filters)
    _write_rows(
        workbook.create_sheet("DanhSachKhachHang"),
        repository.all_customer_rows(filters),
        CUSTOMER_LIST_COLUMNS,
        repository=repository,
    )
    previous_period, current_period = _resolve_compare_periods(repository, previous_period, current_period)
    growth_rows: list[dict[str, object]] = []
    if previous_period and current_period:
        growth_rows = repository.all_movement_rows(previous_period, current_period, filters)
    _write_rows(workbook.create_sheet("BienDongDuNo"), growth_rows, MOVEMENT_COLUMNS, repository=repository)
    multi_rows: list[dict[str, object]] = []
    page = 1
    while True:
        result = repository.multiple_officer_rows(filters, page=page, page_size=1000)
        multi_rows.extend(result.rows)
        if page * result.page_size >= result.total_rows:
            break
        page += 1
    _write_rows(workbook.create_sheet("NhieuCanBoQuanLy"), multi_rows, MULTIPLE_OFFICER_COLUMNS, repository=repository)
    _write_rows(workbook.create_sheet("LichSuImport"), repository.import_runs(page=1, page_size=5000).rows, IMPORT_RUN_COLUMNS, repository=repository)
    _write_rows(workbook.create_sheet("DanhMucCanBo"), repository.officer_directory(page=1, page_size=5000).rows, OFFICER_DIRECTORY_COLUMNS, repository=repository)
    workbook.save(destination)
    return Path(destination)


def _write_dashboard_sheet(worksheet, repository: CustomerRepository, filters: CustomerFilters) -> None:
    report_period = filters.current_period or filters.period_to
    metrics = repository.get_dashboard_kpis(filters, report_period)
    rows = [
        {"metric": "Số khách hàng còn dư nợ", "value": metrics.get("customer_count", 0)},
        {"metric": "Tổng dư nợ", "value": metrics.get("total_balance", 0)},
        {"metric": "Dư nợ ngắn hạn", "value": metrics.get("short_term_balance", 0)},
        {"metric": "Dư nợ trung/dài hạn", "value": metrics.get("medium_long_term_balance", 0)},
        {"metric": "Dư nợ chưa phân loại", "value": metrics.get("other_balance", 0)},
        {"metric": "Tỷ lệ trung/dài hạn", "value": metrics.get("medium_long_ratio", 0)},
        {"metric": "Lãi suất bình quân", "value": metrics.get("average_rate", 0)},
        {"metric": "NIM trước ĐC", "value": metrics.get("nim_before", 0)},
        {"metric": "NIM sau ĐC", "value": metrics.get("nim_after", 0)},
        {"metric": "Khách hàng nhiều cán bộ quản lý", "value": metrics.get("multiple_officer_customer_count", 0)},
        {"metric": "Khách hàng có override cán bộ", "value": metrics.get("override_customer_count", 0)},
    ]
    _write_rows(worksheet, rows, (("metric", "Chỉ tiêu", "text"), ("value", "Giá trị", "raw")))
    trends = _dashboard_export_trends(repository, filters)
    start_row = worksheet.max_row + 3
    worksheet.cell(start_row, 1, "Xu hướng theo kỳ")
    worksheet.cell(start_row, 1).font = Font(bold=True)
    headers = ("Kỳ", "Tổng dư nợ", "Số khách hàng còn dư nợ", "Lãi suất bình quân", "NIM trước ĐC", "NIM sau ĐC")
    for offset, header in enumerate(headers, start=1):
        worksheet.cell(start_row + 1, offset, header)
    for row_index, row in enumerate(trends, start=start_row + 2):
        values = [
            row.get("period", ""),
            _number(row.get("total_balance")),
            _number(row.get("active_customer_count")),
            _number(row.get("average_rate")),
            _number(row.get("nim_before")),
            _number(row.get("nim_after")),
        ]
        for col_index, value in enumerate(values, start=1):
            worksheet.cell(row_index, col_index, value)
    for cell in worksheet[start_row + 1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=start_row + 2, max_row=worksheet.max_row, min_col=2, max_col=2):
        row[0].number_format = MONEY_FORMAT
    for row in worksheet.iter_rows(min_row=start_row + 2, max_row=worksheet.max_row, min_col=3, max_col=3):
        row[0].number_format = "#,##0"
    for row in worksheet.iter_rows(min_row=start_row + 2, max_row=worksheet.max_row, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = PERCENT_FORMAT
    _style_sheet(worksheet)


def _dashboard_export_trends(repository: CustomerRepository, filters: CustomerFilters) -> list[dict[str, object]]:
    period_from = filters.period_from
    period_to = filters.period_to
    rows_by_period: dict[str, dict[str, object]] = {}
    for row in repository.get_total_balance_trend(filters, period_from, period_to, group_by="total"):
        period = str(row.get("period") or "")
        if period:
            rows_by_period.setdefault(period, {"period": period})["total_balance"] = row.get("value", 0)
    for row in repository.get_active_customer_count_trend(filters, period_from, period_to):
        period = str(row.get("period") or "")
        if period:
            rows_by_period.setdefault(period, {"period": period})["active_customer_count"] = row.get("active_customer_count", 0)
    for metric in ("average_rate", "nim_before", "nim_after"):
        for row in repository.get_customer_metric_trend(filters, period_from, period_to, metric=metric):
            period = str(row.get("period") or "")
            if period:
                rows_by_period.setdefault(period, {"period": period})[metric] = row.get("value", 0)
    return [rows_by_period[period] for period in sorted(rows_by_period)]


def _write_rows(
    worksheet,
    rows: Iterable[dict[str, object]],
    columns: tuple[ColumnSpec, ...],
    *,
    repository: CustomerRepository | None = None,
) -> None:
    rows = list(rows)
    worksheet.append([label for _field, label, _kind in columns])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        worksheet.append([_excel_value(row.get(field), kind, row=row, repository=repository) for field, _label, kind in columns])
    for column_index, (_field, _label, kind) in enumerate(columns, start=1):
        letter = get_column_letter(column_index)
        for cell in worksheet[letter][1:]:
            if kind in {"money", "money_or_blank"} and cell.value != "":
                cell.number_format = MONEY_FORMAT
            elif kind in {"percent", "percent_or_blank"} and cell.value != "":
                cell.number_format = PERCENT_FORMAT
            elif kind in {"text", "branch_display"}:
                cell.number_format = TEXT_FORMAT
        alignment = "right" if kind in {"money", "money_or_blank", "percent", "percent_or_blank", "integer", "raw"} else "left"
        if kind == "center":
            alignment = "center"
        for cell in worksheet[letter][1:]:
            cell.alignment = Alignment(horizontal=alignment, vertical="center", wrap_text=True)
    _style_sheet(worksheet)


def _rank_rows(rows: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    return [dict(row, rank=index) for index, row in enumerate(rows, start=1)]


def _ensure_rank_rows(rows: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    output_rows = [dict(row) for row in rows]
    if all("rank" in row for row in output_rows):
        return output_rows
    return _rank_rows(output_rows)


def _write_movement_metadata_sheet(
    worksheet,
    previous_period: str,
    current_period: str,
    filters: CustomerFilters,
    repository: CustomerRepository,
) -> None:
    rows = [
        ("Kỳ trước", previous_period),
        ("Kỳ hiện tại", current_period),
        ("Chi nhánh", _branch_filter_label(repository, filters.branch_code)),
        ("Loại khách hàng", format_customer_type(filters.customer_type) if filters.customer_type else "Tất cả"),
        ("Cán bộ", filters.officer or "Tất cả"),
        ("Loại biến động", filters.movement_status or "Tất cả"),
        ("Từ khóa tìm kiếm", filters.search_text or ""),
        ("Thời gian xuất", datetime.now().astimezone().isoformat(timespec="seconds")),
    ]
    worksheet.append(["Thông tin", "Giá trị"])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for label, value in rows:
        worksheet.append([label, value])
    _style_sheet(worksheet)


def _style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    if worksheet.max_column and worksheet.max_row:
        worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        width = min(48, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _excel_value(
    value: object,
    kind: str,
    *,
    row: dict[str, object] | None = None,
    repository: CustomerRepository | None = None,
) -> object:
    if kind == "customer_type":
        return format_customer_type(value)
    if kind == "branch_display":
        code = "" if value is None else str(value).strip()
        if repository is not None and code:
            return repository.unit_directory.get_branch_display_name(code)
        return code
    if kind == "active_status":
        return "Đang sử dụng" if int(value or 0) == 1 else "Ngừng sử dụng"
    if kind == "yes_no":
        return "Có" if int(value or 0) == 1 else "Không"
    if kind in {"money", "percent", "integer", "raw"}:
        return _number(value)
    if kind in {"money_or_blank", "percent_or_blank"}:
        return "" if value in (None, "") else _number(value)
    return "" if value is None else str(value)


def _branch_filter_label(repository: CustomerRepository, branch_code: object) -> str:
    code = "" if branch_code is None else str(branch_code).strip()
    if not code:
        return "Tất cả"
    return repository.unit_directory.get_branch_display_name(code)


def _office_filter_label(repository: CustomerRepository, office_code: object) -> str:
    code = "" if office_code is None else str(office_code).strip()
    if not code:
        return "Tất cả"
    office = repository.unit_directory.get_office_by_code(code)
    if office is not None:
        return repository.unit_directory.get_office_display_name(office.branch_code, office.trctcd)
    if "-" in code:
        branch, trctcd = code.split("-", 1)
        return repository.unit_directory.get_office_display_name(branch, trctcd)
    return code


def _number(value: object) -> float | int:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        return 0
    if number.is_integer():
        return int(number)
    return number


def _resolve_compare_periods(
    repository: CustomerRepository,
    previous_period: str,
    current_period: str,
) -> tuple[str, str]:
    periods = repository.distinct_periods()
    current = current_period or (periods[-1] if periods else "")
    previous = previous_period
    if current and not previous and current in periods:
        index = periods.index(current)
        previous = periods[index - 1] if index > 0 else ""
    return previous, current


def _detail_compare_rows(history: list[dict[str, object]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row in history:
        if row.get("difference") == "":
            continue
        rows.append(row)
    return rows
