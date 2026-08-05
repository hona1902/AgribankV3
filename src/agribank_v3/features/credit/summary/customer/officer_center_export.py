from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from agribank_v3.features.credit.summary.customer.export_service import (
    _write_rows,
    suggested_customer_export_name,
)
from agribank_v3.features.credit.summary.customer.officer_center_repository import (
    OfficerCenterFilters,
    OfficerCenterRepository,
)
from agribank_v3.features.credit.summary.customer.table_models import ColumnSpec


OFFICER_DASHBOARD_COLUMNS: tuple[ColumnSpec, ...] = (
    ("metric", "Chỉ tiêu", "text"),
    ("value", "Giá trị", "raw"),
    ("mode_label", "Chế độ phân tích", "text"),
)

OFFICER_LIST_COLUMNS: tuple[ColumnSpec, ...] = (
    ("rank", "STT", "integer"),
    ("officer_code", "Mã CBTD", "text"),
    ("officer_name", "Tên CBTD", "text"),
    ("branch_name", "Chi nhánh", "text"),
    ("office_name", "Phòng GD", "text"),
    ("officer_status", "Trạng thái", "text"),
    ("customer_count", "Số KH quản lý", "integer"),
    ("total_balance", "Tổng dư nợ", "money"),
    ("short_term_balance", "Dư nợ ngắn hạn", "term_money_or_dash"),
    ("medium_long_term_balance", "Dư nợ trung/dài hạn", "term_money_or_dash"),
    ("other_balance", "Dư nợ chưa phân loại", "term_money_or_dash"),
    ("medium_long_ratio", "Tỷ lệ trung/dài hạn", "term_percent_or_dash"),
    ("average_rate", "Lãi suất bình quân", "percent_or_blank"),
    ("nim_before", "NIM trước ĐC", "percent_or_blank"),
    ("nim_after", "NIM sau ĐC", "percent_or_blank"),
    ("attention_balance", "Nợ nhóm 2", "money"),
    ("bad_debt_balance", "Nợ xấu", "money"),
    ("attention_ratio", "Tỷ lệ nhóm 2", "percent_or_blank"),
    ("bad_debt_ratio", "Tỷ lệ nợ xấu", "percent_or_blank"),
    ("multiple_officer_customer_count", "Số KH nhiều CBTD", "integer"),
    ("override_customer_count", "Số KH có override", "integer"),
    ("data_warning", "Cảnh báo dữ liệu", "text"),
    ("mode_label", "Chế độ phân tích", "text"),
)

OFFICER_MOVEMENT_COLUMNS: tuple[ColumnSpec, ...] = (
    ("rank", "STT", "integer"),
    ("officer_code", "Mã CBTD", "text"),
    ("officer_name", "Tên CBTD", "text"),
    ("branch_name", "Chi nhánh", "text"),
    ("office_name", "Phòng GD", "text"),
    ("previous_balance", "Dư nợ kỳ trước", "money"),
    ("current_balance", "Dư nợ kỳ hiện tại", "money"),
    ("balance_change", "Tăng/giảm", "money_signed"),
    ("growth_rate", "Tăng trưởng", "percent_or_blank"),
    ("previous_customer_count", "KH kỳ trước", "integer"),
    ("current_customer_count", "KH kỳ hiện tại", "integer"),
    ("new_system_customer_count", "Mới phát sinh dư nợ", "integer"),
    ("paid_off_customer_count", "Tất toán", "integer"),
    ("transfer_in_customer_count", "Chuyển đến CBTD", "integer"),
    ("transfer_out_customer_count", "Chuyển đi CBTD", "integer"),
    ("attention_change", "Nợ nhóm 2 tăng/giảm", "money_signed"),
    ("bad_debt_change", "Nợ xấu tăng/giảm", "money_signed"),
    ("previous_nim_before", "NIM trước ĐC kỳ trước", "percent_or_blank"),
    ("previous_nim_after", "NIM sau ĐC kỳ trước", "percent_or_blank"),
    ("current_nim_before", "NIM trước ĐC kỳ hiện tại", "percent_or_blank"),
    ("current_nim_after", "NIM sau ĐC kỳ hiện tại", "percent_or_blank"),
    ("nim_before_change_pp", "Thay đổi NIM trước ĐC (điểm %)", "percent_point_signed"),
    ("nim_after_change_pp", "Thay đổi NIM sau ĐC (điểm %)", "percent_point_signed"),
)

OFFICER_COMPARE_COLUMNS: tuple[ColumnSpec, ...] = (
    *OFFICER_LIST_COLUMNS[:19],
    ("benchmark_nim_after", "NIM sau bình quân phạm vi lọc", "percent_or_blank"),
    ("nim_after_difference", "Chênh lệch NIM sau", "percent_signed"),
    ("benchmark_bad_debt_ratio", "Tỷ lệ nợ xấu bình quân", "percent_or_blank"),
    ("benchmark_officer_count", "Số CBTD so sánh", "integer"),
)

OFFICER_TOP_COLUMNS: tuple[ColumnSpec, ...] = (
    ("rank", "STT", "integer"),
    ("officer_code", "Mã CBTD", "text"),
    ("officer_name", "Tên CBTD", "text"),
    ("branch_name", "Chi nhánh", "text"),
    ("office_name", "Phòng GD", "text"),
    ("customer_count", "Số KH quản lý", "integer"),
    ("total_balance", "Dư nợ kỳ hiện tại", "money"),
    ("balance_change", "Tăng/giảm dư nợ", "money_or_blank"),
    ("nim_after", "NIM sau ĐC", "percent_or_blank"),
    ("attention_balance", "Nợ nhóm 2", "money"),
    ("bad_debt_balance", "Nợ xấu", "money"),
    ("bad_debt_ratio", "Tỷ lệ nợ xấu", "percent_or_blank"),
)

OFFICER_CUSTOMER_COLUMNS: tuple[ColumnSpec, ...] = (
    ("period", "Kỳ", "text"),
    ("customer_code", "Mã KH", "text"),
    ("customer_name", "Tên khách hàng", "text"),
    ("customer_type", "Loại KH", "customer_type"),
    ("branch_name", "Chi nhánh", "text"),
    ("office_name", "PGD", "text"),
    ("imported_officer_code", "Mã CBTD", "text"),
    ("imported_officer_name", "CBTD", "text"),
    ("total_customer_balance", "Tổng dư nợ KH", "money"),
    ("officer_balance", "Dư nợ thuộc CBTD", "money"),
    ("officer_share", "Tỷ trọng CBTD quản lý", "percent_or_blank"),
    ("short_term_balance", "Dư nợ ngắn hạn", "term_money_or_dash"),
    ("medium_long_term_balance", "Dư nợ trung/dài hạn", "term_money_or_dash"),
    ("other_balance", "Dư nợ chưa phân loại", "term_money_or_dash"),
    ("worst_debt_group", "Nhóm nợ cao nhất", "text"),
    ("attention_balance", "Nợ nhóm 2", "money"),
    ("bad_debt_balance", "Nợ xấu", "money"),
    ("average_rate", "Lãi suất bình quân", "percent_or_blank"),
    ("nim_before", "NIM trước", "percent_or_blank"),
    ("nim_after", "NIM sau", "percent_or_blank"),
    ("has_multiple_officers", "Nhiều CBTD", "yes_no"),
    ("has_override", "Có override", "yes_no"),
)


def export_officer_center_workbook(
    repository: OfficerCenterRepository,
    filters: OfficerCenterFilters,
    destination: Path,
) -> Path:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "TongQuanCBTD"
    payload = repository.dashboard_payload(filters)
    _write_rows(overview, _dashboard_rows(payload), OFFICER_DASHBOARD_COLUMNS)
    _write_rows(workbook.create_sheet("DanhSachCBTD"), _all_officer_rows(repository, filters), OFFICER_LIST_COLUMNS)
    _write_rows(workbook.create_sheet("BienDongCBTD"), _all_movement_rows(repository, filters), OFFICER_MOVEMENT_COLUMNS)
    _write_rows(workbook.create_sheet("SoSanhCBTD"), _all_compare_rows(repository, filters), OFFICER_COMPARE_COLUMNS)
    _write_rows(
        workbook.create_sheet("ChatLuongTinDung"),
        _all_officer_rows(repository, filters, sort_by="bad_debt_ratio"),
        OFFICER_LIST_COLUMNS,
    )
    _write_rows(workbook.create_sheet("KhachHangTheoCBTD"), _all_customer_rows(repository, filters), OFFICER_CUSTOMER_COLUMNS)
    _write_rows(workbook.create_sheet("DanhMucCBTD"), repository.repository.officer_directory(page=1, page_size=5000).rows, _directory_columns())
    _write_metadata(workbook.create_sheet("ThongTin"), repository, filters)
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def export_officer_top_workbook(
    rows: list[dict[str, object]] | tuple[dict[str, object], ...],
    destination: Path,
    *,
    metric_label: str,
    limit: int,
    mode_label: str = "",
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TopCBTD"
    _write_rows(worksheet, rows, OFFICER_TOP_COLUMNS)
    metadata = workbook.create_sheet("ThongTin")
    metadata.append(("Thông tin", "Giá trị"))
    metadata.append(("Chức năng", "Top CBTD"))
    metadata.append(("Chỉ tiêu", metric_label))
    metadata.append(("Số dòng Top", limit))
    metadata.append(("Chế độ phân tích", mode_label))
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def suggested_officer_export_name(tab_name: str = "QuanLyCBTD") -> str:
    return suggested_customer_export_name(tab_name)


def _dashboard_rows(payload: dict[str, object]) -> list[dict[str, object]]:
    kpis = dict(payload.get("kpis") or {})
    mode = str(payload.get("mode_label") or kpis.get("mode_label") or "")
    labels = (
        ("active_officer_count", "Số CBTD có dư nợ"),
        ("total_balance", "Tổng dư nợ"),
        ("officer_customer_occurrence_count", "Tổng lượt khách hàng theo CBTD"),
        ("unique_customer_count", "Số khách hàng duy nhất"),
        ("average_balance_per_officer", "Dư nợ bình quân/CBTD"),
        ("average_customer_per_officer", "Khách hàng bình quân/CBTD"),
        ("average_rate", "Lãi suất bình quân"),
        ("nim_before", "NIM trước ĐC"),
        ("nim_after", "NIM sau ĐC"),
        ("attention_balance", "Nợ cần chú ý"),
        ("bad_debt_balance", "Nợ xấu"),
        ("attention_ratio", "Tỷ lệ nợ cần chú ý"),
        ("bad_debt_ratio", "Tỷ lệ nợ xấu"),
        ("attention_officer_count", "Số CBTD có nợ nhóm 2"),
        ("bad_debt_officer_count", "Số CBTD có nợ xấu"),
    )
    return [{"metric": label, "value": kpis.get(key), "mode_label": mode} for key, label in labels]


def _all_officer_rows(
    repository: OfficerCenterRepository,
    filters: OfficerCenterFilters,
    *,
    sort_by: str = "total_balance",
) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    page = 1
    while True:
        result = repository.officer_list(filters, page=page, page_size=1000, sort_by=sort_by, sort_desc=True)
        output.extend(dict(row, rank=len(output) + index) for index, row in enumerate(result.rows, start=1))
        if page * result.page_size >= result.total_rows:
            return output
        page += 1


def _all_movement_rows(repository: OfficerCenterRepository, filters: OfficerCenterFilters) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    page = 1
    while True:
        result = repository.officer_movement(filters, page=page, page_size=1000)
        output.extend(dict(row, rank=len(output) + index) for index, row in enumerate(result.rows, start=1))
        if page * result.page_size >= result.total_rows:
            return output
        page += 1


def _all_compare_rows(repository: OfficerCenterRepository, filters: OfficerCenterFilters) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    page = 1
    while True:
        result = repository.compare_officers(filters, page=page, page_size=1000)
        output.extend(dict(row, rank=len(output) + index) for index, row in enumerate(result.rows, start=1))
        if page * result.page_size >= result.total_rows:
            return output
        page += 1


def _all_customer_rows(repository: OfficerCenterRepository, filters: OfficerCenterFilters) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    page = 1
    while True:
        result = repository.officer_customers(filters, page=page, page_size=1000)
        output.extend(result.rows)
        if page * result.page_size >= result.total_rows:
            return output
        page += 1


def _directory_columns() -> tuple[ColumnSpec, ...]:
    from agribank_v3.features.credit.summary.customer.export_service import OFFICER_DIRECTORY_COLUMNS

    return OFFICER_DIRECTORY_COLUMNS


def _write_metadata(worksheet, repository: OfficerCenterRepository, filters: OfficerCenterFilters) -> None:
    filters = filters.normalized()
    rows = (
        ("Chức năng", "Quản lý cán bộ tín dụng"),
        ("Chế độ phân tích", _mode_label_for_export(filters)),
        ("Kỳ báo cáo", filters.report_period or "Tất cả"),
        ("Từ kỳ", filters.period_from or "Tất cả"),
        ("Đến kỳ", filters.period_to or "Tất cả"),
        ("Chi nhánh", repository.unit_directory.get_branch_display_name(filters.branch_code) if filters.branch_code else "Tất cả"),
        ("Phòng GD", repository.unit_directory.get_office_name(filters.branch_code, filters.transaction_office) if filters.branch_code and filters.transaction_office else "Tất cả"),
        ("Loại KH", filters.customer_type or "Tất cả"),
        ("Loại thời hạn", filters.loan_term or "Tất cả"),
        ("Nhóm nợ", filters.debt_group or "Tất cả"),
        ("Trạng thái CBTD", filters.officer_status or "Tất cả"),
        ("Tìm kiếm", filters.search_text or ""),
        ("Số CBTD đã chọn", len(filters.selected_officers)),
    )
    worksheet.append(("Thông tin", "Giá trị"))
    for key, value in rows:
        worksheet.append((key, value))


def _mode_label_for_export(filters: OfficerCenterFilters) -> str:
    return dict((value, label) for label, value in (("Theo phân bổ dữ liệu import", "imported"), ("Theo cán bộ quản lý hiệu lực", "effective"))).get(
        filters.mode,
        filters.mode,
    )
