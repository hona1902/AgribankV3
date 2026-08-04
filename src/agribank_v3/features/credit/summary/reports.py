from __future__ import annotations

from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from agribank_v3.features.credit.summary.models import (
    CreditLimitRow,
    DashboardMetric,
    LoanSnapshotRow,
    SummaryDataType,
    SummaryError,
)


VBA_HEADER_FILL = PatternFill("solid", fgColor="C8C8C8")
GENERIC_MONEY_HEADERS = {
    "Dư nợ kỳ trước",
    "Dư nợ kỳ này",
    "Tăng/giảm",
    "Dư nợ",
    "Số dư nguồn vốn",
    "Tổng dư nợ",
    "Tổng dư nợ HĐTD",
    "Tổng dư nỢ HđTD",
    "Tổng hạn mức",
    "Hạn mức tín dụng",
    "Hạn mức TD",
}
GENERIC_COUNT_HEADERS = {
    "HĐTD đã hết hạn",
    "HĐTD sắp hết hạn",
    "Tổng HĐTD cảnh báo",
}
GENERIC_TEXT_HEADERS = {
    "Mã KH",
    "customer_code",
}

LOAN_COMPARE_VBA_HEADERS = (
    "Ma KH",
    "Ten KH",
    "Du no ky truoc",
    "Du no ky nay",
    "Loai KH",
    "CBTD quan ly",
)

CREDIT_LIMIT_VBA_HEADERS = (
    "Mã KH",
    "Tên KH",
    "Số HđTD",
    "Ngày HđTD",
    "Hạn mức TD",
    "Tổng dư nỢ HđTD",
    "Ngày hết hạn",
    "Địa chỉ",
    "Cán bộ TD",
    "Ghi chú",
)

NIM_DN_DETAIL_HEADERS = (
    "Kỳ Báo Cáo",
    "Chi nhánh (BRCD)",
    "Phòng Giao dịch (TRCTCD)",
    "Loại Khách Hàng",
    "Cán bộ tín dụng",
    "Lãi suất bình quân (%)",
    "NIM DN trước ĐC (%)",
    "NIM DN sau ĐC (%)",
)

NIM_DN_BRANCH_HEADERS = (
    "Kỳ Báo Cáo",
    "Chi nhánh",
    "Phòng Giao dịch",
    "Loại Khách Hàng",
    "Lãi suất bình quân (%)",
    "NIM DN trước ĐC (%)",
    "NIM DN sau ĐC (%)",
)

NIM_NV_DETAIL_HEADERS = (
    "Kỳ Báo Cáo",
    "Chi nhánh (BRCD)",
    "Phòng Giao dịch (TRCTCD)",
    "Loại Khách Hàng",
    "Cán bộ tín dụng",
    "NIM NV trước ĐC (%)",
    "NIM NV sau ĐC (%)",
)

NIM_NV_BRANCH_HEADERS = (
    "Kỳ Báo Cáo",
    "Chi nhánh",
    "Phòng Giao dịch",
    "Loại Khách Hàng",
    "NIM NV trước ĐC (%)",
    "NIM NV sau ĐC (%)",
)


def export_rows_to_excel(
    rows: list[dict[str, object]],
    destination: Path,
    *,
    sheet_name: str,
    title: str,
) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or "BaoCao"
    headers = list(rows[0].keys()) if rows else ["Thông báo"]
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    if rows:
        for row in rows:
            worksheet.append([_generic_excel_value(header, row.get(header, "")) for header in headers])
    else:
        worksheet.append(["Không có dữ liệu"])
    for column_index, header in enumerate(headers, start=1):
        if header in GENERIC_MONEY_HEADERS:
            for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=3):
                for item in cell:
                    item.number_format = "#,##0"
        elif header in GENERIC_TEXT_HEADERS:
            for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=3):
                for item in cell:
                    item.number_format = "@"
    for column_cells in worksheet.columns:
        width = min(42, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(destination)
    return destination


def _generic_excel_value(header: str, value: object) -> object:
    if header in GENERIC_TEXT_HEADERS:
        text = "" if value is None else str(value).strip()
        return text[1:] if text.startswith("'") else text
    if header in GENERIC_MONEY_HEADERS:
        return _number_for_excel(value)
    if header in GENERIC_COUNT_HEADERS:
        return int(round(_number_for_excel(value) or 0))
    return value


def export_credit_limit_view_report(
    rows: list[dict[str, object]],
    metrics: tuple[DashboardMetric, ...],
    destination: Path,
    *,
    sheet_name: str,
    title: str,
) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "TongHop"
    summary["A1"] = title
    summary["A1"].font = Font(bold=True, size=14)
    summary.append(("Chỉ tiêu", "Giá trị", "Ghi chú"))
    for cell in summary[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for metric in metrics:
        summary.append((metric.label, _generic_excel_value(metric.label, metric.value), metric.detail))
    for row in summary.iter_rows(min_row=3, min_col=2, max_col=2):
        for cell in row:
            label = summary.cell(cell.row, 1).value
            if label in GENERIC_MONEY_HEADERS:
                cell.number_format = "#,##0"
            elif label in GENERIC_COUNT_HEADERS:
                cell.number_format = "0"
    data_sheet = workbook.create_sheet((sheet_name[:31] or "BaoCao"))
    headers = list(rows[0].keys()) if rows else ["Thông báo"]
    data_sheet["A1"] = title
    data_sheet["A1"].font = Font(bold=True, size=14)
    data_sheet.append(headers)
    for cell in data_sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    if rows:
        for row in rows:
            data_sheet.append([_generic_excel_value(header, row.get(header, "")) for header in headers])
    else:
        data_sheet.append(["Không có dữ liệu"])
    for column_index, header in enumerate(headers, start=1):
        if header in GENERIC_MONEY_HEADERS:
            for cell in data_sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=3):
                for item in cell:
                    item.number_format = "#,##0"
        elif header in GENERIC_TEXT_HEADERS:
            for cell in data_sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=3):
                for item in cell:
                    item.number_format = "@"
    for worksheet in (summary, data_sheet):
        for column_cells in worksheet.columns:
            width = min(48, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(destination)
    return destination


def _number_for_excel(value: object) -> int | float:
    if value in (None, "", "—"):
        return 0
    if isinstance(value, int | float):
        number = float(value)
    else:
        text = str(value).strip().replace("đồng", "").replace("%", "").replace(" ", "")
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "." in text:
            parts = text.split(".")
            if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]) and len(parts[0]) <= 3:
                text = "".join(parts)
        elif "," in text:
            parts = text.split(",")
            if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]) and len(parts[0]) <= 3:
                text = "".join(parts)
            else:
                text = text.replace(",", ".")
        try:
            number = float(text)
        except ValueError:
            return 0
    return int(number) if number.is_integer() else number


def export_loan_compare_vba(rows: list[LoanSnapshotRow], destination: Path) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SoSanh_DuNo"
    worksheet.append(list(LOAN_COMPARE_VBA_HEADERS))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = VBA_HEADER_FILL
    for row in rows:
        worksheet.append(
            [
                row.customer_code,
                row.customer_name,
                row.previous_balance,
                row.current_balance,
                row.category,
                row.officer,
            ]
        )
    for column_letter in ("C", "D"):
        for cell in worksheet[column_letter]:
            cell.number_format = "#,##0"
    _autofit_like_excel(worksheet)
    workbook.create_sheet("Sheet2")
    workbook.create_sheet("Sheet3")
    workbook.save(destination)
    return destination


def export_credit_limit_vba(rows: list[CreditLimitRow], destination: Path) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "HanMuc_HetHan"
    worksheet.append(list(CREDIT_LIMIT_VBA_HEADERS))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = VBA_HEADER_FILL
    for row in rows:
        worksheet.append(
            [
                row.customer_code,
                row.customer_name,
                row.contract_number,
                row.approved_date,
                row.approved_amount,
                row.outstanding_balance,
                row.expiry_date,
                row.address,
                row.officer,
                row.note,
            ]
        )
    for column_letter in ("D", "G"):
        for cell in worksheet[column_letter]:
            cell.number_format = "dd/mm/yyyy"
    for column_letter in ("E", "F"):
        for cell in worksheet[column_letter]:
            cell.number_format = "#,##0"
    _autofit_like_excel(worksheet)
    workbook.save(destination)
    return destination


def export_nim_vba(
    detail_rows: list[list[object]],
    branch_rows: list[list[object]],
    destination: Path,
    *,
    data_type: SummaryDataType,
) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Sheet1"
    workbook.create_sheet("Sheet2")
    workbook.create_sheet("Sheet3")
    cache = workbook.create_sheet(_nim_cache_sheet(data_type), 0)
    cache.sheet_state = "hidden"
    report = workbook.create_sheet(_nim_report_sheet(data_type), 1)

    detail_headers, branch_headers = _nim_headers(data_type)
    for column_index, value in enumerate(detail_headers, start=1):
        cache.cell(1, column_index, value)
    for row_index, row in enumerate(detail_rows, start=2):
        for column_index, value in enumerate(row, start=1):
            cache.cell(row_index, column_index, value)
    for column_index, value in enumerate(branch_headers, start=14):
        cache.cell(1, column_index, value)
    for row_index, row in enumerate(branch_rows, start=2):
        for column_index, value in enumerate(row, start=14):
            cache.cell(row_index, column_index, value)

    sorted_detail = _sort_nim_detail(detail_rows)
    sorted_branch = _sort_nim_branch(branch_rows)
    report["A1"] = "BỘ LỌC DỮ LIỆU"
    report["A2"] = "▼ Kỳ Báo Cáo"
    report["B2"] = "▼ Chi nhánh"
    report["D2"] = "▼ Phòng Giao Dịch"
    report["F2"] = "▼ Cán bộ tín dụng"
    report["A14"] = f"BÁO CÁO TỔNG HỢP NIM {_nim_kind(data_type)} CHI TIẾT"
    report["J14"] = f"NIM {_nim_kind(data_type)} TỔNG HỢP THEO CHI NHÁNH"
    for column_index, value in enumerate(detail_headers, start=1):
        report.cell(15, column_index, value)
    for row_index, row in enumerate(sorted_detail, start=16):
        for column_index, value in enumerate(row, start=1):
            report.cell(row_index, column_index, value)
    for column_index, value in enumerate(branch_headers, start=10):
        report.cell(15, column_index, value)
    for row_index, row in enumerate(sorted_branch, start=16):
        for column_index, value in enumerate(row, start=10):
            report.cell(row_index, column_index, value)

    for worksheet in (cache, report):
        _style_nim_worksheet(worksheet, data_type)
        _autofit_like_excel(worksheet)
    workbook.save(destination)
    return destination


def export_rows_to_csv(rows: list[dict[str, object]], destination: Path) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys()) if rows else ["Thông báo"]
    with destination.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        if rows:
            writer.writerows(rows)
        else:
            writer.writerow({"Thông báo": "Không có dữ liệu"})
    return destination


def export_rows_to_pdf(
    rows: list[dict[str, object]],
    destination: Path,
    *,
    title: str,
    max_rows: int = 80,
) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    lines = [title, ""]
    headers = list(rows[0].keys()) if rows else ["Thông báo"]
    lines.append(" | ".join(headers[:8]))
    for row in rows[:max_rows]:
        lines.append(" | ".join(_pdf_text(row.get(header, "")) for header in headers[:8]))
    if len(rows) > max_rows:
        lines.append(f"... còn {len(rows) - max_rows:,} dòng, hãy xuất Excel/CSV để xem đầy đủ.")
    if not rows:
        lines.append("Không có dữ liệu")
    _write_minimal_pdf(destination, lines)
    return destination


def export_rows(
    rows: list[dict[str, object]],
    destination: Path,
    *,
    title: str,
    sheet_name: str = "BaoCao",
) -> Path:
    suffix = Path(destination).suffix.casefold()
    if suffix == ".xlsx":
        return export_rows_to_excel(rows, destination, sheet_name=sheet_name, title=title)
    if suffix == ".csv":
        return export_rows_to_csv(rows, destination)
    if suffix == ".pdf":
        return export_rows_to_pdf(rows, destination, title=title)
    raise SummaryError("Chỉ hỗ trợ xuất .xlsx, .csv hoặc .pdf.")


def _write_minimal_pdf(destination: Path, lines: list[str]) -> None:
    content_lines = ["BT", "/F1 10 Tf", "40 800 Td"]
    first = True
    for line in lines:
        if first:
            first = False
        else:
            content_lines.append("0 -14 Td")
        content_lines.append(f"({_escape_pdf(line[:135])}) Tj")
    content_lines.append("ET")
    stream = "\n".join(content_lines).encode("utf-8", errors="replace")
    objects: list[bytes] = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    objects.append(b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>")
    objects.append(
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>"
    )
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    objects.append(b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream")
    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{index} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")
    xref_offset = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    destination.write_bytes(output)


def _escape_pdf(text: object) -> str:
    return _pdf_text(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_text(text: object) -> str:
    return str(text or "").replace("\r", " ").replace("\n", " ")


def _autofit_like_excel(worksheet) -> None:
    for column_cells in worksheet.columns:
        width = min(60, max(8, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _nim_headers(data_type: SummaryDataType) -> tuple[tuple[str, ...], tuple[str, ...]]:
    if data_type == SummaryDataType.NIM_DN:
        return NIM_DN_DETAIL_HEADERS, NIM_DN_BRANCH_HEADERS
    return NIM_NV_DETAIL_HEADERS, NIM_NV_BRANCH_HEADERS


def _nim_kind(data_type: SummaryDataType) -> str:
    return "DN" if data_type == SummaryDataType.NIM_DN else "NV"


def _nim_cache_sheet(data_type: SummaryDataType) -> str:
    return "Cache_Nim" if data_type == SummaryDataType.NIM_DN else "Cache_Nim_NV"


def _nim_report_sheet(data_type: SummaryDataType) -> str:
    return "Báo Cáo NIM DN" if data_type == SummaryDataType.NIM_DN else "Báo Cáo NIM NV"


def _sort_nim_detail(rows: list[list[object]]) -> list[list[object]]:
    sorted_rows = list(rows)
    sorted_rows.sort(key=lambda row: _sort_text(row[4]))
    sorted_rows.sort(key=lambda row: _sort_text(row[3]))
    sorted_rows.sort(key=lambda row: _sort_text(row[2]))
    sorted_rows.sort(key=lambda row: _sort_text(row[1]))
    sorted_rows.sort(key=lambda row: _sort_text(row[0]), reverse=True)
    return sorted_rows


def _sort_nim_branch(rows: list[list[object]]) -> list[list[object]]:
    sorted_rows = list(rows)
    sorted_rows.sort(key=lambda row: _sort_text(row[3]))
    sorted_rows.sort(key=lambda row: _sort_text(row[2]), reverse=True)
    sorted_rows.sort(key=lambda row: _sort_text(row[1]))
    sorted_rows.sort(key=lambda row: _sort_text(row[0]), reverse=True)
    return sorted_rows


def _sort_text(value: object) -> str:
    return str(value or "").casefold()


def _style_nim_worksheet(worksheet, data_type: SummaryDataType) -> None:
    for cell in worksheet[1]:
        if cell.value is not None:
            cell.font = Font(bold=True)
    for row_index in (14, 15):
        for cell in worksheet[row_index]:
            if cell.value is not None:
                cell.font = Font(bold=True)
    worksheet["A1"].font = Font(bold=True, size=16, color="0070C0")
    is_cache = worksheet.title.startswith("Cache_")
    if data_type == SummaryDataType.NIM_DN and is_cache:
        numeric_columns = ("F", "G", "H", "R", "S", "T")
    elif data_type == SummaryDataType.NIM_DN:
        numeric_columns = ("F", "G", "H", "N", "O", "P")
    elif is_cache:
        numeric_columns = ("F", "G", "R", "S")
    else:
        numeric_columns = ("F", "G", "N", "O")
    for column_letter in numeric_columns:
        for cell in worksheet[column_letter]:
            cell.number_format = "0.00"
