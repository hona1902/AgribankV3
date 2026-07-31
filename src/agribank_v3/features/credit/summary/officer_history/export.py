from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


SHEET_NAMES = {
    "overview": "TongQuanNIM",
    "growth": "DuNoTangTruong",
    "officer_compare": "SoSanhCBTD",
    "branch_compare": "SoSanhChiNhanh",
}


def export_analysis_rows(
    rows: list[dict[str, object]],
    destination: Path,
    *,
    tab_key: str,
    metadata: dict[str, object] | None = None,
) -> Path:
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAMES.get(tab_key, "TongQuanNIM")
    headers = list(rows[0].keys()) if rows else ["Thông báo"]
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    if rows:
        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])
    else:
        worksheet.append(["Không có dữ liệu"])
    for column_cells in worksheet.columns:
        width = min(42, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        worksheet.column_dimensions[column_cells[0].column_letter].width = width
    if metadata:
        info = workbook.create_sheet("ThongTin")
        info.append(["Thông tin", "Giá trị"])
        for cell in info[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for key, value in metadata.items():
            info.append([key, value])
        for column_cells in info.columns:
            width = min(60, max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2))
            info.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(destination)
    return destination
