from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
import csv
import os
from typing import Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xlrd import open_workbook, xldate_as_datetime

from agribank_v3.features.credit.auto_interest.settings import (
    AutoInterestSettings,
)


class AutoInterestError(RuntimeError):
    pass


COLLECT_ALL_INTEREST = "all_interest"
NOT_DUE_INTEREST = "not_due"
OVERDUE_CENTER_INTEREST = "overdue_center"
NOT_DUE_AND_OVERDUE_INTEREST = "not_due_and_overdue"
LEGACY_DUE_IN_MONTH = "due_in_month"

COLLECTION_MODE_LABELS: dict[str, str] = {
    COLLECT_ALL_INTEREST: "Thu toàn bộ lãi",
    NOT_DUE_INTEREST: "Chỉ thu lãi chưa đến hạn trong tháng",
    OVERDUE_CENTER_INTEREST: "Chỉ thu lãi đến hạn (đã qua ngày chạy tự động)",
    NOT_DUE_AND_OVERDUE_INTEREST: "Thu lãi chưa đến hạn trong tháng và đã quá hạn chạy tự động",
}

COLLECTION_MODE_VBA_MAP: dict[str, str] = {
    COLLECT_ALL_INTEREST: "ThuToanBoLai",
    NOT_DUE_INTEREST: "ThuLaiKoDenHan",
    OVERDUE_CENTER_INTEREST: "ThuLaiQuaHanCenTer",
    NOT_DUE_AND_OVERDUE_INTEREST: "ThuLaiChuaDenHanQuaHan",
}

LOAN_REQUIRED_HEADERS = (
    "chonIn",
    "STT",
    "MaKH",
    "TenKH",
    "NhomNo",
    "MaGN",
    "CCY",
)
LOAN_OPTIONAL_HEADERS = (
    "DuNo",
    "LaiDK",
    "CBTD",
    "NgayLaiC",
    "addr",
    "nxtintrpmt",
)
MSIT_REQUIRED_HEADERS = (
    "Customer_No",
    "Customer_Name",
    "DP_TypeName",
    "Ccy",
    "Curent_Balance",
    "Rate",
    "Account_Number",
)
MSIT_OPTIONAL_HEADERS = ("Tel", "Maturity_Date")
DPDA_REQUIRED_HEADERS = (
    "idxacno",
    "custseq",
    "custnm",
    "stscd",
    "custid",
    "ccycd",
    "intrt",
    "curbal",
)
COLLATERAL_REQUIRED_HEADERS = (
    "clno",
    "clcustno",
    "clcustnm",
    "cltpcd",
    "cldtltpcd",
    "acctkey",
)

LOAN_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "print_flag": ("chonIn", "chon in"),
    "sequence": ("STT", "stt"),
    "customer_code": ("MaKH", "Mã KH", "Ma KH", "Customer_No", "Customer No"),
    "customer_name": ("TenKH", "Tên KH", "Ten KH", "Customer_Name", "Customer Name"),
    "debt_group": ("NhomNo", "Nhóm nợ", "Nhom No"),
    "loan_number": ("MaGN", "Mã GN", "Ma GN", "SoGiaiNgan", "Số giải ngân", "Disburse No"),
    "currency": ("CCY", "Ccy", "Loại tiền"),
    "principal_balance": ("DuNo", "Dư nợ", "Du No", "Principal Balance"),
    "interest_due": ("LaiDK", "Lãi DK", "Lai DK", "Interest Due", "Total", "Sum of LaiDK"),
    "staff": ("CBTD", "Cán bộ tín dụng"),
    "last_interest_date": ("NgayLaiC", "Ngày lãi C", "Ngay Lai C"),
    "address": ("addr", "DiaChi", "Địa chỉ"),
    "next_interest_payment_date": (
        "nxtintrpmt",
        "Next Interest Payment",
        "Next Interest Payment Date",
        "NgayDenHan",
        "Ngày đến hạn",
        "NgayTraLaiTiepTheo",
    ),
    "interest_rate": ("LaiSuat", "Lãi suất", "LS", "Rate", "intrt"),
    "interest_start_date_1": (
        "NgayTinhLaiTu1",
        "NgayTraGocCuoiCung",
        "Last Principal Payment Date",
    ),
    "interest_start_date_2": (
        "NgayTinhLaiTu2",
        "NgayTraLaiCuoiCung",
        "Last Interest Payment Date",
    ),
}

LOAN_COLUMN_FALLBACKS: dict[str, int] = {
    "customer_code": 2,  # VBA FileGoc column C
    "customer_name": 3,  # D
    "loan_number": 5,  # F
    "principal_balance": 7,  # H
    "interest_due": 8,  # I
    "staff": 9,  # J in many lnlr23 exports
    "next_interest_payment_date": 20,  # U
    "interest_start_date_1": 17,  # R
    "interest_start_date_2": 18,  # S
    "interest_rate": 30,  # AE
}

DEPOSIT_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "customer_code": ("Customer_No", "Customer No", "custseq", "MaKH", "Mã KH"),
    "customer_name": ("Customer_Name", "Customer Name", "custnm", "TenKH", "Tên KH"),
    "account_number": ("Account_Number", "Account Number", "idxacno", "SoTK", "Số TK"),
    "balance": ("Curent_Balance", "Current_Balance", "curbal", "Balance", "Số dư"),
    "maturity_date": ("Maturity_Date", "Maturity Date", "NgayDaoHan", "Ngày đáo hạn"),
}
COLLATERAL_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "customer_code": ("clcustno", "MaKH", "Mã KH", "Customer_No", "Customer No"),
    "account_key": ("acctkey", "Account Key"),
}

COLLATERAL_COLUMN_FALLBACKS: dict[str, int] = {
    "customer_code": 1,  # B
    "account_key": 28,  # AC
}

OUTPUT_HEADERS = (
    "STT",
    "MaKH",
    "TenKH",
    "SoGiaiNgan",
    "ChuTK",
    "LaiDK",
    "ThuLaiDen",
    "ThuTuTK",
)
LEDGER_HEADERS = (
    "STT",
    "Mã KH",
    "Tên KH",
    "Số giải ngân",
    "Chủ TK",
    "Lãi DK",
    "Thu lãi đến",
    "Thu từ TK",
)
REPORT_HEADERS = (
    "Disburse No(1)",
    "Principle/ Int Type(2)",
    "Principle Amt(3)",
    "Interest to (4)",
    "Settlement type (5)",
    "TEN",
    "NGAY THU NO",
)


@dataclass(frozen=True, slots=True)
class AutoInterestInput:
    loan_file: Path
    deposit_file: Path
    collection_date: date
    deposit_statement_type: str
    collection_mode: str
    output_folder: Path
    include_weekend_interest: bool = True
    collateral_file: Path | None = None


AutoInterestCreateRequest = AutoInterestInput


@dataclass(frozen=True, slots=True)
class AutoInterestReportRequest:
    source_file: Path
    settings: AutoInterestSettings
    report_date: date
    collection_mode: str | None = None


@dataclass(frozen=True, slots=True)
class AutoInterestResult:
    output_file: Path
    row_count: int
    skipped_count: int = 0
    warnings: tuple[str, ...] = ()
    summary: Mapping[str, object] = field(default_factory=dict)


@dataclass(frozen=True, slots=True)
class TabularData:
    headers: list[str]
    rows: list[tuple[object, ...]]


@dataclass(frozen=True, slots=True)
class ResolvedColumns:
    positions: dict[str, int]
    recognized: dict[str, str]


@dataclass(frozen=True, slots=True)
class DepositAccount:
    customer_code: str
    account_number: str
    balance: float


@dataclass(frozen=True, slots=True)
class InterestRow:
    source_row: tuple[object, ...]
    customer_code: str
    customer_name: str
    loan_number: str
    principal_balance: float
    interest_due: float
    next_interest_payment_date: date | None
    collection_interest_amount: float
    interest_to_date: date
    account_number: str = ""
    skipped_reason: str = ""


def default_interest_file_name(
    today: date | None = None,
    collection_mode: str = COLLECT_ALL_INTEREST,
) -> str:
    current = today or date.today()
    prefix = {
        COLLECT_ALL_INTEREST: "ThuToanBoLai",
        NOT_DUE_INTEREST: "ThuLaiKhongDenHan",
        OVERDUE_CENTER_INTEREST: "ThuLaiQuaHanCenter",
        NOT_DUE_AND_OVERDUE_INTEREST: "ThuLaiChuaDenHanQuaHan",
        LEGACY_DUE_IN_MONTH: "ThuLaiQuaHanCenter",
    }.get(collection_mode, "ThuLaiBanTuDong")
    return f"{prefix}_{current:%Y%m%d}.xlsx"


def validate_auto_interest_inputs(
    loan_file: Path,
    deposit_file: Path,
    deposit_statement_type: str,
    collateral_file: Path | None = None,
) -> tuple[str, ...]:
    warnings: list[str] = []
    loan_data = load_loan_statement(Path(loan_file))
    missing_loan = _missing_headers(loan_data.headers, LOAN_REQUIRED_HEADERS)
    if missing_loan:
        raise AutoInterestError(
            "File sao kê lãi dự kiến không đúng cấu trúc Loan/lnlr23. "
            "Thiếu cột: " + ", ".join(missing_loan)
        )
    resolved_loan = validate_loan_columns(loan_data)
    required_business_fields = ("customer_code", "customer_name", "loan_number", "interest_due")
    missing_business_fields = [
        field_name for field_name in required_business_fields if field_name not in resolved_loan.positions
    ]
    if missing_business_fields:
        raise AutoInterestError(
            "File sao kê lãi dự kiến thiếu cột nghiệp vụ: "
            + ", ".join(missing_business_fields)
        )
    optional_loan = _missing_headers(loan_data.headers, LOAN_OPTIONAL_HEADERS)
    if optional_loan:
        warnings.append(
            "File sao kê lãi dự kiến thiếu cột tùy chọn: "
            + ", ".join(optional_loan)
        )

    deposit_data = load_deposit_statement(Path(deposit_file))
    statement_type = _normalize_deposit_statement_type(deposit_statement_type)
    if statement_type == "msit":
        required = MSIT_REQUIRED_HEADERS
        optional = MSIT_OPTIONAL_HEADERS
    else:
        required = DPDA_REQUIRED_HEADERS
        optional = ()
    missing_deposit = _missing_headers(deposit_data.headers, required)
    if missing_deposit:
        raise AutoInterestError(
            "File sao kê tiền gửi không đúng cấu trúc "
            + statement_type.upper()
            + ". Thiếu cột: "
            + ", ".join(missing_deposit)
        )
    validate_deposit_columns(deposit_data, statement_type)
    optional_deposit = _missing_headers(deposit_data.headers, optional)
    if optional_deposit:
        warnings.append(
            "File sao kê tiền gửi thiếu cột tùy chọn: "
            + ", ".join(optional_deposit)
        )
    if collateral_file is not None:
        collateral_data = load_collateral_statement(Path(collateral_file))
        missing_collateral = _missing_headers(
            collateral_data.headers,
            COLLATERAL_REQUIRED_HEADERS,
        )
        if missing_collateral:
            raise AutoInterestError(
                "File sao kê tài sản bảo đảm không đúng cấu trúc. Thiếu cột: "
                + ", ".join(missing_collateral)
            )
        validate_collateral_columns(collateral_data)
    return tuple(warnings)


def load_loan_statement(path: Path | str) -> TabularData:
    headers, rows = _read_tabular_file(Path(path))
    return TabularData(headers=headers, rows=rows)


def load_deposit_statement(path: Path | str) -> TabularData:
    headers, rows = _read_tabular_file(Path(path))
    return TabularData(headers=headers, rows=rows)


def load_collateral_statement(path: Path | str) -> TabularData:
    headers, rows = _read_tabular_file(Path(path))
    return TabularData(headers=headers, rows=rows)


def validate_loan_columns(data: TabularData) -> ResolvedColumns:
    return _resolve_columns(data.headers, LOAN_COLUMN_ALIASES, LOAN_COLUMN_FALLBACKS)


def validate_deposit_columns(
    data: TabularData,
    statement_type: str,
) -> ResolvedColumns:
    statement = _normalize_deposit_statement_type(statement_type)
    resolved = _resolve_columns(data.headers, DEPOSIT_COLUMN_ALIASES)
    required = ("customer_code", "account_number", "balance")
    missing = [field_name for field_name in required if field_name not in resolved.positions]
    if missing:
        raise AutoInterestError(
            f"File sao kê tiền gửi {statement.upper()} thiếu cột nghiệp vụ: "
            + ", ".join(missing)
        )
    return resolved


def validate_collateral_columns(data: TabularData) -> ResolvedColumns:
    resolved = _resolve_columns(
        data.headers,
        COLLATERAL_COLUMN_ALIASES,
        COLLATERAL_COLUMN_FALLBACKS,
    )
    required = ("customer_code", "account_key")
    missing = [field_name for field_name in required if field_name not in resolved.positions]
    if missing:
        raise AutoInterestError(
            "File sao kê tài sản bảo đảm thiếu cột nghiệp vụ: "
            + ", ".join(missing)
        )
    return resolved


def build_collect_all_interest_file(request: AutoInterestInput) -> AutoInterestResult:
    return _build_interest_file(request, COLLECT_ALL_INTEREST)


def build_not_due_interest_file(request: AutoInterestInput) -> AutoInterestResult:
    return _build_interest_file(request, NOT_DUE_INTEREST)


def build_overdue_center_interest_file(request: AutoInterestInput) -> AutoInterestResult:
    return _build_interest_file(request, OVERDUE_CENTER_INTEREST)


def build_not_due_and_overdue_interest_file(request: AutoInterestInput) -> AutoInterestResult:
    return _build_interest_file(request, NOT_DUE_AND_OVERDUE_INTEREST)


def create_auto_interest_file(request: AutoInterestCreateRequest) -> AutoInterestResult:
    mode = _normalize_collection_mode(request.collection_mode)
    normalized_request = AutoInterestInput(
        loan_file=Path(request.loan_file),
        deposit_file=Path(request.deposit_file),
        collection_date=request.collection_date,
        deposit_statement_type=request.deposit_statement_type,
        collection_mode=mode,
        output_folder=Path(request.output_folder),
        collateral_file=Path(request.collateral_file) if request.collateral_file else None,
        include_weekend_interest=request.include_weekend_interest,
    )
    if mode == COLLECT_ALL_INTEREST:
        return build_collect_all_interest_file(normalized_request)
    if mode == NOT_DUE_INTEREST:
        return build_not_due_interest_file(normalized_request)
    if mode == OVERDUE_CENTER_INTEREST:
        return build_overdue_center_interest_file(normalized_request)
    if mode == NOT_DUE_AND_OVERDUE_INTEREST:
        return build_not_due_and_overdue_interest_file(normalized_request)
    raise AutoInterestError("Hình thức thu lãi không hợp lệ.")


def create_auto_interest_report(request: AutoInterestReportRequest) -> AutoInterestResult:
    return build_auto_interest_report(request)


def build_auto_interest_report(request: AutoInterestReportRequest) -> AutoInterestResult:
    if not request.source_file.exists():
        raise AutoInterestError(f"Không tìm thấy file nguồn: {request.source_file}")
    workbook = load_workbook(request.source_file, data_only=True)
    try:
        if "SaoKeTrichLai" not in workbook.sheetnames:
            raise AutoInterestError(
                "File nguồn chưa có sheet SaoKeTrichLai. Vui lòng tạo file thu lãi trước."
            )
        month_folder = request.settings.report_folder / f"{request.report_date:%m-%Y}"
        month_folder.mkdir(parents=True, exist_ok=True)
        base_name = (
            f"{os.environ.get('USERNAME', 'AgribankV3')} - "
            f"{datetime.now():%d-%m-%Y %H-%M-%S}.xlsx"
        )
        output_file = _unique_output_path(month_folder / base_name)

        report = Workbook()
        sheet = report.active
        sheet.title = "BaoCaoThuBanTuDong"
        sheet.append(REPORT_HEADERS)
        collection_mode = _resolve_report_collection_mode(
            request.source_file,
            workbook,
            request.collection_mode,
        )
        rows_written, total_interest = _append_report_rows_from_interest_workbook(
            workbook,
            sheet,
            request.report_date,
            collection_mode,
        )
        _style_report_sheet(sheet)
        report.save(output_file)
        return AutoInterestResult(
            output_file=output_file,
            row_count=rows_written,
            summary={
                "mode": collection_mode,
                "mode_label": COLLECTION_MODE_LABELS.get(collection_mode, ""),
                "total_interest": total_interest,
            },
        )
    finally:
        workbook.close()


def _build_interest_file(
    request: AutoInterestInput,
    collection_mode: str,
) -> AutoInterestResult:
    warnings = list(
        validate_auto_interest_inputs(
            request.loan_file,
            request.deposit_file,
            request.deposit_statement_type,
            request.collateral_file,
        )
    )
    loan_data = load_loan_statement(request.loan_file)
    deposit_data = load_deposit_statement(request.deposit_file)
    loan_columns = validate_loan_columns(loan_data)
    deposit_columns = validate_deposit_columns(
        deposit_data,
        request.deposit_statement_type,
    )
    deposits = _build_deposit_account_lookup(
        deposit_data,
        deposit_columns,
        request.deposit_statement_type,
    )
    all_interest_rows = _loan_rows_for_mode(
        loan_data,
        loan_columns,
        collection_mode,
        request.collection_date,
        request.include_weekend_interest,
        warnings,
    )
    pledged_customer_codes: set[str] = set()
    if request.collateral_file is not None:
        collateral_data = load_collateral_statement(request.collateral_file)
        collateral_columns = validate_collateral_columns(collateral_data)
        pledged_customer_codes = _pledged_collateral_customer_codes(
            collateral_data,
            collateral_columns,
        )
    interest_rows, pledged_skipped_rows = _filter_pledged_collateral(
        all_interest_rows,
        pledged_customer_codes,
    )
    selected_rows, skipped_rows = _filter_by_deposit_balance(
        interest_rows,
        deposits,
        collection_mode,
    )
    skipped_rows = pledged_skipped_rows + skipped_rows
    if pledged_skipped_rows:
        warnings.append(
            f"Đã bỏ qua {len(pledged_skipped_rows)} khoản vay của "
            f"{len(pledged_customer_codes)} khách hàng có tài sản cầm cố 994003."
        )
    request.output_folder.mkdir(parents=True, exist_ok=True)
    output_file = _unique_output_path(
        request.output_folder
        / default_interest_file_name(request.collection_date, collection_mode)
    )

    workbook = Workbook()
    file_sheet = workbook.active
    file_sheet.title = "FileGoc"
    _write_source_sheet(file_sheet, loan_data.headers, loan_data.rows)

    technical_sheet = workbook.create_sheet("SaoKeTrichLai")
    _write_auto_interest_technical_sheet(
        technical_sheet,
        selected_rows,
    )
    ledger_sheet = workbook.create_sheet("BangKeTheoLo")
    _write_auto_interest_ledger_sheet(
        ledger_sheet,
        selected_rows,
        collection_mode,
        request.collection_date,
    )
    workbook.save(output_file)

    total_interest = sum(row.collection_interest_amount for row in selected_rows)
    return AutoInterestResult(
        output_file=output_file,
        row_count=len(selected_rows),
        skipped_count=len(skipped_rows),
        warnings=tuple(warnings),
        summary={
            "mode": collection_mode,
            "mode_label": COLLECTION_MODE_LABELS[collection_mode],
            "vba_procedure": COLLECTION_MODE_VBA_MAP[collection_mode],
            "total_interest": total_interest,
            "customers": len({row.customer_code for row in selected_rows}),
            "pledged_collateral_customers": len(pledged_customer_codes),
            "pledged_collateral_rows": len(pledged_skipped_rows),
        },
    )


def _loan_rows_for_mode(
    data: TabularData,
    columns: ResolvedColumns,
    collection_mode: str,
    collection_date: date,
    include_weekend_interest: bool,
    warnings: list[str],
) -> list[InterestRow]:
    rows: list[InterestRow] = []
    today = date.today()
    month_end = _end_of_month(today)
    warned_overdue_fallback = False
    warned_missing_due_date = False
    for source_row in data.rows:
        customer_code = _cell_text(source_row, columns, "customer_code")
        if not customer_code:
            continue
        interest_due = _cell_number(source_row, columns, "interest_due")
        if interest_due < 0 or (collection_mode != COLLECT_ALL_INTEREST and interest_due <= 0):
            continue
        due_date = _cell_date(source_row, columns, "next_interest_payment_date")
        if collection_mode == NOT_DUE_INTEREST:
            if due_date is None:
                warned_missing_due_date = True
                continue
            if due_date <= collection_date:
                continue
            interest_amount = interest_due
            interest_to_date = collection_date
        elif collection_mode == OVERDUE_CENTER_INTEREST:
            if due_date is None:
                warned_missing_due_date = True
                continue
            if due_date > today:
                continue
            interest_due_date = (
                _next_workday(due_date)
                if include_weekend_interest
                else due_date
            )
            calculated = _calculate_overdue_center_interest(
                source_row,
                columns,
                interest_due_date,
            )
            if calculated is None:
                warned_overdue_fallback = True
                interest_amount = interest_due
            else:
                interest_amount = calculated
            if interest_amount <= 0:
                continue
            interest_to_date = interest_due_date
        elif collection_mode == NOT_DUE_AND_OVERDUE_INTEREST:
            if due_date is None:
                warned_missing_due_date = True
                continue
            if not (due_date > month_end or due_date <= today):
                continue
            interest_amount = interest_due
            interest_to_date = collection_date
        else:
            interest_amount = interest_due
            interest_to_date = collection_date
        rows.append(
            InterestRow(
                source_row=source_row,
                customer_code=customer_code,
                customer_name=_cell_display_text(source_row, columns, "customer_name"),
                loan_number=_cell_text(source_row, columns, "loan_number"),
                principal_balance=_cell_number(source_row, columns, "principal_balance"),
                interest_due=interest_due,
                next_interest_payment_date=due_date,
                collection_interest_amount=round(interest_amount, 2),
                interest_to_date=interest_to_date,
            )
        )
    if warned_missing_due_date:
        warnings.append(
            "Một số khoản vay bị bỏ qua vì thiếu/không đọc được cột ngày đến hạn nxtintrpmt."
        )
    if warned_overdue_fallback:
        warnings.append(
            "Một số khoản Thu lãi quá hạn Center chưa đủ cột DuNo/LaiSuat/ngày tính lãi; "
            "đã dùng LaiDK gốc thay công thức VBA."
        )
    return rows


def _calculate_overdue_center_interest(
    source_row: tuple[object, ...],
    columns: ResolvedColumns,
    due_date: date,
) -> float | None:
    principal = _cell_number(source_row, columns, "principal_balance")
    rate = _cell_number(source_row, columns, "interest_rate")
    start_1 = _cell_date(source_row, columns, "interest_start_date_1")
    start_2 = _cell_date(source_row, columns, "interest_start_date_2")
    start_candidates = [value for value in (start_1, start_2) if value is not None]
    if principal <= 0 or rate <= 0 or not start_candidates:
        return None
    start_date = max(start_candidates)
    # Chỉ thu đến trước ngày đến hạn thu lãi.
    day_count = (due_date - start_date).days - 1
    if day_count <= 0:
        return 0.0
    return principal * (rate / 100) * day_count / 365


def _next_workday(value: date) -> date:
    if value.weekday() == 5:
        return value + timedelta(days=2)
    if value.weekday() == 6:
        return value + timedelta(days=1)
    return value


def _ledger_interest_to_date(value: date, collection_mode: str) -> date:
    if collection_mode == OVERDUE_CENTER_INTEREST:
        return value - timedelta(days=1)
    return value


def _pledged_collateral_customer_codes(
    data: TabularData,
    columns: ResolvedColumns,
) -> set[str]:
    customer_codes: set[str] = set()
    for row in data.rows:
        account_key = _cell_text(row, columns, "account_key")
        if not account_key.startswith("994003"):
            continue
        customer_code = _cell_text(row, columns, "customer_code")
        if customer_code:
            customer_codes.add(customer_code)
    return customer_codes


def _filter_pledged_collateral(
    rows: list[InterestRow],
    pledged_customer_codes: set[str],
) -> tuple[list[InterestRow], list[InterestRow]]:
    if not pledged_customer_codes:
        return rows, []
    selected: list[InterestRow] = []
    skipped: list[InterestRow] = []
    for row in rows:
        if row.customer_code in pledged_customer_codes:
            skipped.extend(
                _mark_skipped(
                    [row],
                    "Khách hàng có tài sản bảo đảm cầm cố 994003.",
                )
            )
        else:
            selected.append(row)
    return selected, skipped


def _filter_by_deposit_balance(
    rows: list[InterestRow],
    deposits: Mapping[str, DepositAccount],
    collection_mode: str,
) -> tuple[list[InterestRow], list[InterestRow]]:
    grouped: dict[str, list[InterestRow]] = defaultdict(list)
    for row in rows:
        grouped[row.customer_code].append(row)
    selected: list[InterestRow] = []
    skipped: list[InterestRow] = []
    for customer_code, customer_rows in grouped.items():
        deposit = _match_deposit(customer_code, deposits)
        total_interest = sum(row.collection_interest_amount for row in customer_rows)
        if deposit is None:
            skipped.extend(
                _mark_skipped(customer_rows, "Không tìm thấy tài khoản tiền gửi theo mã KH.")
            )
            continue
        remaining_balance = deposit.balance - total_interest
        if collection_mode == COLLECT_ALL_INTEREST:
            insufficient = remaining_balance <= 0
            reason = (
                f"Số dư tiền gửi {deposit.balance:,.0f} không lớn hơn tổng lãi "
                f"{total_interest:,.0f}."
            )
        else:
            insufficient = remaining_balance < 0
            reason = (
                f"Số dư tiền gửi {deposit.balance:,.0f} nhỏ hơn tổng lãi "
                f"{total_interest:,.0f}."
            )
        if insufficient:
            skipped.extend(
                _mark_skipped(
                    customer_rows,
                    reason,
                )
            )
            continue
        selected.extend(
            InterestRow(
                source_row=row.source_row,
                customer_code=row.customer_code,
                customer_name=row.customer_name,
                loan_number=row.loan_number,
                principal_balance=row.principal_balance,
                interest_due=row.interest_due,
                next_interest_payment_date=row.next_interest_payment_date,
                collection_interest_amount=row.collection_interest_amount,
                interest_to_date=row.interest_to_date,
                account_number=deposit.account_number,
            )
            for row in customer_rows
        )
    return selected, skipped


def _mark_skipped(rows: list[InterestRow], reason: str) -> list[InterestRow]:
    return [
        InterestRow(
            source_row=row.source_row,
            customer_code=row.customer_code,
            customer_name=row.customer_name,
            loan_number=row.loan_number,
            principal_balance=row.principal_balance,
            interest_due=row.interest_due,
            next_interest_payment_date=row.next_interest_payment_date,
            collection_interest_amount=row.collection_interest_amount,
            interest_to_date=row.interest_to_date,
            account_number=row.account_number,
            skipped_reason=reason,
        )
        for row in rows
    ]


def _build_deposit_account_lookup(
    data: TabularData,
    columns: ResolvedColumns,
    statement_type: str,
) -> dict[str, DepositAccount]:
    result: dict[str, DepositAccount] = {}
    statement = _normalize_deposit_statement_type(statement_type)
    for row in data.rows:
        customer_code = _cell_text(row, columns, "customer_code")
        if statement == "dpda":
            customer_code = _strip_branch_code(customer_code)
        account_number = _cell_text(row, columns, "account_number")
        balance = _cell_number(row, columns, "balance")
        maturity_date = _cell_raw(row, columns, "maturity_date")
        if statement == "msit" and not _is_demand_deposit_maturity(maturity_date):
            continue
        if not customer_code or not account_number:
            continue
        current = result.get(customer_code)
        if current is None or balance > current.balance:
            result[customer_code] = DepositAccount(
                customer_code=customer_code,
                account_number=account_number,
                balance=balance,
            )
    return result


def _append_report_rows_from_interest_workbook(
    workbook,
    sheet,
    report_date: date,
    collection_mode: str | None,
) -> tuple[int, float]:
    source = workbook["SaoKeTrichLai"]
    header_row = _find_header_row(source, "SoGiaiNgan")
    if header_row is not None:
        return _append_report_rows_from_legacy_table(source, sheet, header_row, report_date)
    if "FileGoc" not in workbook.sheetnames:
        raise AutoInterestError(
            "File nguồn chưa có sheet FileGoc để tra cứu lãi dự kiến và tên khách hàng."
        )
    return _append_report_rows_from_vba_technical_sheet(
        source,
        workbook["FileGoc"],
        workbook["BangKeTheoLo"] if "BangKeTheoLo" in workbook.sheetnames else None,
        sheet,
        report_date,
        collection_mode,
    )


def _resolve_report_collection_mode(
    source_file: Path,
    workbook,
    requested_mode: str | None,
) -> str | None:
    if requested_mode:
        return _normalize_collection_mode(requested_mode)

    stem = source_file.stem.casefold()
    if stem.startswith("thutoanbolai"):
        return COLLECT_ALL_INTEREST
    if stem.startswith("thulaikhongdenhan"):
        return NOT_DUE_INTEREST
    if stem.startswith("thulaiquahancenter"):
        return OVERDUE_CENTER_INTEREST
    if stem.startswith("thulaichuadenhanquahan"):
        return NOT_DUE_AND_OVERDUE_INTEREST

    if "BangKeTheoLo" in workbook.sheetnames:
        return None
    return COLLECT_ALL_INTEREST


def _append_report_rows_from_legacy_table(
    source,
    sheet,
    header_row: int,
    report_date: date,
) -> tuple[int, float]:
    headers = [_text(cell.value) for cell in source[header_row]]
    index = _header_index(headers)
    rows_written = 0
    total_interest = 0.0
    for row in source.iter_rows(min_row=header_row + 1, values_only=True):
        loan_number = _row_value(row, index, "SoGiaiNgan")
        if not _text(loan_number):
            continue
        amount = _number(_row_value(row, index, "LaiDK"))
        sheet.append(
            (
                _text(loan_number),
                "02",
                amount,
                _row_value(row, index, "ThuLaiDen") or _format_ipcas_date(report_date),
                _text(_row_value(row, index, "ThuTuTK")),
                _text(_row_value(row, index, "TenKH")),
                _format_vietnamese_date(report_date),
            )
        )
        rows_written += 1
        total_interest += amount
    return rows_written, total_interest


def _append_report_rows_from_vba_technical_sheet(
    source,
    file_sheet,
    ledger_sheet,
    sheet,
    report_date: date,
    collection_mode: str | None,
) -> tuple[int, float]:
    file_headers = [_text(cell.value) for cell in file_sheet[1]]
    loan_columns = validate_loan_columns(
        TabularData(
            headers=file_headers,
            rows=[
                tuple(cell.value for cell in row)
                for row in file_sheet.iter_rows(min_row=2, values_only=False)
            ],
        )
    )
    file_lookup: dict[str, tuple[object, ...]] = {}
    for row in file_sheet.iter_rows(min_row=2, values_only=True):
        loan_number = _cell_text(row, loan_columns, "loan_number")
        if loan_number:
            file_lookup[loan_number] = tuple(row)
    ledger_lookup = _build_ledger_lookup(ledger_sheet) if ledger_sheet is not None else {}
    use_ledger_amount = collection_mode != COLLECT_ALL_INTEREST
    if use_ledger_amount and not ledger_lookup:
        raise AutoInterestError(
            "File nguồn chưa có sheet BangKeTheoLo để lấy số lãi đã tính lại cho hình thức thu lãi này."
        )

    rows_written = 0
    total_interest = 0.0
    for row in source.iter_rows(values_only=True):
        loan_number = _text(row[0] if len(row) > 0 else "")
        if not loan_number:
            continue
        file_row = file_lookup.get(loan_number, ())
        ledger_row = ledger_lookup.get(loan_number, {})
        amount = (
            _number(ledger_row.get("interest_due"))
            if use_ledger_amount
            else _cell_number(file_row, loan_columns, "interest_due")
            if file_row
            else 0.0
        )
        customer_name = _cell_display_text(file_row, loan_columns, "customer_name") if file_row else ""
        if not customer_name:
            customer_name = _text(ledger_row.get("customer_name"))
        interest_to = (
            _text(ledger_row.get("interest_to")) if use_ledger_amount else ""
        ) or _text(row[3] if len(row) > 3 else "") or _format_ipcas_date(report_date)
        account_number = (
            _text(ledger_row.get("account_number")) if use_ledger_amount else ""
        ) or _text(row[4] if len(row) > 4 else "")
        sheet.append(
            (
                loan_number,
                "02",
                amount,
                interest_to,
                account_number,
                customer_name,
                _format_vietnamese_date(report_date),
            )
        )
        rows_written += 1
        total_interest += amount
    return rows_written, total_interest


def _build_ledger_lookup(sheet) -> dict[str, dict[str, object]]:
    header_row = None
    for loan_header in ("Số giải ngân", "SoGiaiNgan", "MaGN", "Mã GN", "Disburse No"):
        header_row = _find_header_row(sheet, loan_header)
        if header_row is not None:
            break
    if header_row is None:
        return {}
    headers = [_text(cell.value) for cell in sheet[header_row]]
    index = _header_index(headers)
    lookup: dict[str, dict[str, object]] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        loan_number = _text(
            _row_first_value(
                row,
                index,
                "Số giải ngân",
                "SoGiaiNgan",
                "MaGN",
                "Mã GN",
                "Disburse No",
            )
        )
        if not loan_number:
            continue
        lookup[loan_number] = {
            "customer_name": _row_first_value(
                row,
                index,
                "Tên KH",
                "TenKH",
                "Customer_Name",
                "Customer Name",
            ),
            "interest_due": _row_first_value(
                row,
                index,
                "Lãi DK",
                "LaiDK",
                "Lai DK",
                "Interest Due",
            ),
            "interest_to": _row_first_value(
                row,
                index,
                "Thu lãi đến",
                "ThuLaiDen",
                "NgayThuLai",
                "Interest to",
            ),
            "account_number": _row_first_value(
                row,
                index,
                "Thu từ TK",
                "ThuTuTK",
                "SoTK",
                "Số TK",
            ),
        }
    return lookup


def _write_auto_interest_technical_sheet(
    sheet,
    rows: list[InterestRow],
) -> None:
    sheet.append(("", "", "", "", ""))
    sheet.append(("", "", "", "", ""))
    for row in rows:
        sheet.append(
            (
                row.loan_number,
                "01",
                0,
                _format_ipcas_date(row.interest_to_date),
                row.account_number,
            )
        )
    _style_technical_sheet(sheet)


def _write_auto_interest_ledger_sheet(
    sheet,
    rows: list[InterestRow],
    collection_mode: str,
    collection_date: date,
) -> None:
    sheet.append(("DANH SÁCH TRÍCH LÃI BÁN TỰ ĐỘNG",))
    sheet.append(("", "", "", "", "", "", "", ""))
    sheet.append(LEDGER_HEADERS)
    for index, row in enumerate(rows, start=1):
        sheet.append(
            (
                index,
                row.customer_code,
                row.customer_name,
                row.loan_number,
                row.customer_name,
                row.collection_interest_amount,
                _format_vietnamese_date(
                    _ledger_interest_to_date(row.interest_to_date, collection_mode)
                ),
                row.account_number,
            )
        )
    total_row = sheet.max_row + 1
    sheet.cell(total_row, 6).value = sum(row.collection_interest_amount for row in rows)
    sheet.append(("", "", "", "", "", "", "", ""))
    sheet.append(("", "", "", "", "", "", "", ""))
    sheet.append(("", "LẬP BẢNG", "", "", "", "KIỂM SOÁT", "", ""))
    _style_ledger_sheet(sheet, total_row)


def _write_note_sheet(
    workbook: Workbook,
    collection_mode: str,
    loan_columns: ResolvedColumns,
    deposit_columns: ResolvedColumns,
    warnings: list[str],
    skipped_rows: list[InterestRow],
    request: AutoInterestInput,
) -> None:
    sheet = workbook.create_sheet("GhiChu")
    sheet.append(("Muc", "Noi dung"))
    sheet.append(("VBA procedure", COLLECTION_MODE_VBA_MAP[collection_mode]))
    sheet.append(("Che do", COLLECTION_MODE_LABELS[collection_mode]))
    sheet.append(("File loan", str(request.loan_file)))
    sheet.append(("File tien gui", str(request.deposit_file)))
    sheet.append(("Loai sao ke tien gui", _normalize_deposit_statement_type(request.deposit_statement_type)))
    sheet.append(("Cot loan da nhan dien", _format_recognized_columns(loan_columns.recognized)))
    sheet.append(("Cot tien gui da nhan dien", _format_recognized_columns(deposit_columns.recognized)))
    for warning in warnings:
        sheet.append(("Canh bao", warning))
    if skipped_rows:
        sheet.append(())
        sheet.append(("MaKH", "TenKH", "SoGiaiNgan", "LaiDK", "Ly do bo qua"))
        for row in skipped_rows:
            sheet.append(
                (
                    row.customer_code,
                    row.customer_name,
                    row.loan_number,
                    row.collection_interest_amount,
                    row.skipped_reason,
                )
            )
    _auto_width(sheet)


def _write_source_sheet(
    sheet,
    headers: list[str],
    rows: list[tuple[object, ...]],
) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _auto_width(sheet)


def _style_technical_sheet(sheet) -> None:
    widths = (20, 8, 14, 14, 18)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=3, max_col=5):
        row[1].number_format = "@"
        row[2].number_format = "#,##0"
        row[3].number_format = "@"
        row[4].number_format = "@"


def _style_ledger_sheet(sheet, total_row: int) -> None:
    sheet.merge_cells("A1:H1")
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")
    for cell in sheet[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )
    for row in sheet.iter_rows(min_row=3, max_row=total_row, max_col=8):
        for cell in row:
            cell.border = border
    for row in sheet.iter_rows(min_row=4, max_row=total_row):
        row[5].number_format = "#,##0"
        row[6].number_format = "@"
    sheet.cell(total_row, 6).font = Font(bold=True)
    sheet.cell(sheet.max_row, 2).font = Font(bold=True)
    sheet.cell(sheet.max_row, 6).font = Font(bold=True)
    sheet.cell(sheet.max_row, 2).alignment = Alignment(horizontal="center")
    sheet.cell(sheet.max_row, 6).alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A4"
    widths = (6, 15, 24, 18, 24, 16, 14, 20)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _style_report_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2):
        row[1].number_format = "@"
        row[2].number_format = "#,##0"
        row[3].number_format = "dd/mm/yyyy"
        row[6].number_format = "dd/mm/yyyy"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(sheet.max_row, 1)}"
    _auto_width(sheet)


def _read_tabular_file(path: Path) -> tuple[list[str], list[tuple[object, ...]]]:
    if not path.exists():
        raise AutoInterestError(f"Không tìm thấy file: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            sheet = workbook.worksheets[0]
            rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    elif suffix == ".xls":
        with open(os.devnull, "w") as devnull:
            workbook = open_workbook(str(path), logfile=devnull)
        sheet = workbook.sheet_by_index(0)
        rows = [
            tuple(_xlrd_cell_value(sheet, row_index, column_index, workbook.datemode) for column_index in range(sheet.ncols))
            for row_index in range(sheet.nrows)
        ]
    elif suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            rows = [tuple(row) for row in csv.reader(file)]
    else:
        raise AutoInterestError("Chỉ hỗ trợ file .xlsx, .xlsm, .xls hoặc .csv.")
    header_position = next(
        (index for index, row in enumerate(rows) if any(_text(value) for value in row)),
        None,
    )
    if header_position is None:
        raise AutoInterestError(f"File không có dữ liệu: {path}")
    headers = [_text(value) for value in rows[header_position]]
    data_rows = rows[header_position + 1 :]
    return headers, data_rows


def _xlrd_cell_value(sheet, row_index: int, column_index: int, datemode: int) -> object:
    cell = sheet.cell(row_index, column_index)
    if cell.ctype == 3:
        return xldate_as_datetime(cell.value, datemode).date()
    return cell.value


def _resolve_columns(
    headers: Iterable[str],
    aliases: Mapping[str, tuple[str, ...]],
    fallbacks: Mapping[str, int] | None = None,
) -> ResolvedColumns:
    normalized_headers = {_normalize_header(header): index for index, header in enumerate(headers)}
    header_list = list(headers)
    positions: dict[str, int] = {}
    recognized: dict[str, str] = {}
    for field_name, names in aliases.items():
        for name in names:
            normalized = _normalize_header(name)
            if normalized in normalized_headers:
                position = normalized_headers[normalized]
                positions[field_name] = position
                recognized[field_name] = header_list[position]
                break
        if field_name not in positions and fallbacks and field_name in fallbacks:
            position = fallbacks[field_name]
            if position < len(header_list):
                positions[field_name] = position
                recognized[field_name] = f"{header_list[position]} (fallback {get_column_letter(position + 1)})"
    return ResolvedColumns(positions=positions, recognized=recognized)


def _missing_headers(headers: Iterable[str], required: Iterable[str]) -> list[str]:
    available = {_normalize_header(header) for header in headers}
    return [header for header in required if _normalize_header(header) not in available]


def _header_index(headers: Iterable[str]) -> dict[str, int]:
    return {_normalize_header(header): index for index, header in enumerate(headers)}


def _row_value(row: tuple[object, ...], index: Mapping[str, int], header: str) -> object:
    position = index.get(_normalize_header(header))
    if position is None or position >= len(row):
        return ""
    return row[position]


def _row_first_value(
    row: tuple[object, ...],
    index: Mapping[str, int],
    *headers: str,
) -> object:
    for header in headers:
        value = _row_value(row, index, header)
        if _text(value):
            return value
    return ""


def _cell_raw(
    row: tuple[object, ...],
    columns: ResolvedColumns,
    field_name: str,
) -> object:
    position = columns.positions.get(field_name)
    if position is None or position >= len(row):
        return ""
    return row[position]


def _cell_text(
    row: tuple[object, ...],
    columns: ResolvedColumns,
    field_name: str,
) -> str:
    return _text(_cell_raw(row, columns, field_name))


def _cell_display_text(
    row: tuple[object, ...],
    columns: ResolvedColumns,
    field_name: str,
) -> str:
    return _display_text(_cell_raw(row, columns, field_name))


def _cell_number(
    row: tuple[object, ...],
    columns: ResolvedColumns,
    field_name: str,
) -> float:
    return _number(_cell_raw(row, columns, field_name))


def _cell_date(
    row: tuple[object, ...],
    columns: ResolvedColumns,
    field_name: str,
) -> date | None:
    return _to_date(_cell_raw(row, columns, field_name))


def _match_deposit(
    customer_code: str,
    deposits: Mapping[str, DepositAccount],
) -> DepositAccount | None:
    if customer_code in deposits:
        return deposits[customer_code]
    suffix = customer_code[-9:] if len(customer_code) >= 9 else customer_code
    return deposits.get(suffix)


def _strip_branch_code(value: str) -> str:
    return value[4:] if len(value) > 4 else value


def _is_demand_deposit_maturity(value: object) -> bool:
    text = _text(value)
    if not text:
        return True
    normalized = text.replace("-", "/").strip()
    return normalized in {"0", "00/00/0000", "0/0/0000", "00/00/00"}


def _to_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        try:
            return datetime(1899, 12, 30).date() + timedelta(days=int(value))
        except OverflowError:
            return None
    text = str(value).strip()
    if not text or text in {"00/00/0000", "0"}:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y%m%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _end_of_month(value: date) -> date:
    if value.month == 12:
        return date(value.year, 12, 31)
    return date(value.year, value.month + 1, 1) - timedelta(days=1)


def _format_ipcas_date(value: date) -> str:
    return f"{value:%Y%m%d}"


def _format_vietnamese_date(value: date) -> str:
    return f"{value:%d/%m/%Y}"


def _find_header_row(sheet, required_header: str) -> int | None:
    normalized = _normalize_header(required_header)
    for row in range(1, min(sheet.max_row, 20) + 1):
        values = [_normalize_header(cell.value) for cell in sheet[row]]
        if normalized in values:
            return row
    return None


def _auto_width(sheet) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max(len(_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(width + 2, 10), 48)


def _unique_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    for index in range(1, 1000):
        candidate = parent / f"{stem}_{index:02d}{suffix}"
        if not candidate.exists():
            return candidate
    raise AutoInterestError("Không thể tạo tên file kết quả không trùng.")


def _normalize_collection_mode(value: str) -> str:
    normalized = (value or "").strip()
    if normalized == LEGACY_DUE_IN_MONTH:
        return OVERDUE_CENTER_INTEREST
    if normalized in COLLECTION_MODE_LABELS:
        return normalized
    raise AutoInterestError("Hình thức thu lãi không hợp lệ.")


def _normalize_deposit_statement_type(value: str) -> str:
    normalized = (value or "").strip().casefold()
    if normalized in {"msit", "dpda"}:
        return normalized
    raise AutoInterestError("Loại sao kê tiền gửi không hợp lệ.")


def _normalize_header(value: object) -> str:
    return _text(value).casefold().replace(" ", "").replace("_", "").replace("-", "")


def _format_recognized_columns(columns: Mapping[str, str]) -> str:
    return "; ".join(f"{field_name}={header}" for field_name, header in sorted(columns.items()))


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _display_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0
