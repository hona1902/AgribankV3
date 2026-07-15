from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from agribank_v3.features.credit.tovayvon.models import (
    COMMISSION_EXPORT_HEADERS,
    COMMISSION_RULE_EXPORT_HEADERS,
    DATA_TVV_HEADERS,
    DATA_TVV_TEMPLATE_HEADERS,
)


DATA_TVV_TEXT_COLUMNS: frozenset[str] = frozenset(
    {
        "MaTo",
        "MaToTruong",
        "TK_ToTruong",
        "SoDienThoai",
        "TK_ToHoiXa",
        "TK_HUYEN",
        "TK_TINH",
        "TK_TW",
        "uyquyen",
        "TTLN_TW",
        "TTLN_Tinh",
    }
)


DATA_TVV_SAMPLE_ROW: tuple[str | int, ...] = (
    1,
    "TVV001",
    "Tổ vay vốn mẫu",
    "Tổ vay vốn mẫu - Xã mẫu",
    "Xã mẫu",
    "KH001",
    "Nguyễn Văn A",
    "Thôn mẫu, Xã mẫu",
    "5491000000000",
    "0912345678",
    "Hội Nông dân",
    "5491000000001",
    "Hội Nông dân xã",
    "Huyện mẫu",
    "5491000000002",
    "Tỉnh mẫu",
    "5491000000003",
    "Trung ương Hội",
    "5491000000004",
    "",
    "",
    "",
)


COMMISSION_SAMPLE_ROW: tuple[int | float, ...] = (
    80,
    13,
    3.8,
    2.5,
    0.7,
    100,
    90,
    10,
    0,
    0,
    0,
    100,
)


COMMISSION_RULE_SAMPLE_ROW: tuple[int | float, ...] = (
    85,
    90,
    50,
    90,
    95,
    90,
    95,
    100,
    2,
    0,
)


def create_data_tvv_template(output_path: Path) -> Path:
    """Create a Data_TVV import template with commission columns."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data_TVV"

    sheet.append(DATA_TVV_TEMPLATE_HEADERS)
    sheet.append(DATA_TVV_SAMPLE_ROW + COMMISSION_SAMPLE_ROW + COMMISSION_RULE_SAMPLE_ROW)

    header_font = Font(bold=True)
    for column_index, header in enumerate(DATA_TVV_TEMPLATE_HEADERS, start=1):
        column_letter = get_column_letter(column_index)
        header_cell = sheet.cell(row=1, column=column_index)
        header_cell.font = header_font
        header_cell.fill = _header_fill(header)
        sheet.column_dimensions[column_letter].width = _column_width(header)
        if header in DATA_TVV_TEXT_COLUMNS:
            sheet.column_dimensions[column_letter].number_format = "@"
            for row_index in range(2, 202):
                sheet.cell(row=row_index, column=column_index).number_format = "@"
        else:
            for row_index in range(2, 202):
                sheet.cell(row=row_index, column=column_index).number_format = "0.##"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(DATA_TVV_TEMPLATE_HEADERS))}1"

    guide_sheet = workbook.create_sheet("HuongDan")
    guide_sheet.append(("Hướng dẫn nhập dữ liệu Data_TVV",))
    guide_sheet.append(("Không đổi tên sheet Data_TVV.",))
    guide_sheet.append(("Không đổi tên hoặc sắp xếp lại 22 cột Data_TVV gốc.",))
    guide_sheet.append(("22 cột đầu là dữ liệu tổ vay vốn gốc.",))
    guide_sheet.append(("Các cột HH_* dùng để nhập tỷ lệ hoa hồng riêng theo từng tổ.",))
    guide_sheet.append(("Nếu để trống cột HH_* thì hệ thống dùng tỷ lệ mặc định.",))
    guide_sheet.append(("Tổng hoa hồng không BĐ và có BĐTS phải bằng 100.",))
    guide_sheet.append(("Các cột DK_* là điều kiện chi hoa hồng chung.",))
    guide_sheet.append(("Nếu để trống DK_* thì hệ thống dùng cấu hình điều kiện mặc định.",))
    guide_sheet.append(("MaTo là bắt buộc và không được trùng.",))
    guide_sheet.append(
        ("Các mã, tài khoản và số điện thoại nên nhập dạng Text để giữ số 0 đầu.",)
    )
    guide_sheet.append(("Không xóa dòng tiêu đề.",))
    guide_sheet.append(("Sau khi nhập xong, dùng nút Import Excel trong app.",))
    guide_sheet["A1"].font = Font(bold=True)
    guide_sheet.column_dimensions["A"].width = 88

    workbook.save(output_path)
    workbook.close()
    return output_path


def _column_width(header: str) -> int:
    widths = {
        "STT": 8,
        "MaTo": 14,
        "TenTo": 24,
        "TenTVV_DayDu": 34,
        "Xa": 18,
        "MaToTruong": 16,
        "Ten_ToTruong": 24,
        "DiaChi": 30,
        "TK_ToTruong": 20,
        "SoDienThoai": 16,
        "ToHoi": 20,
        "TK_ToHoiXa": 20,
        "ToChuc": 24,
        "Ten_Huyen": 18,
        "TK_HUYEN": 18,
        "Ten_Tinh": 18,
        "TK_TINH": 18,
        "Ten_TW": 20,
        "TK_TW": 18,
        "uyquyen": 14,
        "TTLN_TW": 14,
        "TTLN_Tinh": 14,
    }
    if header in COMMISSION_EXPORT_HEADERS or header in COMMISSION_RULE_EXPORT_HEADERS:
        return max(16, len(header) + 2)
    return widths.get(header, max(12, len(header) + 2))


def _header_fill(header: str) -> PatternFill:
    if header.startswith("HH_KhongBD"):
        return PatternFill("solid", fgColor="FFF2CC")
    if header.startswith("HH_CoBDTS"):
        return PatternFill("solid", fgColor="D9EAD3")
    if header.startswith("DK_"):
        return PatternFill("solid", fgColor="FCE4D6")
    return PatternFill("solid", fgColor="EAF2F8")
