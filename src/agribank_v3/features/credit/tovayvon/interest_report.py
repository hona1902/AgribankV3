from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable
import io
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import xlrd

from agribank_v3.features.credit.tovayvon.models import (
    CreditCommissionRuleSettings,
    CreditGroup,
    CreditGroupCommissionRate,
)
from agribank_v3.features.credit.tovayvon.repository import CreditGroupRepository


class InterestReportError(RuntimeError):
    pass


INTEREST_REPORT_TITLE = "Bảng kê thu lãi tổ vay vốn"
DEFAULT_BRANCH_REPORT_NAME = "CHI NHÁNH LỘC PHÁT LÂM ĐỒNG"

REQUIRED_INTEREST_COLUMNS: tuple[str, ...] = (
    "MaKH",
    "TenKH",
    "SoGiaiNgan",
    "DuNo",
    "SoLaiDaThuTrongKy",
    "TyLeBaoBam",
    "MaToVayVon",
)

REQUIRED_DEBT_COLUMNS: tuple[str, ...] = REQUIRED_INTEREST_COLUMNS + ("NhomNo",)

OPTIONAL_INTEREST_COLUMNS: tuple[str, ...] = (
    "LaiTon",
    "ThuLaiTuNgay",
    "ThuLaiDenNgay",
    "IsYellowFlag",
    "SoLaiQuaHanDaThuTrongKy",
    "LaiQuaHanDaThuTrongKy",
    "LaiQuaHan",
)

SKTL_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "customer_code": ("MaKH", "Mã KH", "Ma Khach Hang", "Mã khách hàng"),
    "customer_name": ("TenKH", "Tên KH", "Ten Khach Hang", "Tên khách hàng"),
    "loan_number": ("SoGiaiNgan", "Số giải ngân", "So Giai Ngan"),
    "group_code": ("MaToVayVon", "MaTo", "Mã tổ vay vốn", "GRPNO"),
    "interest_collected": ("SoLaiDaThuTrongKy", "Số lãi đã thu trong kỳ", "LaiDaThuTrongKy"),
    "overdue_interest_collected": ("SoLaiQuaHanTrongKy", "Lãi quá hạn trong kỳ", "LaiQuaHanTrongKy"),
    "secured_ratio": ("TyLeBaoBam", "TyLeBaoDam", "Tỷ lệ bảo đảm", "TyLeKhongBaoDam"),
    "outstanding_balance": ("DuNo", "Dư nợ", "Du No"),
    "remaining_interest": ("LaiTon", "Lãi tồn", "LaiTonCuoiKy", "Lãi tồn cuối kỳ"),
    "interest_rate": ("LaiSuat", "Lãi suất", "Lai Suat"),
    "disbursement_date": ("NgayGiaiNgan", "Ngày giải ngân", "Ngay Giai Ngan"),
    "last_interest_date": ("NgayTraLaiCuoiCung", "Ngày trả lãi cuối cùng"),
    "last_principal_date": ("NgayTraGocCuoiCung", "Ngày trả gốc cuối cùng"),
    "principal_collected_period": ("SoGocDaThuTrongKy", "Số gốc đã thu trong kỳ"),
    "period_from_date": ("ThuLaiTuNgay", "Thu lãi từ ngày"),
    "period_to_date": ("ThuLaiDenNgay", "Thu lãi đến ngày"),
}

SKCK_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "customer_code": ("MaKH", "Mã KH", "Ma Khach Hang", "Mã khách hàng"),
    "customer_name": ("TenKH", "Tên KH", "Ten Khach Hang", "Tên khách hàng"),
    "loan_number": ("SoGiaiNgan", "Số giải ngân", "So Giai Ngan"),
    "group_code": ("MaToVayVon", "MaTo", "Mã tổ vay vốn", "GRPNO"),
    "outstanding_balance": ("DuNo", "Dư nợ", "Du No"),
    "debt_group": ("NhomNo", "Nhóm nợ", "Nhom No"),
    "remaining_interest": ("LaiTon", "Lãi tồn", "LaiTonCuoiKy", "Lãi tồn cuối kỳ"),
    "interest_receivable": ("LaiPhaiThu", "Lãi phải thu", "SoLaiPhaiThu"),
    "overdue_interest": ("SoLaiQuaHan", "Lãi quá hạn", "SoLaiQuaHanTrongKy"),
    "secured_ratio": ("TyLeBaoBam", "TyLeBaoDam", "Tỷ lệ bảo đảm", "TyLeKhongBaoDam"),
    "interest_rate": ("LaiSuat", "Lãi suất", "Lai Suat"),
    "disbursement_date": ("NgayGiaiNgan", "Ngày giải ngân", "Ngay Giai Ngan"),
    "last_interest_date": ("NgayTraLaiCuoiCung", "Ngày trả lãi cuối cùng"),
    "last_principal_date": ("NgayTraGocCuoiCung", "Ngày trả gốc cuối cùng"),
}

SKTL_REQUIRED_FIELDS: tuple[str, ...] = (
    "customer_code",
    "customer_name",
    "loan_number",
    "group_code",
    "interest_collected",
    "secured_ratio",
)

SKCK_REQUIRED_FIELDS: tuple[str, ...] = (
    "customer_code",
    "customer_name",
    "loan_number",
    "group_code",
    "outstanding_balance",
    "debt_group",
    "secured_ratio",
)

SKCK_LAITON_FORMULA_FIELDS: tuple[str, ...] = (
    "interest_rate",
    "disbursement_date",
    "last_interest_date",
)

DETAIL_HEADERS: tuple[str, ...] = (
    "STT",
    "Mã Tổ Vay Vốn",
    "Mã khách hàng",
    "Tên khách hàng",
    "Số giải ngân",
    "Dư nợ",
    "Thu lãi từ ngày",
    "Thu lãi đến ngày",
    "Lãi phải thu",
    "Lãi có bảo đảm",
    "Lãi không bảo đảm",
    "Lãi tồn",
    "Hoa hồng thu lãi",
    "Ghi chú",
)

SUMMARY_HEADERS: tuple[str, ...] = (
    "MaTo",
    "TenTo",
    "TenToTruong",
    "SoDong",
    "TongDuNo",
    "LaiCoBD",
    "LaiKhongBD",
    "LaiTon",
    "LaiPhaiThu",
    "TyLeThuLai",
    "TyLeNoXau",
    "TyLeChi",
    "HHCoBD",
    "HHKhongBD",
    "TongHH",
    "HH_ToTruong",
    "HH_CapXa",
    "HH_CapHuyen",
    "HH_CapTinh",
    "HH_TW",
    "CanhBao",
)

BANGKE_SHEET_NAME = "BangKe"
SUMMARY_SHEET_NAME = "TongHopTheoTo"
WARNING_SHEET_NAME = "CanhBao"
VBA_SECURED_COMMISSION_BASE_RATE = 0.02
VBA_NO_SECURED_COMMISSION_BASE_RATE = 0.03


@dataclass(frozen=True, slots=True)
class InterestReportRequest:
    interest_file: Path
    debt_file: Path
    output_path: Path
    from_date: date
    to_date: date
    ma_to: str = ""
    selected_group_codes: tuple[str, ...] = ()
    include_overdue_interest: bool = False


@dataclass(frozen=True, slots=True)
class ColumnDetectionResult:
    file_kind: str
    sheet_name: str
    headers: tuple[str, ...]
    field_to_header: dict[str, str]
    missing_required: tuple[str, ...]
    missing_optional: tuple[str, ...]

    @property
    def can_create_report(self) -> bool:
        return not self.missing_required


@dataclass(slots=True)
class InterestRow:
    ma_to: str
    ma_kh: str
    ten_kh: str
    so_giai_ngan: str
    du_no: float
    ty_le_bao_bam: float
    so_lai_da_thu: float
    lai_ton: float = 0.0
    thu_lai_tu_ngay: Any = ""
    thu_lai_den_ngay: Any = ""
    is_yellow_flag: bool = False
    source: str = "thu_lai"
    debt_group: str = ""
    loan_status: str = ""
    note: str = ""

    @property
    def lai_co_bd(self) -> float:
        return self.so_lai_da_thu if self.ty_le_bao_bam > 0 else 0.0

    @property
    def lai_khong_bd(self) -> float:
        return self.so_lai_da_thu if self.ty_le_bao_bam <= 0 else 0.0

    @property
    def lai_phai_thu(self) -> float:
        return self.lai_co_bd + self.lai_khong_bd + self.lai_ton


@dataclass(slots=True)
class InterestGroupSummary:
    group: CreditGroup
    rate: CreditGroupCommissionRate
    rows: list[InterestRow] = field(default_factory=list)
    total_du_no: float = 0.0
    bad_du_no: float = 0.0
    warning: str = ""
    rule_settings: CreditCommissionRuleSettings | None = None
    uses_custom_rule: bool = False

    @property
    def total_lai_co_bd(self) -> float:
        return sum(row.lai_co_bd for row in self.rows)

    @property
    def total_lai_khong_bd(self) -> float:
        return sum(row.lai_khong_bd for row in self.rows)

    @property
    def total_lai_ton(self) -> float:
        return sum(row.lai_ton for row in self.rows)

    @property
    def total_lai_phai_thu(self) -> float:
        return self.total_lai_co_bd + self.total_lai_khong_bd + self.total_lai_ton

    @property
    def collection_rate(self) -> float:
        denominator = self.total_lai_co_bd + self.total_lai_khong_bd + self.total_lai_ton
        return (self.total_lai_co_bd + self.total_lai_khong_bd) / denominator if denominator else 0.0

    @property
    def bad_debt_rate(self) -> float:
        return self.bad_du_no / self.total_du_no if self.total_du_no else 0.0


@dataclass(frozen=True, slots=True)
class InterestReportResult:
    output_path: Path
    group_count: int
    detail_count: int
    warnings: tuple[str, ...]
    info_messages: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class WarningDetail:
    category: str
    ma_kh: str = ""
    so_giai_ngan: str = ""
    ma_to: str = ""
    ten_kh: str = ""
    interest_collected: float = 0.0
    outstanding_balance: float = 0.0
    note: str = ""


def create_interest_report(
    request: InterestReportRequest,
    repository: CreditGroupRepository,
) -> InterestReportResult:
    if not isinstance(request.from_date, date) or not isinstance(request.to_date, date):
        raise InterestReportError("Vui lòng nhập kỳ thu lãi từ ngày và đến ngày.")
    if request.from_date > request.to_date:
        raise InterestReportError("Từ ngày không được lớn hơn đến ngày.")

    sktl_detection = detect_sktl_columns(request.interest_file)
    skck_detection = detect_skck_columns(request.debt_file)
    if sktl_detection.missing_required:
        raise InterestReportError(
            "File SKTL thiếu cột bắt buộc: "
            + ", ".join(sktl_detection.missing_required)
        )
    if skck_detection.missing_required:
        raise InterestReportError(
            "File SKCK thiếu cột bắt buộc: "
            + ", ".join(skck_detection.missing_required)
        )

    interest_rows, read_warnings = _read_sktl_rows(
        request.interest_file,
        sktl_detection,
        request.to_date,
        include_overdue_interest=request.include_overdue_interest,
    )
    debt_rows, debt_warnings = _read_skck_rows(
        request.debt_file,
        skck_detection,
        request.to_date,
    )

    warnings = list(read_warnings)
    warnings.extend(debt_warnings)

    interest_by_key = {_join_key(row.ma_kh, row.so_giai_ngan): row for row in interest_rows}
    debt_by_key = {_join_key(row.ma_kh, row.so_giai_ngan): row for row in debt_rows}
    rows_by_key: dict[tuple[str, str], InterestRow] = {}
    warning_details: list[WarningDetail] = []

    for key, interest_row in interest_by_key.items():
        debt_row = debt_by_key.get(key)
        if debt_row is None:
            note = "Có thu lãi trong kỳ nhưng không còn dư nợ cuối kỳ"
            rows_by_key[key] = InterestRow(
                ma_to=interest_row.ma_to,
                ma_kh=interest_row.ma_kh,
                ten_kh=interest_row.ten_kh,
                so_giai_ngan=interest_row.so_giai_ngan,
                du_no=0.0,
                ty_le_bao_bam=interest_row.ty_le_bao_bam,
                so_lai_da_thu=interest_row.so_lai_da_thu,
                lai_ton=interest_row.lai_ton,
                thu_lai_tu_ngay=interest_row.thu_lai_tu_ngay,
                thu_lai_den_ngay=interest_row.thu_lai_den_ngay,
                is_yellow_flag=interest_row.is_yellow_flag,
                source="SKTL_ONLY",
                debt_group="",
                loan_status="Không còn trong SKCK",
                note=note,
            )
            warning_details.append(
                WarningDetail(
                    category="SKTL-only",
                    ma_kh=interest_row.ma_kh,
                    so_giai_ngan=interest_row.so_giai_ngan,
                    ma_to=interest_row.ma_to,
                    ten_kh=interest_row.ten_kh,
                    interest_collected=interest_row.so_lai_da_thu,
                    outstanding_balance=0.0,
                    note=note,
                )
            )
            continue
        rows_by_key[key] = InterestRow(
            ma_to=debt_row.ma_to or interest_row.ma_to,
            ma_kh=debt_row.ma_kh or interest_row.ma_kh,
            ten_kh=debt_row.ten_kh or interest_row.ten_kh,
            so_giai_ngan=debt_row.so_giai_ngan or interest_row.so_giai_ngan,
            du_no=debt_row.du_no,
            ty_le_bao_bam=debt_row.ty_le_bao_bam,
            so_lai_da_thu=interest_row.so_lai_da_thu,
            lai_ton=debt_row.lai_ton,
            thu_lai_tu_ngay=interest_row.thu_lai_tu_ngay,
            thu_lai_den_ngay=interest_row.thu_lai_den_ngay,
            is_yellow_flag=debt_row.is_yellow_flag,
            source="SKTL_SKCK",
            debt_group=debt_row.debt_group,
            loan_status="Còn trong SKCK",
        )

    for debt_row in debt_rows:
        key = _join_key(debt_row.ma_kh, debt_row.so_giai_ngan)
        if key not in rows_by_key:
            note = "Có trong SKCK nhưng không phát sinh thu lãi trong kỳ"
            debt_row.source = "SKCK_ONLY"
            debt_row.loan_status = "Không có trong SKTL"
            debt_row.note = note
            rows_by_key[key] = debt_row
            warning_details.append(
                WarningDetail(
                    category="SKCK-only",
                    ma_kh=debt_row.ma_kh,
                    so_giai_ngan=debt_row.so_giai_ngan,
                    ma_to=debt_row.ma_to,
                    ten_kh=debt_row.ten_kh,
                    interest_collected=0.0,
                    outstanding_balance=debt_row.du_no,
                    note=note,
                )
            )

    selected_group_codes = _selected_group_codes(request)
    selected_rows = list(rows_by_key.values())
    if selected_group_codes:
        selected_rows = [row for row in selected_rows if row.ma_to in selected_group_codes]
    selected_warning_details = [
        detail
        for detail in warning_details
        if not selected_group_codes or detail.ma_to in selected_group_codes
    ]
    sktl_only_count = sum(1 for detail in selected_warning_details if detail.category == "SKTL-only")
    if sktl_only_count:
        warnings.append(
            f"Có {sktl_only_count} khoản có trong SKTL nhưng không có trong SKCK; "
            "đã đưa vào bảng kê với dư nợ cuối kỳ = 0."
        )
    skck_only_count = sum(1 for detail in selected_warning_details if detail.category == "SKCK-only")
    if skck_only_count:
        warnings.append(
            f"Có {skck_only_count} khoản có trong SKCK nhưng không có trong SKTL; "
            "đã đưa vào bảng kê với lãi đã thu = 0 để đối chiếu cuối kỳ."
        )

    if not selected_rows:
        raise InterestReportError("Không có dữ liệu tổ vay vốn phù hợp để tạo bảng kê.")

    groups_by_ma_to = {group.ma_to: group for group in repository.list_groups()}
    summaries: dict[str, InterestGroupSummary] = {}

    debt_totals = _debt_totals_from_rows(debt_rows)
    for row in selected_rows:
        if not row.ma_to:
            warnings.append(f"Dòng khoản vay {row.so_giai_ngan or row.ma_kh}: thiếu MaToVayVon.")
            continue
        group = groups_by_ma_to.get(row.ma_to)
        if group is None:
            warnings.append(f"MaTo {row.ma_to} chưa có trong dữ liệu credit_groups.")
            continue
        if row.ma_to not in summaries:
            rate = repository.get_or_create_commission_rate(row.ma_to)
            group_rule_settings, uses_custom_rule = repository.get_effective_commission_rule(row.ma_to)
            total_du_no, bad_du_no = debt_totals.get(row.ma_to, (0.0, 0.0))
            summaries[row.ma_to] = InterestGroupSummary(
                group=group,
                rate=rate,
                total_du_no=total_du_no,
                bad_du_no=bad_du_no,
                rule_settings=group_rule_settings,
                uses_custom_rule=uses_custom_rule,
            )
        summaries[row.ma_to].rows.append(row)

    if not summaries:
        raise InterestReportError("Không có MaTo nào khớp với dữ liệu tổ vay vốn đã import.")

    branch_report_name = repository.get_branch_report_name()
    info_messages = tuple(
        _summary_info_message(summary)
        for summary in sorted(summaries.values(), key=lambda item: item.group.ma_to)
    )
    _export_interest_report(
        request,
        summaries.values(),
        branch_report_name,
        tuple(dict.fromkeys(warnings)),
        tuple(selected_warning_details),
    )
    detail_count = sum(len(summary.rows) for summary in summaries.values())
    return InterestReportResult(
        output_path=request.output_path,
        group_count=len(summaries),
        detail_count=detail_count,
        warnings=tuple(dict.fromkeys(warnings)),
        info_messages=info_messages,
    )


def _summary_info_message(
    summary: InterestGroupSummary,
) -> str:
    rule_settings = summary.rule_settings or CreditCommissionRuleSettings()
    pay_rate = calculate_commission_pay_rate(
        summary.collection_rate,
        summary.bad_debt_rate,
        rule_settings,
    )
    commission = _commission_breakdown(summary, pay_rate)
    return "\n".join(
        (
            f"Tổ {summary.group.ma_to}:",
            f"- Tỷ lệ thu lãi: {summary.collection_rate:.2%}",
            f"- Tỷ lệ nợ xấu: {summary.bad_debt_rate:.2%}",
            f"- Lãi không BĐ: {_format_money(summary.total_lai_khong_bd)}",
            f"- Tỷ lệ hoa hồng không BĐ: {_format_percent(_no_secured_base_rate(summary.rate))}",
            f"- Hoa hồng không BĐ thực nhận: {_format_money(commission['unsecured'])}",
            f"- Lãi có BĐTS: {_format_money(summary.total_lai_co_bd)}",
            f"- Tỷ lệ hoa hồng có BĐTS: {_format_percent(_secured_base_rate(summary.rate))}",
            f"- Hoa hồng có BĐTS thực nhận: {_format_money(commission['secured'])}",
            f"- Tỷ lệ chi điều kiện: {pay_rate:.0%}",
            f"- Điều kiện chi: {'riêng theo tổ' if summary.uses_custom_rule else 'chung'}",
        )
    )


def _format_money(value: float) -> str:
    return f"{round_money(value):,.0f}"


def _format_percent(value: float) -> str:
    return f"{value:.0%}"


def _selected_group_codes(request: InterestReportRequest) -> set[str]:
    if request.selected_group_codes:
        return {code for code in request.selected_group_codes if code}
    return {request.ma_to} if request.ma_to else set()


def _branch_report_title(branch_report_name: str) -> str:
    name = branch_report_name.strip()
    if not name:
        return DEFAULT_BRANCH_REPORT_NAME
    return name.upper()


def validate_interest_report_columns(path: Path, required_columns: Iterable[str]) -> None:
    table = _read_table(path)
    headers = {header: index for index, header in enumerate(table.headers)}
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise InterestReportError(
            "File thiếu cột bắt buộc: " + ", ".join(missing)
        )


def detect_sktl_columns(path: Path) -> ColumnDetectionResult:
    return _detect_columns(
        path,
        file_kind="SKTL",
        aliases=SKTL_COLUMN_ALIASES,
        required_fields=SKTL_REQUIRED_FIELDS,
        optional_fields=("overdue_interest_collected",),
    )


def detect_skck_columns(path: Path) -> ColumnDetectionResult:
    detection = _detect_columns(
        path,
        file_kind="SKCK",
        aliases=SKCK_COLUMN_ALIASES,
        required_fields=SKCK_REQUIRED_FIELDS,
        optional_fields=("remaining_interest", "interest_receivable", "overdue_interest"),
    )
    missing_required = list(detection.missing_required)
    if "remaining_interest" not in detection.field_to_header:
        for field_name in SKCK_LAITON_FORMULA_FIELDS:
            if field_name not in detection.field_to_header and field_name not in missing_required:
                missing_required.append(field_name)
    return ColumnDetectionResult(
        file_kind=detection.file_kind,
        sheet_name=detection.sheet_name,
        headers=detection.headers,
        field_to_header=detection.field_to_header,
        missing_required=tuple(missing_required),
        missing_optional=detection.missing_optional,
    )


def calculate_commission_pay_rate(
    collection_rate: float,
    bad_debt_rate: float,
    settings: CreditCommissionRuleSettings,
) -> float:
    bad_threshold = settings.bad_debt_threshold / 100
    if bad_debt_rate >= bad_threshold:
        return settings.bad_debt_pay / 100
    if collection_rate >= settings.interest_min_3 / 100:
        return settings.interest_pay_3 / 100
    if settings.interest_min_2 / 100 <= collection_rate < settings.interest_max_2 / 100:
        return settings.interest_pay_2 / 100
    if settings.interest_min_1 / 100 <= collection_rate < settings.interest_max_1 / 100:
        return settings.interest_pay_1 / 100
    return 0.0


@dataclass(frozen=True, slots=True)
class _TableData:
    sheet_name: str
    headers: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]
    datemode: int = 0


def _read_sktl_rows(
    path: Path,
    detection: ColumnDetectionResult,
    to_date: date,
    *,
    include_overdue_interest: bool,
) -> tuple[list[InterestRow], tuple[str, ...]]:
    table = _read_table(path)
    warnings: list[str] = []
    overdue_header = detection.field_to_header.get("overdue_interest_collected")
    if include_overdue_interest and overdue_header is None:
        warnings.append("File SKTL không có cột lãi quá hạn trong kỳ; chỉ dùng SoLaiDaThuTrongKy.")
    rows: list[InterestRow] = []
    for source_row in table.rows:
        group_code = _mapped_text(source_row, detection, "group_code")
        if not group_code:
            continue
        interest_collected = _mapped_number(source_row, detection, "interest_collected")
        if include_overdue_interest and overdue_header is not None:
            interest_collected += _number(source_row.get(overdue_header))
        lai_ton = (
            _calculate_lai_ton_from_vba_formula(source_row, detection, to_date, table.datemode)
            if "remaining_interest" not in detection.field_to_header
            else _mapped_number(source_row, detection, "remaining_interest")
        )
        period_from_date, period_to_date = _collection_period_dates(
            source_row,
            detection,
            table.datemode,
            interest_collected,
        )
        rows.append(
            InterestRow(
                ma_to=group_code,
                ma_kh=_mapped_text(source_row, detection, "customer_code"),
                ten_kh=_mapped_text(source_row, detection, "customer_name"),
                so_giai_ngan=_mapped_text(source_row, detection, "loan_number"),
                du_no=0.0,
                ty_le_bao_bam=_mapped_number(source_row, detection, "secured_ratio"),
                so_lai_da_thu=interest_collected,
                lai_ton=lai_ton,
                thu_lai_tu_ngay=period_from_date,
                thu_lai_den_ngay=period_to_date,
                source="SKTL",
            )
        )
    return rows, tuple(warnings)


def _read_skck_rows(
    path: Path,
    detection: ColumnDetectionResult,
    to_date: date,
) -> tuple[list[InterestRow], tuple[str, ...]]:
    table = _read_table(path)
    warnings: list[str] = []
    uses_formula_lai_ton = "remaining_interest" not in detection.field_to_header
    if uses_formula_lai_ton:
        warnings.append(
            "File SKCK không có cột LaiTon; LaiTon được tính theo công thức VBA từ "
            "DuNo, LaiSuat, NgayGiaiNgan, NgayTraLaiCuoiCung."
        )
    rows: list[InterestRow] = []
    for source_row in table.rows:
        group_code = _mapped_text(source_row, detection, "group_code")
        if not group_code:
            continue
        lai_ton = (
            _calculate_lai_ton_from_vba_formula(source_row, detection, to_date, table.datemode)
            if uses_formula_lai_ton
            else _mapped_number(source_row, detection, "remaining_interest")
        )
        rows.append(
            InterestRow(
                ma_to=group_code,
                ma_kh=_mapped_text(source_row, detection, "customer_code"),
                ten_kh=_mapped_text(source_row, detection, "customer_name"),
                so_giai_ngan=_mapped_text(source_row, detection, "loan_number"),
                du_no=_mapped_number(source_row, detection, "outstanding_balance"),
                ty_le_bao_bam=_mapped_number(source_row, detection, "secured_ratio"),
                so_lai_da_thu=0.0,
                lai_ton=lai_ton,
                source="SKCK",
                debt_group=_mapped_text(source_row, detection, "debt_group"),
            )
        )
    return rows, tuple(warnings)


def _debt_totals_from_rows(rows: Iterable[InterestRow]) -> dict[str, tuple[float, float]]:
    totals: dict[str, list[float]] = {}
    for row in rows:
        bucket = totals.setdefault(row.ma_to, [0.0, 0.0])
        bucket[0] += row.du_no
        if any(value in row.debt_group for value in ("3", "4", "5")):
            bucket[1] += row.du_no
    return {ma_to: (values[0], values[1]) for ma_to, values in totals.items()}


def _detect_columns(
    path: Path,
    *,
    file_kind: str,
    aliases: dict[str, tuple[str, ...]],
    required_fields: tuple[str, ...],
    optional_fields: tuple[str, ...],
) -> ColumnDetectionResult:
    table = _read_table(path)
    normalized_headers = {_normalize_header(header): header for header in table.headers}
    field_to_header: dict[str, str] = {}
    for field_name, candidates in aliases.items():
        for candidate in candidates:
            header = normalized_headers.get(_normalize_header(candidate))
            if header:
                field_to_header[field_name] = header
                break
    missing_required = tuple(
        field_name for field_name in required_fields if field_name not in field_to_header
    )
    missing_optional = tuple(
        field_name for field_name in optional_fields if field_name not in field_to_header
    )
    return ColumnDetectionResult(
        file_kind=file_kind,
        sheet_name=table.sheet_name,
        headers=table.headers,
        field_to_header=field_to_header,
        missing_required=missing_required,
        missing_optional=missing_optional,
    )


def _read_table(path: Path) -> _TableData:
    suffix = path.suffix.casefold()
    if suffix == ".xls":
        return _read_xls_table(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_table(path)
    raise InterestReportError(f"Không hỗ trợ định dạng file {path.suffix}.")


def _read_xls_table(path: Path) -> _TableData:
    workbook = xlrd.open_workbook(path, logfile=io.StringIO())
    sheet = workbook.sheet_by_index(0)
    headers = tuple(str(sheet.cell_value(0, col) or "").strip() for col in range(sheet.ncols))
    rows: list[dict[str, Any]] = []
    for row_index in range(1, sheet.nrows):
        row: dict[str, Any] = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            cell = sheet.cell(row_index, col_index)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate_as_datetime(value, workbook.datemode).date()
            row[header] = value
        rows.append(row)
    return _TableData(
        sheet_name=sheet.name,
        headers=tuple(header for header in headers if header),
        rows=tuple(rows),
        datemode=workbook.datemode,
    )


def _read_xlsx_table(path: Path) -> _TableData:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = tuple(str(value or "").strip() for value in raw_headers)
        rows: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
            rows.append(row)
        return _TableData(
            sheet_name=sheet.title,
            headers=tuple(header for header in headers if header),
            rows=tuple(rows),
            datemode=0,
        )
    finally:
        workbook.close()


def _mapped_value(
    row: dict[str, Any],
    detection: ColumnDetectionResult,
    field_name: str,
) -> Any:
    header = detection.field_to_header.get(field_name)
    return row.get(header) if header else None


def _mapped_text(
    row: dict[str, Any],
    detection: ColumnDetectionResult,
    field_name: str,
) -> str:
    return _text(_mapped_value(row, detection, field_name))


def _mapped_number(
    row: dict[str, Any],
    detection: ColumnDetectionResult,
    field_name: str,
) -> float:
    return _number(_mapped_value(row, detection, field_name))


def _calculate_lai_ton_from_vba_formula(
    row: dict[str, Any],
    detection: ColumnDetectionResult,
    to_date: date,
    datemode: int,
) -> float:
    interest_rate = _mapped_number(row, detection, "interest_rate")
    outstanding_balance = _mapped_number(row, detection, "outstanding_balance")
    disbursement_date = _to_date(_mapped_value(row, detection, "disbursement_date"), datemode)
    last_interest_date = _to_date(_mapped_value(row, detection, "last_interest_date"), datemode)
    if interest_rate <= 0 or outstanding_balance <= 0 or disbursement_date is None:
        return 0.0
    if last_interest_date is None:
        days = (to_date - disbursement_date).days
    elif last_interest_date > to_date:
        days = 0
    else:
        days = (to_date - max(disbursement_date, last_interest_date)).days
    if days <= 0:
        return 0.0
    return outstanding_balance * (interest_rate / 100) * days / 365


def _collection_period_dates(
    row: dict[str, Any],
    detection: ColumnDetectionResult,
    datemode: int,
    interest_collected: float,
) -> tuple[date | None, date | None]:
    existing_from = _to_date(_mapped_value(row, detection, "period_from_date"), datemode)
    existing_to = _to_date(_mapped_value(row, detection, "period_to_date"), datemode)
    if existing_from is not None or existing_to is not None:
        return existing_from, existing_to
    return _calculate_collection_period_dates(row, detection, datemode, interest_collected)


def _calculate_collection_period_dates(
    row: dict[str, Any],
    detection: ColumnDetectionResult,
    datemode: int,
    interest_collected: float,
) -> tuple[date | None, date | None]:
    interest_rate = _mapped_number(row, detection, "interest_rate")
    outstanding_balance = _mapped_number(row, detection, "outstanding_balance")
    principal_collected = _mapped_number(row, detection, "principal_collected_period")
    last_principal_date = _to_date(_mapped_value(row, detection, "last_principal_date"), datemode)
    last_interest_date = _to_date(_mapped_value(row, detection, "last_interest_date"), datemode)

    if outstanding_balance > 0:
        if last_interest_date is None:
            return None, None
        total_days = 0.0
        if interest_rate > 0:
            daily_current_balance = outstanding_balance * (interest_rate / 100) / 365
            daily_before_principal = (outstanding_balance + principal_collected) * (interest_rate / 100) / 365
            if (
                principal_collected > 0
                and last_principal_date is not None
                and last_interest_date >= last_principal_date
            ):
                days_after_principal = (last_interest_date - last_principal_date).days + 1
                interest_after_principal = daily_current_balance * days_after_principal
                if interest_collected >= interest_after_principal and daily_before_principal:
                    interest_before_principal = interest_collected - interest_after_principal
                    total_days = days_after_principal + interest_before_principal / daily_before_principal
                elif daily_current_balance:
                    total_days = interest_collected / daily_current_balance
            elif daily_current_balance:
                total_days = interest_collected / daily_current_balance
        if total_days > 0:
            start_date = last_interest_date - timedelta(days=round(total_days) - 1)
            return start_date, last_interest_date
        return None, last_interest_date

    if last_principal_date is None:
        return None, None
    to_date = last_principal_date - timedelta(days=1)
    total_days = 0.0
    if principal_collected * interest_rate > 0:
        daily_before_principal = principal_collected * (interest_rate / 100) / 365
        total_days = interest_collected / daily_before_principal if daily_before_principal else 0.0
    if total_days > 0:
        return last_principal_date - timedelta(days=round(total_days)), to_date
    return None, to_date


def _to_date(value: Any, datemode: int) -> date | None:
    if value in (None, "", 0):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return xlrd.xldate_as_datetime(value, datemode).date()
        except (OverflowError, ValueError, xlrd.XLDateError):
            return None
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _join_key(customer_code: str, loan_number: str) -> tuple[str, str]:
    return customer_code.strip().casefold(), loan_number.strip().casefold()


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().casefold()
    decomposed = unicodedata.normalize("NFD", text)
    text = "".join(char for char in decomposed if unicodedata.category(char) != "Mn")
    return "".join(char for char in text if char.isalnum())


def _export_interest_report(
    request: InterestReportRequest,
    summaries: Iterable[InterestGroupSummary],
    branch_report_name: str,
    warnings: tuple[str, ...],
    warning_details: tuple[WarningDetail, ...] = (),
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    summary_sheet.append(SUMMARY_HEADERS)

    for summary in sorted(summaries, key=lambda item: item.group.ma_to):
        rule_settings = summary.rule_settings or CreditCommissionRuleSettings()
        pay_rate = calculate_commission_pay_rate(
            summary.collection_rate,
            summary.bad_debt_rate,
            rule_settings,
        )
        commission = _commission_breakdown(summary, pay_rate)
        _write_group_sheet(
            workbook,
            summary,
            pay_rate,
            commission,
            request,
            branch_report_name,
        )
        summary_sheet.append(
            (
                summary.group.ma_to,
                summary.group.ten_to,
                summary.group.ten_to_truong,
                len(summary.rows),
                summary.total_du_no,
                summary.total_lai_co_bd,
                summary.total_lai_khong_bd,
                summary.total_lai_ton,
                summary.total_lai_phai_thu,
                summary.collection_rate,
                summary.bad_debt_rate,
                pay_rate,
                commission["secured"],
                commission["unsecured"],
                commission["total"],
                commission["to_truong"],
                commission["cap_xa"],
                commission["cap_huyen"],
                commission["cap_tinh"],
                commission["cap_tw"],
                summary.warning,
            )
        )

    if warnings or warning_details:
        warning_sheet = workbook.create_sheet(WARNING_SHEET_NAME)
        warning_sheet.append(
            (
                "STT",
                "Loai",
                "MaKH",
                "SoGiaiNgan",
                "MaTo",
                "TenKH",
                "LaiDaThu",
                "DuNoCuoiKy",
                "GhiChu",
            )
        )
        row_index = 1
        for detail in warning_details:
            warning_sheet.append(
                (
                    row_index,
                    detail.category,
                    detail.ma_kh,
                    detail.so_giai_ngan,
                    detail.ma_to,
                    detail.ten_kh,
                    detail.interest_collected,
                    detail.outstanding_balance,
                    detail.note,
                )
            )
            row_index += 1
        if warning_details and warnings:
            warning_sheet.append(())
        warning_sheet.append(("Thông tin tổng hợp",))
        for index, warning in enumerate(warnings, start=1):
            warning_sheet.append((index, "TongHop", "", "", "", "", "", "", warning))
        _format_sheet(warning_sheet)

    _format_sheet(summary_sheet)
    request.output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(request.output_path)
    workbook.close()


def _write_group_sheet(
    workbook: Workbook,
    summary: InterestGroupSummary,
    pay_rate: float,
    commission: dict[str, float],
    request: InterestReportRequest,
    branch_report_name: str,
) -> None:
    title = _safe_sheet_title(summary.group.ma_to)
    sheet = workbook.create_sheet(title)
    secured_base_rate = _secured_base_rate(summary.rate)
    no_secured_base_rate = _no_secured_base_rate(summary.rate)

    group_name = summary.group.ten_to or "DỊCH VỤ TÍN DỤNG"
    title_prefix = (
        "BẢNG KÊ THU LÃI TRÍCH HOA HỒNG "
        if group_name.casefold().startswith("tổ")
        else "BẢNG KÊ THU LÃI TRÍCH HOA HỒNG TỔ "
    )
    sheet["A1"] = "NGÂN HÀNG NÔNG NGHIỆP"
    sheet["A2"] = "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM"
    sheet["A3"] = _branch_report_title(branch_report_name)
    sheet["A5"] = title_prefix + group_name.upper()
    sheet["A6"] = f"(Từ ngày {request.from_date:%d/%m/%Y} đến ngày {request.to_date:%d/%m/%Y})"
    sheet["B7"] = "Mã chi hội:"
    sheet["C7"] = summary.group.ma_to
    sheet["B8"] = "Họ tên tổ trưởng:"
    sheet["C8"] = summary.group.ten_to_truong
    sheet["B9"] = (
        "Tỷ lệ trích hoa hồng không bảo đảm tài sản: "
        f"{_format_percent(no_secured_base_rate)}"
    )
    sheet["B10"] = (
        "Tỷ lệ trích hoa hồng có bảo đảm tài sản: "
        f"{_format_percent(secured_base_rate)}"
    )
    sheet["M11"] = "Đơn vị tính: đồng."
    for merge_range in ("A1:D1", "A2:D2", "A3:D3", "A5:N5", "A6:N6"):
        sheet.merge_cells(merge_range)

    for row_index in range(1, 12):
        for cell in sheet[row_index]:
            cell.font = Font(name="Times New Roman", size=12)
    for address in ("A3", "A5", "A6", "B7", "B8", "B9", "B10"):
        sheet[address].font = Font(name="Times New Roman", size=12, bold=True)
    for row_index in (1, 2, 3, 5, 6):
        sheet.cell(row_index, 1).alignment = Alignment(horizontal="center", vertical="center")
    sheet["M11"].font = Font(name="Times New Roman", size=12, italic=True)

    header_row = 12
    for column_index, header in enumerate(DETAIL_HEADERS, start=1):
        cell = sheet.cell(header_row, column_index, header)
        cell.font = Font(name="Times New Roman", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    sorted_rows = sorted(
        summary.rows,
        key=lambda item: (item.ma_kh.casefold(), item.so_giai_ngan.casefold()),
    )
    detail_commissions = _detail_commissions_matching_group_total(
        sorted_rows,
        secured_base_rate=secured_base_rate,
        no_secured_base_rate=no_secured_base_rate,
        expected_total=commission["secured_base"] + commission["unsecured_base"],
    )
    for index, row in enumerate(sorted_rows, start=1):
        excel_row = header_row + index
        sheet.append(
            (
                index,
                row.ma_to,
                row.ma_kh,
                row.ten_kh,
                row.so_giai_ngan,
                row.du_no,
                row.thu_lai_tu_ngay,
                row.thu_lai_den_ngay,
                row.lai_phai_thu,
                row.lai_co_bd,
                row.lai_khong_bd,
                row.lai_ton,
                detail_commissions[index - 1],
                "",
            )
        )
        if row.is_yellow_flag:
            sheet.cell(excel_row, 7).fill = PatternFill("solid", fgColor="FFF2CC")

    total_row = sheet.max_row + 1
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    sheet.cell(total_row, 1, "Tổng cộng")
    sheet.cell(total_row, 6, summary.total_du_no)
    sheet.cell(total_row, 9, summary.total_lai_phai_thu)
    sheet.cell(total_row, 10, summary.total_lai_co_bd)
    sheet.cell(total_row, 11, summary.total_lai_khong_bd)
    sheet.cell(total_row, 12, summary.total_lai_ton)
    sheet.cell(total_row, 13, sum(detail_commissions))
    for cell in sheet[total_row][:14]:
        cell.font = Font(bold=True)
    sheet.cell(total_row, 1).alignment = Alignment(horizontal="center")
    _apply_table_border(sheet, header_row, total_row, 14)

    metric_row = total_row + 2
    sheet.cell(metric_row, 2, "Tổng lãi phải thu:")
    sheet.cell(metric_row, 4, summary.total_lai_phai_thu)
    sheet.cell(metric_row, 5, "Tỷ lệ thu lãi:")
    sheet.cell(metric_row, 7, summary.collection_rate)
    sheet.cell(metric_row, 9, f"Tỷ lệ chi: {pay_rate:.0%}")
    sheet.cell(metric_row + 1, 2, f"Tổng lãi thực thu đến {request.to_date:%d/%m/%Y}:")
    sheet.cell(metric_row + 1, 4, summary.total_lai_co_bd + summary.total_lai_khong_bd)
    sheet.cell(metric_row + 1, 5, "Tỷ lệ nợ xấu:")
    sheet.cell(metric_row + 1, 7, summary.bad_debt_rate)
    sheet.cell(metric_row + 1, 9, "Đủ điều kiện chi" if pay_rate else "Không đủ điều kiện chi")
    for cell in (sheet.cell(metric_row, 4), sheet.cell(metric_row + 1, 4)):
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    for cell in (sheet.cell(metric_row, 7), sheet.cell(metric_row + 1, 7)):
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for merge_range in (
        f"B{metric_row}:C{metric_row}",
        f"E{metric_row}:F{metric_row}",
        f"G{metric_row}:H{metric_row}",
        f"I{metric_row}:J{metric_row}",
        f"B{metric_row + 1}:C{metric_row + 1}",
        f"E{metric_row + 1}:F{metric_row + 1}",
        f"G{metric_row + 1}:H{metric_row + 1}",
        f"I{metric_row + 1}:J{metric_row + 1}",
    ):
        sheet.merge_cells(merge_range)
    _apply_table_border(sheet, metric_row, metric_row + 1, 10, start_column=2)

    commission_row = total_row + 5
    commission_lines = (
        (f"*HOA HỒNG ĐỐI VỚI KHOẢN KHÔNG CÓ TSĐB ({_format_percent(no_secured_base_rate)})", commission["unsecured_base"]),
        (f"* HOA HỒNG THỰC NHẬN ({pay_rate:.0%})", commission["unsecured"]),
        ("1", "Tổ trưởng", summary.rate.no_secured_to_truong / 100, commission["no_secured_to_truong"]),
        ("2", "Cấp trung ương", summary.rate.no_secured_cap_tw / 100, commission["no_secured_cap_tw"]),
        ("3", "Cấp tỉnh", summary.rate.no_secured_cap_tinh / 100, commission["no_secured_cap_tinh"]),
        ("4", "Cấp huyện", summary.rate.no_secured_cap_huyen / 100, commission["no_secured_cap_huyen"]),
        ("5", "Cấp xã", summary.rate.no_secured_cap_xa / 100, commission["no_secured_cap_xa"]),
        (f"*HOA HỒNG ĐỐI VỚI KHOẢN CÓ TSĐB ({_format_percent(secured_base_rate)})", commission["secured_base"]),
        (f"* HOA HỒNG THỰC NHẬN ({pay_rate:.0%})", commission["secured"]),
        ("1", "Tổ trưởng", summary.rate.secured_to_truong / 100, commission["secured_to_truong"]),
        ("2", "Cấp xã", summary.rate.secured_cap_xa / 100, commission["secured_cap_xa"]),
    )
    for offset, line in enumerate(commission_lines):
        row_index = commission_row + offset
        if len(line) == 2:
            sheet.cell(row_index, 2, line[0])
            sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=9)
            sheet.cell(row_index, 10, line[1])
            sheet.cell(row_index, 11, "đồng")
            sheet.cell(row_index, 2).font = Font(name="Times New Roman", size=12, bold=True)
        else:
            sheet.cell(row_index, 2, line[0])
            sheet.cell(row_index, 3, line[1])
            sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=4)
            sheet.cell(row_index, 5, line[2])
            sheet.merge_cells(start_row=row_index, start_column=5, end_row=row_index, end_column=9)
            sheet.cell(row_index, 10, line[3])
            sheet.cell(row_index, 11, "đồng")
    _apply_table_border(sheet, commission_row, commission_row + len(commission_lines) - 1, 11, start_column=2)

    signature_row = commission_row + len(commission_lines) + 2
    sheet.merge_cells(start_row=signature_row, start_column=12, end_row=signature_row, end_column=14)
    sheet.cell(signature_row, 12, "Ngày.....tháng....năm 20.......")
    sheet.cell(signature_row, 12).font = Font(name="Times New Roman", size=12, italic=True)
    sheet.cell(signature_row, 12).alignment = Alignment(horizontal="center", vertical="center")
    signature_title_row = signature_row + 1
    sheet.merge_cells(start_row=signature_title_row, start_column=1, end_row=signature_title_row, end_column=4)
    sheet.merge_cells(start_row=signature_title_row, start_column=7, end_row=signature_title_row, end_column=9)
    sheet.merge_cells(start_row=signature_title_row, start_column=12, end_row=signature_title_row, end_column=14)
    sheet.cell(signature_title_row, 1, "TỔ TRƯỞNG")
    sheet.cell(signature_title_row, 7, "NGƯỜI KIỂM TRA")
    sheet.cell(signature_title_row, 12, "KIỂM SOÁT")
    for column in (1, 7, 12):
        cell = sheet.cell(signature_title_row, column)
        cell.font = Font(name="Times New Roman", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _format_group_sheet(sheet)


def _apply_table_border(
    sheet: Any,
    start_row: int,
    end_row: int,
    end_column: int,
    *,
    start_column: int = 1,
) -> None:
    thin = Side(style="thin", color="000000")
    hair = Side(style="hair", color="000000")
    for row in sheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
    ):
        for cell in row:
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin if cell.row in (start_row, end_row) else hair,
                bottom=thin if cell.row in (start_row, end_row) else hair,
            )


def _format_group_sheet(sheet: Any) -> None:
    widths = {
        "A": 6,
        "B": 20,
        "C": 18,
        "D": 24,
        "E": 21,
        "F": 14,
        "G": 15,
        "H": 17,
        "I": 14,
        "J": 15,
        "K": 18,
        "L": 12,
        "M": 17,
        "N": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.font = Font(
                name="Times New Roman",
                size=12,
                bold=cell.font.bold,
                italic=cell.font.italic,
            )
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="center",
                wrap_text=False,
            )

    center_columns = ("A", "B", "C", "E", "G", "H")
    for column in center_columns:
        for cell in sheet[column]:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for cell in sheet["D"]:
        if cell.value is not None:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for cell in sheet["E"]:
        if cell.value is not None:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for row_index in range(7, 11):
        for column_index in (2, 3):
            cell = sheet.cell(row_index, column_index)
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    money_columns = ("F", "I", "J", "K", "L", "M")
    for column in money_columns:
        for cell in sheet[column]:
            if cell.row > 12:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    for column in ("G", "H"):
        for cell in sheet[column]:
            if cell.row > 12:
                cell.number_format = "dd/mm/yyyy"
    for column in ("E", "G"):
        for cell in sheet[column]:
            if cell.row > 12 and isinstance(cell.value, float) and 0 <= cell.value <= 1:
                cell.number_format = "0.00%"
    for row in sheet.iter_rows(min_col=10, max_col=10):
        for cell in row:
            if cell.row > 12:
                cell.number_format = '#,##0'

    for row_index in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row_index, 2).value or "")
        if label.startswith("Tổng lãi phải thu:") or label.startswith("Tổng lãi thực thu đến"):
            money_cell = sheet.cell(row_index, 4)
            money_cell.number_format = '#,##0'
            money_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        if label.startswith("Tổng lãi phải thu:") or label.startswith("Tổng lãi thực thu đến"):
            percent_cell = sheet.cell(row_index, 7)
            percent_cell.number_format = "0.00%"
            percent_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        if "Ngày.....tháng....năm 20......." in values or "KIỂM SOÁT" in values:
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    sheet.freeze_panes = "A13"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "12:12"


def _commission_breakdown(
    summary: InterestGroupSummary,
    pay_rate: float,
) -> dict[str, float]:
    rate = summary.rate
    unsecured_base = round_money(
        summary.total_lai_khong_bd * _no_secured_base_rate(rate)
    )
    secured_base = round_money(
        summary.total_lai_co_bd * _secured_base_rate(rate)
    )
    unsecured = round_money(unsecured_base * pay_rate)
    secured = round_money(secured_base * pay_rate)

    no_secured_to_truong = round_money(unsecured * rate.no_secured_to_truong / 100)
    no_secured_cap_huyen = round_money(unsecured * rate.no_secured_cap_huyen / 100)
    no_secured_cap_tinh = round_money(unsecured * rate.no_secured_cap_tinh / 100)
    no_secured_cap_tw = round_money(unsecured * rate.no_secured_cap_tw / 100)
    no_secured_cap_xa = round_money(
        unsecured
        - no_secured_to_truong
        - no_secured_cap_huyen
        - no_secured_cap_tinh
        - no_secured_cap_tw
    )

    secured_to_truong = round_money(secured * rate.secured_to_truong / 100)
    secured_cap_huyen = round_money(secured * rate.secured_cap_huyen / 100)
    secured_cap_tinh = round_money(secured * rate.secured_cap_tinh / 100)
    secured_cap_tw = round_money(secured * rate.secured_cap_tw / 100)
    secured_cap_xa = round_money(
        secured
        - secured_to_truong
        - secured_cap_huyen
        - secured_cap_tinh
        - secured_cap_tw
    )

    total = secured + unsecured
    to_truong = secured_to_truong + no_secured_to_truong
    cap_huyen = secured_cap_huyen + no_secured_cap_huyen
    cap_tinh = secured_cap_tinh + no_secured_cap_tinh
    cap_tw = secured_cap_tw + no_secured_cap_tw
    cap_xa = secured_cap_xa + no_secured_cap_xa
    return {
        "secured_base": secured_base,
        "unsecured_base": unsecured_base,
        "secured": secured,
        "unsecured": unsecured,
        "total": total,
        "to_truong": to_truong,
        "cap_xa": cap_xa,
        "cap_huyen": cap_huyen,
        "cap_tinh": cap_tinh,
        "cap_tw": cap_tw,
        "no_secured_to_truong": no_secured_to_truong,
        "no_secured_cap_xa": no_secured_cap_xa,
        "no_secured_cap_huyen": no_secured_cap_huyen,
        "no_secured_cap_tinh": no_secured_cap_tinh,
        "no_secured_cap_tw": no_secured_cap_tw,
        "secured_to_truong": secured_to_truong,
        "secured_cap_xa": secured_cap_xa,
        "secured_cap_huyen": secured_cap_huyen,
        "secured_cap_tinh": secured_cap_tinh,
        "secured_cap_tw": secured_cap_tw,
    }


def _secured_base_rate(rate: CreditGroupCommissionRate | None = None) -> float:
    if rate is None:
        return VBA_SECURED_COMMISSION_BASE_RATE
    return rate.base_secured_rate / 100


def _no_secured_base_rate(rate: CreditGroupCommissionRate | None = None) -> float:
    if rate is None:
        return VBA_NO_SECURED_COMMISSION_BASE_RATE
    return rate.base_no_secured_rate / 100


def round_money(value: float | int | Decimal) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _round_money(value: float | int | Decimal) -> int:
    return round_money(value)


def _detail_commissions_matching_group_total(
    rows: list[InterestRow],
    *,
    secured_base_rate: float,
    no_secured_base_rate: float,
    expected_total: float | int,
) -> list[int]:
    commissions = [
        round_money(
            row.lai_co_bd * secured_base_rate
            + row.lai_khong_bd * no_secured_base_rate
        )
        for row in rows
    ]
    difference = round_money(expected_total) - sum(commissions)
    if difference:
        for index in range(len(commissions) - 1, -1, -1):
            if commissions[index] != 0:
                # Adjust rounding residue so the detail column matches
                # the VBA-style group totals computed from summed interest.
                commissions[index] += difference
                break
    return commissions


def _headers_from_sheet(sheet: Any) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return {
        str(value or "").strip(): index
        for index, value in enumerate(header_row)
        if str(value or "").strip()
    }


def _row_dict(headers: dict[str, int], values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        header: values[index] if index < len(values) else None
        for header, index in headers.items()
    }


def _first_existing_header(headers: dict[str, int], candidates: tuple[str, ...]) -> str | None:
    for candidate in candidates:
        if candidate in headers:
            return candidate
    return None


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _text(value: Any) -> str:
    return str(value or "").strip()


def _safe_sheet_title(value: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if char in invalid else char for char in value).strip()
    return (cleaned or "ToVayVon")[:31]


def _format_sheet(sheet: Any) -> None:
    sheet.freeze_panes = sheet.freeze_panes or "A2"
    money_headers = {
        "TongDuNo",
        "DuNo",
        "LaiCoBD",
        "LaiKhongBD",
        "LaiTon",
        "LaiPhaiThu",
        "HoaHongGoc",
        "HoaHongThucNhan",
        "HHCoBD",
        "HHKhongBD",
        "TongHH",
        "HH_ToTruong",
        "HH_CapXa",
        "HH_CapHuyen",
        "HH_CapTinh",
        "HH_TW",
    }
    percent_headers = {"TyLeThuLai", "TyLeNoXau", "TyLeChi"}
    header_rows = [row for row in sheet.iter_rows() if any(cell.value for cell in row)]
    if not header_rows:
        return
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        if any(value in (*SUMMARY_HEADERS, *DETAIL_HEADERS, "STT", "Nội dung") for value in values):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            for column_index, cell in enumerate(row, start=1):
                header = str(cell.value or "")
                if header in money_headers:
                    for data_cell in sheet[get_column_letter(column_index)]:
                        if data_cell.row > cell.row:
                            data_cell.number_format = '#,##0'
                if header in percent_headers:
                    for data_cell in sheet[get_column_letter(column_index)]:
                        if data_cell.row > cell.row:
                            data_cell.number_format = '0.00%'
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def default_interest_report_output_path(directory: Path | None = None) -> Path:
    target_dir = directory or Path.cwd()
    return target_dir / f"BangKeThuLaiTo_{datetime.now():%Y%m%d}.xlsx"
