from __future__ import annotations

from contextlib import contextmanager
from dataclasses import replace
from pathlib import Path
import sqlite3
from typing import Iterator

from openpyxl import Workbook, load_workbook

from agribank_v3.features.credit.tovayvon.models import (
    COMMISSION_EXPORT_HEADERS,
    COMMISSION_RULE_EXPORT_HEADERS,
    DATA_TVV_HEADERS,
    CreditGroup,
    CreditGroupCommissionRate,
    CreditCommissionRuleSettings,
    now_text,
)


class CreditGroupRepositoryError(RuntimeError):
    pass


class CreditGroupRepository:
    """SQLite repository for Tổ vay vốn data and per-group commission rates."""

    def __init__(self, database_path: Path) -> None:
        self.database_path = Path(database_path)
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self.ensure_schema()

    def connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.database_path, timeout=15)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        connection.execute("PRAGMA journal_mode = WAL")
        connection.execute("PRAGMA busy_timeout = 15000")
        return connection

    @contextmanager
    def _database(self) -> Iterator[sqlite3.Connection]:
        database = self.connect()
        try:
            with database:
                yield database
        finally:
            database.close()

    def ensure_schema(self) -> None:
        try:
            with self._database() as database:
                database.executescript(
                    """
                    CREATE TABLE IF NOT EXISTS credit_groups (
                        ma_to TEXT PRIMARY KEY,
                        stt INTEGER NOT NULL DEFAULT 0,
                        ten_to TEXT NOT NULL DEFAULT '',
                        ten_tvv_day_du TEXT NOT NULL DEFAULT '',
                        xa TEXT NOT NULL DEFAULT '',
                        ma_to_truong TEXT NOT NULL DEFAULT '',
                        ten_to_truong TEXT NOT NULL DEFAULT '',
                        dia_chi TEXT NOT NULL DEFAULT '',
                        tk_to_truong TEXT NOT NULL DEFAULT '',
                        so_dien_thoai TEXT NOT NULL DEFAULT '',
                        to_hoi TEXT NOT NULL DEFAULT '',
                        tk_to_hoi_xa TEXT NOT NULL DEFAULT '',
                        to_chuc TEXT NOT NULL DEFAULT '',
                        ten_huyen TEXT NOT NULL DEFAULT '',
                        tk_huyen TEXT NOT NULL DEFAULT '',
                        ten_tinh TEXT NOT NULL DEFAULT '',
                        tk_tinh TEXT NOT NULL DEFAULT '',
                        ten_tw TEXT NOT NULL DEFAULT '',
                        tk_tw TEXT NOT NULL DEFAULT '',
                        uy_quyen TEXT NOT NULL DEFAULT '',
                        ttln_tw TEXT NOT NULL DEFAULT '',
                        ttln_tinh TEXT NOT NULL DEFAULT '',
                        active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0, 1)),
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL
                    );

                    CREATE TABLE IF NOT EXISTS credit_group_commission_rates (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        ma_to TEXT NOT NULL UNIQUE,
                        no_secured_to_truong REAL DEFAULT 80,
                        no_secured_cap_xa REAL DEFAULT 13,
                        no_secured_cap_huyen REAL DEFAULT 3.8,
                        no_secured_cap_tinh REAL DEFAULT 2.5,
                        no_secured_cap_tw REAL DEFAULT 0.7,
                        secured_to_truong REAL DEFAULT 90,
                        secured_cap_xa REAL DEFAULT 10,
                        secured_cap_huyen REAL DEFAULT 0,
                        secured_cap_tinh REAL DEFAULT 0,
                        secured_cap_tw REAL DEFAULT 0,
                        created_at TEXT,
                        updated_at TEXT,
                        FOREIGN KEY(ma_to) REFERENCES credit_groups(ma_to)
                    );

                    CREATE TABLE IF NOT EXISTS credit_commission_rules (
                        id INTEGER PRIMARY KEY CHECK(id = 1),
                        interest_min_1 REAL NOT NULL DEFAULT 85,
                        interest_max_1 REAL NOT NULL DEFAULT 90,
                        interest_pay_1 REAL NOT NULL DEFAULT 50,
                        interest_min_2 REAL NOT NULL DEFAULT 90,
                        interest_max_2 REAL NOT NULL DEFAULT 95,
                        interest_pay_2 REAL NOT NULL DEFAULT 90,
                        interest_min_3 REAL NOT NULL DEFAULT 95,
                        interest_pay_3 REAL NOT NULL DEFAULT 100,
                        bad_debt_threshold REAL NOT NULL DEFAULT 2,
                        bad_debt_pay REAL NOT NULL DEFAULT 0,
                        updated_at TEXT NOT NULL
                    );
                    """
                )
        except sqlite3.Error as exc:
            raise CreditGroupRepositoryError(
                f"Không thể khởi tạo dữ liệu tổ vay vốn: {exc}"
            ) from exc

    def save_group(self, group: CreditGroup) -> None:
        ma_to = group.ma_to.strip()
        if not ma_to:
            raise CreditGroupRepositoryError("Mã tổ không được để trống.")
        now = now_text()
        try:
            with self._database() as database:
                existing = database.execute(
                    "SELECT created_at FROM credit_groups WHERE ma_to = ?",
                    (ma_to,),
                ).fetchone()
                created_at = str(existing["created_at"]) if existing else now
                existing_rows = database.execute(
                    """
                    SELECT * FROM credit_groups
                    WHERE ma_to <> ?
                    ORDER BY
                        CASE WHEN stt <= 0 THEN 2147483647 ELSE stt END,
                        created_at,
                        ma_to COLLATE NOCASE
                    """,
                    (ma_to,),
                ).fetchall()
                ordered_groups = [self._group_from_row(row) for row in existing_rows]
                insert_index = self._insert_index_for_stt(group.stt, len(ordered_groups))
                ordered_groups.insert(insert_index, replace(group, ma_to=ma_to))
                for index, ordered_group in enumerate(ordered_groups, start=1):
                    row_created_at = created_at if ordered_group.ma_to == ma_to else ordered_group.created_at or now
                    self._upsert_group(
                        database,
                        replace(ordered_group, stt=index),
                        created_at=row_created_at,
                        updated_at=now,
                    )
                self._ensure_commission_rate(database, ma_to)
        except sqlite3.Error as exc:
            raise CreditGroupRepositoryError(f"Không thể lưu tổ vay vốn: {exc}") from exc

    def list_groups(self) -> list[CreditGroup]:
        with self._database() as database:
            rows = database.execute(
                "SELECT * FROM credit_groups ORDER BY stt, ma_to COLLATE NOCASE"
            ).fetchall()
        return [self._group_from_row(row) for row in rows]

    def resequence_group_stt(self) -> None:
        now = now_text()
        try:
            with self._database() as database:
                rows = database.execute(
                    """
                    SELECT * FROM credit_groups
                    ORDER BY
                        CASE WHEN stt <= 0 THEN 2147483647 ELSE stt END,
                        created_at,
                        ma_to COLLATE NOCASE
                    """
                ).fetchall()
                for index, row in enumerate(rows, start=1):
                    group = replace(self._group_from_row(row), stt=index)
                    self._upsert_group(
                        database,
                        group,
                        created_at=group.created_at or now,
                        updated_at=now,
                    )
        except sqlite3.Error as exc:
            raise CreditGroupRepositoryError(
                f"Không thể sắp xếp lại STT tổ vay vốn: {exc}"
            ) from exc

    def get_group(self, ma_to: str) -> CreditGroup | None:
        with self._database() as database:
            row = database.execute(
                "SELECT * FROM credit_groups WHERE ma_to = ?",
                (ma_to,),
            ).fetchone()
        return self._group_from_row(row) if row else None

    def get_commission_rate(self, ma_to: str) -> CreditGroupCommissionRate | None:
        with self._database() as database:
            row = database.execute(
                "SELECT * FROM credit_group_commission_rates WHERE ma_to = ?",
                (ma_to,),
            ).fetchone()
        return self._commission_from_row(row) if row else None

    def get_or_create_commission_rate(self, ma_to: str) -> CreditGroupCommissionRate:
        with self._database() as database:
            self._ensure_commission_rate(database, ma_to)
            row = database.execute(
                "SELECT * FROM credit_group_commission_rates WHERE ma_to = ?",
                (ma_to,),
            ).fetchone()
        if row is None:
            raise CreditGroupRepositoryError(
                f"Không thể tạo tỷ lệ hoa hồng mặc định cho tổ {ma_to}."
            )
        return self._commission_from_row(row)

    def save_commission_rate(self, rate: CreditGroupCommissionRate) -> None:
        errors = rate.validate()
        if errors:
            raise CreditGroupRepositoryError("\n".join(errors))
        if self.get_group(rate.ma_to) is None:
            raise CreditGroupRepositoryError(
                f"Không tìm thấy tổ vay vốn có mã {rate.ma_to}."
            )
        now = now_text()
        existing = self.get_commission_rate(rate.ma_to)
        created_at = existing.created_at if existing else now
        try:
            with self._database() as database:
                database.execute(
                    """
                    INSERT INTO credit_group_commission_rates(
                        ma_to, no_secured_to_truong, no_secured_cap_xa,
                        no_secured_cap_huyen, no_secured_cap_tinh,
                        no_secured_cap_tw, secured_to_truong, secured_cap_xa,
                        secured_cap_huyen, secured_cap_tinh, secured_cap_tw,
                        created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(ma_to) DO UPDATE SET
                        no_secured_to_truong = excluded.no_secured_to_truong,
                        no_secured_cap_xa = excluded.no_secured_cap_xa,
                        no_secured_cap_huyen = excluded.no_secured_cap_huyen,
                        no_secured_cap_tinh = excluded.no_secured_cap_tinh,
                        no_secured_cap_tw = excluded.no_secured_cap_tw,
                        secured_to_truong = excluded.secured_to_truong,
                        secured_cap_xa = excluded.secured_cap_xa,
                        secured_cap_huyen = excluded.secured_cap_huyen,
                        secured_cap_tinh = excluded.secured_cap_tinh,
                        secured_cap_tw = excluded.secured_cap_tw,
                        updated_at = excluded.updated_at
                    """,
                    (
                        rate.ma_to,
                        rate.no_secured_to_truong,
                        rate.no_secured_cap_xa,
                        rate.no_secured_cap_huyen,
                        rate.no_secured_cap_tinh,
                        rate.no_secured_cap_tw,
                        rate.secured_to_truong,
                        rate.secured_cap_xa,
                        rate.secured_cap_huyen,
                        rate.secured_cap_tinh,
                        rate.secured_cap_tw,
                        created_at,
                        now,
                    ),
                )
        except sqlite3.Error as exc:
            raise CreditGroupRepositoryError(
                f"Không thể lưu tỷ lệ hoa hồng: {exc}"
            ) from exc

    def delete_commission_rate(self, ma_to: str) -> None:
        with self._database() as database:
            database.execute(
                "DELETE FROM credit_group_commission_rates WHERE ma_to = ?",
                (ma_to,),
            )

    def reset_commission_rate_to_default(self, ma_to: str) -> CreditGroupCommissionRate:
        rate = CreditGroupCommissionRate.default_for_group(ma_to)
        self.save_commission_rate(rate)
        return self.get_or_create_commission_rate(ma_to)

    def copy_commission_rate(self, from_ma_to: str, to_ma_to: str) -> None:
        source = self.get_or_create_commission_rate(from_ma_to)
        self.save_commission_rate(
            replace(
                source,
                ma_to=to_ma_to,
                created_at="",
                updated_at="",
            )
        )

    def list_groups_with_commission_status(self) -> list[tuple[CreditGroup, bool]]:
        with self._database() as database:
            rows = database.execute(
                """
                SELECT g.*, r.ma_to AS rate_ma_to
                FROM credit_groups g
                LEFT JOIN credit_group_commission_rates r ON r.ma_to = g.ma_to
                ORDER BY g.stt, g.ma_to COLLATE NOCASE
                """
            ).fetchall()
        return [(self._group_from_row(row), bool(row["rate_ma_to"])) for row in rows]

    def get_commission_rule_settings(self) -> CreditCommissionRuleSettings:
        with self._database() as database:
            row = database.execute(
                "SELECT * FROM credit_commission_rules WHERE id = 1"
            ).fetchone()
            if row is None:
                settings = CreditCommissionRuleSettings(updated_at=now_text())
                self._save_commission_rule_settings(database, settings)
                return settings
        return self._commission_rule_settings_from_row(row)

    def save_commission_rule_settings(
        self,
        settings: CreditCommissionRuleSettings,
    ) -> None:
        errors = settings.validate()
        if errors:
            raise CreditGroupRepositoryError("\n".join(errors))
        with self._database() as database:
            self._save_commission_rule_settings(
                database,
                replace(settings, updated_at=now_text()),
            )

    def import_data_tvv(
        self,
        source_path: Path,
        *,
        update_commission_rules: bool = False,
    ) -> int:
        workbook = load_workbook(source_path, data_only=True)
        try:
            if "Data_TVV" not in workbook.sheetnames:
                raise CreditGroupRepositoryError("File không có sheet Data_TVV.")
            sheet = workbook["Data_TVV"]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(value or "").strip() for value in header_row]
            header_map = {
                header: index for index, header in enumerate(headers) if header
            }
            count = 0
            rule_settings: CreditCommissionRuleSettings | None = None
            max_col = max(len(DATA_TVV_HEADERS), len(headers))
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=2, max_col=max_col, values_only=True),
                start=2,
            ):
                if not row or not str(row[1] or "").strip():
                    continue
                group = CreditGroup.from_data_tvv_row(row[: len(DATA_TVV_HEADERS)])
                self.save_group(group)
                rate = self._commission_rate_from_import_row(row, header_map, group.ma_to)
                if rate is not None:
                    try:
                        self.save_commission_rate(rate)
                    except CreditGroupRepositoryError as exc:
                        raise CreditGroupRepositoryError(
                            f"Dòng {row_number}: {exc}"
                        ) from exc
                if update_commission_rules and rule_settings is None:
                    rule_settings = self._commission_rules_from_import_row(
                        row,
                        header_map,
                    )
                count += 1
            if update_commission_rules and rule_settings is not None:
                self.save_commission_rule_settings(rule_settings)
            return count
        finally:
            workbook.close()

    def export_data_tvv(self, output_path: Path, *, include_commission: bool = False) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data_TVV"
        headers = list(DATA_TVV_HEADERS)
        if include_commission:
            headers.extend(COMMISSION_EXPORT_HEADERS)
            headers.extend(COMMISSION_RULE_EXPORT_HEADERS)
        sheet.append(headers)
        rule_settings = (
            self.get_commission_rule_settings() if include_commission else None
        )
        for group in self.list_groups():
            row = list(group.to_data_tvv_row())
            if include_commission:
                rate = self.get_or_create_commission_rate(group.ma_to)
                row.extend(
                    (
                        rate.no_secured_to_truong,
                        rate.no_secured_cap_xa,
                        rate.no_secured_cap_huyen,
                        rate.no_secured_cap_tinh,
                        rate.no_secured_cap_tw,
                        rate.total_no_secured(),
                        rate.secured_to_truong,
                        rate.secured_cap_xa,
                        rate.secured_cap_huyen,
                        rate.secured_cap_tinh,
                        rate.secured_cap_tw,
                        rate.total_secured(),
                    )
                )
                if rule_settings is not None:
                    row.extend(
                        (
                            rule_settings.interest_min_1,
                            rule_settings.interest_max_1,
                            rule_settings.interest_pay_1,
                            rule_settings.interest_min_2,
                            rule_settings.interest_max_2,
                            rule_settings.interest_pay_2,
                            rule_settings.interest_min_3,
                            rule_settings.interest_pay_3,
                            rule_settings.bad_debt_threshold,
                            rule_settings.bad_debt_pay,
                        )
                    )
            sheet.append(row)
        workbook.save(output_path)
        workbook.close()
        return output_path

    def has_commission_rule_columns(self, source_path: Path) -> bool:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            if "Data_TVV" not in workbook.sheetnames:
                return False
            sheet = workbook["Data_TVV"]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = {str(value or "").strip() for value in header_row}
            return any(header in headers for header in COMMISSION_RULE_EXPORT_HEADERS)
        finally:
            workbook.close()

    @staticmethod
    def _insert_index_for_stt(stt: int, existing_count: int) -> int:
        if stt <= 0:
            return existing_count
        return min(stt - 1, existing_count)

    @staticmethod
    def _upsert_group(
        database: sqlite3.Connection,
        group: CreditGroup,
        *,
        created_at: str,
        updated_at: str,
    ) -> None:
        database.execute(
            """
            INSERT INTO credit_groups(
                ma_to, stt, ten_to, ten_tvv_day_du, xa, ma_to_truong,
                ten_to_truong, dia_chi, tk_to_truong, so_dien_thoai,
                to_hoi, tk_to_hoi_xa, to_chuc, ten_huyen, tk_huyen,
                ten_tinh, tk_tinh, ten_tw, tk_tw, uy_quyen, ttln_tw,
                ttln_tinh, active, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(ma_to) DO UPDATE SET
                stt = excluded.stt,
                ten_to = excluded.ten_to,
                ten_tvv_day_du = excluded.ten_tvv_day_du,
                xa = excluded.xa,
                ma_to_truong = excluded.ma_to_truong,
                ten_to_truong = excluded.ten_to_truong,
                dia_chi = excluded.dia_chi,
                tk_to_truong = excluded.tk_to_truong,
                so_dien_thoai = excluded.so_dien_thoai,
                to_hoi = excluded.to_hoi,
                tk_to_hoi_xa = excluded.tk_to_hoi_xa,
                to_chuc = excluded.to_chuc,
                ten_huyen = excluded.ten_huyen,
                tk_huyen = excluded.tk_huyen,
                ten_tinh = excluded.ten_tinh,
                tk_tinh = excluded.tk_tinh,
                ten_tw = excluded.ten_tw,
                tk_tw = excluded.tk_tw,
                uy_quyen = excluded.uy_quyen,
                ttln_tw = excluded.ttln_tw,
                ttln_tinh = excluded.ttln_tinh,
                active = excluded.active,
                updated_at = excluded.updated_at
            """,
            (
                group.ma_to,
                group.stt,
                group.ten_to,
                group.ten_tvv_day_du,
                group.xa,
                group.ma_to_truong,
                group.ten_to_truong,
                group.dia_chi,
                group.tk_to_truong,
                group.so_dien_thoai,
                group.to_hoi,
                group.tk_to_hoi_xa,
                group.to_chuc,
                group.ten_huyen,
                group.tk_huyen,
                group.ten_tinh,
                group.tk_tinh,
                group.ten_tw,
                group.tk_tw,
                group.uy_quyen,
                group.ttln_tw,
                group.ttln_tinh,
                1 if group.active else 0,
                created_at,
                updated_at,
            ),
        )

    def _ensure_commission_rate(self, database: sqlite3.Connection, ma_to: str) -> None:
        if not ma_to.strip():
            raise CreditGroupRepositoryError("Mã tổ không được để trống.")
        existing_group = database.execute(
            "SELECT ma_to FROM credit_groups WHERE ma_to = ?",
            (ma_to,),
        ).fetchone()
        if existing_group is None:
            raise CreditGroupRepositoryError(
                f"Không tìm thấy tổ vay vốn có mã {ma_to}."
            )
        existing_rate = database.execute(
            "SELECT ma_to FROM credit_group_commission_rates WHERE ma_to = ?",
            (ma_to,),
        ).fetchone()
        if existing_rate is not None:
            return
        now = now_text()
        rate = CreditGroupCommissionRate.default_for_group(ma_to)
        database.execute(
            """
            INSERT INTO credit_group_commission_rates(
                ma_to, no_secured_to_truong, no_secured_cap_xa,
                no_secured_cap_huyen, no_secured_cap_tinh, no_secured_cap_tw,
                secured_to_truong, secured_cap_xa, secured_cap_huyen,
                secured_cap_tinh, secured_cap_tw, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                ma_to,
                rate.no_secured_to_truong,
                rate.no_secured_cap_xa,
                rate.no_secured_cap_huyen,
                rate.no_secured_cap_tinh,
                rate.no_secured_cap_tw,
                rate.secured_to_truong,
                rate.secured_cap_xa,
                rate.secured_cap_huyen,
                rate.secured_cap_tinh,
                rate.secured_cap_tw,
                now,
                now,
            ),
        )

    @staticmethod
    def _save_commission_rule_settings(
        database: sqlite3.Connection,
        settings: CreditCommissionRuleSettings,
    ) -> None:
        database.execute(
            """
            INSERT INTO credit_commission_rules(
                id, interest_min_1, interest_max_1, interest_pay_1,
                interest_min_2, interest_max_2, interest_pay_2,
                interest_min_3, interest_pay_3, bad_debt_threshold,
                bad_debt_pay, updated_at
            )
            VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                interest_min_1 = excluded.interest_min_1,
                interest_max_1 = excluded.interest_max_1,
                interest_pay_1 = excluded.interest_pay_1,
                interest_min_2 = excluded.interest_min_2,
                interest_max_2 = excluded.interest_max_2,
                interest_pay_2 = excluded.interest_pay_2,
                interest_min_3 = excluded.interest_min_3,
                interest_pay_3 = excluded.interest_pay_3,
                bad_debt_threshold = excluded.bad_debt_threshold,
                bad_debt_pay = excluded.bad_debt_pay,
                updated_at = excluded.updated_at
            """,
            (
                settings.interest_min_1,
                settings.interest_max_1,
                settings.interest_pay_1,
                settings.interest_min_2,
                settings.interest_max_2,
                settings.interest_pay_2,
                settings.interest_min_3,
                settings.interest_pay_3,
                settings.bad_debt_threshold,
                settings.bad_debt_pay,
                settings.updated_at,
            ),
        )

    @staticmethod
    def _cell_by_header(
        row: tuple[object, ...],
        header_map: dict[str, int],
        header: str,
    ) -> object:
        index = header_map.get(header)
        if index is None or index >= len(row):
            return None
        return row[index]

    @classmethod
    def _float_by_header(
        cls,
        row: tuple[object, ...],
        header_map: dict[str, int],
        header: str,
    ) -> float | None:
        value = cls._cell_by_header(row, header_map, header)
        if value is None or str(value).strip() == "":
            return None
        try:
            return float(str(value).strip().replace(",", "."))
        except ValueError as exc:
            raise CreditGroupRepositoryError(
                f"Giá trị không hợp lệ ở cột {header}: {value}"
            ) from exc

    @classmethod
    def _commission_rate_from_import_row(
        cls,
        row: tuple[object, ...],
        header_map: dict[str, int],
        ma_to: str,
    ) -> CreditGroupCommissionRate | None:
        if not all(header in header_map for header in COMMISSION_EXPORT_HEADERS[:5]):
            return None
        values = [
            cls._float_by_header(row, header_map, header)
            for header in COMMISSION_EXPORT_HEADERS[:5]
        ] + [
            cls._float_by_header(row, header_map, header)
            for header in COMMISSION_EXPORT_HEADERS[6:11]
        ]
        if all(value is None for value in values):
            return None
        default = CreditGroupCommissionRate.default_for_group(ma_to)
        return CreditGroupCommissionRate(
            ma_to=ma_to,
            no_secured_to_truong=values[0] if values[0] is not None else default.no_secured_to_truong,
            no_secured_cap_xa=values[1] if values[1] is not None else default.no_secured_cap_xa,
            no_secured_cap_huyen=values[2] if values[2] is not None else default.no_secured_cap_huyen,
            no_secured_cap_tinh=values[3] if values[3] is not None else default.no_secured_cap_tinh,
            no_secured_cap_tw=values[4] if values[4] is not None else default.no_secured_cap_tw,
            secured_to_truong=values[5] if values[5] is not None else default.secured_to_truong,
            secured_cap_xa=values[6] if values[6] is not None else default.secured_cap_xa,
            secured_cap_huyen=values[7] if values[7] is not None else default.secured_cap_huyen,
            secured_cap_tinh=values[8] if values[8] is not None else default.secured_cap_tinh,
            secured_cap_tw=values[9] if values[9] is not None else default.secured_cap_tw,
        )

    @classmethod
    def _commission_rules_from_import_row(
        cls,
        row: tuple[object, ...],
        header_map: dict[str, int],
    ) -> CreditCommissionRuleSettings | None:
        if not all(header in header_map for header in COMMISSION_RULE_EXPORT_HEADERS):
            return None
        values = [
            cls._float_by_header(row, header_map, header)
            for header in COMMISSION_RULE_EXPORT_HEADERS
        ]
        if all(value is None for value in values):
            return None
        default = CreditCommissionRuleSettings()
        return CreditCommissionRuleSettings(
            interest_min_1=values[0] if values[0] is not None else default.interest_min_1,
            interest_max_1=values[1] if values[1] is not None else default.interest_max_1,
            interest_pay_1=values[2] if values[2] is not None else default.interest_pay_1,
            interest_min_2=values[3] if values[3] is not None else default.interest_min_2,
            interest_max_2=values[4] if values[4] is not None else default.interest_max_2,
            interest_pay_2=values[5] if values[5] is not None else default.interest_pay_2,
            interest_min_3=values[6] if values[6] is not None else default.interest_min_3,
            interest_pay_3=values[7] if values[7] is not None else default.interest_pay_3,
            bad_debt_threshold=values[8] if values[8] is not None else default.bad_debt_threshold,
            bad_debt_pay=values[9] if values[9] is not None else default.bad_debt_pay,
        )

    @staticmethod
    def _group_from_row(row: sqlite3.Row) -> CreditGroup:
        return CreditGroup(
            stt=int(row["stt"] or 0),
            ma_to=str(row["ma_to"] or ""),
            ten_to=str(row["ten_to"] or ""),
            ten_tvv_day_du=str(row["ten_tvv_day_du"] or ""),
            xa=str(row["xa"] or ""),
            ma_to_truong=str(row["ma_to_truong"] or ""),
            ten_to_truong=str(row["ten_to_truong"] or ""),
            dia_chi=str(row["dia_chi"] or ""),
            tk_to_truong=str(row["tk_to_truong"] or ""),
            so_dien_thoai=str(row["so_dien_thoai"] or ""),
            to_hoi=str(row["to_hoi"] or ""),
            tk_to_hoi_xa=str(row["tk_to_hoi_xa"] or ""),
            to_chuc=str(row["to_chuc"] or ""),
            ten_huyen=str(row["ten_huyen"] or ""),
            tk_huyen=str(row["tk_huyen"] or ""),
            ten_tinh=str(row["ten_tinh"] or ""),
            tk_tinh=str(row["tk_tinh"] or ""),
            ten_tw=str(row["ten_tw"] or ""),
            tk_tw=str(row["tk_tw"] or ""),
            uy_quyen=str(row["uy_quyen"] or ""),
            ttln_tw=str(row["ttln_tw"] or ""),
            ttln_tinh=str(row["ttln_tinh"] or ""),
            active=bool(row["active"]),
            created_at=str(row["created_at"] or ""),
            updated_at=str(row["updated_at"] or ""),
        )

    @staticmethod
    def _commission_from_row(row: sqlite3.Row) -> CreditGroupCommissionRate:
        return CreditGroupCommissionRate(
            ma_to=str(row["ma_to"] or ""),
            no_secured_to_truong=float(row["no_secured_to_truong"] or 0),
            no_secured_cap_xa=float(row["no_secured_cap_xa"] or 0),
            no_secured_cap_huyen=float(row["no_secured_cap_huyen"] or 0),
            no_secured_cap_tinh=float(row["no_secured_cap_tinh"] or 0),
            no_secured_cap_tw=float(row["no_secured_cap_tw"] or 0),
            secured_to_truong=float(row["secured_to_truong"] or 0),
            secured_cap_xa=float(row["secured_cap_xa"] or 0),
            secured_cap_huyen=float(row["secured_cap_huyen"] or 0),
            secured_cap_tinh=float(row["secured_cap_tinh"] or 0),
            secured_cap_tw=float(row["secured_cap_tw"] or 0),
            created_at=str(row["created_at"] or ""),
            updated_at=str(row["updated_at"] or ""),
        )

    @staticmethod
    def _commission_rule_settings_from_row(
        row: sqlite3.Row,
    ) -> CreditCommissionRuleSettings:
        return CreditCommissionRuleSettings(
            interest_min_1=float(row["interest_min_1"] or 0),
            interest_max_1=float(row["interest_max_1"] or 0),
            interest_pay_1=float(row["interest_pay_1"] or 0),
            interest_min_2=float(row["interest_min_2"] or 0),
            interest_max_2=float(row["interest_max_2"] or 0),
            interest_pay_2=float(row["interest_pay_2"] or 0),
            interest_min_3=float(row["interest_min_3"] or 0),
            interest_pay_3=float(row["interest_pay_3"] or 0),
            bad_debt_threshold=float(row["bad_debt_threshold"] or 0),
            bad_debt_pay=float(row["bad_debt_pay"] or 0),
            updated_at=str(row["updated_at"] or ""),
        )
