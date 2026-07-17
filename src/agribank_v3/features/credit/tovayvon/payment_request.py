from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable
import os
import platform
import re
import subprocess
import unicodedata

from openpyxl import load_workbook

from agribank_v3.features.credit.tovayvon.models import (
    CreditGroup,
    CreditGroupCommissionRate,
)
from agribank_v3.features.credit.tovayvon.repository import (
    CreditGroupRepository,
    CreditGroupRepositoryError,
)
from agribank_v3.features.credit.tovayvon.word_template import (
    replace_word_placeholders,
    scan_word_placeholders,
)
from agribank_v3.features.credit.tovayvon.interest_report import round_money
from agribank_v3.runtime_paths import application_root


PAYMENT_REQUEST_TITLE = "Đề nghị thanh toán hoa hồng tổ vay vốn"
PAYMENT_TEMPLATE_EDIT_TITLE = "Chỉnh sửa mẫu biểu Đề nghị thanh toán"
SUMMARY_SHEET_NAME = "TongHopTheoTo"

REQUIRED_PAYMENT_SUMMARY_COLUMNS: tuple[str, ...] = (
    "MaTo",
    "TenTo",
    "TenToTruong",
    "SoDong",
    "TongDuNo",
    "LaiCoBD",
    "LaiKhongBD",
    "LaiTon",
    "LaiPhaiThu",
    "TyLeThuLai",
    "TyLeNoXau",
    "TyLeChi",
    "HHCoBD",
    "HHKhongBD",
    "TongHH",
    "HH_ToTruong",
    "HH_CapXa",
    "HH_CapHuyen",
    "HH_CapTinh",
    "HH_TW",
    "CanhBao",
)

_MONEY_COLUMNS = {
    "TongDuNo",
    "LaiCoBD",
    "LaiKoBD",
    "LaiKhongBD",
    "LaiTon",
    "LaiPhaiThu",
    "HHCoBD",
    "HHKhongBD",
    "TongHH",
    "HH_ToTruong",
    "HH_CapXa",
    "HH_CapHuyen",
    "HH_CapTinh",
    "HH_TW",
}
_PERCENT_COLUMNS = {"TyLeThuLai", "TyLeNoXau", "TyLeChi"}


class PaymentRequestError(RuntimeError):
    pass


@dataclass(frozen=True, slots=True)
class PaymentSummaryRow:
    ma_to: str
    ten_to: str
    ten_to_truong: str
    so_dong: int
    tong_du_no: float
    lai_co_bd: float
    lai_khong_bd: float
    lai_ton: float
    lai_phai_thu: float
    ty_le_thu_lai: float
    ty_le_no_xau: float
    ty_le_chi: float
    hh_co_bd: float
    hh_khong_bd: float
    tong_hh: float
    hh_to_truong: float
    hh_cap_xa: float
    hh_cap_huyen: float
    hh_cap_tinh: float
    hh_tw: float
    canh_bao: str = ""
    base_no_secured_rate: float = 0.0
    base_secured_rate: float = 0.0


@dataclass(frozen=True, slots=True)
class PaymentReportData:
    rows: tuple[PaymentSummaryRow, ...]
    period_from: str = ""
    period_to: str = ""
    warnings: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class PaymentExportResult:
    output_paths: tuple[Path, ...]
    warnings: tuple[str, ...]
    placeholders: tuple[str, ...]
    logs: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class PaymentEligibilitySummary:
    total: int
    eligible: tuple[PaymentSummaryRow, ...]
    ineligible: tuple[PaymentSummaryRow, ...]
    logs: tuple[str, ...]


def default_payment_template_path(root: Path | None = None) -> Path:
    base = Path(root) if root is not None else application_root()
    return base / "DuLieuTEST" / "TOVAYVON" / "DeNghiThanhToan.docx"


def default_payment_output_folder(root: Path | None = None) -> Path:
    base = Path(root) if root is not None else application_root()
    return base / "KetQua"


def open_payment_template_for_edit(template_path: Path | None = None) -> None:
    path = Path(template_path) if template_path is not None else default_payment_template_path()
    if not path.is_file():
        raise PaymentRequestError(
            "Không tìm thấy file mẫu DeNghiThanhToan.docx. "
            "Vui lòng kiểm tra thư mục DuLieuTEST\\TOVAYVON."
        )
    _open_path(path)


def open_payment_output_folder(path: Path) -> None:
    folder = Path(path)
    folder.mkdir(parents=True, exist_ok=True)
    _open_path(folder)


def load_payment_report_data(report_path: Path) -> PaymentReportData:
    report_path = Path(report_path)
    if not report_path.is_file():
        raise PaymentRequestError(f"Không tìm thấy file bảng kê: {report_path}")

    workbook = load_workbook(report_path, data_only=True)
    try:
        if SUMMARY_SHEET_NAME not in workbook.sheetnames:
            raise PaymentRequestError(
                f"File bảng kê không có sheet {SUMMARY_SHEET_NAME}."
            )
        sheet = workbook[SUMMARY_SHEET_NAME]
        headers = _headers_from_sheet(sheet)
        missing = [
            column
            for column in REQUIRED_PAYMENT_SUMMARY_COLUMNS
            if column not in headers
        ]
        if missing:
            raise PaymentRequestError(
                f"Sheet {SUMMARY_SHEET_NAME} thiếu cột bắt buộc: "
                + ", ".join(missing)
            )
        rows: list[PaymentSummaryRow] = []
        warnings: list[str] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            data = _row_dict(headers, values)
            ma_to = str(data.get("MaTo") or "").strip()
            if not ma_to:
                continue
            rows.append(_summary_row_from_dict(data))
        if not rows:
            warnings.append(f"Sheet {SUMMARY_SHEET_NAME} không có tổ vay vốn để xuất.")
        period_from, period_to = _read_report_period(workbook)
        return PaymentReportData(
            rows=tuple(rows),
            period_from=period_from,
            period_to=period_to,
            warnings=tuple(warnings),
        )
    finally:
        workbook.close()


def build_payment_context(
    summary: PaymentSummaryRow,
    repository: CreditGroupRepository,
    *,
    period_from: str = "",
    period_to: str = "",
    today: date | None = None,
) -> tuple[dict[str, str], tuple[str, ...]]:
    group = repository.get_group(summary.ma_to)
    rate = repository.get_commission_rate(summary.ma_to)
    group_exists = group is not None
    warnings: list[str] = []
    if group is None:
        group = CreditGroup(
            ma_to=summary.ma_to,
            ten_to=summary.ten_to,
            ten_to_truong=summary.ten_to_truong,
        )
        warnings.append(
            f"Không tìm thấy MaTo {summary.ma_to} trong credit_groups; "
            "xuất theo dữ liệu tổng hợp trong bảng kê."
        )
    if rate is None:
        rate = CreditGroupCommissionRate.default_for_group(summary.ma_to)
        warnings.append(
            f"MaTo {summary.ma_to} chưa có tỷ lệ phân bổ riêng; dùng tỷ lệ mặc định VBA."
        )

    base_no_secured_rate, base_secured_rate, rate_source, rate_warnings = (
        _payment_base_rates(summary, repository, group_exists=group_exists)
    )
    warnings.extend(rate_warnings)
    split = _split_commission(summary, rate)
    current_date = today or date.today()
    period_text = (
        f"từ ngày {period_from} đến ngày {period_to}"
        if period_from and period_to
        else ""
    )
    den_ngay = period_to
    tong_lai_thu = summary.lai_co_bd + summary.lai_khong_bd

    context: dict[str, object] = {
        "KyThuLai": period_text,
        "NgayThangNam": (
            f"ngày {current_date.day} tháng {current_date.month} năm {current_date.year}"
        ),
        "DenNgay": den_ngay,
        "MaToVayVon": summary.ma_to,
        "MaTo": summary.ma_to,
        "TenToVV_Xa": group.ten_tvv_day_du or summary.ten_to or group.ten_to,
        "TenTo": summary.ten_to or group.ten_to,
        "TenToTruong": group.ten_to_truong or summary.ten_to_truong,
        "DiaChi_TT": group.dia_chi,
        "TongLaiThu": tong_lai_thu,
        "LaiTon": summary.lai_ton,
        "TyLeThuLai": summary.ty_le_thu_lai,
        "LaiCoBD": summary.lai_co_bd,
        "LaiKoBD": summary.lai_khong_bd,
        "LaiKhongBD": summary.lai_khong_bd,
        "TongHH": summary.tong_hh,
        "TongHH_BC": amount_to_vietnamese_words(summary.tong_hh),
        "HHCoBD": summary.hh_co_bd,
        "HHKoBD": summary.hh_khong_bd,
        "HHKhongBD": summary.hh_khong_bd,
        "HH_ToTr": summary.hh_to_truong,
        "HH_ToTruong": summary.hh_to_truong,
        "HH_CapXa": summary.hh_cap_xa,
        "HH_CapHuyen": summary.hh_cap_huyen,
        "HH_CapTinh": summary.hh_cap_tinh,
        "HH_TW": summary.hh_tw,
        "HH_CacCap": summary.tong_hh - summary.hh_to_truong,
        "TK_ToTr": group.tk_to_truong,
        "Ten_ToHoiXa": _to_hoi_xa_name(group),
        "TK_ToHoiXa": group.tk_to_hoi_xa,
        "Ten_ToHoiHuyen": group.ten_huyen,
        "TK_ToHoiHuyen": group.tk_huyen,
        "Ten_ToHoiTinh": group.ten_tinh,
        "TK_ToHoiTinh": group.tk_tinh,
        "Ten_ToHoiTW": group.ten_tw,
        "TK_ToHoiTW": group.tk_tw,
        "ToHoi": group.to_hoi,
        "TTLN_TW": group.ttln_tw,
        "TTLN_Tinh": group.ttln_tinh,
        "TyLeNoXau": summary.ty_le_no_xau,
        "TyLeChi": summary.ty_le_chi,
        "TL_KoBD": base_no_secured_rate / 100,
        "TL_CoBD": base_secured_rate / 100,
        "Nguon_TyLeHH": rate_source,
        "TL_ToTr_KoBD": rate.no_secured_to_truong / 100,
        "TL_Xa_KoBD": rate.no_secured_cap_xa / 100,
        "TL_Huyen_KoBD": rate.no_secured_cap_huyen / 100,
        "TL_Tinh_KoBD": rate.no_secured_cap_tinh / 100,
        "TL_TW_KoBD": rate.no_secured_cap_tw / 100,
        "TL_ToTr_CoBD": rate.secured_to_truong / 100,
        "TL_Xa_CoBD": rate.secured_cap_xa / 100,
        "TL_Huyen_CoBD": rate.secured_cap_huyen / 100,
        "TL_Tinh_CoBD": rate.secured_cap_tinh / 100,
        "TL_TW_CoBD": rate.secured_cap_tw / 100,
    }
    context.update(split)
    for key, value in list(context.items()):
        if key.startswith("HH") and not key.endswith("_BC") and _is_number(value):
            context[f"{key}_BC"] = amount_to_vietnamese_words(float(value))

    return _format_context(context), tuple(warnings)


def export_payment_requests(
    *,
    report_path: Path,
    template_path: Path,
    output_folder: Path,
    repository: CreditGroupRepository,
    ma_to: str = "",
    selected_group_codes: tuple[str, ...] = (),
    export_all: bool = False,
    today: date | None = None,
) -> PaymentExportResult:
    report_data = load_payment_report_data(report_path)
    template_path = Path(template_path)
    if not template_path.is_file():
        raise PaymentRequestError(f"Không tìm thấy file mẫu Word: {template_path}")

    selected_rows = list(report_data.rows)
    selected_codes = _selected_payment_codes(ma_to, selected_group_codes)
    if not export_all and not selected_codes:
        raise PaymentRequestError(
            "Vui lòng chọn ít nhất một tổ vay vốn hoặc tích Xuất tất cả các tổ trong sheet TongHopTheoTo."
        )
    if not export_all and selected_codes:
        selected_rows = [row for row in selected_rows if row.ma_to in selected_codes]
    if not selected_rows:
        raise PaymentRequestError("Không có tổ vay vốn phù hợp để xuất đề nghị thanh toán.")

    eligibility = analyze_payment_rows(selected_rows)
    if not eligibility.eligible:
        raise PaymentRequestError(
            "Không có tổ nào đủ điều kiện chi hoa hồng để tạo đề nghị thanh toán."
        )

    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)
    template_placeholders = scan_word_placeholders(template_path)
    warnings: list[str] = list(report_data.warnings)
    output_paths: list[Path] = []
    logs: list[str] = list(eligibility.logs)
    unmapped_all: set[str] = set()
    export_date = today or date.today()

    for summary in eligibility.eligible:
        context, context_warnings = build_payment_context(
            summary,
            repository,
            period_from=report_data.period_from,
            period_to=report_data.period_to,
            today=export_date,
        )
        warnings.extend(context_warnings)
        logs.append(
            "Tổ "
            f"{summary.ma_to}: TL_KoBD={context.get('TL_KoBD', '')}, "
            f"TL_CoBD={context.get('TL_CoBD', '')} từ "
            f"{context.get('Nguon_TyLeHH', '')}."
        )
        context_keys = {f"[{key}]" for key in context}
        for placeholder in sorted(template_placeholders - context_keys):
            warnings.append(
                f"Placeholder {placeholder} chưa có mapping; đã để trống khi xuất {summary.ma_to}."
            )
        output_path = _unique_output_path(output_folder, summary.ma_to, export_date)
        unmapped = replace_word_placeholders(template_path, output_path, context)
        unmapped_all.update(unmapped)
        output_paths.append(output_path)

    logs.append(
        "Hoàn thành: "
        f"tổng số tổ trong phạm vi xuất: {eligibility.total}; "
        f"đã tạo file: {len(output_paths)}; "
        f"bỏ qua do không đủ điều kiện chi: {len(eligibility.ineligible)}; "
        "lỗi khác: 0."
    )

    return PaymentExportResult(
        output_paths=tuple(output_paths),
        warnings=tuple(dict.fromkeys(warnings)),
        placeholders=tuple(sorted(template_placeholders)),
        logs=tuple(dict.fromkeys(logs)),
    )


def analyze_payment_rows(rows: Iterable[PaymentSummaryRow]) -> PaymentEligibilitySummary:
    eligible: list[PaymentSummaryRow] = []
    ineligible: list[PaymentSummaryRow] = []
    logs: list[str] = []
    row_list = list(rows)
    for row in row_list:
        reason = payment_ineligible_reason(row)
        if reason:
            ineligible.append(row)
            logs.append(f"Bỏ qua tổ {row.ma_to} - {row.ten_to}: {reason}.")
        else:
            eligible.append(row)
    return PaymentEligibilitySummary(
        total=len(row_list),
        eligible=tuple(eligible),
        ineligible=tuple(ineligible),
        logs=tuple(logs),
    )


def is_payment_summary_eligible(row: PaymentSummaryRow) -> bool:
    return payment_ineligible_reason(row) == ""


def payment_ineligible_reason(row: PaymentSummaryRow) -> str:
    normalized_warning = _normalize_text(row.canh_bao)
    explicit_ineligible = (
        "khong du dieu kien" in normalized_warning
        or "khong du dk" in normalized_warning
    )
    if explicit_ineligible:
        return "Không đủ điều kiện chi hoa hồng"
    if row.ty_le_chi <= 0:
        return "TyLeChi = 0%"
    if row.tong_hh <= 0:
        return "TongHH = 0"
    return ""


def amount_to_vietnamese_words(amount: float | int) -> str:
    value = round_money(amount)
    if value == 0:
        return "không đồng"
    if value < 0:
        return "âm " + amount_to_vietnamese_words(abs(value))

    units = ["", "nghìn", "triệu", "tỷ"]
    groups: list[int] = []
    while value:
        groups.append(value % 1000)
        value //= 1000

    parts: list[str] = []
    highest = len(groups) - 1
    for index in range(highest, -1, -1):
        group_value = groups[index]
        if group_value == 0:
            continue
        text = _read_three_digits(group_value, full=index < highest)
        unit = units[index % 4]
        if unit:
            text = f"{text} {unit}"
        parts.append(text)
    return " ".join(parts) + " đồng"


def _read_three_digits(value: int, *, full: bool) -> str:
    digits = [
        "không",
        "một",
        "hai",
        "ba",
        "bốn",
        "năm",
        "sáu",
        "bảy",
        "tám",
        "chín",
    ]
    hundred = value // 100
    ten = (value % 100) // 10
    unit = value % 10
    parts: list[str] = []
    if hundred:
        parts.extend([digits[hundred], "trăm"])
    elif full and (ten or unit):
        parts.extend(["không", "trăm"])
    if ten > 1:
        parts.extend([digits[ten], "mươi"])
        if unit == 1:
            parts.append("mốt")
        elif unit == 5:
            parts.append("lăm")
        elif unit:
            parts.append(digits[unit])
    elif ten == 1:
        parts.append("mười")
        if unit == 5:
            parts.append("lăm")
        elif unit:
            parts.append(digits[unit])
    elif unit:
        if parts:
            parts.append("lẻ")
        parts.append(digits[unit])
    return " ".join(parts)


def _split_commission(
    summary: PaymentSummaryRow,
    rate: CreditGroupCommissionRate,
) -> dict[str, float]:
    no_to_truong = _round_money(summary.hh_khong_bd * rate.no_secured_to_truong / 100)
    no_cap_huyen = _round_money(summary.hh_khong_bd * rate.no_secured_cap_huyen / 100)
    no_cap_tinh = _round_money(summary.hh_khong_bd * rate.no_secured_cap_tinh / 100)
    no_tw = _round_money(summary.hh_khong_bd * rate.no_secured_cap_tw / 100)
    no_cap_xa = _round_money(
        summary.hh_khong_bd - no_to_truong - no_cap_huyen - no_cap_tinh - no_tw
    )

    secured_to_truong = _round_money(summary.hh_co_bd * rate.secured_to_truong / 100)
    secured_cap_huyen = _round_money(summary.hh_co_bd * rate.secured_cap_huyen / 100)
    secured_cap_tinh = _round_money(summary.hh_co_bd * rate.secured_cap_tinh / 100)
    secured_tw = _round_money(summary.hh_co_bd * rate.secured_cap_tw / 100)
    secured_cap_xa = _round_money(
        summary.hh_co_bd
        - secured_to_truong
        - secured_cap_huyen
        - secured_cap_tinh
        - secured_tw
    )
    return {
        "HH_ToTr_KoBD": no_to_truong,
        "HH_CapXa_KoBD": no_cap_xa,
        "HH_CapHuyen_KoBD": no_cap_huyen,
        "HH_CapTinh_KoBD": no_cap_tinh,
        "HH_TW_KoBD": no_tw,
        "HH_ToTr_CoBD": secured_to_truong,
        "HH_CapXa_CoBD": secured_cap_xa,
        "HH_CapHuyen_CoBD": secured_cap_huyen,
        "HH_CapTinh_CoBD": secured_cap_tinh,
        "HH_TW_CoBD": secured_tw,
    }


def _payment_base_rates(
    summary: PaymentSummaryRow,
    repository: CreditGroupRepository,
    *,
    group_exists: bool,
) -> tuple[float, float, str, list[str]]:
    """Resolve base commission rates for Word placeholders by MaTo.

    Priority: current database configuration, optional TongHopTheoTo columns,
    then the VBA defaults 3%/2%. Distribution rates are intentionally not used
    for TL_KoBD/TL_CoBD.
    """

    rate = repository.get_commission_rate(summary.ma_to)
    if rate is not None:
        return (
            rate.base_no_secured_rate,
            rate.base_secured_rate,
            "database",
            [],
        )

    sheet_has_rates = (
        summary.base_no_secured_rate > 0 or summary.base_secured_rate > 0
    )
    if sheet_has_rates:
        default_rate = CreditGroupCommissionRate.default_for_group(summary.ma_to)
        return (
            summary.base_no_secured_rate or default_rate.base_no_secured_rate,
            summary.base_secured_rate or default_rate.base_secured_rate,
            "TongHopTheoTo",
            [
                f"Tổ {summary.ma_to} chưa có cấu hình tỷ lệ hoa hồng trong database; "
                "đã dùng tỷ lệ từ sheet TongHopTheoTo."
            ],
        )

    default_rate = CreditGroupCommissionRate.default_for_group(summary.ma_to)
    if group_exists:
        try:
            default_rate = repository.get_or_create_commission_rate(summary.ma_to)
        except CreditGroupRepositoryError:
            default_rate = CreditGroupCommissionRate.default_for_group(summary.ma_to)
    return (
        default_rate.base_no_secured_rate,
        default_rate.base_secured_rate,
        "mặc định",
        [
            f"Tổ {summary.ma_to} chưa có cấu hình tỷ lệ hoa hồng riêng, "
            "đã dùng mặc định 3%/2%."
        ],
    )


def _format_context(context: dict[str, object]) -> dict[str, str]:
    formatted: dict[str, str] = {}
    for key, value in context.items():
        if value is None:
            formatted[key] = ""
        elif key.startswith("TL_") or key in _PERCENT_COLUMNS:
            formatted[key] = _format_percent(value)
        elif key.endswith("_BC"):
            formatted[key] = str(value)
        elif _is_money_key(key):
            formatted[key] = _format_money(value)
        else:
            formatted[key] = str(value)
    return formatted


def _is_money_key(key: str) -> bool:
    return (
        key in _MONEY_COLUMNS
        or key.startswith("HH_")
        or key in {"TongLaiThu", "HHKoBD", "HHKhongBD"}
    )


def _format_money(value: object) -> str:
    if not _is_number(value):
        return str(value or "")
    number = round_money(value)
    return f"{number:,}".replace(",", ".")


def _format_percent(value: object) -> str:
    if not _is_number(value):
        return str(value or "")
    percent = float(value) * 100
    text = f"{percent:.2f}".rstrip("0").rstrip(".").replace(".", ",")
    return text + "%"


def _summary_row_from_dict(data: dict[str, object]) -> PaymentSummaryRow:
    return PaymentSummaryRow(
        ma_to=str(data.get("MaTo") or "").strip(),
        ten_to=str(data.get("TenTo") or "").strip(),
        ten_to_truong=str(data.get("TenToTruong") or "").strip(),
        so_dong=int(_parse_number(data.get("SoDong"))),
        tong_du_no=_parse_number(data.get("TongDuNo")),
        lai_co_bd=_parse_number(data.get("LaiCoBD")),
        lai_khong_bd=_parse_number(data.get("LaiKhongBD")),
        lai_ton=_parse_number(data.get("LaiTon")),
        lai_phai_thu=_parse_number(data.get("LaiPhaiThu")),
        ty_le_thu_lai=_parse_percent(data.get("TyLeThuLai")),
        ty_le_no_xau=_parse_percent(data.get("TyLeNoXau")),
        ty_le_chi=_parse_percent(data.get("TyLeChi")),
        hh_co_bd=_parse_number(data.get("HHCoBD")),
        hh_khong_bd=_parse_number(data.get("HHKhongBD")),
        tong_hh=_parse_number(data.get("TongHH")),
        hh_to_truong=_parse_number(data.get("HH_ToTruong")),
        hh_cap_xa=_parse_number(data.get("HH_CapXa")),
        hh_cap_huyen=_parse_number(data.get("HH_CapHuyen")),
        hh_cap_tinh=_parse_number(data.get("HH_CapTinh")),
        hh_tw=_parse_number(data.get("HH_TW")),
        canh_bao=str(data.get("CanhBao") or "").strip(),
        base_no_secured_rate=_parse_number(
            data.get("base_no_secured_rate")
            or data.get("HH_TyLeChung_KhongTSBD")
        ),
        base_secured_rate=_parse_number(
            data.get("base_secured_rate")
            or data.get("HH_TyLeChung_CoTSBD")
        ),
    )


def _parse_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return 0.0
    text = text.replace("%", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_percent(value: object) -> float:
    if isinstance(value, str) and "%" in value:
        return _parse_number(value) / 100
    number = _parse_number(value)
    return number / 100 if number > 1 else number


def _headers_from_sheet(sheet) -> dict[str, int]:
    values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return {
        str(value or "").strip(): index
        for index, value in enumerate(values)
        if str(value or "").strip()
    }


def _row_dict(headers: dict[str, int], values: tuple[object, ...]) -> dict[str, object]:
    return {
        header: values[index] if index < len(values) else None
        for header, index in headers.items()
    }


def _read_report_period(workbook) -> tuple[str, str]:
    pattern = re.compile(
        r"Từ ngày\s+(\d{1,2}/\d{1,2}/\d{4})\s+đến ngày\s+(\d{1,2}/\d{1,2}/\d{4})",
        re.IGNORECASE,
    )
    for sheet_name in workbook.sheetnames:
        if sheet_name in {SUMMARY_SHEET_NAME, "CanhBao"}:
            continue
        value = str(workbook[sheet_name]["A6"].value or "")
        match = pattern.search(value)
        if match:
            return _normalize_date_text(match.group(1)), _normalize_date_text(match.group(2))
    return "", ""


def _normalize_date_text(value: str) -> str:
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(value.strip(), fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return value.strip()


def _unique_output_path(output_folder: Path, ma_to: str, export_date: date) -> Path:
    safe_ma_to = _safe_filename(ma_to or "ToVayVon")
    base = output_folder / f"{safe_ma_to}_{export_date:%Y%m%d}.docx"
    if not base.exists():
        return base
    index = 1
    while True:
        candidate = output_folder / f"{safe_ma_to}_{export_date:%Y%m%d}_{index}.docx"
        if not candidate.exists():
            return candidate
        index += 1


def _selected_payment_codes(
    ma_to: str,
    selected_group_codes: tuple[str, ...],
) -> set[str]:
    if selected_group_codes:
        return {code for code in selected_group_codes if code}
    return {ma_to} if ma_to else set()


def _safe_filename(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = "".join(char for char in normalized if ord(char) < 128)
    text = re.sub(r"[^A-Za-z0-9_-]+", "_", ascii_text or value)
    return text.strip("_") or "ToVayVon"


def _normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value.casefold())
    text = "".join(char for char in normalized if not unicodedata.combining(char))
    return text.replace("đ", "d")


def _to_hoi_xa_name(group: CreditGroup) -> str:
    if group.to_hoi and group.xa:
        return f"{group.to_hoi} xã {group.xa}"
    return group.to_hoi or group.xa


def _round_money(value: float | int | Decimal) -> int:
    return round_money(value)


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def _open_path(path: Path) -> None:
    system = platform.system().lower()
    if system == "windows" and hasattr(os, "startfile"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif system == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])
