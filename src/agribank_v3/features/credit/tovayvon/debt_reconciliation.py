from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable
import csv
import io
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import xlrd

from agribank_v3.features.credit.tovayvon.models import CreditGroup
from agribank_v3.features.credit.tovayvon.repository import CreditGroupRepository


DEBT_RECONCILIATION_TITLE = "Đối chiếu dư nợ theo tổ vay vốn"
SUMMARY_SHEET_NAME = "TongHopTheoTo"
DETAIL_SHEET_NAME = "ChiTietDuNo"
MISSING_GROUP_SHEET_NAME = "ThieuMaTo"
UNKNOWN_GROUP_SHEET_NAME = "MaToKhongTonTai"
GROUP_WITHOUT_BALANCE_SHEET_NAME = "ToKhongCoDuNo"
WARNING_SHEET_NAME = "CanhBao"

DEBT_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "customer_code": ("MaKH", "Mã KH", "CUSTSEQ", "Customer Code"),
    "customer_name": ("TenKH", "Tên KH", "CUSTNM", "Customer Name"),
    "loan_number": ("SoGiaiNgan", "Số giải ngân", "APPRSEQ", "Loan Number", "Số HĐ"),
    "group_code": ("MaTo", "Mã tổ", "MaToVayVon", "GRPNO", "Group No"),
    "outstanding_balance": ("DuNo", "Dư nợ", "DU_NO", "OUTSTANDING_BALANCE", "Principal Balance"),
    "debt_group": ("NhomNo", "Nhóm nợ", "Debt Group"),
    "interest_amount": ("INTEREST_AMOUNT", "Lãi", "Lãi phải thu", "Lãi tồn"),
    "disbursement_date": ("Ngày giải ngân", "DISBURSEMENT_DATE", "Disbursement Date"),
    "maturity_date": ("Ngày đến hạn", "Maturity Date"),
}

REQUIRED_DEBT_FIELDS: tuple[str, ...] = (
    "customer_code",
    "loan_number",
    "group_code",
    "outstanding_balance",
)
OPTIONAL_DEBT_FIELDS: tuple[str, ...] = (
    "customer_name",
    "debt_group",
    "interest_amount",
    "disbursement_date",
    "maturity_date",
)


class DebtReconciliationError(RuntimeError):
    pass


@dataclass(frozen=True, slots=True)
class DebtColumnDetection:
    sheet_name: str
    headers: tuple[str, ...]
    field_to_header: dict[str, str]
    missing_required: tuple[str, ...]
    missing_optional: tuple[str, ...]
    row_count: int = 0
    rows_with_group: int = 0
    rows_missing_group: int = 0
    group_count: int = 0
    known_group_count: int = 0
    unknown_group_count: int = 0

    @property
    def can_reconcile(self) -> bool:
        return not self.missing_required


@dataclass(frozen=True, slots=True)
class DebtReconciliationRequest:
    input_file: Path
    output_path: Path
    reconciliation_date: date
    selected_group_codes: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class DebtRow:
    row_number: int
    ma_to: str
    ma_kh: str
    ten_kh: str
    so_giai_ngan: str
    du_no: float
    nhom_no: str = ""
    interest_amount: float = 0.0
    disbursement_date: Any = ""
    maturity_date: Any = ""
    note: str = ""


@dataclass(frozen=True, slots=True)
class DebtGroupSummary:
    group: CreditGroup
    rows: tuple[DebtRow, ...]

    @property
    def customer_count(self) -> int:
        return len({row.ma_kh for row in self.rows if row.ma_kh})

    @property
    def loan_count(self) -> int:
        return len({(row.ma_kh, row.so_giai_ngan) for row in self.rows})

    @property
    def total_outstanding(self) -> float:
        return sum(row.du_no for row in self.rows)

    @property
    def total_interest(self) -> float:
        return sum(row.interest_amount for row in self.rows)

    def outstanding_by_group(self, debt_group: int) -> float:
        return sum(row.du_no for row in self.rows if normalize_debt_group(row.nhom_no) == debt_group)

    @property
    def bad_debt(self) -> float:
        return sum(row.du_no for row in self.rows if normalize_debt_group(row.nhom_no) in {3, 4, 5})

    @property
    def bad_debt_ratio(self) -> float:
        return self.bad_debt / self.total_outstanding if self.total_outstanding else 0.0


@dataclass(frozen=True, slots=True)
class DebtReconciliationResult:
    output_path: Path
    detail_count: int
    group_count: int
    warning_count: int
    warnings: tuple[str, ...]


def default_debt_reconciliation_output_path(
    output_folder: Path,
    reconciliation_date: date | None = None,
) -> Path:
    value = reconciliation_date or date.today()
    return Path(output_folder) / f"DoiChieuDuNoToVayVon_{value:%Y%m%d}.xlsx"


def detect_debt_columns(
    path: Path,
    repository: CreditGroupRepository | None = None,
) -> DebtColumnDetection:
    table = _read_best_table(path)
    field_to_header = _detect_field_headers(table.headers)
    missing_required = tuple(
        field for field in REQUIRED_DEBT_FIELDS if field not in field_to_header
    )
    missing_optional = tuple(
        field for field in OPTIONAL_DEBT_FIELDS if field not in field_to_header
    )

    rows_with_group = 0
    rows_missing_group = 0
    groups: set[str] = set()
    if "group_code" in field_to_header:
        group_header = field_to_header["group_code"]
        for row in table.rows:
            ma_to = normalize_text_code(row.get(group_header))
            if ma_to:
                rows_with_group += 1
                groups.add(ma_to)
            else:
                rows_missing_group += 1

    known_group_count = 0
    unknown_group_count = 0
    if repository is not None and groups:
        known = {group.ma_to for group in repository.list_groups()}
        known_group_count = len(groups & known)
        unknown_group_count = len(groups - known)

    return DebtColumnDetection(
        sheet_name=table.sheet_name,
        headers=table.headers,
        field_to_header=field_to_header,
        missing_required=missing_required,
        missing_optional=missing_optional,
        row_count=len(table.rows),
        rows_with_group=rows_with_group,
        rows_missing_group=rows_missing_group,
        group_count=len(groups),
        known_group_count=known_group_count,
        unknown_group_count=unknown_group_count,
    )


def create_debt_reconciliation(
    request: DebtReconciliationRequest,
    repository: CreditGroupRepository,
) -> DebtReconciliationResult:
    detection = detect_debt_columns(request.input_file, repository)
    if detection.missing_required:
        raise DebtReconciliationError(
            "File sao kê thiếu cột bắt buộc: " + ", ".join(detection.missing_required)
        )

    rows = _read_debt_rows(request.input_file, detection)
    selected_codes = {code for code in request.selected_group_codes if code}
    if selected_codes:
        rows = [row for row in rows if row.ma_to in selected_codes]
    vba_style_rows = _aggregate_vba_contract_rows(
        row for row in rows if _is_assigned_group(row.ma_to)
    )

    groups_by_code = {group.ma_to: group for group in repository.list_groups()}
    managed_codes = set(groups_by_code)
    if selected_codes:
        managed_codes = managed_codes & selected_codes

    missing_group_rows = [row for row in rows if not row.ma_to]
    unknown_group_rows = [
        row for row in rows
        if row.ma_to and row.ma_to not in groups_by_code and (not selected_codes or row.ma_to in selected_codes)
    ]
    valid_rows = [
        row for row in rows
        if row.ma_to and row.ma_to in groups_by_code and (not selected_codes or row.ma_to in selected_codes)
    ]

    summaries: list[DebtGroupSummary] = []
    for ma_to in sorted({row.ma_to for row in valid_rows}):
        group_rows = tuple(row for row in valid_rows if row.ma_to == ma_to)
        summaries.append(DebtGroupSummary(groups_by_code[ma_to], group_rows))

    groups_with_balance = {summary.group.ma_to for summary in summaries}
    groups_without_balance = [
        groups_by_code[ma_to]
        for ma_to in sorted(managed_codes)
        if ma_to not in groups_with_balance
    ]

    warnings = _build_warnings(
        missing_group_rows,
        unknown_group_rows,
        groups_without_balance,
        rows,
    )
    _export_debt_reconciliation(
        request,
        vba_style_rows,
        summaries,
        valid_rows,
        missing_group_rows,
        unknown_group_rows,
        groups_without_balance,
        warnings,
        detection,
    )
    return DebtReconciliationResult(
        output_path=request.output_path,
        detail_count=len(vba_style_rows),
        group_count=len({row.ma_to for row in vba_style_rows}),
        warning_count=len(warnings),
        warnings=tuple(warnings),
    )


def normalize_debt_group(value: object) -> int:
    text = normalize_text_code(value)
    digits = "".join(char for char in text if char.isdigit())
    if not digits:
        return 0
    try:
        return int(digits)
    except ValueError:
        return 0


def normalize_text_code(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.startswith("'"):
        text = text[1:]
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.strip()


def clean_debt_reconciliation_sheet_name(value: str) -> str:
    name = str(value or "")
    for old in ("\t", "\n", "\r", "\u00a0"):
        name = name.replace(old, " ")
    for char in "\\/?*:[]":
        name = name.replace(char, "")

    quote_chars = {"'", "\u2018", "\u2019", "`", '"', "\u201c", "\u201d"}
    changed = True
    while changed and name:
        changed = False
        name = name.strip()
        if name and name[0] in quote_chars:
            name = name[1:]
            changed = True
        name = name.strip()
        if name and name[-1] in quote_chars:
            name = name[:-1]
            changed = True

    if not name.strip():
        name = "Group"
    if len(name) > 31:
        name = name[:31]

    changed = True
    while changed and name:
        changed = False
        name = name.strip()
        if name and name[-1] in quote_chars:
            name = name[:-1]
            changed = True

    if not name.strip():
        name = "Group"
    if name.casefold() == "history":
        name = "History_"
    return name


def format_debt_reconciliation_date(value: date | str) -> str:
    if isinstance(value, date):
        day = f"{value.day:02d}"
        month = f"0{value.month}" if value.month < 3 else str(value.month)
        return f"đến ngày {day} tháng {month} năm {value.year}"

    raw_date = str(value or "").strip()
    clean_date = raw_date
    lowered = clean_date.casefold()
    if lowered.startswith("den ngay"):
        clean_date = clean_date[8:].strip()
    elif lowered.startswith("đến ngày"):
        clean_date = clean_date[8:].strip()
    clean_date = clean_date.replace("-", "/").replace(".", "/")
    parts = clean_date.split("/")
    if len(parts) == 3 and all(part.strip().isdigit() for part in parts):
        day, month, year = (int(part.strip()) for part in parts)
        if year < 100:
            year += 2000
        day_text = f"{day:02d}"
        month_text = f"0{month}" if month < 3 else str(month)
        return f"đến ngày {day_text} tháng {month_text} năm {year}"
    if "đến ngày" not in raw_date.casefold() and "den ngay" not in raw_date.casefold():
        return f"đến ngày {raw_date}"
    return raw_date


@dataclass(frozen=True, slots=True)
class _TableData:
    sheet_name: str
    headers: tuple[str, ...]
    rows: tuple[dict[str, object], ...]


def _read_debt_rows(path: Path, detection: DebtColumnDetection) -> list[DebtRow]:
    table = _read_best_table(path)
    header = detection.field_to_header
    rows: list[DebtRow] = []
    for index, raw in enumerate(table.rows, start=2):
        ma_to = normalize_text_code(raw.get(header["group_code"]))
        ma_kh = normalize_text_code(raw.get(header["customer_code"]))
        so_giai_ngan = normalize_text_code(raw.get(header["loan_number"]))
        du_no = parse_money(raw.get(header["outstanding_balance"]))
        notes: list[str] = []
        if not ma_to:
            notes.append("Thiếu mã tổ")
        if not ma_kh:
            notes.append("Thiếu mã khách hàng")
        if not so_giai_ngan:
            notes.append("Thiếu số giải ngân")
        if du_no < 0:
            notes.append("Dư nợ âm")
        row = DebtRow(
            row_number=index,
            ma_to=ma_to,
            ma_kh=ma_kh,
            ten_kh=str(raw.get(header.get("customer_name", ""), "") or "").strip(),
            so_giai_ngan=so_giai_ngan,
            du_no=du_no,
            nhom_no=normalize_text_code(raw.get(header.get("debt_group", ""))),
            interest_amount=parse_money(raw.get(header.get("interest_amount", ""))),
            disbursement_date=raw.get(header.get("disbursement_date", ""), ""),
            maturity_date=raw.get(header.get("maturity_date", ""), ""),
            note="; ".join(notes),
        )
        rows.append(row)
    return rows


def _is_assigned_group(value: str) -> bool:
    return bool(value) and value.casefold() != "chua_phan_to"


def _aggregate_vba_contract_rows(rows: Iterable[DebtRow]) -> list[DebtRow]:
    grouped: dict[tuple[str, str, str], DebtRow] = {}
    order: list[tuple[str, str, str]] = []
    for row in rows:
        key = (row.ma_to, row.ma_kh, row.so_giai_ngan)
        existing = grouped.get(key)
        if existing is None:
            grouped[key] = row
            order.append(key)
            continue
        grouped[key] = DebtRow(
            row_number=existing.row_number,
            ma_to=existing.ma_to,
            ma_kh=existing.ma_kh,
            ten_kh=existing.ten_kh,
            so_giai_ngan=existing.so_giai_ngan,
            du_no=existing.du_no + row.du_no,
            nhom_no=existing.nhom_no,
            interest_amount=existing.interest_amount + row.interest_amount,
            disbursement_date=existing.disbursement_date,
            maturity_date=existing.maturity_date,
            note=existing.note,
        )
    return [grouped[key] for key in order]


def _read_best_table(path: Path) -> _TableData:
    path = Path(path)
    if not path.is_file():
        raise DebtReconciliationError(f"Không tìm thấy file sao kê: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_table(path)
    if suffix == ".xls":
        return _read_xls_table(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        candidates = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = tuple(str(value or "").strip() for value in rows[0])
            data_rows = _records_from_rows(headers, rows[1:])
            score = len(_detect_field_headers(headers))
            candidates.append((score, sheet.title, headers, data_rows))
        if not candidates:
            raise DebtReconciliationError("File sao kê không có sheet dữ liệu.")
        _, sheet_name, headers, rows = max(candidates, key=lambda item: item[0])
        return _TableData(sheet_name, headers, rows)
    finally:
        workbook.close()


def _read_csv_table(path: Path) -> _TableData:
    encodings = ("utf-8-sig", "cp1258", "cp1252")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with open(path, newline="", encoding=encoding) as handle:
                sample = handle.read(4096)
                handle.seek(0)
                delimiter = "," if sample.count(",") >= sample.count(";") else ";"
                reader = csv.reader(handle, delimiter=delimiter)
                rows = list(reader)
            if not rows:
                raise DebtReconciliationError("File CSV không có dữ liệu.")
            headers = tuple(str(value or "").strip() for value in rows[0])
            return _TableData(path.name, headers, _records_from_rows(headers, rows[1:]))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise DebtReconciliationError(f"Không đọc được file CSV: {last_error}")


def _read_xls_table(path: Path) -> _TableData:
    workbook = xlrd.open_workbook(path, logfile=io.StringIO())
    candidates = []
    for sheet in workbook.sheets():
        if sheet.nrows == 0:
            continue
        headers = tuple(str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols))
        row_values = [
            tuple(sheet.cell_value(row, col) for col in range(sheet.ncols))
            for row in range(1, sheet.nrows)
        ]
        score = len(_detect_field_headers(headers))
        candidates.append((score, sheet.name, headers, _records_from_rows(headers, row_values)))
    if not candidates:
        raise DebtReconciliationError("File XLS không có dữ liệu.")
    _, sheet_name, headers, rows = max(candidates, key=lambda item: item[0])
    return _TableData(sheet_name, headers, rows)


def _records_from_rows(headers: Iterable[str], rows: Iterable[Iterable[object]]) -> tuple[dict[str, object], ...]:
    header_list = list(headers)
    records: list[dict[str, object]] = []
    for row in rows:
        values = list(row)
        if not any(value not in (None, "") for value in values):
            continue
        records.append({
            header: values[index] if index < len(values) else None
            for index, header in enumerate(header_list)
            if header
        })
    return tuple(records)


def _detect_field_headers(headers: Iterable[str]) -> dict[str, str]:
    normalized = {_normalize_header(header): header for header in headers}
    found: dict[str, str] = {}
    for field, aliases in DEBT_COLUMN_ALIASES.items():
        for alias in aliases:
            header = normalized.get(_normalize_header(alias))
            if header:
                found[field] = header
                break
    return found


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().casefold()
    decomposed = unicodedata.normalize("NFD", text)
    text = "".join(char for char in decomposed if unicodedata.category(char) != "Mn")
    return "".join(char for char in text if char.isalnum())


def parse_money(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return 0.0
    text = text.replace("'", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        text = text.replace(".", "")
    elif text.count(",") > 1:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _build_warnings(
    missing_group_rows: list[DebtRow],
    unknown_group_rows: list[DebtRow],
    groups_without_balance: list[CreditGroup],
    rows: list[DebtRow],
) -> list[str]:
    warnings: list[str] = []
    if missing_group_rows:
        warnings.append(f"Có {len(missing_group_rows)} dòng thiếu mã tổ vay vốn.")
    if unknown_group_rows:
        codes = sorted({row.ma_to for row in unknown_group_rows})
        warnings.append(
            f"Có {len(codes)} MaTo trong sao kê chưa có trong credit_groups: "
            + ", ".join(codes)
        )
    if groups_without_balance:
        warnings.append(f"Có {len(groups_without_balance)} tổ đang quản lý nhưng không có dư nợ.")
    missing_customer = sum(1 for row in rows if not row.ma_kh)
    missing_loan = sum(1 for row in rows if not row.so_giai_ngan)
    negative_balance = sum(1 for row in rows if row.du_no < 0)
    if missing_customer:
        warnings.append(f"Có {missing_customer} dòng thiếu mã khách hàng.")
    if missing_loan:
        warnings.append(f"Có {missing_loan} dòng thiếu số giải ngân.")
    if negative_balance:
        warnings.append(f"Có {negative_balance} dòng dư nợ âm.")
    return warnings


def _export_debt_reconciliation(
    request: DebtReconciliationRequest,
    vba_style_rows: list[DebtRow],
    summaries: list[DebtGroupSummary],
    detail_rows: list[DebtRow],
    missing_group_rows: list[DebtRow],
    unknown_group_rows: list[DebtRow],
    groups_without_balance: list[CreditGroup],
    warnings: list[str],
    detection: DebtColumnDetection,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    groups_by_code = {summary.group.ma_to: summary.group for summary in summaries}
    _write_vba_style_group_sheets(workbook, vba_style_rows, request)
    _write_summary_sheet(workbook, summaries)
    _write_detail_sheet(workbook, DETAIL_SHEET_NAME, detail_rows, groups_by_code)
    _write_detail_sheet(workbook, MISSING_GROUP_SHEET_NAME, missing_group_rows)
    _write_detail_sheet(workbook, UNKNOWN_GROUP_SHEET_NAME, unknown_group_rows)
    _write_groups_without_balance_sheet(workbook, groups_without_balance)
    _write_warning_sheet(workbook, warnings, detection, request)
    request.output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(request.output_path)
    workbook.close()


def _write_vba_style_group_sheets(
    workbook: Workbook,
    rows: list[DebtRow],
    request: DebtReconciliationRequest,
) -> None:
    rows_by_group: dict[str, list[DebtRow]] = {}
    for row in rows:
        rows_by_group.setdefault(row.ma_to, []).append(row)

    used_names = {sheet.title.casefold() for sheet in workbook.worksheets}
    for group_code, group_rows in rows_by_group.items():
        sheet_name = _unique_sheet_name(clean_debt_reconciliation_sheet_name(group_code), used_names)
        used_names.add(sheet_name.casefold())
        sheet = workbook.create_sheet(sheet_name)
        _write_vba_style_group_sheet(
            sheet,
            group_code=group_code,
            rows=group_rows,
            reconciliation_date=format_debt_reconciliation_date(request.reconciliation_date),
        )


def _unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    base_name = clean_debt_reconciliation_sheet_name(base_name)
    suffix = 1
    candidate = base_name
    while candidate.casefold() in used_names:
        suffix_text = f"_{suffix}"
        if len(base_name) + len(suffix_text) > 31:
            candidate = base_name[: 31 - len(suffix_text)] + suffix_text
        else:
            candidate = base_name + suffix_text
        suffix += 1
    return candidate


def _write_vba_style_group_sheet(
    sheet: Any,
    *,
    group_code: str,
    rows: list[DebtRow],
    reconciliation_date: str,
) -> None:
    sheet.sheet_view.showGridLines = True
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Times New Roman", size=10)

    sheet["K1"] = "Mẫu số: 22/ĐCN-CN"
    sheet["K1"].font = Font(name="Times New Roman", size=10, bold=True)
    sheet["K1"].alignment = Alignment(horizontal="right")

    sheet.merge_cells("A2:C2")
    sheet["A2"] = "NGÂN HÀNG NÔNG NGHIỆP"
    sheet.merge_cells("H2:K2")
    sheet["H2"] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"

    sheet.merge_cells("A3:C3")
    sheet["A3"] = "VÀ PHÁT TRIỂN NÔNG THÔN VIỆT NAM"
    sheet.merge_cells("H3:K3")
    sheet["H3"] = "Độc lập - Tự do - Hạnh phúc"
    sheet["H3"].font = Font(name="Times New Roman", size=10, bold=True, underline="single")

    sheet.merge_cells("A4:C4")
    sheet["A4"] = "CHI NHÁNH LỘC PHÁT LÂM ĐỒNG"
    sheet["A4"].font = Font(name="Times New Roman", size=10, bold=True, underline="single")

    sheet.merge_cells("H5:K5")
    sheet["H5"] = "......, ngày ...... tháng ...... năm ......"
    sheet["H5"].font = Font(name="Times New Roman", size=10, italic=True)

    sheet.merge_cells("A7:K7")
    sheet["A7"] = "BẢNG ĐỐI CHIẾU DƯ NỢ"
    sheet["A7"].font = Font(name="Times New Roman", size=14, bold=True)

    sheet.merge_cells("A8:K8")
    sheet["A8"] = reconciliation_date
    sheet["A8"].font = Font(name="Times New Roman", size=10, italic=True)

    sheet.merge_cells("A9:K9")
    sheet["A9"] = f"Tổ vay vốn: {group_code}"
    sheet["A9"].font = Font(name="Times New Roman", size=11, bold=True)

    for address in ("A2", "A3", "A4", "H2", "H3", "H5", "A7", "A8", "A9"):
        sheet[address].alignment = Alignment(horizontal="center", vertical="center")
    for address in ("A2", "A3", "H2"):
        sheet[address].font = Font(name="Times New Roman", size=10, bold=True)

    sheet.merge_cells("A11:A12")
    sheet["A11"] = "TT"
    sheet.merge_cells("B11:B12")
    sheet["B11"] = "Họ tên khách hàng"
    sheet.merge_cells("C11:C12")
    sheet["C11"] = "Mã khách hàng"
    sheet.merge_cells("D11:D12")
    sheet["D11"] = "Số hợp đồng tín dụng"
    sheet.merge_cells("E11:F11")
    sheet["E11"] = "Số liệu tại Agribank"
    sheet["E12"] = "Gốc"
    sheet["F12"] = "Lãi đã thu"
    sheet.merge_cells("G11:H11")
    sheet["G11"] = "Số liệu tại hồ sơ lưu khách hàng"
    sheet["G12"] = "Gốc"
    sheet["H12"] = "Lãi đã thu"
    sheet.merge_cells("I11:J11")
    sheet["I11"] = "Chênh lệch"
    sheet["I12"] = "Gốc"
    sheet["J12"] = "Lãi đã thu"
    sheet.merge_cells("K11:K12")
    sheet["K11"] = "Chữ ký ngưi vay"

    header_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in sheet["A11:K12"]:
        for cell in row:
            cell.font = Font(name="Times New Roman", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border

    first_data_row = 13
    for index, row in enumerate(rows, start=1):
        excel_row = first_data_row + index - 1
        sheet.cell(excel_row, 1, index)
        sheet.cell(excel_row, 2, row.ten_kh)
        sheet.cell(excel_row, 3, "'" + row.ma_kh if row.ma_kh else "")
        sheet.cell(excel_row, 4, row.so_giai_ngan)
        sheet.cell(excel_row, 5, row.du_no)
        sheet.cell(excel_row, 6, row.interest_amount)
        for column in range(7, 12):
            sheet.cell(excel_row, column, "")
        sheet.row_dimensions[excel_row].height = 35

    last_data_row = first_data_row + len(rows) - 1
    if rows:
        for row in sheet.iter_rows(min_row=first_data_row, max_row=last_data_row, min_col=1, max_col=11):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Times New Roman", size=10)
        for row in sheet.iter_rows(min_row=first_data_row, max_row=last_data_row, min_col=5, max_col=10):
            for cell in row:
                cell.number_format = "#,##0"

    total_row = first_data_row + len(rows)
    sheet.cell(total_row, 2, "Tổng cộng")
    sheet.cell(total_row, 2).font = Font(name="Times New Roman", size=10, bold=True)
    if rows:
        sheet.cell(total_row, 5, f"=SUM(E{first_data_row}:E{last_data_row})")
        sheet.cell(total_row, 6, f"=SUM(F{first_data_row}:F{last_data_row})")
    else:
        sheet.cell(total_row, 5, 0)
        sheet.cell(total_row, 6, 0)
    for cell in sheet[total_row]:
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="double"),
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Times New Roman", size=10, bold=cell.column in {2, 5, 6})
    sheet.cell(total_row, 5).number_format = "#,##0"
    sheet.cell(total_row, 6).number_format = "#,##0"

    signature_row = total_row + 2
    sheet.merge_cells(start_row=signature_row, start_column=2, end_row=signature_row, end_column=4)
    sheet.cell(signature_row, 2, "Cán bộ đối chiếu")
    sheet.merge_cells(start_row=signature_row + 1, start_column=2, end_row=signature_row + 1, end_column=4)
    sheet.cell(signature_row + 1, 2, "(Ký, ghi rõ họ tên)")
    sheet.merge_cells(start_row=signature_row, start_column=8, end_row=signature_row, end_column=10)
    sheet.cell(signature_row, 8, "Tổ trưởng (nếu có)")
    sheet.merge_cells(start_row=signature_row + 1, start_column=8, end_row=signature_row + 1, end_column=10)
    sheet.cell(signature_row + 1, 8, "(Ký, ghi rõ họ tên)")
    for address in (
        (signature_row, 2),
        (signature_row, 8),
    ):
        cell = sheet.cell(*address)
        cell.font = Font(name="Times New Roman", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for address in (
        (signature_row + 1, 2),
        (signature_row + 1, 8),
    ):
        cell = sheet.cell(*address)
        cell.font = Font(name="Times New Roman", size=10, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 8,
        "B": 22,
        "C": 15,
        "D": 20,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _write_summary_sheet(workbook: Workbook, summaries: list[DebtGroupSummary]) -> None:
    sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    headers = (
        "STT", "MaTo", "TenTo", "TenTVV_DayDu", "Xa", "Ten_ToTruong",
        "SoKhachHang", "SoKhoanVay", "TongDuNo", "DuNoNhom1", "DuNoNhom2",
        "DuNoNhom3", "DuNoNhom4", "DuNoNhom5", "DuNoXau", "TyLeNoXau",
        "TongLai", "GhiChu",
    )
    sheet.append(headers)
    for index, summary in enumerate(sorted(summaries, key=lambda item: item.group.ma_to), start=1):
        sheet.append((
            index,
            summary.group.ma_to,
            summary.group.ten_to,
            summary.group.ten_tvv_day_du,
            summary.group.xa,
            summary.group.ten_to_truong,
            summary.customer_count,
            summary.loan_count,
            summary.total_outstanding,
            summary.outstanding_by_group(1),
            summary.outstanding_by_group(2),
            summary.outstanding_by_group(3),
            summary.outstanding_by_group(4),
            summary.outstanding_by_group(5),
            summary.bad_debt,
            summary.bad_debt_ratio,
            summary.total_interest,
            "",
        ))
    _format_sheet(sheet, money_columns={9, 10, 11, 12, 13, 14, 15, 17}, percent_columns={16})


def _write_detail_sheet(
    workbook: Workbook,
    title: str,
    rows: list[DebtRow],
    groups_by_code: dict[str, CreditGroup] | None = None,
) -> None:
    sheet = workbook.create_sheet(title)
    groups_by_code = groups_by_code or {}
    headers = (
        "STT", "MaTo", "TenTo", "MaKH", "TenKH", "SoGiaiNgan",
        "DuNo", "NhomNo", "INTEREST_AMOUNT", "NgayGiaiNgan", "NgayDenHan", "GhiChu",
    )
    sheet.append(headers)
    for index, row in enumerate(rows, start=1):
        group = groups_by_code.get(row.ma_to)
        sheet.append((
            index,
            row.ma_to,
            group.ten_to if group is not None else "",
            row.ma_kh,
            row.ten_kh,
            row.so_giai_ngan,
            row.du_no,
            row.nhom_no,
            row.interest_amount,
            row.disbursement_date,
            row.maturity_date,
            row.note,
        ))
    _format_sheet(sheet, money_columns={7, 9})


def _write_groups_without_balance_sheet(workbook: Workbook, groups: list[CreditGroup]) -> None:
    sheet = workbook.create_sheet(GROUP_WITHOUT_BALANCE_SHEET_NAME)
    sheet.append(("STT", "MaTo", "TenTo", "TenTVV_DayDu", "Xa", "Ten_ToTruong", "GhiChu"))
    for index, group in enumerate(groups, start=1):
        sheet.append((
            index,
            group.ma_to,
            group.ten_to,
            group.ten_tvv_day_du,
            group.xa,
            group.ten_to_truong,
            "Tổ đang quản lý nhưng không có dư nợ trong file sao kê.",
        ))
    _format_sheet(sheet)


def _write_warning_sheet(
    workbook: Workbook,
    warnings: list[str],
    detection: DebtColumnDetection,
    request: DebtReconciliationRequest,
) -> None:
    sheet = workbook.create_sheet(WARNING_SHEET_NAME)
    sheet.append(("Thông tin", "Giá trị"))
    sheet.append(("File", str(request.input_file)))
    sheet.append(("Ngày đối chiếu", request.reconciliation_date.strftime("%d/%m/%Y")))
    sheet.append(("Sheet", detection.sheet_name))
    sheet.append(("Số dòng dữ liệu", detection.row_count))
    sheet.append(("Số dòng có MaTo", detection.rows_with_group))
    sheet.append(("Số dòng thiếu MaTo", detection.rows_missing_group))
    sheet.append(("Số MaTo trong file", detection.group_count))
    sheet.append(("Số MaTo có trong SQLite", detection.known_group_count))
    sheet.append(("Số MaTo chưa có trong SQLite", detection.unknown_group_count))
    sheet.append(())
    sheet.append(("Cảnh báo", ""))
    for index, warning in enumerate(warnings, start=1):
        sheet.append((index, warning))
    _format_sheet(sheet)


def _format_sheet(sheet: Any, money_columns: set[int] | None = None, percent_columns: set[int] | None = None) -> None:
    money_columns = money_columns or set()
    percent_columns = percent_columns or set()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="A6A6A6"),
        right=Side(style="thin", color="A6A6A6"),
        top=Side(style="thin", color="A6A6A6"),
        bottom=Side(style="thin", color="A6A6A6"),
    )
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            if cell.column in money_columns and cell.row > 1:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
            elif cell.column in percent_columns and cell.row > 1:
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 40)
