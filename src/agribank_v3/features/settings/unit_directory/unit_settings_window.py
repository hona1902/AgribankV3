from __future__ import annotations

from dataclasses import replace
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from agribank_v3.features.settings.unit_directory import UNIT_SETTINGS_TITLE
from agribank_v3.features.settings.unit_directory.models import (
    AppUnitSettings,
    BranchDirectoryEntry,
    HEAD_OFFICE,
    OFFICE_TYPES,
    OfficeDirectoryEntry,
    OTHER_OFFICE,
    TRANSACTION_OFFICE,
)
from agribank_v3.features.settings.unit_directory.repository import (
    UnitDirectoryError,
    build_office_code,
    default_office_type,
    normalize_trctcd,
)
from agribank_v3.features.settings.unit_directory.service import (
    UnitDirectoryService,
    get_unit_directory_service,
)
from agribank_v3.features.settings.unit_directory.ui_helpers import (
    populate_branch_combo,
    populate_office_combo,
)
from agribank_v3.ui.components.controls import (
    combo_box,
    current_data,
    danger_button,
    make_compact_control,
    primary_button,
    secondary_button,
)


class UnitSettingsWindow(QDialog):
    def __init__(self, database_path: Path | str, parent=None) -> None:
        super().__init__(parent)
        self.service = get_unit_directory_service(database_path)
        self.setWindowTitle(f"Cài đặt thông tin đơn vị - AgribankV3")
        self.setWindowFlag(Qt.WindowType.WindowMinMaxButtonsHint, True)
        self.setModal(False)
        self.resize(1120, 720)
        self.setMinimumSize(900, 580)
        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        self.tabs.addTab(self._build_settings_tab(), "Đơn vị sử dụng")
        self.tabs.addTab(self._build_branch_tab(), "Danh mục chi nhánh")
        self.tabs.addTab(self._build_office_tab(), "Hội sở/Phòng giao dịch")
        layout.addWidget(self.tabs)
        self.service.add_listener(self.refresh_all)
        self.refresh_all()

    def closeEvent(self, event) -> None:
        self.service.remove_listener(self.refresh_all)
        super().closeEvent(event)

    def _build_settings_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.home_branch_combo = combo_box("Chọn chi nhánh", minimum_width=320, maximum_width=None, searchable=True)
        self.default_office_combo = combo_box("Chọn đơn vị mặc định", minimum_width=320, maximum_width=None, searchable=True)
        self.organization_name_input = QLineEdit()
        self.organization_name_input.setMinimumWidth(420)
        self.organization_name_input.setClearButtonEnabled(True)
        make_compact_control(self.organization_name_input)
        self.current_info_label = QLabel()
        self.current_info_label.setObjectName("MutedText")
        self.current_info_label.setWordWrap(True)
        self.home_branch_combo.currentIndexChanged.connect(self._refresh_default_office_combo)
        form.addRow("Chi nhánh đang sử dụng", self.home_branch_combo)
        form.addRow("Đơn vị mặc định", self.default_office_combo)
        form.addRow("Tên đơn vị hiển thị", self.organization_name_input)
        form.addRow("Thông tin hiện tại", self.current_info_label)
        actions = QHBoxLayout()
        save_button = primary_button("Lưu")
        restore_button = secondary_button("Khôi phục")
        refresh_button = secondary_button("Làm mới")
        save_button.clicked.connect(self._save_settings)
        restore_button.clicked.connect(self._load_settings_controls)
        refresh_button.clicked.connect(self.refresh_all)
        actions.addWidget(save_button)
        actions.addWidget(restore_button)
        actions.addWidget(refresh_button)
        actions.addStretch()
        layout.addLayout(form)
        layout.addLayout(actions)
        layout.addStretch()
        return page

    def _build_branch_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        self.branch_table = QTableWidget(0, 8)
        self.branch_table.setHorizontalHeaderLabels(
            ("STT", "Mã chi nhánh", "Tên chi nhánh", "Tên ngắn", "Tên hiển thị", "Tỉnh/Thành phố", "Trạng thái", "Ngày cập nhật")
        )
        self._configure_table(self.branch_table)
        self.branch_table.doubleClicked.connect(lambda _index: self._edit_selected_branch())
        actions = QHBoxLayout()
        add_button = primary_button("Thêm")
        edit_button = secondary_button("Sửa")
        inactive_button = danger_button("Ngừng sử dụng")
        active_button = secondary_button("Kích hoạt lại")
        export_button = secondary_button("Xuất Excel")
        refresh_button = secondary_button("Làm mới")
        add_button.clicked.connect(self._add_branch)
        edit_button.clicked.connect(self._edit_selected_branch)
        inactive_button.clicked.connect(lambda: self._set_selected_branch_active(False))
        active_button.clicked.connect(lambda: self._set_selected_branch_active(True))
        export_button.clicked.connect(self._export_directory)
        refresh_button.clicked.connect(self.refresh_all)
        for button in (add_button, edit_button, inactive_button, active_button, export_button, refresh_button):
            actions.addWidget(button)
        actions.addStretch()
        layout.addLayout(actions)
        layout.addWidget(self.branch_table)
        return page

    def _build_office_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        filters = QHBoxLayout()
        self.office_branch_filter = combo_box("Tất cả chi nhánh", minimum_width=220, maximum_width=320, searchable=True)
        self.office_type_filter = combo_box("Tất cả loại đơn vị", minimum_width=180, maximum_width=240)
        self.office_type_filter.addItem("Hội sở", HEAD_OFFICE)
        self.office_type_filter.addItem("Phòng giao dịch", TRANSACTION_OFFICE)
        self.office_type_filter.addItem("Khác", OTHER_OFFICE)
        self.office_status_filter = combo_box("Tất cả trạng thái", minimum_width=160, maximum_width=220)
        self.office_status_filter.addItem("Đang sử dụng", "active")
        self.office_status_filter.addItem("Ngừng sử dụng", "inactive")
        self.office_search_input = QLineEdit()
        self.office_search_input.setPlaceholderText("Tìm mã hoặc tên đơn vị")
        self.office_search_input.setClearButtonEnabled(True)
        make_compact_control(self.office_search_input)
        for widget in (self.office_branch_filter, self.office_type_filter, self.office_status_filter, self.office_search_input):
            filters.addWidget(widget)
        filters.addStretch()
        self.office_branch_filter.currentIndexChanged.connect(self._render_offices)
        self.office_type_filter.currentIndexChanged.connect(self._render_offices)
        self.office_status_filter.currentIndexChanged.connect(self._render_offices)
        self.office_search_input.textChanged.connect(self._render_offices)
        self.office_table = QTableWidget(0, 9)
        self.office_table.setHorizontalHeaderLabels(
            ("STT", "Mã chi nhánh", "Mã TRCTCD", "Mã đơn vị", "Loại đơn vị", "Tên đơn vị", "Tên ngắn", "Trạng thái", "Ngày cập nhật")
        )
        self._configure_table(self.office_table)
        self.office_table.doubleClicked.connect(lambda _index: self._edit_selected_office())
        actions = QHBoxLayout()
        add_button = primary_button("Thêm")
        edit_button = secondary_button("Sửa")
        inactive_button = danger_button("Ngừng sử dụng")
        active_button = secondary_button("Kích hoạt lại")
        export_button = secondary_button("Xuất Excel")
        refresh_button = secondary_button("Làm mới")
        add_button.clicked.connect(self._add_office)
        edit_button.clicked.connect(self._edit_selected_office)
        inactive_button.clicked.connect(lambda: self._set_selected_office_active(False))
        active_button.clicked.connect(lambda: self._set_selected_office_active(True))
        export_button.clicked.connect(self._export_directory)
        refresh_button.clicked.connect(self.refresh_all)
        for button in (add_button, edit_button, inactive_button, active_button, export_button, refresh_button):
            actions.addWidget(button)
        actions.addStretch()
        layout.addLayout(filters)
        layout.addLayout(actions)
        layout.addWidget(self.office_table)
        return page

    @staticmethod
    def _configure_table(table: QTableWidget) -> None:
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        table.setSortingEnabled(True)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.horizontalHeader().setStretchLastSection(True)

    def refresh_all(self) -> None:
        self.service.refresh()
        self._load_settings_controls()
        self._render_branches()
        self._refresh_office_filters()
        self._render_offices()

    def _load_settings_controls(self) -> None:
        settings = self.service.get_settings()
        self.home_branch_combo.blockSignals(True)
        populate_branch_combo(self.home_branch_combo, self.service, include_all=False)
        index = self.home_branch_combo.findData(settings.home_branch_code)
        self.home_branch_combo.setCurrentIndex(index if index >= 0 else 0)
        self.home_branch_combo.blockSignals(False)
        self._refresh_default_office_combo()
        office_index = self.default_office_combo.findData(settings.default_office_code)
        self.default_office_combo.setCurrentIndex(office_index if office_index >= 0 else 0)
        self.organization_name_input.setText(settings.organization_name)
        home = self.service.format_branch_display(settings.home_branch_code) if settings.home_branch_code else "Chưa chọn"
        office = self.service.get_office_by_code(settings.default_office_code)
        office_text = (
            self.service.format_office_display(office.branch_code, office.trctcd)
            if office is not None
            else "Chưa chọn"
        )
        self.current_info_label.setText(
            f"Chi nhánh: {home}\nĐơn vị mặc định: {office_text}\nCập nhật: {settings.updated_at or 'Chưa có'}"
        )

    def _refresh_default_office_combo(self) -> None:
        branch_code = current_data(self.home_branch_combo)
        populate_office_combo(self.default_office_combo, self.service, branch_code=branch_code)

    def _save_settings(self) -> None:
        try:
            self.service.save_settings(
                AppUnitSettings(
                    home_branch_code=current_data(self.home_branch_combo),
                    default_office_code=current_data(self.default_office_combo),
                    organization_name=self.organization_name_input.text(),
                )
            )
        except UnitDirectoryError as exc:
            QMessageBox.warning(self, UNIT_SETTINGS_TITLE, str(exc))
            return
        QMessageBox.information(self, UNIT_SETTINGS_TITLE, "Đã lưu thông tin đơn vị.")

    def _render_branches(self) -> None:
        rows = self.service.repository.list_branches(active_only=False)
        self.branch_table.setSortingEnabled(False)
        self.branch_table.setRowCount(len(rows))
        for row_index, branch in enumerate(rows):
            values = (
                row_index + 1,
                branch.branch_code,
                branch.branch_name,
                branch.short_name,
                self.service.format_branch_display(branch.branch_code),
                branch.province_name,
                "Đang sử dụng" if branch.is_active else "Ngừng sử dụng",
                branch.updated_at,
            )
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if column == 1:
                    item.setData(Qt.ItemDataRole.UserRole, branch.branch_code)
                if column == 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.branch_table.setItem(row_index, column, item)
        self.branch_table.setSortingEnabled(True)

    def _refresh_office_filters(self) -> None:
        current = current_data(self.office_branch_filter)
        populate_branch_combo(self.office_branch_filter, self.service)
        if current and self.office_branch_filter.findData(current) >= 0:
            self.office_branch_filter.setCurrentIndex(self.office_branch_filter.findData(current))

    def _render_offices(self) -> None:
        branch = current_data(self.office_branch_filter)
        office_type = current_data(self.office_type_filter)
        status = current_data(self.office_status_filter)
        search = self.office_search_input.text().strip().casefold()
        rows = self.service.repository.list_offices(branch_code=branch, office_type=office_type, active_only=False)
        if status == "active":
            rows = [item for item in rows if item.is_active]
        elif status == "inactive":
            rows = [item for item in rows if not item.is_active]
        if search:
            rows = [
                item for item in rows
                if search in " ".join((item.branch_code, item.trctcd, item.office_code, item.office_name, item.short_name)).casefold()
            ]
        self.office_table.setSortingEnabled(False)
        self.office_table.setRowCount(len(rows))
        for row_index, office in enumerate(rows):
            values = (
                row_index + 1,
                office.branch_code,
                office.trctcd,
                office.office_code,
                _office_type_label(office.office_type),
                office.office_name,
                office.short_name,
                "Đang sử dụng" if office.is_active else "Ngừng sử dụng",
                office.updated_at,
            )
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if column == 3:
                    item.setData(Qt.ItemDataRole.UserRole, office.office_code)
                if column == 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.office_table.setItem(row_index, column, item)
        self.office_table.setSortingEnabled(True)

    def _selected_branch_code(self) -> str:
        row = self.branch_table.currentRow()
        if row < 0:
            return ""
        item = self.branch_table.item(row, 1)
        return str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "") if item is not None else ""

    def _selected_office_code(self) -> str:
        row = self.office_table.currentRow()
        if row < 0:
            return ""
        item = self.office_table.item(row, 3)
        return str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "") if item is not None else ""

    def _add_branch(self) -> None:
        dialog = BranchDialog(self.service, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_all()

    def _edit_selected_branch(self) -> None:
        branch = self.service.get_branch(self._selected_branch_code())
        if branch is None:
            return
        dialog = BranchDialog(self.service, branch, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_all()

    def _set_selected_branch_active(self, active: bool) -> None:
        code = self._selected_branch_code()
        if not code:
            return
        try:
            self.service.set_branch_active(code, active)
        except UnitDirectoryError as exc:
            QMessageBox.warning(self, UNIT_SETTINGS_TITLE, str(exc))

    def _add_office(self) -> None:
        dialog = OfficeDialog(self.service, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_all()

    def _edit_selected_office(self) -> None:
        office = self.service.get_office_by_code(self._selected_office_code())
        if office is None:
            return
        dialog = OfficeDialog(self.service, office, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.refresh_all()

    def _set_selected_office_active(self, active: bool) -> None:
        code = self._selected_office_code()
        if not code:
            return
        try:
            self.service.set_office_active(code, active)
        except UnitDirectoryError as exc:
            QMessageBox.warning(self, UNIT_SETTINGS_TITLE, str(exc))

    def _export_directory(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Xuất danh mục đơn vị",
            "DanhMucDonVi.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return
        try:
            output = export_unit_directory_excel(self.service, Path(path))
        except OSError as exc:
            QMessageBox.warning(self, "Xuất Excel", str(exc))
            return
        QMessageBox.information(self, "Xuất Excel", f"Đã xuất: {output}")


class BranchDialog(QDialog):
    def __init__(self, service: UnitDirectoryService, branch: BranchDirectoryEntry | None = None, parent=None) -> None:
        super().__init__(parent)
        self.service = service
        self.branch = branch
        self.setWindowTitle("Chi nhánh")
        self.setMinimumWidth(560)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.code_input = QLineEdit(branch.branch_code if branch else "")
        self.name_input = QLineEdit(branch.branch_name if branch else "")
        self.short_input = QLineEdit(branch.short_name if branch else "")
        self.province_input = QLineEdit(branch.province_name if branch else "")
        self.active_check = QCheckBox("Đang sử dụng")
        self.active_check.setChecked(branch.is_active if branch else True)
        self.preview_label = QLabel()
        for editor in (self.code_input, self.name_input, self.short_input, self.province_input):
            editor.setClearButtonEnabled(True)
            make_compact_control(editor)
            editor.textChanged.connect(self._update_preview)
        if branch is not None:
            self.code_input.setReadOnly(True)
        form.addRow("Mã chi nhánh", self.code_input)
        form.addRow("Tên chi nhánh", self.name_input)
        form.addRow("Tên ngắn", self.short_input)
        form.addRow("Tỉnh/Thành phố", self.province_input)
        form.addRow("Trạng thái", self.active_check)
        form.addRow("Preview", self.preview_label)
        actions = QHBoxLayout()
        save_button = primary_button("Lưu")
        cancel_button = secondary_button("Hủy")
        save_button.clicked.connect(self._save)
        cancel_button.clicked.connect(self.reject)
        actions.addWidget(save_button)
        actions.addWidget(cancel_button)
        actions.addStretch()
        layout.addLayout(form)
        layout.addLayout(actions)
        self._update_preview()

    def _update_preview(self) -> None:
        code = self.code_input.text().strip()
        label = self.short_input.text().strip() or self.name_input.text().strip()
        self.preview_label.setText(f"{code} - {label}" if code and label else "")

    def _save(self) -> None:
        try:
            entry = BranchDirectoryEntry(
                branch_code=self.code_input.text(),
                branch_name=self.name_input.text(),
                short_name=self.short_input.text(),
                province_name=self.province_input.text(),
                is_active=self.active_check.isChecked(),
            )
            if self.branch is None:
                self.service.create_branch(entry)
            else:
                self.service.save_branch(entry)
        except UnitDirectoryError as exc:
            QMessageBox.warning(self, "Chi nhánh", str(exc))
            return
        self.accept()


class OfficeDialog(QDialog):
    def __init__(self, service: UnitDirectoryService, office: OfficeDirectoryEntry | None = None, parent=None) -> None:
        super().__init__(parent)
        self.service = service
        self.office = office
        self.setWindowTitle("Hội sở/Phòng giao dịch")
        self.setMinimumWidth(620)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.branch_combo = combo_box("Chọn chi nhánh", minimum_width=340, maximum_width=None, searchable=True)
        populate_branch_combo(self.branch_combo, service, include_all=False)
        self.trctcd_input = QLineEdit(office.trctcd if office else "")
        self.type_combo = QComboBox()
        for label, value in (
            ("Hội sở", HEAD_OFFICE),
            ("Phòng giao dịch", TRANSACTION_OFFICE),
            ("Khác", OTHER_OFFICE),
        ):
            self.type_combo.addItem(label, value)
        self.name_input = QLineEdit(office.office_name if office else "")
        self.short_input = QLineEdit(office.short_name if office else "")
        self.active_check = QCheckBox("Đang sử dụng")
        self.active_check.setChecked(office.is_active if office else True)
        self.preview_label = QLabel()
        for editor in (self.trctcd_input, self.name_input, self.short_input):
            editor.setClearButtonEnabled(True)
            make_compact_control(editor)
            editor.textChanged.connect(self._update_preview)
        self.branch_combo.currentIndexChanged.connect(self._update_preview)
        self.trctcd_input.editingFinished.connect(self._normalize_trctcd_and_type)
        if office is not None:
            branch_index = self.branch_combo.findData(office.branch_code)
            if branch_index >= 0:
                self.branch_combo.setCurrentIndex(branch_index)
            type_index = self.type_combo.findData(office.office_type)
            if type_index >= 0:
                self.type_combo.setCurrentIndex(type_index)
            self.branch_combo.setEnabled(False)
            self.trctcd_input.setReadOnly(True)
        form.addRow("Chi nhánh", self.branch_combo)
        form.addRow("Mã TRCTCD", self.trctcd_input)
        form.addRow("Loại đơn vị", self.type_combo)
        form.addRow("Tên đơn vị", self.name_input)
        form.addRow("Tên ngắn", self.short_input)
        form.addRow("Trạng thái", self.active_check)
        form.addRow("Preview", self.preview_label)
        actions = QHBoxLayout()
        save_button = primary_button("Lưu")
        cancel_button = secondary_button("Hủy")
        save_button.clicked.connect(self._save)
        cancel_button.clicked.connect(self.reject)
        actions.addWidget(save_button)
        actions.addWidget(cancel_button)
        actions.addStretch()
        layout.addLayout(form)
        layout.addLayout(actions)
        self._update_preview()

    def _normalize_trctcd_and_type(self) -> None:
        code = normalize_trctcd(self.trctcd_input.text())
        self.trctcd_input.setText(code)
        if self.office is None:
            type_index = self.type_combo.findData(default_office_type(code))
            if type_index >= 0:
                self.type_combo.setCurrentIndex(type_index)
        self._update_preview()

    def _update_preview(self) -> None:
        branch = current_data(self.branch_combo)
        code = normalize_trctcd(self.trctcd_input.text())
        office_code = build_office_code(branch, code)
        label = self.short_input.text().strip() or self.name_input.text().strip()
        self.preview_label.setText(f"{office_code} - {label}" if office_code and label else "")

    def _save(self) -> None:
        self._normalize_trctcd_and_type()
        try:
            entry = OfficeDirectoryEntry(
                id=self.office.id if self.office else None,
                branch_code=current_data(self.branch_combo),
                trctcd=self.trctcd_input.text(),
                office_code="",
                office_name=self.name_input.text(),
                short_name=self.short_input.text(),
                office_type=str(self.type_combo.currentData() or OTHER_OFFICE),
                is_active=self.active_check.isChecked(),
            )
            if self.office is None:
                self.service.create_office(entry)
            else:
                self.service.save_office(entry)
        except UnitDirectoryError as exc:
            QMessageBox.warning(self, "Hội sở/Phòng giao dịch", str(exc))
            return
        self.accept()


def export_unit_directory_excel(service: UnitDirectoryService, destination: Path) -> Path:
    workbook = Workbook()
    branch_sheet = workbook.active
    branch_sheet.title = "ChiNhanh"
    _write_rows(
        branch_sheet,
        (
            ("Mã chi nhánh", "Tên chi nhánh", "Tên ngắn", "Tên hiển thị", "Tỉnh/Thành phố", "Trạng thái", "Ngày cập nhật"),
            *(
                (
                    item.branch_code,
                    item.branch_name,
                    item.short_name,
                    service.format_branch_display(item.branch_code),
                    item.province_name,
                    "Đang sử dụng" if item.is_active else "Ngừng sử dụng",
                    item.updated_at,
                )
                for item in service.repository.list_branches(active_only=False)
            ),
        ),
    )
    office_sheet = workbook.create_sheet("PhongGiaoDich")
    _write_rows(
        office_sheet,
        (
            ("Mã chi nhánh", "Mã TRCTCD", "Mã đơn vị", "Loại đơn vị", "Tên đơn vị", "Tên ngắn", "Trạng thái", "Ngày cập nhật"),
            *(
                (
                    item.branch_code,
                    item.trctcd,
                    item.office_code,
                    item.office_type,
                    item.office_name,
                    item.short_name,
                    "Đang sử dụng" if item.is_active else "Ngừng sử dụng",
                    item.updated_at,
                )
                for item in service.repository.list_offices(active_only=False)
            ),
        ),
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def _write_rows(worksheet, rows) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        worksheet.append(list(row))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        width = min(48, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        worksheet.column_dimensions[column[0].column_letter].width = width


def _office_type_label(value: str) -> str:
    labels = {
        HEAD_OFFICE: "Hội sở",
        TRANSACTION_OFFICE: "Phòng giao dịch",
        OTHER_OFFICE: "Khác",
    }
    if value not in OFFICE_TYPES:
        return value
    return labels.get(value, value)
