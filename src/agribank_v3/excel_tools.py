from __future__ import annotations

import csv
from copy import copy
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
import re
import sys
import tempfile

from openpyxl import Workbook, load_workbook
from xlrd import XL_CELL_BOOLEAN, XL_CELL_DATE, XL_CELL_EMPTY, XL_CELL_ERROR, xldate_as_datetime
import xlrd


class ExcelToolError(RuntimeError):
    """Raised when an Excel utility operation cannot be completed."""


@dataclass(frozen=True, slots=True)
class CsvConversionResult:
    output_path: Path
    row_count: int
    column_count: int


@dataclass(frozen=True, slots=True)
class SheetSplitResult:
    source_path: Path
    output_directory: Path
    output_paths: tuple[Path, ...]


def convert_csv_to_xlsx(source_path: Path, output_path: Path) -> CsvConversionResult:
    return convert_csv_to_excel(source_path, output_path, output_format="xlsx")


def convert_csv_to_excel(
    source_path: Path,
    output_path: Path,
    *,
    output_format: str = "xlsx",
) -> CsvConversionResult:
    source = Path(source_path)
    if source.suffix.casefold() != ".csv":
        raise ExcelToolError(f"File không phải CSV: {source.name}")
    if not source.is_file():
        raise ExcelToolError(f"Không tìm thấy file CSV: {source}")
    normalized_format = output_format.casefold().lstrip(".")
    if normalized_format not in {"xlsx", "xls"}:
        raise ExcelToolError("Định dạng xuất chỉ hỗ trợ XLSX hoặc XLS.")
    output = Path(output_path)
    expected_suffix = f".{normalized_format}"
    if output.suffix.casefold() != expected_suffix:
        output = output.with_suffix(expected_suffix)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook_output = output
    temporary_xlsx: Path | None = None
    if normalized_format == "xls":
        temporary_xlsx = Path(tempfile.gettempdir()) / f"{output.stem}.agribankv3.tmp.xlsx"
        workbook_output = temporary_xlsx

    rows = _read_csv_rows(source)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(source.stem)
    row_count = 0
    column_count = 0
    try:
        for row in rows:
            sheet.append([_normalize_csv_cell(value) for value in row])
            row_count += 1
            column_count = max(column_count, len(row))
        workbook.save(workbook_output)
    except Exception:
        workbook.close()
        for cleanup_path in (workbook_output, output):
            if cleanup_path.exists():
                cleanup_path.unlink()
        raise
    workbook.close()
    if normalized_format == "xls":
        if temporary_xlsx is None:
            raise ExcelToolError("Không tạo được file tạm để chuyển XLS.")
        try:
            _save_xlsx_as_xls(temporary_xlsx, output)
        finally:
            if temporary_xlsx.exists():
                temporary_xlsx.unlink()
    return CsvConversionResult(output, row_count, column_count)


def split_workbook_sheets_to_files(
    source_path: Path,
    output_directory: Path | None = None,
    sheet_names: tuple[str, ...] | None = None,
) -> SheetSplitResult:
    source = Path(source_path)
    if source.suffix.casefold() not in {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ExcelToolError("Chức năng tách sheet hỗ trợ file .xls/.xlsx/.xlsm/.xltx/.xltm.")
    if not source.is_file():
        raise ExcelToolError(f"Không tìm thấy file Excel: {source}")
    destination = (
        Path(output_directory)
        if output_directory is not None
        else source.with_name(f"{source.stem}_tach_sheet")
    )
    destination.mkdir(parents=True, exist_ok=True)
    if source.suffix.casefold() == ".xls":
        return _split_xls_sheets_to_files(source, destination, sheet_names)

    workbook = load_workbook(source, data_only=False)
    output_paths: list[Path] = []
    try:
        selected_sheet_names = tuple(sheet_names) if sheet_names is not None else tuple(workbook.sheetnames)
        missing_sheets = [name for name in selected_sheet_names if name not in workbook.sheetnames]
        if missing_sheets:
            raise ExcelToolError("Không tìm thấy sheet: " + ", ".join(missing_sheets))
        if not selected_sheet_names:
            raise ExcelToolError("Chưa chọn sheet để tách.")
        for sheet_name in selected_sheet_names:
            new_workbook = Workbook()
            default_sheet = new_workbook.active
            new_workbook.remove(default_sheet)
            source_sheet = workbook[sheet_name]
            copied_sheet = new_workbook.create_sheet(title=_safe_sheet_title(sheet_name))
            _copy_sheet_values(source_sheet, copied_sheet)
            output_path = _unique_output_path(
                destination / f"{source.stem}_{_safe_filename(sheet_name)}.xlsx"
            )
            new_workbook.save(output_path)
            new_workbook.close()
            output_paths.append(output_path)
    finally:
        workbook.close()
    return SheetSplitResult(source, destination, tuple(output_paths))


def list_workbook_sheet_names(source_path: Path) -> tuple[str, ...]:
    source = Path(source_path)
    if source.suffix.casefold() not in {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ExcelToolError("Chức năng này hỗ trợ file .xls/.xlsx/.xlsm/.xltx/.xltm.")
    if not source.is_file():
        raise ExcelToolError(f"Không tìm thấy file Excel: {source}")
    if source.suffix.casefold() == ".xls":
        workbook = xlrd.open_workbook(str(source), logfile=StringIO())
        return tuple(workbook.sheet_names())
    workbook = load_workbook(source, read_only=True, data_only=False)
    try:
        return tuple(workbook.sheetnames)
    finally:
        workbook.close()


def _read_csv_rows(path: Path) -> list[list[str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1258", "mbcs"):
        try:
            text = path.read_text(encoding=encoding)
            dialect = _sniff_dialect(text)
            return [list(row) for row in csv.reader(text.splitlines(), dialect)]
        except (OSError, UnicodeError, csv.Error) as exc:
            last_error = exc
    raise ExcelToolError(f"Không thể đọc file CSV {path.name}: {last_error}")


def _normalize_csv_cell(value: str) -> object:
    if value == "":
        return None
    if value.startswith("'"):
        return value
    text = value.strip()
    if text == "":
        return None
    if re.fullmatch(r"[+-]?\d+", text):
        return int(text)
    if re.fullmatch(r"[+-]?(?:\d+\.\d*|\d*\.\d+)", text):
        return float(text)
    return value


def _sniff_dialect(text: str) -> csv.Dialect:
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _copy_sheet_values(source_sheet, target_sheet) -> None:
    for row in source_sheet.iter_rows():
        for cell in row:
            target = target_sheet.cell(row=cell.row, column=cell.column)
            target.value = cell.value
            if cell.has_style:
                target._style = copy(cell._style)
            if cell.number_format:
                target.number_format = cell.number_format
    for column_letter, dimension in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[column_letter].width = dimension.width
    for row_index, dimension in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_index].height = dimension.height
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))


def _split_xls_sheets_to_files(
    source: Path,
    destination: Path,
    sheet_names: tuple[str, ...] | None,
) -> SheetSplitResult:
    workbook = xlrd.open_workbook(str(source), logfile=StringIO(), formatting_info=True)
    workbook_sheet_names = tuple(workbook.sheet_names())
    selected_sheet_names = tuple(sheet_names) if sheet_names is not None else workbook_sheet_names
    missing_sheets = [name for name in selected_sheet_names if name not in workbook_sheet_names]
    if missing_sheets:
        raise ExcelToolError("Không tìm thấy sheet: " + ", ".join(missing_sheets))
    if not selected_sheet_names:
        raise ExcelToolError("Chưa chọn sheet để tách.")

    output_paths: list[Path] = []
    for sheet_name in selected_sheet_names:
        source_sheet = workbook.sheet_by_name(sheet_name)
        new_workbook = Workbook()
        target_sheet = new_workbook.active
        target_sheet.title = _safe_sheet_title(sheet_name)
        for row_index in range(source_sheet.nrows):
            for column_index in range(source_sheet.ncols):
                cell = source_sheet.cell(row_index, column_index)
                target_sheet.cell(
                    row=row_index + 1,
                    column=column_index + 1,
                    value=_xls_cell_value(cell, workbook.datemode),
                )
        output_path = _unique_output_path(
            destination / f"{source.stem}_{_safe_filename(sheet_name)}.xlsx"
        )
        new_workbook.save(output_path)
        new_workbook.close()
        output_paths.append(output_path)
    return SheetSplitResult(source, destination, tuple(output_paths))


def _xls_cell_value(cell, datemode: int) -> object:
    if cell.ctype == XL_CELL_EMPTY:
        return None
    if cell.ctype == XL_CELL_DATE:
        try:
            return xldate_as_datetime(cell.value, datemode)
        except Exception:
            return cell.value
    if cell.ctype == XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == XL_CELL_ERROR:
        return None
    return cell.value


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", value).strip(" .")
    return cleaned or "Sheet"


def _safe_sheet_title(value: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]+", "_", value).strip()
    return (cleaned or "Sheet")[:31]


def _unique_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(1, 1000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise ExcelToolError(f"Không tạo được tên file không trùng cho: {path.name}")


def _save_xlsx_as_xls(source_path: Path, output_path: Path) -> None:
    if not sys.platform.startswith("win"):
        raise ExcelToolError("Xuất XLS chỉ hỗ trợ trên Windows có Microsoft Excel.")
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise ExcelToolError("Xuất XLS cần Microsoft Excel và pywin32.") from exc

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(source_path))
        workbook.SaveAs(str(output_path), FileFormat=56)
    except Exception as exc:
        raise ExcelToolError(f"Không thể lưu file XLS: {exc}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
