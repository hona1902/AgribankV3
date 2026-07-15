from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
import xlrd


class FileMergeError(RuntimeError):
    """Raised when same-structure files cannot be merged."""


@dataclass(frozen=True, slots=True)
class MergeResult:
    output_path: Path
    source_count: int
    row_count: int
    column_count: int


def merge_same_structure_csv_to_xlsx(
    source_paths: Iterable[Path],
    output_path: Path,
    *,
    sheet_name: str = "Data",
    include_source_filename: bool = False,
    source_filename_column: str = "File gốc",
) -> MergeResult:
    sources = tuple(Path(path) for path in source_paths)
    if not sources:
        raise FileMergeError("Chưa chọn file CSV để nối.")
    for source in sources:
        if source.suffix.casefold() != ".csv":
            raise FileMergeError(f"File không phải CSV: {source.name}")
        if not source.is_file():
            raise FileMergeError(f"Không tìm thấy file: {source}")

    output = Path(output_path)
    if output.suffix.casefold() != ".xlsx":
        output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    try:
        expected_header: tuple[str, ...] | None = None
        total_rows = 0
        column_count = 0
        for source in sources:
            header, rows = _read_csv(source)
            if not header:
                raise FileMergeError(f"File không có dòng tiêu đề: {source.name}")
            if expected_header is None:
                expected_header = header
                column_count = len(header)
                output_header = list(header)
                if include_source_filename:
                    output_header.append(source_filename_column)
                sheet.append(output_header)
            elif header != expected_header:
                raise FileMergeError(
                    f"File {source.name} không cùng cấu trúc với file {sources[0].name}."
                )
            for row in rows:
                if not any(value != "" for value in row):
                    continue
                normalized = list(row[:column_count])
                if len(normalized) < column_count:
                    normalized.extend([""] * (column_count - len(normalized)))
                if include_source_filename:
                    normalized.append(source.name)
                sheet.append(normalized)
                total_rows += 1

        workbook.save(output)
    except Exception:
        workbook.close()
        if output.exists():
            output.unlink()
        raise
    workbook.close()
    return MergeResult(
        output_path=output,
        source_count=len(sources),
        row_count=total_rows,
        column_count=_merged_column_count(column_count, include_source_filename),
    )


def merge_same_structure_csv_to_csv(
    source_paths: Iterable[Path],
    output_path: Path,
    *,
    include_source_filename: bool = False,
    source_filename_column: str = "File gốc",
) -> MergeResult:
    sources = tuple(Path(path) for path in source_paths)
    if not sources:
        raise FileMergeError("Chưa chọn file CSV để nối.")
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    expected_header: tuple[str, ...] | None = None
    total_rows = 0
    column_count = 0
    try:
        with output.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(stream)
            for source in sources:
                if source.suffix.casefold() != ".csv":
                    raise FileMergeError(f"File không phải CSV: {source.name}")
                if not source.is_file():
                    raise FileMergeError(f"Không tìm thấy file: {source}")
                header, rows = _read_csv(source)
                if not header:
                    raise FileMergeError(f"File không có dòng tiêu đề: {source.name}")
                if expected_header is None:
                    expected_header = header
                    column_count = len(header)
                    output_header = list(header)
                    if include_source_filename:
                        output_header.append(source_filename_column)
                    writer.writerow(output_header)
                elif header != expected_header:
                    raise FileMergeError(
                        f"File {source.name} không cùng cấu trúc với file {sources[0].name}."
                    )
                for row in rows:
                    if not any(value != "" for value in row):
                        continue
                    normalized = list(row[:column_count])
                    if len(normalized) < column_count:
                        normalized.extend([""] * (column_count - len(normalized)))
                    if include_source_filename:
                        normalized.append(source.name)
                    writer.writerow(normalized)
                    total_rows += 1
    except Exception:
        if output.exists():
            output.unlink()
        raise
    return MergeResult(
        output_path=output,
        source_count=len(sources),
        row_count=total_rows,
        column_count=_merged_column_count(column_count, include_source_filename),
    )


def merge_same_structure_excel_to_xlsx(
    source_paths: Iterable[Path],
    output_path: Path,
    *,
    sheet_name: str | None = None,
    include_source_filename: bool = False,
    source_filename_column: str = "File gốc",
) -> MergeResult:
    sources = tuple(Path(path) for path in source_paths)
    if not sources:
        raise FileMergeError("Chưa chọn file Excel để nối.")
    for source in sources:
        if source.suffix.casefold() not in {".xls", ".xlsm", ".xlsx", ".xltx", ".xltm"}:
            raise FileMergeError(f"File không phải Excel: {source.name}")
        if not source.is_file():
            raise FileMergeError(f"Không tìm thấy file: {source}")

    output = Path(output_path)
    if output.suffix.casefold() != ".xlsx":
        output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name or "Data"
    expected_header: tuple[str, ...] | None = None
    total_rows = 0
    column_count = 0

    try:
        for source in sources:
            header, rows = _read_excel_sheet(source)
            if not header:
                raise FileMergeError(f"File không có dòng tiêu đề: {source.name}")
            if expected_header is None:
                expected_header = header
                column_count = len(header)
                output_header = list(header)
                if include_source_filename:
                    output_header.append(source_filename_column)
                sheet.append(output_header)
            elif header != expected_header:
                raise FileMergeError(
                    f"File {source.name} không cùng cấu trúc với file {sources[0].name}."
                )
            for row in rows:
                if not any(value not in (None, "") for value in row):
                    continue
                normalized = list(row[:column_count])
                if len(normalized) < column_count:
                    normalized.extend([None] * (column_count - len(normalized)))
                if include_source_filename:
                    normalized.append(source.name)
                sheet.append(normalized)
                total_rows += 1
        workbook.save(output)
    except Exception:
        workbook.close()
        if output.exists():
            output.unlink()
        raise
    workbook.close()
    return MergeResult(
        output_path=output,
        source_count=len(sources),
        row_count=total_rows,
        column_count=_merged_column_count(column_count, include_source_filename),
    )


def _read_csv(path: Path) -> tuple[tuple[str, ...], list[list[str]]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1258", "mbcs"):
        try:
            text = path.read_text(encoding=encoding)
            dialect = _sniff_dialect(text)
            reader = csv.reader(text.splitlines(), dialect)
            rows = [list(row) for row in reader]
            if not rows:
                return (), []
            return tuple(cell.strip() for cell in rows[0]), rows[1:]
        except (OSError, UnicodeError, csv.Error) as exc:
            last_error = exc
    raise FileMergeError(f"Không thể đọc file CSV {path.name}: {last_error}")


def _sniff_dialect(text: str) -> csv.Dialect:
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _read_excel_sheet(path: Path) -> tuple[tuple[str, ...], list[list[object]]]:
    if path.suffix.casefold() == ".xls":
        return _read_xls_sheet(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise FileMergeError(f"Không thể đọc file Excel {path.name}: {exc}") from exc
    try:
        sheet = workbook.worksheets[0]
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header = _normalize_excel_header(header_values)
        if not header:
            return (), []
        rows: list[list[object]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(list(row[: len(header)]))
        return header, rows
    finally:
        workbook.close()


def _normalize_excel_header(values: tuple[object, ...]) -> tuple[str, ...]:
    cells = list(values)
    while cells and cells[-1] in (None, ""):
        cells.pop()
    return tuple("" if value is None else str(value).strip() for value in cells)


def _read_xls_sheet(path: Path) -> tuple[tuple[str, ...], list[list[object]]]:
    try:
        workbook = xlrd.open_workbook(str(path))
    except Exception as exc:
        raise FileMergeError(f"Không thể đọc file Excel {path.name}: {exc}") from exc
    if workbook.nsheets < 1:
        return (), []
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows < 1:
        return (), []
    header = _normalize_excel_header(tuple(sheet.row_values(0)))
    rows = [sheet.row_values(index)[: len(header)] for index in range(1, sheet.nrows)]
    return header, rows


def _merged_column_count(column_count: int, include_source_filename: bool) -> int:
    return column_count + 1 if include_source_filename and column_count else column_count
