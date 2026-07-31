from __future__ import annotations

from contextlib import closing
import os
from pathlib import Path
import sqlite3
from tempfile import TemporaryDirectory
from time import perf_counter, sleep
import threading
import unittest
import zipfile
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from PySide6.QtCharts import QValueAxis
from PySide6.QtCore import QEvent, QRect, Qt
from PySide6.QtWidgets import QApplication, QAbstractItemView, QComboBox, QDialog, QLabel, QPushButton, QScrollArea, QSizePolicy, QStyleOptionViewItem, QTableView, QTableWidget, QWidget

from agribank_v3.features.catalog import SECTIONS
from agribank_v3.features.credit.summary.customer import chart_service
from agribank_v3.features.credit.summary.customer.async_query import AsyncQueryController, LruQueryCache
from agribank_v3.features.credit.summary.customer.charts import (
    ChartTooltip,
    CustomerDonutChart,
    CustomerHorizontalBarChart,
    CustomerLineChart,
)
from agribank_v3.features.credit.summary.customer.charts.chart_formatters import (
    format_money_axis,
    format_money_full,
    format_percentage,
)
from agribank_v3.features.credit.summary.models import LOAN_COMPARE_TITLE, NIM_DN_CONFIG, SummaryDataType, SummaryError
from agribank_v3.features.credit.summary.repository import SummaryRepository
from agribank_v3.features.credit.summary.services import _parse_nim_file, import_nim_dn
from agribank_v3.features.credit.summary.customer.database import (
    CUSTOMER_DATABASE_NAME,
    CUSTOMER_SCHEMA_MIGRATION_NAME,
    CUSTOMER_SCHEMA_VERSION,
    CustomerDatabaseOperationLock,
    customer_database_path,
    get_customer_database_connection,
)
from agribank_v3.features.credit.summary.customer.export_service import (
    export_all_customer_sheets,
    export_cross_branch_customers,
    export_cross_branch_customer_detail,
    export_customer_dashboard,
    export_customer_detail,
    export_customer_growth,
    export_customer_list,
    export_import_history,
    export_multiple_officers,
    export_officer_directory,
    export_top_customer_balance,
    export_top_customer_movement,
)
from agribank_v3.features.credit.summary.customer.filters import CustomerFilters
from agribank_v3.features.credit.summary.customer.formatters import format_money_vn, format_percent_vn
from agribank_v3.features.credit.summary.customer.cross_branch_detail_dialog import CrossBranchCustomerDetailDialog
from agribank_v3.features.credit.summary.customer.cross_branch_tab import CrossBranchCustomersTab
from agribank_v3.features.credit.summary.customer.customer_detail_window import CustomerDetailWindow
from agribank_v3.features.credit.summary.customer.dashboard_tab import DASHBOARD_MAIN_KPI_LABELS, CustomerDashboardTab
from agribank_v3.features.credit.summary.customer.management_window import CustomerManagementWindow
from agribank_v3.features.credit.summary.customer.maintenance_dialog import CustomerMaintenanceDialog
from agribank_v3.features.credit.summary.customer.models import LoanTerm
from agribank_v3.features.credit.summary.customer.movement_tab import CustomerMovementTab
from agribank_v3.features.credit.summary.customer.officer_lookup import OfficerLookupWidget
from agribank_v3.features.credit.summary.customer.officer_management_tab import OfficerDirectoryDialog
from agribank_v3.features.credit.summary.customer.officer_override_dialog import OfficerOverrideDialog
from agribank_v3.features.credit.summary.customer.period_validation import validate_dashboard_period_filters
from agribank_v3.features.credit.summary.customer.repository import CustomerRepository, _cross_branch_candidate_sql, _movement_base_sql
from agribank_v3.features.credit.summary.customer.services import (
    build_office_code,
    build_customer_code,
    classify_office_type,
    classify_loan_term,
    classify_customer_movement,
    growth_rate,
    normalize_trctcd,
    normalize_customer_sequence,
    resolve_representative_office,
    split_officer,
)
from agribank_v3.features.credit.summary.customer.table_models import CustomerTableModel
from agribank_v3.features.credit.summary.customer.export_service import CUSTOMER_LIST_COLUMNS
from agribank_v3.features.credit.summary.customer.routes import (
    CUSTOMER_DATA_ROUTE,
    CUSTOMER_DATA_TITLE,
)
from agribank_v3.features.credit.summary.customer.widgets import (
    AGRIBANK_TABLE_SELECTION_STYLE,
    CompactKpiCard,
    CompactToolbar,
    CustomerTableView,
    KpiMetric,
    OFFICER_NAME_ROLE,
    QueryStateBanner,
    ResponsiveKpiGrid,
    ScopeFilterComboBox,
    SearchBox,
    compact_money_vn,
    combo_box,
    configure_combo_popup_width,
    current_data,
    ensure_geometry_visible,
    populate_officer_combo,
)
from agribank_v3.features.credit.summary.customer.window_controller import open_customer_management_window
from agribank_v3.features.credit.summary.windows import NimTab
from agribank_v3.features.settings.unit_directory.models import (
    BranchDirectoryEntry,
    OfficeDirectoryEntry,
    TRANSACTION_OFFICE,
)
from agribank_v3.ui.main_window import MainWindow


FTPLN_HEADER = "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP,FTPCD,CUSTSEQ,CUSTNM"


def _wait_until(predicate, timeout: float = 3.0) -> bool:
    app = QApplication.instance() or QApplication([])
    deadline = perf_counter() + timeout
    while perf_counter() < deadline:
        app.processEvents()
        if predicate():
            return True
        sleep(0.01)
    app.processEvents()
    return bool(predicate())


def _drain_customer_query_threads(timeout_ms: int = 5000) -> None:
    app = QApplication.instance()
    if app is None:
        return
    for widget in list(app.allWidgets()):
        for controller in widget.findChildren(AsyncQueryController):
            controller.wait_for_idle(timeout_ms)
    app.processEvents()


class CustomerDatabaseTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.main_database_path = self.root / "DuLieuV3.db"
        self.repository = CustomerRepository(self.main_database_path)
        self.summary_repository = SummaryRepository(self.main_database_path)

    def tearDown(self) -> None:
        _drain_customer_query_threads()
        self.temporary_directory.cleanup()

    def test_customer_database_created_separately(self) -> None:
        self.assertEqual(self.repository.database_path, self.root / CUSTOMER_DATABASE_NAME)
        self.assertEqual(customer_database_path(self.main_database_path), self.root / CUSTOMER_DATABASE_NAME)
        self.assertTrue(self.repository.database_path.is_file())
        self.assertNotEqual(self.repository.database_path, self.main_database_path)

    def test_customer_schema_tables_and_indexes(self) -> None:
        expected_tables = {
            "customer_action_log",
            "customer_import_files",
            "customer_import_runs",
            "customer_master",
            "customer_officer_directory",
            "customer_officer_override",
            "customer_officer_period",
            "customer_office_period",
            "customer_period_summary",
            "customer_schema_migrations",
        }
        expected_indexes = {
            "idx_customer_period_summary_period",
            "idx_customer_period_summary_branch_period",
            "idx_customer_period_summary_officer_period",
            "idx_customer_period_summary_type_period",
            "idx_customer_period_summary_code_period",
            "idx_customer_period_summary_period_sequence_branch",
            "idx_customer_period_summary_period_officer_sequence",
            "idx_customer_officer_period_customer_period",
            "idx_customer_officer_period_officer_period",
            "idx_customer_officer_period_period_primary",
            "idx_customer_office_period_customer_period",
            "idx_customer_office_period_period_customer_branch_trctcd",
            "idx_customer_office_period_period_branch_office",
            "idx_customer_office_period_period_customer_branch_office",
            "idx_customer_office_period_period_branch_type_sequence",
            "idx_customer_officer_override_customer_effective",
        }
        with closing(self.repository.connect()) as connection:
            tables = {
                str(row[0])
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'table'"
                ).fetchall()
            }
            indexes = {
                str(row[0])
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'index'"
                ).fetchall()
            }

        self.assertTrue(expected_tables.issubset(tables))
        self.assertTrue(expected_indexes.issubset(indexes))

    def test_customer_schema_migration_idempotent(self) -> None:
        self.repository.ensure_schema()
        self.repository.ensure_schema()

        with closing(self.repository.connect()) as connection:
            count_rows = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM customer_schema_migrations
                    WHERE version = ? AND migration_name = ?
                    """,
                    (CUSTOMER_SCHEMA_VERSION, CUSTOMER_SCHEMA_MIGRATION_NAME),
                ).fetchone()[0]
            )

        self.assertEqual(count_rows, 1)

    def test_customer_repository_status_empty_database(self) -> None:
        status = self.repository.maintenance_status()

        self.assertEqual(status.master_count, 0)
        self.assertEqual(status.period_count, 0)
        self.assertEqual(status.period_summary_count, 0)
        self.assertEqual(status.officer_period_count, 0)
        self.assertEqual(status.override_count, 0)
        self.assertEqual(status.first_period, "")
        self.assertEqual(status.last_period, "")

    def test_customer_code_combines_branch_and_sequence(self) -> None:
        self.assertEqual(build_customer_code("5491", "'177616932"), "5491177616932")

    def test_customer_code_from_branch_and_sequence(self) -> None:
        self.assertEqual(build_customer_code("5491", "'177616932"), "5491177616932")

    def test_customer_sequence_preserves_leading_zero(self) -> None:
        self.assertEqual(normalize_customer_sequence("'00177616932"), "00177616932")
        self.assertEqual(build_customer_code("5491", "'00177616932"), "549100177616932")

    def test_customer_sequence_removes_excel_float_suffix_only(self) -> None:
        self.assertEqual(normalize_customer_sequence("177616932.0"), "177616932")
        self.assertEqual(normalize_customer_sequence("177616932.50"), "177616932.50")

    def test_ftp_code_short_term_mapping(self) -> None:
        for code in ("DN1", "DN2", "DN3", "DN4", "DN5", "DN6", "DN13", "DN14", "DN15", "DN16"):
            self.assertEqual(classify_loan_term(code), LoanTerm.SHORT_TERM)

    def test_ftp_code_medium_long_mapping(self) -> None:
        for code in ("DN7", "DN8", "DN9", "DN10", "DN11", "DN12"):
            self.assertEqual(classify_loan_term(code), LoanTerm.MEDIUM_LONG_TERM)

    def test_unknown_ftp_code_goes_to_other_balance(self) -> None:
        self.assertEqual(classify_loan_term("DN99"), LoanTerm.UNKNOWN)

    def test_split_officer_code_and_name(self) -> None:
        officer = split_officer("[540400321] Nguyen Van A")

        self.assertEqual(officer.officer_code, "540400321")
        self.assertEqual(officer.officer_name, "Nguyen Van A")

    def test_customer_backup_restore(self) -> None:
        with closing(self.repository.connect()) as connection:
            connection.execute(
                """
                INSERT INTO customer_master(
                    customer_code, branch_code, customer_sequence, customer_name,
                    customer_type, created_at, updated_at
                )
                VALUES ('5491001', '5491', '001', 'Khach hang A', 'CN', 'now', 'now')
                """
            )
            connection.commit()
        backup_path = self.repository.backup_database(self.root / "customer-backup.zip")
        with zipfile.ZipFile(backup_path) as archive:
            self.assertIn(CUSTOMER_DATABASE_NAME, archive.namelist())
        with closing(self.repository.connect()) as connection:
            connection.execute("DELETE FROM customer_master")
            connection.commit()

        self.repository.restore_database(backup_path)

        with closing(self.repository.connect()) as connection:
            count_rows = int(connection.execute("SELECT COUNT(*) FROM customer_master").fetchone()[0])
        self.assertEqual(count_rows, 1)

    def test_normalized_loan_row_from_ftpln_headers(self) -> None:
        path = self.root / "5491_FTPLN_20260331.csv"
        path.write_text(
            "\n".join(
                [
                    "CUSTNM,CBTD,LDRBAL,CUSTSEQ,BRCD,FTPCD,CUSTTP,FTP,INTRT,MUCFTPDC,TRCTCD",
                    "Khach A,[540000321] Nguyen Van A,1000,'177616932,5491,DN1,CN,2,10,1,00",
                ]
            ),
            encoding="utf-8",
        )

        parsed = _parse_nim_file(path, NIM_DN_CONFIG, credit_card_rate=0)
        row = parsed.customer_rows[0]

        self.assertEqual(row.period, "2026-03")
        self.assertEqual(row.customer_sequence, "177616932")
        self.assertEqual(row.customer_code, "5491177616932")
        self.assertEqual(row.customer_name, "Khach A")
        self.assertEqual(row.ftp_code, "DN1")
        self.assertAlmostEqual(row.balance, 1000)
        self.assertEqual(row.officer_code, "540000321")
        self.assertEqual(row.trctcd, "00")
        self.assertEqual(row.office_code, "5491-00")
        self.assertEqual(row.office_name, "Hội sở")
        self.assertEqual(row.office_type, "HEAD_OFFICE")

    def test_normalized_loan_row_contains_trctcd(self) -> None:
        path = self.root / "5491_FTPLN_20260331.csv"
        path.write_text(
            "\n".join(
                [
                    "CUSTNM,CBTD,LDRBAL,CUSTSEQ,BRCD,FTPCD,CUSTTP,FTP,INTRT,MUCFTPDC,TRCTCD",
                    "Khach A,[540000321] Nguyen Van A,1000,'001,5491,DN1,CN,2,10,1,01",
                ]
            ),
            encoding="utf-8",
        )

        row = _parse_nim_file(path, NIM_DN_CONFIG, credit_card_rate=0).customer_rows[0]

        self.assertEqual(row.trctcd, "01")
        self.assertEqual(row.office_code, "5491-01")
        self.assertEqual(row.office_type, "TRANSACTION_OFFICE")

    def test_head_office_code_00(self) -> None:
        self.assertEqual(normalize_trctcd("0.0"), "00")
        self.assertEqual(build_office_code("5405", "00"), "5405-00")
        self.assertEqual(classify_office_type("00").value, "HEAD_OFFICE")

    def test_transaction_office_non_zero_code(self) -> None:
        self.assertEqual(normalize_trctcd("1"), "01")
        self.assertEqual(build_office_code("5405", "1"), "5405-01")
        self.assertEqual(classify_office_type("1").value, "TRANSACTION_OFFICE")

    def test_customer_office_one_row_per_period_customer_office(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", trctcd="01"),
                self._row("001", "Khach A", 3000, "DN7", trctcd="01"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)

        with closing(self.repository.connect()) as connection:
            row = connection.execute(
                """
                SELECT COUNT(*) AS row_count, SUM(source_loan_count) AS source_count
                FROM customer_office_period
                WHERE period = '2026-03'
                  AND customer_code = '5491001'
                  AND office_code = '5491-01'
                """
            ).fetchone()
        self.assertEqual(int(row["row_count"] or 0), 1)
        self.assertEqual(int(row["source_count"] or 0), 2)

    def test_customer_office_total_matches_customer_summary(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", trctcd="00"),
                self._row("001", "Khach A", 3000, "DN7", trctcd="01"),
                self._row("002", "Khach B", 2000, "DN1", trctcd="02"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)

        with closing(self.repository.connect()) as connection:
            mismatches = connection.execute(
                """
                SELECT s.customer_code
                FROM customer_period_summary s
                LEFT JOIN (
                    SELECT period, customer_code, SUM(total_balance) AS total_balance
                    FROM customer_office_period
                    GROUP BY period, customer_code
                ) o
                    ON o.period = s.period
                   AND o.customer_code = s.customer_code
                WHERE s.period = '2026-03'
                  AND ABS(s.total_balance - COALESCE(o.total_balance, 0)) > 0.0001
                """
            ).fetchall()
        self.assertEqual(mismatches, [])

    def test_customer_office_does_not_store_raw_loans(self) -> None:
        self._import_basic_customer_fixture()

        with closing(self.repository.connect()) as connection:
            tables = {
                str(row[0])
                for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
            }
        self.assertNotIn("customer_raw_loans", tables)
        self.assertNotIn("customer_loan_details", tables)
        self.assertNotIn("ftpln_raw", tables)

    def test_same_customer_two_pgds_two_office_rows(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", trctcd="01"),
                self._row("001", "Khach A", 3000, "DN7", trctcd="02"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)

        self.assertEqual(self._count("customer_period_summary"), 1)
        with closing(self.repository.connect()) as connection:
            offices = [
                str(row["office_code"])
                for row in connection.execute(
                    """
                    SELECT office_code
                    FROM customer_office_period
                    WHERE period = '2026-03' AND customer_code = '5491001'
                    ORDER BY office_code
                    """
                ).fetchall()
            ]
        self.assertEqual(offices, ["5491-01", "5491-02"])

    def test_import_file_still_read_once(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [self._row("001", "Khach A", 1000, "DN1", trctcd="01")],
        )
        calls: list[str] = []
        original_read_bytes = Path.read_bytes

        def counted_read_bytes(path: Path) -> bytes:
            calls.append(path.name)
            return original_read_bytes(path)

        with patch.object(Path, "read_bytes", counted_read_bytes):
            import_nim_dn(self.summary_repository, self.root)

        self.assertEqual(calls.count("5491_FTPLN_20260331.csv"), 1)

    def test_one_customer_many_loans_one_summary_row(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1"),
                self._row("001", "Khach A", 3000, "DN7"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)

        self.assertEqual(self._count("customer_period_summary"), 1)
        summary = self._customer_summary("5491001", "2026-03")
        self.assertAlmostEqual(float(summary["total_balance"]), 4000)
        self.assertEqual(int(summary["source_loan_count"]), 2)

    def test_two_customers_two_summary_rows(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1"),
                self._row("002", "Khach B", 2000, "DN7"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)

        self.assertEqual(self._count("customer_period_summary"), 2)

    def test_customer_total_balance_matches_source(self) -> None:
        self._import_basic_customer_fixture()

        self.assertAlmostEqual(self._customer_total("2026-03"), 6000)

    def test_customer_short_term_balance(self) -> None:
        self._import_basic_customer_fixture()

        summary = self._customer_summary("5491001", "2026-03")
        self.assertAlmostEqual(float(summary["short_term_balance"]), 1000)

    def test_customer_medium_long_term_balance(self) -> None:
        self._import_basic_customer_fixture()

        summary = self._customer_summary("5491001", "2026-03")
        self.assertAlmostEqual(float(summary["medium_long_term_balance"]), 3000)

    def test_customer_other_balance(self) -> None:
        self._import_basic_customer_fixture()

        summary = self._customer_summary("5491002", "2026-03")
        self.assertAlmostEqual(float(summary["other_balance"]), 2000)

    def test_customer_medium_long_ratio(self) -> None:
        self._import_basic_customer_fixture()

        summary = self._customer_summary("5491001", "2026-03")
        self.assertAlmostEqual(float(summary["medium_long_ratio"]), 75.0)

    def test_customer_weighted_average_rate(self) -> None:
        self._import_basic_customer_fixture()

        summary = self._customer_summary("5491001", "2026-03")
        self.assertAlmostEqual(float(summary["average_rate"]), 8.5)

    def test_customer_weighted_nim_before(self) -> None:
        self._import_basic_customer_fixture()

        summary = self._customer_summary("5491001", "2026-03")
        self.assertAlmostEqual(float(summary["nim_before"]), 5.0)

    def test_customer_weighted_nim_after(self) -> None:
        self._import_basic_customer_fixture()

        summary = self._customer_summary("5491001", "2026-03")
        self.assertAlmostEqual(float(summary["nim_after"]), 4.0)

    def test_customer_zero_balance_safe(self) -> None:
        self._write_ftpln("5491_FTPLN_20260331.csv", [self._row("001", "Khach A", 0, "DN1")])

        import_nim_dn(self.summary_repository, self.root)

        summary = self._customer_summary("5491001", "2026-03")
        self.assertEqual(float(summary["average_rate"]), 0)
        self.assertEqual(float(summary["nim_before"]), 0)
        self.assertEqual(float(summary["nim_after"]), 0)
        self.assertEqual(float(summary["medium_long_ratio"]), 0)

    def test_customer_multiple_officers_summary_not_duplicated(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", officer="[540000321] Nguyen Van A"),
                self._row("001", "Khach A", 2000, "DN1", officer="[540000322] Nguyen Van B"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)

        self.assertEqual(self._count("customer_period_summary"), 1)
        summary = self._customer_summary("5491001", "2026-03")
        self.assertEqual(int(summary["officer_count"]), 2)
        self.assertEqual(int(summary["has_multiple_officers"]), 1)

    def test_customer_officer_relation_rows(self) -> None:
        self.test_customer_multiple_officers_summary_not_duplicated()

        self.assertEqual(self._count("customer_officer_period"), 2)

    def test_primary_officer_largest_balance(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", officer="[540000321] Nguyen Van A"),
                self._row("001", "Khach A", 2000, "DN1", officer="[540000322] Nguyen Van B"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)

        summary = self._customer_summary("5491001", "2026-03")
        self.assertEqual(summary["primary_officer_code"], "540000322")

    def test_primary_officer_tie_is_stable(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", officer="[540000321] Nguyen Van A"),
                self._row("001", "Khach A", 1000, "DN1", officer="[540000322] Nguyen Van B"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)

        summary = self._customer_summary("5491001", "2026-03")
        self.assertEqual(summary["primary_officer_code"], "540000321")

    def test_customer_master_upsert(self) -> None:
        self._write_ftpln("5491_FTPLN_20260131.csv", [self._row("001", "Khach A", 1000, "DN1")])
        self._write_ftpln("5491_FTPLN_20260228.csv", [self._row("001", "Khach A Moi", 2000, "DN7")])

        import_nim_dn(self.summary_repository, self.root)

        self.assertEqual(self._count("customer_master"), 1)
        master = self._customer_master("5491001")
        self.assertEqual(master["customer_name"], "Khach A Moi")
        self.assertEqual(master["last_seen_period"], "2026-02")

    def test_customer_first_last_seen_period(self) -> None:
        self.test_customer_master_upsert()

        master = self._customer_master("5491001")
        self.assertEqual(master["first_seen_period"], "2026-01")
        self.assertEqual(master["last_seen_period"], "2026-02")

    def test_customer_import_run_and_files(self) -> None:
        self._import_basic_customer_fixture()

        self.assertEqual(self._count("customer_import_runs"), 1)
        self.assertEqual(self._count("customer_import_files"), 1)
        with closing(self.repository.connect()) as connection:
            run = connection.execute("SELECT * FROM customer_import_runs LIMIT 1").fetchone()
        self.assertEqual(run["status"], "COMPLETED")
        self.assertEqual(int(run["customer_count"]), 2)

    def test_duplicate_file_hash_detected(self) -> None:
        self._write_ftpln("5491_FTPLN_20260331.csv", [self._row("001", "Khach A", 1000, "DN1")])
        import_nim_dn(self.summary_repository, self.root)

        with self.assertRaises(SummaryError):
            import_nim_dn(self.summary_repository, self.root)

    def test_reimport_period_requires_replace(self) -> None:
        folder_a = self.root / "a"
        folder_b = self.root / "b"
        folder_a.mkdir()
        folder_b.mkdir()
        self._write_ftpln("5491_FTPLN_20260331.csv", [self._row("001", "Khach A", 1000, "DN1")], folder=folder_a)
        self._write_ftpln("5491_FTPLN_20260330.csv", [self._row("001", "Khach A", 2000, "DN1")], folder=folder_b)
        import_nim_dn(self.summary_repository, folder_a)

        with self.assertRaises(Exception):
            import_nim_dn(self.summary_repository, folder_b)

    def test_replace_period_preserves_old_data_on_failure(self) -> None:
        folder_a = self.root / "a"
        folder_b = self.root / "b"
        folder_a.mkdir()
        folder_b.mkdir()
        self._write_ftpln("5491_FTPLN_20260331.csv", [self._row("001", "Khach A", 1000, "DN1")], folder=folder_a)
        self._write_ftpln(
            "5491_FTPLN_20260330.csv",
            ["5491,2,10,1,[540000321] Nguyen Van A,00,999,,CN,DN1,,Khach Loi"],
            folder=folder_b,
        )
        import_nim_dn(self.summary_repository, folder_a)

        with self.assertRaises(Exception):
            import_nim_dn(self.summary_repository, folder_b, replace_existing_periods=True)

        self.assertAlmostEqual(self._customer_total("2026-03"), 1000)

    def test_replace_period_updates_nim_and_customer_without_append(self) -> None:
        folder_a = self.root / "a"
        folder_b = self.root / "b"
        folder_a.mkdir()
        folder_b.mkdir()
        self._write_ftpln("5491_FTPLN_20260331.csv", [self._row("001", "Khach A", 1000, "DN1")], folder=folder_a)
        self._write_ftpln("5491_FTPLN_20260330.csv", [self._row("001", "Khach A", 2000, "DN1")], folder=folder_b)
        import_nim_dn(self.summary_repository, folder_a)

        import_nim_dn(self.summary_repository, folder_b, replace_existing_periods=True)

        self.assertAlmostEqual(self._customer_total("2026-03"), 2000)
        with closing(self.summary_repository.connect()) as connection:
            nim_total = float(
                connection.execute(
                    """
                    SELECT SUM(balance)
                    FROM nim_period_summary
                    WHERE data_type = ? AND period = ?
                    """,
                    (SummaryDataType.NIM_DN.value, "2026-03"),
                ).fetchone()[0]
                or 0
            )
        self.assertAlmostEqual(nim_total, 2000)

    def test_import_nim_dn_updates_customer_db(self) -> None:
        result = self._import_basic_customer_fixture()

        self.assertIn("Customer.db", result.message)
        self.assertEqual(self._count("customer_period_summary"), 2)

    def test_nim_dn_result_unchanged(self) -> None:
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", ftp=2, intrt=10, adjustment=1),
                self._row("002", "Khach B", 3000, "DN7", ftp=4, intrt=8, adjustment=1, customer_type="TC"),
            ],
        )

        import_nim_dn(self.summary_repository, self.root)
        row = self.summary_repository.query_nim(SummaryDataType.NIM_DN).rows[0]

        self.assertAlmostEqual(float(row["balance"]), 4000)
        self.assertAlmostEqual(float(row["average_rate"]), 8.5)
        self.assertAlmostEqual(float(row["nim_before"]), 5.0)
        self.assertAlmostEqual(float(row["nim_after"]), 4.0)

    def test_cross_database_failure_does_not_leave_partial_import(self) -> None:
        self._write_ftpln("5491_FTPLN_20260331.csv", [self._row("001", "Khach A", 1000, "DN1")])

        with patch.object(CustomerRepository, "save_aggregation", side_effect=RuntimeError("boom")):
            with self.assertRaises(RuntimeError):
                import_nim_dn(self.summary_repository, self.root)

        self.assertEqual(self._count("customer_period_summary"), 0)
        with closing(self.summary_repository.connect()) as connection:
            count_rows = int(
                connection.execute(
                    "SELECT COUNT(*) FROM nim_period_summary WHERE data_type = ?",
                    (SummaryDataType.NIM_DN.value,),
                ).fetchone()[0]
            )
        self.assertEqual(count_rows, 0)

    def test_large_source_stores_customer_aggregates_only(self) -> None:
        rows = [
            self._row(f"{index % 5000:05d}", f"Khach {index % 5000:05d}", 100, "DN1")
            for index in range(100000)
        ]
        self._write_ftpln("5491_FTPLN_20260331.csv", rows)
        started = perf_counter()

        import_nim_dn(self.summary_repository, self.root)
        elapsed = perf_counter() - started

        self.assertEqual(self._count("customer_period_summary"), 5000)
        self.assertEqual(self._count("customer_officer_period"), 5000)
        self.assertAlmostEqual(self._customer_total("2026-03"), 10000000)
        self.assertLess(elapsed, 60)

    def test_no_raw_loan_table_created(self) -> None:
        self._import_basic_customer_fixture()

        with closing(self.repository.connect()) as connection:
            tables = {
                str(row[0])
                for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
            }
        self.assertNotIn("customer_raw_loans", tables)
        self.assertNotIn("customer_loan_details", tables)
        self.assertNotIn("ftpln_raw", tables)

    def test_sum_customer_balance_equals_nim_balance(self) -> None:
        self._import_basic_customer_fixture()

        with closing(self.summary_repository.connect()) as connection:
            nim_total = float(
                connection.execute(
                    """
                    SELECT SUM(balance)
                    FROM nim_period_summary
                    WHERE data_type = ? AND period = ?
                    """,
                    (SummaryDataType.NIM_DN.value, "2026-03"),
                ).fetchone()[0]
                or 0
            )
        self.assertAlmostEqual(self._customer_total("2026-03"), nim_total)

    def test_unknown_ftp_codes_reported(self) -> None:
        result = self._import_basic_customer_fixture()

        self.assertIn("FTPCD lạ 1", result.message)
        summary = self._customer_summary("5491002", "2026-03")
        self.assertAlmostEqual(float(summary["other_balance"]), 2000)

    def test_customer_branch_combo_uses_unit_directory(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Động", short_name="CN Động")
        )
        self._write_ftpln(
            "6501_FTPLN_20260331.csv",
            [self._row("001", "Khach A", 1000, "DN1", branch_code="6501")],
        )
        import_nim_dn(self.summary_repository, self.root)
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = CrossBranchCustomersTab(self.repository)
        self.addCleanup(tab.deleteLater)

        labels = [tab.branch_combo.itemText(index) for index in range(tab.branch_combo.count())]

        self.assertIn("6501 - CN Động", labels)
        self.assertEqual(tab.branch_combo.itemData(labels.index("6501 - CN Động")), "6501")

    def test_cross_branch_office_combo_uses_unit_directory(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Động", short_name="CN Động")
        )
        self.repository.unit_directory.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="6501",
                trctcd="02",
                office_code="",
                office_name="Phòng giao dịch Động",
                short_name="PGD Động",
                office_type=TRANSACTION_OFFICE,
            )
        )
        self._write_ftpln(
            "6501_FTPLN_20260331.csv",
            [self._row("001", "Khach A", 1000, "DN1", branch_code="6501", trctcd="02")],
        )
        import_nim_dn(self.summary_repository, self.root)
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = CrossBranchCustomersTab(self.repository)
        self.addCleanup(tab.deleteLater)

        labels = [tab.office_combo.itemText(index) for index in range(tab.office_combo.count())]

        self.assertIn("6501-02 - PGD Động", labels)
        self.assertEqual(tab.office_combo.itemData(labels.index("6501-02 - PGD Động")), "6501-02")

    def test_customer_code_unchanged_after_branch_rename(self) -> None:
        self._write_ftpln(
            "6501_FTPLN_20260331.csv",
            [self._row("001", "Khach A", 1000, "DN1", branch_code="6501")],
        )
        import_nim_dn(self.summary_repository, self.root)
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Đổi tên", short_name="CN Đổi tên")
        )

        summary = self._customer_summary("6501001", "2026-03")

        self.assertEqual(summary["customer_code"], "6501001")
        self.assertEqual(summary["branch_code"], "6501")

    def test_customer_total_balance_unchanged(self) -> None:
        self._import_basic_customer_fixture()
        before = self._customer_total("2026-03")
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="5491", branch_name="Chi nhánh Đổi tên", short_name="CN Đổi tên")
        )
        after = self._customer_total("2026-03")

        self.assertAlmostEqual(after, before)

    def test_cross_branch_detection_unchanged(self) -> None:
        self._write_ftpln(
            "6501_FTPLN_20260331.csv",
            [self._row("001", "Khach A", 1000, "DN1", branch_code="6501")],
        )
        self._write_ftpln(
            "6502_FTPLN_20260331.csv",
            [self._row("001", "Khach A", 2000, "DN1", branch_code="6502")],
        )
        import_nim_dn(self.summary_repository, self.root)
        before = self.repository.count_cross_branch_customers("2026-03")
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Đổi tên", short_name="CN Đổi tên")
        )

        self.assertEqual(before, 1)
        self.assertEqual(self.repository.count_cross_branch_customers("2026-03"), 1)

    def test_office_head_pgd_detection_unchanged(self) -> None:
        self._write_ftpln(
            "6501_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", branch_code="6501", trctcd="00"),
                self._row("001", "Khach A", 2000, "DN1", branch_code="6501", trctcd="01"),
            ],
        )
        import_nim_dn(self.summary_repository, self.root)
        before = self.repository.get_cross_branch_customers("2026-03", scope_type="head_and_pgd")
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Đổi tên", short_name="CN Đổi tên")
        )
        after = self.repository.get_cross_branch_customers("2026-03", scope_type="head_and_pgd")

        self.assertEqual(int(before[0]["has_head_and_pgd"]), 1)
        self.assertEqual(int(after[0]["has_head_and_pgd"]), 1)

    def test_missing_customer_headers_does_not_write_partial_customer_db(self) -> None:
        path = self.root / "5491_FTPLN_20260331.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyen Van A,00,1000,,CN",
                ]
            ),
            encoding="utf-8",
        )

        result = import_nim_dn(self.summary_repository, self.root)

        self.assertIn("Customer.db chưa cập nhật", result.message)
        self.assertEqual(self._count("customer_period_summary"), 0)

    def _import_basic_customer_fixture(self):
        self._write_ftpln(
            "5491_FTPLN_20260331.csv",
            [
                self._row("001", "Khach A", 1000, "DN1", ftp=2, intrt=10, adjustment=1),
                self._row("001", "Khach A", 3000, "DN7", ftp=4, intrt=8, adjustment=1),
                self._row("002", "Khach B", 2000, "DN99", ftp=1, intrt=6, adjustment=0, customer_type="TC"),
            ],
        )
        return import_nim_dn(self.summary_repository, self.root)

    def _write_ftpln(self, filename: str, rows: list[str], *, folder: Path | None = None) -> Path:
        target_folder = folder or self.root
        target_folder.mkdir(parents=True, exist_ok=True)
        path = target_folder / filename
        path.write_text("\n".join([FTPLN_HEADER, *rows]), encoding="utf-8")
        return path

    def _row(
        self,
        customer_sequence: str,
        customer_name: str,
        balance: float,
        ftp_code: str,
        *,
        ftp: float = 2,
        intrt: float = 10,
        adjustment: float = 1,
        officer: str = "[540000321] Nguyen Van A",
        customer_type: str = "CN",
        branch_code: str = "5491",
        trctcd: str = "00",
    ) -> str:
        return (
            f"{branch_code},{ftp},{intrt},{adjustment},{officer},{trctcd},{balance},,"
            f"{customer_type},{ftp_code},{customer_sequence},{customer_name}"
        )

    def _count(self, table_name: str) -> int:
        with closing(self.repository.connect()) as connection:
            return int(connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0] or 0)

    def _customer_summary(self, customer_code: str, period: str):
        with closing(self.repository.connect()) as connection:
            row = connection.execute(
                """
                SELECT *
                FROM customer_period_summary
                WHERE customer_code = ? AND period = ?
                """,
                (customer_code, period),
            ).fetchone()
        self.assertIsNotNone(row)
        return row

    def _customer_master(self, customer_code: str):
        with closing(self.repository.connect()) as connection:
            row = connection.execute(
                "SELECT * FROM customer_master WHERE customer_code = ?",
                (customer_code,),
            ).fetchone()
        self.assertIsNotNone(row)
        return row

    def _customer_total(self, period: str) -> float:
        with closing(self.repository.connect()) as connection:
            return float(
                connection.execute(
                    "SELECT SUM(total_balance) FROM customer_period_summary WHERE period = ?",
                    (period,),
                ).fetchone()[0]
                or 0
            )


class CustomerPhaseCTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.main_database_path = self.root / "DuLieuV3.db"
        self.repository = CustomerRepository(self.main_database_path)
        self.summary_repository = SummaryRepository(self.main_database_path)
        self._seed_phase_c_fixture()

    def tearDown(self) -> None:
        _drain_customer_query_threads()
        self.temporary_directory.cleanup()

    def test_customer_movement_uses_unit_directory(self) -> None:
        self._insert_period_summary("2026-03", "6501001", "001", "Khach A", 1000, branch_code="6501")
        self._insert_period_summary("2026-04", "6501001", "001", "Khach A", 1500, branch_code="6501")
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh cũ", short_name="CN cũ")
        )
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh mới", short_name="CN mới")
        )
        output = self.root / "customer-movement.xlsx"

        export_customer_growth(
            self.repository,
            "2026-03",
            "2026-04",
            CustomerFilters(branch_code="6501"),
            output,
        )
        workbook = load_workbook(output, data_only=True)
        try:
            movement_sheet = workbook["BienDongDuNo"]
            metadata_sheet = workbook["ThongTinLoc"]
            self.assertEqual(movement_sheet.cell(2, 4).value, "6501 - CN mới")
            metadata = {
                metadata_sheet.cell(row, 1).value: metadata_sheet.cell(row, 2).value
                for row in range(2, metadata_sheet.max_row + 1)
            }
            self.assertEqual(metadata["Chi nhánh"], "6501 - CN mới")
        finally:
            workbook.close()

    def test_cross_branch_detail_uses_unit_directory(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh cũ", short_name="CN cũ")
        )
        self.repository.unit_directory.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="6501",
                trctcd="03",
                office_code="6501-03",
                office_name="Phòng giao dịch cũ",
                short_name="PGD cũ",
                office_type=TRANSACTION_OFFICE,
            )
        )
        self._insert_customer_with_offices(sequence="U03", branch_code="6501", offices=(("03", 1000),))
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh mới", short_name="CN mới")
        )
        self.repository.unit_directory.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="6501",
                trctcd="03",
                office_code="6501-03",
                office_name="Phòng giao dịch mới",
                short_name="PGD mới",
                office_type=TRANSACTION_OFFICE,
            )
        )

        detail = self.repository.get_cross_branch_customer_detail("2026-04", "U03")

        self.assertEqual(detail[0]["branch_name"], "6501 - CN mới")
        self.assertEqual(detail[0]["office_name"], "PGD mới")
        self.assertEqual(detail[0]["office_display"], "6501-03 - PGD mới")
        output = self.root / "cross-branch-detail.xlsx"
        export_cross_branch_customer_detail(
            self.repository,
            "U03",
            output,
            report_period="2026-04",
            branch_code="6501",
            office_code="6501-03",
        )
        workbook = load_workbook(output, data_only=True)
        try:
            detail_sheet = workbook["ChiTietTheoDonVi"]
            overview = workbook["TongQuanLienChiNhanh"]
            metadata = {
                overview.cell(row, 1).value: overview.cell(row, 2).value
                for row in range(2, overview.max_row + 1)
            }
            self.assertEqual(metadata["Chi nhánh lọc"], "6501 - CN mới")
            self.assertEqual(metadata["PGD/Đơn vị lọc"], "6501-03 - PGD mới")
            self.assertEqual(detail_sheet.cell(2, 4).value, "6501 - CN mới")
            self.assertEqual(detail_sheet.cell(2, 7).value, "PGD mới")
        finally:
            workbook.close()

    def test_all_exports_use_unit_directory(self) -> None:
        self._insert_period_summary("2026-03", "6501002", "002", "Khach B", 1000, branch_code="6501")
        self._insert_period_summary(
            "2026-04",
            "6501002",
            "002",
            "Khach B",
            2000,
            branch_code="6501",
            officer_rows=(("A01", "Officer A", 1500), ("A02", "Officer B", 500)),
        )
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh export", short_name="CN export")
        )
        output = self.root / "customer-all.xlsx"

        export_all_customer_sheets(
            self.repository,
            CustomerFilters(branch_code="6501"),
            output,
            previous_period="2026-03",
            current_period="2026-04",
        )
        workbook = load_workbook(output, data_only=True)
        try:
            self.assertEqual(workbook["BienDongDuNo"].cell(2, 4).value, "6501 - CN export")
            self.assertEqual(workbook["NhieuCanBoQuanLy"].cell(2, 5).value, "6501 - CN export")
            officer_branches = [
                workbook["DanhMucCanBo"].cell(row, 3).value
                for row in range(2, workbook["DanhMucCanBo"].max_row + 1)
            ]
            self.assertIn("6501 - CN export", officer_branches)
        finally:
            workbook.close()

    def test_customer_management_window_opens(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self.assertEqual(window.windowTitle(), "Quản lý dữ liệu khách hàng - AgribankV3")
        self.assertEqual(window.tabs.count(), 7)
        window.close()

    def test_customer_maintenance_window_opens(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        dialog = CustomerMaintenanceDialog(window.repository, window)
        self.assertEqual(dialog.windowTitle(), "Bảo trì dữ liệu khách hàng - AgribankV3")
        self.assertIn("Customer.db", dialog.status_labels["database_path"].text())
        dialog.close()
        window.close()

    def test_customer_maintenance_statistics(self) -> None:
        status = self.repository.maintenance_status()
        self.assertEqual(status.period_count, 2)
        self.assertEqual(status.first_period, "2026-03")
        self.assertEqual(status.last_period, "2026-04")
        self.assertEqual(status.import_run_count, 2)
        self.assertGreater(status.period_summary_count, 0)
        self.assertGreater(status.page_count, 0)

    def test_customer_database_diagnostics_reports_sqlite_storage(self) -> None:
        diagnostics = self.repository.database_diagnostics()

        self.assertEqual(diagnostics["journal_mode"], "wal")
        self.assertGreater(int(diagnostics["page_size"]), 0)
        self.assertGreater(int(diagnostics["page_count"]), 0)
        self.assertIn("customer_office_period", diagnostics["table_counts"])
        self.assertIn("customer_officer_override", diagnostics["retained_table_counts"])

    def test_customer_quick_check(self) -> None:
        result = self.repository.check_database()
        self.assertTrue(result["ok"])
        self.assertEqual(result["mode"], "quick_check")

    def test_customer_optimize(self) -> None:
        result = self.repository.optimize_database(vacuum=False)
        self.assertFalse(result["vacuum"])
        self.assertIn("before_size_bytes", result)
        self.assertIn("after_size_bytes", result)

    def test_customer_vacuum_preserves_data(self) -> None:
        before = self.repository.query_customer_list(CustomerFilters(current_period="2026-04")).total_rows
        self.repository.optimize_database(vacuum=True)
        after = self.repository.query_customer_list(CustomerFilters(current_period="2026-04")).total_rows
        self.assertEqual(after, before)

    def test_customer_vacuum_reports_size_before_after(self) -> None:
        result = self.repository.optimize_database(vacuum=True)
        self.assertIsInstance(result["before_size_bytes"], int)
        self.assertIsInstance(result["after_size_bytes"], int)
        self.assertIsInstance(result["recovered_bytes"], int)

    def test_customer_vacuum_creates_backup(self) -> None:
        result = self.repository.optimize_database(vacuum=True)
        backup_path = Path(str(result["backup_path"]))
        self.assertTrue(backup_path.is_file())
        with zipfile.ZipFile(backup_path) as archive:
            self.assertIn(CUSTOMER_DATABASE_NAME, archive.namelist())

    def test_customer_maintenance_runs_outside_ui_thread(self) -> None:
        _controller, payload = self._worker_probe(
            "customer_maintenance",
            lambda: self.repository.check_database()["ok"],
        )
        self.assertTrue(payload["value"])

    def test_customer_maintenance_prevents_concurrent_write(self) -> None:
        with CustomerDatabaseOperationLock("thu hồi dung lượng Customer.db"):
            with self.assertRaisesRegex(Exception, "Customer.db đang"):
                self.repository.upsert_officer_directory(
                    officer_code="999",
                    officer_name="Officer Locked",
                )

    def test_customer_maintenance_refreshes_statistics(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = CustomerMaintenanceDialog(self.repository)
        before = dialog.status_labels["officer_directory_count"].text()
        self.repository.upsert_officer_directory(officer_code="777", officer_name="Officer Refresh")
        dialog.reload()
        after = dialog.status_labels["officer_directory_count"].text()
        self.assertNotEqual(after, before)
        dialog.close()

    def test_customer_management_window_is_non_modal(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self.assertFalse(window.isModal())
        window.close()

    def test_customer_management_default_size_fits_available_screen(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        available = window.screen().availableGeometry()
        self.assertLessEqual(window.width(), available.width())
        self.assertLessEqual(window.height(), available.height())
        window.close()

    def test_customer_detail_default_size_fits_available_screen(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerDetailWindow(self.repository, "5491001", period="2026-04")
        available = window.screen().availableGeometry()
        self.assertLessEqual(window.width(), available.width())
        self.assertLessEqual(window.height(), available.height())
        window.close()

    def test_customer_management_minimum_width_not_based_on_table_columns(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        total_configured_width = sum(window.list_tab.table.columnWidth(index) for index in range(window.list_tab.table.model().columnCount()))
        self.assertLess(window.minimumWidth(), total_configured_width)
        self.assertLessEqual(window.minimumWidth(), window.screen().availableGeometry().width())
        window.close()

    def test_customer_detail_minimum_width_not_excessive(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerDetailWindow(self.repository, "5491001", period="2026-04")
        self.assertLessEqual(window.minimumWidth(), min(850, window.screen().availableGeometry().width()))
        window.close()

    def test_customer_detail_officer_double_click_uses_clicked_period_primary_import(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerDetailWindow(self.repository, "5491001", period="2026-04")
        window.query_controller.cancel_pending()
        window.wait_for_queries()
        captured: dict[str, object] = {}

        class FakeOfficerDialog:
            def __init__(self, repository, **kwargs) -> None:
                captured.update(kwargs)

            def exec(self):
                return QDialog.DialogCode.Rejected

        window.officer_model.set_rows(
            [
                {
                    "period": "2026-03",
                    "imported_officer_code": "ROW_SECONDARY",
                    "imported_officer_name": "Row Secondary",
                }
            ]
        )
        with patch("agribank_v3.features.credit.summary.customer.customer_detail_window.OfficerOverrideDialog", FakeOfficerDialog):
            window._officer_row_double_clicked(window.officer_model.index(0, 0))
        self.assertEqual(captured["period"], "2026-03")
        self.assertEqual(captured["imported_officer_code"], "001")
        self.assertEqual(captured["imported_officer_name"], "Officer A")
        window.close()

    def test_customer_management_table_allows_horizontal_scroll(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self.assertEqual(window.list_tab.table.horizontalScrollBarPolicy(), Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.assertEqual(window.list_tab.table.horizontalScrollMode(), QAbstractItemView.ScrollMode.ScrollPerPixel)
        window.close()

    def test_customer_tables_use_shared_selection_style(self) -> None:
        table = CustomerTableView()
        self.assertEqual(table.styleSheet(), AGRIBANK_TABLE_SELECTION_STYLE)
        table.deleteLater()

    def test_selected_row_text_has_readable_contrast(self) -> None:
        self.assertIn("background-color: rgba(174, 28, 63, 38)", AGRIBANK_TABLE_SELECTION_STYLE)
        self.assertIn("color: #202020", AGRIBANK_TABLE_SELECTION_STYLE)
        self.assertNotIn("color: white", AGRIBANK_TABLE_SELECTION_STYLE.casefold())

    def test_selected_row_style_applied_to_detail_tables(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerDetailWindow(self.repository, "5491001", period="2026-04")
        tables = window.findChildren(CustomerTableView)
        self.assertTrue(tables)
        self.assertTrue(all(table.styleSheet() == AGRIBANK_TABLE_SELECTION_STYLE for table in tables))
        window.close()

    def test_hover_style_does_not_override_selection(self) -> None:
        hover_index = AGRIBANK_TABLE_SELECTION_STYLE.index("item:hover")
        selected_index = AGRIBANK_TABLE_SELECTION_STYLE.index("item:selected")
        self.assertLess(hover_index, selected_index)
        self.assertIn("item:selected:active", AGRIBANK_TABLE_SELECTION_STYLE)

    def test_table_selection_uses_full_rows(self) -> None:
        table = CustomerTableView()
        self.assertEqual(table.selectionBehavior(), QAbstractItemView.SelectionBehavior.SelectRows)
        self.assertEqual(table.selectionMode(), QAbstractItemView.SelectionMode.SingleSelection)
        table.deleteLater()

    def test_multiple_officers_long_text_does_not_expand_window(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.multiple_tab.model.set_rows(
            [
                {
                    "period": "2026-04",
                    "customer_code": "5491001",
                    "customer_name": "Khach hang A",
                    "officer_list": "Officer A\n" * 30,
                }
            ]
        )
        self.assertLessEqual(window.multiple_tab.table.columnWidth(9), 300)
        self.assertLessEqual(window.minimumWidth(), window.screen().availableGeometry().width())
        window.close()

    def test_customer_filter_bar_wraps_or_uses_two_rows(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        panel = window.findChild(QWidget, "CustomerFilterPanel")
        self.assertIsNotNone(panel)
        self.assertGreaterEqual(panel.layout().count(), 2)
        window.close()

    def test_saved_geometry_outside_screen_is_recovered(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        available = window.screen().availableGeometry()
        recovered = ensure_geometry_visible(window, QRect(available.right() + 500, available.bottom() + 500, 900, 650))
        self.assertTrue(available.contains(recovered.center()))
        window.close()

    def test_saved_geometry_larger_than_screen_is_clamped(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        available = window.screen().availableGeometry()
        recovered = ensure_geometry_visible(window, QRect(available.left(), available.top(), available.width() * 2, available.height() * 2))
        self.assertLessEqual(recovered.width(), available.width())
        self.assertLessEqual(recovered.height(), available.height())
        window.close()

    def test_customer_window_centered_on_available_screen(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        available = window.screen().availableGeometry()
        center_delta = window.geometry().center() - available.center()
        self.assertLessEqual(abs(center_delta.x()), 2)
        self.assertLessEqual(abs(center_delta.y()), 2)
        window.close()

    def test_credit_menu_contains_customer_data_card(self) -> None:
        titles = [feature.title for feature in SECTIONS["Tín dụng"]]
        self.assertIn(CUSTOMER_DATA_TITLE, titles)
        self.assertEqual(CUSTOMER_DATA_ROUTE, "credit.customer_data")

    def test_customer_data_card_route_opens_management_window(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = MainWindow()
        window.settings_database.database_path = self.main_database_path
        window.open_feature(CUSTOMER_DATA_TITLE)
        self.assertIsNotNone(window._customer_management_window)
        window._customer_management_window.close()
        window.close()

    def test_nim_dn_customer_button_uses_same_open_function(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = NimTab(self.summary_repository, SummaryDataType.NIM_DN)
        with patch("agribank_v3.features.credit.summary.customer.window_controller.open_customer_management_window") as opener:
            tab.open_customer_management()
        self.assertTrue(opener.called)
        tab.close()

    def test_customer_window_single_instance(self) -> None:
        app = QApplication.instance() or QApplication([])
        host = QWidget()
        host._customer_management_window = None
        first = open_customer_management_window(host, self.main_database_path)
        second = open_customer_management_window(host, self.main_database_path)
        self.assertIs(first, second)
        first.close()
        host.close()

    def test_existing_customer_window_is_raised(self) -> None:
        app = QApplication.instance() or QApplication([])
        host = QWidget()
        host._customer_management_window = None
        first = open_customer_management_window(host, self.main_database_path)
        second = open_customer_management_window(host, self.main_database_path)
        self.assertIs(first, second)
        self.assertTrue(second.isVisible())
        first.close()
        host.close()

    def test_customer_window_restores_from_minimized(self) -> None:
        app = QApplication.instance() or QApplication([])
        host = QWidget()
        host._customer_management_window = None
        window = open_customer_management_window(host, self.main_database_path)
        window.showMinimized()
        restored = open_customer_management_window(host, self.main_database_path)
        self.assertFalse(restored.windowState() & Qt.WindowState.WindowMinimized)
        window.close()
        host.close()

    def test_customer_empty_database_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        empty_root = self.root / "empty_customer"
        empty_root.mkdir()
        window = CustomerManagementWindow(empty_root / "DuLieuV3.db")
        self.assertFalse(window.empty_data_banner.isHidden())
        window.close()

    def test_customer_empty_state_can_open_nim_dn(self) -> None:
        app = QApplication.instance() or QApplication([])
        empty_root = self.root / "empty_customer_signal"
        empty_root.mkdir()
        window = CustomerManagementWindow(empty_root / "DuLieuV3.db")
        triggered = {"value": False}
        window.openNimDnRequested.connect(lambda: triggered.__setitem__("value", True))
        button = next(button for button in window.empty_data_banner.findChildren(QPushButton) if button.text() == "Mở NIM dư nợ")
        button.click()
        self.assertTrue(triggered["value"])
        window.close()

    def test_customer_window_opens_with_empty_database(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty"))

        self.assertFalse(window.empty_data_banner.isHidden())
        self.assertEqual(window.dashboard_tab.state_banner.state, "empty")
        self.assertTrue(window.tabs.isEnabled())
        window.close()

    def test_empty_database_does_not_start_dashboard_queries(self) -> None:
        app = QApplication.instance() or QApplication([])
        blocked_methods = (
            "get_dashboard_kpis",
            "get_total_balance_trend",
            "get_customer_metric_trend",
            "get_active_customer_count_trend",
            "get_top_customers_by_balance",
            "get_top_customer_movements",
        )
        patchers = [
            patch.object(CustomerRepository, name, side_effect=AssertionError(f"{name} should not run"))
            for name in blocked_methods
        ]
        for patcher in patchers:
            patcher.start()
        try:
            window = CustomerManagementWindow(self._empty_main_database_path("window_no_dashboard_queries"))
            self.assertEqual(window.dashboard_tab.state_banner.state, "empty")
            window.close()
        finally:
            for patcher in reversed(patchers):
                patcher.stop()

    def test_empty_database_shows_empty_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_state"))

        self.assertFalse(window.empty_data_banner.isHidden())
        self.assertEqual(window.dashboard_tab.state_banner.label.text(), "Chưa có dữ liệu khách hàng.")
        window.close()

    def test_empty_database_kpis_show_dash(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_kpis"))

        values = [card.value_label.text() for card in window.dashboard_tab.metrics._main_cards]
        self.assertTrue(values)
        self.assertEqual(set(values), {"—"})
        window.close()

    def test_empty_database_charts_are_cleared(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_charts"))

        for chart in window.dashboard_tab._chart_widgets:
            self.assertEqual(chart.state, "empty")
            self.assertEqual(chart.chart.series(), [])
            self.assertEqual(chart.chart.axes(), [])
        window.close()

    def test_empty_database_tables_are_cleared(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_tables"))

        self.assertEqual(window.dashboard_tab.top_balance_model.rowCount(), 0)
        self.assertEqual(window.dashboard_tab.top_movement_model.rowCount(), 0)
        self.assertEqual(window.list_tab.model.rowCount(), 0)
        self.assertEqual(window.movement_tab.model.rowCount(), 0)
        self.assertEqual(window.multiple_tab.model.rowCount(), 0)
        self.assertEqual(window.cross_branch_tab.model.rowCount(), 0)
        window.close()

    def test_empty_database_export_disabled(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_export"))

        self.assertFalse(window.export_all_button.isEnabled())
        self.assertFalse(window.dashboard_tab.export_button.isEnabled())
        self.assertFalse(window.list_tab.export_button.isEnabled())
        self.assertFalse(window.movement_tab.export_button.isEnabled())
        self.assertFalse(window.multiple_tab.export_button.isEnabled())
        self.assertFalse(window.cross_branch_tab.export_button.isEnabled())
        window.close()

    def test_empty_database_delete_disabled(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_delete"))

        self.assertFalse(window.delete_period_button.isEnabled())
        self.assertIn("Chưa có dữ liệu kỳ", window.delete_period_button.toolTip())
        window.close()

    def test_empty_database_maintenance_enabled(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_maintenance"))

        self.assertTrue(window.maintenance_button.isEnabled())
        window.close()

    def test_empty_database_tabs_remain_enabled(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_tabs"))

        self.assertTrue(window.tabs.isEnabled())
        for index in range(window.tabs.count()):
            self.assertTrue(window.tabs.isTabEnabled(index))
            window.tabs.setCurrentIndex(index)
            self.assertEqual(window.tabs.currentIndex(), index)
        window.close()

    def test_delete_last_period_enters_empty_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()

        self._delete_all_customer_periods(window.repository)
        window.refresh_all()

        self.assertFalse(window.empty_data_banner.isHidden())
        self.assertEqual(window.dashboard_tab.state_banner.state, "empty")
        window.close()

    def test_delete_last_period_resets_period_combos(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()

        self._delete_all_customer_periods(window.repository)
        window.refresh_all()

        for combo in (window.period_from_combo, window.period_to_combo, window.current_period_combo):
            self.assertEqual(combo.count(), 1)
            self.assertEqual(combo.itemText(0), "Chưa có dữ liệu")
            self.assertFalse(combo.isEnabled())
        window.close()

    def test_delete_last_period_removes_generated_customer_master(self) -> None:
        self._delete_all_customer_periods(self.repository)

        with closing(self.repository.connect()) as connection:
            master_count = int(connection.execute("SELECT COUNT(*) FROM customer_master").fetchone()[0])
            period_count = int(connection.execute("SELECT COUNT(*) FROM customer_period_summary").fetchone()[0])

        self.assertEqual(period_count, 0)
        self.assertEqual(master_count, 0)

    def test_delete_last_period_recommends_vacuum_without_running_it(self) -> None:
        info = self.repository.delete_customer_period("2026-04")
        self.assertFalse(info["vacuum_recommended"])

        info = self.repository.delete_customer_period("2026-03")
        diagnostics = self.repository.database_diagnostics()

        self.assertTrue(info["vacuum_recommended"])
        self.assertTrue(diagnostics["vacuum_recommended"])
        self.assertEqual(diagnostics["table_counts"]["customer_period_summary"], 0)

    def test_delete_last_period_invalidates_cache(self) -> None:
        app = QApplication.instance() or QApplication([])
        self._delete_all_customer_periods(self.repository)
        window = CustomerManagementWindow(self.main_database_path)
        window.dashboard_tab.query_controller.cache.set(("old",), {"value": 1})

        window.refresh_all()

        self.assertEqual(len(window.dashboard_tab.query_controller.cache), 0)
        window.close()

    def test_delete_last_period_invalidates_old_workers(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()
        controller = window.dashboard_tab.query_controller
        generation_before = controller.generation

        window.handle_customer_data_became_empty()

        self.assertGreater(controller.generation, generation_before)
        window.close()

    def test_old_worker_result_not_rendered_after_delete(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()
        controller = window.dashboard_tab.query_controller
        controller.run(
            "old_dashboard",
            lambda: (sleep(0.08), "old")[1],
            lambda _payload: window.dashboard_tab.state_banner.clear(),
            use_cache=False,
        )

        window.handle_customer_data_became_empty()

        self.assertTrue(_wait_until(lambda: controller.stale_result_count >= 1))
        self.assertEqual(window.dashboard_tab.state_banner.state, "empty")
        window.close()

    def test_delete_last_period_does_not_refresh_with_old_period(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()
        self._delete_all_customer_periods(window.repository)

        with patch.object(window.dashboard_tab, "refresh", side_effect=AssertionError("Dashboard refresh should not run")) as refresh:
            window.refresh_all()

        self.assertFalse(refresh.called)
        window.close()

    def test_delete_last_period_window_remains_responsive(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()

        self._delete_all_customer_periods(window.repository)
        window.refresh_all()
        window.tabs.setCurrentWidget(window.officer_tab)
        app.processEvents()

        self.assertIs(window.tabs.currentWidget(), window.officer_tab)
        window.close()

    def test_clear_period_combos_does_not_trigger_refresh_loop(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_signal_loop"))
        emitted = {"count": 0}
        window.period_from_combo.currentIndexChanged.connect(lambda _index: emitted.__setitem__("count", emitted["count"] + 1))

        window._filter_timer.stop()
        window._reset_period_combos_empty()
        app.processEvents()

        self.assertEqual(emitted["count"], 0)
        self.assertFalse(window._filter_timer.isActive())
        window.close()

    def test_linked_filters_use_signal_blocker(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_signal_blocker"))
        emitted = {"count": 0}
        for combo in (window.period_from_combo, window.period_to_combo, window.current_period_combo):
            combo.currentIndexChanged.connect(lambda _index: emitted.__setitem__("count", emitted["count"] + 1))

        window._reset_period_combos_empty()
        app.processEvents()

        self.assertEqual(emitted["count"], 0)
        window.close()

    def test_empty_filters_do_not_query_repository(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_empty_filters"))
        with patch.object(window.dashboard_tab.query_controller, "run", side_effect=AssertionError("query should not run")):
            window._apply_filter_changed()

        self.assertEqual(window.dashboard_tab.state_banner.state, "empty")
        window.close()

    def test_filter_refresh_called_once(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        calls = {"count": 0}
        original = window.refresh_filters

        def counted_refresh_filters():
            calls["count"] += 1
            return original()

        window.refresh_filters = counted_refresh_filters
        for _index in range(5):
            window._filter_changed()

        self.assertTrue(_wait_until(lambda: calls["count"] == 1, timeout=1.5))
        self.assertEqual(calls["count"], 1)
        window.close()

    def test_report_period_none_handled_safely(self) -> None:
        validation = validate_dashboard_period_filters([], CustomerFilters())

        self.assertFalse(validation.valid)
        self.assertTrue(validation.no_data)
        self.assertEqual(validation.report_period, "")

    def test_worker_error_always_emits_finished(self) -> None:
        controller = AsyncQueryController()
        errors: list[str] = []
        finished: list[bool] = []
        with self.assertLogs("agribank_v3.features.credit.summary.customer.async_query", level="ERROR"):
            controller.run(
                "boom",
                lambda: (sleep(0.05), (_ for _ in ()).throw(RuntimeError("boom")))[1],
                lambda _payload: None,
                lambda exc: errors.append(str(exc)),
                use_cache=False,
            )
            controller._threads[-1].finished.connect(lambda: finished.append(True))

            self.assertTrue(_wait_until(lambda: bool(errors) and bool(finished)))
            controller.wait_for_idle()
        self.assertEqual(errors, ["boom"])

    def test_stale_worker_releases_loading_state(self) -> None:
        banner = QueryStateBanner()
        controller = AsyncQueryController()
        controller.run(
            "stale",
            lambda: (sleep(0.08), "old")[1],
            lambda _payload: banner.clear(),
            use_cache=False,
            state_callback=lambda state, _message: banner.set_loading() if state == "loading" else None,
        )

        controller.cancel_pending()
        banner.set_empty("Chưa có dữ liệu khách hàng.")

        self.assertTrue(_wait_until(lambda: controller.stale_result_count >= 1))
        controller.wait_for_idle()
        self.assertEqual(banner.state, "empty")
        banner.deleteLater()

    def test_loading_state_ends_on_empty_result(self) -> None:
        banner = QueryStateBanner()
        controller = AsyncQueryController()
        controller.run(
            "empty",
            lambda: [],
            lambda rows: banner.set_empty("Không có dữ liệu.") if not rows else banner.clear(),
            use_cache=False,
            state_callback=lambda state, _message: banner.set_loading() if state == "loading" else None,
        )

        self.assertTrue(_wait_until(lambda: banner.state == "empty"))
        controller.wait_for_idle()
        banner.deleteLater()

    def test_loading_state_ends_on_exception(self) -> None:
        banner = QueryStateBanner()
        controller = AsyncQueryController()
        with self.assertLogs("agribank_v3.features.credit.summary.customer.async_query", level="ERROR"):
            controller.run(
                "error",
                lambda: (_ for _ in ()).throw(RuntimeError("broken")),
                lambda _payload: None,
                lambda _exc: None,
                use_cache=False,
                state_callback=lambda state, message: banner.set_loading()
                if state == "loading"
                else banner.set_error(message)
                if state == "error"
                else None,
            )

            self.assertTrue(_wait_until(lambda: banner.state == "error"))
            controller.wait_for_idle()
        banner.deleteLater()

    def test_ui_thread_does_not_call_thread_wait(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self._empty_main_database_path("window_no_wait_close"))

        with patch.object(AsyncQueryController, "wait_for_idle", side_effect=AssertionError("UI close must not wait")):
            window.close()
            app.processEvents()

    def test_window_can_close_while_worker_running(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()
        window.dashboard_tab.query_controller.run(
            "slow_close",
            lambda: (sleep(0.15), "done")[1],
            lambda _payload: None,
            use_cache=False,
        )

        with patch.object(AsyncQueryController, "wait_for_idle", side_effect=AssertionError("UI close must not wait")):
            window.close()
            app.processEvents()

    def test_empty_kpi_query(self) -> None:
        repository = CustomerRepository(self._empty_main_database_path("repo_empty_kpi"))

        metrics = repository.get_dashboard_kpis(CustomerFilters(), "")

        self.assertEqual(int(metrics["customer_count"] or 0), 0)
        self.assertEqual(float(metrics["total_balance"] or 0), 0)
        self.assertEqual(float(metrics["average_rate"] or 0), 0)

    def test_empty_trend_queries(self) -> None:
        repository = CustomerRepository(self._empty_main_database_path("repo_empty_trends"))

        self.assertEqual(repository.get_total_balance_trend(CustomerFilters(), "", ""), [])
        self.assertEqual(repository.get_customer_metric_trend(CustomerFilters(), "", ""), [])
        self.assertEqual(repository.get_active_customer_count_trend(CustomerFilters(), "", ""), [])

    def test_empty_top_queries(self) -> None:
        repository = CustomerRepository(self._empty_main_database_path("repo_empty_top"))

        self.assertEqual(repository.top_customers(CustomerFilters(), limit=10), [])
        self.assertEqual(repository.get_top_customers_by_balance(CustomerFilters(), "", 10), [])
        self.assertEqual(repository.get_top_customer_movements(CustomerFilters(), "", "", limit=10), [])

    def test_empty_cross_branch_query(self) -> None:
        repository = CustomerRepository(self._empty_main_database_path("repo_empty_cross"))

        self.assertEqual(repository.get_cross_branch_customers("", CustomerFilters()), [])
        self.assertEqual(repository.count_cross_branch_customers("", CustomerFilters()), 0)

    def test_chart_empty_dataset_does_not_calculate_min_max(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("Empty Chart", value_kind="percent")

        chart.set_series(())

        self.assertEqual(chart.state, "empty")
        self.assertEqual(chart.chart.series(), [])
        self.assertEqual(chart.chart.axes(), [])
        chart.deleteLater()

    def test_chart_can_reload_after_empty_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("Reload Chart", value_kind="percent")

        chart.set_empty("Chưa có dữ liệu để hiển thị.")
        chart.set_series((("NIM", (("2026-04", 7.2),)),))

        self.assertEqual(chart.state, "ready")
        self.assertGreater(len(chart.chart.series()), 0)
        self.assertGreater(len(chart.chart.axes()), 0)
        chart.deleteLater()

    def test_delete_closes_write_connection_before_refresh(self) -> None:
        self.repository.delete_customer_period("2026-04")

        with closing(self.repository.connect()) as connection:
            connection.execute("BEGIN IMMEDIATE")
            connection.rollback()

    def test_database_locked_does_not_retry_forever(self) -> None:
        with CustomerDatabaseOperationLock("kiểm thử khóa Customer.db"):
            with self.assertRaises(RuntimeError):
                self.repository.delete_customer_period("2026-04")

    def test_customer_connection_has_reasonable_busy_timeout(self) -> None:
        with closing(get_customer_database_connection(self.main_database_path)) as connection:
            timeout_ms = int(connection.execute("PRAGMA busy_timeout").fetchone()[0])

        self.assertEqual(timeout_ms, 10000)

    def test_all_connections_closed_after_delete(self) -> None:
        self.repository.delete_customer_period("2026-04")

        with closing(sqlite3.connect(self.repository.database_path, timeout=0.1)) as connection:
            connection.execute("BEGIN IMMEDIATE")
            connection.rollback()

    def test_import_after_empty_database_exits_empty_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        main_database_path = self._empty_main_database_path("import_after_empty")
        repository = CustomerRepository(main_database_path)
        summary_repository = SummaryRepository(main_database_path)
        import_folder = main_database_path.parent
        self._write_single_customer_ftpln(import_folder, "20260430")
        window = CustomerManagementWindow(main_database_path)
        self.assertFalse(window.empty_data_banner.isHidden())

        import_nim_dn(summary_repository, import_folder)
        window.refresh_all()

        self.assertTrue(repository.has_period_data())
        self.assertTrue(window.empty_data_banner.isHidden())
        self.assertGreaterEqual(window.current_period_combo.findData("2026-04"), 0)
        window.close()

    def test_import_after_delete_refreshes_periods(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()
        self._delete_all_customer_periods(window.repository)
        self._write_single_customer_ftpln(self.root / "reimport_periods", "20260630")

        import_nim_dn(self.summary_repository, self.root / "reimport_periods")
        window.refresh_all()

        self.assertGreaterEqual(window.current_period_combo.findData("2026-06"), 0)
        self.assertTrue(window.current_period_combo.isEnabled())
        window.close()

    def test_import_after_delete_dashboard_loads_normally(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.cancel_period_data_queries()
        self._delete_all_customer_periods(window.repository)
        self._write_single_customer_ftpln(self.root / "reimport_dashboard", "20260731")

        import_nim_dn(self.summary_repository, self.root / "reimport_dashboard")
        window.refresh_all()

        self.assertTrue(window.empty_data_banner.isHidden())
        self.assertTrue(window.dashboard_tab.export_button.isEnabled())
        self.assertNotEqual(window.dashboard_tab.state_banner.state, "empty")
        window.close()

    def test_nim_dn_result_unchanged(self) -> None:
        before = dict(self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04")))
        window = CustomerManagementWindow(self.main_database_path)
        window.handle_customer_data_became_empty()
        after = dict(self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04")))

        self.assertEqual(before, after)
        window.close()

    def test_customer_period_unique_constraint_unchanged(self) -> None:
        with closing(self.repository.connect()) as connection:
            row = connection.execute(
                """
                SELECT *
                FROM customer_period_summary
                WHERE period = '2026-04'
                LIMIT 1
                """
            ).fetchone()
            self.assertIsNotNone(row)
            with self.assertRaises(sqlite3.IntegrityError):
                connection.execute(
                    """
                    INSERT INTO customer_period_summary(
                        period, customer_code, branch_code, customer_sequence,
                        customer_name, customer_type, primary_officer_code,
                        primary_officer_name, total_balance, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'now', 'now')
                    """,
                    (
                        row["period"],
                        row["customer_code"],
                        row["branch_code"],
                        row["customer_sequence"],
                        row["customer_name"],
                        row["customer_type"],
                        row["primary_officer_code"],
                        row["primary_officer_name"],
                        row["total_balance"],
                    ),
                )
            connection.rollback()

    def test_compact_kpi_card_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        card = CompactKpiCard(KpiMetric("Tổng dư nợ", 455_866_810_772, "money"))
        self.assertLessEqual(card.sizeHint().height(), 80)
        self.assertLessEqual(card.maximumHeight(), 80)
        card.deleteLater()

    def test_compact_kpi_card_money_display(self) -> None:
        app = QApplication.instance() or QApplication([])
        card = CompactKpiCard(KpiMetric("Tổng dư nợ", 455_866_810_772, "money"))
        self.assertEqual(card.value_label.text(), "455,87 tỷ")
        self.assertEqual(compact_money_vn(127_554_800_000), "127,55 tỷ")
        card.deleteLater()

    def test_compact_kpi_card_full_value_tooltip(self) -> None:
        app = QApplication.instance() or QApplication([])
        card = CompactKpiCard(KpiMetric("Tổng dư nợ", 455_866_810_772, "money"))
        self.assertIn("Tổng dư nợ", card.toolTip())
        self.assertIn("455.866.810.772 đồng", card.toolTip())
        card.deleteLater()

    def test_compact_kpi_percentage_display(self) -> None:
        app = QApplication.instance() or QApplication([])
        card = CompactKpiCard(KpiMetric("Tỷ lệ trung/dài hạn", 27.98, "percentage"))
        self.assertEqual(card.value_label.text(), "27,98%")
        card.deleteLater()

    def test_compact_kpi_zero_value(self) -> None:
        app = QApplication.instance() or QApplication([])
        count_card = CompactKpiCard(KpiMetric("Số khách hàng còn dư nợ", 0, "count"))
        percent_card = CompactKpiCard(KpiMetric("NIM sau ĐC", 0, "percentage"))
        self.assertEqual(count_card.value_label.text(), "0")
        self.assertEqual(percent_card.value_label.text(), "0,00%")
        count_card.deleteLater()
        percent_card.deleteLater()

    def test_compact_kpi_no_data_value(self) -> None:
        app = QApplication.instance() or QApplication([])
        card = CompactKpiCard(KpiMetric("Tổng dư nợ", None, "money"))
        self.assertEqual(card.value_label.text(), "—")
        self.assertNotIn("None", card.toolTip())
        card.deleteLater()

    def test_responsive_kpi_grid_wide_window(self) -> None:
        app = QApplication.instance() or QApplication([])
        grid = ResponsiveKpiGrid()
        grid.resize(1600, 200)
        grid.set_metrics([KpiMetric(f"KPI {index}", index, "count") for index in range(8)])
        self.assertEqual(grid.main_column_count(), 8)
        grid.deleteLater()

    def test_responsive_kpi_grid_medium_window(self) -> None:
        app = QApplication.instance() or QApplication([])
        grid = ResponsiveKpiGrid()
        grid.resize(1250, 200)
        grid.set_metrics([KpiMetric(f"KPI {index}", index, "count") for index in range(8)])
        self.assertEqual(grid.main_column_count(), 6)
        grid.deleteLater()

    def test_responsive_kpi_grid_small_window(self) -> None:
        app = QApplication.instance() or QApplication([])
        grid = ResponsiveKpiGrid()
        grid.resize(900, 200)
        grid.set_metrics([KpiMetric(f"KPI {index}", index, "count") for index in range(8)])
        self.assertEqual(grid.main_column_count(), 3)
        grid.deleteLater()

    def test_kpi_grid_has_no_horizontal_scroll(self) -> None:
        app = QApplication.instance() or QApplication([])
        grid = ResponsiveKpiGrid()
        grid.set_metrics([KpiMetric(f"KPI {index}", 999_999_999_999, "money") for index in range(12)])
        self.assertEqual(grid.findChildren(QScrollArea), [])
        grid.deleteLater()

    def test_long_money_value_does_not_expand_window(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        minimum_before = window.minimumWidth()
        window.dashboard_tab.metrics.set_metrics([KpiMetric("Tổng dư nợ", 999_999_999_999_999, "money")])
        cards = window.dashboard_tab.metrics.findChildren(CompactKpiCard)
        self.assertTrue(cards)
        self.assertLessEqual(cards[0].minimumWidth(), 170)
        self.assertEqual(window.minimumWidth(), minimum_before)
        window.close()

    def test_dashboard_toolbar_compact_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        toolbar = tab.dashboard_toolbar
        self.assertIsNotNone(toolbar)
        toolbar.resize(1000, 34)
        self.assertLessEqual(toolbar.sizeHint().height(), 42)
        tab.query_controller.cancel_pending()
        tab.wait_for_queries()
        tab.close()

    def test_dashboard_export_button_not_expanding(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            self.assertNotEqual(tab.export_button.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
            tab.dashboard_toolbar.resize(1200, 40)
            tab.dashboard_toolbar.layout().activate()
            self.assertLessEqual(tab.export_button.sizeHint().width(), 150)
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_dashboard_toolbar_buttons_fit_content(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            for button in (tab.refresh_button, tab.export_button):
                self.assertNotEqual(button.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
                self.assertLessEqual(button.sizeHint().width(), 150)
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_dashboard_toolbar_has_stretch_after_buttons(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            toolbar_layout = tab.dashboard_toolbar.layout()
            self.assertEqual(toolbar_layout.indexOf(tab.refresh_button), 0)
            self.assertEqual(toolbar_layout.indexOf(tab.export_button), 1)
            self.assertIsNotNone(toolbar_layout.itemAt(2).spacerItem())
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_dashboard_export_button_height_matches_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            self.assertEqual(tab.export_button.minimumHeight(), tab.refresh_button.minimumHeight())
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_top_combo_uses_shared_compact_style(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            combos = (tab.top_balance_limit_combo, tab.top_movement_limit_combo, tab.top_movement_mode_combo)
            self.assertTrue(all(combo.objectName() == "AgribankComboBox" for combo in combos))
            self.assertTrue(all(combo.view().objectName() == "AgribankComboPopup" for combo in combos))
            self.assertTrue(all(combo.view().textElideMode() == Qt.TextElideMode.ElideNone for combo in combos))
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_top_combo_has_limits_10_20_50(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            labels = [tab.top_balance_limit_combo.itemText(index) for index in range(tab.top_balance_limit_combo.count())]
            self.assertEqual(labels, ["Top 10", "Top 20", "Top 50"])
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_top_combo_item_data_is_numeric(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            values = [tab.top_balance_limit_combo.itemData(index) for index in range(tab.top_balance_limit_combo.count())]
            self.assertEqual(values, [10, 20, 50])
            self.assertTrue(all(isinstance(value, int) for value in values))
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_top_combo_closed_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            self.assertGreaterEqual(tab.top_balance_limit_combo.minimumHeight(), 30)
            self.assertLessEqual(tab.top_balance_limit_combo.maximumHeight(), 34)
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_top_combo_popup_item_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            option = QStyleOptionViewItem()
            height = tab.top_balance_limit_combo.view().itemDelegate().sizeHint(
                option,
                tab.top_balance_limit_combo.model().index(0, 0),
            ).height()
            self.assertGreaterEqual(height, 24)
            self.assertLessEqual(height, 32)
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_top_combo_updates_table_immediately(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            with patch.object(tab.query_controller, "run") as run:
                tab.top_balance_limit_combo.setCurrentIndex(tab.top_balance_limit_combo.findData(20))
            self.assertTrue(run.called)
            self.assertEqual(run.call_args.kwargs["cache_key"][5], 20)
        finally:
            tab.close()

    def test_top_balance_and_movement_limits_independent(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        try:
            tab.top_balance_limit_combo.blockSignals(True)
            tab.top_balance_limit_combo.setCurrentIndex(tab.top_balance_limit_combo.findData(20))
            tab.top_balance_limit_combo.blockSignals(False)
            self.assertEqual(tab.top_balance_limit_combo.currentData(), 20)
            self.assertEqual(tab.top_movement_limit_combo.currentData(), 10)
            tab.top_movement_limit_combo.blockSignals(True)
            tab.top_movement_limit_combo.setCurrentIndex(tab.top_movement_limit_combo.findData(50))
            tab.top_movement_limit_combo.blockSignals(False)
            self.assertEqual(tab.top_balance_limit_combo.currentData(), 20)
            self.assertEqual(tab.top_movement_limit_combo.currentData(), 50)
        finally:
            tab.query_controller.cancel_pending()
            tab.wait_for_queries()
            tab.close()

    def test_dashboard_top_customer_widgets_are_tables(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        self.assertIsInstance(tab.top_balance_table, CustomerTableView)
        self.assertIsInstance(tab.top_movement_table, CustomerTableView)
        self.assertFalse(hasattr(tab, "top_balance_chart"))
        self.assertFalse(hasattr(tab, "top_movement_chart"))
        tab.query_controller.cancel_pending()
        tab.wait_for_queries()
        tab.close()

    def test_dashboard_top_limit_is_part_of_query_cache_key(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        with patch.object(tab.query_controller, "run") as run:
            tab.top_balance_limit_combo.setCurrentIndex(1)
        self.assertTrue(run.called)
        cache_key = run.call_args.kwargs["cache_key"]
        self.assertEqual(cache_key[5], 20)
        self.assertEqual(cache_key[6], 10)
        tab.close()

    def test_dashboard_top_balance_export_uses_visible_rows(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        path = self.root / "visible_top_balance.xlsx"
        tab.top_balance_model.set_rows(
            [
                {
                    "rank": 1,
                    "customer_code": "VISIBLE01",
                    "customer_name": "Visible Balance",
                    "total_balance": 123,
                }
            ]
        )
        with (
            patch("PySide6.QtWidgets.QFileDialog.getSaveFileName", return_value=(str(path), "Excel (*.xlsx)")),
            patch("PySide6.QtWidgets.QMessageBox.information"),
        ):
            tab.export_top_balance_excel()
        worksheet = load_workbook(path)["TopKhachHangDuNo"]
        self.assertEqual(worksheet.cell(2, 2).value, "VISIBLE01")
        tab.close()

    def test_dashboard_top_movement_export_uses_visible_rows(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        path = self.root / "visible_top_movement.xlsx"
        tab.top_movement_mode_combo.blockSignals(True)
        tab.top_movement_mode_combo.setCurrentIndex(1)
        tab.top_movement_mode_combo.blockSignals(False)
        tab.top_movement_model.set_rows(
            [
                {
                    "rank": 1,
                    "customer_code": "VISIBLE_DEC",
                    "customer_name": "Visible Movement",
                    "difference": -123,
                }
            ]
        )
        with (
            patch("PySide6.QtWidgets.QFileDialog.getSaveFileName", return_value=(str(path), "Excel (*.xlsx)")),
            patch("PySide6.QtWidgets.QMessageBox.information"),
        ):
            tab.export_top_movement_excel()
        worksheet = load_workbook(path)["TopGiamDuNo"]
        self.assertEqual(worksheet.cell(2, 2).value, "VISIBLE_DEC")
        tab.query_controller.cancel_pending()
        tab.wait_for_queries()
        tab.close()

    def test_each_chart_has_own_header_controls(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        self.assertIs(tab.balance_group_combo.parent(), tab.balance_chart)
        self.assertIs(tab.metric_combo.parent(), tab.nim_chart)
        self.assertIs(tab.balance_chart.save_button.parent(), tab.balance_chart)
        self.assertIs(tab.customer_count_chart.save_button.parent(), tab.customer_count_chart)
        tab.close()

    def test_nim_metric_combo_is_left_of_save_button(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        self.assertLess(tab.nim_chart.toolbar.indexOf(tab.metric_combo), tab.nim_chart.toolbar.indexOf(tab.nim_chart.save_button))
        tab.close()

    def test_balance_group_combo_is_left_of_save_button(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        self.assertLess(tab.balance_chart.toolbar.indexOf(tab.balance_group_combo), tab.balance_chart.toolbar.indexOf(tab.balance_chart.save_button))
        tab.close()

    def test_unused_global_detail_combo_removed(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        self.assertFalse(hasattr(tab, "movement_metric_combo"))
        tab.close()

    def test_chart_stale_request_is_ignored(self) -> None:
        controller = AsyncQueryController()
        applied: list[str] = []
        controller.run("old_chart", lambda: (sleep(0.08), "old")[1], applied.append, use_cache=False)
        controller.run("new_chart", lambda: "new", applied.append, use_cache=False)
        self.assertTrue(_wait_until(lambda: applied == ["new"]))
        controller.wait_for_idle()
        self.assertTrue(_wait_until(lambda: controller.stale_result_count >= 1))

    def test_chart_save_disabled_while_loading(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("Loading", value_kind="money")
        chart.set_series((("Tổng dư nợ", (("2026-04", 1000),)),))
        self.assertTrue(chart.save_button.isEnabled())
        chart.set_loading()
        self.assertFalse(chart.save_button.isEnabled())
        chart.deleteLater()

    def test_chart_tooltip_behavior_unchanged(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("Tooltip", value_kind="money")
        chart.resize(520, 340)
        tooltip = ChartTooltip(label="2026-04", series_name="Dư nợ", value=1000, value_kind="money")
        chart.show_tooltip(tooltip, chart.mapToGlobal(chart.rect().center()))
        chart.eventFilter(chart.chart_view, QEvent(QEvent.Type.MouseMove))
        self.assertFalse(chart.tooltip_label.isHidden())
        chart.eventFilter(chart.chart_view, QEvent(QEvent.Type.Leave))
        self.assertTrue(chart.tooltip_label.isHidden())
        chart.deleteLater()

    def test_kpi_values_unchanged_after_ui_refactor(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        tab.query_controller.cancel_pending()
        tab.wait_for_queries()
        before = dict(self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04")))
        payload = {
            "metrics": before,
            "movement": self.repository.movement_kpis("2026-03", "2026-04", CustomerFilters()),
            "trends": self.repository.dashboard_trends(CustomerFilters(current_period="2026-04")),
            "movement_metric": "count",
            "current_period": "2026-04",
            "top_balance_rows": [],
            "top_movement_rows": [],
            "top_mode": "increase",
        }
        tab._apply_payload(payload)
        after = dict(self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04")))
        total_card = next(card for card in tab.metrics.findChildren(CompactKpiCard) if card.title_label.text() == "Tổng dư nợ")
        self.assertEqual(before, after)
        self.assertIn(format_money_vn(before["total_balance"]), total_card.toolTip())
        tab.close()

    def test_dashboard_chart_gets_more_vertical_space(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        tab.query_controller.cancel_pending()
        tab.wait_for_queries()
        tab.resize(1600, 900)
        metrics = dict(self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04")))
        payload = {
            "metrics": metrics,
            "movement": self.repository.movement_kpis("2026-03", "2026-04", CustomerFilters()),
            "trends": self.repository.dashboard_trends(CustomerFilters(current_period="2026-04")),
            "movement_metric": "count",
            "current_period": "2026-04",
            "top_balance_rows": [],
            "top_movement_rows": [],
            "top_mode": "increase",
        }
        tab._apply_payload(payload)
        toolbar_height = tab.dashboard_toolbar.sizeHint().height()
        kpi_height = tab.metrics.estimated_height_for_width(1600)
        self.assertLessEqual(toolbar_height + kpi_height, 230)
        tab.close()

    def test_customer_list_filters(self) -> None:
        result = self.repository.query_customer_list(CustomerFilters(current_period="2026-04", customer_type="TC"))
        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.rows[0]["customer_code"], "5491003")

    def test_customer_list_search_by_code(self) -> None:
        result = self.repository.query_customer_list(CustomerFilters(search_text="5491001"))
        self.assertEqual({row["customer_code"] for row in result.rows}, {"5491001"})

    def test_customer_list_search_by_name(self) -> None:
        result = self.repository.query_customer_list(CustomerFilters(search_text="Khach C"))
        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.rows[0]["customer_code"], "5491003")

    def test_customer_list_pagination(self) -> None:
        result = self.repository.query_customer_list(CustomerFilters(), page=2, page_size=3)
        self.assertEqual(result.page, 2)
        self.assertEqual(len(result.rows), 3)
        self.assertGreater(result.total_rows, len(result.rows))

    def test_customer_list_count_matches_filter(self) -> None:
        result = self.repository.query_customer_list(CustomerFilters(current_period="2026-04"))
        self.assertEqual(result.total_rows, 5)

    def test_customer_list_sorting(self) -> None:
        result = self.repository.query_customer_list(
            CustomerFilters(current_period="2026-04"),
            sort_by="total_balance",
            sort_desc=True,
        )
        self.assertEqual(result.rows[0]["customer_code"], "5491006")

    def test_customer_list_vietnamese_headers(self) -> None:
        model = CustomerTableModel(CUSTOMER_LIST_COLUMNS)
        headers = [
            model.headerData(index, Qt.Orientation.Horizontal)
            for index in range(model.columnCount())
        ]
        self.assertIn("Mã khách hàng", headers)
        self.assertIn("Lãi suất bình quân", headers)

    def test_customer_money_display_uses_dot_separator(self) -> None:
        self.assertEqual(format_money_vn(1_000_000_000), "1.000.000.000")

    def test_customer_code_preserves_text(self) -> None:
        self._insert_period_summary("2026-04", "5491000123", "00123", "Khach Zero", 1)
        rows = self.repository.query_customer_list(CustomerFilters(search_text="5491000123")).rows
        self.assertEqual(rows[0]["customer_code"], "5491000123")

    def test_customer_percentage_display(self) -> None:
        self.assertEqual(format_percent_vn(2.3567), "2,36%")

    def test_customer_detail_history(self) -> None:
        history = self.repository.customer_history("5491001")
        self.assertEqual([row["period"] for row in history], ["2026-03", "2026-04"])

    def test_customer_detail_balance_trend(self) -> None:
        history = self.repository.customer_history("5491001")
        self.assertAlmostEqual(float(history[-1]["difference"]), 500)

    def test_customer_detail_term_structure(self) -> None:
        detail = self.repository.customer_detail("5491006", "2026-04")
        self.assertAlmostEqual(float(detail["medium_long_ratio"]), 25.0)

    def test_customer_detail_nim_history(self) -> None:
        history = self.repository.customer_history("5491001")
        self.assertAlmostEqual(float(history[-1]["nim_before"]), 8.0)

    def test_customer_growth_new_customer(self) -> None:
        statuses = self._movement_statuses()
        self.assertEqual(statuses["5491003"], "Vay mới")

    def test_customer_growth_paid_off(self) -> None:
        statuses = self._movement_statuses()
        self.assertEqual(statuses["5491002"], "Tất toán")

    def test_customer_growth_increase(self) -> None:
        statuses = self._movement_statuses()
        self.assertEqual(statuses["5491001"], "Tăng dư nợ")

    def test_customer_growth_decrease(self) -> None:
        statuses = self._movement_statuses()
        self.assertEqual(statuses["5491004"], "Giảm dư nợ")

    def test_customer_growth_unchanged(self) -> None:
        statuses = self._movement_statuses()
        self.assertEqual(statuses["5491005"], "Không thay đổi")

    def test_customer_growth_previous_zero(self) -> None:
        self.assertIsNone(growth_rate(0, 1000))
        row = next(row for row in self._movement_rows() if row["customer_code"] == "5491003")
        self.assertIsNone(row["growth_rate"])

    def test_customer_growth_full_outer_join_behavior(self) -> None:
        codes = {row["customer_code"] for row in self._movement_rows()}
        self.assertIn("5491002", codes)
        self.assertIn("5491003", codes)

    def test_comparison_no_longer_requires_two_files(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = MainWindow()
        window.settings_database.database_path = self.main_database_path
        with patch("agribank_v3.ui.main_window.LoanCompareWindow") as legacy_window:
            window.open_feature(LOAN_COMPARE_TITLE)
        self.assertFalse(legacy_window.called)
        self.assertIsNotNone(window._customer_management_window)
        self.assertIs(window._customer_management_window.tabs.currentWidget(), window._customer_management_window.movement_tab)
        window._customer_management_window.close()
        window.close()

    def test_movement_tab_has_no_chart(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self.assertEqual(window.movement_tab.findChildren(CustomerHorizontalBarChart), [])
        self.assertFalse(hasattr(window.movement_tab, "chart"))
        window.close()

    def test_movement_tab_has_no_save_chart_button(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        texts = [button.text() for button in window.movement_tab.findChildren(QPushButton)]
        self.assertNotIn("Lưu biểu đồ", texts)
        self.assertFalse(any(text.startswith("Top ") for text in texts))
        window.close()

    def test_movement_tab_does_not_query_top_chart_data(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        with patch.object(tab.repository, "top_movement_customers", side_effect=AssertionError("chart query")):
            payload = tab._load_payload(
                "2026-03",
                "2026-04",
                CustomerFilters(),
                1,
                100,
                "difference",
                True,
            )
        self.assertIn("kpis", payload)
        self.assertNotIn("top_rows", payload)
        tab.close()

    def test_movement_table_expands_after_chart_removed(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self.assertEqual(window.movement_tab.table.sizePolicy().verticalPolicy(), QSizePolicy.Policy.Expanding)
        self.assertGreaterEqual(window.movement_tab.layout().stretch(window.movement_tab.layout().indexOf(window.movement_tab.table)), 1)
        window.close()

    def test_movement_kpi_values_unchanged(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        before = self.repository.movement_kpis("2026-03", "2026-04", CustomerFilters())
        payload = window.movement_tab._load_payload("2026-03", "2026-04", CustomerFilters(), 1, 100, "difference", True)
        window.movement_tab._apply_payload(payload)
        after = self.repository.movement_kpis("2026-03", "2026-04", CustomerFilters())
        self.assertEqual(before, after)
        window.close()

    def test_movement_export_unchanged(self) -> None:
        path = self.root / "movement-no-chart.xlsx"
        export_customer_growth(self.repository, "2026-03", "2026-04", CustomerFilters(), path)
        ws = load_workbook(path)["BienDongDuNo"]
        self.assertEqual(ws.max_row - 1, 6)

    def test_comparison_periods_from_customer_db(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        periods = self.repository.distinct_periods()
        combo_values = [
            window.movement_tab.current_combo.itemData(index)
            for index in range(1, window.movement_tab.current_combo.count())
        ]
        self.assertEqual(combo_values, periods)
        window.close()

    def test_comparison_defaults_to_latest_two_periods(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self.assertEqual(window.movement_tab.previous_combo.currentData(), "2026-03")
        self.assertEqual(window.movement_tab.current_combo.currentData(), "2026-04")
        window.close()

    def test_comparison_uses_customer_period_summary(self) -> None:
        with closing(self.repository.connect()) as connection:
            row_count = int(connection.execute("SELECT COUNT(*) FROM customer_period_summary").fetchone()[0])
        self.assertGreater(row_count, 0)
        self.assertEqual(self.repository.movement_rows("2026-03", "2026-04", CustomerFilters()).total_rows, 6)

    def test_comparison_does_not_use_legacy_batch_table(self) -> None:
        with closing(self.repository.connect()) as connection:
            legacy = connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' AND name LIKE 'loan_compare%'"
            ).fetchall()
        self.assertEqual(legacy, [])
        self.assertGreater(self.repository.movement_rows("2026-03", "2026-04", CustomerFilters()).total_rows, 0)

    def test_comparison_new_customer(self) -> None:
        self.assertEqual(self._movement_statuses()["5491003"], "Vay mới")

    def test_comparison_paid_off_customer(self) -> None:
        self.assertEqual(self._movement_statuses()["5491002"], "Tất toán")

    def test_comparison_increase(self) -> None:
        self.assertEqual(self._movement_statuses()["5491001"], "Tăng dư nợ")

    def test_comparison_decrease(self) -> None:
        self.assertEqual(self._movement_statuses()["5491004"], "Giảm dư nợ")

    def test_comparison_unchanged(self) -> None:
        self.assertEqual(self._movement_statuses()["5491005"], "Không thay đổi")

    def test_comparison_zero_previous_balance(self) -> None:
        row = next(row for row in self._movement_rows() if row["customer_code"] == "5491003")
        self.assertIsNone(row["growth_rate"])

    def test_comparison_customer_only_in_previous_period(self) -> None:
        row = next(row for row in self._movement_rows() if row["customer_code"] == "5491002")
        self.assertEqual(row["current_balance"], 0)
        self.assertEqual(row["movement_status"], "Tất toán")

    def test_comparison_customer_only_in_current_period(self) -> None:
        row = next(row for row in self._movement_rows() if row["customer_code"] == "5491003")
        self.assertEqual(row["previous_balance"], 0)
        self.assertEqual(row["movement_status"], "Vay mới")

    def test_comparison_filters(self) -> None:
        result = self.repository.movement_rows(
            "2026-03",
            "2026-04",
            CustomerFilters(customer_type="TC"),
            sort_by="customer_code",
            sort_desc=False,
        )
        self.assertEqual({row["customer_code"] for row in result.rows}, {"5491002", "5491003"})

    def test_comparison_pagination(self) -> None:
        result = self.repository.movement_rows("2026-03", "2026-04", CustomerFilters(), page=2, page_size=2)
        self.assertEqual(result.page, 2)
        self.assertEqual(len(result.rows), 2)
        self.assertEqual(result.total_rows, 6)

    def test_comparison_kpi(self) -> None:
        kpis = self.repository.movement_kpis("2026-03", "2026-04", CustomerFilters())
        self.assertEqual(int(kpis["new_customer_count"]), 2)
        self.assertEqual(int(kpis["paid_off_customer_count"]), 1)
        self.assertEqual(int(kpis["unchanged_customer_count"]), 1)

    def test_comparison_export_all_filtered_rows(self) -> None:
        path = self.root / "comparison-filtered.xlsx"
        export_customer_growth(
            self.repository,
            "2026-03",
            "2026-04",
            CustomerFilters(customer_type="TC"),
            path,
            sort_by="customer_code",
            sort_desc=False,
        )
        ws = load_workbook(path)["BienDongDuNo"]
        self.assertEqual(ws.max_row - 1, 2)
        self.assertEqual({ws.cell(row, 1).value for row in range(2, ws.max_row + 1)}, {"5491002", "5491003"})

    def test_comparison_effective_officer_override(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            effective_to_period="2026-04",
            officer_code="OV01",
            officer_name="Officer Override",
            reason="comparison",
        )
        row = next(row for row in self._movement_rows() if row["customer_code"] == "5491001")
        self.assertEqual(row["effective_officer_code"], "OV01")
        self.assertEqual(row["effective_officer_name"], "Officer Override")

    def test_comparison_menu_opens_customer_movement_tab(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = MainWindow()
        window.settings_database.database_path = self.main_database_path
        window.open_feature(LOAN_COMPARE_TITLE)
        self.assertIs(window._customer_management_window.tabs.currentWidget(), window._customer_management_window.movement_tab)
        window._customer_management_window.close()
        window.close()

    def test_existing_customer_window_switches_to_movement_tab(self) -> None:
        app = QApplication.instance() or QApplication([])
        host = QWidget()
        host._customer_management_window = None
        window = open_customer_management_window(host, self.main_database_path)
        window.select_tab("dashboard")
        same = open_customer_management_window(host, self.main_database_path, initial_tab="movement")
        self.assertIs(window, same)
        self.assertIs(same.tabs.currentWidget(), same.movement_tab)
        window.close()
        host.close()

    def test_single_period_empty_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        self.repository.delete_customer_period("2026-03")
        window = CustomerManagementWindow(self.main_database_path)
        try:
            window.select_tab("movement")
            self.assertEqual(window.movement_tab.state_banner.state, "empty")
            self.assertIn("Cần tối thiểu hai kỳ", window.movement_tab.state_banner.label.text())
            self.assertFalse(window.movement_tab.open_nim_button.isHidden())
        finally:
            window.close()

    def test_import_new_period_refreshes_comparison_periods(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self._insert_period_summary("2026-05", "5491001", "001", "Khach A", 1700)
        self._refresh_master_rows()
        window.refresh_filters()
        self.assertGreaterEqual(window.movement_tab.current_combo.findData("2026-05"), 0)
        window.close()

    def test_delete_period_refreshes_comparison_periods(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self.repository.delete_customer_period("2026-04")
        window.refresh_all()
        self.assertEqual(window.movement_tab.current_combo.findData("2026-04"), -1)
        window.close()

    def test_restore_customer_db_refreshes_comparison(self) -> None:
        app = QApplication.instance() or QApplication([])
        backup = self.repository.backup_database(self.root / "customer-before-restore.zip")
        self._insert_period_summary("2026-05", "5491001", "001", "Khach A", 1700)
        self._refresh_master_rows()
        window = CustomerManagementWindow(self.main_database_path)
        self.repository.restore_database(backup)
        window.refresh_all()
        self.assertEqual(window.movement_tab.current_combo.findData("2026-05"), -1)
        window.close()

    def test_nim_dn_result_unchanged(self) -> None:
        before = self.summary_repository.maintenance_status()
        self.repository.movement_rows("2026-03", "2026-04", CustomerFilters())
        self.repository.optimize_database(vacuum=False)
        after = self.summary_repository.maintenance_status()
        self.assertEqual(before, after)

    def test_customer_dashboard_weighted_average_rate(self) -> None:
        metrics = self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04"))
        self.assertAlmostEqual(float(metrics["average_rate"]), 8.5)

    def test_customer_dashboard_weighted_nim_before(self) -> None:
        metrics = self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04"))
        self.assertAlmostEqual(float(metrics["nim_before"]), 6.5)

    def test_customer_dashboard_weighted_nim_after(self) -> None:
        metrics = self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04"))
        self.assertAlmostEqual(float(metrics["nim_after"]), 5.5)

    def test_customer_dashboard_weighted_medium_long_ratio(self) -> None:
        metrics = self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04"))
        self.assertAlmostEqual(float(metrics["medium_long_ratio"]), 40.0)

    def test_customer_dashboard_filters(self) -> None:
        metrics = self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04", branch_code="5491"))
        self.assertEqual(int(metrics["customer_count"]), 5)

    def test_total_customer_kpi_counts_positive_balance_only(self) -> None:
        self._insert_period_summary("2026-04", "5491098", "098", "Khach Am", -50)
        self._insert_period_summary("2026-04", "5491099", "099", "Khach Zero", 0)
        metrics = self.repository.get_dashboard_kpis(CustomerFilters(current_period="2026-04"), "2026-04")
        self.assertEqual(int(metrics["customer_count"]), 5)
        self.assertEqual(int(metrics["active_customer_count"]), 5)

    def test_total_customer_kpi_excludes_zero_balance(self) -> None:
        self._insert_period_summary("2026-04", "5491099", "099", "Khach Zero", 0)
        metrics = self.repository.get_dashboard_kpis(CustomerFilters(current_period="2026-04"), "2026-04")
        self.assertEqual(int(metrics["customer_count"]), 5)

    def test_total_customer_kpi_label_active_customers(self) -> None:
        self.assertIn("Số khách hàng còn dư nợ", DASHBOARD_MAIN_KPI_LABELS)
        self.assertNotIn("Tổng số khách hàng", DASHBOARD_MAIN_KPI_LABELS)

    def test_total_customer_kpi_uses_report_period(self) -> None:
        filters = CustomerFilters(period_from="2026-03", period_to="2026-04", current_period="2026-03")
        self.assertEqual(int(self.repository.get_dashboard_kpis(filters, "2026-03")["customer_count"]), 4)
        self.assertEqual(int(self.repository.get_dashboard_kpis(filters, "2026-04")["customer_count"]), 5)

    def test_total_customer_kpi_respects_filters(self) -> None:
        self.assertEqual(int(self.repository.get_dashboard_kpis(CustomerFilters(customer_type="TC"), "2026-04")["customer_count"]), 1)
        self.assertEqual(int(self.repository.get_dashboard_kpis(CustomerFilters(loan_term="SHORT_TERM"), "2026-04")["customer_count"]), 4)
        self.assertEqual(int(self.repository.get_dashboard_kpis(CustomerFilters(search_text="Khach C"), "2026-04")["customer_count"]), 1)

    def test_total_customer_kpi_matches_chart_report_period_point(self) -> None:
        filters = CustomerFilters(period_from="2026-03", period_to="2026-04", current_period="2026-04")
        metrics = self.repository.get_dashboard_kpis(filters, "2026-04")
        rows = self.repository.get_active_customer_count_trend(filters, "2026-03", "2026-04")
        points = {row["period"]: row["active_customer_count"] for row in rows}
        self.assertEqual(int(metrics["customer_count"]), int(points["2026-04"]))

    def test_balance_trend_returns_all_periods_in_range(self) -> None:
        self._insert_active_periods(["2026-05", "2026-06"])
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-03", period_to="2026-06"), group_by="total")
        self.assertEqual([row["period"] for row in rows], ["2026-03", "2026-04", "2026-05", "2026-06"])

    def test_metric_trend_returns_all_periods_in_range(self) -> None:
        self._insert_active_periods(["2026-05"])
        rows = self.repository.get_customer_metric_trend(CustomerFilters(period_from="2026-03", period_to="2026-05"), metric="nim_before")
        self.assertEqual([row["period"] for row in rows], ["2026-03", "2026-04", "2026-05"])

    def test_active_customer_trend_returns_all_periods_in_range(self) -> None:
        self._insert_active_periods(["2026-05"])
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-03", period_to="2026-05"))
        self.assertEqual([row["period"] for row in rows], ["2026-03", "2026-04", "2026-05"])

    def test_period_range_six_months_returns_six_points(self) -> None:
        self._insert_active_periods(["2026-01", "2026-02", "2026-05", "2026-06"])
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-01", period_to="2026-06"))
        self.assertEqual([row["period"] for row in rows], ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"])

    def test_period_range_does_not_only_return_endpoints(self) -> None:
        self._insert_active_periods(["2026-01", "2026-02", "2026-05", "2026-06"])
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-01", period_to="2026-06"), group_by="total")
        periods = [row["period"] for row in rows]
        self.assertGreater(len(periods), 2)
        self.assertIn("2026-03", periods)
        self.assertIn("2026-04", periods)

    def test_periods_sorted_ascending(self) -> None:
        self._insert_active_periods(["2026-01", "2026-06", "2026-02"])
        rows = self.repository.get_customer_metric_trend(CustomerFilters(period_from="2026-01", period_to="2026-06"), metric="average_rate")
        self.assertEqual([row["period"] for row in rows], sorted(row["period"] for row in rows))

    def test_missing_unimported_period_not_filled_with_fake_zero(self) -> None:
        self._insert_active_periods(["2026-01", "2026-06"])
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-01", period_to="2026-06"))
        self.assertEqual([row["period"] for row in rows], ["2026-01", "2026-03", "2026-04", "2026-06"])

    def test_single_period_range(self) -> None:
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), group_by="total")
        self.assertEqual([row["period"] for row in rows], ["2026-04"])

    def test_three_period_range(self) -> None:
        self._insert_active_periods(["2026-05"])
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-03", period_to="2026-05"))
        self.assertEqual([row["period"] for row in rows], ["2026-03", "2026-04", "2026-05"])

    def test_twelve_period_range(self) -> None:
        self._insert_active_periods([f"2026-{month:02d}" for month in range(1, 13)])
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-01", period_to="2026-12"))
        self.assertEqual(len(rows), 12)
        self.assertEqual(rows[0]["period"], "2026-01")
        self.assertEqual(rows[-1]["period"], "2026-12")

    def test_period_combos_use_database_periods(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        periods = self.repository.distinct_periods()
        for combo in (window.period_from_combo, window.period_to_combo, window.current_period_combo):
            combo_values = [combo.itemData(index) for index in range(1, combo.count())]
            self.assertEqual(combo_values, periods)
        window.close()

    def test_period_from_cannot_exceed_period_to(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window._set_combo_current_data(window.period_from_combo, "2026-04")
        window._set_combo_current_data(window.period_to_combo, "2026-03")
        window._normalize_period_combos(self.repository.distinct_periods())
        self.assertLessEqual(current_data(window.period_from_combo), current_data(window.period_to_combo))
        window.close()

    def test_report_period_stays_inside_range(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window._set_combo_current_data(window.period_from_combo, "2026-03")
        window._set_combo_current_data(window.period_to_combo, "2026-03")
        window._set_combo_current_data(window.current_period_combo, "2026-04")
        window._normalize_period_combos(self.repository.distinct_periods())
        self.assertEqual(current_data(window.current_period_combo), "2026-03")
        window.close()

    def test_changing_period_range_refreshes_charts(self) -> None:
        app = QApplication.instance() or QApplication([])
        holder = {"filters": CustomerFilters(period_from="2026-03", period_to="2026-04", current_period="2026-04")}
        tab = CustomerDashboardTab(self.repository, lambda: holder["filters"])
        with (
            patch.object(tab.query_controller, "run"),
            patch.object(tab.balance_chart_controller, "run") as balance_run,
            patch.object(tab.metric_chart_controller, "run") as metric_run,
            patch.object(tab.customer_count_chart_controller, "run") as count_run,
        ):
            tab.refresh()
            holder["filters"] = CustomerFilters(period_from="2026-03", period_to="2026-03", current_period="2026-03")
            tab.refresh()
        self.assertEqual(balance_run.call_count, 2)
        self.assertEqual(metric_run.call_count, 2)
        self.assertEqual(count_run.call_count, 2)
        tab.close()

    def test_changing_report_period_refreshes_kpi(self) -> None:
        app = QApplication.instance() or QApplication([])
        holder = {"filters": CustomerFilters(period_from="2026-03", period_to="2026-04", current_period="2026-03")}
        tab = CustomerDashboardTab(self.repository, lambda: holder["filters"])
        with (
            patch.object(tab.query_controller, "run") as dashboard_run,
            patch.object(tab.balance_chart_controller, "run"),
            patch.object(tab.metric_chart_controller, "run"),
            patch.object(tab.customer_count_chart_controller, "run"),
        ):
            tab.refresh()
            holder["filters"] = CustomerFilters(period_from="2026-03", period_to="2026-04", current_period="2026-04")
            tab.refresh()
        keys = [call.kwargs["cache_key"] for call in dashboard_run.call_args_list]
        self.assertNotEqual(keys[0], keys[1])
        self.assertIn("2026-03", keys[0])
        self.assertIn("2026-04", keys[1])
        tab.close()

    def test_chart_cache_key_includes_full_period_range(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(period_from="2026-03", period_to="2026-04", current_period="2026-04"))
        with patch.object(tab.balance_chart_controller, "run") as run:
            tab.refresh_balance_chart()
        cache_key = run.call_args.kwargs["cache_key"]
        self.assertEqual(cache_key[2], "2026-03")
        self.assertEqual(cache_key[3], "2026-04")
        tab.close()

    def test_kpi_cache_key_includes_report_period(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(period_from="2026-03", period_to="2026-04", current_period="2026-04"))
        with (
            patch.object(tab.query_controller, "run") as dashboard_run,
            patch.object(tab.balance_chart_controller, "run"),
            patch.object(tab.metric_chart_controller, "run"),
            patch.object(tab.customer_count_chart_controller, "run"),
        ):
            tab.refresh()
        cache_key = dashboard_run.call_args.kwargs["cache_key"]
        self.assertIn("report_period", cache_key)
        self.assertIn("2026-04", cache_key)
        tab.close()

    def test_dashboard_export_contains_all_periods(self) -> None:
        self._insert_active_periods(["2026-05", "2026-06"])
        path = self.root / "dashboard-all-periods.xlsx"
        export_customer_dashboard(self.repository, CustomerFilters(period_from="2026-03", period_to="2026-06", current_period="2026-06"), path)
        worksheet = load_workbook(path)["TongQuanKhachHang"]
        self.assertEqual(self._dashboard_export_periods(worksheet), ["2026-03", "2026-04", "2026-05", "2026-06"])

    def test_active_customer_export_contains_all_periods(self) -> None:
        self._insert_active_periods(["2026-05", "2026-06"])
        path = self.root / "dashboard-active-periods.xlsx"
        export_customer_dashboard(self.repository, CustomerFilters(period_from="2026-03", period_to="2026-06", current_period="2026-06"), path)
        worksheet = load_workbook(path)["TongQuanKhachHang"]
        header_row = self._dashboard_trend_header_row(worksheet)
        self.assertEqual(worksheet.cell(header_row, 3).value, "Số khách hàng còn dư nợ")
        values = [worksheet.cell(row, 3).value for row in range(header_row + 1, worksheet.max_row + 1)]
        self.assertEqual(len(values), 4)
        self.assertTrue(all(value and value > 0 for value in values))

    def test_export_uses_active_customer_label(self) -> None:
        path = self.root / "dashboard-active-label.xlsx"
        export_customer_dashboard(self.repository, CustomerFilters(current_period="2026-04"), path)
        worksheet = load_workbook(path)["TongQuanKhachHang"]
        labels = [worksheet.cell(row, 1).value for row in range(1, 12)]
        self.assertIn("Số khách hàng còn dư nợ", labels)
        self.assertNotIn("Tổng số khách hàng", labels)

    def test_cross_branch_tab_exists(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        labels = [window.tabs.tabText(index) for index in range(window.tabs.count())]
        self.assertIn("Khách hàng vay liên chi nhánh", labels)
        self.assertIsInstance(window.cross_branch_tab, CrossBranchCustomersTab)
        window.close()

    def test_cross_branch_scope_popup_applies_only_on_apply(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = ScopeFilterComboBox((
            ("Liên chi nhánh", "cross_branch"),
            ("Hội sở và PGD", "head_and_pgd"),
        ))
        applied = {"count": 0}
        combo.applied.connect(lambda: applied.__setitem__("count", applied["count"] + 1))

        combo._handle_pressed(combo.model().index(2, 0))
        self.assertEqual(applied["count"], 0)
        combo._handle_pressed(combo.model().index(combo.model().rowCount() - 2, 0))

        self.assertEqual(applied["count"], 1)
        combo.deleteLater()

    def test_cross_branch_scope_popup_clear_selects_all(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = ScopeFilterComboBox((
            ("Liên chi nhánh", "cross_branch"),
            ("Hội sở và PGD", "head_and_pgd"),
        ))

        combo._handle_pressed(combo.model().index(combo.model().rowCount() - 1, 0))

        self.assertEqual(set(combo.selected_values()), {"cross_branch", "head_and_pgd"})
        combo.deleteLater()

    def test_cross_branch_scope_popup_uses_agribank_style(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = ScopeFilterComboBox((("Liên chi nhánh", "cross_branch"),))

        self.assertIn("selection-background-color", combo.view().styleSheet())
        self.assertIn("174, 28, 63", combo.view().styleSheet())
        combo.deleteLater()

    def test_compact_combo_closed_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Từ 2 chi nhánh")

        self.assertGreaterEqual(combo.minimumHeight(), 30)
        self.assertLessEqual(combo.maximumHeight(), 34)
        combo.deleteLater()

    def test_compact_combo_popup_item_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Từ 2 chi nhánh")
        combo.addItem("Tất cả trường hợp liên chi nhánh", 2)
        option = QStyleOptionViewItem()
        height = combo.view().itemDelegate().sizeHint(option, combo.model().index(0, 0)).height()

        self.assertGreaterEqual(height, 24)
        self.assertLessEqual(height, 32)
        combo.deleteLater()

    def test_compact_multiselect_item_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = ScopeFilterComboBox((("Liên chi nhánh", "cross_branch"),))
        option = QStyleOptionViewItem()
        height = combo.view().itemDelegate().sizeHint(option, combo.model().index(0, 0)).height()

        self.assertGreaterEqual(height, 24)
        self.assertLessEqual(height, 32)
        combo.deleteLater()

    def test_combo_item_text_not_clipped(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Từ 2 chi nhánh", minimum_width=150, maximum_width=210)
        long_text = "Tất cả trường hợp liên chi nhánh"
        combo.addItem(long_text, 2)
        width = configure_combo_popup_width(combo, minimum_popup_width=260)

        self.assertGreaterEqual(width, combo.fontMetrics().horizontalAdvance(long_text) + 40)
        combo.deleteLater()

    def test_combo_popup_width_displays_longest_text(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Chọn", minimum_width=100, maximum_width=140)
        combo.addItem("Tất cả trường hợp liên chi nhánh", 2)

        width = configure_combo_popup_width(combo, minimum_popup_width=280)

        self.assertGreater(width, combo.minimumWidth())
        combo.deleteLater()

    def test_combo_popup_does_not_wrap_items(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Chọn")

        self.assertEqual(combo.view().textElideMode(), Qt.TextElideMode.ElideNone)
        combo.deleteLater()

    def test_combo_view_uses_uniform_item_sizes(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Chọn")

        self.assertTrue(combo.view().uniformItemSizes())
        combo.deleteLater()

    def test_combo_style_preserves_item_data(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Chọn")
        combo.addItem("Từ 3 chi nhánh", 3)

        self.assertEqual(combo.itemData(1), 3)
        combo.deleteLater()

    def test_combo_style_preserves_filter_signals(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Chọn")
        combo.addItem("Từ 3 chi nhánh", 3)
        changes = {"count": 0}
        combo.currentIndexChanged.connect(lambda _index: changes.__setitem__("count", changes["count"] + 1))

        combo.setCurrentIndex(1)

        self.assertEqual(changes["count"], 1)
        combo.deleteLater()

    def test_combo_compact_at_125_percent_scaling(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Chọn")
        font = combo.font()
        font.setPointSizeF(font.pointSizeF() * 1.25)
        combo.setFont(font)
        option = QStyleOptionViewItem()
        height = combo.view().itemDelegate().sizeHint(option, combo.model().index(0, 0)).height()

        self.assertLessEqual(height, 32)
        combo.deleteLater()

    def test_combo_compact_at_150_percent_scaling(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Chọn")
        font = combo.font()
        font.setPointSizeF(font.pointSizeF() * 1.50)
        combo.setFont(font)
        option = QStyleOptionViewItem()
        height = combo.view().itemDelegate().sizeHint(option, combo.model().index(0, 0)).height()

        self.assertLessEqual(height, 32)
        combo.deleteLater()

    def test_cross_branch_tab_filter_debounce_creates_single_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CrossBranchCustomersTab(self.repository)
        tab.query_controller.cancel_pending()
        count = {"value": 0}
        tab._filter_timer.timeout.connect(lambda: count.__setitem__("value", count["value"] + 1))

        for _index in range(5):
            tab._filter_changed()

        self.assertTrue(_wait_until(lambda: count["value"] == 1, timeout=1.5))
        self.assertEqual(count["value"], 1)
        tab.close()

    def test_cross_branch_tab_short_circuits_period_without_office_detail(self) -> None:
        app = QApplication.instance() or QApplication([])
        self._insert_period_summary("2026-05", "5400996", "996", "Khach Legacy", 1000, branch_code="5400", insert_office=False)
        self._insert_period_summary("2026-05", "5491996", "996", "Khach Legacy", 2000, branch_code="5491", insert_office=False)
        tab = CrossBranchCustomersTab(self.repository)
        tab.query_controller.cancel_pending()
        tab.period_combo.setCurrentIndex(tab.period_combo.findData("2026-05"))

        with patch.object(self.repository, "query_cross_branch_customers", side_effect=AssertionError("heavy query must not run")):
            tab.refresh()

        self.assertEqual(tab.state_banner.state, "empty")
        self.assertIn("chưa có dữ liệu chi tiết", tab.state_banner.label.text())
        tab.close()

    def test_cross_branch_detected_by_customer_sequence(self) -> None:
        self._insert_cross_branch_fixture(sequence="177616932")
        rows = self.repository.get_cross_branch_customers("2026-04")
        self.assertIn("177616932", {row["customer_sequence"] for row in rows})

    def test_cross_branch_not_grouped_by_full_customer_code(self) -> None:
        self._insert_cross_branch_fixture(sequence="177616932")
        rows = [row for row in self.repository.get_cross_branch_customers("2026-04") if row["customer_sequence"] == "177616932"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["branch_count"], 2)

    def test_same_sequence_two_branches_is_cross_branch(self) -> None:
        self._insert_cross_branch_fixture(sequence="222")
        self.assertEqual(self.repository.count_cross_branch_customers("2026-04", CustomerFilters(search_text="222")), 1)

    def test_same_sequence_one_branch_is_not_cross_branch(self) -> None:
        self._insert_period_summary("2026-04", "5400333", "333", "Khach One Branch", 1000, branch_code="5400")
        self.assertEqual(self.repository.count_cross_branch_customers("2026-04", CustomerFilters(search_text="333")), 0)

    def test_different_periods_not_treated_as_same_period_cross_branch(self) -> None:
        self._insert_period_summary("2026-03", "5400444", "444", "Khach Move", 1000, branch_code="5400")
        self._insert_period_summary("2026-04", "5491444", "444", "Khach Move", 1000, branch_code="5491")
        self.assertEqual(self.repository.count_cross_branch_customers("2026-03", CustomerFilters(search_text="444")), 0)
        self.assertEqual(self.repository.count_cross_branch_customers("2026-04", CustomerFilters(search_text="444")), 0)

    def test_only_positive_balance_branches_counted(self) -> None:
        self._insert_cross_branch_fixture(sequence="555", balances=(1000, 2000))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="555"))[0]
        self.assertEqual(row["branch_count"], 2)

    def test_zero_balance_branch_not_counted(self) -> None:
        self._insert_cross_branch_fixture(sequence="556", balances=(1000, 0))
        self.assertEqual(self.repository.count_cross_branch_customers("2026-04", CustomerFilters(search_text="556")), 0)

    def test_three_branch_customer(self) -> None:
        self._insert_cross_branch_fixture(sequence="557", branches=("5400", "5401", "5491"), balances=(1000, 2000, 3000))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="557"))[0]
        self.assertEqual(row["branch_count"], 3)

    def test_minimum_branch_count_filter(self) -> None:
        self._insert_cross_branch_fixture(sequence="558", branches=("5400", "5401", "5491"), balances=(1000, 2000, 3000))
        self._insert_cross_branch_fixture(sequence="559", branches=("5400", "5491"), balances=(1000, 2000))
        rows = self.repository.get_cross_branch_customers("2026-04", minimum_branch_count=3)
        sequences = {row["customer_sequence"] for row in rows}
        self.assertIn("558", sequences)
        self.assertNotIn("559", sequences)

    def test_cross_branch_total_balance(self) -> None:
        self._insert_cross_branch_fixture(sequence="560", balances=(1000, 3000))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="560"))[0]
        self.assertEqual(row["total_balance"], 4000)

    def test_cross_branch_short_term_balance(self) -> None:
        self._insert_cross_branch_fixture(sequence="561", balances=(1000, 3000), shorts=(700, 1300), mediums=(300, 1700))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="561"))[0]
        self.assertEqual(row["short_term_balance"], 2000)

    def test_cross_branch_medium_long_balance(self) -> None:
        self._insert_cross_branch_fixture(sequence="562", balances=(1000, 3000), shorts=(700, 1300), mediums=(300, 1700))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="562"))[0]
        self.assertEqual(row["medium_long_term_balance"], 2000)

    def test_cross_branch_weighted_average_rate(self) -> None:
        self._insert_cross_branch_fixture(sequence="563", balances=(1000, 3000), average_rates=(10, 8))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="563"))[0]
        self.assertAlmostEqual(row["average_rate"], 8.5)

    def test_cross_branch_weighted_nim_before(self) -> None:
        self._insert_cross_branch_fixture(sequence="564", balances=(1000, 3000), nim_befores=(8, 6))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="564"))[0]
        self.assertAlmostEqual(row["nim_before"], 6.5)

    def test_cross_branch_weighted_nim_after(self) -> None:
        self._insert_cross_branch_fixture(sequence="565", balances=(1000, 3000), nim_afters=(7, 5))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="565"))[0]
        self.assertAlmostEqual(row["nim_after"], 5.5)

    def test_cross_branch_medium_long_ratio(self) -> None:
        self._insert_cross_branch_fixture(sequence="566", balances=(1000, 3000), shorts=(1000, 2000), mediums=(0, 1000))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="566"))[0]
        self.assertAlmostEqual(row["medium_long_ratio"], 25.0)

    def test_cross_branch_name_conflict_flag(self) -> None:
        self._insert_cross_branch_fixture(sequence="567", names=("Cong Ty A", "Cong ty B"))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="567"))[0]
        self.assertEqual(row["name_conflict"], 1)
        self.assertIn("xung đột tên", row["conflict_status"])

    def test_cross_branch_customer_type_conflict_flag(self) -> None:
        self._insert_cross_branch_fixture(sequence="568", customer_types=("CN", "TC"))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="568"))[0]
        self.assertEqual(row["customer_type_conflict"], 1)
        self.assertEqual(row["customer_type_display"], "Không thống nhất")

    def test_cross_branch_branch_filter_uses_exists_after_grouping(self) -> None:
        self._insert_cross_branch_fixture(sequence="569", branches=("5400", "5491"))
        rows = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(branch_code="5491"))
        self.assertIn("569", {row["customer_sequence"] for row in rows})

    def test_cross_branch_branch_filter_does_not_break_having(self) -> None:
        self._insert_cross_branch_fixture(sequence="570", branches=("5400", "5491"))
        row = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(branch_code="5400", search_text="570"))[0]
        self.assertEqual(row["branch_count"], 2)

    def test_cross_branch_customer_type_filter(self) -> None:
        self._insert_cross_branch_fixture(sequence="571", customer_types=("TC", "TC"))
        rows = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(customer_type="TC"))
        self.assertIn("571", {row["customer_sequence"] for row in rows})

    def test_cross_branch_officer_filter_uses_effective_officer(self) -> None:
        self._insert_cross_branch_fixture(sequence="572", officer_codes=("A01", "A02"), officer_names=("Officer A01", "Officer A02"))
        self.repository.create_officer_override(
            customer_code="5400572",
            effective_from_period="2026-04",
            officer_code="OV572",
            officer_name="Override 572",
            reason="cross",
        )
        rows = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(officer="OV572"))
        self.assertIn("572", {row["customer_sequence"] for row in rows})

    def test_cross_branch_override_is_scoped_by_full_customer_code(self) -> None:
        self._insert_cross_branch_fixture(sequence="573", officer_codes=("A01", "A02"), officer_names=("Officer A01", "Officer A02"))
        self.repository.create_officer_override(
            customer_code="5400573",
            effective_from_period="2026-04",
            officer_code="OV573",
            officer_name="Override 573",
            reason="cross",
        )
        detail = self.repository.get_cross_branch_customer_detail("2026-04", "573")
        overrides = {row["customer_code"]: row["override_status"] for row in detail}
        self.assertEqual(overrides["5400573"], "Có override")
        self.assertEqual(overrides["5491573"], "Không override")

    def test_cross_branch_search_by_sequence(self) -> None:
        self._insert_cross_branch_fixture(sequence="574")
        rows = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="574"))
        self.assertEqual([row["customer_sequence"] for row in rows], ["574"])

    def test_cross_branch_search_by_name(self) -> None:
        self._insert_cross_branch_fixture(sequence="575", names=("Khach Lien CN Search", "Khach Lien CN Search"))
        rows = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="Lien CN Search"))
        self.assertIn("575", {row["customer_sequence"] for row in rows})

    def test_cross_branch_pagination(self) -> None:
        for sequence in ("576", "577", "578"):
            self._insert_cross_branch_fixture(sequence=sequence)
        result = self.repository.query_cross_branch_customers("2026-04", page=1, page_size=2)
        self.assertGreaterEqual(result.total_rows, 3)
        self.assertEqual(len(result.rows), 2)

    def test_cross_branch_count_query(self) -> None:
        self._insert_cross_branch_fixture(sequence="579")
        self.assertEqual(self.repository.count_cross_branch_customers("2026-04", CustomerFilters(search_text="579")), 1)

    def test_cross_branch_sort_branch_count(self) -> None:
        self._insert_cross_branch_fixture(sequence="580", branches=("5400", "5491"), balances=(1000, 1000))
        self._insert_cross_branch_fixture(sequence="581", branches=("5400", "5401", "5491"), balances=(1000, 1000, 1000))
        rows = self.repository.get_cross_branch_customers("2026-04", sort_by="branch_count", sort_order="desc")
        ordered = [row["customer_sequence"] for row in rows if row["customer_sequence"] in {"580", "581"}]
        self.assertEqual(ordered[0], "581")

    def test_cross_branch_sort_total_balance(self) -> None:
        self._insert_cross_branch_fixture(sequence="582", balances=(1000, 1000))
        self._insert_cross_branch_fixture(sequence="583", balances=(5000, 5000))
        rows = self.repository.get_cross_branch_customers("2026-04", sort_by="total_balance", sort_order="desc")
        ordered = [row["customer_sequence"] for row in rows if row["customer_sequence"] in {"582", "583"}]
        self.assertEqual(ordered[0], "583")

    def test_cross_branch_detail_rows(self) -> None:
        self._insert_cross_branch_fixture(sequence="584", branches=("5400", "5491"))
        detail = self.repository.get_cross_branch_customer_detail("2026-04", "584")
        self.assertEqual(len(detail), 2)
        self.assertEqual({row["customer_code"] for row in detail}, {"5400584", "5491584"})

    def test_cross_branch_detail_opens_standard_customer_detail(self) -> None:
        app = QApplication.instance() or QApplication([])
        self._insert_cross_branch_fixture(sequence="585", branches=("5400", "5491"))
        dialog = CrossBranchCustomerDetailDialog(self.repository, "2026-04", "585")
        dialog.query_controller.cancel_pending()
        dialog.wait_for_queries()
        dialog.detail_model.set_rows(self.repository.get_cross_branch_customer_detail("2026-04", "585"))
        captured: dict[str, object] = {}

        class FakeCustomerDetailWindow:
            def __init__(self, repository, customer_code, *, period="", parent=None) -> None:
                captured["customer_code"] = customer_code
                captured["period"] = period

            def setAttribute(self, *args):
                return None

            @property
            def finished(self):
                class Finished:
                    def connect(self, *_args):
                        return None
                return Finished()

            def show(self):
                return None

            def raise_(self):
                return None

        with patch("agribank_v3.features.credit.summary.customer.cross_branch_detail_dialog.CustomerDetailWindow", FakeCustomerDetailWindow):
            dialog._branch_row_double_clicked(dialog.detail_model.index(0, 0))
        self.assertIn(captured["customer_code"], {"5400585", "5491585"})
        self.assertEqual(captured["period"], "2026-04")
        dialog.close()

    def test_cross_branch_history(self) -> None:
        self._insert_period_summary("2026-03", "5400586", "586", "Khach History", 1000, branch_code="5400")
        self._insert_cross_branch_fixture(sequence="586", branches=("5400", "5491"), balances=(1000, 2000))
        history = self.repository.get_cross_branch_customer_history("586")
        self.assertEqual([row["period"] for row in history], ["2026-03", "2026-04"])
        self.assertEqual(history[-1]["branch_count"], 2)
        self.assertEqual(history[-1]["difference"], 2000)

    def test_cross_branch_export_all_filtered_rows(self) -> None:
        self._insert_cross_branch_fixture(sequence="587")
        path = self.root / "cross-branch.xlsx"
        export_cross_branch_customers(self.repository, "2026-04", CustomerFilters(search_text="587"), 2, path)
        workbook = load_workbook(path)
        worksheet = workbook["KhachHangLienChiNhanh"]
        self.assertEqual(worksheet.max_row - 1, 1)
        self.assertEqual(worksheet.cell(2, 3).value, "587")
        self.assertIn("ChiTietTheoDonVi", workbook.sheetnames)

    def test_cross_branch_export_customer_sequence_as_text(self) -> None:
        self._insert_cross_branch_fixture(sequence="001234")
        path = self.root / "cross-branch-text.xlsx"
        export_cross_branch_customers(self.repository, "2026-04", CustomerFilters(search_text="001234"), 2, path)
        worksheet = load_workbook(path)["KhachHangLienChiNhanh"]
        self.assertEqual(worksheet.cell(2, 3).value, "001234")
        self.assertEqual(worksheet.cell(2, 3).number_format, "@")

    def test_cross_branch_import_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        self._insert_cross_branch_fixture(period="2026-05", sequence="588")
        window.refresh_all()
        self.assertGreaterEqual(window.cross_branch_tab.period_combo.findData("2026-05"), 0)
        window.close()

    def test_cross_branch_delete_period_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        self._insert_cross_branch_fixture(period="2026-05", sequence="589")
        window = CustomerManagementWindow(self.main_database_path)
        self.repository.delete_customer_period("2026-05")
        window.refresh_all()
        self.assertEqual(window.cross_branch_tab.period_combo.findData("2026-05"), -1)
        window.close()

    def test_cross_branch_restore_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        backup = self.repository.backup_database(self.root / "cross-before-restore.zip")
        self._insert_cross_branch_fixture(period="2026-05", sequence="590")
        window = CustomerManagementWindow(self.main_database_path)
        window.wait_for_all_queries()
        self.repository.restore_database(backup)
        window.refresh_all()
        self.assertEqual(window.cross_branch_tab.period_combo.findData("2026-05"), -1)
        window.close()

    def test_cross_branch_query_plan_uses_index_if_added(self) -> None:
        self._insert_cross_branch_fixture(sequence="591")
        plan = self.repository.explain_cross_branch_query_plan("2026-04", CustomerFilters(search_text="591"))
        self.assertTrue(any("idx_customer_period_summary_period" in detail or "period" in detail.casefold() for detail in plan))

    def test_cross_branch_candidate_sql_has_no_group_concat(self) -> None:
        sql, _params = _cross_branch_candidate_sql("2026-04", CustomerFilters(search_text="591"))

        self.assertNotIn("GROUP_CONCAT", sql.upper())
        self.assertIn("customer_office_period", sql)

    def test_cross_branch_one_period_fast_path(self) -> None:
        self._insert_cross_branch_fixture(sequence="CBFAST")

        payload = self.repository.query_cross_branch_tab_payload("2026-04", CustomerFilters(current_period="2026-04"))

        self.assertGreaterEqual(payload["result"].total_rows, 1)
        self.assertIn("fast_scope_branch_count", payload["benchmark"])
        self.assertLessEqual(payload["sql_statement_count"], 14)

    def test_cross_branch_period_filter_applied_before_grouping(self) -> None:
        sql, params = _cross_branch_candidate_sql("2026-04", CustomerFilters(current_period="2026-04"))

        self.assertIn("FROM customer_office_period o", sql)
        self.assertIn("WHERE o.period = ?", sql)
        self.assertEqual(params[:2], ["2026-04", "2026-04"])

    def test_cross_branch_no_n_plus_one_queries(self) -> None:
        self._insert_cross_branch_fixture(sequence="NPLUS1A")
        self._insert_cross_branch_fixture(sequence="NPLUS1B")

        with patch.object(self.repository, "get_cross_branch_customer_detail", side_effect=AssertionError("detail must be bulk-loaded")):
            payload = self.repository.query_cross_branch_tab_payload("2026-04", page=1, page_size=2)

        self.assertEqual(len(payload["result"].rows), 2)

    def test_cross_branch_sql_statement_count_bounded(self) -> None:
        self._insert_cross_branch_fixture(sequence="SQLBOUND")

        payload = self.repository.query_cross_branch_tab_payload("2026-04", page=1, page_size=100)

        self.assertLessEqual(payload["sql_statement_count"], 14)

    def test_cross_branch_count_does_not_repeat_enrichment(self) -> None:
        self._insert_cross_branch_fixture(sequence="NOENRICH")

        with patch.object(
            self.repository,
            "_cross_branch_detail_rows_for_keys",
            side_effect=AssertionError("count must not enrich rows"),
        ):
            count = self.repository.count_cross_branch_customers("2026-04", CustomerFilters(search_text="NOENRICH"))

        self.assertEqual(count, 1)

    def test_cross_branch_kpi_has_no_group_concat(self) -> None:
        sql, _params = _cross_branch_candidate_sql("2026-04", CustomerFilters())

        self.assertNotIn("GROUP_CONCAT", sql.upper())

    def test_cross_branch_page_enrichment_only_current_page(self) -> None:
        for sequence in ("PAGEENRICH1", "PAGEENRICH2", "PAGEENRICH3"):
            self._insert_cross_branch_fixture(sequence=sequence)
        captured: dict[str, object] = {}
        original = CustomerRepository._cross_branch_detail_rows_for_keys_from_database

        def wrapper(database, keys):
            captured["keys"] = list(keys)
            return original(database, keys)

        with patch.object(CustomerRepository, "_cross_branch_detail_rows_for_keys_from_database", side_effect=wrapper):
            payload = self.repository.query_cross_branch_tab_payload("2026-04", page=1, page_size=2)

        self.assertEqual(len(payload["result"].rows), 2)
        self.assertLessEqual(len(captured["keys"]), 2)

    def test_cross_branch_default_filters_skip_unused_joins(self) -> None:
        sql, _params = _cross_branch_candidate_sql("2026-04", CustomerFilters())

        self.assertNotIn("selected_officer", sql)
        self.assertNotIn("customer_officer_override", sql)

    def test_cross_branch_officer_join_only_when_needed(self) -> None:
        default_sql, _params = _cross_branch_candidate_sql("2026-04", CustomerFilters())
        officer_sql, _params = _cross_branch_candidate_sql("2026-04", CustomerFilters(officer="001"))

        self.assertNotIn("customer_officer_override", default_sql)
        self.assertIn("customer_officer_override", officer_sql)

    def test_cross_branch_zero_result_fast_path(self) -> None:
        payload = self.repository.query_cross_branch_tab_payload(
            "2026-04",
            CustomerFilters(search_text="__NO_MATCH_FOR_CROSS_BRANCH__"),
        )

        self.assertEqual(payload["result"].total_rows, 0)
        self.assertLessEqual(payload["sql_statement_count"], 14)

    def test_cross_branch_single_worker_per_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CrossBranchCustomersTab(self.repository)
        tab.query_controller.cancel_pending()

        with patch.object(tab.query_controller, "run") as run:
            tab.refresh()

        self.assertEqual(run.call_count, 1)
        tab.close()

    def test_cross_branch_one_apply_one_request(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CrossBranchCustomersTab(self.repository)
        tab.query_controller.cancel_pending()
        tab._filter_timer.stop()

        with patch.object(tab, "refresh") as refresh:
            tab.scope_combo._handle_pressed(tab.scope_combo.model().index(2, 0))
            self.assertEqual(refresh.call_count, 0)
            tab.scope_combo._handle_pressed(tab.scope_combo.model().index(tab.scope_combo.model().rowCount() - 2, 0))
            self.assertTrue(_wait_until(lambda: refresh.call_count == 1, timeout=1.5))

        tab.close()

    def test_cross_branch_stale_request_cleanup(self) -> None:
        controller = AsyncQueryController()
        applied: list[str] = []
        controller.run("slow-cross", lambda: (sleep(0.08), "old")[1], applied.append, use_cache=False)
        controller.run("fast-cross", lambda: "new", applied.append, use_cache=False)

        self.assertTrue(_wait_until(lambda: applied == ["new"]))
        controller.wait_for_idle()
        self.assertTrue(_wait_until(lambda: controller.stale_result_count >= 1 or not controller._threads))
        self.assertFalse(controller._threads)

    def test_cross_branch_connection_created_in_worker(self) -> None:
        calls = {"count": 0}
        original = self.repository.connect

        def wrapped_connect():
            calls["count"] += 1
            return original()

        with patch.object(self.repository, "connect", side_effect=wrapped_connect):
            payload = self.repository.query_cross_branch_tab_payload("2026-04")

        self.assertGreaterEqual(calls["count"], 1)
        self.assertIsInstance(payload["result"].rows, list)

    def test_cross_branch_model_updated_on_ui_thread(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CrossBranchCustomersTab(self.repository)
        ui_thread = threading.get_ident()
        payload = self.repository.query_cross_branch_tab_payload("2026-04", page=1, page_size=2)

        tab._apply_payload(payload)

        self.assertEqual(threading.get_ident(), ui_thread)
        self.assertEqual(tab.model.rowCount(), len(payload["result"].rows))
        tab.close()

    def test_cross_branch_does_not_resize_columns_to_contents_each_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CrossBranchCustomersTab(self.repository)
        payload = self.repository.query_cross_branch_tab_payload("2026-04", page=1, page_size=2)

        with patch.object(tab.table, "resizeColumnsToContents", side_effect=AssertionError("must not auto-resize columns")):
            tab._apply_payload(payload)

        tab.close()

    def test_cross_branch_query_empty_when_period_has_no_office_detail(self) -> None:
        self._insert_period_summary("2026-05", "5400992", "992", "Khach Legacy", 1000, branch_code="5400", insert_office=False)
        self._insert_period_summary("2026-05", "5491992", "992", "Khach Legacy", 2000, branch_code="5491", insert_office=False)

        self.assertFalse(self.repository.has_office_detail_for_period("2026-05"))
        self.assertEqual(self.repository.count_cross_branch_customers("2026-05"), 0)
        self.assertEqual(self.repository.get_cross_branch_customers("2026-05"), [])
        self.assertEqual(self.repository.get_cross_branch_kpis("2026-05")["cross_customer_count"], 0)

    def test_cross_branch_multiscope_filter_uses_any_selected_scope(self) -> None:
        self._insert_cross_branch_fixture(sequence="592", branches=("5400", "5491"))
        self._insert_customer_with_offices(sequence="593", offices=(("00", 1000), ("01", 2000)))

        rows = self.repository.get_cross_branch_customers(
            "2026-04",
            CustomerFilters(search_text="59"),
            scope_type=("cross_branch", "head_and_pgd"),
        )

        self.assertIn("592", {row["customer_sequence"] for row in rows})
        self.assertIn("593", {row["customer_sequence"] for row in rows})

    def test_cross_branch_page_uses_bulk_detail_enrichment(self) -> None:
        self._insert_cross_branch_fixture(sequence="594")
        self._insert_cross_branch_fixture(sequence="595")

        with patch.object(
            self.repository,
            "_cross_branch_detail_rows_for_keys",
            wraps=self.repository._cross_branch_detail_rows_for_keys,
        ) as bulk_enrich:
            rows = self.repository.get_cross_branch_customers("2026-04", limit=2)

        self.assertEqual(len(rows), 2)
        self.assertEqual(bulk_enrich.call_count, 1)

    def test_customer_period_unique_constraint_unchanged(self) -> None:
        with closing(self.repository.connect()) as connection:
            with self.assertRaises(sqlite3.IntegrityError):
                connection.execute(
                    """
                    INSERT INTO customer_period_summary(
                        period, customer_code, branch_code, customer_sequence, customer_name,
                        customer_type, created_at, updated_at
                    )
                    VALUES ('2026-04', '5491001', '5491', '001', 'Duplicate', 'CN', 'now', 'now')
                    """
                )
                connection.commit()

    def test_same_branch_head_office_and_pgd_detected(self) -> None:
        self._insert_customer_with_offices(sequence="901", offices=(("00", 1000), ("01", 2000)))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "head_and_pgd",
            filters=CustomerFilters(search_text="901"),
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0]["has_head_and_pgd"]), 1)
        self.assertEqual(int(rows[0]["branch_count"]), 1)

    def test_same_branch_only_head_office_not_detected(self) -> None:
        self._insert_customer_with_offices(sequence="902", offices=(("00", 1000),))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "head_and_pgd",
            filters=CustomerFilters(search_text="902"),
        )

        self.assertEqual(rows, [])

    def test_same_branch_only_one_pgd_not_detected(self) -> None:
        self._insert_customer_with_offices(sequence="903", offices=(("01", 1000),))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "multi_pgd",
            filters=CustomerFilters(search_text="903"),
        )

        self.assertEqual(rows, [])

    def test_same_branch_two_pgds_detected_as_multi_pgd(self) -> None:
        self._insert_customer_with_offices(sequence="904", offices=(("01", 1000), ("02", 2000)))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "multi_pgd",
            filters=CustomerFilters(search_text="904"),
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0]["has_multi_pgd"]), 1)
        self.assertEqual(int(rows[0]["pgd_count"]), 2)

    def test_two_branches_detected_as_cross_branch(self) -> None:
        self._insert_cross_branch_fixture(sequence="905", branches=("5405", "5491"))

        rows = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="905"))

        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0]["branch_count"]), 2)

    def test_cross_branch_and_head_office_pgd_can_both_be_true(self) -> None:
        self._insert_customer_with_offices(sequence="906", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        self._insert_customer_with_offices(sequence="906", branch_code="5491", offices=(("01", 3000),))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "all_multi_unit",
            filters=CustomerFilters(search_text="906"),
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0]["branch_count"]), 2)
        self.assertEqual(int(rows[0]["has_head_and_pgd"]), 1)

    def test_zero_balance_office_not_counted(self) -> None:
        self._insert_customer_with_offices(sequence="907", offices=(("00", 1000), ("01", 0)))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "head_and_pgd",
            filters=CustomerFilters(search_text="907"),
        )
        kpis = self.repository.get_cross_branch_customer_filtered_kpis("907", "2026-04")

        self.assertEqual(rows, [])
        self.assertEqual(int(kpis["head_office_count"]), 1)
        self.assertEqual(int(kpis["pgd_count"]), 0)

    def test_different_periods_not_combined(self) -> None:
        self._insert_customer_with_offices(period="2026-03", sequence="908", branch_code="5405", offices=(("00", 1000),))
        self._insert_customer_with_offices(period="2026-04", sequence="908", branch_code="5491", offices=(("00", 1000),))

        self.assertEqual(
            self.repository.get_multi_unit_customers("2026-04", "cross_branch", filters=CustomerFilters(search_text="908")),
            [],
        )

    def test_customer_head_office_and_one_pgd_assigned_to_head_office(self) -> None:
        self._insert_customer_with_offices(sequence="909", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        rows = self.repository.get_cross_branch_customer_offices("909", "2026-04")

        representative = resolve_representative_office("2026-04", "909", "5405", rows)

        self.assertEqual(representative.representative_office_code, "5405-00")
        self.assertEqual(representative.reason, "HAS_HEAD_OFFICE")

    def test_customer_head_office_and_multiple_pgds_assigned_to_head_office(self) -> None:
        self._insert_customer_with_offices(sequence="910", branch_code="5405", offices=(("00", 1000), ("01", 2000), ("02", 3000)))
        rows = self.repository.get_cross_branch_customer_offices("910", "2026-04")

        representative = resolve_representative_office("2026-04", "910", "5405", rows)

        self.assertEqual(representative.representative_office_code, "5405-00")
        self.assertEqual(representative.reason, "HAS_HEAD_OFFICE")

    def test_customer_without_head_office_single_pgd_assigned_to_that_pgd(self) -> None:
        self._insert_customer_with_offices(sequence="911", branch_code="5405", offices=(("01", 2000),))
        rows = self.repository.get_cross_branch_customer_offices("911", "2026-04")

        representative = resolve_representative_office("2026-04", "911", "5405", rows)

        self.assertEqual(representative.representative_office_code, "5405-01")
        self.assertEqual(representative.reason, "SINGLE_PGD")

    def test_customer_without_head_office_multiple_pgds_uses_largest_balance_pgd(self) -> None:
        self._insert_customer_with_offices(sequence="912", branch_code="5405", offices=(("01", 2000), ("02", 5000), ("03", 1000)))
        rows = self.repository.get_cross_branch_customer_offices("912", "2026-04")

        representative = resolve_representative_office("2026-04", "912", "5405", rows)

        self.assertEqual(representative.representative_office_code, "5405-02")
        self.assertEqual(representative.reason, "MULTIPLE_PGD_LARGEST_BALANCE")

    def test_multiple_pgd_tie_uses_lowest_trctcd(self) -> None:
        self._insert_customer_with_offices(sequence="913", branch_code="5405", offices=(("02", 5000), ("01", 5000)))
        rows = self.repository.get_cross_branch_customer_offices("913", "2026-04")

        representative = resolve_representative_office("2026-04", "913", "5405", rows)

        self.assertEqual(representative.representative_office_code, "5405-01")

    def test_customer_with_head_office_counted_once_in_branch(self) -> None:
        self._insert_customer_with_offices(sequence="914", branch_code="5405", offices=(("00", 1000), ("01", 2000), ("02", 3000)))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "all_multi_unit",
            branch_code="5405",
            filters=CustomerFilters(search_text="914"),
            office_filter_mode="representative",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0]["branch_count"]), 1)
        self.assertEqual(int(rows[0]["office_count"]), 3)

    def test_customer_with_head_office_not_counted_again_for_pgds(self) -> None:
        self._insert_customer_with_offices(sequence="915", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "all_multi_unit",
            branch_code="5405",
            office_code="5405-01",
            filters=CustomerFilters(search_text="915"),
            office_filter_mode="representative",
        )

        self.assertEqual(rows, [])

    def test_customer_pgd_balances_remain_at_actual_pgds(self) -> None:
        self._insert_customer_with_offices(sequence="916", branch_code="5405", offices=(("00", 1000), ("01", 2000), ("02", 3000)))

        rows = self.repository.get_cross_branch_customer_offices("916", "2026-04", office_code="5405-01")

        self.assertEqual(len(rows), 1)
        self.assertEqual(float(rows[0]["total_balance"]), 2000)

    def test_branch_total_balance_includes_head_office_and_all_pgds(self) -> None:
        self._insert_customer_with_offices(sequence="917", branch_code="5405", offices=(("00", 1000), ("01", 2000), ("02", 3000)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("917", "2026-04", branch_code="5405")

        self.assertEqual(float(kpis["total_balance"]), 6000)
        self.assertEqual(float(kpis["head_office_balance"]), 1000)
        self.assertEqual(float(kpis["pgd_balance"]), 5000)

    def test_customer_two_branches_counted_once_per_branch(self) -> None:
        self._insert_customer_with_offices(sequence="918", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        self._insert_customer_with_offices(sequence="918", branch_code="5491", offices=(("01", 3000), ("02", 4000)))

        rows = self.repository.get_cross_branch_customers("2026-04", CustomerFilters(search_text="918"))

        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0]["branch_count"]), 2)
        self.assertEqual(int(rows[0]["office_count"]), 4)

    def test_cross_branch_customer_not_counted_by_number_of_offices(self) -> None:
        self._insert_customer_with_offices(sequence="919", branch_code="5405", offices=(("00", 1000), ("01", 2000), ("02", 3000)))

        self.assertEqual(
            self.repository.count_cross_branch_customers("2026-04", CustomerFilters(search_text="919")),
            0,
        )

    def test_representative_office_filter(self) -> None:
        self._insert_customer_with_offices(sequence="920", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        head_rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "all_multi_unit",
            office_code="5405-00",
            filters=CustomerFilters(search_text="920"),
            office_filter_mode="representative",
        )
        pgd_rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "all_multi_unit",
            office_code="5405-01",
            filters=CustomerFilters(search_text="920"),
            office_filter_mode="representative",
        )

        self.assertEqual(len(head_rows), 1)
        self.assertEqual(pgd_rows, [])

    def test_actual_debt_office_filter(self) -> None:
        self._insert_customer_with_offices(sequence="921", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "all_multi_unit",
            office_code="5405-01",
            filters=CustomerFilters(search_text="921"),
            office_filter_mode="actual",
        )

        self.assertEqual(len(rows), 1)

    def test_pgd_actual_debt_search_still_finds_head_office_assigned_customer(self) -> None:
        self._insert_customer_with_offices(sequence="922", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        detail = self.repository.get_cross_branch_customer_offices("922", "2026-04", office_code="5405-01")
        representative_rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "all_multi_unit",
            office_code="5405-01",
            filters=CustomerFilters(search_text="922"),
            office_filter_mode="representative",
        )

        self.assertEqual(len(detail), 1)
        self.assertEqual(representative_rows, [])

    def test_status_detected_before_office_filter(self) -> None:
        self._insert_customer_with_offices(sequence="923", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "head_and_pgd",
            office_code="5405-01",
            filters=CustomerFilters(search_text="923"),
            office_filter_mode="actual",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0]["has_head_and_pgd"]), 1)

    def test_branch_filter_does_not_break_multi_unit_detection(self) -> None:
        self._insert_customer_with_offices(sequence="924", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        rows = self.repository.get_multi_unit_customers(
            "2026-04",
            "head_and_pgd",
            branch_code="5405",
            filters=CustomerFilters(search_text="924"),
        )

        self.assertEqual(len(rows), 1)

    def test_detail_period_filter(self) -> None:
        for period in ("2026-02", "2026-03", "2026-04"):
            self._insert_customer_with_offices(period=period, sequence="925", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        history = self.repository.get_cross_branch_customer_unit_history("925", "2026-03", "2026-04")

        self.assertEqual([row["period"] for row in history], ["2026-03", "2026-04"])

    def test_detail_branch_filter(self) -> None:
        self._insert_customer_with_offices(sequence="926", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        self._insert_customer_with_offices(sequence="926", branch_code="5491", offices=(("01", 3000),))

        detail = self.repository.get_cross_branch_customer_offices("926", "2026-04", branch_code="5405")

        self.assertEqual({row["branch_code"] for row in detail}, {"5405"})
        self.assertEqual(float(sum(row["total_balance"] for row in detail)), 3000)

    def test_detail_office_filter(self) -> None:
        self._insert_customer_with_offices(sequence="927", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        detail = self.repository.get_cross_branch_customer_offices("927", "2026-04", office_code="5405-01")

        self.assertEqual(len(detail), 1)
        self.assertEqual(detail[0]["office_code"], "5405-01")

    def test_office_combo_depends_on_branch(self) -> None:
        self._insert_customer_with_offices(sequence="928", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        self._insert_customer_with_offices(sequence="928", branch_code="5491", offices=(("01", 3000),))

        offices = self.repository.get_customer_available_offices("928", "2026-04", branch_code="5405")

        self.assertEqual({row["office_code"] for row in offices}, {"5405-00", "5405-01"})

    def test_invalid_office_selection_resets(self) -> None:
        app = QApplication.instance() or QApplication([])
        self._insert_customer_with_offices(sequence="929", branch_code="5405", offices=(("00", 1000),))
        self._insert_customer_with_offices(sequence="929", branch_code="5491", offices=(("01", 2000),))
        dialog = CrossBranchCustomerDetailDialog(self.repository, "2026-04", "929")
        dialog.query_controller.cancel_pending()
        dialog.wait_for_queries()

        dialog._set_combo_current_data(dialog.office_combo, "5405-00")
        dialog._set_combo_current_data(dialog.branch_combo, "5491")
        dialog._refresh_office_options()

        self.assertEqual(current_data(dialog.office_combo), "")
        dialog.close()

    def test_kpis_recalculate_for_selected_branch(self) -> None:
        self._insert_customer_with_offices(sequence="930", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        self._insert_customer_with_offices(sequence="930", branch_code="5491", offices=(("01", 3000),))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("930", "2026-04", branch_code="5405")

        self.assertEqual(float(kpis["total_balance"]), 3000)
        self.assertEqual(int(kpis["branch_count"]), 1)

    def test_kpis_recalculate_for_selected_office(self) -> None:
        self._insert_customer_with_offices(sequence="931", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("931", "2026-04", office_code="5405-01")

        self.assertEqual(float(kpis["total_balance"]), 2000)
        self.assertEqual(int(kpis["office_count"]), 1)

    def test_filtered_office_total_balance(self) -> None:
        self._insert_customer_with_offices(sequence="932", branch_code="5405", offices=(("00", 1000), ("01", 2500)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("932", "2026-04", office_code="5405-01")

        self.assertEqual(float(kpis["total_balance"]), 2500)

    def test_filtered_head_office_balance(self) -> None:
        self._insert_customer_with_offices(sequence="933", branch_code="5405", offices=(("00", 1000), ("01", 2500)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("933", "2026-04", branch_code="5405")

        self.assertEqual(float(kpis["head_office_balance"]), 1000)

    def test_filtered_pgd_balance(self) -> None:
        self._insert_customer_with_offices(sequence="934", branch_code="5405", offices=(("00", 1000), ("01", 2500)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("934", "2026-04", branch_code="5405")

        self.assertEqual(float(kpis["pgd_balance"]), 2500)

    def test_filtered_weighted_average_rate(self) -> None:
        self._insert_customer_with_offices(sequence="935", branch_code="5405", offices=(("00", 1000, 10, 8, 7), ("01", 3000, 6, 4, 3)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("935", "2026-04", branch_code="5405")

        self.assertAlmostEqual(float(kpis["average_rate"]), 7.0)

    def test_filtered_weighted_nim_before(self) -> None:
        self._insert_customer_with_offices(sequence="936", branch_code="5405", offices=(("00", 1000, 10, 8, 7), ("01", 3000, 6, 4, 3)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("936", "2026-04", branch_code="5405")

        self.assertAlmostEqual(float(kpis["nim_before"]), 5.0)

    def test_filtered_weighted_nim_after(self) -> None:
        self._insert_customer_with_offices(sequence="937", branch_code="5405", offices=(("00", 1000, 10, 8, 7), ("01", 3000, 6, 4, 3)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("937", "2026-04", branch_code="5405")

        self.assertAlmostEqual(float(kpis["nim_after"]), 4.0)

    def test_office_count(self) -> None:
        self._insert_customer_with_offices(sequence="938", branch_code="5405", offices=(("00", 1000), ("01", 2000), ("02", 3000)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("938", "2026-04")

        self.assertEqual(int(kpis["office_count"]), 3)

    def test_head_office_count(self) -> None:
        self._insert_customer_with_offices(sequence="939", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("939", "2026-04")

        self.assertEqual(int(kpis["head_office_count"]), 1)

    def test_pgd_count(self) -> None:
        self._insert_customer_with_offices(sequence="940", branch_code="5405", offices=(("00", 1000), ("01", 2000), ("02", 3000)))

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("940", "2026-04")

        self.assertEqual(int(kpis["pgd_count"]), 2)

    def test_unit_history_returns_all_periods_in_range(self) -> None:
        for period, balance in (("2026-02", 1000), ("2026-03", 2000), ("2026-04", 3000)):
            self._insert_customer_with_offices(period=period, sequence="941", branch_code="5405", offices=(("00", balance), ("01", balance)))

        history = self.repository.get_cross_branch_customer_unit_history("941", "2026-02", "2026-04")

        self.assertEqual([row["period"] for row in history], ["2026-02", "2026-03", "2026-04"])

    def test_unit_history_not_only_endpoints(self) -> None:
        for period in ("2026-02", "2026-03", "2026-04"):
            self._insert_customer_with_offices(period=period, sequence="942", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        history = self.repository.get_cross_branch_customer_unit_history("942", "2026-02", "2026-04")

        self.assertIn("2026-03", [row["period"] for row in history])

    def test_history_head_office_pgd_flags(self) -> None:
        self._insert_customer_with_offices(period="2026-03", sequence="943", branch_code="5405", offices=(("00", 1000),))
        self._insert_customer_with_offices(period="2026-04", sequence="943", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        history = self.repository.get_cross_branch_customer_unit_history("943", "2026-03", "2026-04")

        self.assertEqual(int(history[0]["has_head_and_pgd"]), 0)
        self.assertEqual(int(history[1]["has_head_and_pgd"]), 1)

    def test_history_balance_difference(self) -> None:
        self._insert_customer_with_offices(period="2026-03", sequence="944", branch_code="5405", offices=(("00", 1000),))
        self._insert_customer_with_offices(period="2026-04", sequence="944", branch_code="5405", offices=(("00", 2500),))

        history = self.repository.get_cross_branch_customer_unit_history("944", "2026-03", "2026-04")

        self.assertEqual(float(history[1]["difference"]), 1500)

    def test_history_first_period_difference_na(self) -> None:
        self._insert_customer_with_offices(period="2026-03", sequence="945", branch_code="5405", offices=(("00", 1000),))
        self._insert_customer_with_offices(period="2026-04", sequence="945", branch_code="5405", offices=(("00", 2500),))

        history = self.repository.get_cross_branch_customer_unit_history("945", "2026-03", "2026-04")

        self.assertEqual(history[0]["difference"], "")

    def test_old_period_without_office_detail_shows_warning(self) -> None:
        self._insert_period_summary("2026-04", "5405946", "946", "Khach Old", 1000, branch_code="5405", insert_office=False)

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("946", "2026-04")

        self.assertEqual(int(kpis["office_detail_missing"]), 1)

    def test_old_period_not_assigned_fake_head_office(self) -> None:
        self._insert_period_summary("2026-04", "5405947", "947", "Khach Old", 1000, branch_code="5405", insert_office=False)

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("947", "2026-04")

        self.assertEqual(int(kpis["head_office_count"]), 0)
        self.assertEqual(float(kpis["head_office_balance"]), 0)

    def test_reimport_period_creates_office_detail(self) -> None:
        root = self.root / "reimport-office"
        root.mkdir()
        (root / "5405_FTPLN_20260630.csv").write_text(
            "\n".join(
                [
                    FTPLN_HEADER,
                    "5405,2,10,1,[540500321] Nguyen Van A,01,1000,,CN,DN1,948,Khach Reimport",
                ]
            ),
            encoding="utf-8",
        )

        import_nim_dn(self.summary_repository, root)

        self.assertTrue(self.repository.has_office_detail_for_customer_period("948", "2026-06"))

    def test_cross_branch_detail_export_three_sheets(self) -> None:
        self._insert_customer_with_offices(sequence="949", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        path = self.root / "cross-detail.xlsx"

        export_cross_branch_customer_detail(
            self.repository,
            "949",
            path,
            period_from="2026-04",
            period_to="2026-04",
            report_period="2026-04",
        )

        self.assertEqual(load_workbook(path).sheetnames, ["TongQuanLienChiNhanh", "ChiTietTheoDonVi", "LichSuTheoKy"])

    def test_office_codes_export_as_text(self) -> None:
        self._insert_customer_with_offices(sequence="950", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        path = self.root / "cross-detail-text.xlsx"

        export_cross_branch_customer_detail(self.repository, "950", path, report_period="2026-04")

        worksheet = load_workbook(path)["ChiTietTheoDonVi"]
        self.assertEqual(worksheet.cell(2, 5).number_format, "@")
        self.assertEqual(worksheet.cell(2, 5).value, "5405-00")

    def test_delete_period_removes_office_summary(self) -> None:
        self._insert_customer_with_offices(sequence="951", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        self.repository.delete_customer_period("2026-04")

        with closing(self.repository.connect()) as connection:
            count_rows = int(connection.execute("SELECT COUNT(*) FROM customer_office_period WHERE period = '2026-04'").fetchone()[0] or 0)
        self.assertEqual(count_rows, 0)

    def test_import_rollback_restores_office_summary(self) -> None:
        root_a = self.root / "rollback-office-a"
        root_b = self.root / "rollback-office-b"
        root_a.mkdir()
        root_b.mkdir()
        (root_a / "5405_FTPLN_20260731.csv").write_text(
            "\n".join([FTPLN_HEADER, "5405,2,10,1,[540500321] Nguyen Van A,01,1000,,CN,DN1,952,Khach Rollback"]),
            encoding="utf-8",
        )
        (root_b / "5405_FTPLN_20260730.csv").write_text(
            "\n".join([FTPLN_HEADER, "5405,2,10,1,[540500321] Nguyen Van A,01,2000,,CN,DN1,,Khach Loi"]),
            encoding="utf-8",
        )
        import_nim_dn(self.summary_repository, root_a)

        with self.assertRaises(Exception):
            import_nim_dn(self.summary_repository, root_b, replace_existing_periods=True)

        kpis = self.repository.get_cross_branch_customer_filtered_kpis("952", "2026-07")
        self.assertEqual(float(kpis["total_balance"]), 1000)
        self.assertEqual(int(kpis["office_detail_missing"]), 0)

    def test_customer_period_summary_row_count_unchanged(self) -> None:
        self._insert_customer_with_offices(sequence="953", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        with closing(self.repository.connect()) as connection:
            summary_count = int(
                connection.execute(
                    "SELECT COUNT(*) FROM customer_period_summary WHERE period = '2026-04' AND customer_sequence = '953'"
                ).fetchone()[0]
                or 0
            )
            office_count = int(
                connection.execute(
                    "SELECT COUNT(*) FROM customer_office_period WHERE period = '2026-04' AND customer_sequence = '953'"
                ).fetchone()[0]
                or 0
            )
        self.assertEqual(summary_count, 1)
        self.assertEqual(office_count, 2)

    def test_customer_total_balance_unchanged(self) -> None:
        self._insert_customer_with_offices(sequence="954", branch_code="5405", offices=(("00", 1000), ("01", 2000)))

        with closing(self.repository.connect()) as connection:
            summary_total = float(
                connection.execute(
                    "SELECT total_balance FROM customer_period_summary WHERE period = '2026-04' AND customer_sequence = '954'"
                ).fetchone()[0]
            )
            office_total = float(
                connection.execute(
                    "SELECT SUM(total_balance) FROM customer_office_period WHERE period = '2026-04' AND customer_sequence = '954'"
                ).fetchone()[0]
            )

        self.assertEqual(summary_total, office_total)

    def test_export_contains_representative_and_actual_offices(self) -> None:
        self._insert_customer_with_offices(sequence="955", branch_code="5405", offices=(("00", 1000), ("01", 2000)))
        path = self.root / "cross-branch-representative.xlsx"

        export_cross_branch_customers(
            self.repository,
            "2026-04",
            CustomerFilters(search_text="955"),
            2,
            path,
            scope_type="all_multi_unit",
            office_filter_mode="representative",
        )

        workbook = load_workbook(path)
        summary_headers = [cell.value for cell in workbook["KhachHangLienChiNhanh"][1]]
        detail_headers = [cell.value for cell in workbook["ChiTietTheoDonVi"][1]]
        self.assertIn("Đơn vị đại diện", summary_headers)
        self.assertIn("Danh sách đơn vị thực tế", summary_headers)
        self.assertIn("Mã đơn vị", detail_headers)

    def test_multiple_officers_same_period(self) -> None:
        result = self.repository.multiple_officer_rows(CustomerFilters(current_period="2026-04", multi_status="same_period"))
        self.assertEqual(result.total_rows, 1)
        self.assertEqual(result.rows[0]["customer_code"], "5491006")

    def test_officer_change_between_periods_not_same_period_multi(self) -> None:
        same_period = self.repository.multiple_officer_rows(CustomerFilters(current_period="2026-04", multi_status="same_period"))
        changed = self.repository.multiple_officer_rows(CustomerFilters(current_period="2026-04", multi_status="changed_period"))
        self.assertNotIn("5491004", {row["customer_code"] for row in same_period.rows})
        self.assertIn("5491004", {row["customer_code"] for row in changed.rows})

    def test_effective_officer_without_override(self) -> None:
        detail = self.repository.customer_detail("5491001", "2026-04")
        self.assertEqual(detail["effective_officer_name"], "Officer A")

    def test_officer_override_one_period(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            effective_to_period="2026-04",
            officer_code="900",
            officer_name="Officer Override",
            reason="one",
        )
        self.assertEqual(self.repository.customer_detail("5491001", "2026-04")["effective_officer_name"], "Officer Override")
        self.assertEqual(self.repository.customer_detail("5491001", "2026-03")["effective_officer_name"], "Officer A")

    def test_officer_override_from_period_forward(self) -> None:
        self._insert_period_summary("2026-05", "5491001", "001", "Khach A", 1700)
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            officer_code="901",
            officer_name="Forward Officer",
            reason="forward",
        )
        self.assertEqual(self.repository.customer_detail("5491001", "2026-05")["effective_officer_name"], "Forward Officer")

    def test_officer_override_date_range(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-03",
            effective_to_period="2026-04",
            officer_code="902",
            officer_name="Range Officer",
            reason="range",
        )
        self.assertEqual(self.repository.customer_detail("5491001", "2026-03")["effective_officer_name"], "Range Officer")
        self.assertEqual(self.repository.customer_detail("5491001", "2026-04")["effective_officer_name"], "Range Officer")

    def test_officer_override_open_ended(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            effective_to_period="",
            officer_code="903",
            officer_name="Open Officer",
            reason="open",
        )
        self.assertEqual(self.repository.customer_detail("5491001", "2026-04")["effective_officer_name"], "Open Officer")

    def test_override_overlap_resolution(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-03",
            effective_to_period="2026-04",
            officer_code="904",
            officer_name="Old Override",
            reason="old",
        )
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            effective_to_period="",
            officer_code="905",
            officer_name="New Override",
            reason="new",
        )
        with closing(self.repository.connect()) as connection:
            active_count = int(
                connection.execute(
                    "SELECT COUNT(*) FROM customer_officer_override WHERE customer_code = '5491001' AND is_active = 1"
                ).fetchone()[0]
            )
        self.assertEqual(active_count, 1)
        self.assertEqual(self.repository.customer_detail("5491001", "2026-04")["effective_officer_name"], "New Override")

    def test_restore_imported_officer(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            officer_code="906",
            officer_name="Restore Me",
            reason="restore",
        )
        self.repository.restore_imported_officer(customer_code="5491001", period="2026-04")
        self.assertEqual(self.repository.customer_detail("5491001", "2026-04")["effective_officer_name"], "Officer A")

    def test_officer_override_action_log(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            officer_code="907",
            officer_name="Log Officer",
            reason="log",
        )
        logs = self.repository.action_logs("5491001")
        self.assertEqual(logs[0]["action_type"], "OFFICER_OVERRIDE")

    def test_officer_directory_normalization(self) -> None:
        self.repository.upsert_officer_directory(officer_code="  abc  ", officer_name="  Nguyen   Van   A  ")
        rows = self.repository.officer_directory(search_text="abc").rows
        self.assertEqual(rows[0]["officer_code"], "abc")
        self.assertEqual(rows[0]["officer_name"], "Nguyen Van A")

    def test_officer_directory_does_not_merge_same_name_different_code(self) -> None:
        self.repository.upsert_officer_directory(officer_code="a01", officer_name="Nguyen Van A")
        self.repository.upsert_officer_directory(officer_code="a02", officer_name="Nguyen Van A")
        rows = self.repository.officer_directory(search_text="Nguyen Van A").rows
        self.assertGreaterEqual(len([row for row in rows if row["officer_code"] in {"a01", "a02"}]), 2)

    def test_officer_dialog_reasonable_default_width(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        self.assertGreaterEqual(dialog.width(), 520)
        self.assertGreaterEqual(dialog.minimumWidth(), 520)
        dialog.close()

    def test_officer_dialog_fields_expand(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        self.assertEqual(dialog.name_input.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
        self.assertEqual(dialog.branch_input.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
        dialog.close()

    def test_officer_dialog_name_field_not_too_short(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        self.assertGreaterEqual(dialog.name_input.minimumWidth(), 340)
        dialog.close()

    def test_officer_dialog_branch_combo_not_too_short(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        self.assertGreaterEqual(dialog.branch_input.minimumWidth(), 340)
        dialog.close()

    def test_officer_dialog_has_no_vertical_expanding_spacer(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        try:
            def assert_no_vertical_expanding(layout) -> None:
                for index in range(layout.count()):
                    item = layout.itemAt(index)
                    if item.spacerItem() is not None:
                        self.assertNotEqual(item.spacerItem().sizePolicy().verticalPolicy(), QSizePolicy.Policy.Expanding)
                    if item.layout() is not None:
                        assert_no_vertical_expanding(item.layout())

            assert_no_vertical_expanding(dialog.layout())
        finally:
            dialog.close()

    def test_officer_dialog_height_is_compact(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        try:
            self.assertLessEqual(dialog.height(), 360)
            self.assertLessEqual(dialog.minimumHeight(), 280)
        finally:
            dialog.close()

    def test_officer_dialog_buttons_close_to_form(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        try:
            dialog.show()
            app.processEvents()
            gap = dialog.save_button.geometry().top() - dialog.active_check.geometry().bottom()
            self.assertLessEqual(gap, 80)
            self.assertGreaterEqual(gap, 8)
        finally:
            dialog.close()

    def test_officer_dialog_fields_expand_horizontally(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        try:
            for field in (dialog.code_input, dialog.name_input, dialog.branch_input, dialog.office_input):
                self.assertEqual(field.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
                self.assertGreaterEqual(field.minimumWidth(), 340)
                self.assertLessEqual(field.maximumHeight(), 34)
        finally:
            dialog.close()

    def test_officer_dialog_save_cancel_same_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        try:
            self.assertEqual(dialog.save_button.minimumHeight(), dialog.cancel_button.minimumHeight())
            self.assertEqual(dialog.save_button.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Preferred)
            self.assertEqual(dialog.cancel_button.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Preferred)
        finally:
            dialog.close()

    def test_officer_dialog_no_large_blank_area(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        try:
            dialog.show()
            app.processEvents()
            gap = dialog.save_button.geometry().top() - dialog.active_check.geometry().bottom()
            self.assertLessEqual(gap, 80)
            self.assertLessEqual(dialog.height(), 360)
        finally:
            dialog.close()

    def test_officer_dialog_preserves_leading_zero_code(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerDirectoryDialog(self.repository)
        dialog.code_input.setText("  00123  ")
        dialog.name_input.setText("Officer Leading Zero")
        dialog._save()
        rows = self.repository.officer_directory(search_text="00123").rows
        self.assertEqual(rows[0]["officer_code"], "00123")

    def test_officer_dialog_duplicate_validation(self) -> None:
        app = QApplication.instance() or QApplication([])
        self.repository.upsert_officer_directory(officer_code="DUP01", officer_name="Officer Dup")
        dialog = OfficerDirectoryDialog(self.repository)
        dialog.code_input.setText("dup01")
        dialog.name_input.setText("Officer Dup 2")
        dialog._save()
        self.assertIn("đã tồn tại", dialog.error_label.text())
        dialog.close()

    def test_officer_combo_minimum_width(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Tất cả cán bộ", minimum_width=220, maximum_width=320, searchable=True)
        populate_officer_combo(combo, self.repository.distinct_officers(CustomerFilters()))
        self.assertGreaterEqual(combo.minimumWidth(), 220)
        combo.deleteLater()

    def test_officer_combo_popup_expands_for_long_names(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Tất cả cán bộ", minimum_width=220, maximum_width=320, searchable=True)
        populate_officer_combo(
            combo,
            [{"officer_code": "540600123", "officer_name": "Nguyen Van A co ten rat dai de kiem tra popup"}],
        )
        width = configure_combo_popup_width(combo, minimum_popup_width=360)
        self.assertGreaterEqual(width, combo.minimumWidth())
        combo.deleteLater()

    def test_officer_combo_popup_fits_screen(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Tất cả cán bộ", minimum_width=220, maximum_width=320, searchable=True)
        populate_officer_combo(
            combo,
            [{"officer_code": "540600123", "officer_name": "Nguyen Van A co ten rat dai de kiem tra popup"}],
        )
        width = configure_combo_popup_width(combo, minimum_popup_width=360, maximum_screen_ratio=0.5)
        self.assertLessEqual(width, int(combo.screen().availableGeometry().width() * 0.5))
        combo.deleteLater()

    def test_officer_combo_uses_code_as_item_data(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Tất cả cán bộ", minimum_width=220, maximum_width=320, searchable=True)
        populate_officer_combo(combo, [{"officer_code": "540600123", "officer_name": "Nguyen Van A"}])
        self.assertEqual(combo.itemData(1), "540600123")
        self.assertEqual(combo.itemData(1, OFFICER_NAME_ROLE), "Nguyen Van A")
        combo.deleteLater()

    def test_same_name_different_officer_code_are_distinct(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Tất cả cán bộ", minimum_width=220, maximum_width=320, searchable=True)
        populate_officer_combo(
            combo,
            [
                {"officer_code": "A01", "officer_name": "Nguyen Van A"},
                {"officer_code": "A02", "officer_name": "Nguyen Van A"},
            ],
        )
        self.assertEqual(combo.itemData(1), "A01")
        self.assertEqual(combo.itemData(2), "A02")
        self.assertNotEqual(combo.itemText(1), combo.itemText(2))
        combo.deleteLater()

    def test_officer_combo_full_name_tooltip(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Tất cả cán bộ", minimum_width=220, maximum_width=320, searchable=True)
        populate_officer_combo(combo, [{"officer_code": "A01", "officer_name": "Nguyen Van A rat dai"}])
        self.assertIn("Nguyen Van A rat dai", combo.itemData(1, Qt.ItemDataRole.ToolTipRole))
        combo.deleteLater()

    def test_officer_combo_search_by_code(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Tất cả cán bộ", minimum_width=220, maximum_width=320, searchable=True)
        populate_officer_combo(combo, [{"officer_code": "540600123", "officer_name": "Nguyen Van A"}])
        combo.completer().setCompletionPrefix("540600123")
        self.assertGreaterEqual(combo.completer().completionCount(), 1)
        combo.deleteLater()

    def test_officer_combo_search_by_name(self) -> None:
        app = QApplication.instance() or QApplication([])
        combo = combo_box("Tất cả cán bộ", minimum_width=220, maximum_width=320, searchable=True)
        populate_officer_combo(combo, [{"officer_code": "540600123", "officer_name": "Nguyen Van A"}])
        combo.completer().setCompletionPrefix("Van A")
        self.assertGreaterEqual(combo.completer().completionCount(), 1)
        combo.deleteLater()

    def test_officer_lookup_by_exact_code_fills_name(self) -> None:
        app = QApplication.instance() or QApplication([])
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("001", "")
        widget.lookup_by_code()
        self.assertEqual(widget.name_combo.currentText(), "Officer A")
        self.assertEqual(widget.selected_officer()["officer_code"], "001")
        widget.deleteLater()

    def test_officer_lookup_by_code_prefix_returns_suggestions(self) -> None:
        app = QApplication.instance() or QApplication([])
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("00", "")
        widget.lookup_by_code()
        self.assertGreaterEqual(widget.code_combo.count(), 1)
        self.assertIsNone(widget.selected_officer())
        widget.deleteLater()

    def test_officer_lookup_suggestion_shows_code_but_selection_displays_name(self) -> None:
        app = QApplication.instance() or QApplication([])
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("00", "")
        widget.lookup_by_code()
        index = widget.name_combo.findData("001")
        self.assertGreaterEqual(index, 0)
        self.assertIn("[001]", widget.name_combo.itemText(index))
        widget.name_combo.setCurrentIndex(index)
        widget._select_from_combo(widget.name_combo)
        self.assertEqual(widget.name_combo.currentText(), "Officer A")
        self.assertEqual(widget.selected_officer()["officer_code"], "001")
        widget.deleteLater()

    def test_officer_lookup_by_name_fills_code(self) -> None:
        app = QApplication.instance() or QApplication([])
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("", "Officer A")
        widget.lookup_by_name()
        self.assertEqual(widget.name_combo.currentText(), "Officer A")
        self.assertEqual(widget.selected_officer()["officer_code"], "001")
        self.assertEqual(widget.selected_officer()["officer_name"], "Officer A")
        widget.deleteLater()

    def test_officer_lookup_name_case_insensitive(self) -> None:
        app = QApplication.instance() or QApplication([])
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("", "officer a")
        widget.lookup_by_name()
        self.assertEqual(widget.name_combo.currentText(), "Officer A")
        self.assertEqual(widget.selected_officer()["officer_code"], "001")
        widget.deleteLater()

    def test_officer_lookup_same_name_requires_selection(self) -> None:
        app = QApplication.instance() or QApplication([])
        self.repository.upsert_officer_directory(officer_code="D01", officer_name="Same Officer")
        self.repository.upsert_officer_directory(officer_code="D02", officer_name="Same Officer")
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("", "Same Officer")
        widget.lookup_by_name()
        self.assertIsNone(widget.selected_officer())
        self.assertIn("trùng tên", widget.status_label.text())
        widget.deleteLater()

    def test_officer_lookup_same_name_different_codes(self) -> None:
        self.repository.upsert_officer_directory(officer_code="E01", officer_name="Duplicate Name")
        self.repository.upsert_officer_directory(officer_code="E02", officer_name="Duplicate Name")
        rows = self.repository.find_officers_by_name("Duplicate Name")
        self.assertEqual({row["officer_code"] for row in rows if row["officer_name"] == "Duplicate Name"}, {"E01", "E02"})

    def test_officer_lookup_preserves_leading_zero_code(self) -> None:
        app = QApplication.instance() or QApplication([])
        self.repository.upsert_officer_directory(officer_code="00123", officer_name="Officer Zero")
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("00123", "")
        widget.lookup_by_code()
        self.assertEqual(widget.name_combo.currentText(), "Officer Zero")
        self.assertEqual(widget.selected_officer()["officer_code"], "00123")
        widget.deleteLater()

    def test_officer_lookup_code_name_mismatch_rejected(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerOverrideDialog(
            self.repository,
            customer_code="5491001",
            customer_name="Khach A",
            period="2026-04",
        )
        dialog.officer_lookup.set_pending_identity("001", "Officer B")
        with patch("PySide6.QtWidgets.QMessageBox.warning") as warning:
            dialog.save_override()
        self.assertTrue(warning.called)
        self.assertNotEqual(dialog.result(), QDialog.DialogCode.Accepted)
        dialog.close()

    def test_override_dialog_uses_single_officer_name_lookup_field(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerOverrideDialog(
            self.repository,
            customer_code="5491001",
            customer_name="Khach A",
            period="2026-04",
        )
        lookup_labels = [label.text() for label in dialog.officer_lookup.findChildren(QLabel)]
        self.assertIn("Tên cán bộ mới", lookup_labels)
        self.assertNotIn("Mã cán bộ mới", lookup_labels)
        self.assertIs(dialog.officer_code_input, dialog.officer_name_input)
        self.assertEqual(len(dialog.officer_lookup.findChildren(QComboBox)), 1)
        dialog.close()

    def test_officer_lookup_inactive_officer_not_selectable(self) -> None:
        app = QApplication.instance() or QApplication([])
        self.repository.upsert_officer_directory(officer_code="IN01", officer_name="Inactive Officer", is_active=False)
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("IN01", "")
        widget.lookup_by_code()
        self.assertIsNone(widget.selected_officer())
        self.assertIn("Không tìm thấy", widget.status_label.text())
        widget.deleteLater()

    def test_officer_lookup_add_new_officer_flow(self) -> None:
        app = QApplication.instance() or QApplication([])
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("NEW01", "New Officer")

        def accept_dialog(dialog):
            dialog.repository.upsert_officer_directory(
                officer_code=dialog.code_input.text(),
                officer_name=dialog.name_input.text(),
            )
            return QDialog.DialogCode.Accepted

        with patch("agribank_v3.features.credit.summary.customer.officer_management_tab.OfficerDirectoryDialog.exec", accept_dialog):
            widget.add_new_officer()
        self.assertEqual(widget.selected_officer()["officer_code"], "NEW01")
        widget.deleteLater()

    def test_officer_lookup_refresh_after_add(self) -> None:
        app = QApplication.instance() or QApplication([])
        widget = OfficerLookupWidget(self.repository)
        widget.set_pending_identity("NEW02", "New Officer 2")

        def accept_dialog(dialog):
            dialog.repository.upsert_officer_directory(
                officer_code=dialog.code_input.text(),
                officer_name=dialog.name_input.text(),
            )
            return QDialog.DialogCode.Accepted

        with patch("agribank_v3.features.credit.summary.customer.officer_management_tab.OfficerDirectoryDialog.exec", accept_dialog):
            widget.add_new_officer()
        self.assertEqual(self.repository.get_officer_by_code("NEW02")["officer_name"], "New Officer 2")
        self.assertEqual(widget.name_combo.currentText(), "New Officer 2")
        widget.deleteLater()

    def test_officer_lookup_debounce(self) -> None:
        app = QApplication.instance() or QApplication([])
        widget = OfficerLookupWidget(self.repository)
        widget.code_combo.setEditText("00")
        self.assertTrue(widget.code_timer.isActive())
        self.assertEqual(widget.code_timer.interval(), 300)
        widget.code_timer.stop()
        widget.name_timer.stop()
        widget.deleteLater()

    def test_override_saves_selected_officer_identity(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerOverrideDialog(
            self.repository,
            customer_code="5491001",
            customer_name="Khach A",
            period="2026-04",
        )
        dialog.officer_lookup.set_pending_identity("002", "")
        dialog.officer_lookup.lookup_by_code()
        dialog.reason_input.setText("lookup save")
        dialog.save_override()
        self.assertEqual(self.repository.customer_detail("5491001", "2026-04")["effective_officer_name"], "Officer B")
        dialog.close()

    def test_override_does_not_modify_imported_officer(self) -> None:
        app = QApplication.instance() or QApplication([])
        before = self.repository.customer_detail("5491001", "2026-04")
        dialog = OfficerOverrideDialog(
            self.repository,
            customer_code="5491001",
            customer_name="Khach A",
            period="2026-04",
        )
        dialog.officer_lookup.set_pending_identity("002", "")
        dialog.officer_lookup.lookup_by_code()
        dialog.reason_input.setText("import preserved")
        dialog.save_override()
        after = self.repository.customer_detail("5491001", "2026-04")
        self.assertEqual(after["imported_officer_code"], before["imported_officer_code"])
        self.assertEqual(after["imported_officer_name"], before["imported_officer_name"])
        dialog.close()

    def test_override_action_log_unchanged(self) -> None:
        app = QApplication.instance() or QApplication([])
        dialog = OfficerOverrideDialog(
            self.repository,
            customer_code="5491001",
            customer_name="Khach A",
            period="2026-04",
        )
        dialog.officer_lookup.set_pending_identity("002", "")
        dialog.officer_lookup.lookup_by_code()
        dialog.reason_input.setText("lookup log")
        dialog.save_override()
        logs = self.repository.action_logs("5491001")
        self.assertEqual(logs[0]["action_type"], "OFFICER_OVERRIDE")
        dialog.close()

    def test_export_customer_dashboard(self) -> None:
        path = self.root / "dashboard.xlsx"
        export_customer_dashboard(self.repository, CustomerFilters(current_period="2026-04"), path)
        self.assertIn("TongQuanKhachHang", load_workbook(path).sheetnames)

    def test_export_customer_list_all_filtered_rows(self) -> None:
        path = self.root / "list.xlsx"
        export_customer_list(self.repository, CustomerFilters(current_period="2026-04"), path)
        ws = load_workbook(path)["DanhSachKhachHang"]
        self.assertEqual(ws.max_row - 1, 5)

    def test_export_customer_growth(self) -> None:
        path = self.root / "growth.xlsx"
        export_customer_growth(self.repository, "2026-03", "2026-04", CustomerFilters(), path)
        self.assertIn("BienDongDuNo", load_workbook(path).sheetnames)

    def test_export_top_customer_balance_uses_filter_and_limit(self) -> None:
        path = self.root / "top_balance.xlsx"
        export_top_customer_balance(
            self.repository,
            CustomerFilters(current_period="2026-04", customer_type="TC"),
            "2026-04",
            1,
            path,
        )
        worksheet = load_workbook(path)["TopKhachHangDuNo"]
        self.assertEqual(worksheet.max_row - 1, 1)
        self.assertEqual(worksheet.cell(2, 2).value, "5491003")

    def test_export_top_customer_movement_uses_direction_and_limit(self) -> None:
        path = self.root / "top_movement.xlsx"
        export_top_customer_movement(
            self.repository,
            CustomerFilters(),
            "2026-03",
            "2026-04",
            "decrease",
            1,
            path,
        )
        worksheet = load_workbook(path)["TopGiamDuNo"]
        self.assertEqual(worksheet.max_row - 1, 1)
        self.assertEqual(worksheet.cell(2, 2).value, "5491004")

    def test_export_multiple_officers(self) -> None:
        path = self.root / "multiple.xlsx"
        export_multiple_officers(self.repository, CustomerFilters(current_period="2026-04", multi_status="same_period"), path)
        self.assertIn("NhieuCanBoQuanLy", load_workbook(path).sheetnames)

    def test_export_import_history(self) -> None:
        path = self.root / "history.xlsx"
        export_import_history(self.repository, path)
        self.assertIn("LichSuImport", load_workbook(path).sheetnames)

    def test_export_officer_directory(self) -> None:
        path = self.root / "officers.xlsx"
        export_officer_directory(self.repository, path)
        self.assertIn("DanhMucCanBo", load_workbook(path).sheetnames)

    def test_export_customer_detail(self) -> None:
        path = self.root / "detail.xlsx"
        export_customer_detail(self.repository, "5491001", path)
        self.assertTrue({"LichSuDuNo", "CoCauKyHan", "NIMLaiSuat", "LichSuCanBo", "SoSanhCacKy"}.issubset(load_workbook(path).sheetnames))

    def test_export_all_sheets(self) -> None:
        path = self.root / "all.xlsx"
        export_all_customer_sheets(
            self.repository,
            CustomerFilters(current_period="2026-04"),
            path,
            previous_period="2026-03",
            current_period="2026-04",
        )
        expected = {"TongQuanKhachHang", "DanhSachKhachHang", "BienDongDuNo", "NhieuCanBoQuanLy", "LichSuImport", "DanhMucCanBo"}
        self.assertTrue(expected.issubset(load_workbook(path).sheetnames))

    def test_export_uses_current_filters(self) -> None:
        path = self.root / "filtered.xlsx"
        export_customer_list(self.repository, CustomerFilters(current_period="2026-03", customer_type="TC"), path)
        ws = load_workbook(path)["DanhSachKhachHang"]
        self.assertEqual(ws.max_row - 1, 1)
        self.assertEqual(ws.cell(2, 2).value, "5491002")

    def test_delete_customer_period(self) -> None:
        info = self.repository.delete_customer_period("2026-04")
        self.assertEqual(int(info["customer_count"]), 5)
        self.assertEqual(self.repository.query_customer_list(CustomerFilters(current_period="2026-04")).total_rows, 0)

    def test_delete_period_preserves_override(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            officer_code="908",
            officer_name="Preserved Override",
            reason="preserve",
        )
        self.repository.delete_customer_period("2026-04")
        with closing(self.repository.connect()) as connection:
            count_rows = int(connection.execute("SELECT COUNT(*) FROM customer_officer_override").fetchone()[0])
        self.assertGreater(count_rows, 0)

    def test_delete_period_preserves_action_log(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491001",
            effective_from_period="2026-04",
            officer_code="909",
            officer_name="Log Preserved",
            reason="preserve log",
        )
        self.repository.delete_customer_period("2026-04")
        self.assertGreater(len(self.repository.action_logs("5491001")), 0)

    def test_delete_period_updates_customer_master(self) -> None:
        self.repository.delete_customer_period("2026-04")
        with closing(self.repository.connect()) as connection:
            row_a = connection.execute("SELECT last_seen_period, is_active FROM customer_master WHERE customer_code = '5491001'").fetchone()
            row_c = connection.execute("SELECT last_seen_period, is_active FROM customer_master WHERE customer_code = '5491003'").fetchone()
        self.assertEqual(row_a["last_seen_period"], "2026-03")
        self.assertEqual(int(row_a["is_active"]), 1)
        self.assertEqual(int(row_c["is_active"]), 0)

    def test_customer_ui_uses_pagination_not_load_all(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window.list_tab.page_size = 2
        window.list_tab.refresh()
        self.assertTrue(_wait_until(lambda: window.list_tab.pager.total_rows > 0))
        self.assertLessEqual(window.list_tab.model.rowCount(), 2)
        self.assertGreater(window.list_tab.pager.total_rows, window.list_tab.model.rowCount())
        window.close()

    def test_customer_period_summary_unique_unchanged(self) -> None:
        with closing(self.repository.connect()) as connection:
            with self.assertRaises(sqlite3.IntegrityError):
                connection.execute(
                    """
                    INSERT INTO customer_period_summary(
                        period, customer_code, branch_code, customer_sequence, customer_name,
                        customer_type, created_at, updated_at
                    )
                    VALUES ('2026-04', '5491001', '5491', '001', 'Duplicate', 'CN', 'now', 'now')
                    """
                )
                connection.commit()

    def test_customer_db_has_no_raw_loan_table(self) -> None:
        with closing(self.repository.connect()) as connection:
            tables = {
                str(row[0])
                for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
            }
        self.assertNotIn("customer_raw_loans", tables)
        self.assertNotIn("customer_loan_details", tables)
        self.assertNotIn("ftpln_raw", tables)

    def test_nim_dn_result_unchanged_after_phase_c(self) -> None:
        root = self.root / "nim_unchanged"
        root.mkdir()
        (root / "5491_FTPLN_20260630.csv").write_text(
            "\n".join(
                [
                    FTPLN_HEADER,
                    "5491,2,10,1,[540000321] Nguyen Van A,00,1000,,CN,DN1,001,Khach A",
                    "5491,4,8,1,[540000322] Nguyen Van B,00,3000,,TC,DN7,002,Khach B",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.summary_repository, root)
        rows = self.summary_repository.query_nim(SummaryDataType.NIM_DN).rows
        balance = sum(float(row["balance"]) for row in rows)
        average_rate = sum(float(row["balance"]) * float(row["average_rate"]) for row in rows) / balance
        nim_before = sum(float(row["balance"]) * float(row["nim_before"]) for row in rows) / balance
        nim_after = sum(float(row["balance"]) * float(row["nim_after"]) for row in rows) / balance
        self.assertAlmostEqual(balance, 4000)
        self.assertAlmostEqual(average_rate, 8.5)
        self.assertAlmostEqual(nim_before, 5.0)
        self.assertAlmostEqual(nim_after, 4.0)

    def test_nim_dn_tab_has_customer_data_button(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = NimTab(self.summary_repository, SummaryDataType.NIM_DN)
        button_texts = [button.text() for button in tab.findChildren(QPushButton)]
        self.assertIn("Dữ liệu khách hàng", button_texts)
        tab.close()

    def test_dashboard_chart_dataset_balance_trend(self) -> None:
        trends = self.repository.dashboard_trends(CustomerFilters(current_period="2026-04"))
        dataset = chart_service.dashboard_chart_dataset_balance_trend(trends)
        self.assertEqual(dataset[0][0], "Tổng dư nợ")
        self.assertEqual(dataset[0][1][-1], ("2026-04", 10000.0))

    def test_balance_trend_has_total_balance_only(self) -> None:
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), group_by="total")
        dataset = chart_service.dashboard_chart_dataset_balance_trend(rows)
        self.assertEqual([name for name, _points in dataset], ["Tổng dư nợ"])

    def test_balance_trend_has_no_short_term_series(self) -> None:
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), group_by="total")
        dataset = chart_service.dashboard_chart_dataset_balance_trend(rows)
        self.assertNotIn("Dư nợ ngắn hạn", [name for name, _points in dataset])

    def test_balance_trend_has_no_medium_long_series(self) -> None:
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), group_by="total")
        dataset = chart_service.dashboard_chart_dataset_balance_trend(rows)
        self.assertNotIn("Dư nợ trung/dài hạn", [name for name, _points in dataset])

    def test_balance_trend_no_duplicate_legend(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("Balance", value_kind="money")
        chart.set_series((("5491 - CN Lộc Phát", (("2026-04", 1000),)), ("5400 - CN Lâm Đồng", (("2026-04", 2000),))))
        labels = [marker.label() for marker in chart.chart.legend().markers() if marker.isVisible()]
        self.assertEqual(len(labels), len(set(labels)))
        chart.deleteLater()

    def test_balance_trend_respects_branch_filter(self) -> None:
        self._insert_period_summary("2026-04", "5400001", "001", "Khach Branch", 700, branch_code="5400")
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-04", period_to="2026-04", branch_code="5400"), group_by="total")
        self.assertEqual(rows[0]["value"], 700)

    def test_balance_trend_respects_customer_type_filter(self) -> None:
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-04", period_to="2026-04", customer_type="TC"), group_by="total")
        self.assertEqual(rows[0]["value"], 3000)

    def test_balance_trend_group_by_branch(self) -> None:
        self._insert_period_summary("2026-04", "5400001", "001", "Khach Branch", 700, branch_code="5400")
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), group_by="branch")
        dataset = chart_service.dashboard_chart_dataset_balance_trend(rows)
        self.assertEqual([name for name, _points in dataset], ["5400 - CN Lâm Đồng", "5491 - CN Lộc Phát"])

    def test_balance_trend_group_by_customer_type(self) -> None:
        self._insert_period_summary("2026-04", "5491999", "999", "Khach Other", 100, customer_type="")
        rows = self.repository.get_total_balance_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), group_by="customer_type")
        dataset = chart_service.dashboard_chart_dataset_balance_trend(rows)
        self.assertEqual([name for name, _points in dataset], ["Cá nhân", "Tổ chức/Pháp nhân", "Khác"])

    def test_balance_trend_period_sorted(self) -> None:
        rows = self.repository.get_total_balance_trend(CustomerFilters(), group_by="total")
        self.assertEqual([row["period"] for row in rows], sorted(row["period"] for row in rows))

    def test_balance_group_combo_updates_chart(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        with patch.object(tab.balance_chart_controller, "run") as run:
            tab.balance_group_combo.setCurrentIndex(1)
        self.assertTrue(run.called)
        self.assertEqual(run.call_args.kwargs["cache_key"][-1], "branch")
        tab.close()

    def test_dashboard_chart_dataset_term_structure(self) -> None:
        metrics = self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04"))
        slices = dict(chart_service.dashboard_chart_dataset_term_structure(metrics))
        self.assertEqual(slices["Ngắn hạn"], 6000.0)
        self.assertEqual(slices["Trung/dài hạn"], 4000.0)

    def test_dashboard_chart_dataset_customer_movements(self) -> None:
        kpis = self.repository.movement_kpis("2026-03", "2026-04", CustomerFilters())
        dataset = chart_service.dashboard_chart_dataset_customer_movements(kpis, value_mode="count", period="2026-04")
        self.assertIn(("Tăng dư nợ", (("2026-04", 1.0),)), dataset)

    def test_dashboard_chart_dataset_nim_rates(self) -> None:
        trends = self.repository.dashboard_trends(CustomerFilters(current_period="2026-04"))
        dataset = chart_service.dashboard_chart_dataset_nim_rates(trends)
        self.assertEqual(dataset[1][0], "NIM trước ĐC")
        self.assertAlmostEqual(dataset[1][1][-1][1], 6.5)

    def test_metric_combo_has_three_metrics(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        self.assertEqual([tab.metric_combo.itemData(index) for index in range(tab.metric_combo.count())], ["average_rate", "nim_before", "nim_after"])
        tab.close()

    def test_average_rate_metric_weighted(self) -> None:
        rows = self.repository.get_customer_metric_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), metric="average_rate")
        self.assertAlmostEqual(rows[0]["value"], 8.5)

    def test_nim_before_metric_weighted(self) -> None:
        rows = self.repository.get_customer_metric_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), metric="nim_before")
        self.assertAlmostEqual(rows[0]["value"], 6.5)

    def test_nim_after_metric_weighted(self) -> None:
        rows = self.repository.get_customer_metric_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), metric="nim_after")
        self.assertAlmostEqual(rows[0]["value"], 5.5)

    def test_metric_chart_has_one_series_only(self) -> None:
        rows = self.repository.get_customer_metric_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), metric="nim_before")
        dataset = chart_service.dashboard_chart_dataset_metric_trend(rows, "nim_before")
        self.assertEqual(len(dataset), 1)
        self.assertEqual(dataset[0][0], "NIM trước điều chỉnh")

    def test_metric_combo_updates_title(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        with patch.object(tab.metric_chart_controller, "run"):
            tab.metric_combo.setCurrentIndex(1)
        self.assertEqual(tab.nim_chart.title, "Xu hướng NIM trước điều chỉnh")
        tab.close()

    def test_metric_combo_updates_without_full_dashboard_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        with patch.object(tab.query_controller, "run") as dashboard_run, patch.object(tab.metric_chart_controller, "run"):
            tab.metric_combo.setCurrentIndex(1)
        self.assertFalse(dashboard_run.called)
        tab.close()

    def test_metric_cache_key_includes_metric(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2026-04"))
        with patch.object(tab.metric_chart_controller, "run") as run:
            tab.metric_combo.setCurrentIndex(2)
        self.assertEqual(run.call_args.kwargs["cache_key"][-1], "nim_after")
        tab.close()

    def test_metric_percentage_not_multiplied_twice(self) -> None:
        rows = self.repository.get_customer_metric_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"), metric="average_rate")
        dataset = chart_service.dashboard_chart_dataset_metric_trend(rows, "average_rate")
        self.assertAlmostEqual(dataset[0][1][0][1], 8.5)
        self.assertLess(dataset[0][1][0][1], 100)

    def test_active_customer_count_total_balance_positive(self) -> None:
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"))
        self.assertEqual(rows[0]["active_customer_count"], 5)

    def test_active_customer_count_excludes_zero_balance(self) -> None:
        self._insert_period_summary("2026-04", "5491099", "099", "Khach Zero", 0)
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"))
        self.assertEqual(rows[0]["active_customer_count"], 5)

    def test_active_customer_count_one_customer_once_per_period(self) -> None:
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"))
        self.assertEqual(rows[0]["active_customer_count"], 5)

    def test_active_customer_count_respects_branch_filter(self) -> None:
        self._insert_period_summary("2026-04", "5400001", "001", "Khach Branch", 700, branch_code="5400")
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04", branch_code="5400"))
        self.assertEqual(rows[0]["active_customer_count"], 1)

    def test_active_customer_count_respects_customer_type_filter(self) -> None:
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04", customer_type="TC"))
        self.assertEqual(rows[0]["active_customer_count"], 1)

    def test_active_customer_count_respects_officer_override(self) -> None:
        self.repository.create_officer_override(
            customer_code="5491003",
            effective_from_period="2026-04",
            officer_code="OV1",
            officer_name="Override Officer",
            reason="test",
        )
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04", officer="OV1"))
        self.assertEqual(rows[0]["active_customer_count"], 1)

    def test_active_customer_count_short_term_filter(self) -> None:
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04", loan_term="SHORT_TERM"))
        self.assertEqual(rows[0]["active_customer_count"], 4)

    def test_active_customer_count_medium_long_filter(self) -> None:
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04", loan_term="MEDIUM_LONG_TERM"))
        self.assertEqual(rows[0]["active_customer_count"], 2)

    def test_active_customer_count_other_term_filter(self) -> None:
        self._insert_period_summary("2026-04", "5491888", "888", "Khach Other Term", 100, short=0, other=100)
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04", loan_term="OTHER"))
        self.assertEqual(rows[0]["active_customer_count"], 1)

    def test_active_customer_count_period_sorted(self) -> None:
        rows = self.repository.get_active_customer_count_trend(CustomerFilters())
        self.assertEqual([row["period"] for row in rows], sorted(row["period"] for row in rows))

    def test_active_customer_count_single_period(self) -> None:
        rows = self.repository.get_active_customer_count_trend(CustomerFilters(period_from="2026-04", period_to="2026-04"))
        dataset = chart_service.dashboard_chart_dataset_active_customer_count(rows)
        self.assertEqual(dataset[0][1], (("2026-04", 5.0),))

    def test_active_customer_count_no_data_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerDashboardTab(self.repository, lambda: CustomerFilters(current_period="2099-01"))
        tab._apply_customer_count_chart([])
        self.assertEqual(tab.customer_count_chart.state, "empty")
        self.assertEqual(tab.customer_count_chart.state_label.text(), "Không có khách hàng còn dư nợ phù hợp với bộ lọc.")
        tab.close()

    def test_customer_chart_uses_integer_axis(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("Số lượng", value_kind="number")
        chart.set_series((("Số khách hàng còn dư nợ", (("2026-04", 5),)),))
        value_axes = [axis for axis in chart.chart.axes(Qt.Orientation.Vertical) if isinstance(axis, QValueAxis)]
        self.assertTrue(value_axes)
        self.assertEqual(value_axes[0].labelFormat(), "%d")
        chart.deleteLater()

    def test_dashboard_top_balance_dataset(self) -> None:
        rows = self.repository.top_customers(CustomerFilters(current_period="2026-04"), limit=2)
        dataset = chart_service.dashboard_top_balance_dataset(rows)
        self.assertEqual(dataset[0]["customer_code"], "5491006")

    def test_dashboard_top_increase_dataset(self) -> None:
        rows = self.repository.top_movement_customers(
            "2026-03",
            "2026-04",
            CustomerFilters(),
            movement_status="Tăng dư nợ",
            limit=10,
        )
        dataset = chart_service.dashboard_top_increase_dataset(rows)
        self.assertEqual([row["customer_code"] for row in dataset], ["5491001"])

    def test_dashboard_top_decrease_dataset(self) -> None:
        rows = self.repository.top_movement_customers(
            "2026-03",
            "2026-04",
            CustomerFilters(),
            movement_status="Giảm dư nợ",
            limit=10,
        )
        dataset = chart_service.dashboard_top_decrease_dataset(rows)
        self.assertEqual([row["customer_code"] for row in dataset], ["5491004"])

    def test_chart_periods_are_sorted(self) -> None:
        trends = self.repository.dashboard_trends(CustomerFilters())
        dataset = chart_service.dashboard_chart_dataset_balance_trend(trends)
        self.assertTrue(chart_service.chart_periods_are_sorted(dataset))

    def test_chart_money_axis_formatter(self) -> None:
        self.assertEqual(format_money_axis(1_234_567_890), "1,23 tỷ")

    def test_chart_money_full_formatter(self) -> None:
        self.assertEqual(format_money_full(-1_234_567_890), "-1.234.567.890")

    def test_chart_percentage_formatter(self) -> None:
        self.assertEqual(format_percentage(2.3567), "2,36%")

    def test_chart_no_data_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("No data", value_kind="money")
        chart.set_series(())
        self.assertEqual(chart.state, "empty")
        chart.deleteLater()

    def test_chart_single_period_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("Single period", value_kind="money")
        chart.set_series((("Tổng dư nợ", (("2026-04", 1000.0),)),))
        self.assertEqual(chart.state, "ready")
        chart.deleteLater()

    def test_chart_zero_total_donut(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerDonutChart("Zero")
        chart.set_slices((("Ngắn hạn", 0.0), ("Trung/dài hạn", 0.0)))
        self.assertEqual(chart.state, "empty")
        chart.deleteLater()

    def test_chart_tooltip_contains_full_customer_name(self) -> None:
        tooltip = ChartTooltip(
            label="Khach hang co ten rat dai khong bi cat trong tooltip",
            series_name="Dư nợ",
            value=1000,
            value_kind="money",
            detail="5491001\nKhach hang co ten rat dai khong bi cat trong tooltip",
        )
        self.assertIn("Khach hang co ten rat dai", tooltip.text())

    def test_chart_tooltip_stays_visible_until_chart_leave(self) -> None:
        app = QApplication.instance() or QApplication([])
        chart = CustomerLineChart("Tooltip", value_kind="money")
        chart.resize(520, 340)
        tooltip = ChartTooltip(label="2026-04", series_name="Dư nợ", value=1000, value_kind="money")
        chart.show_tooltip(tooltip, chart.mapToGlobal(chart.rect().center()))
        self.assertFalse(chart.tooltip_label.isHidden())
        chart.eventFilter(chart.chart_view, QEvent(QEvent.Type.MouseMove))
        self.assertFalse(chart.tooltip_label.isHidden())
        chart.eventFilter(chart.chart_view, QEvent(QEvent.Type.Leave))
        self.assertTrue(chart.tooltip_label.isHidden())
        chart.deleteLater()

    def test_customer_detail_chart_datasets(self) -> None:
        history = self.repository.customer_history("5491001")
        datasets = chart_service.customer_detail_chart_datasets(history)
        self.assertIn("balance", datasets)
        self.assertEqual(datasets["compare_money"][0][1][-1], ("2026-04", 500.0))

    def test_chart_values_match_dashboard_kpi(self) -> None:
        metrics = self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04"))
        trends = self.repository.dashboard_trends(CustomerFilters(current_period="2026-04"))
        dataset = chart_service.dashboard_chart_dataset_balance_trend(trends)
        self.assertEqual(dataset[0][1][-1][1], float(metrics["total_balance"]))

    def test_chart_uses_weighted_nim(self) -> None:
        trends = self.repository.dashboard_trends(CustomerFilters(current_period="2026-04"))
        dataset = chart_service.dashboard_chart_dataset_nim_rates(trends)
        nim_before = dataset[1][1][-1][1]
        self.assertAlmostEqual(nim_before, 6.5)
        self.assertNotAlmostEqual(nim_before, 6.0)

    def test_dashboard_query_runs_outside_ui_thread(self) -> None:
        _controller, payload = self._worker_probe(
            "dashboard_query",
            lambda: self.repository.dashboard_metrics(CustomerFilters(current_period="2026-04")),
        )
        self.assertIn("customer_count", payload["value"])

    def test_movement_query_runs_outside_ui_thread(self) -> None:
        _controller, payload = self._worker_probe(
            "movement_query",
            lambda: self.repository.movement_rows("2026-03", "2026-04", CustomerFilters()).total_rows,
        )
        self.assertEqual(payload["value"], 6)

    def test_multiple_officers_query_runs_outside_ui_thread(self) -> None:
        _controller, payload = self._worker_probe(
            "multiple_officers_query",
            lambda: self.repository.multiple_officer_rows(CustomerFilters(current_period="2026-04")).total_rows,
        )
        self.assertEqual(payload["value"], 1)

    def test_customer_detail_query_runs_outside_ui_thread(self) -> None:
        _controller, payload = self._worker_probe(
            "customer_detail_query",
            lambda: len(self.repository.customer_history("5491001")),
        )
        self.assertEqual(payload["value"], 2)

    def test_stale_request_does_not_update_ui(self) -> None:
        controller = AsyncQueryController()
        applied: list[str] = []
        controller.run("slow", lambda: (sleep(0.08), "old")[1], applied.append, use_cache=False)
        controller.run("fast", lambda: "new", applied.append, use_cache=False)
        self.assertTrue(_wait_until(lambda: applied == ["new"]))
        controller.wait_for_idle()
        self.assertTrue(_wait_until(lambda: controller.stale_result_count >= 1))
        self.assertEqual(applied, ["new"])
        self.assertGreaterEqual(controller.stale_result_count, 1)

    def test_latest_request_updates_ui(self) -> None:
        controller = AsyncQueryController()
        applied: list[str] = []
        controller.run("latest", lambda: "value", applied.append, use_cache=False)
        self.assertTrue(_wait_until(lambda: applied == ["value"]))
        controller.wait_for_idle()
        self.assertEqual(controller.latest_applied_generation, 1)

    def test_filter_debounce_creates_single_query(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        count = {"value": 0}
        window._filter_timer.timeout.connect(lambda: count.__setitem__("value", count["value"] + 1))
        for _index in range(5):
            window._filter_changed()
        self.assertTrue(_wait_until(lambda: count["value"] == 1, timeout=1.5))
        self.assertEqual(count["value"], 1)
        window.close()

    def test_linked_combobox_update_does_not_repeat_query(self) -> None:
        app = QApplication.instance() or QApplication([])
        window = CustomerManagementWindow(self.main_database_path)
        window._filter_timer.stop()
        window._updating_filters = True
        window._filter_changed()
        self.assertFalse(window._filter_timer.isActive())
        window._updating_filters = False
        window.close()

    def test_loading_state_visible_during_query(self) -> None:
        banner = QueryStateBanner()
        banner.set_loading()
        self.assertEqual(banner.state, "loading")
        banner.deleteLater()

    def test_no_data_state(self) -> None:
        banner = QueryStateBanner()
        banner.set_empty()
        self.assertEqual(banner.state, "empty")
        banner.deleteLater()

    def test_error_state_and_retry(self) -> None:
        banner = QueryStateBanner()
        triggered = {"value": False}
        banner.retryRequested.connect(lambda: triggered.__setitem__("value", True))
        banner.set_error("Lỗi thử nghiệm")
        banner.retry_button.click()
        self.assertEqual(banner.state, "error")
        self.assertTrue(triggered["value"])
        banner.deleteLater()

    def test_window_remains_responsive_during_slow_query(self) -> None:
        controller = AsyncQueryController()
        applied: list[str] = []
        controller.run("slow", lambda: (sleep(0.15), "done")[1], applied.append, use_cache=False)
        processed = 0
        deadline = perf_counter() + 1.0
        app = QApplication.instance() or QApplication([])
        while not applied and perf_counter() < deadline:
            app.processEvents()
            processed += 1
            sleep(0.01)
        controller.wait_for_idle()
        self.assertEqual(applied, ["done"])
        self.assertGreater(processed, 2)

    def test_customer_query_explain_plan_available(self) -> None:
        plans = self.repository.explain_customer_query_plans(
            CustomerFilters(current_period="2026-04"),
            previous_period="2026-03",
            current_period="2026-04",
        )
        self.assertIn("customer_list_page", plans)
        self.assertTrue(any("INDEX" in detail.upper() or "SEARCH" in detail.upper() for detail in plans["customer_list_page"]))

    def test_customer_ui_cache_is_limited(self) -> None:
        cache = LruQueryCache(max_entries=3)
        for index in range(5):
            cache.set(("key", index), index)
        self.assertEqual(len(cache), 3)
        self.assertIsNone(cache.get(("key", 0), None))
        self.assertEqual(cache.get(("key", 4)), 4)

    def test_movement_full_outer_join_new_customer(self) -> None:
        self.assertEqual(self._movement_statuses()["5491003"], "Vay mới")

    def test_movement_full_outer_join_paid_off_customer(self) -> None:
        self.assertEqual(self._movement_statuses()["5491002"], "Tất toán")

    def test_movement_full_outer_join_increase(self) -> None:
        self.assertEqual(self._movement_statuses()["5491001"], "Tăng dư nợ")

    def test_movement_full_outer_join_decrease(self) -> None:
        self.assertEqual(self._movement_statuses()["5491004"], "Giảm dư nợ")

    def test_movement_full_outer_join_unchanged(self) -> None:
        self.assertEqual(self._movement_statuses()["5491005"], "Không thay đổi")

    def test_movement_query_limits_periods_before_join(self) -> None:
        sql, params = _movement_base_sql("2026-03", "2026-04", resolve_officer=False)
        self.assertIn("FROM customer_period_summary\n            WHERE period = ?", sql)
        self.assertEqual(params, ["2026-03", "2026-04"])

    def test_movement_query_uses_union_all(self) -> None:
        sql, _params = _movement_base_sql("2026-03", "2026-04", resolve_officer=False)
        self.assertIn("UNION ALL", sql)
        self.assertNotIn("\n            UNION\n", sql)

    def test_movement_page_query_uses_limit_offset(self) -> None:
        statements: list[str] = []
        with closing(self.repository.connect()) as connection:
            connection.set_trace_callback(statements.append)
            with patch.object(self.repository, "connect", return_value=connection):
                self.repository.movement_rows("2026-03", "2026-04", CustomerFilters(), page=2, page_size=2)
        page_sql = " ".join(statements).upper()
        self.assertIn("LIMIT", page_sql)
        self.assertIn("OFFSET", page_sql)

    def test_movement_does_not_load_all_rows_for_page(self) -> None:
        payload = self.repository.movement_payload("2026-03", "2026-04", CustomerFilters(), page=1, page_size=2)
        result = payload["result"]
        self.assertEqual(len(result.rows), 2)
        self.assertGreater(result.total_rows, len(result.rows))
        self.assertEqual(
            [stage["rows"] for stage in payload["stage_stats"] if stage["stage"] == "page"],
            [2],
        )

    def test_movement_count_query_is_separate(self) -> None:
        statements: list[str] = []
        with closing(self.repository.connect()) as connection:
            connection.set_trace_callback(statements.append)
            with patch.object(self.repository, "connect", return_value=connection):
                self.repository.movement_payload("2026-03", "2026-04", CustomerFilters(), page=1, page_size=2)
        self.assertTrue(any("SELECT COUNT(*) FROM temp_customer_movements" in statement for statement in statements))

    def test_movement_candidate_not_recomputed_for_kpi_count_page_when_materialized(self) -> None:
        payload = self.repository.movement_payload("2026-03", "2026-04", CustomerFilters(), page=1, page_size=2)
        stages = [stage["stage"] for stage in payload["stage_stats"]]
        self.assertEqual(stages.count("candidate_join"), 1)
        self.assertIn("kpi", stages)
        self.assertIn("count", stages)
        self.assertIn("page", stages)

    def test_movement_no_n_plus_one_officer_queries(self) -> None:
        statements: list[str] = []
        with closing(self.repository.connect()) as connection:
            connection.set_trace_callback(statements.append)
            with patch.object(self.repository, "connect", return_value=connection):
                self.repository.movement_payload("2026-03", "2026-04", CustomerFilters(), page=1, page_size=4)
        override_selects = [
            statement
            for statement in statements
            if "FROM customer_officer_override" in statement and "customer_code IN" in statement
        ]
        self.assertLessEqual(len(override_selects), 1)

    def test_movement_officer_resolution_is_bulk(self) -> None:
        statements: list[str] = []
        with closing(self.repository.connect()) as connection:
            connection.set_trace_callback(statements.append)
            with patch.object(self.repository, "connect", return_value=connection):
                rows = self.repository.movement_rows("2026-03", "2026-04", CustomerFilters(), page=1, page_size=3).rows
        self.assertEqual(len(rows), 3)
        self.assertTrue(any("customer_code IN" in statement for statement in statements if "customer_officer_override" in statement))

    def test_movement_unit_directory_uses_cache(self) -> None:
        rows = self.repository.movement_rows(
            "2026-03",
            "2026-04",
            CustomerFilters(),
            page=1,
            page_size=2,
        ).rows
        self.assertTrue(all("branch_display" in row for row in rows))
        self.assertTrue(all(str(row["branch_display"]).startswith(str(row["branch_code"])) for row in rows))

    def test_movement_table_render_has_no_cell_widgets(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        self.assertIsInstance(tab.table, QTableView)
        self.assertEqual(tab.findChildren(QTableWidget), [])
        tab.close()

    def test_movement_does_not_resize_columns_to_contents_each_refresh(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        payload = tab._load_payload("2026-03", "2026-04", CustomerFilters(), 1, 2, "difference", True)
        with patch.object(tab.table, "resizeColumnsToContents", side_effect=AssertionError("no auto resize")):
            tab._apply_payload(payload)
        tab.close()

    def test_movement_worker_creates_connection_in_run(self) -> None:
        controller, payload = self._worker_probe(
            "movement_connection_thread",
            lambda: self.repository.movement_rows("2026-03", "2026-04", CustomerFilters()).total_rows,
        )
        self.assertEqual(payload["value"], 6)
        self.assertNotEqual(controller.last_worker_thread_id, controller.last_ui_thread_id)

    def test_movement_worker_closes_connection_in_finally(self) -> None:
        connection = self.repository.connect()
        closed = {"value": False}

        class TrackingConnection:
            def __enter__(self):
                connection.__enter__()
                return self

            def __exit__(self, exc_type, exc, traceback):
                return connection.__exit__(exc_type, exc, traceback)

            def execute(self, *args, **kwargs):
                return connection.execute(*args, **kwargs)

            def close(self):
                closed["value"] = True
                return connection.close()

        with patch.object(self.repository, "connect", return_value=TrackingConnection()):
            self.repository.movement_rows("2026-03", "2026-04", CustomerFilters(), page=1, page_size=1)
        self.assertTrue(closed["value"])

    def test_movement_worker_returns_python_data_only(self) -> None:
        payload = self.repository.movement_payload("2026-03", "2026-04", CustomerFilters(), page=1, page_size=2)
        self.assertIsInstance(payload["kpis"], dict)
        self.assertIsInstance(payload["result"].rows[0], dict)

    def test_movement_model_updates_on_ui_thread(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        thread_ids: list[int] = []
        original = tab.model.set_rows

        def tracking_set_rows(rows):
            thread_ids.append(threading.get_ident())
            original(rows)

        tab.model.set_rows = tracking_set_rows
        payload = tab._load_payload("2026-03", "2026-04", CustomerFilters(), 1, 2, "difference", True)
        tab._apply_payload(payload)
        self.assertEqual(thread_ids, [threading.get_ident()])
        tab.close()

    def test_movement_stale_result_ignored(self) -> None:
        self.test_stale_request_does_not_update_ui()

    def test_movement_stale_request_releases_loading(self) -> None:
        controller = AsyncQueryController()
        states: list[str] = []
        controller.run("slow", lambda: (sleep(0.05), "old")[1], lambda _payload: None, use_cache=False, state_callback=lambda state, _message: states.append(state))
        controller.cancel_pending()
        self.assertTrue(_wait_until(lambda: states[-1:] == ["ready"]))
        controller.wait_for_idle()

    def test_movement_only_latest_generation_updates_ui(self) -> None:
        self.test_latest_request_updates_ui()

    def test_movement_window_close_while_loading(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        tab.query_controller.run("slow_close", lambda: (sleep(0.05), "done")[1], lambda _payload: None, use_cache=False)
        tab.close()
        app.processEvents()
        tab.query_controller.wait_for_idle()

    def test_movement_rapid_filter_changes(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        calls = {"value": 0}

        def counted_refresh(*args, **kwargs):
            calls["value"] += 1

        tab.refresh = counted_refresh
        for _index in range(30):
            tab._filter_changed()
        self.assertEqual(calls["value"], 30)
        tab.close()

    def test_movement_rapid_page_changes(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        calls = {"value": 0}
        tab.refresh = lambda *args, **kwargs: calls.__setitem__("value", calls["value"] + 1)
        for index in range(20):
            tab._page_changed(index + 1)
        self.assertEqual(calls["value"], 20)
        tab.close()

    def test_movement_worker_error_finishes_loading(self) -> None:
        controller = AsyncQueryController()
        states: list[str] = []
        errors: list[str] = []
        controller.run(
            "movement_error",
            lambda: (_ for _item in ()).throw(RuntimeError("boom")),
            lambda _payload: None,
            lambda exc: errors.append(str(exc)),
            use_cache=False,
            state_callback=lambda state, _message: states.append(state),
        )
        self.assertTrue(_wait_until(lambda: states[-1:] == ["error"]))
        controller.wait_for_idle()
        self.assertEqual(errors, ["boom"])

    def test_movement_database_locked_does_not_retry_forever(self) -> None:
        controller = AsyncQueryController()
        errors: list[str] = []
        with patch.object(self.repository, "connect", side_effect=sqlite3.OperationalError("database is locked")):
            controller.run(
                "movement_locked",
                lambda: self.repository.movement_rows("2026-03", "2026-04", CustomerFilters()).total_rows,
                lambda _payload: None,
                lambda exc: errors.append(str(exc)),
                use_cache=False,
            )
            self.assertTrue(_wait_until(lambda: bool(errors)))
        controller.wait_for_idle()
        self.assertIn("database is locked", errors[0])

    def test_movement_filter_repopulation_uses_signal_blocker(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        calls = {"value": 0}
        tab.refresh = lambda *args, **kwargs: calls.__setitem__("value", calls["value"] + 1)
        tab._updating_filters = True
        tab._filter_changed()
        tab._updating_filters = False
        self.assertEqual(calls["value"], 0)
        tab.close()

    def test_movement_one_filter_action_one_request(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        calls = {"value": 0}
        tab.refresh = lambda *args, **kwargs: calls.__setitem__("value", calls["value"] + 1)
        tab._filter_changed()
        self.assertEqual(calls["value"], 1)
        tab.close()

    def test_movement_search_debounce(self) -> None:
        app = QApplication.instance() or QApplication([])
        search = SearchBox()
        emitted: list[str] = []
        search.debouncedTextChanged.connect(emitted.append)
        for value in ("5", "54", "5491"):
            search.setText(value)
        self.assertTrue(_wait_until(lambda: emitted == ["5491"], timeout=1.0))
        search.deleteLater()

    def test_movement_page_change_does_not_requery_kpi_unnecessarily(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        tab.previous_combo.addItem("2026-03", "2026-03")
        tab.current_combo.addItem("2026-04", "2026-04")
        tab.previous_combo.setCurrentIndex(tab.previous_combo.findData("2026-03"))
        tab.current_combo.setCurrentIndex(tab.current_combo.findData("2026-04"))
        with patch.object(tab.repository, "movement_kpis", side_effect=AssertionError("no separate KPI query")):
            tab._load_payload("2026-03", "2026-04", CustomerFilters(), 2, 2, "difference", True)
        tab.close()

    def test_movement_export_runs_in_background(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        tab.previous_combo.addItem("2026-03", "2026-03")
        tab.current_combo.addItem("2026-04", "2026-04")
        tab.previous_combo.setCurrentIndex(tab.previous_combo.findData("2026-03"))
        tab.current_combo.setCurrentIndex(tab.current_combo.findData("2026-04"))
        path = self.root / "movement-background.xlsx"
        with (
            patch("PySide6.QtWidgets.QFileDialog.getSaveFileName", return_value=(str(path), "Excel (*.xlsx)")),
            patch.object(tab.export_controller, "run", return_value=1) as run,
        ):
            tab.export_excel()
        self.assertTrue(run.called)
        tab.close()

    def test_movement_export_all_filtered_rows(self) -> None:
        self.test_comparison_export_all_filtered_rows()

    def test_movement_export_does_not_use_current_page_model(self) -> None:
        app = QApplication.instance() or QApplication([])
        tab = CustomerMovementTab(self.repository, lambda: CustomerFilters())
        tab.model.set_rows([])
        path = self.root / "movement-direct-export.xlsx"
        export_customer_growth(self.repository, "2026-03", "2026-04", CustomerFilters(), path)
        worksheet = load_workbook(path, data_only=True)["BienDongDuNo"]
        self.assertEqual(worksheet.max_row - 1, 6)
        tab.close()

    def test_movement_export_connection_created_in_worker(self) -> None:
        path = self.root / "movement-worker-export.xlsx"
        _controller, payload = self._worker_probe(
            "movement_export_worker",
            lambda: str(export_customer_growth(self.repository, "2026-03", "2026-04", CustomerFilters(), path)),
        )
        self.assertTrue(Path(payload["value"]).is_file())

    def _empty_main_database_path(self, name: str) -> Path:
        root = self.root / name
        root.mkdir(parents=True, exist_ok=True)
        return root / "DuLieuV3.db"

    def _delete_all_customer_periods(self, repository: CustomerRepository) -> None:
        for period in list(repository.distinct_periods()):
            repository.delete_customer_period(period)

    def _write_single_customer_ftpln(self, folder: Path, date_suffix: str) -> Path:
        folder.mkdir(parents=True, exist_ok=True)
        path = folder / f"5491_FTPLN_{date_suffix}.csv"
        path.write_text(
            "\n".join(
                [
                    FTPLN_HEADER,
                    "5491,2,10,1,[001] Officer A,00,1000,TR1,CN,DN1,'001,Khach A",
                ]
            ),
            encoding="utf-8",
        )
        return path

    def _worker_probe(self, name: str, function):
        ui_thread = threading.get_ident()
        controller = AsyncQueryController()
        output: dict[str, object] = {}

        def task():
            return {"thread_id": threading.get_ident(), "value": function()}

        controller.run(name, task, lambda payload: output.update(payload), use_cache=False)
        self.assertTrue(_wait_until(lambda: "thread_id" in output))
        controller.wait_for_idle()
        self.assertNotEqual(output["thread_id"], ui_thread)
        return controller, output

    def _movement_rows(self) -> list[dict[str, object]]:
        return self.repository.movement_rows(
            "2026-03",
            "2026-04",
            CustomerFilters(),
            page=1,
            page_size=100,
            sort_by="customer_code",
            sort_desc=False,
        ).rows

    def _movement_statuses(self) -> dict[str, str]:
        return {row["customer_code"]: row["movement_status"] for row in self._movement_rows()}

    def _insert_cross_branch_fixture(
        self,
        *,
        period: str = "2026-04",
        sequence: str = "777",
        branches: tuple[str, ...] = ("5400", "5491"),
        balances: tuple[float, ...] = (1000, 3000),
        shorts: tuple[float, ...] | None = None,
        mediums: tuple[float, ...] | None = None,
        names: tuple[str, ...] | None = None,
        customer_types: tuple[str, ...] | None = None,
        average_rates: tuple[float, ...] | None = None,
        nim_befores: tuple[float, ...] | None = None,
        nim_afters: tuple[float, ...] | None = None,
        officer_codes: tuple[str, ...] | None = None,
        officer_names: tuple[str, ...] | None = None,
        trctcds: tuple[str, ...] | None = None,
    ) -> None:
        for index, branch in enumerate(branches):
            total = balances[index] if index < len(balances) else balances[-1]
            medium = mediums[index] if mediums is not None and index < len(mediums) else 0
            short = shorts[index] if shorts is not None and index < len(shorts) else total - medium
            self._insert_period_summary(
                period,
                f"{branch}{sequence}",
                sequence,
                names[index] if names is not None and index < len(names) else f"Khach Lien Chi Nhanh {sequence}",
                total,
                branch_code=branch,
                customer_type=customer_types[index] if customer_types is not None and index < len(customer_types) else "CN",
                short=short,
                medium=medium,
                average_rate=average_rates[index] if average_rates is not None and index < len(average_rates) else 10,
                nim_before=nim_befores[index] if nim_befores is not None and index < len(nim_befores) else 8,
                nim_after=nim_afters[index] if nim_afters is not None and index < len(nim_afters) else 7,
                officer_code=officer_codes[index] if officer_codes is not None and index < len(officer_codes) else f"CB{index + 1:03d}",
                officer_name=officer_names[index] if officer_names is not None and index < len(officer_names) else f"Officer Cross {index + 1}",
                trctcd=trctcds[index] if trctcds is not None and index < len(trctcds) else "00",
            )

    def _insert_customer_with_offices(
        self,
        *,
        period: str = "2026-04",
        sequence: str = "900",
        branch_code: str = "5405",
        offices: tuple[tuple[object, ...], ...] = (("00", 1000), ("01", 2000)),
        average_rate: float = 10,
        nim_before: float = 8,
        nim_after: float = 7,
        customer_name: str | None = None,
    ) -> None:
        customer_code = f"{branch_code}{sequence}"
        total = sum(float(office[1]) for office in offices)
        self._insert_period_summary(
            period,
            customer_code,
            sequence,
            customer_name or f"Khach Multi Unit {sequence}",
            total,
            branch_code=branch_code,
            average_rate=average_rate,
            nim_before=nim_before,
            nim_after=nim_after,
            insert_office=False,
        )
        with closing(self.repository.connect()) as connection:
            connection.execute(
                "DELETE FROM customer_office_period WHERE period = ? AND customer_code = ?",
                (period, customer_code),
            )
            for index, office in enumerate(offices, start=1):
                trctcd = str(office[0])
                balance = float(office[1])
                row_average_rate = float(office[2]) if len(office) > 2 else average_rate
                row_nim_before = float(office[3]) if len(office) > 3 else nim_before
                row_nim_after = float(office[4]) if len(office) > 4 else nim_after
                clean_trctcd = normalize_trctcd(trctcd)
                office_code = build_office_code(branch_code, clean_trctcd)
                office_type = classify_office_type(clean_trctcd).value
                office_name = "Hội sở" if clean_trctcd == "00" else f"PGD {clean_trctcd}"
                connection.execute(
                    """
                    INSERT INTO customer_office_period(
                        period, customer_code, customer_sequence, branch_code,
                        trctcd, office_code, office_name, office_type,
                        primary_officer_code, primary_officer_name, officer_count,
                        total_balance, short_term_balance, medium_long_term_balance,
                        other_balance, interest_rate_numerator, nim_before_numerator,
                        nim_after_numerator, source_loan_count, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, 0, 0, ?, ?, ?, 1, 'now', 'now')
                    """,
                    (
                        period,
                        customer_code,
                        sequence,
                        branch_code,
                        clean_trctcd,
                        office_code,
                        office_name,
                        office_type,
                        f"OF{index:03d}",
                        f"Officer Office {index}",
                        balance,
                        balance,
                        row_average_rate * balance,
                        row_nim_before * balance,
                        row_nim_after * balance,
                    ),
                )
            connection.commit()

    def _insert_active_periods(self, periods: list[str]) -> None:
        for index, period in enumerate(periods, start=1):
            self._insert_period_summary(
                period,
                f"RANGE{index:05d}",
                f"R{index:05d}",
                f"Khach Range {index}",
                1000 + index,
                average_rate=7 + index / 10,
                nim_before=5 + index / 10,
                nim_after=4 + index / 10,
            )

    def _dashboard_trend_header_row(self, worksheet) -> int:
        for row in range(1, worksheet.max_row + 1):
            if worksheet.cell(row, 1).value == "Xu hướng theo kỳ":
                return row + 1
        self.fail("Không tìm thấy phần Xu hướng theo kỳ trong file export Dashboard.")

    def _dashboard_export_periods(self, worksheet) -> list[str]:
        header_row = self._dashboard_trend_header_row(worksheet)
        return [
            str(worksheet.cell(row, 1).value)
            for row in range(header_row + 1, worksheet.max_row + 1)
            if worksheet.cell(row, 1).value
        ]

    def _seed_phase_c_fixture(self) -> None:
        run_202603 = self._insert_import_run("2026-03")
        run_202604 = self._insert_import_run("2026-04")
        self._insert_period_summary("2026-03", "5491001", "001", "Khach A", 1000, run_id=run_202603, average_rate=10, nim_before=8, nim_after=7)
        self._insert_period_summary("2026-03", "5491002", "002", "Khach B", 2000, run_id=run_202603, customer_type="TC", short=0, medium=2000, average_rate=6, nim_before=4, nim_after=3, officer_code="002", officer_name="Officer B")
        self._insert_period_summary("2026-03", "5491004", "004", "Khach D", 1000, run_id=run_202603, average_rate=5, nim_before=3, nim_after=2, officer_code="004", officer_name="Officer D")
        self._insert_period_summary("2026-03", "5491005", "005", "Khach E", 1000, run_id=run_202603, average_rate=7, nim_before=5, nim_after=4, officer_code="005", officer_name="Officer E")
        self._insert_period_summary("2026-04", "5491001", "001", "Khach A", 1500, run_id=run_202604, average_rate=10, nim_before=8, nim_after=7)
        self._insert_period_summary("2026-04", "5491003", "003", "Khach C", 3000, run_id=run_202604, customer_type="TC", short=0, medium=3000, average_rate=8, nim_before=6, nim_after=5, officer_code="003", officer_name="Officer C")
        self._insert_period_summary("2026-04", "5491004", "004", "Khach D", 500, run_id=run_202604, average_rate=6, nim_before=4, nim_after=3, officer_code="044", officer_name="Officer D2")
        self._insert_period_summary("2026-04", "5491005", "005", "Khach E", 1000, run_id=run_202604, average_rate=7, nim_before=5, nim_after=4, officer_code="005", officer_name="Officer E")
        self._insert_period_summary(
            "2026-04",
            "5491006",
            "006",
            "Khach M",
            4000,
            run_id=run_202604,
            short=3000,
            medium=1000,
            average_rate=9,
            nim_before=7,
            nim_after=6,
            officer_code="006",
            officer_name="Officer M1",
            officer_rows=(("006", "Officer M1", 3000), ("066", "Officer M2", 1000)),
        )
        self._refresh_master_rows()

    def _insert_import_run(self, period: str) -> int:
        with closing(self.repository.connect()) as connection:
            cursor = connection.execute(
                """
                INSERT INTO customer_import_runs(
                    period, source_folder, data_type, file_count, source_row_count,
                    customer_count, started_at, completed_at, status, created_by, computer_name
                )
                VALUES (?, ?, 'DN', 1, 10, 5, 'start', 'done', 'COMPLETED', 'tester', 'machine')
                """,
                (period, str(self.root)),
            )
            run_id = int(cursor.lastrowid)
            connection.execute(
                """
                INSERT INTO customer_import_files(
                    run_id, file_name, file_path, file_hash, branch_code, period,
                    source_row_count, customer_count, status
                )
                VALUES (?, ?, ?, ?, '5491', ?, 10, 5, 'COMPLETED')
                """,
                (run_id, f"5491_FTPLN_{period}.csv", str(self.root), f"hash-{period}", period),
            )
            connection.commit()
        return run_id

    def _insert_period_summary(
        self,
        period: str,
        customer_code: str,
        customer_sequence: str,
        customer_name: str,
        total: float,
        *,
        run_id: int | None = None,
        branch_code: str = "5491",
        customer_type: str = "CN",
        short: float | None = None,
        medium: float = 0,
        other: float = 0,
        average_rate: float = 10,
        nim_before: float = 8,
        nim_after: float = 7,
        officer_code: str = "001",
        officer_name: str = "Officer A",
        officer_rows: tuple[tuple[str, str, float], ...] | None = None,
        trctcd: str = "00",
        office_name: str | None = None,
        insert_office: bool = True,
    ) -> None:
        short_balance = total - medium - other if short is None else short
        medium_long_ratio = (medium / total * 100) if total else 0
        officer_rows = officer_rows or ((officer_code, officer_name, total),)
        clean_trctcd = normalize_trctcd(trctcd)
        office_code = build_office_code(branch_code, clean_trctcd)
        office_type = classify_office_type(clean_trctcd).value
        office_label = office_name if office_name is not None else "Hội sở" if clean_trctcd == "00" else f"PGD {clean_trctcd or 'UNKNOWN'}"
        with closing(self.repository.connect()) as connection:
            connection.execute(
                """
                INSERT OR REPLACE INTO customer_period_summary(
                    run_id, period, customer_code, branch_code, customer_sequence,
                    customer_name, customer_type, primary_officer_code,
                    primary_officer_name, officer_count, has_multiple_officers,
                    total_balance, short_term_balance, medium_long_term_balance,
                    other_balance, medium_long_ratio, interest_rate_numerator,
                    nim_before_numerator, nim_after_numerator, average_rate,
                    nim_before, nim_after, source_loan_count, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'now', 'now')
                """,
                (
                    run_id,
                    period,
                    customer_code,
                    branch_code,
                    customer_sequence,
                    customer_name,
                    customer_type,
                    officer_rows[0][0],
                    officer_rows[0][1],
                    len(officer_rows),
                    1 if len(officer_rows) > 1 else 0,
                    total,
                    short_balance,
                    medium,
                    other,
                    medium_long_ratio,
                    average_rate * total,
                    nim_before * total,
                    nim_after * total,
                    average_rate,
                    nim_before,
                    nim_after,
                    len(officer_rows),
                ),
            )
            connection.execute(
                "DELETE FROM customer_officer_period WHERE period = ? AND customer_code = ?",
                (period, customer_code),
            )
            for index, (code, name, balance) in enumerate(officer_rows):
                connection.execute(
                    """
                    INSERT INTO customer_officer_period(
                        period, customer_code, officer_code, officer_name,
                        balance_managed, source_loan_count, interest_rate_numerator,
                        nim_before_numerator, nim_after_numerator, is_primary, created_at
                    )
                    VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, ?, 'now')
                    """,
                    (
                        period,
                        customer_code,
                        code,
                        name,
                        balance,
                        average_rate * balance,
                        nim_before * balance,
                        nim_after * balance,
                        1 if index == 0 else 0,
                    ),
                )
                connection.execute(
                    """
                    INSERT INTO customer_officer_directory(
                        officer_code, officer_name, branch_code, transaction_office, is_active, updated_at
                    )
                    VALUES (?, ?, ?, '', 1, 'now')
                    ON CONFLICT(officer_code) DO UPDATE SET officer_name = excluded.officer_name
                    """,
                    (code, name, branch_code),
                )
            if insert_office:
                connection.execute(
                    """
                    INSERT OR REPLACE INTO customer_office_period(
                        run_id, period, customer_code, customer_sequence, branch_code,
                        trctcd, office_code, office_name, office_type,
                        primary_officer_code, primary_officer_name, officer_count,
                        total_balance, short_term_balance, medium_long_term_balance,
                        other_balance, interest_rate_numerator, nim_before_numerator,
                        nim_after_numerator, source_loan_count, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'now', 'now')
                    """,
                    (
                        run_id,
                        period,
                        customer_code,
                        customer_sequence,
                        branch_code,
                        clean_trctcd,
                        office_code,
                        office_label,
                        office_type,
                        officer_rows[0][0],
                        officer_rows[0][1],
                        len(officer_rows),
                        total,
                        short_balance,
                        medium,
                        other,
                        average_rate * total,
                        nim_before * total,
                        nim_after * total,
                        len(officer_rows),
                    ),
                )
            connection.commit()

    def _refresh_master_rows(self) -> None:
        with closing(self.repository.connect()) as connection:
            codes = [
                str(row[0])
                for row in connection.execute("SELECT DISTINCT customer_code FROM customer_period_summary").fetchall()
            ]
            for code in codes:
                latest = connection.execute(
                    """
                    SELECT *
                    FROM customer_period_summary
                    WHERE customer_code = ?
                    ORDER BY period DESC, id DESC
                    LIMIT 1
                    """,
                    (code,),
                ).fetchone()
                first_last = connection.execute(
                    "SELECT MIN(period), MAX(period) FROM customer_period_summary WHERE customer_code = ?",
                    (code,),
                ).fetchone()
                connection.execute(
                    """
                    INSERT INTO customer_master(
                        customer_code, branch_code, customer_sequence, customer_name,
                        customer_type, latest_officer_code, latest_officer_name,
                        first_seen_period, last_seen_period, is_active, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 'now', 'now')
                    ON CONFLICT(customer_code) DO UPDATE SET
                        latest_officer_code = excluded.latest_officer_code,
                        latest_officer_name = excluded.latest_officer_name,
                        first_seen_period = excluded.first_seen_period,
                        last_seen_period = excluded.last_seen_period,
                        is_active = 1,
                        updated_at = 'now'
                    """,
                    (
                        code,
                        latest["branch_code"],
                        latest["customer_sequence"],
                        latest["customer_name"],
                        latest["customer_type"],
                        latest["primary_officer_code"],
                        latest["primary_officer_name"],
                        first_last[0],
                        first_last[1],
                    ),
                )
            connection.commit()


if __name__ == "__main__":
    unittest.main()
