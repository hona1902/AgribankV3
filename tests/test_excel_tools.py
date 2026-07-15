from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

from agribank_v3.excel_tools import (
    convert_csv_to_excel,
    convert_csv_to_xlsx,
    list_workbook_sheet_names,
    split_workbook_sheets_to_files,
)


class ExcelToolsTests(unittest.TestCase):
    def test_convert_csv_to_xlsx_preserves_rows(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.csv"
            output = root / "output.xlsx"
            source.write_text("A,B\n1,2\n3,4\n", encoding="utf-8")

            result = convert_csv_to_xlsx(source, output)
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["source"]

            self.assertEqual(result.row_count, 3)
            self.assertEqual(result.column_count, 2)
            self.assertEqual(sheet["A1"].value, "A")
            self.assertEqual(sheet["B3"].value, 4)
            workbook.close()

    def test_convert_csv_to_xlsx_converts_numeric_cells_like_excel(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "5491_rt13.csv"
            output = root / "5491_rt13.xlsx"
            source.write_text(
                "NGAY,MA_CN,MA_KH,SO_DU,MA_TEXT,TEN_KH,EMPTY\n"
                "20251231,5491,00000000,\"115.61                  \",\"'000  \",\"Nguyen Van A  \",\n",
                encoding="utf-8",
            )

            convert_csv_to_xlsx(source, output)
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["5491_rt13"]

            self.assertEqual(sheet["A2"].value, 20251231)
            self.assertEqual(sheet["C2"].value, 0)
            self.assertEqual(sheet["D2"].value, 115.61)
            self.assertEqual(sheet["E2"].value, "'000  ")
            self.assertEqual(sheet["F2"].value, "Nguyen Van A  ")
            self.assertIsNone(sheet["G2"].value)
            self.assertEqual(sheet["D2"].data_type, "n")
            workbook.close()

    def test_convert_csv_to_excel_honors_xlsx_format(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.csv"
            output = root / "custom_name"
            source.write_text("A,B\n1,2\n", encoding="utf-8")

            result = convert_csv_to_excel(source, output, output_format="xlsx")

            self.assertEqual(result.output_path, root / "custom_name.xlsx")
            self.assertTrue(result.output_path.is_file())

    def test_split_workbook_sheets_to_files_creates_one_file_per_sheet(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "book.xlsx"
            output_dir = root / "sheets"
            workbook = Workbook()
            first = workbook.active
            first.title = "Sheet A"
            first["A1"] = "A"
            second = workbook.create_sheet("Sheet B")
            second["B2"] = "B"
            workbook.save(source)
            workbook.close()

            result = split_workbook_sheets_to_files(source, output_dir)

            self.assertEqual(len(result.output_paths), 2)
            self.assertTrue((output_dir / "book_Sheet A.xlsx").is_file())
            self.assertTrue((output_dir / "book_Sheet B.xlsx").is_file())
            split_workbook = load_workbook(output_dir / "book_Sheet B.xlsx", data_only=True)
            self.assertEqual(split_workbook["Sheet B"]["B2"].value, "B")
            split_workbook.close()

    def test_list_workbook_sheet_names_reads_xlsx(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "book.xlsx"
            workbook = Workbook()
            workbook.active.title = "A"
            workbook.create_sheet("B")
            workbook.save(source)
            workbook.close()

            self.assertEqual(list_workbook_sheet_names(source), ("A", "B"))

    def test_split_workbook_sheets_to_files_respects_selected_sheets(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "book.xlsx"
            output_dir = root / "selected"
            workbook = Workbook()
            workbook.active.title = "A"
            workbook.create_sheet("B")
            workbook.create_sheet("C")
            workbook.save(source)
            workbook.close()

            result = split_workbook_sheets_to_files(source, output_dir, sheet_names=("B",))

            self.assertEqual(len(result.output_paths), 1)
            self.assertTrue((output_dir / "book_B.xlsx").is_file())
            self.assertFalse((output_dir / "book_A.xlsx").exists())
            self.assertFalse((output_dir / "book_C.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
