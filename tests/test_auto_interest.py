from __future__ import annotations

from datetime import date
import os
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication

from agribank_v3.features.catalog import SECTIONS
from agribank_v3.features.credit.auto_interest.menu import (
    AUTO_INTEREST_FEATURES,
    AUTO_INTEREST_TITLE,
    CREATE_INTEREST_FILE_TITLE,
    CREATE_REPORT_FILE_TITLE,
    REPORT_FOLDER_SETTINGS_TITLE,
)
from agribank_v3.features.credit.auto_interest.processor import (
    AutoInterestCreateRequest,
    AutoInterestError,
    AutoInterestReportRequest,
    COLLECT_ALL_INTEREST,
    NOT_DUE_AND_OVERDUE_INTEREST,
    NOT_DUE_INTEREST,
    OVERDUE_CENTER_INTEREST,
    build_collect_all_interest_file,
    build_not_due_and_overdue_interest_file,
    build_not_due_interest_file,
    build_overdue_center_interest_file,
    create_auto_interest_file,
    create_auto_interest_report,
    load_deposit_statement,
    load_loan_statement,
    validate_auto_interest_inputs,
    validate_deposit_columns,
    validate_loan_columns,
)
from agribank_v3.features.credit.auto_interest.settings import (
    AutoInterestSettings,
    load_auto_interest_settings,
    save_auto_interest_settings,
)
from agribank_v3.features.credit.auto_interest.placeholder_windows import (
    AutoInterestFolderSettingsWindow,
    CreateAutoInterestFileWindow,
)
from agribank_v3.settings import AppSettingsDatabase


class AutoInterestMigrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.database_path = self.root / "DuLieuV3.db"
        AppSettingsDatabase(self.database_path)
        self.qt_app = QApplication.instance() or QApplication([])
        self.loan_file = self.root / "lnlr23.xlsx"
        self.deposit_file = self.root / "msit.xlsx"
        self.collateral_file = self.root / "collateral.xlsx"
        self._write_loan_file(self.loan_file)
        self._write_msit_file(self.deposit_file)

    def tearDown(self) -> None:
        self.temporary_directory.cleanup()

    def test_credit_menu_has_single_auto_interest_card(self) -> None:
        titles = [feature.title for feature in SECTIONS["Tín dụng"]]
        self.assertEqual(titles.count(AUTO_INTEREST_TITLE), 1)
        self.assertIn(AUTO_INTEREST_TITLE, titles)

    def test_auto_interest_menu_has_3_child_cards(self) -> None:
        titles = [feature.title for feature in AUTO_INTEREST_FEATURES]
        self.assertEqual(
            titles,
            [
                CREATE_INTEREST_FILE_TITLE,
                REPORT_FOLDER_SETTINGS_TITLE,
                CREATE_REPORT_FILE_TITLE,
            ],
        )

    def test_auto_interest_settings_save_load(self) -> None:
        settings = AutoInterestSettings(
            report_folder=self.root / "BaoCao",
            output_folder=self.root / "ThuLai",
            loan_folder=self.root / "Loan",
            deposit_folder=self.root / "TienGui",
            backup_folder=self.root / "Backup",
        )
        save_auto_interest_settings(settings, self.database_path)

        loaded = load_auto_interest_settings(self.database_path)

        self.assertEqual(loaded.report_folder, settings.report_folder)
        self.assertEqual(loaded.output_folder, settings.output_folder)
        self.assertEqual(loaded.loan_folder, settings.loan_folder)
        self.assertEqual(loaded.deposit_folder, settings.deposit_folder)
        self.assertEqual(loaded.backup_folder, settings.backup_folder)

    def test_create_file_window_defaults_to_create_report(self) -> None:
        window = CreateAutoInterestFileWindow(database_path=self.database_path)
        try:
            self.assertTrue(window.create_report_radio.isChecked())
            self.assertFalse(window.skip_report_radio.isChecked())
            self.assertEqual(
                window.collateral_file_edit.placeholderText(),
                "File sao kê tài sản bảo đảm",
            )
        finally:
            window.close()

    def test_create_file_window_weekend_interest_options_only_for_overdue_center(self) -> None:
        window = CreateAutoInterestFileWindow(database_path=self.database_path)
        try:
            self.assertTrue(window.include_weekend_interest_radio.isChecked())
            self.assertFalse(window.include_weekend_interest_radio.isEnabled())
            self.assertFalse(window.exclude_weekend_interest_radio.isEnabled())

            overdue_index = window.collection_mode_combo.findData(OVERDUE_CENTER_INTEREST)
            window.collection_mode_combo.setCurrentIndex(overdue_index)

            self.assertTrue(window.include_weekend_interest_radio.isEnabled())
            self.assertTrue(window.exclude_weekend_interest_radio.isEnabled())
        finally:
            window.close()

    def test_folder_settings_window_shows_report_and_output_folders(self) -> None:
        window = AutoInterestFolderSettingsWindow(database_path=self.database_path)
        try:
            self.assertEqual(
                Path(window.report_folder_edit.text()),
                load_auto_interest_settings(self.database_path).report_folder,
            )
            self.assertEqual(
                Path(window.output_folder_edit.text()),
                load_auto_interest_settings(self.database_path).output_folder,
            )
        finally:
            window.close()

    def test_validate_loan_headers_success(self) -> None:
        data = load_loan_statement(self.loan_file)
        resolved = validate_loan_columns(data)

        self.assertIn("customer_code", resolved.positions)
        self.assertIn("customer_name", resolved.positions)
        self.assertIn("loan_number", resolved.positions)
        self.assertIn("interest_due", resolved.positions)

    def test_validate_loan_headers_missing_required(self) -> None:
        bad_loan = self.root / "bad_lnlr23.xlsx"
        self._write_loan_file(bad_loan, headers=("chonIn", "STT", "MaKH"))

        with self.assertRaises(AutoInterestError):
            validate_auto_interest_inputs(bad_loan, self.deposit_file, "msit")

    def test_validate_deposit_headers_success(self) -> None:
        data = load_deposit_statement(self.deposit_file)
        resolved = validate_deposit_columns(data, "msit")

        self.assertIn("customer_code", resolved.positions)
        self.assertIn("account_number", resolved.positions)
        self.assertIn("balance", resolved.positions)

    def test_build_collect_all_interest_file(self) -> None:
        result = build_collect_all_interest_file(self._request(COLLECT_ALL_INTEREST))

        self.assertEqual(result.row_count, 2)
        self.assertEqual(result.skipped_count, 1)
        self.assertEqual(result.summary["vba_procedure"], "ThuToanBoLai")
        workbook = load_workbook(result.output_file, data_only=False)
        self.assertEqual(workbook.sheetnames[:3], ["FileGoc", "SaoKeTrichLai", "BangKeTheoLo"])
        technical = workbook["SaoKeTrichLai"]
        self.assertEqual(technical["A3"].value, "GN_FUTURE")
        self.assertEqual(technical["B3"].value, "01")
        self.assertEqual(technical["D3"].value, "20260731")
        self.assertEqual(technical["E3"].value, "123456789")
        ledger = workbook["BangKeTheoLo"]
        self.assertEqual(ledger["A1"].value, "DANH SÁCH TRÍCH LÃI BÁN TỰ ĐỘNG")
        self.assertEqual(ledger["D4"].value, "GN_FUTURE")
        self.assertEqual(ledger["F4"].value, 100000)
        self.assertEqual(ledger["G4"].value, "31/07/2026")
        self.assertEqual(ledger["H4"].value, "123456789")
        workbook.close()

    def test_collect_all_interest_requires_positive_remaining_balance(self) -> None:
        self._write_msit_file(self.deposit_file, customer_a_balance=300000)

        result = build_collect_all_interest_file(self._request(COLLECT_ALL_INTEREST))

        self.assertEqual(result.row_count, 0)
        self.assertEqual(result.skipped_count, 3)

    def test_not_due_interest_allows_zero_remaining_balance(self) -> None:
        self._write_msit_file(self.deposit_file, customer_a_balance=100000)

        result = build_not_due_interest_file(self._request(NOT_DUE_INTEREST))

        self.assertEqual(result.row_count, 1)
        self.assertEqual(result.skipped_count, 1)

    def test_build_not_due_interest_file(self) -> None:
        result = build_not_due_interest_file(self._request(NOT_DUE_INTEREST))

        self.assertEqual(result.row_count, 1)
        workbook = load_workbook(result.output_file, data_only=True)
        sheet = workbook["SaoKeTrichLai"]
        self.assertEqual(sheet["A3"].value, "GN_FUTURE")
        self.assertEqual(sheet["D3"].value, "20260731")
        workbook.close()

    def test_build_overdue_center_interest_file(self) -> None:
        result = build_overdue_center_interest_file(self._request(OVERDUE_CENTER_INTEREST))

        self.assertEqual(result.row_count, 1)
        self.assertEqual(result.summary["vba_procedure"], "ThuLaiQuaHanCenTer")
        workbook = load_workbook(result.output_file, data_only=True)
        technical = workbook["SaoKeTrichLai"]
        self.assertEqual(technical["A3"].value, "GN_PAST")
        self.assertEqual(technical["D3"].value, "20000103")
        ledger = workbook["BangKeTheoLo"]
        self.assertEqual(ledger["D4"].value, "GN_PAST")
        self.assertEqual(ledger["F4"].value, 1000)
        self.assertEqual(ledger["G4"].value, "02/01/2000")
        workbook.close()

    def test_overdue_center_weekend_due_date_can_include_weekend_interest(self) -> None:
        self._write_weekend_due_loan_file(self.loan_file)

        with_weekend = build_overdue_center_interest_file(
            self._request(OVERDUE_CENTER_INTEREST)
        )
        without_weekend = build_overdue_center_interest_file(
            AutoInterestCreateRequest(
                loan_file=self.loan_file,
                deposit_file=self.deposit_file,
                collection_date=date(2026, 7, 31),
                deposit_statement_type="msit",
                collection_mode=OVERDUE_CENTER_INTEREST,
                output_folder=self.root / "output",
                include_weekend_interest=False,
            )
        )

        with_workbook = load_workbook(with_weekend.output_file, data_only=True)
        without_workbook = load_workbook(without_weekend.output_file, data_only=True)
        self.assertEqual(with_workbook["SaoKeTrichLai"]["D3"].value, "20000110")
        self.assertEqual(with_workbook["BangKeTheoLo"]["F4"].value, 300)
        self.assertEqual(with_workbook["BangKeTheoLo"]["G4"].value, "09/01/2000")
        self.assertEqual(without_workbook["SaoKeTrichLai"]["D3"].value, "20000108")
        self.assertEqual(without_workbook["BangKeTheoLo"]["F4"].value, 100)
        self.assertEqual(without_workbook["BangKeTheoLo"]["G4"].value, "07/01/2000")
        with_workbook.close()
        without_workbook.close()

    def test_build_not_due_and_overdue_interest_file(self) -> None:
        result = build_not_due_and_overdue_interest_file(
            self._request(NOT_DUE_AND_OVERDUE_INTEREST)
        )

        self.assertEqual(result.row_count, 2)
        workbook = load_workbook(result.output_file, data_only=True)
        loans = [row[0] for row in workbook["SaoKeTrichLai"].iter_rows(min_row=3, max_row=4, values_only=True)]
        self.assertEqual(loans, ["GN_FUTURE", "GN_PAST"])
        workbook.close()

    def test_auto_interest_output_does_not_modify_input(self) -> None:
        loan_bytes = self.loan_file.read_bytes()
        deposit_bytes = self.deposit_file.read_bytes()

        create_auto_interest_file(self._request(COLLECT_ALL_INTEREST))

        self.assertEqual(self.loan_file.read_bytes(), loan_bytes)
        self.assertEqual(self.deposit_file.read_bytes(), deposit_bytes)

    def test_auto_interest_excludes_pledged_collateral_customers(self) -> None:
        self._write_msit_file(self.deposit_file, customer_b_balance=1000000)
        self._write_collateral_file(self.collateral_file)

        interest = create_auto_interest_file(
            AutoInterestCreateRequest(
                loan_file=self.loan_file,
                deposit_file=self.deposit_file,
                collection_date=date(2026, 7, 31),
                deposit_statement_type="msit",
                collection_mode=COLLECT_ALL_INTEREST,
                output_folder=self.root / "output",
                collateral_file=self.collateral_file,
            )
        )

        self.assertEqual(interest.row_count, 1)
        self.assertEqual(interest.skipped_count, 2)
        self.assertEqual(interest.summary["pledged_collateral_rows"], 2)
        workbook = load_workbook(interest.output_file, data_only=True)
        technical_loans = [
            row[0]
            for row in workbook["SaoKeTrichLai"].iter_rows(
                min_row=3,
                values_only=True,
            )
            if row[0]
        ]
        ledger_loans = [
            row[3]
            for row in workbook["BangKeTheoLo"].iter_rows(
                min_row=4,
                values_only=True,
            )
            if row[3]
        ]
        workbook.close()
        self.assertEqual(technical_loans, ["GN_SKIP"])
        self.assertEqual(ledger_loans, ["GN_SKIP"])

        report = create_auto_interest_report(
            AutoInterestReportRequest(
                source_file=interest.output_file,
                settings=AutoInterestSettings(
                    report_folder=self.root / "reports",
                    output_folder=self.root / "output",
                ),
                report_date=date(2026, 7, 31),
            )
        )
        report_workbook = load_workbook(report.output_file, data_only=True)
        report_sheet = report_workbook["BaoCaoThuBanTuDong"]
        self.assertEqual(report_sheet["A2"].value, "GN_SKIP")
        self.assertIsNone(report_sheet["A3"].value)
        report_workbook.close()

    def test_auto_interest_report_created(self) -> None:
        interest = create_auto_interest_file(self._request(COLLECT_ALL_INTEREST))
        report = create_auto_interest_report(
            AutoInterestReportRequest(
                source_file=interest.output_file,
                settings=AutoInterestSettings(
                    report_folder=self.root / "reports",
                    output_folder=self.root / "output",
                ),
                report_date=date(2026, 7, 31),
            )
        )

        self.assertTrue(report.output_file.exists())
        self.assertEqual(report.row_count, 2)
        workbook = load_workbook(report.output_file, data_only=True)
        sheet = workbook["BaoCaoThuBanTuDong"]
        self.assertEqual(sheet["A1"].value, "Disburse No(1)")
        self.assertEqual(sheet["B2"].value, "02")
        self.assertEqual(sheet["A2"].value, "GN_FUTURE")
        workbook.close()

    def test_auto_interest_report_summary_matches_input(self) -> None:
        interest = create_auto_interest_file(self._request(COLLECT_ALL_INTEREST))
        report = create_auto_interest_report(
            AutoInterestReportRequest(
                source_file=interest.output_file,
                settings=AutoInterestSettings(
                    report_folder=self.root / "reports",
                    output_folder=self.root / "output",
                ),
                report_date=date(2026, 7, 31),
            )
        )

        self.assertEqual(report.summary["total_interest"], interest.summary["total_interest"])
        self.assertEqual(report.summary["total_interest"], 300000)

    def test_auto_interest_report_collect_all_uses_filegoc_laidk(self) -> None:
        interest = create_auto_interest_file(self._request(COLLECT_ALL_INTEREST))
        source_workbook = load_workbook(interest.output_file)
        source_workbook["BangKeTheoLo"]["F4"].value = 999
        source_workbook.save(interest.output_file)
        source_workbook.close()

        report = create_auto_interest_report(
            AutoInterestReportRequest(
                source_file=interest.output_file,
                settings=AutoInterestSettings(
                    report_folder=self.root / "reports",
                    output_folder=self.root / "output",
                ),
                report_date=date(2026, 7, 31),
            )
        )

        workbook = load_workbook(report.output_file, data_only=True)
        sheet = workbook["BaoCaoThuBanTuDong"]
        self.assertEqual(sheet["A2"].value, "GN_FUTURE")
        self.assertEqual(sheet["C2"].value, 100000)
        workbook.close()

    def test_auto_interest_report_other_modes_use_bang_ke_lai_dk(self) -> None:
        interest = create_auto_interest_file(self._request(OVERDUE_CENTER_INTEREST))

        report = create_auto_interest_report(
            AutoInterestReportRequest(
                source_file=interest.output_file,
                settings=AutoInterestSettings(
                    report_folder=self.root / "reports",
                    output_folder=self.root / "output",
                ),
                report_date=date(2026, 7, 31),
            )
        )

        workbook = load_workbook(report.output_file, data_only=True)
        sheet = workbook["BaoCaoThuBanTuDong"]
        self.assertEqual(sheet["A2"].value, "GN_PAST")
        self.assertEqual(sheet["C2"].value, 1000)
        self.assertEqual(sheet["D2"].value, "02/01/2000")
        workbook.close()
        self.assertEqual(report.summary["total_interest"], 1000)

    def _request(self, mode: str) -> AutoInterestCreateRequest:
        return AutoInterestCreateRequest(
            loan_file=self.loan_file,
            deposit_file=self.deposit_file,
            collection_date=date(2026, 7, 31),
            deposit_statement_type="msit",
            collection_mode=mode,
            output_folder=self.root / "output",
        )

    def _write_loan_file(
        self,
        path: Path,
        headers: tuple[str, ...] | None = None,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        actual_headers = headers or (
            "chonIn",
            "STT",
            "MaKH",
            "TenKH",
            "NhomNo",
            "MaGN",
            "CCY",
            "DuNo",
            "LaiDK",
            "CBTD",
            "NgayLaiC",
            "addr",
            "nxtintrpmt",
            "LaiSuat",
            "NgayTinhLaiTu1",
            "NgayTinhLaiTu2",
        )
        sheet.append(actual_headers)
        if headers is None:
            sheet.append(
                (
                    "",
                    1,
                    "549100000001",
                    "Nguyen Van A",
                    "1",
                    "GN_FUTURE",
                    "VND",
                    10000000,
                    100000,
                    "CBTD",
                    date(2026, 7, 31),
                    "Dia chi A",
                    date(2099, 12, 31),
                    10,
                    date(2099, 1, 1),
                    date(2099, 1, 1),
                )
            )
            sheet.append(
                (
                    "",
                    2,
                    "549100000001",
                    "Nguyen Van A",
                    "1",
                    "GN_PAST",
                    "VND",
                    3650000,
                    200000,
                    "CBTD",
                    date(2000, 1, 3),
                    "Dia chi A",
                    date(2000, 1, 3),
                    10,
                    date(2000, 1, 1),
                    date(1999, 12, 31),
                )
            )
            sheet.append(
                (
                    "",
                    3,
                    "549100000002",
                    "Tran Van B",
                    "1",
                    "GN_SKIP",
                    "VND",
                    5000000,
                    300000,
                    "CBTD",
                    date(2026, 7, 31),
                    "Dia chi B",
                    date(2099, 12, 31),
                    10,
                    date(2099, 1, 1),
                    date(2099, 1, 1),
                )
            )
        workbook.save(path)

    def _write_weekend_due_loan_file(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            (
                "chonIn",
                "STT",
                "MaKH",
                "TenKH",
                "NhomNo",
                "MaGN",
                "CCY",
                "DuNo",
                "LaiDK",
                "CBTD",
                "NgayLaiC",
                "addr",
                "nxtintrpmt",
                "LaiSuat",
                "NgayTinhLaiTu1",
                "NgayTinhLaiTu2",
            )
        )
        sheet.append(
            (
                "",
                1,
                "549100000001",
                "Nguyen Van A",
                "1",
                "GN_WEEKEND",
                "VND",
                365000,
                999999,
                "CBTD",
                date(2000, 1, 8),
                "Dia chi A",
                date(2000, 1, 8),
                10,
                date(2000, 1, 6),
                date(2000, 1, 6),
            )
        )
        workbook.save(path)

    def _write_msit_file(
        self,
        path: Path,
        customer_a_balance: int = 1000000,
        customer_b_balance: int = 100,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            (
                "Customer_No",
                "Customer_Name",
                "DP_TypeName",
                "Ccy",
                "Curent_Balance",
                "Rate",
                "Account_Number",
                "Tel",
                "Maturity_Date",
            )
        )
        sheet.append(
            (
                "549100000001",
                "Nguyen Van A",
                "TGTT",
                "VND",
                customer_a_balance,
                0.1,
                "123456789",
                "",
                "",
            )
        )
        sheet.append(
            (
                "549100000001",
                "Nguyen Van A",
                "TKCKH",
                "VND",
                9000000,
                0.1,
                "TERM999",
                "",
                date(2099, 1, 1),
            )
        )
        sheet.append(
            (
                "549100000002",
                "Tran Van B",
                "TGTT",
                "VND",
                customer_b_balance,
                0.1,
                "222222222",
                "",
                "",
            )
        )
        workbook.save(path)

    def _write_collateral_file(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        headers = [""] * 29
        headers[0] = "clno"
        headers[1] = "clcustno"
        headers[2] = "clcustnm"
        headers[5] = "cltpcd"
        headers[6] = "cldtltpcd"
        headers[28] = "acctkey"
        sheet.append(headers)

        pledged = [""] * 29
        pledged[0] = "CL001"
        pledged[1] = "549100000001"
        pledged[2] = "Nguyen Van A"
        pledged[5] = "TSBD"
        pledged[6] = "CC"
        pledged[28] = "994003 - [Cầm Cố] Pledged assets"
        sheet.append(pledged)

        mortgage = [""] * 29
        mortgage[0] = "CL002"
        mortgage[1] = "549100000002"
        mortgage[2] = "Tran Van B"
        mortgage[5] = "TSBD"
        mortgage[6] = "TC"
        mortgage[28] = "994001 - [Thế chấp]Mortgage assets"
        sheet.append(mortgage)
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()
