from __future__ import annotations

import os
from pathlib import Path
from contextlib import closing
import sqlite3
from tempfile import TemporaryDirectory
import unittest
import zipfile
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from PySide6.QtWidgets import QApplication, QPushButton

from agribank_v3.features.settings.unit_directory import UNIT_SETTINGS_TITLE
from agribank_v3.features.settings.unit_directory.models import (
    AppUnitSettings,
    BranchDirectoryEntry,
    HEAD_OFFICE,
    OfficeDirectoryEntry,
    TRANSACTION_OFFICE,
)
from agribank_v3.features.settings.unit_directory.repository import (
    UnitDirectoryError,
    UnitDirectoryRepository,
    build_office_code,
    normalize_trctcd,
)
from agribank_v3.features.settings.unit_directory.service import (
    get_unit_directory_service,
    invalidate_unit_directory_cache,
)
from agribank_v3.features.settings.unit_directory.unit_settings_window import (
    export_unit_directory_excel,
)
from agribank_v3.settings import AddinMode, AppSettingsDatabase, BranchProfile
from agribank_v3.ui.settings import SettingsWidget


class SettingsDatabaseTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = TemporaryDirectory()
        self.database_path = (
            Path(self.temporary_directory.name) / "DuLieuV3.db"
        )
        self.database = AppSettingsDatabase(self.database_path)
        with closing(
            sqlite3.connect(self.database.quiz_database_path)
        ) as connection:
            connection.execute(
                """
                CREATE TABLE questions (
                    id INTEGER PRIMARY KEY,
                    question_text TEXT NOT NULL
                )
                """
            )
            connection.commit()

    def tearDown(self) -> None:
        invalidate_unit_directory_cache(self.database_path)
        self.temporary_directory.cleanup()

    def test_branch_profile_is_structured_versioned_and_durable(self) -> None:
        saved = self.database.save_branch_profile(
            BranchProfile(
                branch_code="  1234 ",
                branch_name="Agribank Chi nhánh Trung tâm",
                address="  01   Nguyễn Huệ  ",
                report_preparer="Nguyễn Văn An",
            )
        )
        updated = self.database.save_branch_profile(
            BranchProfile(
                branch_code="1234",
                branch_name="Agribank Chi nhánh Trung tâm",
                address="01 Nguyễn Huệ",
                report_preparer="Trần Thị Bình",
            )
        )

        loaded = self.database.load_branch_profile()
        self.assertEqual(saved.revision, 1)
        self.assertEqual(updated.revision, 2)
        self.assertEqual(loaded.address, "01 Nguyễn Huệ")
        self.assertEqual(loaded.report_preparer, "Trần Thị Bình")

        with closing(sqlite3.connect(self.database_path)) as connection:
            history_count = connection.execute(
                "SELECT COUNT(*) FROM branch_profile_history"
            ).fetchone()[0]
        self.assertEqual(history_count, 2)

    def test_backup_and_restore_include_branch_profile(self) -> None:
        self.database.save_branch_profile(
            BranchProfile(branch_code="1001", branch_name="Chi nhánh A")
        )
        backup = self.database.create_backup()
        self.database.save_branch_profile(
            BranchProfile(branch_code="2002", branch_name="Chi nhánh B")
        )

        safety_backup = self.database.restore_backup(backup)

        restored = self.database.load_branch_profile()
        self.assertEqual(restored.branch_code, "1001")
        self.assertEqual(restored.branch_name, "Chi nhánh A")
        self.assertTrue(safety_backup.is_file())
        self.assertEqual(self.database.status().integrity, "ok")

    def test_backup_bundle_includes_and_restores_both_databases(self) -> None:
        self.database.save_branch_profile(
            BranchProfile(branch_code="1001", branch_name="Chi nhánh A")
        )
        with closing(
            sqlite3.connect(self.database.quiz_database_path)
        ) as connection:
            connection.execute(
                """
                INSERT INTO questions(id, question_text)
                VALUES (1, 'Nội dung trước sao lưu')
                """
            )
            connection.commit()

        backup = self.database.create_backup()
        self.assertEqual(backup.suffix, ".zip")
        with zipfile.ZipFile(backup) as archive:
            self.assertEqual(
                set(archive.namelist()),
                {"manifest.json", "DuLieuV3.db", "quiz.db"},
            )

        self.database.save_branch_profile(
            BranchProfile(branch_code="2002", branch_name="Chi nhánh B")
        )
        with closing(
            sqlite3.connect(self.database.quiz_database_path)
        ) as connection:
            connection.execute(
                "UPDATE questions SET question_text = 'Nội dung đã thay đổi'"
            )
            connection.commit()

        self.database.restore_backup(backup)

        self.assertEqual(
            self.database.load_branch_profile().branch_code,
            "1001",
        )
        with closing(
            sqlite3.connect(self.database.quiz_database_path)
        ) as connection:
            restored_question = connection.execute(
                "SELECT question_text FROM questions WHERE id = 1"
            ).fetchone()[0]
        self.assertEqual(restored_question, "Nội dung trước sao lưu")

    def test_addin_mode_defaults_to_permanent_and_is_durable(self) -> None:
        self.assertEqual(
            self.database.load_addin_mode(),
            AddinMode.PERMANENT,
        )

        saved = self.database.save_addin_mode(AddinMode.SESSION)
        reopened = AppSettingsDatabase(self.database_path)

        self.assertEqual(saved, AddinMode.SESSION)
        self.assertEqual(reopened.load_addin_mode(), AddinMode.SESSION)

    def test_each_addin_enabled_state_is_durable_and_new_files_default_on(
        self,
    ) -> None:
        initial = self.database.load_addin_states(
            ["FunctionsA.xlam", "FunctionsB.xla"]
        )
        self.assertEqual(
            initial,
            {"FunctionsA.xlam": True, "FunctionsB.xla": True},
        )

        self.database.save_addin_enabled("FunctionsA.xlam", False)
        reopened = AppSettingsDatabase(self.database_path)

        self.assertEqual(
            reopened.load_addin_states(
                ["FunctionsA.xlam", "FunctionsB.xla", "NewFunctions.xlam"]
            ),
            {
                "FunctionsA.xlam": False,
                "FunctionsB.xla": True,
                "NewFunctions.xlam": True,
            },
        )

    def test_quick_access_items_are_durable_and_ignore_unknown_ids(self) -> None:
        default_ids = ("convert_case", "vlookup_extended")
        valid_ids = ("convert_case", "merge_same_structure", "vlookup_extended")

        self.assertEqual(
            self.database.load_quick_access_items(default_ids, valid_ids),
            default_ids,
        )

        saved = self.database.save_quick_access_items(
            ["merge_same_structure", "unknown", "convert_case", "convert_case"],
            valid_ids,
        )
        reopened = AppSettingsDatabase(self.database_path)

        self.assertEqual(saved, ("merge_same_structure", "convert_case"))
        self.assertEqual(
            reopened.load_quick_access_items(default_ids, valid_ids),
            ("merge_same_structure", "convert_case"),
        )


class UnitDirectoryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = TemporaryDirectory()
        self.database_path = Path(self.temporary_directory.name) / "DuLieuV3.db"
        self.database = AppSettingsDatabase(self.database_path)
        with closing(sqlite3.connect(self.database.quiz_database_path)) as connection:
            connection.execute(
                """
                CREATE TABLE questions (
                    id INTEGER PRIMARY KEY,
                    question_text TEXT NOT NULL
                )
                """
            )
            connection.commit()
        self.service = get_unit_directory_service(self.database_path)

    def tearDown(self) -> None:
        invalidate_unit_directory_cache(self.database_path)
        self.temporary_directory.cleanup()

    def _create_branch(self, code: str = "6501", name: str = "Chi nhánh Mới") -> BranchDirectoryEntry:
        return self.service.create_branch(
            BranchDirectoryEntry(
                branch_code=code,
                branch_name=name,
                short_name=f"CN {name.split()[-1]}",
                province_name="Lâm Đồng",
            )
        )

    def _create_office(
        self,
        branch_code: str = "6501",
        trctcd: str = "00",
        name: str = "Hội sở Chi nhánh Mới",
        office_type: str = HEAD_OFFICE,
    ) -> OfficeDirectoryEntry:
        if self.service.get_branch(branch_code) is None:
            self._create_branch(branch_code)
        return self.service.create_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code=branch_code,
                trctcd=trctcd,
                office_code="",
                office_name=name,
                short_name="Hội sở" if normalize_trctcd(trctcd) == "00" else f"PGD {normalize_trctcd(trctcd)}",
                office_type=office_type,
            )
        )

    def test_branch_directory_migration(self) -> None:
        with closing(sqlite3.connect(self.database_path)) as connection:
            columns = {row[1] for row in connection.execute("PRAGMA table_info(branch_directory)")}

        self.assertIn("branch_code", columns)
        self.assertIn("display_name", columns)
        self.assertIn("is_active", columns)

    def test_office_directory_migration(self) -> None:
        with closing(sqlite3.connect(self.database_path)) as connection:
            columns = {row[1] for row in connection.execute("PRAGMA table_info(office_directory)")}
            indexes = {row[1] for row in connection.execute("PRAGMA index_list(office_directory)")}

        self.assertIn("office_code", columns)
        self.assertIn("office_type", columns)
        self.assertTrue(any("office_directory" in index for index in indexes))

    def test_app_unit_settings_migration(self) -> None:
        settings = self.service.get_settings()

        self.assertEqual(settings.home_branch_code, "5491")
        self.assertEqual(settings.default_office_code, "5491-00")

    def test_unit_directory_seed_from_legacy_mapping(self) -> None:
        self.assertEqual(self.service.get_branch_display_name("5491"), "5491 - CN Lộc Phát")
        self.assertEqual(self.service.get_office_display_name("5405", "01"), "5405-01 - PGD Ka Đô")

    def test_unit_directory_migration_idempotent(self) -> None:
        with closing(sqlite3.connect(self.database_path)) as connection:
            before = connection.execute("SELECT COUNT(*) FROM branch_directory").fetchone()[0]
        UnitDirectoryRepository(self.database_path).initialize_schema()
        with closing(sqlite3.connect(self.database_path)) as connection:
            after = connection.execute("SELECT COUNT(*) FROM branch_directory").fetchone()[0]

        self.assertEqual(before, after)

    def test_legacy_seed_does_not_overwrite_changed_branch_name(self) -> None:
        self.service.save_branch(
            BranchDirectoryEntry(
                branch_code="5491",
                branch_name="Chi nhánh người dùng đổi",
                short_name="CN người dùng",
            )
        )

        UnitDirectoryRepository(self.database_path).initialize_schema()
        reopened = get_unit_directory_service(self.database_path)

        self.assertEqual(reopened.get_branch_display_name("5491"), "5491 - CN người dùng")

    def test_legacy_seed_does_not_overwrite_changed_office_name(self) -> None:
        self.service.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="5405",
                trctcd="01",
                office_code="5405-01",
                office_name="Phòng giao dịch người dùng đổi",
                short_name="PGD người dùng",
                office_type=TRANSACTION_OFFICE,
            )
        )

        UnitDirectoryRepository(self.database_path).initialize_schema()
        reopened = get_unit_directory_service(self.database_path)

        self.assertEqual(reopened.get_office_display_name("5405", "01"), "5405-01 - PGD người dùng")

    def test_legacy_seed_does_not_reset_home_branch(self) -> None:
        self._create_office("6501", "00")
        self._create_office("6501", "01", "Phòng giao dịch 01", TRANSACTION_OFFICE)
        self.service.save_settings(
            AppUnitSettings(
                home_branch_code="6501",
                default_office_code="6501-01",
                organization_name="Agribank Chi nhánh 6501",
            )
        )

        UnitDirectoryRepository(self.database_path).initialize_schema()
        reopened = get_unit_directory_service(self.database_path)
        settings = reopened.get_settings()

        self.assertEqual(settings.home_branch_code, "6501")
        self.assertEqual(settings.default_office_code, "6501-01")
        self.assertEqual(settings.organization_name, "Agribank Chi nhánh 6501")

    def test_runtime_source_has_no_legacy_branch_mapping(self) -> None:
        source_root = Path(__file__).resolve().parents[1] / "src" / "agribank_v3"
        offenders: list[str] = []
        for path in source_root.rglob("*.py"):
            if path.name == "legacy_seed.py":
                continue
            text = path.read_text(encoding="utf-8")
            if any(token in text for token in ("BRANCH_NAMES", "CUSTOMER_BRANCH_NAMES", "branch_name_map")):
                offenders.append(str(path.relative_to(source_root)))

        self.assertEqual(offenders, [])

    def test_runtime_source_has_no_legacy_pgd_mapping(self) -> None:
        source_root = Path(__file__).resolve().parents[1] / "src" / "agribank_v3"
        offenders: list[str] = []
        for path in source_root.rglob("*.py"):
            if path.name == "legacy_seed.py":
                continue
            text = path.read_text(encoding="utf-8")
            if any(token in text for token in ("PGD_NAMES", "office_name_map")):
                offenders.append(str(path.relative_to(source_root)))

        self.assertEqual(offenders, [])

    def test_update_does_not_overwrite_user_unit_settings(self) -> None:
        self._create_office("6501", "00")
        self.service.save_settings(
            AppUnitSettings(
                home_branch_code="6501",
                default_office_code="6501-00",
                organization_name="Agribank Chi nhánh Mới",
            )
        )

        AppSettingsDatabase(self.database_path)
        reopened = get_unit_directory_service(self.database_path)

        self.assertEqual(reopened.get_settings().home_branch_code, "6501")
        self.assertEqual(reopened.get_settings().organization_name, "Agribank Chi nhánh Mới")

    def test_create_branch(self) -> None:
        branch = self._create_branch("6501", "Chi nhánh Mới")

        self.assertEqual(branch.branch_code, "6501")
        self.assertEqual(self.service.get_branch("6501").branch_name, "Chi nhánh Mới")

    def test_update_branch_name(self) -> None:
        self._create_branch("6501", "Chi nhánh Cũ")
        self.service.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Đổi tên", short_name="CN Đổi tên")
        )

        self.assertEqual(self.service.get_branch_display_name("6501"), "6501 - CN Đổi tên")

    def test_branch_code_is_text(self) -> None:
        self._create_branch("0123", "Chi nhánh Có số 0")
        with closing(sqlite3.connect(self.database_path)) as connection:
            stored_type = connection.execute(
                "SELECT typeof(branch_code) FROM branch_directory WHERE branch_code = '0123'"
            ).fetchone()[0]

        self.assertEqual(stored_type, "text")
        self.assertTrue(self.service.get_branch_display_name("0123").startswith("0123 - "))

    def test_duplicate_branch_code_rejected(self) -> None:
        self._create_branch("6501")

        with self.assertRaises(UnitDirectoryError):
            self._create_branch("6501", "Chi nhánh Trùng")

    def test_inactive_branch_still_resolves_historical_name(self) -> None:
        self._create_branch("6501", "Chi nhánh Lịch sử")
        self.service.set_branch_active("6501", False)

        self.assertEqual(self.service.get_branch_display_name("6501"), "6501 - CN sử")

    def test_branch_display_name(self) -> None:
        self.service.create_branch(
            BranchDirectoryEntry(
                branch_code="6501",
                branch_name="Chi nhánh Mới",
                short_name="CN Mới",
                display_name="6501 - Chi nhánh triển khai",
            )
        )

        self.assertEqual(self.service.get_branch_display_name("6501"), "6501 - Chi nhánh triển khai")

    def test_unknown_branch_fallback(self) -> None:
        self.assertEqual(self.service.get_branch_display_name("6501"), "6501 - CN chưa khai báo")

    def test_create_head_office(self) -> None:
        office = self._create_office("6501", "00")

        self.assertEqual(office.office_type, HEAD_OFFICE)
        self.assertEqual(office.office_code, "6501-00")

    def test_create_transaction_office(self) -> None:
        office = self._create_office("6501", "01", "Phòng giao dịch Mới", TRANSACTION_OFFICE)

        self.assertEqual(office.office_type, TRANSACTION_OFFICE)
        self.assertEqual(office.office_code, "6501-01")

    def test_trctcd_preserves_leading_zero(self) -> None:
        office = self._create_office("6501", "1", "Phòng giao dịch 01", TRANSACTION_OFFICE)

        self.assertEqual(office.trctcd, "01")

    def test_trctcd_zero_resolves_head_office(self) -> None:
        office = self._create_office("6501", "0")

        self.assertEqual(office.trctcd, "00")
        self.assertEqual(office.office_type, HEAD_OFFICE)

    def test_unique_branch_trctcd(self) -> None:
        self._create_office("6501", "01", "Phòng giao dịch 01", TRANSACTION_OFFICE)

        with self.assertRaises(UnitDirectoryError):
            self._create_office("6501", "1", "Phòng giao dịch trùng", TRANSACTION_OFFICE)

    def test_office_code_generation(self) -> None:
        self.assertEqual(build_office_code("6501", "1"), "6501-01")

    def test_unknown_office_fallback(self) -> None:
        self.assertEqual(self.service.get_office_display_name("6501", "03"), "6501-03 - PGD 03")

    def test_inactive_office_historical_lookup(self) -> None:
        self._create_office("6501", "01", "Phòng giao dịch Cũ", TRANSACTION_OFFICE)
        self.service.set_office_active("6501-01", False)

        self.assertEqual(self.service.get_office_display_name("6501", "01"), "6501-01 - PGD 01")

    def test_inactive_office_resolves_historical_name(self) -> None:
        self.service.save_branch(BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh lịch sử"))
        self.service.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="6501",
                trctcd="01",
                office_code="6501-01",
                office_name="Phòng giao dịch lịch sử",
                short_name="PGD lịch sử",
                office_type=TRANSACTION_OFFICE,
            )
        )
        self.service.set_office_active("6501-01", False)

        self.assertEqual(self.service.get_office_display_name("6501", "01"), "6501-01 - PGD lịch sử")

    def test_inactive_office_excluded_from_active_combo(self) -> None:
        self._create_office("6501", "01", "Phòng giao dịch 01", TRANSACTION_OFFICE)
        self.service.set_office_active("6501-01", False)

        active_codes = {office.office_code for office in self.service.get_active_offices("6501")}

        self.assertNotIn("6501-01", active_codes)

    def test_save_home_branch(self) -> None:
        self._create_office("6501", "00")
        saved = self.service.save_settings(AppUnitSettings(home_branch_code="6501"))

        self.assertEqual(saved.home_branch_code, "6501")

    def test_save_default_office(self) -> None:
        self._create_office("6501", "00")
        saved = self.service.save_settings(
            AppUnitSettings(home_branch_code="6501", default_office_code="6501-00")
        )

        self.assertEqual(saved.default_office_code, "6501-00")

    def test_default_office_belongs_to_home_branch(self) -> None:
        self._create_office("6501", "00")

        with self.assertRaises(UnitDirectoryError):
            self.service.save_settings(
                AppUnitSettings(home_branch_code="5491", default_office_code="6501-00")
            )

    def test_home_branch_persists_after_restart(self) -> None:
        self._create_office("6501", "00")
        self.service.save_settings(AppUnitSettings(home_branch_code="6501", default_office_code="6501-00"))
        invalidate_unit_directory_cache(self.database_path)
        reopened = get_unit_directory_service(self.database_path)

        self.assertEqual(reopened.get_settings().home_branch_code, "6501")

    def test_new_non_legacy_branch_after_restart(self) -> None:
        self._create_office("6501", "00")
        self.service.save_settings(AppUnitSettings(home_branch_code="6501", default_office_code="6501-00"))

        invalidate_unit_directory_cache(self.database_path)
        reopened = get_unit_directory_service(self.database_path)

        self.assertEqual(reopened.get_settings().home_branch_code, "6501")
        self.assertEqual(reopened.get_branch_display_name("6501"), "6501 - CN Mới")

    def test_new_non_legacy_offices_after_restart(self) -> None:
        self._create_office("6501", "00")
        self._create_office("6501", "01", "Phòng giao dịch 01", TRANSACTION_OFFICE)
        self._create_office("6501", "02", "Phòng giao dịch 02", TRANSACTION_OFFICE)
        self.service.save_settings(AppUnitSettings(home_branch_code="6501", default_office_code="6501-01"))

        invalidate_unit_directory_cache(self.database_path)
        reopened = get_unit_directory_service(self.database_path)

        self.assertEqual(reopened.get_settings().default_office_code, "6501-01")
        self.assertEqual(
            {office.office_code for office in reopened.get_active_offices("6501")},
            {"6501-00", "6501-01", "6501-02"},
        )

    def test_home_branch_not_hardcoded_at_runtime(self) -> None:
        self._create_office("6501", "00")
        self.service.save_settings(AppUnitSettings(home_branch_code="6501", default_office_code="6501-00"))

        self.assertEqual(self.service.get_home_branch().branch_code, "6501")

    def test_unit_directory_cache(self) -> None:
        self.assertEqual(self.service.get_branch_display_name("6501"), "6501 - CN chưa khai báo")
        UnitDirectoryRepository(self.database_path).save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Ngoài cache", short_name="CN Ngoài")
        )

        self.assertEqual(self.service.get_branch_display_name("6501"), "6501 - CN chưa khai báo")

    def test_unit_directory_cache_invalidated_after_update(self) -> None:
        self.service.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Ngoài cache", short_name="CN Ngoài")
        )

        self.assertEqual(self.service.get_branch_display_name("6501"), "6501 - CN Ngoài")

    def test_open_windows_refresh_after_directory_change(self) -> None:
        calls: list[str] = []
        def callback() -> None:
            calls.append("refresh")

        self.service.add_listener(callback)
        self.addCleanup(self.service.remove_listener, callback)

        self.service.save_branch(BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Listener"))

        self.assertEqual(calls, ["refresh"])

    def test_restore_database_invalidates_unit_cache(self) -> None:
        self.service.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Gốc", short_name="CN Gốc")
        )
        backup = self.database.create_backup()
        self.service.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Tạm", short_name="CN Tạm")
        )

        self.database.restore_backup(backup)

        self.assertEqual(self.service.get_branch_display_name("6501"), "6501 - CN Gốc")

    def test_restore_invalidates_unit_directory_cache(self) -> None:
        self.test_restore_database_invalidates_unit_cache()

    def test_unknown_branch_placeholder_editable(self) -> None:
        self.service.ensure_known_unit("6502")

        self.service.save_branch(
            BranchDirectoryEntry(branch_code="6502", branch_name="Chi nhánh đã khai báo", short_name="CN đã khai báo")
        )

        self.assertEqual(self.service.get_branch_display_name("6502"), "6502 - CN đã khai báo")

    def test_unknown_office_placeholder_editable(self) -> None:
        self.service.ensure_known_unit("6502", "03")

        self.service.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="6502",
                trctcd="03",
                office_code="6502-03",
                office_name="Phòng giao dịch 03 đã khai báo",
                short_name="PGD 03 khai báo",
                office_type=TRANSACTION_OFFICE,
            )
        )

        self.assertEqual(self.service.get_office_display_name("6502", "03"), "6502-03 - PGD 03 khai báo")

    def test_export_uses_current_branch_name(self) -> None:
        self._create_branch("6501", "Chi nhánh Xuất")
        self.service.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Xuất đổi", short_name="CN Xuất đổi")
        )
        output = export_unit_directory_excel(self.service, Path(self.temporary_directory.name) / "units.xlsx")
        workbook = load_workbook(output)

        self.assertIn(
            "6501 - CN Xuất đổi",
            [workbook["ChiNhanh"].cell(row, 4).value for row in range(2, workbook["ChiNhanh"].max_row + 1)],
        )

    def test_update_manager_preserves_unit_directory(self) -> None:
        self._create_office("6501", "00")
        self.service.save_settings(
            AppUnitSettings(
                home_branch_code="6501",
                default_office_code="6501-00",
                organization_name="Đơn vị triển khai",
            )
        )

        AppSettingsDatabase(self.database_path)

        self.assertEqual(self.service.get_settings().organization_name, "Đơn vị triển khai")

    def test_settings_page_exposes_unit_directory_tab(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        with patch("agribank_v3.ui.settings.AppSettingsDatabase", lambda: AppSettingsDatabase(self.database_path)):
            widget = SettingsWidget()
        self.addCleanup(widget.deleteLater)

        labels = [widget.tabs.tabText(index) for index in range(widget.tabs.count())]
        self.assertIn(UNIT_SETTINGS_TITLE, labels)
        widget.show_tab_for_feature(UNIT_SETTINGS_TITLE)
        self.assertEqual(widget.tabs.tabText(widget.tabs.currentIndex()), UNIT_SETTINGS_TITLE)

        emitted: list[bool] = []
        widget.unit_settings_requested.connect(lambda: emitted.append(True))
        buttons = [button for button in widget.findChildren(QPushButton) if button.text() == "Mở danh mục chi nhánh/PGD"]
        self.assertEqual(len(buttons), 1)
        buttons[0].click()

        self.assertEqual(emitted, [True])


if __name__ == "__main__":
    unittest.main()
