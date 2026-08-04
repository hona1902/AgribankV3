from __future__ import annotations

import os
from pathlib import Path
from contextlib import closing
import tempfile
import unittest

from openpyxl import load_workbook
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QApplication, QSizePolicy, QWidget

from agribank_v3.features.credit.summary.customer.officer_center_controller import open_officer_center_window
from agribank_v3.features.credit.summary.customer.officer_center_export import export_officer_center_workbook
from agribank_v3.features.credit.summary.customer.officer_center_repository import (
    OFFICER_MODE_EFFECTIVE,
    OFFICER_MODE_IMPORTED,
    OfficerCenterFilters,
    OfficerCenterRepository,
)
from agribank_v3.features.credit.summary.customer.officer_center_window import OfficerDashboardTab, OfficerCenterWindow
from agribank_v3.features.credit.summary.customer.officer_detail_registry import open_shared_officer_detail
from agribank_v3.features.credit.summary.customer.officer_detail_window import OfficerDetailWindow
from agribank_v3.features.credit.summary.customer.repository import CustomerRepository
from agribank_v3.features.credit.summary.customer.table_models import CustomerTableModel
from agribank_v3.features.credit.summary.models import PageResult, now_text
from agribank_v3.features.credit.summary.repository import SummaryRepository
from agribank_v3.features.credit.summary.services import import_nim_dn
from agribank_v3.ui.components.kpi import KpiMetric, kpi_display_values


FTPLN_HEADER = "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP,FTPCD,CUSTSEQ,CUSTNM"


os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")


class OfficerCenterTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.main_database_path = self.root / "DuLieuV3.db"
        self.customer_repository = CustomerRepository(self.main_database_path)
        self.summary_repository = SummaryRepository(self.main_database_path)
        self.repository = OfficerCenterRepository(self.customer_repository)
        self._seed_data()

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_officer_import_mode_uses_customer_officer_period(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        result = self.repository.officer_list(filters, page=1, page_size=10)
        by_code = {row["officer_code"]: row for row in result.rows}

        self.assertEqual(float(by_code["CB01"]["total_balance"]), 600)
        self.assertEqual(float(by_code["CB02"]["total_balance"]), 400)

    def test_import_mode_split_balance_between_officers(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        kpis = self.repository.kpis(filters)

        self.assertEqual(float(kpis["total_balance"]), 1000)
        self.assertEqual(int(kpis["officer_customer_occurrence_count"]), 2)

    def test_officer_effective_mode_uses_customer_period_summary(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_EFFECTIVE)
        result = self.repository.officer_list(filters, page=1, page_size=10)
        by_code = {row["officer_code"]: row for row in result.rows}

        self.assertNotIn("CB01", by_code)
        self.assertEqual(float(by_code["CB03"]["total_balance"]), 1000)

    def test_effective_mode_assigns_full_customer_to_effective_officer(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_EFFECTIVE)
        kpis = self.repository.kpis(filters)

        self.assertEqual(float(kpis["total_balance"]), 1000)
        self.assertEqual(int(kpis["officer_customer_occurrence_count"]), 1)

    def test_override_does_not_change_import_allocation(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        rows = {row["officer_code"]: row for row in self.repository.officer_list(filters, page=1, page_size=10).rows}

        self.assertEqual(float(rows["CB01"]["total_balance"]), 600)
        self.assertEqual(float(rows["CB02"]["total_balance"]), 400)
        self.assertNotIn("CB03", rows)

    def test_officer_modes_not_mixed(self) -> None:
        imported = self.repository.kpis(OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED))
        effective = self.repository.kpis(OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_EFFECTIVE))

        self.assertEqual(int(imported["active_officer_count"]), 2)
        self.assertEqual(int(effective["active_officer_count"]), 1)

    def test_customer_unique_count_vs_officer_customer_occurrences(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        kpis = self.repository.kpis(filters)

        self.assertEqual(int(kpis["officer_customer_occurrence_count"]), 2)
        self.assertEqual(int(kpis["unique_customer_count"]), 1)

    def test_officer_code_is_primary_identity(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-04", mode=OFFICER_MODE_IMPORTED)
        result = self.repository.officer_list(filters, page=1, page_size=10)
        codes = sorted(row["officer_code"] for row in result.rows)

        self.assertEqual(codes, ["CB01", "CB04"])

    def test_same_name_different_codes_not_merged(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-04", mode=OFFICER_MODE_IMPORTED)
        result = self.repository.officer_list(filters, page=1, page_size=10)

        self.assertEqual(result.total_rows, 2)

    def test_missing_code_has_unresolved_identity(self) -> None:
        self._insert_summary("2026-05", "CUST3", "003", "Khach 3", "Nguyen Van A", "", 500)
        self._insert_officer("2026-05", "CUST3", "", "Nguyen Van A", 500)
        filters = OfficerCenterFilters(report_period="2026-05", mode=OFFICER_MODE_IMPORTED)
        row = self.repository.officer_list(filters, page=1, page_size=10).rows[0]

        self.assertTrue(str(row["officer_key"]).startswith("NAME:"))
        self.assertEqual(row["officer_status"], "Chưa có mã cán bộ")

    def test_officer_weighted_nim_after(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        row = next(item for item in self.repository.officer_list(filters, page=1, page_size=10).rows if item["officer_code"] == "CB01")

        self.assertAlmostEqual(float(row["nim_after"]), 2.0)

    def test_officer_attention_and_bad_debt_ratio(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        row = next(item for item in self.repository.officer_debt_quality(filters, page=1, page_size=10).rows if item["officer_code"] == "CB02")

        self.assertEqual(float(row["attention_balance"]), 400)
        self.assertAlmostEqual(float(row["attention_ratio"]), 100.0)

    def test_compare_more_than_eight_officers_no_error(self) -> None:
        for index in range(10):
            code = f"CX{index:02d}"
            customer = f"CX-CUST-{index:02d}"
            self._insert_summary("2026-06", customer, f"{index:03d}", f"Khach {index}", code, code, 100 + index)
            self._insert_officer("2026-06", customer, code, code, 100 + index)
        selected = tuple(f"CX{index:02d}" for index in range(10))
        filters = OfficerCenterFilters(report_period="2026-06", mode=OFFICER_MODE_IMPORTED, selected_officers=selected)
        result = self.repository.compare_officers(filters, page=1, page_size=20)

        self.assertEqual(result.total_rows, 10)

    def test_compare_more_than_eight_officers_table(self) -> None:
        self._seed_many_officers("2026-10", 20)
        selected = tuple(f"CM{index:02d}" for index in range(20))
        filters = OfficerCenterFilters(report_period="2026-10", mode=OFFICER_MODE_IMPORTED, selected_officers=selected)
        first_page = self.repository.compare_officers(filters, page=1, page_size=8)
        third_page = self.repository.compare_officers(filters, page=3, page_size=8)

        self.assertEqual(first_page.total_rows, 20)
        self.assertEqual(len(first_page.rows), 8)
        self.assertEqual(len(third_page.rows), 4)

    def test_compare_more_than_eight_officers_export(self) -> None:
        self._seed_many_officers("2026-11", 20)
        output = self.root / "compare-20.xlsx"
        selected = tuple(f"CM{index:02d}" for index in range(20))
        export_officer_center_workbook(
            self.repository,
            OfficerCenterFilters(report_period="2026-11", mode=OFFICER_MODE_IMPORTED, selected_officers=selected),
            output,
        )
        workbook = load_workbook(output, data_only=True)
        try:
            sheet = workbook["SoSanhCBTD"]
            exported_codes = {sheet.cell(row, 2).value for row in range(2, sheet.max_row + 1)}
            self.assertEqual(exported_codes, set(selected))
            self.assertEqual(workbook["ThongTin"]["B14"].value, 20)
        finally:
            workbook.close()

    def test_chart_visible_series_limit_is_disclosed(self) -> None:
        app = QApplication.instance() or QApplication([])
        parent = _Host(self.main_database_path)
        window = open_officer_center_window(parent, self.main_database_path)
        try:
            window.compare_tab._apply_result(PageResult(rows=[{} for _ in range(8)], total_rows=20, page=1, page_size=8))
            self.assertEqual(window.compare_tab.series_status_label.text(), "Biểu đồ đang hiển thị 20/20 CBTD; bảng giữ đầy đủ dữ liệu.")
        finally:
            window.close()

    def test_officer_movement_transfer_not_counted_as_new_customer(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-04", compare_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        rows = {row["officer_code"]: row for row in self.repository.officer_movement(filters, page=1, page_size=20).rows}

        self.assertEqual(int(rows["CB04"]["transfer_in_customer_count"]), 1)
        self.assertEqual(int(rows["CB04"]["new_system_customer_count"]), 0)

    def test_customer_transfer_from_officer(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-04", compare_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        rows = {row["officer_code"]: row for row in self.repository.officer_movement(filters, page=1, page_size=20).rows}

        self.assertEqual(int(rows["CB01"]["transfer_out_customer_count"]), 1)

    def test_paid_off_not_transfer(self) -> None:
        self._insert_summary("2026-07", "PAID", "777", "Paid Customer", "Old Officer", "OLD", 700)
        self._insert_officer("2026-07", "PAID", "OLD", "Old Officer", 700)
        self._insert_summary("2026-08", "LIVE", "778", "Live Customer", "Live Officer", "LIVE", 800)
        self._insert_officer("2026-08", "LIVE", "LIVE", "Live Officer", 800)

        filters = OfficerCenterFilters(report_period="2026-08", compare_period="2026-07", mode=OFFICER_MODE_IMPORTED)
        rows = {row["officer_code"]: row for row in self.repository.officer_movement(filters, page=1, page_size=20).rows}

        self.assertEqual(int(rows["OLD"]["paid_off_customer_count"]), 1)
        self.assertEqual(int(rows["OLD"]["transfer_out_customer_count"]), 0)

    def test_branch_weighted_average_rate(self) -> None:
        self._seed_weighted_period()
        filters = OfficerCenterFilters(report_period="2026-09", mode=OFFICER_MODE_IMPORTED)
        kpis = self.repository.kpis(filters)

        self.assertAlmostEqual(float(kpis["average_rate"]), 1.9)

    def test_branch_weighted_nim_before(self) -> None:
        self._seed_weighted_period()
        filters = OfficerCenterFilters(report_period="2026-09", mode=OFFICER_MODE_IMPORTED)
        kpis = self.repository.kpis(filters)

        self.assertAlmostEqual(float(kpis["nim_before"]), 1.4)

    def test_branch_weighted_nim_after(self) -> None:
        self._seed_weighted_period()
        filters = OfficerCenterFilters(report_period="2026-09", mode=OFFICER_MODE_IMPORTED)
        kpis = self.repository.kpis(filters)

        self.assertAlmostEqual(float(kpis["nim_after"]), 1.3)

    def test_branch_weighted_bad_debt_ratio(self) -> None:
        self._seed_weighted_period()
        filters = OfficerCenterFilters(report_period="2026-09", mode=OFFICER_MODE_IMPORTED)
        kpis = self.repository.kpis(filters)

        self.assertAlmostEqual(float(kpis["bad_debt_ratio"]), 90.0)

    def test_officer_customer_total_and_allocated_balance(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED, selected_officers=("CB01",))
        result = self.repository.officer_customers(filters, page=1, page_size=10)
        row = result.rows[0]

        self.assertEqual(float(row["total_customer_balance"]), 1000)
        self.assertEqual(float(row["officer_balance"]), 600)

    def test_officer_customer_allocation_ratio(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED, selected_officers=("CB02",))
        row = self.repository.officer_customers(filters, page=1, page_size=10).rows[0]

        self.assertAlmostEqual(float(row["officer_share"]), 40.0)

    def test_debt_groups_sum_to_officer_total(self) -> None:
        filters = OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED)
        for row in self.repository.officer_debt_quality(filters, page=1, page_size=10).rows:
            group_total = sum(float(row.get(f"debt_group_{suffix}_balance") or 0) for suffix in ("1", "2", "3", "4", "5", "unknown"))
            self.assertAlmostEqual(group_total, float(row["total_balance"]))

    def test_officer_detail_has_required_tabs(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = OfficerDetailWindow(
            self.main_database_path,
            officer_code="CB01",
            officer_name="Nguyen Van A",
            filters=OfficerCenterFilters(report_period="2026-03", mode=OFFICER_MODE_IMPORTED),
        )
        try:
            labels = [window.tabs.tabText(index) for index in range(window.tabs.count())]
            self.assertEqual(labels, list(OfficerDetailWindow.REQUIRED_TAB_LABELS))
            self.assertGreaterEqual(window.overview_model.rowCount(), 1)
            self.assertGreaterEqual(window.customer_model.rowCount(), 1)
        finally:
            window.close()

    def test_same_officer_detail_single_instance(self) -> None:
        app = QApplication.instance() or QApplication([])
        parent = QWidget()
        row = {"officer_code": "CB01", "officer_name": "Nguyen Van A", "branch_code": "5491", "transaction_office": "00"}
        first = open_shared_officer_detail(parent, self.main_database_path, row, filters=OfficerCenterFilters(report_period="2026-03"))
        second = open_shared_officer_detail(parent, self.main_database_path, row, filters=OfficerCenterFilters(report_period="2026-03"))
        try:
            self.assertIs(first, second)
        finally:
            if first is not None:
                first.close()
            parent.close()

    def test_different_officers_can_open_separately(self) -> None:
        app = QApplication.instance() or QApplication([])
        parent = QWidget()
        first = open_shared_officer_detail(parent, self.main_database_path, {"officer_code": "CB01", "officer_name": "Nguyen Van A"})
        second = open_shared_officer_detail(parent, self.main_database_path, {"officer_code": "CB02", "officer_name": "Tran Van B"})
        try:
            self.assertIsNot(first, second)
        finally:
            if first is not None:
                first.close()
            if second is not None:
                second.close()
            parent.close()

    def test_credit_menu_opens_officer_management(self) -> None:
        app = QApplication.instance() or QApplication([])
        parent = _Host(self.main_database_path)
        window = open_officer_center_window(parent, self.main_database_path)

        self.assertEqual(window.windowTitle(), "Quản lý cán bộ tín dụng - AgribankV3")
        self.assertFalse(window.isModal())
        self.assertIs(open_officer_center_window(parent, self.main_database_path), window)
        window.close()

    def test_export_all_officer_sheets(self) -> None:
        output = self.root / "officers.xlsx"
        export_officer_center_workbook(
            self.repository,
            OfficerCenterFilters(report_period="2026-03", period_from="2026-03", period_to="2026-04"),
            output,
        )
        workbook = load_workbook(output)
        try:
            self.assertIn("TongQuanCBTD", workbook.sheetnames)
            self.assertIn("DanhSachCBTD", workbook.sheetnames)
            self.assertIn("BienDongCBTD", workbook.sheetnames)
            self.assertIn("SoSanhCBTD", workbook.sheetnames)
            self.assertIn("ChatLuongDuNo", workbook.sheetnames)
            self.assertIn("KhachHangTheoCBTD", workbook.sheetnames)
            self.assertIn("DanhMucCBTD", workbook.sheetnames)
            self.assertIn("ThongTin", workbook.sheetnames)
            self.assertEqual(workbook["DanhSachCBTD"]["B2"].data_type, "s")
            self.assertIsInstance(workbook["DanhSachCBTD"]["H2"].value, (int, float))
            self.assertEqual(workbook["ThongTin"]["B3"].value, "Theo phân bổ dữ liệu import")
        finally:
            workbook.close()

    def test_officer_period_term_balances_written_on_import(self) -> None:
        self._import_officer_term_fixture("2026-12")

        rows = self._officer_period_rows("2026-12")

        self.assertEqual(sum(float(row["balance_managed"]) for row in rows), 10_000_000_000)
        self.assertEqual(sum(float(row["short_term_balance"]) for row in rows), 7_000_000_000)
        self.assertEqual(sum(float(row["medium_long_term_balance"]) for row in rows), 2_000_000_000)
        self.assertEqual(sum(float(row["other_balance"]) for row in rows), 1_000_000_000)

    def test_officer_short_term_balance_mapping(self) -> None:
        self._import_officer_term_fixture("2026-12")

        rows = {row["officer_code"]: row for row in self.repository.officer_list(OfficerCenterFilters(report_period="2026-12"), page=1, page_size=10).rows}

        self.assertEqual(float(rows["A"]["short_term_balance"]), 6_000_000_000)
        self.assertEqual(float(rows["B"]["short_term_balance"]), 1_000_000_000)

    def test_officer_medium_long_term_balance_mapping(self) -> None:
        self._import_officer_term_fixture("2026-12")

        row = next(item for item in self.repository.officer_list(OfficerCenterFilters(report_period="2026-12"), page=1, page_size=10).rows if item["officer_code"] == "A")

        self.assertEqual(float(row["medium_long_term_balance"]), 2_000_000_000)
        self.assertAlmostEqual(float(row["medium_long_ratio"]), 25.0)

    def test_officer_other_term_balance_mapping(self) -> None:
        self._import_officer_term_fixture("2026-12")

        row = next(item for item in self.repository.officer_list(OfficerCenterFilters(report_period="2026-12"), page=1, page_size=10).rows if item["officer_code"] == "B")

        self.assertEqual(float(row["other_balance"]), 1_000_000_000)
        self.assertAlmostEqual(float(row["medium_long_ratio"]), 0.0)

    def test_officer_term_balances_sum_to_total(self) -> None:
        self._import_officer_term_fixture("2026-12")

        rows = self.repository.officer_list(OfficerCenterFilters(report_period="2026-12"), page=1, page_size=10).rows

        for row in rows:
            term_total = sum(float(row.get(field) or 0) for field in ("short_term_balance", "medium_long_term_balance", "other_balance"))
            self.assertAlmostEqual(term_total, float(row["total_balance"]))
            self.assertTrue(row["term_structure_available"])

    def test_old_officer_period_without_term_data_not_displayed_as_zero(self) -> None:
        self._insert_summary("2027-01", "OLDTERM", "901", "Old Term", "Officer Old", "OLD", 1000)
        self._insert_officer("2027-01", "OLDTERM", "OLD", "Officer Old", 1000, short_term_balance=0, medium_long_term_balance=0, other_balance=0)
        row = self.repository.officer_list(OfficerCenterFilters(report_period="2027-01"), page=1, page_size=10).rows[0]
        model = CustomerTableModel((("short_term_balance", "Dư nợ ngắn hạn", "term_money_or_dash"), ("medium_long_ratio", "Tỷ lệ", "term_percent_or_dash")))
        model.set_rows([row])

        self.assertFalse(row["term_structure_available"])
        self.assertIsNone(row["short_term_balance"])
        self.assertEqual(model.data(model.index(0, 0), Qt.ItemDataRole.DisplayRole), "—")
        self.assertEqual(model.data(model.index(0, 1), Qt.ItemDataRole.DisplayRole), "—")
        self.assertIn("chưa có dữ liệu phân bổ kỳ hạn", model.data(model.index(0, 0), Qt.ItemDataRole.ToolTipRole))

    def test_compare_officer_term_columns(self) -> None:
        self._import_officer_term_fixture("2026-12")

        rows = {row["officer_code"]: row for row in self.repository.compare_officers(OfficerCenterFilters(report_period="2026-12"), page=1, page_size=10).rows}

        self.assertEqual(float(rows["A"]["total_balance"]), 8_000_000_000)
        self.assertEqual(float(rows["A"]["short_term_balance"]), 6_000_000_000)
        self.assertEqual(float(rows["A"]["medium_long_term_balance"]), 2_000_000_000)
        self.assertEqual(float(rows["A"]["other_balance"]), 0)
        self.assertAlmostEqual(float(rows["A"]["medium_long_ratio"]), 25.0)

    def test_debt_quality_officer_term_columns(self) -> None:
        self._import_officer_term_fixture("2026-12")

        rows = {row["officer_code"]: row for row in self.repository.officer_debt_quality(OfficerCenterFilters(report_period="2026-12"), page=1, page_size=10).rows}

        self.assertEqual(float(rows["B"]["total_balance"]), 2_000_000_000)
        self.assertEqual(float(rows["B"]["short_term_balance"]), 1_000_000_000)
        self.assertEqual(float(rows["B"]["medium_long_term_balance"]), 0)
        self.assertEqual(float(rows["B"]["other_balance"]), 1_000_000_000)

    def test_import_mode_term_structure_uses_officer_period(self) -> None:
        self._insert_summary(
            "2027-02",
            "MODETERM",
            "902",
            "Mode Term",
            "Officer Term",
            "TERM",
            1000,
            short_term_balance=700,
            medium_long_term_balance=300,
        )
        self._insert_officer("2027-02", "MODETERM", "TERM", "Officer Term", 300, short_term_balance=0, medium_long_term_balance=300)

        row = self.repository.officer_list(OfficerCenterFilters(report_period="2027-02", mode=OFFICER_MODE_IMPORTED), page=1, page_size=10).rows[0]

        self.assertEqual(float(row["total_balance"]), 300)
        self.assertEqual(float(row["short_term_balance"]), 0)
        self.assertEqual(float(row["medium_long_term_balance"]), 300)
        self.assertAlmostEqual(float(row["medium_long_ratio"]), 100.0)

    def test_effective_mode_term_structure_uses_customer_summary(self) -> None:
        self._insert_summary(
            "2027-03",
            "EFFECTTERM",
            "903",
            "Effective Term",
            "Officer Effective",
            "EFF",
            1000,
            short_term_balance=700,
            medium_long_term_balance=300,
        )
        self._insert_officer("2027-03", "EFFECTTERM", "EFF", "Officer Effective", 300, short_term_balance=0, medium_long_term_balance=300)

        row = self.repository.officer_list(OfficerCenterFilters(report_period="2027-03", mode=OFFICER_MODE_EFFECTIVE), page=1, page_size=10).rows[0]

        self.assertEqual(float(row["total_balance"]), 1000)
        self.assertEqual(float(row["short_term_balance"]), 700)
        self.assertEqual(float(row["medium_long_term_balance"]), 300)
        self.assertAlmostEqual(float(row["medium_long_ratio"]), 30.0)

    def test_reimport_old_period_creates_officer_term_data(self) -> None:
        self._insert_summary("2027-04", "REIMPORT", "904", "Reimport Term", "Officer Reimport", "R01", 1000)
        self._insert_officer("2027-04", "REIMPORT", "R01", "Officer Reimport", 1000, short_term_balance=0, medium_long_term_balance=0, other_balance=0)
        before = self.repository.officer_list(OfficerCenterFilters(report_period="2027-04"), page=1, page_size=10).rows[0]
        self.assertFalse(before["term_structure_available"])

        folder = self.root / "reimport-term"
        self._write_ftpln(
            folder,
            "5491_FTPLN_20270430.csv",
            [self._ftpln_row("904", "Reimport Term", 1000, "DN8", officer="[R01] Officer Reimport")],
        )
        import_nim_dn(self.summary_repository, folder, replace_existing_periods=True)

        after = self.repository.officer_list(OfficerCenterFilters(report_period="2027-04"), page=1, page_size=10).rows[0]
        self.assertTrue(after["term_structure_available"])
        self.assertEqual(float(after["medium_long_term_balance"]), 1000)

    def test_officer_search_field_minimum_width(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = OfficerCenterWindow(self.main_database_path)
        try:
            self.assertGreaterEqual(window.search_box.minimumWidth(), 320)
            self.assertEqual(window.search_box.placeholderText(), "Tìm mã hoặc tên CBTD")
        finally:
            window.close()

    def test_officer_search_field_expands(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = OfficerCenterWindow(self.main_database_path)
        try:
            self.assertEqual(window.search_box.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
        finally:
            window.close()

    def test_officer_search_field_not_overlapped(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = OfficerCenterWindow(self.main_database_path)
        try:
            layout = window.search_box.parentWidget().layout()
            self.assertIs(layout.itemAtPosition(1, 3).widget(), window.search_box)
            self.assertIs(layout.itemAtPosition(1, 5).widget(), window.refresh_button)
            self.assertIs(layout.itemAtPosition(1, 6).widget(), window.clear_button)
            self.assertIs(layout.itemAtPosition(1, 7).widget(), window.export_button)
        finally:
            window.close()

    def test_dashboard_top_table_visible(self) -> None:
        tab = self._dashboard_tab()
        try:
            tab._apply_payload(self._dashboard_payload_rows(10))
            self.assertTrue(tab.top_table.isVisibleTo(tab))
            self.assertEqual(tab.dashboard_content_layout.indexOf(tab.top_table), 3)
        finally:
            tab.close()

    def test_dashboard_top_table_has_minimum_height(self) -> None:
        tab = self._dashboard_tab()
        try:
            self.assertGreaterEqual(tab.top_table.minimumHeight(), 320)
        finally:
            tab.close()

    def test_dashboard_scroll_reaches_top_table(self) -> None:
        tab = self._dashboard_tab()
        try:
            tab._apply_payload(self._dashboard_payload_rows(10))
            self.assertIs(tab.dashboard_scroll.widget(), tab.dashboard_content)
            self.assertGreaterEqual(tab.dashboard_content.sizeHint().height(), tab.top_table.minimumHeight())
        finally:
            tab.close()

    def test_top_table_scrolls_for_top_50(self) -> None:
        tab = self._dashboard_tab()
        try:
            tab._apply_payload(self._dashboard_payload_rows(50))
            self.assertEqual(tab.top_model.rowCount(), 50)
            self.assertEqual(tab.top_table.verticalScrollBarPolicy(), Qt.ScrollBarPolicy.ScrollBarAsNeeded)
            self.assertEqual(tab.top_table.minimumHeight(), 320)
        finally:
            tab.close()

    def test_dashboard_no_stretch_before_top_table(self) -> None:
        tab = self._dashboard_tab()
        try:
            top_index = tab.dashboard_content_layout.indexOf(tab.top_table)
            for index in range(top_index):
                self.assertIsNone(tab.dashboard_content_layout.itemAt(index).spacerItem())
        finally:
            tab.close()

    def test_dashboard_resize_does_not_hide_top_table(self) -> None:
        tab = self._dashboard_tab()
        try:
            tab.resize(900, 600)
            tab._apply_payload(self._dashboard_payload_rows(20))
            self.assertTrue(tab.top_table.isVisibleTo(tab))
            self.assertGreaterEqual(tab.top_table.minimumHeight(), 320)
        finally:
            tab.close()

    def test_average_customer_kpi_formatted_two_decimals(self) -> None:
        display, tooltip = kpi_display_values(KpiMetric("KH bình quân/CBTD", 231.45544554455446, "number"))

        self.assertEqual(display, "231,46")
        self.assertEqual(tooltip, "231,46")

    def test_movement_nim_before_previous_weighted(self) -> None:
        row = self._weighted_movement_row()

        self.assertAlmostEqual(float(row["previous_nim_before"]), 4.0)

    def test_movement_nim_after_previous_weighted(self) -> None:
        row = self._weighted_movement_row()

        self.assertAlmostEqual(float(row["previous_nim_after"]), 2.5)

    def test_movement_nim_before_current_weighted(self) -> None:
        row = self._weighted_movement_row()

        self.assertAlmostEqual(float(row["current_nim_before"]), 1.8)

    def test_movement_nim_after_current_weighted(self) -> None:
        row = self._weighted_movement_row()

        self.assertAlmostEqual(float(row["current_nim_after"]), 2.2)

    def test_movement_nim_before_change_percentage_points(self) -> None:
        row = self._weighted_movement_row()

        self.assertAlmostEqual(float(row["nim_before_change_pp"]), -2.2)

    def test_movement_nim_after_change_percentage_points(self) -> None:
        row = self._weighted_movement_row()

        self.assertAlmostEqual(float(row["nim_after_change_pp"]), -0.3)

    def test_movement_nim_previous_zero_balance_na(self) -> None:
        self._insert_summary("2027-07", "NEWNIM", "907", "New Nim", "New Nim Officer", "NEWNIM", 500, nim_before=6, nim_after=5)
        self._insert_officer("2027-07", "NEWNIM", "NEWNIM", "New Nim Officer", 500, nim_before=6, nim_after=5)

        row = self.repository.officer_movement(OfficerCenterFilters(report_period="2027-07", compare_period="2027-06"), page=1, page_size=10).rows[0]

        self.assertIsNone(row["previous_nim_before"])
        self.assertIsNone(row["previous_nim_after"])
        self.assertIsNone(row["nim_after_change_pp"])

    def test_movement_nim_current_zero_balance_na(self) -> None:
        self._insert_summary("2027-08", "OLDNIM", "908", "Old Nim", "Old Nim Officer", "OLDNIM", 500, nim_before=6, nim_after=5)
        self._insert_officer("2027-08", "OLDNIM", "OLDNIM", "Old Nim Officer", 500, nim_before=6, nim_after=5)

        row = self.repository.officer_movement(OfficerCenterFilters(report_period="2027-09", compare_period="2027-08"), page=1, page_size=10).rows[0]

        self.assertIsNone(row["current_nim_before"])
        self.assertIsNone(row["current_nim_after"])
        self.assertIsNone(row["nim_after_change_pp"])

    def test_movement_nim_import_mode(self) -> None:
        self._seed_mode_movement_fixture()

        row = next(
            item
            for item in self.repository.officer_movement(
                OfficerCenterFilters(report_period="2027-11", compare_period="2027-10", mode=OFFICER_MODE_IMPORTED),
                page=1,
                page_size=10,
            ).rows
            if item["officer_code"] == "MODEA"
        )

        self.assertEqual(float(row["current_balance"]), 100)
        self.assertAlmostEqual(float(row["current_nim_after"]), 20.0)

    def test_movement_nim_effective_mode(self) -> None:
        self._seed_mode_movement_fixture()

        row = self.repository.officer_movement(
            OfficerCenterFilters(report_period="2027-11", compare_period="2027-10", mode=OFFICER_MODE_EFFECTIVE),
            page=1,
            page_size=10,
        ).rows[0]

        self.assertEqual(row["officer_code"], "MODEA")
        self.assertEqual(float(row["current_balance"]), 1000)
        self.assertAlmostEqual(float(row["current_nim_after"]), 7.0)

    def test_movement_nim_export_headers(self) -> None:
        self._seed_weighted_movement_fixture()
        output = self.root / "movement-nim-headers.xlsx"

        export_officer_center_workbook(
            self.repository,
            OfficerCenterFilters(report_period="2027-06", compare_period="2027-05"),
            output,
        )
        workbook = load_workbook(output, data_only=True)
        try:
            headers = [cell.value for cell in workbook["BienDongCBTD"][1]]
            self.assertIn("NIM trước ĐC kỳ trước", headers)
            self.assertIn("NIM sau ĐC kỳ trước", headers)
            self.assertIn("NIM trước ĐC kỳ hiện tại", headers)
            self.assertIn("NIM sau ĐC kỳ hiện tại", headers)
            self.assertIn("Thay đổi NIM trước ĐC (điểm %)", headers)
            self.assertIn("Thay đổi NIM sau ĐC (điểm %)", headers)
        finally:
            workbook.close()

    def test_movement_nim_export_numeric_values(self) -> None:
        self._seed_weighted_movement_fixture()
        output = self.root / "movement-nim-values.xlsx"

        export_officer_center_workbook(
            self.repository,
            OfficerCenterFilters(report_period="2027-06", compare_period="2027-05"),
            output,
        )
        workbook = load_workbook(output, data_only=True)
        try:
            sheet = workbook["BienDongCBTD"]
            headers = [cell.value for cell in sheet[1]]
            code_column = headers.index("Mã CBTD") + 1
            change_column = headers.index("Thay đổi NIM sau ĐC (điểm %)") + 1
            row_index = next(index for index in range(2, sheet.max_row + 1) if sheet.cell(index, code_column).value == "WNIM")
            self.assertIsInstance(sheet.cell(row_index, change_column).value, (int, float))
            self.assertAlmostEqual(float(sheet.cell(row_index, change_column).value), -0.3)
        finally:
            workbook.close()

    def _dashboard_tab(self) -> OfficerDashboardTab:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = OfficerDashboardTab(
            self.repository,
            lambda: OfficerCenterFilters(report_period="2026-03", period_from="2026-03", period_to="2026-04"),
        )
        tab.show()
        app.processEvents()
        return tab

    def _dashboard_payload_rows(self, count: int) -> dict[str, object]:
        rows = [
            {
                "officer_code": f"T{index:03d}",
                "officer_name": f"Officer {index:03d}",
                "branch_name": "Chi nhánh",
                "office_name": "PGD",
                "customer_count": index,
                "total_balance": 1_000_000 + index,
                "balance_change": 10_000 + index,
                "nim_after": 2.0,
                "attention_balance": 0,
                "bad_debt_balance": 0,
                "bad_debt_ratio": 0,
            }
            for index in range(1, count + 1)
        ]
        return {
            "kpis": {
                "active_officer_count": 2,
                "total_balance": 1000,
                "officer_customer_occurrence_count": 3,
                "unique_customer_count": 2,
                "average_customer_per_officer": 1.5,
            },
            "balance_trend": [],
            "officer_count_trend": [],
            "metric_trend": [],
            "debt_structure": {},
            "top_rows": rows,
            "top_limit": count,
            "mode_label": "Theo phân bổ import",
        }

    def _import_officer_term_fixture(self, period: str) -> None:
        folder = self.root / f"term-{period}"
        compact_period = period.replace("-", "")
        self._write_ftpln(
            folder,
            f"5491_FTPLN_{compact_period}31.csv",
            [
                self._ftpln_row("101", "Khach A", 6_000_000_000, "DN1", officer="[A] Officer A"),
                self._ftpln_row("101", "Khach A", 2_000_000_000, "DN8", officer="[A] Officer A"),
                self._ftpln_row("102", "Khach B", 1_000_000_000, "DN15", officer="[B] Officer B"),
                self._ftpln_row("102", "Khach B", 1_000_000_000, "DN99", officer="[B] Officer B"),
            ],
        )
        import_nim_dn(self.summary_repository, folder, replace_existing_periods=True)

    def _write_ftpln(self, folder: Path, filename: str, rows: list[str]) -> Path:
        folder.mkdir(parents=True, exist_ok=True)
        path = folder / filename
        path.write_text("\n".join([FTPLN_HEADER, *rows]), encoding="utf-8")
        return path

    def _ftpln_row(
        self,
        customer_sequence: str,
        customer_name: str,
        balance: float,
        ftp_code: str,
        *,
        officer: str,
        ftp: float = 2,
        intrt: float = 10,
        adjustment: float = 1,
    ) -> str:
        return f"5491,{ftp},{intrt},{adjustment},{officer},00,{balance},,CN,{ftp_code},{customer_sequence},{customer_name}"

    def _officer_period_rows(self, period: str) -> list:
        with closing(self.customer_repository.connect()) as connection:
            return list(
                connection.execute(
                    """
                    SELECT *
                    FROM customer_officer_period
                    WHERE period = ?
                    ORDER BY officer_code
                    """,
                    (period,),
                ).fetchall()
            )

    def _weighted_movement_row(self) -> dict[str, object]:
        self._seed_weighted_movement_fixture()
        rows = self.repository.officer_movement(
            OfficerCenterFilters(report_period="2027-06", compare_period="2027-05"),
            page=1,
            page_size=20,
        ).rows
        return next(row for row in rows if row["officer_code"] == "WNIM")

    def _seed_weighted_movement_fixture(self) -> None:
        existing = self.repository.officer_movement(
            OfficerCenterFilters(report_period="2027-06", compare_period="2027-05"),
            page=1,
            page_size=1,
        )
        if any(row.get("officer_code") == "WNIM" for row in existing.rows):
            return
        self._insert_summary("2027-05", "WNIM1", "905", "Weighted Nim 1", "Weighted Nim", "WNIM", 100, nim_before=10, nim_after=4)
        self._insert_officer("2027-05", "WNIM1", "WNIM", "Weighted Nim", 100, nim_before=10, nim_after=4)
        self._insert_summary("2027-05", "WNIM2", "906", "Weighted Nim 2", "Weighted Nim", "WNIM", 300, nim_before=2, nim_after=2)
        self._insert_officer("2027-05", "WNIM2", "WNIM", "Weighted Nim", 300, nim_before=2, nim_after=2)
        self._insert_summary("2027-06", "WNIM1", "905", "Weighted Nim 1", "Weighted Nim", "WNIM", 200, nim_before=5, nim_after=3)
        self._insert_officer("2027-06", "WNIM1", "WNIM", "Weighted Nim", 200, nim_before=5, nim_after=3)
        self._insert_summary("2027-06", "WNIM2", "906", "Weighted Nim 2", "Weighted Nim", "WNIM", 800, nim_before=1, nim_after=2)
        self._insert_officer("2027-06", "WNIM2", "WNIM", "Weighted Nim", 800, nim_before=1, nim_after=2)

    def _seed_mode_movement_fixture(self) -> None:
        existing = self.repository.officer_movement(
            OfficerCenterFilters(report_period="2027-11", compare_period="2027-10"),
            page=1,
            page_size=1,
        )
        if existing.total_rows:
            return
        self._insert_summary("2027-10", "MODECUST", "910", "Mode Customer", "Officer Mode A", "MODEA", 1000, nim_after=5)
        self._insert_officer("2027-10", "MODECUST", "MODEA", "Officer Mode A", 100, nim_after=10)
        self._insert_officer("2027-10", "MODECUST", "MODEB", "Officer Mode B", 900, nim_after=4.4444444444)
        self._insert_summary("2027-11", "MODECUST", "910", "Mode Customer", "Officer Mode A", "MODEA", 1000, nim_after=7)
        self._insert_officer("2027-11", "MODECUST", "MODEA", "Officer Mode A", 100, nim_after=20)
        self._insert_officer("2027-11", "MODECUST", "MODEB", "Officer Mode B", 900, nim_after=5.5555555556)

    def _seed_data(self) -> None:
        self.customer_repository.upsert_officer_directory(
            officer_code="CB01", officer_name="Nguyen Van A", branch_code="5491", transaction_office="00"
        )
        self.customer_repository.upsert_officer_directory(
            officer_code="CB02", officer_name="Tran Van B", branch_code="5491", transaction_office="01"
        )
        self.customer_repository.upsert_officer_directory(
            officer_code="CB03", officer_name="Le Van C", branch_code="5491", transaction_office="00"
        )
        self.customer_repository.upsert_officer_directory(
            officer_code="CB04", officer_name="Nguyen Van A", branch_code="5491", transaction_office="01"
        )
        self._insert_summary("2026-03", "CUST1", "001", "Khach 1", "Nguyen Van A", "CB01", 1000, has_multiple=1)
        self._insert_officer("2026-03", "CUST1", "CB01", "Nguyen Van A", 600, debt_group_1=600)
        self._insert_officer("2026-03", "CUST1", "CB02", "Tran Van B", 400, debt_group_2=400)
        self._insert_override("CUST1", "2026-03", "2026-12", "CB03", "Le Van C")
        self._insert_summary("2026-04", "CUST1", "001", "Khach 1", "Nguyen Van A", "CB01", 1000, has_multiple=1)
        self._insert_officer("2026-04", "CUST1", "CB04", "Nguyen Van A", 1000, debt_group_3=1000)
        self._insert_summary("2026-04", "CUST2", "002", "Khach 2", "Nguyen Van A", "CB01", 300)
        self._insert_officer("2026-04", "CUST2", "CB01", "Nguyen Van A", 300, debt_group_1=300)

    def _insert_summary(
        self,
        period: str,
        customer_code: str,
        sequence: str,
        customer_name: str,
        officer_name: str,
        officer_code: str,
        balance: float,
        *,
        has_multiple: int = 0,
        average_rate: float = 7,
        nim_before: float = 3,
        nim_after: float = 2,
        debt_group_1: float | None = None,
        debt_group_2: float = 0,
        debt_group_3: float = 0,
        short_term_balance: float | None = None,
        medium_long_term_balance: float = 0,
        other_balance: float = 0,
    ) -> None:
        short_balance = balance if short_term_balance is None else short_term_balance
        medium_long_ratio = medium_long_term_balance / balance * 100 if balance else 0
        now = now_text()
        with closing(self.customer_repository.connect()) as connection:
            connection.execute(
                """
                INSERT INTO customer_period_summary(
                    period, customer_code, branch_code, customer_sequence, customer_name, customer_type,
                    primary_officer_code, primary_officer_name, officer_count, has_multiple_officers,
                    total_balance, short_term_balance, medium_long_term_balance, other_balance, medium_long_ratio,
                    interest_rate_numerator, nim_before_numerator, nim_after_numerator,
                    average_rate, nim_before, nim_after, source_loan_count, created_at, updated_at,
                    has_debt_group_data, worst_debt_group, debt_group_1_balance, debt_group_2_balance, debt_group_3_balance
                )
                VALUES (?, ?, '5491', ?, ?, 'CN', ?, ?, ?, ?,
                        ?, ?, ?, ?, ?,
                        ?, ?, ?, 7, 3, 2, 1, ?, ?,
                        1, '01', ?, ?, ?)
                """,
                (
                    period,
                    customer_code,
                    sequence,
                    customer_name,
                    officer_code,
                    officer_name,
                    2 if has_multiple else 1,
                    has_multiple,
                    balance,
                    short_balance,
                    medium_long_term_balance,
                    other_balance,
                    medium_long_ratio,
                    balance * average_rate,
                    balance * nim_before,
                    balance * nim_after,
                    now,
                    now,
                    balance if debt_group_1 is None else debt_group_1,
                    debt_group_2,
                    debt_group_3,
                ),
            )
            connection.commit()

    def _insert_officer(
        self,
        period: str,
        customer_code: str,
        officer_code: str,
        officer_name: str,
        balance: float,
        *,
        debt_group_1: float = 0,
        debt_group_2: float = 0,
        debt_group_3: float = 0,
        average_rate: float = 7,
        nim_before: float = 3,
        nim_after: float = 2,
        short_term_balance: float | None = None,
        medium_long_term_balance: float = 0,
        other_balance: float = 0,
    ) -> None:
        now = now_text()
        with closing(self.customer_repository.connect()) as connection:
            connection.execute(
                """
                INSERT INTO customer_officer_period(
                    period, customer_code, officer_code, officer_name, branch_code, transaction_office,
                    balance_managed, short_term_balance, medium_long_term_balance, other_balance,
                    source_loan_count, interest_rate_numerator, nim_before_numerator,
                    nim_after_numerator, is_primary, created_at, has_debt_group_data,
                    debt_group_1_balance, debt_group_2_balance, debt_group_3_balance
                )
                VALUES (?, ?, ?, ?, '5491', '00', ?, ?, ?, ?, 1, ?, ?, ?, 1, ?, 1, ?, ?, ?)
                """,
                (
                    period,
                    customer_code,
                    officer_code,
                    officer_name,
                    balance,
                    balance if short_term_balance is None else short_term_balance,
                    medium_long_term_balance,
                    other_balance,
                    balance * average_rate,
                    balance * nim_before,
                    balance * nim_after,
                    now,
                    debt_group_1,
                    debt_group_2,
                    debt_group_3,
                ),
            )
            connection.commit()

    def _seed_many_officers(self, period: str, count: int) -> None:
        for index in range(count):
            code = f"CM{index:02d}"
            customer = f"CM-CUST-{index:02d}"
            balance = 100 + index
            self._insert_summary(period, customer, f"8{index:02d}", f"Customer {index}", code, code, balance)
            self._insert_officer(period, customer, code, code, balance)

    def _seed_weighted_period(self) -> None:
        existing = self.repository.officer_list(OfficerCenterFilters(report_period="2026-09"), page=1, page_size=1)
        if existing.total_rows:
            return
        self._insert_summary(
            "2026-09",
            "W1",
            "901",
            "Weighted 1",
            "Weighted A",
            "W01",
            100,
            average_rate=10,
            nim_before=5,
            nim_after=4,
            debt_group_1=100,
        )
        self._insert_officer("2026-09", "W1", "W01", "Weighted A", 100, average_rate=10, nim_before=5, nim_after=4, debt_group_1=100)
        self._insert_summary(
            "2026-09",
            "W2",
            "902",
            "Weighted 2",
            "Weighted B",
            "W02",
            900,
            average_rate=1,
            nim_before=1,
            nim_after=1,
            debt_group_1=0,
            debt_group_3=900,
        )
        self._insert_officer("2026-09", "W2", "W02", "Weighted B", 900, average_rate=1, nim_before=1, nim_after=1, debt_group_3=900)

    def _insert_override(self, customer_code: str, start: str, end: str, officer_code: str, officer_name: str) -> None:
        now = now_text()
        with closing(self.customer_repository.connect()) as connection:
            connection.execute(
                """
                INSERT INTO customer_officer_override(
                    customer_code, effective_from_period, effective_to_period, officer_code,
                    officer_name, reason, is_active, created_by, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, 'test', 1, 'test', ?, ?)
                """,
                (customer_code, start, end, officer_code, officer_name, now, now),
            )
            connection.commit()


class _Host(QWidget):
    def __init__(self, database_path: Path) -> None:
        super().__init__()
        self.settings_database = type("Settings", (), {"database_path": database_path})()


if __name__ == "__main__":
    unittest.main()
