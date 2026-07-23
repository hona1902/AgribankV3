from __future__ import annotations

from pathlib import Path
import os
import tempfile
import unittest
from contextlib import closing
from openpyxl import Workbook, load_workbook
from PySide6.QtWidgets import QApplication, QPushButton

from agribank_v3.quiz import (
    Question,
    QuizDatabase,
    QuizSession,
    export_questions_to_excel,
    import_questions_from_excel,
)
from agribank_v3.ui.dialogs.quiz import QuizWidget


def sample_question(question_id: int, correct: str = "A") -> Question:
    return Question(
        id=question_id,
        legacy_id=float(question_id),
        number=question_id,
        text=f"Câu hỏi {question_id}",
        options={"A": "Đáp án A", "B": "Đáp án B"},
        correct_answer=correct,
        topic_code="TEST",
        topic_name="Kiểm thử",
        source_reference="",
    )


class QuizSessionTests(unittest.TestCase):
    def test_answer_shuffle_preserves_correct_answer_content(self) -> None:
        question = sample_question(1, "B")
        shuffled = QuizWidget._shuffle_question_answers(question)

        self.assertEqual(set(shuffled.options.values()), set(question.options.values()))
        self.assertEqual(
            shuffled.options[shuffled.correct_answer],
            question.options[question.correct_answer],
        )
        self.assertEqual(question.options, {"A": "Đáp án A", "B": "Đáp án B"})

    def test_answer_shuffle_keeps_locked_positions(self) -> None:
        question = Question(
            id=1,
            legacy_id=None,
            number=1,
            text="Câu hỏi",
            options={"A": "A", "B": "B", "C": "C", "D": "D"},
            correct_answer="B",
            topic_code="TEST",
            topic_name="Kiểm thử",
            source_reference="",
            locked_answers="AC",
        )

        for _ in range(10):
            shuffled = QuizWidget._shuffle_question_answers(question)
            self.assertEqual(shuffled.options["A"], "A")
            self.assertEqual(shuffled.options["C"], "C")

    def test_scoring_and_rating(self) -> None:
        questions = [sample_question(1), sample_question(2, "B")]
        session = QuizSession(questions, "Nguyễn Văn A", "Kiểm thử")
        session.answer("A")
        session.current_index = 1
        session.answer("A")

        result = session.finish()

        self.assertEqual(result.total, 2)
        self.assertEqual(result.correct, 1)
        self.assertEqual(result.incorrect, 1)
        self.assertEqual(result.percentage, 50.0)
        self.assertEqual(result.rating, "Trung bình")

    def test_rating_boundaries(self) -> None:
        self.assertEqual(QuizSession.rating_for(49.99), "Yếu")
        self.assertEqual(QuizSession.rating_for(50), "Trung bình")
        self.assertEqual(QuizSession.rating_for(75), "Khá")
        self.assertEqual(QuizSession.rating_for(90), "Xuất sắc")


class QuizDatabaseTests(unittest.TestCase):
    def test_export_questions_to_excel_layout(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "bo-cau-hoi.xlsx"
            question = Question(
                id=1,
                legacy_id=None,
                number=1,
                text="Nội dung câu hỏi",
                options={
                    "A": "Phương án A",
                    "B": "Phương án B",
                    "C": "Phương án C",
                    "D": "Phương án D",
                },
                correct_answer="C",
                topic_code="KTC",
                topic_name="Kiến thức chung",
                source_reference="",
                subject_name="Ngân hàng điện tử",
            )

            export_questions_to_excel(
                output,
                [question],
                subtitle="Chuyên đề: Ngân hàng điện tử (1 Câu)",
            )

            workbook = load_workbook(output)
            sheet = workbook["Bo Cau Hoi"]
            self.assertEqual(sheet["A1"].value, "Bộ Câu Hỏi Ôn Tập, Khảo Sát Kiến Thức Nghiệp Vụ")
            self.assertEqual(
                sheet["A2"].value,
                "Chuyên đề: Ngân hàng điện tử (1 Câu)",
            )
            self.assertEqual(sheet["A4"].value, "Câu Hỏi")
            self.assertEqual(sheet["C4"].value, "Đúng")
            self.assertEqual(sheet["D4"].value, "Chọn")
            self.assertEqual(sheet["A5"].value, "Câu 1:")
            self.assertEqual(sheet["C5"].value, "C")
            self.assertEqual(sheet["A8"].value, "C:")
            self.assertEqual(sheet["B8"].font.color.rgb, "00008000")
            self.assertTrue(sheet.data_validations.count)
            workbook.close()

    def test_business_topic_question_crud_and_saved_quotas(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = QuizDatabase(
                sqlite_path=Path(directory) / "quiz.db",
                access_path=Path(directory) / "missing.mdb",
            )
            database.save_business_topic("TD", "Tín dụng")
            subject_id = database.save_question_topic("TD", "Tín dụng cá nhân")
            question_id = database.save_question(
                question_id=None,
                topic_code="TD",
                subject_id=subject_id,
                text="Câu hỏi mới",
                options={"A": "Đúng", "B": "Sai", "C": "", "D": ""},
                correct_answer="A",
                locked_answers="AC",
            )
            database.save_random_exam_setting({"TD": 12})
            database.save_random_subject_setting({subject_id: 1})

            question = database.question_by_id(question_id)

            self.assertIsNotNone(question)
            self.assertEqual(question.options["A"], "Đúng")
            self.assertEqual(question.subject_name, "Tín dụng cá nhân")
            self.assertEqual(question.locked_answers, "AC")
            self.assertEqual(database.load_random_exam_setting(), {"TD": 12})
            self.assertEqual(
                database.load_random_subject_setting(),
                {subject_id: 1},
            )
            self.assertEqual(database.business_topics(), [("TD", "Tín dụng", 1)])

            database.delete_question(question_id)
            self.assertEqual(database.search_questions(), [])

    def test_import_questions_from_excel(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            database = QuizDatabase(
                sqlite_path=root / "quiz.db",
                access_path=root / "missing.mdb",
            )
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Mã nghiệp vụ", "Nghiệp vụ", "Chuyên đề", "Câu hỏi",
                    "Đáp án A", "Đáp án B", "Đáp án C", "Đáp án D",
                    "Đáp án đúng", "Đáp án không đảo", "Nguồn",
                ]
            )
            sheet.append(
                [
                    "CNTT", "Công nghệ thông tin", "Ngân hàng điện tử",
                    "Câu nhập Excel", "A1", "B1", "C1", "D1",
                    "B", "C,D", "VB 01",
                ]
            )
            excel_path = root / "questions.xlsx"
            workbook.save(excel_path)
            workbook.close()

            result = import_questions_from_excel(excel_path, database)

            self.assertEqual(result.imported, 1)
            question = database.search_questions("Câu nhập Excel")[0]
            self.assertEqual(question.correct_answer, "B")
            self.assertEqual(question.options["D"], "D1")
            self.assertEqual(question.subject_name, "Ngân hàng điện tử")
            self.assertEqual(question.locked_answers, "CD")

    def test_bulk_delete_questions_by_business(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = QuizDatabase(
                sqlite_path=Path(directory) / "quiz.db",
                access_path=Path(directory) / "missing.mdb",
            )
            for code in ("A", "B"):
                database.save_business_topic(code, f"Nghiệp vụ {code}")
                subject_id = database.save_question_topic(code, f"Chuyên đề {code}")
                database.save_question(
                    question_id=None,
                    topic_code=code,
                    subject_id=subject_id,
                    text=f"Câu {code}",
                    options={"A": "Đúng", "B": "Sai"},
                    correct_answer="A",
                )

            deleted = database.delete_questions_bulk("A")

            self.assertEqual(deleted, 1)
            self.assertEqual(
                [question.topic_code for question in database.search_questions()],
                ["B"],
            )

    def test_questions_by_quotas_combines_topic_codes(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = QuizDatabase(
                sqlite_path=Path(directory) / "quiz.db",
                access_path=Path(directory) / "missing.mdb",
            )
            with closing(database.connect()) as connection, connection:
                for question_id, topic_code in enumerate(
                    ("NV", "NV", "CHUNG"),
                    start=1,
                ):
                    connection.execute(
                        """
                        INSERT INTO questions (
                            id, legacy_id, question_number, question_text,
                            option_a, option_b, correct_answer,
                            topic_code, topic_name, source_reference, source_active
                        ) VALUES (?, ?, ?, ?, 'A', 'B', 'A', ?, ?, '', 1)
                        """,
                        (
                            question_id,
                            question_id,
                            question_id,
                            f"Câu {question_id}",
                            topic_code,
                            topic_code,
                        ),
                    )

            questions = database.questions_by_quotas({"NV": 2, "CHUNG": 1})

            self.assertEqual(len(questions), 3)
            self.assertEqual({question.topic_code for question in questions}, {"NV", "CHUNG"})

    def test_questions_by_subject_quotas_keeps_subjects_separate(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = QuizDatabase(
                sqlite_path=Path(directory) / "quiz.db",
                access_path=Path(directory) / "missing.mdb",
            )
            database.save_business_topic("KTC", "Kiến thức chung")
            first = database.save_question_topic("KTC", "Điều lệ")
            second = database.save_question_topic("KTC", "TEST")
            for subject_id, text in ((first, "Câu Điều lệ"), (second, "Câu TEST")):
                database.save_question(
                    question_id=None,
                    topic_code="KTC",
                    subject_id=subject_id,
                    text=text,
                    options={"A": "Đúng", "B": "Sai"},
                    correct_answer="A",
                )

            questions = database.questions_by_subject_quotas({second: 1})

            self.assertEqual(len(questions), 1)
            self.assertEqual(questions[0].subject_name, "TEST")

    def test_attempt_and_answers_are_persisted(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = QuizDatabase(
                sqlite_path=Path(directory) / "quiz.db",
                access_path=Path(directory) / "missing.mdb",
            )
            with closing(database.connect()) as connection, connection:
                connection.execute(
                    """
                    INSERT INTO questions (
                        id, legacy_id, question_number, question_text,
                        option_a, option_b, correct_answer,
                        topic_code, topic_name, source_reference, source_active
                    ) VALUES (1, 1, 1, 'Câu hỏi', 'A', 'B', 'A',
                              'TEST', 'Kiểm thử', '', 1)
                    """
                )
            questions = database.questions("Kiểm thử", 1, False)
            session = QuizSession(questions, "Nguyễn Văn A", "Kiểm thử")
            session.answer("A")
            session.check_current()
            result = session.finish()

            attempt_id = database.save_attempt(session, result)

            with closing(database.connect()) as connection, connection:
                attempt = connection.execute(
                    "SELECT * FROM quiz_attempts WHERE id = ?",
                    (attempt_id,),
                ).fetchone()
                answer = connection.execute(
                    "SELECT * FROM quiz_answers WHERE attempt_id = ?",
                    (attempt_id,),
                ).fetchone()
            self.assertEqual(attempt["correct_answers"], 1)
            self.assertEqual(answer["selected_answer"], "A")
            self.assertEqual(answer["checked_before_finish"], 1)


class QuizWidgetTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        cls.app = QApplication.instance() or QApplication([])

    def test_data_management_button_opens_data_management_tab(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            database = QuizDatabase(
                sqlite_path=Path(directory) / "quiz.db",
                access_path=Path(directory) / "missing.mdb",
            )
            database.save_business_topic("TEST", "Kiểm thử")
            database.save_question(
                question_id=None,
                topic_code="TEST",
                text="Câu hỏi kiểm thử",
                options={"A": "Đúng", "B": "Sai"},
                correct_answer="A",
            )
            widget = QuizWidget(database=database)
            self.addCleanup(widget.deleteLater)

            buttons = [
                button
                for button in widget.findChildren(QPushButton)
                if button.text() == "Quản lý dữ liệu"
            ]
            self.assertTrue(buttons)
            buttons[0].click()

            self.assertEqual(widget.pages.currentIndex(), widget.data_management_tab_index)
            self.assertEqual(widget.pages.currentWidget(), widget.data_admin_tabs.parentWidget())
            self.assertEqual(widget.pages.tabText(widget.pages.currentIndex()), "Quản Lý Dữ Liệu")


if __name__ == "__main__":
    unittest.main()
