from __future__ import annotations

import unittest
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from agribank_v3.file_merge import (
    FileMergeError,
    merge_same_structure_csv_to_csv,
    merge_same_structure_csv_to_xlsx,
    merge_same_structure_excel_to_xlsx,
)
from agribank_v3.settings import BranchProfile
from agribank_v3.settlement import (
    ACTIVE_SETTLEMENT_SPECS,
    LEGACY_SETTLEMENT_SPECS,
    SETTLEMENT_SPECS,
    SettlementEngine,
    SettlementError,
    SettlementOptions,
    SettlementRequest,
)
from agribank_v3.settlement.transforms import (
    decode_telex,
    excel_column_name,
    is_branch_code,
    is_code_like,
    normalize_customer_id,
    parse_vietnamese_amount,
    parse_yyyymmdd,
    vietnamese_report_date,
)
from agribank_v3.settlement.processors import (
    Mau04Processor,
    Mau05Processor,
    Mau06Processor,
    Mau0708Processor,
    Mau09Processor,
    Mau1314Processor,
    Mau1516Processor,
    Mau18Processor,
    Mau20aProcessor,
    Mau22Processor,
    Mau23Processor,
    Mau24Processor,
    Mau30Processor,
)
from agribank_v3.settlement.processors.summary05 import Summary05Processor
from agribank_v3.settlement.processors.mau05 import CollateralRecord
from agribank_v3.settlement.processors.mau13_14 import DepositRecord
from agribank_v3.settlement.processors.mau24 import ReceivablePayableRecord


class SettlementRegistryTests(unittest.TestCase):
    def test_registry_contains_every_active_ribbon_entry(self) -> None:
        self.assertEqual(len(ACTIVE_SETTLEMENT_SPECS), 29)
        self.assertEqual(len(LEGACY_SETTLEMENT_SPECS), 7)
        self.assertEqual(len(SETTLEMENT_SPECS), 36)
        self.assertEqual(
            len(SETTLEMENT_SPECS),
            len(set(SETTLEMENT_SPECS)),
        )

    def test_shared_report_families_are_explicit(self) -> None:
        self.assertEqual(
            SETTLEMENT_SPECS["accounting.13"].processor_family,
            SETTLEMENT_SPECS["accounting.14"].processor_family,
        )
        self.assertEqual(
            SETTLEMENT_SPECS["credit.15a"].processor_family,
            SETTLEMENT_SPECS["credit.16"].processor_family,
        )


class SettlementEngineTests(unittest.TestCase):
    def test_requires_database_branch_code(self) -> None:
        request = SettlementRequest(
            spec=SETTLEMENT_SPECS["credit.05"],
            profile=BranchProfile(),
        )
        with self.assertRaisesRegex(SettlementError, "mã chi nhánh"):
            SettlementEngine().execute(request)

    def test_reports_unmigrated_processor_without_using_vba(self) -> None:
        request = SettlementRequest(
            spec=SETTLEMENT_SPECS["credit.05"],
            profile=BranchProfile(branch_code="5491"),
        )
        with self.assertRaises(SettlementError) as raised:
            SettlementEngine().execute(request)
        self.assertEqual(raised.exception.code, "processor_not_migrated")


class SettlementTransformTests(unittest.TestCase):
    def test_excel_column_name_supports_full_xlsx_range(self) -> None:
        self.assertEqual(excel_column_name(1), "A")
        self.assertEqual(excel_column_name(26), "Z")
        self.assertEqual(excel_column_name(27), "AA")
        self.assertEqual(excel_column_name(16_384), "XFD")

    def test_parses_ipcas_date_without_locale_dependency(self) -> None:
        self.assertEqual(str(parse_yyyymmdd("20251231")), "2025-12-31")
        self.assertEqual(
            vietnamese_report_date("20251231"),
            "Ngày 31 tháng 12 năm 2025",
        )
        self.assertIsNone(parse_yyyymmdd("20250230"))

    def test_normalizes_customer_id_from_vba_text_values(self) -> None:
        self.assertEqual(
            normalize_customer_id("'000123", "5491", include_branch=True),
            "5491000123",
        )
        self.assertEqual(
            normalize_customer_id("'000123", "5491", include_branch=False),
            "000123",
        )

    def test_parses_vietnamese_amount(self) -> None:
        self.assertEqual(
            parse_vietnamese_amount("1.234.567,89"),
            parse_vietnamese_amount(1234567.89),
        )
        self.assertEqual(parse_vietnamese_amount("-"), 0)

    def test_recognizes_balance_codes(self) -> None:
        self.assertTrue(is_branch_code(" 5491 "))
        self.assertFalse(is_branch_code("Tổng cộng"))
        self.assertTrue(is_code_like("211101_VND"))
        self.assertFalse(is_code_like("Cộng chi nhánh"))

    def test_decodes_legacy_telex_labels(self) -> None:
        self.assertEqual(
            decode_telex("Toorng howjp quyeest toasn"),
            "Tổng hợp quyết toán",
        )


class SettlementFixtureParityTests(unittest.TestCase):
    FIXTURE_DIR = (
        Path(__file__).resolve().parents[1] / "DuLieuTEST" / "SoSanh"
    )
    PROFILE = BranchProfile(
        branch_code="5491",
        branch_name="Chi nhánh Lộc Phát Lâm Đồng",
        report_location="Tân Hà",
    )

    @staticmethod
    def _normalized(value: object) -> object:
        if value is None:
            return ""
        if isinstance(value, Decimal):
            return int(value) if value == value.to_integral_value() else float(value)
        return value.date() if isinstance(value, datetime) else value

    def test_mau04_reads_ic_xls_and_builds_report(self) -> None:
        source = self.FIXTURE_DIR / "IC_100435.XLS"
        if not source.exists():
            self.skipTest("Mẫu 04 fixture is not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.04"],
            BranchProfile(
                branch_code="5491",
                branch_name="chi nhánh: Lộc Phát-Lâm Đồng",
                reporting_branch_name="chi nhánh: Lộc Phát-Lâm Đồng",
                report_location="Hà Lâm Hà",
                report_preparer="Thái Thị Mỹ Hằng",
                phone="02633.854.021",
            ),
            options=SettlementOptions(output_prefix="BN"),
            source_paths=(source,),
        )
        processor = Mau04Processor()
        rows = processor.read_source(source)
        sheet = processor.build_workbook(request, processor.summarize(rows))["04"]

        self.assertEqual(len(rows), 15)
        self.assertEqual(sheet["A1"].value, "NGÂN HÀNG NÔNG NGHIỆP")
        self.assertEqual(sheet["S1"].value, "1. Mẫu số 04/QT")
        self.assertEqual(sheet["S1"].font.name, "Times New Roman")
        self.assertEqual(sheet["S2"].font.name, "Times New Roman")
        self.assertEqual(sheet["A5"].font.sz, 16)
        self.assertEqual(sheet["A6"].font.sz, 14)
        self.assertEqual(sheet["B13"].value, "Séc")
        self.assertEqual(sheet["H13"].value, 1000)
        self.assertEqual(sheet["I13"].value, 950000)
        self.assertEqual(sheet["B14"].value, "Sổ tiết kiệm có kỳ hạn")
        self.assertEqual(sheet["H14"].value, 4500)
        self.assertEqual(sheet["I14"].value, 11565000)
        self.assertEqual(sheet["O14"].value, "=D14+F14+G14+H14-J14-K14-L14-M14")
        self.assertEqual(sheet["B15"].value, "Sổ tiết kiệm")
        self.assertEqual(sheet["L15"].value, 75)
        self.assertEqual(sheet["R15"].value, 1226456)
        self.assertEqual(sheet["B19"].value, "TỔNG CỘNG")
        self.assertEqual(sheet["P19"].value, "=SUM(P13:P18)")
        self.assertEqual(sheet["Q19"].value, "=SUM(Q13:Q18)")
        self.assertEqual(sheet["R19"].value, "=SUM(R13:R18)")
        self.assertEqual(sheet["A6"].value, f"Ngày 30 tháng 6 năm {date.today().year}")
        self.assertEqual(sheet["R20"].value, f"Hà Lâm Hà, Ngày 30 tháng 6 năm {date.today().year}")
        self.assertEqual(sheet["A21"].value, "LẬP BIỂU")
        self.assertEqual(sheet["R21"].value, "GIÁM ĐỐC")
        self.assertEqual(sheet["O12"].alignment.wrap_text, True)
        self.assertEqual(sheet["O12"].font.sz, 8)
        self.assertEqual(sheet.column_dimensions["B"].width, 24.5)
        self.assertEqual(sheet.column_dimensions["C"].width, 7.5)
        self.assertEqual(sheet.column_dimensions["O"].width, 14)
        self.assertEqual(sheet["A19"].border.top.style, "thin")
        self.assertEqual(str(sheet.print_area), "'04'!$A$1:$W$30")
        self.assertEqual(sheet.print_title_rows, "$8:$12")
        self.assertEqual(sheet.oddHeader.center.text, "&P/&N")

    def test_mau08_builds_grouped_fixed_asset_inventory(self) -> None:
        source = self.FIXTURE_DIR / "FA_100586.xls"
        if not source.exists():
            self.skipTest("Mẫu 08 fixture is not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.08"],
            self.PROFILE,
            options=SettlementOptions(output_prefix="QT"),
            source_paths=(source,),
        )
        processor = Mau0708Processor()
        records = processor.read_mau08_source(request, source)
        sheet = processor.build_mau08_workbook(request, records)["8"]

        self.assertEqual(len(records), 19)
        self.assertEqual(sheet["B11"].value, "Nhà cửa, vật kiến trúc (TK3012)")
        self.assertEqual(sheet["C11"].value, 1)
        self.assertEqual(sheet["D11"].value, 9329120577)
        self.assertEqual(sheet["B12"].value, "Máy móc, thiết bị (TK 3013)")
        self.assertEqual(sheet["C12"].value, 6)
        self.assertEqual(sheet["D12"].value, 1190624705)
        self.assertEqual(sheet["B18"].value, None)
        self.assertEqual(sheet["A18"].value, "TỔNG CỘNG")
        self.assertEqual(sheet["D18"].value, "=SUM(D11:D17)")
        self.assertEqual(sheet.column_dimensions["B"].width, 50)
        for column in ("G", "H", "I", "J", "K"):
            self.assertEqual(sheet.column_dimensions[column].width, 15)
        self.assertEqual(str(sheet.print_area), "'8'!$A$1:$K$34")
        self.assertEqual(sheet.print_title_rows, "$8:$10")

    def test_mau07a_builds_working_tool_inventory_from_ipcas_source(self) -> None:
        source = self.FIXTURE_DIR / "WT_100642.xls"
        if not source.exists():
            self.skipTest("Mẫu 07a fixture is not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.07a"],
            self.PROFILE,
            options=SettlementOptions(output_prefix="BN"),
            source_paths=(source,),
        )
        processor = Mau0708Processor()
        records = processor.read_mau07_source(source)
        sheet = processor.build_mau07_workbook(request, records)["7a"]

        self.assertEqual(len(records), 80)
        self.assertEqual(sheet["A6"].value, f"Ngày 30 tháng 6 năm {date.today().year}")
        self.assertEqual(sheet["J1"].value, "1. Mẫu số 07a/QT")
        self.assertEqual(sheet["B11"].value, "110101")
        self.assertEqual(sheet["C11"].value, "Bàn làm việc gỗ ép")
        self.assertEqual(sheet["E11"].value, 53250000)
        self.assertIsNone(sheet["L11"].value)
        self.assertEqual(sheet["B90"].value, "200305")
        self.assertEqual(sheet["C90"].value, "Phụ kiện thiết bị vệ sinh")
        self.assertEqual(sheet["A91"].value, "TỔNG CỘNG:")
        self.assertEqual(sheet["D91"].value, "=SUM(D11:D90)")
        self.assertTrue(str(sheet["D94"].value).startswith("Tồn đầu kỳ"))
        self.assertEqual(sheet.print_title_rows, "$8:$10")

    def test_mau09a_09b_09c_build_fixed_asset_movement_reports(self) -> None:
        cases = (
            ("accounting.09a", "TMBCTC_TSCD001.xls", "9a", "BÁO CÁO TÌNH HÌNH TĂNG, GIẢM TSCĐ HỮU HÌNH"),
            ("accounting.09b", "TMBCTC_TSCD002.xls", "9b", "BÁO CÁO TÌNH HÌNH TĂNG, GIẢM TSCĐ VÔ HÌNH"),
            ("accounting.09c", "TMBCTC_TSCD003.xls", "9c", "BÁO CÁO TÌNH HÌNH TĂNG, GIẢM TSCĐ THUÊ TÀI CHÍNH"),
        )
        processor = Mau09Processor()
        for spec_key, file_name, sheet_name, title in cases:
            source = self.FIXTURE_DIR / file_name
            if not source.exists():
                self.skipTest(f"{file_name} fixture is not available.")
            request = SettlementRequest(
                SETTLEMENT_SPECS[spec_key],
                self.PROFILE,
                options=SettlementOptions(output_prefix="QT"),
                source_paths=(source,),
            )
            rows = processor.read_source(source)
            sheet = processor.build_workbook(request, rows)[sheet_name]

            self.assertEqual(sheet["A5"].value, title)
            self.assertEqual(sheet["E1"].value, f"1. Mẫu số 09{sheet_name[-1]}/QT")
            self.assertEqual(sheet["G6"].value, "Đơn vị : VNĐ")
            self.assertTrue(sheet["A1"].font.bold)
            self.assertEqual(sheet.print_title_rows, "$7:$7")
            if sheet_name == "9c":
                self.assertEqual(sheet["B8"].value, "CHI NHÁNH KHÔNG PHÁT SINH")
                self.assertIn("B8:G30", {str(merged) for merged in sheet.merged_cells.ranges})
                self.assertEqual(sheet["E31"].value, f"Tân Hà, Ngày 31 tháng 12 năm {date.today().year}")
            else:
                target_cell = "B18" if sheet_name == "9a" else "B19"
                expected_formula = "=B9+B10-B14" if sheet_name == "9a" else "=B9+B10-B15"
                self.assertEqual(sheet[target_cell].value, expected_formula)
        self.assertEqual(processor.build_workbook(
            SettlementRequest(
                SETTLEMENT_SPECS["accounting.09a"],
                self.PROFILE,
                options=SettlementOptions(output_prefix="QT"),
                source_paths=(self.FIXTURE_DIR / "TMBCTC_TSCD001.xls",),
            ),
            processor.read_source(self.FIXTURE_DIR / "TMBCTC_TSCD001.xls"),
        )["9a"]["B9"].value, 9329120577)

    def test_mau22_builds_deferred_income_expense_report(self) -> None:
        sources = (
            self.FIXTURE_DIR / "5491_rt221.xls",
            self.FIXTURE_DIR / "5491_rt222.xls",
        )
        if not all(source.exists() for source in sources):
            self.skipTest("Mẫu 22 fixtures are not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.22"],
            BranchProfile(
                branch_code="5491",
                branch_name="Chi nhánh Lộc Phát-Lâm Đồng",
                report_location="Tân Hà",
                report_preparer="Thái Thị Mỹ Hằng",
                phone="02633.854.021",
            ),
            options=SettlementOptions(output_prefix="QT"),
            source_paths=sources,
        )
        processor = Mau22Processor()
        records = processor.read_sources(request)
        sheet = processor.build_workbook(request, records)["22"]

        self.assertEqual(len(records), 30)
        self.assertEqual(sheet["A6"].value, "SAO KÊ CHI TIẾT SỐ DƯ TÀI KHOẢN DOANH THU VÀ CHI PHÍ CHỜ PHÂN BỔ")
        self.assertEqual(sheet["A7"].value, f"ngày 31 tháng 12 năm {date.today().year}")
        self.assertEqual(sheet["M8"].value, "Đơn vị : VNĐ")
        self.assertEqual(sheet["M8"].font.name, "Times New Roman")
        self.assertEqual(sheet["A1"].font.name, "Times New Roman")
        self.assertEqual(sheet["L1"].font.name, "Times New Roman")
        self.assertEqual(sheet["A9"].font.name, "Times New Roman")
        self.assertEqual(sheet["A12"].value, "A")
        self.assertEqual(sheet["B12"].value, "DOANH THU CHỜ PHÂN BỔ")
        self.assertEqual(sheet["A17"].value, "I")
        self.assertEqual(sheet["B17"].value, "Lãi huy động trả trước")
        self.assertEqual(sheet["A18"].value, "5491365770796")
        self.assertEqual(sheet["D18"].value, "801007")
        self.assertEqual(sheet["K18"].value, 30001500)
        self.assertEqual(sheet["L18"].value, 431703)
        self.assertEqual(sheet["B23"].value, "Chi phí chờ phân bổ về CCDC")
        self.assertEqual(sheet["B45"].value, "Chi phí chờ phân bổ khác")
        self.assertEqual(sheet["D46"].value, "874001")
        self.assertEqual(sheet["B50"].value, "Cộng")
        self.assertEqual(sheet["M50"].value, "=SUM(M18:M49)")
        self.assertEqual(sheet.print_title_rows, "$9:$11")

    def test_mau22_report_date_uses_selected_period(self) -> None:
        processor = Mau22Processor()
        qt_sheet = processor.build_workbook(
            SettlementRequest(
                SETTLEMENT_SPECS["accounting.22"],
                self.PROFILE,
                options=SettlementOptions(output_prefix="QT"),
            ),
            [],
        )["22"]
        bn_sheet = processor.build_workbook(
            SettlementRequest(
                SETTLEMENT_SPECS["accounting.22"],
                self.PROFILE,
                options=SettlementOptions(output_prefix="BN"),
            ),
            [],
        )["22"]

        self.assertEqual(qt_sheet["A7"].value, f"ngày 31 tháng 12 năm {date.today().year}")
        self.assertEqual(bn_sheet["A7"].value, f"ngày 30 tháng 6 năm {date.today().year}")

    def test_mau23_builds_unusual_income_expense_report(self) -> None:
        sources = (
            self.FIXTURE_DIR / "790008.xls",
            self.FIXTURE_DIR / "790009.xls",
            self.FIXTURE_DIR / "899001.xls",
        )
        if not all(source.exists() for source in sources):
            self.skipTest("Mẫu 23 fixtures are not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.23"],
            BranchProfile(
                branch_code="5491",
                branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                report_location="Tân Hà",
                report_preparer="Thái Thị Mỹ Hằng",
                phone="02633.854.021",
            ),
            options=SettlementOptions(output_prefix="QT"),
            source_paths=sources,
        )
        processor = Mau23Processor()
        records, skipped = processor.read_sources(sources)
        sheet = processor.build_workbook(request, records)["23"]

        self.assertEqual(skipped, ["790008.xls", "899001.xls"])
        self.assertEqual(len(records), 8)
        self.assertEqual(sheet["A3"].value, "Mã chi nhánh: 5491")
        self.assertEqual(sheet["A4"].value, "Tên chi nhánh: Lộc Phát Lâm Đồng")
        self.assertIn("A1:B1", {str(merged) for merged in sheet.merged_cells.ranges})
        self.assertNotIn("A1:C1", {str(merged) for merged in sheet.merged_cells.ranges})
        self.assertEqual(sheet["A6"].value, "SAO KÊ CHI TIẾT TÀI KHOẢN THU NHẬP VÀ CHI PHÍ BẤT THƯỜNG")
        self.assertEqual(sheet["D8"].value, "Đơn vị : VNĐ")
        self.assertEqual(sheet["A12"].value, 790008)
        self.assertEqual(sheet["A14"].value, 790009)
        self.assertEqual(sheet["B15"].value, "Quyết toán thuế GTGT 2024")
        self.assertEqual(sheet["D15"].value, 129661)
        self.assertEqual(sheet["B21"].value, "TK không hoạt động chuyển sang tài khoản ngủ")
        self.assertEqual(sheet["C21"].value, 7)
        self.assertEqual(sheet["B22"].value, "Thu tiền lẻ cuối ngày")
        self.assertEqual(sheet["C22"].number_format, "0")
        self.assertEqual(sheet["A24"].value, 899001)
        self.assertEqual(sheet.print_title_rows, "$9:$11")
        self.assertEqual(sheet.page_setup.orientation, "portrait")
        merged_ranges = {str(merged) for merged in sheet.merged_cells.ranges}
        signature_row = next(
            row for row in range(1, sheet.max_row + 1)
            if sheet.cell(row, 1).value == "LẬP BIỂU"
        )
        self.assertNotIn(f"B{signature_row}:C{signature_row}", merged_ranges)
        self.assertNotIn(f"B{signature_row + 1}:C{signature_row + 1}", merged_ranges)
        self.assertIn(f"C{signature_row - 1}:D{signature_row - 1}", merged_ranges)
        self.assertIn(f"C{signature_row}:D{signature_row}", merged_ranges)
        self.assertIn(f"C{signature_row + 1}:D{signature_row + 1}", merged_ranges)
        note_row = next(
            row for row in range(1, sheet.max_row + 1)
            if sheet.cell(row, 1).value == "Ghi chú: "
        )
        self.assertTrue(sheet.cell(note_row, 1).font.bold)
        self.assertTrue(str(sheet.cell(note_row + 3, 1).value).endswith("không hoạt động,"))
        self.assertTrue(str(sheet.cell(note_row + 4, 1).value).startswith("tài khoản ngủ…"))
        self.assertFalse(sheet.cell(note_row + 3, 1).alignment.wrap_text)
        self.assertFalse(sheet.cell(note_row + 4, 1).alignment.wrap_text)
        self.assertEqual(sheet.column_dimensions["C"].width, 25)
        self.assertEqual(sheet.column_dimensions["D"].width, 25)

    def test_mau05_business_region_matches_vba_fixture(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt05.csv"
        expected = load_workbook(
            self.FIXTURE_DIR / "5491QT05.xlsx",
            data_only=False,
        )["05"]
        fixture_uses_branch_code = str(expected["A16"].value or "").startswith(
            self.PROFILE.branch_code
        )
        fixture_uses_customer_totals = any(
            str(expected.cell(row, 1).value or "").startswith(
                "   Cộng khách hàng:"
            )
            for row in range(12, expected.max_row + 1)
        )
        processor = Mau05Processor()
        generated = None
        records = []
        for use_owner in (False, True):
            request = SettlementRequest(
                SETTLEMENT_SPECS["credit.05"],
                self.PROFILE,
                options=SettlementOptions(
                    include_branch_in_customer_id=fixture_uses_branch_code,
                    include_customer_totals=fixture_uses_customer_totals,
                    use_collateral_owner_for_guarantee=use_owner,
                ),
                source_paths=(source,),
            )
            records, report_date = processor.read_source(request, source)
            generated_workbook = processor.build_workbook(
                request, records, report_date
            )
            buffer = BytesIO()
            generated_workbook.save(buffer)
            buffer.seek(0)
            candidate = load_workbook(buffer, data_only=False)["05"]
            if self._normalized(candidate["A1533"].value) == self._normalized(
                expected["A1533"].value
            ):
                generated = candidate
                break
        self.assertIsNotNone(generated)

        self.assertEqual(len(records), 1837)
        for row in range(12, 1861):
            for column in range(1, 19):
                self.assertEqual(
                    self._normalized(generated.cell(row, column).value),
                    self._normalized(expected.cell(row, column).value),
                    f"Mismatch at {generated.cell(row, column).coordinate}",
                )

    def test_mau06_business_region_matches_vba_fixture(self) -> None:
        csv_source = self.FIXTURE_DIR / "5491_rt05.csv"
        request05 = SettlementRequest(
            SETTLEMENT_SPECS["credit.05"],
            self.PROFILE,
            source_paths=(csv_source,),
        )
        processor05 = Mau05Processor()
        records, report_date = processor05.read_source(request05, csv_source)
        source_sheet = processor05.build_workbook(
            request05, records, report_date
        )["05"]
        request06 = SettlementRequest(
            SETTLEMENT_SPECS["credit.06"],
            self.PROFILE,
            source_paths=(self.FIXTURE_DIR / "5491QT05.xlsx",),
        )
        generated = Mau06Processor().build_workbook(
            request06, source_sheet
        )["06"]
        expected = load_workbook(
            self.FIXTURE_DIR / "5491QT06.xlsx",
            data_only=False,
        )["06"]

        for row in range(12, 25):
            for column in range(1, 12):
                self.assertEqual(
                    generated.cell(row, column).value,
                    expected.cell(row, column).value,
                    f"Mismatch at {generated.cell(row, column).coordinate}",
                )
        self.assertEqual(generated["A6"].value, "Ngày 31 tháng 12 năm 2025")

    def test_mau05_header_and_print_layout_are_ready_to_print(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt05.csv"
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.05"],
            self.PROFILE,
            source_paths=(source,),
        )
        processor = Mau05Processor()
        records, report_date = processor.read_source(request, source)
        sheet = processor.build_workbook(request, records, report_date)["05"]

        merged_ranges = {str(merged) for merged in sheet.merged_cells.ranges}
        self.assertTrue({"A1:D1", "A2:D2", "A3:D3", "A4:D4"} <= merged_ranges)
        for address in ("A1", "A2", "A3", "A4"):
            self.assertEqual(sheet[address].alignment.horizontal, "center")
            self.assertTrue(sheet[address].font.bold)
        self.assertIsNone(sheet["A2"].border.bottom.style)
        self.assertEqual(sheet["R7"].alignment.horizontal, "right")
        self.assertTrue(sheet["R7"].font.italic)
        total_row = next(
            row
            for row in range(12, sheet.max_row + 1)
            if sheet.cell(row, 1).value == "Cộng II:"
        )
        self.assertEqual(sheet.cell(total_row, 10).number_format, "0")
        signature_row = next(
            row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row, 1).value == "LẬP BIỂU"
        )
        for column in (1, 4, 7, 12):
            self.assertTrue(sheet.cell(signature_row, column).font.bold)
            self.assertTrue(sheet.cell(signature_row + 1, column).font.italic)
        self.assertTrue(sheet.cell(signature_row - 1, 12).font.italic)
        self.assertLessEqual(sheet.column_dimensions["A"].width, 18)
        self.assertGreaterEqual(sheet.column_dimensions["B"].width, 32)
        self.assertGreaterEqual(sheet.column_dimensions["E"].width, 17)
        self.assertGreaterEqual(sheet.column_dimensions["H"].width, 17)
        self.assertEqual(sheet.page_setup.paperSize, 9)
        self.assertEqual(sheet.page_setup.orientation, "landscape")
        self.assertEqual(sheet.page_setup.fitToWidth, 1)
        self.assertEqual(sheet.page_setup.fitToHeight, 0)
        self.assertTrue(sheet.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(sheet.print_title_rows, "$8:$10")
        self.assertEqual(str(sheet.print_area), f"'05'!$A$1:$R${sheet.max_row}")
        self.assertTrue(sheet.HeaderFooter.differentFirst)
        self.assertEqual(sheet.firstHeader.right.text, "")
        self.assertEqual(sheet.oddHeader.center.text, "&P/&N")

    def test_mau06_print_layout_is_ready_to_print(self) -> None:
        csv_source = self.FIXTURE_DIR / "5491_rt05.csv"
        request05 = SettlementRequest(
            SETTLEMENT_SPECS["credit.05"],
            self.PROFILE,
            source_paths=(csv_source,),
        )
        processor05 = Mau05Processor()
        records, report_date = processor05.read_source(request05, csv_source)
        source_sheet = processor05.build_workbook(
            request05, records, report_date
        )["05"]
        request06 = SettlementRequest(
            SETTLEMENT_SPECS["credit.06"],
            self.PROFILE,
            source_paths=(self.FIXTURE_DIR / "5491QT05.xlsx",),
        )
        sheet = Mau06Processor().build_workbook(request06, source_sheet)["06"]

        self.assertEqual(sheet["K7"].alignment.horizontal, "right")
        self.assertTrue(sheet["K7"].font.italic)
        for address in ("A26", "B26", "E26", "H26"):
            self.assertTrue(sheet[address].font.bold)
            self.assertEqual(sheet[address].alignment.horizontal, "center")
        for address in ("A27", "B27", "E27", "H27"):
            self.assertTrue(sheet[address].font.italic)
            self.assertEqual(sheet[address].alignment.horizontal, "center")
        self.assertEqual(sheet.page_setup.paperSize, 9)
        self.assertEqual(sheet.page_setup.orientation, "landscape")
        self.assertEqual(sheet.page_setup.fitToWidth, 1)
        self.assertEqual(sheet.page_setup.fitToHeight, 0)
        self.assertTrue(sheet.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(str(sheet.print_area), "'06'!$A$1:$K$40")
        self.assertTrue(sheet.HeaderFooter.differentFirst)
        self.assertEqual(sheet.firstHeader.right.text, "")
        self.assertEqual(sheet.oddHeader.center.text, "&P/&N")

    def test_mau1516_processors_build_expected_report_shapes(self) -> None:
        fixture_dir = Path(__file__).resolve().parents[1] / "DuLieuTEST" / "5491"
        cases = (
            ("credit.15a", "5491_rt15a_20251231.csv", "15a", 20, True),
            ("credit.15b", "5491_rt15b_20251231.csv", "15b", 19, True),
            ("credit.16", "5491_rt16_20251231.csv", "16", 17, False),
        )
        for spec_key, file_name, sheet_name, last_column, has_control in cases:
            with self.subTest(spec_key=spec_key):
                source = fixture_dir / file_name
                request = SettlementRequest(
                    SETTLEMENT_SPECS[spec_key],
                    self.PROFILE,
                    options=SettlementOptions(
                        include_branch_in_customer_id=True,
                        create_control_sheet=True,
                    ),
                    source_paths=(source,),
                )
                processor = Mau1516Processor()
                records, report_date = processor.read_source(request, source)
                workbook = processor.build_workbook(request, records, report_date)
                sheet = workbook[sheet_name]

                self.assertEqual(sheet.max_column, last_column)
                self.assertEqual(str(sheet.print_area), f"'{sheet_name}'!$A$1:${excel_column_name(last_column)}${sheet.max_row}")
                self.assertEqual(sheet.print_title_rows, "$9:$10")
                self.assertEqual(sheet.oddHeader.center.text, "&P/&N")
                self.assertEqual(sheet.cell(10, last_column).value, last_column)
                self.assertTrue(sheet["A1"].font.bold)
                if has_control:
                    self.assertIn("SoLieuTongHop", workbook.sheetnames)
                else:
                    self.assertNotIn("SoLieuTongHop", workbook.sheetnames)
                if records:
                    self.assertTrue(str(sheet["B11"].value).startswith("5491"))
                if sheet_name == "16":
                    self.assertTrue(
                        any(
                            str(sheet.cell(row, 2).value or "").startswith("Cộng KH:")
                            for row in range(11, sheet.max_row + 1)
                        )
                    )

    def test_mau1516_can_keep_unused_source_columns_off_print_area(self) -> None:
        source = (
            Path(__file__).resolve().parents[1]
            / "DuLieuTEST"
            / "5491"
            / "5491_rt15a_20251231.csv"
        )
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.15a"],
            self.PROFILE,
            options=SettlementOptions(remove_unused_columns=False),
            source_paths=(source,),
        )
        processor = Mau1516Processor()
        records, report_date = processor.read_source(request, source)
        sheet = processor.build_workbook(request, records, report_date)["15a"]

        self.assertGreater(sheet.max_column, 20)
        self.assertEqual(sheet.cell(10, 21).value, "NGAY")
        self.assertEqual(str(sheet.print_area), f"'15a'!$A$1:$T${sheet.max_row}")

    def test_mau18_matches_vba_fixture_values(self) -> None:
        source = self.FIXTURE_DIR / "5400_rt18.csv"
        expected = load_workbook(
            self.FIXTURE_DIR / "5491QT18.xlsx",
            data_only=False,
        )
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.18"],
            BranchProfile(
                branch_code="5491",
                branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                report_location="Tân Hà",
                report_preparer="Nam",
            ),
            options=SettlementOptions(
                include_branch_in_customer_id=False,
                create_control_sheet=True,
            ),
            source_paths=(source,),
        )
        processor = Mau18Processor()
        records, report_date = processor.read_source(request, source)
        generated = processor.build_workbook(request, records, report_date)

        self.assertEqual(len(records), 288)
        for sheet_name in ("18", "SoLieuTongHop"):
            with self.subTest(sheet=sheet_name):
                actual_sheet = generated[sheet_name]
                expected_sheet = expected[sheet_name]
                self.assertEqual(actual_sheet.max_row, expected_sheet.max_row)
                self.assertEqual(actual_sheet.max_column, expected_sheet.max_column)
                for row in range(1, expected_sheet.max_row + 1):
                    for column in range(1, expected_sheet.max_column + 1):
                        if sheet_name == "18" and row == 21 and column == 13:
                            self.assertEqual(
                                self._normalized(actual_sheet.cell(row, column).value),
                                719240.85,
                            )
                            self.assertEqual(
                                actual_sheet.cell(row, column).number_format,
                                "#,##0",
                            )
                            continue
                        self.assertEqual(
                            self._normalized(actual_sheet.cell(row, column).value),
                            self._normalized(expected_sheet.cell(row, column).value),
                            f"Mismatch at {sheet_name}!{actual_sheet.cell(row, column).coordinate}",
                        )

    def test_mau20a_matches_vba_xls_transform_shape(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt20a.XLS"
        if not source.exists():
            self.skipTest("Mẫu 20a fixture is not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.20a"],
            BranchProfile(
                branch_code="5491",
                branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                reporting_branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                report_location="Tân Hà",
            ),
            options=SettlementOptions(),
            source_paths=(source,),
        )
        processor = Mau20aProcessor()
        records = processor.read_source(request, source)
        sheet = processor.build_workbook(request, records, source)["20a"]

        self.assertEqual(len(records), 3)
        self.assertEqual(sheet.max_column, 16)
        self.assertEqual(sheet["A9"].value, "STT")
        self.assertEqual(sheet["A11"].value, 1)
        self.assertEqual(sheet["A12"].value, 1)
        self.assertEqual(sheet["D12"].value, "5491147744864")
        self.assertEqual(sheet["F12"].value, "5491LAV201703312")
        self.assertEqual(sheet["G12"].value, "5491LDS201705592")
        self.assertEqual(sheet["I12"].value, "")
        self.assertEqual(sheet["J12"].value, 1600000)
        self.assertEqual(processor._principal_account("Vay ngắn hạn (TK 211)"), "971101")
        self.assertEqual(processor._principal_account("Vay trung hạn (TK 212)"), "791102")
        self.assertEqual(processor._principal_account("Nguồn vốn (TK 252101)"), "791101")
        self.assertEqual(processor._principal_account("Nguồn vốn (TK 252102)"), "791102")
        self.assertEqual(processor._principal_account("Nguồn vốn (TK 252103)"), "791103")
        self.assertEqual(processor._principal_account("Nguồn vốn (TK 271101)"), "791101")
        self.assertEqual(processor._principal_account("Nguồn vốn (TK 271102)"), "791102")
        self.assertEqual(processor._principal_account("Nguồn vốn (TK 271103)"), "791103")
        self.assertEqual(sheet["A15"].value, "TỔNG CỘNG:")
        self.assertEqual(sheet["J15"].value, "=SUM(J12:J14)")
        self.assertEqual(sheet["N18"].value, "Tân Hà, Ngày 31 tháng 12 năm 2025")
        self.assertTrue(sheet["A19"].font.bold)
        self.assertTrue(sheet["A20"].font.italic)
        self.assertEqual(sheet["A20"].alignment.horizontal, "center")
        self.assertFalse(sheet["B29"].alignment.wrap_text)
        self.assertEqual(sheet["B29"].alignment.horizontal, "left")
        self.assertEqual(sheet.print_area, "'20a'!$A$1:$P$32")
        self.assertEqual(sheet.print_title_rows, "$11:$11")

    def test_mau05_control_sheet_has_table_borders(self) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        records = [
            CollateralRecord(
                sort_customer_id="1",
                customer_id="1",
                customer_name="A",
                contract_number="HD1",
                contract_date=None,
                total=100,
                real_estate=100,
                movable_property=0,
                valuable_papers=0,
                other_assets=0,
                legal_valid="X",
                legal_other="",
                dossier_complete="X",
                dossier_in_progress="",
                dossier_impossible="",
                dossier_other="",
                marketable="X",
                less_marketable="",
                not_marketable="",
                account="994001",
                collateral_type="real_estate",
            )
        ]
        Mau05Processor._write_control_sheet(workbook, records)
        sheet = workbook["SoLieuTongHop"]

        self.assertEqual(sheet["A2"].border.left.style, "thin")
        self.assertEqual(sheet["G4"].border.right.style, "thin")
        self.assertTrue(sheet["A2"].font.bold)
        self.assertTrue(sheet["A4"].font.bold)

    def test_mau30_builds_from_selected_qt_workbook(self) -> None:
        source = self.FIXTURE_DIR / "DaTest" / "5491QT15a.xlsx"
        if not source.exists():
            self.skipTest("Mẫu 30 source fixture is not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.30a"],
            BranchProfile(
                branch_code="5491",
                branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                reporting_branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                parent_branch_code="5400",
                report_location="Tân Hà",
            ),
            options=SettlementOptions(
                source_report_code="15a",
                include_accrual_accounts=True,
            ),
            source_paths=(source,),
        )
        workbook = load_workbook(source, data_only=False)
        values = load_workbook(source, data_only=True)
        processor = Mau30Processor()
        balance_source = self.FIXTURE_DIR / "CanDoiNam-30062026.XLS"
        if not balance_source.exists():
            balance_source = self.FIXTURE_DIR / "DaTest" / "CanDoiNam-30062026.XLS"
        if not balance_source.exists():
            self.skipTest("Mẫu 30 balance fixture is not available.")
        balance = processor._read_balance(balance_source)
        sheet = processor.build_sheet(request, workbook, values, "15a", balance)

        self.assertEqual(sheet.title, "Mau30QT-15a")
        self.assertEqual(sheet["A5"].value, "TỔNG HỢP SỐ LIỆU TOÀN CHI NHÁNH MẪU BIỂU QUYẾT TOÁN SỐ 15a")
        self.assertEqual(sheet["B12"].value, "5400")
        self.assertEqual(sheet["C12"].value, "5491")
        self.assertEqual(sheet["D12"].value, "211101")
        self.assertEqual(sheet["E12"].value, 1959747671000)
        self.assertEqual(sheet["F12"].value, 2169363622401)
        self.assertEqual(sheet["G12"].value, "=F12-E12")
        self.assertTrue(sheet["A1"].font.bold)
        self.assertEqual(sheet["A1"].alignment.horizontal, "center")
        self.assertIn("A1:D1", {str(merged) for merged in sheet.merged_cells.ranges})
        self.assertTrue(sheet["A5"].font.bold)
        self.assertEqual(sheet["A5"].alignment.horizontal, "center")
        self.assertEqual(sheet["G27"].value, "Tân Hà, Ngày 31 tháng 12 năm 2025")
        self.assertEqual(sheet["G27"].alignment.horizontal, "center")
        self.assertTrue(sheet["G27"].alignment.shrink_to_fit)
        self.assertTrue(sheet["G27"].font.italic)
        self.assertEqual(sheet.row_dimensions[6].height, 40)
        self.assertEqual(sheet.row_dimensions[27].height, 21)
        self.assertEqual(sheet.print_title_rows, "$11:$11")
        self.assertTrue(str(sheet.print_area).startswith("'Mau30QT-15a'!$A$1:$H$"))

    def test_mau30_builds_20a_without_control_sheet(self) -> None:
        workbook = Workbook()
        source = workbook.active
        source.title = "20a"
        source["A6"] = "BÁO CÁO NỢ ĐƯỢC XỬ LÝ BẰNG NGUỒN DỰ PHÒNG"
        source["A7"] = "ĐẾN NGÀY 31/12/2025"
        source["I12"] = "971101"
        source["J12"] = 1_000_000
        source["K12"] = "971201"
        source["L12"] = 200_000
        source["I13"] = "971101"
        source["J13"] = 500_000
        source["K13"] = "971202"
        source["L13"] = 300_000
        source["A14"] = "TỔNG CỘNG:"
        source["J14"] = 1_500_000
        source["L14"] = 500_000
        source["I17"] = "TRƯỞNG PHÒNG KẾ TOÁN"
        source["K18"] = "(Ký, ghi rõ họ tên)"
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.30a"],
            BranchProfile(
                branch_code="5491",
                branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                reporting_branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                parent_branch_code="5400",
                report_location="Tân Hà",
            ),
            options=SettlementOptions(source_report_code="20a"),
            source_paths=(Path("5491QT20a.xlsx"),),
        )

        sheet = Mau30Processor().build_sheet(
            request,
            workbook,
            workbook,
            "20a",
        )

        self.assertNotIn("SoLieuTongHop", workbook.sheetnames)
        self.assertEqual(sheet["D12"].value, "971101")
        self.assertEqual(sheet["F12"].value, 1_500_000)
        self.assertEqual(sheet["D13"].value, "971201")
        self.assertEqual(sheet["F13"].value, 200_000)
        self.assertEqual(sheet["D14"].value, "971202")
        self.assertEqual(sheet["F14"].value, 300_000)

    def test_mau30_accounting_builds_from_mau23_without_control_sheet(self) -> None:
        workbook = Workbook()
        source = workbook.active
        source.title = "23"
        source["A6"] = "SAO KÊ CHI TIẾT TÀI KHOẢN THU NHẬP VÀ CHI PHÍ BẤT THƯỜNG"
        source["A7"] = "Ngày 31 tháng 12 năm 2025"
        source["A13"] = "Cộng TK 790008"
        source["D13"] = 0
        source["A23"] = "Cộng TK 790009"
        source["D23"] = 6_742_929
        source["A25"] = "Cộng TK 899001"
        source["D25"] = 0
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.30a"],
            BranchProfile(
                branch_code="5491",
                branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                reporting_branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                parent_branch_code="5400",
                report_location="Tân Hà",
            ),
            options=SettlementOptions(source_report_code="23"),
            source_paths=(Path("5491QT23.xlsx"),),
        )

        sheet = Mau30Processor().build_sheet(request, workbook, workbook, "23")

        self.assertEqual(sheet.title, "Mau30QT-23")
        self.assertEqual(sheet["A7"].value, "Ngày 31 tháng 12 năm 2025")
        self.assertEqual(sheet["D12"].value, "790008")
        self.assertEqual(sheet["F12"].value, 0)
        self.assertEqual(sheet["D13"].value, "790009")
        self.assertEqual(sheet["F13"].value, 6_742_929)
        self.assertEqual(sheet["D14"].value, "899001")
        self.assertEqual(sheet["F14"].value, 0)

    def test_mau30_accounting_normalizes_mau22_report_date(self) -> None:
        workbook = Workbook()
        source = workbook.active
        source.title = "22"
        source["A6"] = "SAO KÊ CHI TIẾT SỐ DƯ TÀI KHOẢN DOANH THU VÀ CHI PHÍ CHỜ PHÂN BỔ"
        source["A7"] = "ĐẾN NGÀY 31/12/2025"
        source["C12"] = "388001"
        source["M12"] = 1_250_000
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.30a"],
            BranchProfile(
                branch_code="5491",
                branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                reporting_branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                parent_branch_code="5400",
                report_location="Tân Hà",
            ),
            options=SettlementOptions(source_report_code="22"),
            source_paths=(Path("5491QT22.xlsx"),),
        )

        sheet = Mau30Processor().build_sheet(request, workbook, workbook, "22")

        self.assertEqual(sheet.title, "Mau30QT-22")
        self.assertEqual(sheet["A7"].value, "ngày 31 tháng 12 năm 2025")
        self.assertEqual(sheet["D12"].value, "388001")
        self.assertEqual(sheet["F12"].value, 1_250_000)

    def test_mau13_matches_vba_source_and_control_totals(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt13.csv"
        expected_path = self.FIXTURE_DIR / "5491QT13.xlsx"
        if not source.exists() or not expected_path.exists():
            self.skipTest("Mẫu 13 fixtures are not available.")
        options = SettlementOptions(
            include_branch_in_customer_id=True,
            create_control_sheet=True,
            include_accrual_accounts=True,
        )
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.13"],
            BranchProfile(branch_code="5491"),
            options=options,
            source_paths=(source,),
        )
        processor = Mau1314Processor()
        records, report_date, currencies = processor.read_source(request, source)
        workbook = Workbook()
        processor._write_control_sheet(workbook, options, records)
        actual = workbook["SoLieuTongHop"]
        expected = load_workbook(
            expected_path,
            data_only=True,
            read_only=True,
        )["SoLieuTongHop"]

        self.assertEqual(len(records), 68_093)
        self.assertEqual(report_date, "Ngày 31 tháng 12 năm 2025")
        self.assertEqual(currencies, ("EUR", "USD", "VND"))
        self.assertTrue(
            all(record.ledger_account[:2] not in {"40", "41"} for record in records)
        )
        for address in (
            "B3",
            "C3",
            "B4",
            "C4",
            "D29",
            "E30",
            "E31",
            "D32",
            "E32",
        ):
            self.assertAlmostEqual(
                float(actual[address].value or 0),
                float(expected[address].value or 0),
                places=2,
            )

    def test_mau13_wraps_long_customer_name_and_widens_savings_book(self) -> None:
        long_savings_book = "549120250123456789"
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.13"],
            BranchProfile(branch_code="5491"),
            options=SettlementOptions(),
        )
        workbook = Mau1314Processor().build_workbook(
            request,
            [
                DepositRecord(
                    ledger_account="421101",
                    customer_id="034728393",
                    customer_name="Ban QLDA DT XD Khu Vực Lâm Đồng Có Tên Rất Dài",
                    deposit_account="5491201007177",
                    savings_book=long_savings_book,
                    currency="VND",
                    original_balance=Decimal(10),
                    converted_balance=Decimal(10),
                    term_months="12",
                    deposit_date=date(2025, 1, 1),
                    maturity_date=date(2025, 12, 31),
                    interest_rate=5,
                    last_interest_date=date(2025, 12, 31),
                    prepaid_interest=Decimal(0),
                    accrued_interest=Decimal(0),
                    interest_account="491101",
                    prepaid_interest_display=0,
                    accrued_interest_display=0,
                    extra_values=(),
                )
            ],
            "Ngày 31 tháng 12 năm 2025",
            ("VND",),
        )
        sheet = workbook["13"]

        self.assertTrue(sheet["C11"].alignment.wrap_text)
        self.assertGreater(sheet.row_dimensions[11].height, 15)
        self.assertFalse(bool(sheet["E11"].alignment.wrap_text))
        self.assertGreaterEqual(
            sheet.column_dimensions["E"].width,
            len(long_savings_book) + 2,
        )

    def test_mau14_groups_and_totals_by_customer(self) -> None:
        def record(
            customer_id: str,
            account: str,
            amount: int,
            prepaid: int = 0,
            accrued: int = 0,
        ) -> DepositRecord:
            return DepositRecord(
                ledger_account=account,
                customer_id=customer_id,
                customer_name=f"Khách hàng {customer_id}",
                deposit_account=f"TK{customer_id}",
                savings_book="",
                currency="VND",
                original_balance=Decimal(amount),
                converted_balance=Decimal(amount),
                term_months="12",
                deposit_date=date(2025, 1, 1),
                maturity_date=date(2025, 12, 31),
                interest_rate=5,
                last_interest_date=date(2025, 12, 31),
                prepaid_interest=Decimal(prepaid),
                accrued_interest=Decimal(accrued),
                interest_account="",
                prepaid_interest_display=None,
                accrued_interest_display=None,
                extra_values=("", "", "", "", "", "", ""),
            )

        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.14"],
            BranchProfile(
                branch_code="5491",
                reporting_branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                report_location="Tân Hà",
            ),
            options=SettlementOptions(remove_unused_columns=True),
        )
        workbook = Mau1314Processor().build_workbook(
            request,
            [
                record("001", "423101", 10, 1, 3),
                record("001", "423201", 20, 2, 4),
                record("002", "423101", 30, 5, 6),
            ],
            "Ngày 31 tháng 12 năm 2025",
            ("VND",),
        )
        sheet = workbook["14"]

        self.assertEqual(sheet["A13"].value, "   Cộng khách hàng: 001")
        self.assertEqual(sheet["H13"].value, "=SUM(H11:H12)")
        self.assertEqual(sheet["N13"].value, "=SUM(N11:N12)")
        self.assertEqual(sheet["O13"].value, "=SUM(O11:O12)")
        self.assertEqual(sheet["A15"].value, "   Cộng khách hàng: 002")
        self.assertEqual(sheet["H15"].value, "=SUM(H14:H14)")
        self.assertEqual(sheet["N15"].value, "=SUM(N14:N14)")
        self.assertEqual(sheet["O15"].value, "=SUM(O14:O14)")
        self.assertEqual(sheet["K16"].value, "Tân Hà, Ngày 31 tháng 12 năm 2025")
        self.assertNotIn("SoLieuTongHop", workbook.sheetnames)
        self.assertEqual(sheet.print_title_rows, "$9:$10")
        self.assertEqual(sheet["A1"].value, "NGÂN HÀNG NÔNG NGHIỆP")
        self.assertTrue(sheet["A1"].font.bold)
        self.assertEqual(sheet["A1"].alignment.horizontal, "center")
        self.assertTrue(sheet["A4"].font.bold)
        self.assertEqual(sheet["A4"].alignment.horizontal, "center")
        self.assertFalse(sheet.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(sheet.page_setup.fitToWidth, 0)
        self.assertEqual(sheet.max_column, 15)
        self.assertEqual(sheet.column_dimensions["C"].width, 35)
        self.assertTrue(sheet["C11"].alignment.wrap_text)
        self.assertEqual(sheet.page_setup.scale, 68)
        self.assertEqual(str(sheet.print_area), f"'14'!$A$1:$O${sheet.max_row}")

    def test_mau14_keeps_print_area_through_extra_columns(self) -> None:
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.14"],
            BranchProfile(branch_code="5491"),
            options=SettlementOptions(remove_unused_columns=False),
        )
        workbook = Mau1314Processor().build_workbook(
            request,
            [
                DepositRecord(
                    ledger_account="421101",
                    customer_id="034728393",
                    customer_name="Khách hàng 001",
                    deposit_account="5491201007177",
                    savings_book="",
                    currency="VND",
                    original_balance=Decimal(10),
                    converted_balance=Decimal(10),
                    term_months="12",
                    deposit_date=date(2025, 1, 1),
                    maturity_date=date(2025, 12, 31),
                    interest_rate=5,
                    last_interest_date=date(2025, 12, 31),
                    prepaid_interest=Decimal(0),
                    accrued_interest=Decimal(0),
                    interest_account="491101",
                    prepaid_interest_display=0,
                    accrued_interest_display=0,
                    extra_values=("A", "B", "C", "D", "E", "F", "G"),
                )
            ],
            "Ngày 31 tháng 12 năm 2025",
            ("VND",),
        )
        sheet = workbook["14"]

        self.assertEqual(sheet.max_column, 23)
        self.assertEqual(str(sheet.print_area), f"'14'!$A$1:$W${sheet.max_row}")

    def test_mau14_accepts_short_source_header_names(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt14.csv"
        if not source.exists():
            self.skipTest("Mẫu 14 source fixture is not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.14"],
            BranchProfile(
                branch_code="5491",
                reporting_branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                report_location="Tân Hà",
            ),
            options=SettlementOptions(
                include_branch_in_customer_id=True,
                four_digit_year=True,
                remove_unused_columns=True,
            ),
            source_paths=(source,),
        )
        processor = Mau1314Processor()
        records, report_date, currencies = processor.read_source(request, source)
        workbook = processor.build_workbook(
            request,
            records,
            report_date,
            currencies,
        )
        sheet = workbook["14"]

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].original_balance, Decimal("2170138864"))
        self.assertEqual(records[0].interest_rate, 0.2)
        self.assertEqual(sheet["B11"].value, "5491034728393")
        self.assertEqual(sheet["H13"].value, "=SUM(H11:H12)")
        self.assertEqual(sheet["N13"].value, "=SUM(N11:N12)")
        self.assertEqual(sheet["O13"].value, "=SUM(O11:O12)")
        self.assertEqual(sheet["K14"].value, "Tân Hà, Ngày 31 tháng 12 năm 2025")
        self.assertEqual(sheet.max_column, 15)

    def test_mau24_groups_summary_source_by_account_level_5(self) -> None:
        source = self.FIXTURE_DIR / "5400_rt24.csv"
        expected_path = self.FIXTURE_DIR / "5400QT24a.xlsx"
        if not source.exists() or not expected_path.exists():
            self.skipTest("Mẫu 24 summary fixtures are not available.")
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.24"],
            BranchProfile(
                branch_code="5400",
                reporting_branch_name="Chi nhánh Hội sở",
                report_location="Tân Hà",
                report_preparer="Nam",
            ),
            options=SettlementOptions(create_control_sheet=True),
            source_paths=(source,),
        )
        processor = Mau24Processor()
        records, report_date = processor.read_source(request, source)
        workbook = processor.build_workbook(request, records, report_date)
        sheet = workbook["24"]

        self.assertEqual(len(records), 15)
        self.assertEqual(report_date, "Ngày 31 tháng 12 năm 2025")
        self.assertEqual(sheet["G1"].value, "1. Mẫu số 24a/QT")
        self.assertNotEqual(sheet["A11"].value, "TK 35: Các khoản phải thu bên ngoài")
        self.assertEqual(sheet["A11"].value, 353901)
        self.assertEqual(sheet["A12"].value, "Cộng TK 353901")
        self.assertEqual(sheet["H12"].value, "=SUM(H11:H11)")
        self.assertEqual(sheet["A14"].value, "Cộng TK 355001")
        self.assertEqual(sheet["A17"].value, "Cộng TK 359209")
        self.assertEqual(sheet["H17"].value, "=SUM(H15:H16)")
        self.assertEqual(sheet["A35"].value, "Cộng TK 459901")
        self.assertEqual(sheet["H35"].value, "=SUM(H32:H34)")
        self.assertEqual(sheet["A37"].value, "Cộng TK 461001")
        self.assertEqual(sheet["A48"].value, "Ghi chú:")
        self.assertEqual(sheet["E38"].value, "Tân Hà, Ngày 31 tháng 12 năm 2025")
        self.assertEqual(sheet.column_dimensions["B"].width, 34)
        self.assertEqual(sheet.column_dimensions["C"].width, 36)
        self.assertTrue(sheet["B11"].alignment.wrap_text)
        self.assertTrue(sheet["C11"].alignment.wrap_text)
        self.assertEqual(sheet.print_title_rows, "$9:$9")
        self.assertEqual(str(sheet.print_area), "'24'!$A$1:$H$56")
        self.assertEqual(sheet.page_setup.scale, 98)
        self.assertIn("SoLieuTongHop", workbook.sheetnames)
        control = workbook["SoLieuTongHop"]
        self.assertEqual(control["A3"].value, "353901_VND")
        self.assertEqual(control["C3"].value, 5430655696)
        self.assertEqual(control["A5"].value, "359209_USD")
        self.assertEqual(control["C5"].value, 139003100)
        self.assertEqual(control["A6"].value, "359209_VND")
        self.assertEqual(control["C6"].value, 500000)
        self.assertEqual(control["A16"].value, "Tổng cộng:")
        self.assertEqual(control["C16"].value, "=SUM(C3:C15)")

    def test_mau24_output_name_uses_24a_and_selected_period_prefix(self) -> None:
        source = self.FIXTURE_DIR / "5400_rt24.csv"
        if not source.exists():
            self.skipTest("Mẫu 24 summary source fixture is not available.")
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        test_source = test_dir / "5400_rt24.csv"
        test_source.write_bytes(source.read_bytes())
        request = SettlementRequest(
            SETTLEMENT_SPECS["accounting.24"],
            BranchProfile(branch_code="5400"),
            options=SettlementOptions(output_prefix="BN"),
            source_paths=(test_source,),
        )
        result = Mau24Processor().execute(request)

        self.assertEqual(result.output_path.name, "5400BN24a.xlsx")
        if result.output_path.exists():
            result.output_path.unlink()

    def test_mau05_can_keep_guarantee_borrower_customer(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt05.csv"
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.05"],
            self.PROFILE,
            options=SettlementOptions(
                use_collateral_owner_for_guarantee=False,
            ),
            source_paths=(source,),
        )
        records, _ = Mau05Processor().read_source(request, source)

        self.assertTrue(
            any(
                record.collateral_type.casefold() == "bao lanh"
                and record.customer_name == "Cty TNHH Phú Cường"
                for record in records
            )
        )

    def test_mau05_can_insert_customer_total_rows(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt05.csv"
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.05"],
            self.PROFILE,
            options=SettlementOptions(include_customer_totals=True),
            source_paths=(source,),
        )
        processor = Mau05Processor()
        records, report_date = processor.read_source(request, source)
        sheet = processor.build_workbook(request, records, report_date)["05"]

        self.assertTrue(
            any(
                str(sheet.cell(row, 1).value or "").startswith(
                    "   Cộng khách hàng:"
                )
                for row in range(12, sheet.max_row + 1)
            )
        )

    def test_mau05_can_bold_customer_total_rows(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt05.csv"
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.05"],
            self.PROFILE,
            options=SettlementOptions(
                include_customer_totals=True,
                bold_customer_rows=True,
            ),
            source_paths=(source,),
        )
        processor = Mau05Processor()
        records, report_date = processor.read_source(request, source)
        sheet = processor.build_workbook(request, records, report_date)["05"]
        total_row = next(
            row
            for row in range(12, sheet.max_row + 1)
            if str(sheet.cell(row, 1).value or "").startswith(
                "   Cộng khách hàng:"
            )
        )

        self.assertTrue(sheet.cell(total_row, 1).font.bold)
        self.assertTrue(sheet.cell(total_row, 5).font.bold)

    def test_mau05_can_keep_unused_source_columns_off_print_area(self) -> None:
        source = self.FIXTURE_DIR / "5491_rt05.csv"
        request = SettlementRequest(
            SETTLEMENT_SPECS["credit.05"],
            self.PROFILE,
            options=SettlementOptions(
                include_customer_totals=True,
                remove_unused_columns=False,
            ),
            source_paths=(source,),
        )
        processor = Mau05Processor()
        records, report_date = processor.read_source(request, source)
        sheet = processor.build_workbook(request, records, report_date)["05"]

        self.assertGreater(sheet.max_column, 18)
        self.assertEqual(sheet.cell(11, 19).value, "NGAY")
        self.assertEqual(str(sheet.print_area), f"'05'!$A$1:$R${sheet.max_row}")

    def test_merge_same_structure_csv_to_xlsx_appends_without_repeating_header(self) -> None:
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        source_a = test_dir / "merge_a.csv"
        source_b = test_dir / "merge_b.csv"
        output = test_dir / "merge_output.xlsx"
        source_a.write_text("NGAY,MA_CN,MA_KH\n20251231,5491,001\n", encoding="utf-8")
        source_b.write_text("NGAY,MA_CN,MA_KH\n20251231,5492,002\n", encoding="utf-8")
        try:
            result = merge_same_structure_csv_to_xlsx((source_a, source_b), output)
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["Data"]

            self.assertEqual(result.source_count, 2)
            self.assertEqual(result.row_count, 2)
            self.assertEqual(sheet.max_row, 3)
            self.assertEqual(sheet["A1"].value, "NGAY")
            self.assertEqual(sheet["B2"].value, "5491")
            self.assertEqual(sheet["B3"].value, "5492")
        finally:
            for path in (source_a, source_b, output):
                if path.exists():
                    path.unlink()

    def test_merge_same_structure_csv_to_xlsx_rejects_different_headers(self) -> None:
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        source_a = test_dir / "merge_header_a.csv"
        source_b = test_dir / "merge_header_b.csv"
        output = test_dir / "merge_header_output.xlsx"
        source_a.write_text("NGAY,MA_CN\n20251231,5491\n", encoding="utf-8")
        source_b.write_text("NGAY,KHAC\n20251231,5491\n", encoding="utf-8")
        try:
            with self.assertRaises(FileMergeError):
                merge_same_structure_csv_to_xlsx((source_a, source_b), output)
        finally:
            for path in (source_a, source_b, output):
                if path.exists():
                    path.unlink()

    def test_merge_same_structure_csv_to_xlsx_can_add_source_filename_column(self) -> None:
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        source_a = test_dir / "merge_note_a.csv"
        source_b = test_dir / "merge_note_b.csv"
        output = test_dir / "merge_note_output.xlsx"
        source_a.write_text("NGAY,MA_CN\n20251231,5491\n", encoding="utf-8")
        source_b.write_text("NGAY,MA_CN\n20251231,5492\n", encoding="utf-8")
        try:
            result = merge_same_structure_csv_to_xlsx(
                (source_a, source_b),
                output,
                include_source_filename=True,
            )
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["Data"]

            self.assertEqual(result.column_count, 3)
            self.assertEqual(sheet["C1"].value, "File gốc")
            self.assertEqual(sheet["C2"].value, source_a.name)
            self.assertEqual(sheet["C3"].value, source_b.name)
            workbook.close()
        finally:
            for path in (source_a, source_b, output):
                if path.exists():
                    path.unlink()

    def test_merge_same_structure_excel_to_xlsx_appends_without_repeating_header(self) -> None:
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        source_a = test_dir / "merge_excel_a.xlsx"
        source_b = test_dir / "merge_excel_b.xlsx"
        output = test_dir / "merge_excel_output.xlsx"
        for path, branch_code, customer_id in (
            (source_a, "5491", "001"),
            (source_b, "5492", "002"),
        ):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["NGAY", "MA_CN", "MA_KH"])
            sheet.append(["20251231", branch_code, customer_id])
            workbook.save(path)
            workbook.close()
        try:
            result = merge_same_structure_excel_to_xlsx((source_a, source_b), output)
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["Data"]

            self.assertEqual(result.source_count, 2)
            self.assertEqual(result.row_count, 2)
            self.assertEqual(sheet.max_row, 3)
            self.assertEqual(sheet["A1"].value, "NGAY")
            self.assertEqual(sheet["B2"].value, "5491")
            self.assertEqual(sheet["B3"].value, "5492")
            workbook.close()
        finally:
            for path in (source_a, source_b, output):
                if path.exists():
                    path.unlink()

    def test_merge_same_structure_excel_to_xlsx_rejects_different_headers(self) -> None:
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        source_a = test_dir / "merge_excel_header_a.xlsx"
        source_b = test_dir / "merge_excel_header_b.xlsx"
        output = test_dir / "merge_excel_header_output.xlsx"
        for path, headers in (
            (source_a, ["NGAY", "MA_CN"]),
            (source_b, ["NGAY", "KHAC"]),
        ):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
            sheet.append(["20251231", "5491"])
            workbook.save(path)
            workbook.close()
        try:
            with self.assertRaises(FileMergeError):
                merge_same_structure_excel_to_xlsx((source_a, source_b), output)
        finally:
            for path in (source_a, source_b, output):
                if path.exists():
                    path.unlink()

    def test_merge_same_structure_excel_to_xlsx_can_add_source_filename_column(self) -> None:
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        source_a = test_dir / "merge_excel_note_a.xlsx"
        source_b = test_dir / "merge_excel_note_b.xlsx"
        output = test_dir / "merge_excel_note_output.xlsx"
        for path, branch_code in ((source_a, "5491"), (source_b, "5492")):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["NGAY", "MA_CN"])
            sheet.append(["20251231", branch_code])
            workbook.save(path)
            workbook.close()
        try:
            result = merge_same_structure_excel_to_xlsx(
                (source_a, source_b),
                output,
                include_source_filename=True,
            )
            workbook = load_workbook(output, data_only=True)
            sheet = workbook["Data"]

            self.assertEqual(result.column_count, 3)
            self.assertEqual(sheet["C1"].value, "File gốc")
            self.assertEqual(sheet["C2"].value, source_a.name)
            self.assertEqual(sheet["C3"].value, source_b.name)
            workbook.close()
        finally:
            for path in (source_a, source_b, output):
                if path.exists():
                    path.unlink()

    def test_summary05_builds_consolidation_sheet_from_merged_rt05(self) -> None:
        source_dir = self.FIXTURE_DIR / "TESTTONGHOP"
        sources = (
            source_dir / "5400_rt05_20251231.csv",
            source_dir / "5491_rt05_20251231.csv",
        )
        if not all(source.exists() for source in sources):
            self.skipTest("Tổng hợp mẫu 05 fixtures are not available.")
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        merged_csv = test_dir / "summary05_merged.csv"
        output = test_dir / "summary05_merged.xlsx"
        try:
            merge_same_structure_csv_to_csv(sources, merged_csv)
            request = SettlementRequest(
                SETTLEMENT_SPECS["consolidation.05"],
                self.PROFILE,
                options=SettlementOptions(
                    create_control_sheet=False,
                    include_branch_in_customer_id=True,
                ),
                source_paths=(merged_csv,),
            )
            processed_rows = Summary05Processor().execute(request, merged_csv, output)
            workbook = load_workbook(output, data_only=True)
            self.assertEqual(workbook.sheetnames, ["SoLieu_Mau05", "TongHop_Mau05"])
            detail_sheet = workbook["SoLieu_Mau05"]
            customer_ids = {
                str(detail_sheet.cell(row, 1).value)
                for row in range(11, detail_sheet.max_row + 1)
                if detail_sheet.cell(row, 1).value
            }
            sheet = workbook["TongHop_Mau05"]
            branches = {
                str(sheet.cell(row, 1).value)
                for row in range(3, sheet.max_row)
            }
            headers = {
                str(sheet.cell(2, column).value)
                for column in range(2, sheet.max_column)
            }

            self.assertGreater(processed_rows, 0)
            self.assertTrue(any(value.startswith("5400") for value in customer_ids))
            self.assertTrue(any(value.startswith("5491") for value in customer_ids))
            self.assertIn("5400", branches)
            self.assertIn("5491", branches)
            self.assertIn("Cộng chi nhánh", str(sheet.cell(2, sheet.max_column).value))
            self.assertTrue(any(header.startswith("994") for header in headers))
        finally:
            for path in (merged_csv, output):
                if path.exists():
                    path.unlink()

    def test_consolidation06_uses_summary05_detail_sheet(self) -> None:
        source_dir = self.FIXTURE_DIR / "TESTTONGHOP"
        sources = (
            source_dir / "5400_rt05_20251231.csv",
            source_dir / "5491_rt05_20251231.csv",
        )
        if not all(source.exists() for source in sources):
            self.skipTest("Tổng hợp mẫu 05 fixtures are not available.")
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        merged_csv = test_dir / "summary06_source_merged.csv"
        summary05_output = test_dir / "summary06_source_05.xlsx"
        summary06_output = test_dir / "5491QT06.xlsx"
        try:
            merge_same_structure_csv_to_csv(sources, merged_csv)
            Summary05Processor().execute(
                SettlementRequest(
                    SETTLEMENT_SPECS["consolidation.05"],
                    self.PROFILE,
                    options=SettlementOptions(
                        create_control_sheet=False,
                        include_branch_in_customer_id=True,
                    ),
                    source_paths=(merged_csv,),
                ),
                merged_csv,
                summary05_output,
            )

            result = Mau06Processor().execute(
                SettlementRequest(
                    SETTLEMENT_SPECS["consolidation.06"],
                    self.PROFILE,
                    options=SettlementOptions(output_prefix="QT"),
                    source_paths=(summary05_output,),
                )
            )
            workbook = load_workbook(result.output_path, data_only=False)

            self.assertEqual(result.output_path, summary06_output)
            self.assertIn("06", workbook.sheetnames)
            self.assertEqual(workbook["06"]["A12"].value, "I. TSTC của khách hàng trực tiếp vay vốn")
            self.assertEqual(workbook["06"]["A18"].value, "II. TSTC của đơn vị bảo lãnh cho bên thứ ba vay vốn")
        finally:
            for path in (merged_csv, summary05_output, summary06_output):
                if path.exists():
                    path.unlink()

    def test_consolidation13_uses_source_branch_in_customer_id(self) -> None:
        root = self.FIXTURE_DIR.parent
        sources = (
            root / "5401" / "5401_rt13_20251231.csv",
            root / "5404" / "5404_rt13_20251231.csv",
        )
        if not all(source.exists() for source in sources):
            self.skipTest("Tổng hợp mẫu 13 fixtures are not available.")
        test_dir = self.FIXTURE_DIR / "DaTest"
        test_dir.mkdir(exist_ok=True)
        merged_csv = test_dir / "summary13_merged.csv"
        output = test_dir / "summary13_merged.xlsx"
        try:
            merge_same_structure_csv_to_csv(sources, merged_csv)
            request = SettlementRequest(
                SETTLEMENT_SPECS["consolidation.13"],
                self.PROFILE,
                options=SettlementOptions(
                    include_branch_in_customer_id=True,
                    create_control_sheet=True,
                ),
                source_paths=(merged_csv,),
            )
            processor = Mau1314Processor()
            records, report_date, currency_order = processor.read_source(
                request,
                merged_csv,
            )
            processor.save_mau13_streaming_workbook(
                request,
                records,
                report_date,
                currency_order,
                output,
            )
            output_workbook = load_workbook(output, data_only=False)
            sheet = output_workbook["SoLieu_Mau13"]
            customer_ids = {
                str(sheet.cell(row, 2).value)
                for row in range(11, min(sheet.max_row, 2000) + 1)
                if sheet.cell(row, 2).value
            }

            self.assertTrue(any(value.startswith("5401") for value in customer_ids))
            self.assertTrue(any(value.startswith("5404") for value in customer_ids))
            self.assertIn("TongHop_Mau13", output_workbook.sheetnames)
        finally:
            for path in (merged_csv, output):
                if path.exists():
                    path.unlink()


if __name__ == "__main__":
    unittest.main()
