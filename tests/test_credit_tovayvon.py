from __future__ import annotations

from pathlib import Path
from contextlib import closing
from datetime import date, datetime
import os
import re
import sqlite3
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtCore import QEvent, QUrl
from PySide6.QtGui import QDesktopServices
from PySide6.QtWidgets import QApplication, QComboBox, QFrame, QLabel, QLineEdit, QMessageBox

from agribank_v3.settings import AppSettingsDatabase, BranchProfile
from agribank_v3.features.credit.tovayvon.models import (
    COMMISSION_EXPORT_HEADERS,
    COMMISSION_RULE_EXPORT_HEADERS,
    DATA_TVV_FIELD_LABELS,
    DATA_TVV_HEADERS,
    DATA_TVV_TEMPLATE_HEADERS,
    CreditGroup,
    CreditGroupCommissionRate,
    CreditGroupCommissionRule,
    CreditCommissionRuleSettings,
)
from agribank_v3.features.credit.tovayvon.excel_templates import (
    DATA_TVV_TEXT_COLUMNS,
    create_data_tvv_template,
)
from agribank_v3.features.credit.tovayvon.checkable_combo_box import CheckableComboBox
from agribank_v3.features.credit.tovayvon.interest_report import (
    InterestReportError,
    InterestReportRequest,
    InterestGroupSummary,
    InterestRow,
    VBA_NO_SECURED_COMMISSION_BASE_RATE,
    VBA_SECURED_COMMISSION_BASE_RATE,
    calculate_commission_pay_rate,
    create_interest_report,
    detect_skck_columns,
    detect_sktl_columns,
    validate_interest_report_columns,
    _commission_breakdown,
)
from agribank_v3.features.credit.tovayvon.debt_reconciliation import (
    DETAIL_SHEET_NAME,
    GROUP_WITHOUT_BALANCE_SHEET_NAME,
    MISSING_GROUP_SHEET_NAME,
    SUMMARY_SHEET_NAME,
    UNKNOWN_GROUP_SHEET_NAME,
    WARNING_SHEET_NAME,
    DebtReconciliationError,
    DebtReconciliationRequest,
    clean_debt_reconciliation_sheet_name,
    create_debt_reconciliation,
    detect_debt_columns,
    normalize_debt_group,
)
from agribank_v3.features.credit.tovayvon.debt_reconciliation_window import (
    DebtReconciliationWindow,
)
from agribank_v3.features.credit.tovayvon.payment_request import (
    PaymentSummaryRow,
    PaymentRequestError,
    analyze_payment_rows,
    build_payment_context,
    amount_to_vietnamese_words,
    default_payment_template_path,
    export_payment_requests,
    load_payment_report_data,
    payment_ineligible_reason,
)
from agribank_v3.features.credit.tovayvon.repository import (
    CreditGroupRepository,
    CreditGroupRepositoryError,
)
from agribank_v3.features.credit.tovayvon.word_template import (
    extract_docx_text,
    replace_word_placeholders,
    scan_word_placeholders,
)
from agribank_v3.features.credit.tovayvon.placeholder_windows import (
    CREDIT_TOVAYVON_PLACEHOLDER_TITLES,
    DEFAULT_COMMISSION_FIELD_LABELS,
    CreditGroupDefaultInfoDialog,
    CreditGroupEditDialog,
    CreditGroupManagementPlaceholderDialog,
    FIELD_PLACEHOLDERS,
    get_default_credit_group_info,
    get_suggested_credit_group_default_info,
    load_default_credit_group_info,
    normalize_uy_quyen,
    save_default_credit_group_info,
)
from agribank_v3.features.credit.tovayvon.interest_report_window import InterestReportWindow
from agribank_v3.features.credit.tovayvon.payment_request_window import PaymentRequestWindow
from agribank_v3.features.credit.tovayvon.help_window import (
    TOVAYVON_HELP_TITLE,
    ToVayVonHelpWindow,
)


class CreditGroupRepositoryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = TemporaryDirectory()
        self.database_path = Path(self.temporary_directory.name) / "DuLieuV3.db"
        self.repository = CreditGroupRepository(self.database_path)

    def tearDown(self) -> None:
        self.temporary_directory.cleanup()

    def test_new_group_gets_vba_default_commission_rates(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )

        rate = self.repository.get_or_create_commission_rate("T001")

        self.assertEqual(rate.no_secured_to_truong, 80.0)
        self.assertEqual(rate.no_secured_cap_xa, 13.0)
        self.assertEqual(rate.no_secured_cap_huyen, 3.8)
        self.assertEqual(rate.no_secured_cap_tinh, 2.5)
        self.assertEqual(rate.no_secured_cap_tw, 0.7)
        self.assertEqual(rate.secured_to_truong, 90.0)
        self.assertEqual(rate.secured_cap_xa, 10.0)
        self.assertEqual(rate.total_no_secured(), 100.0)
        self.assertEqual(rate.total_secured(), 100.0)

    def test_data_tvv_field_labels_complete(self) -> None:
        self.assertEqual(set(DATA_TVV_FIELD_LABELS), set(DATA_TVV_HEADERS))
        self.assertEqual(DATA_TVV_FIELD_LABELS["MaTo"], "Mã số tổ")
        self.assertEqual(DATA_TVV_FIELD_LABELS["TenTo"], "Tên tổ vay vốn")
        self.assertEqual(
            DATA_TVV_FIELD_LABELS["TK_HUYEN"],
            "Tài khoản tổ hội cấp huyện",
        )

    def test_default_info_empty_does_not_crash(self) -> None:
        self.assertEqual(get_default_credit_group_info(), {})

    def test_default_info_settings_save_load(self) -> None:
        saved = save_default_credit_group_info(
            {
                "xa": "Đan Phượng",
                "ten_huyen": "Hội nông dân huyện Lâm Hà",
                "uy_quyen": "1",
                "ma_to": "SHOULD_NOT_SAVE",
            },
            self.database_path,
        )

        loaded = load_default_credit_group_info(self.database_path)

        self.assertEqual(saved["xa"], "Đan Phượng")
        self.assertEqual(loaded["ten_huyen"], "Hội nông dân huyện Lâm Hà")
        self.assertEqual(loaded["uy_quyen"], "Có")
        self.assertNotIn("ma_to", loaded)

    def test_apply_default_info_fills_empty_only(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        save_default_credit_group_info(
            {
                "xa": "Đan Phượng",
                "dia_chi": "Địa chỉ mặc định",
                "ten_tinh": "Hội nông dân tỉnh Lâm Đồng",
                "uy_quyen": "Có",
            },
            self.database_path,
        )
        dialog = CreditGroupEditDialog(self.repository, mode="add")
        try:
            assert isinstance(dialog.inputs["xa"], QLineEdit)
            dialog.inputs["xa"].setText("Xã đã nhập")
            with patch(
                "agribank_v3.features.credit.tovayvon.placeholder_windows.QMessageBox.information"
            ):
                dialog.apply_default_info()

            self.assertEqual(dialog.inputs["xa"].text(), "Xã đã nhập")
            self.assertEqual(dialog.inputs["dia_chi"].text(), "Địa chỉ mặc định")
            self.assertEqual(dialog.inputs["ten_tinh"].text(), "Hội nông dân tỉnh Lâm Đồng")
            uy_quyen = dialog.inputs["uy_quyen"]
            self.assertIsInstance(uy_quyen, QComboBox)
            assert isinstance(uy_quyen, QComboBox)
            self.assertEqual(uy_quyen.currentText(), "Có")
        finally:
            dialog.close()

    def test_uyquyen_option_normalization(self) -> None:
        for value in ("Có", "Co", "1", "True", "Yes"):
            self.assertEqual(normalize_uy_quyen(value), "Có")
        for value in ("Không", "Khong", "0", "False", "No", ""):
            self.assertEqual(normalize_uy_quyen(value), "Không")

    def test_form_placeholders_mapping(self) -> None:
        self.assertEqual(FIELD_PLACEHOLDERS["ma_to"], "5491LLG202100003")
        self.assertEqual(FIELD_PLACEHOLDERS["ten_to"], "Tổ dịch vụ TD HND Đan Phượng 01")
        self.assertEqual(
            FIELD_PLACEHOLDERS["ttln_tw"],
            "012025/TTLN-HNDVN-AGRIBANK ngày 26/12/2025",
        )
        self.assertEqual(
            get_suggested_credit_group_default_info()["uy_quyen"],
            "Không",
        )

    def test_default_info_commission_labels_vietnamese(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dialog = CreditGroupDefaultInfoDialog(self.database_path)
        try:
            label_texts = {
                label.text()
                for label in dialog.findChildren(QLabel)
                if label.text()
            }

            self.assertIn("Tỷ lệ hoa hồng không TSBĐ (%)", label_texts)
            self.assertIn("Tỷ lệ hoa hồng có TSBĐ (%)", label_texts)
            self.assertIn("Tổ trưởng (%)", label_texts)
            self.assertIn("Cấp xã (%)", label_texts)
            self.assertIn("Cấp huyện (%)", label_texts)
            self.assertIn("Cấp tỉnh (%)", label_texts)
            self.assertIn("Cấp TW (%)", label_texts)
            self.assertNotIn("base_no_secured_rate", label_texts)
            self.assertNotIn("no_secured_to_truong", label_texts)
            self.assertNotIn("secured_to_truong", label_texts)
        finally:
            dialog.close()

    def test_default_info_save_load_keys_unchanged(self) -> None:
        saved = save_default_credit_group_info(
            {
                "base_no_secured_rate": "4",
                "base_secured_rate": "1",
                "no_secured_to_truong": "70",
                "secured_to_truong": "85",
            },
            self.database_path,
        )

        loaded = load_default_credit_group_info(self.database_path)

        self.assertEqual(
            DEFAULT_COMMISSION_FIELD_LABELS["base_no_secured_rate"],
            "Tỷ lệ hoa hồng không TSBĐ (%)",
        )
        self.assertEqual(saved["base_no_secured_rate"], "4")
        self.assertEqual(saved["no_secured_to_truong"], "70")
        self.assertEqual(loaded["base_secured_rate"], "1")
        self.assertEqual(loaded["secured_to_truong"], "85")

    def test_interest_report_required_columns(self) -> None:
        source_path = Path(self.temporary_directory.name) / "thu_lai.xlsx"
        self._write_interest_workbook(
            source_path,
            rows=[
                {
                    "MaKH": "KH001",
                    "TenKH": "Nguyễn Văn A",
                    "SoGiaiNgan": "GN001",
                    "DuNo": 100_000_000,
                    "SoLaiDaThuTrongKy": 1_000_000,
                    "TyLeBaoBam": 1,
                    "MaToVayVon": "T001",
                    "NhomNo": "1",
                }
            ],
        )

        validate_interest_report_columns(
            source_path,
            (
                "MaKH",
                "TenKH",
                "SoGiaiNgan",
                "DuNo",
                "SoLaiDaThuTrongKy",
                "TyLeBaoBam",
                "MaToVayVon",
            ),
        )

    def test_sktl_header_detection_real_file(self) -> None:
        source_path = Path("AgribankV2-PyThon/DuLieuTEST/TOVAYVON/SKTL.xls")
        if not source_path.is_file():
            self.skipTest("Chưa có file SKTL.xls mẫu.")

        detection = detect_sktl_columns(source_path)

        self.assertEqual(detection.sheet_name, "lai0104")
        self.assertEqual(detection.field_to_header["customer_code"], "MaKH")
        self.assertEqual(detection.field_to_header["loan_number"], "SoGiaiNgan")
        self.assertEqual(detection.field_to_header["group_code"], "MaToVayVon")
        self.assertEqual(detection.field_to_header["interest_collected"], "SoLaiDaThuTrongKy")
        self.assertFalse(detection.missing_required)

    def test_skck_header_detection_real_file(self) -> None:
        source_path = Path("AgribankV2-PyThon/DuLieuTEST/TOVAYVON/SKCK.xls")
        if not source_path.is_file():
            self.skipTest("Chưa có file SKCK.xls mẫu.")

        detection = detect_skck_columns(source_path)

        self.assertEqual(detection.sheet_name, "duno3006")
        self.assertEqual(detection.field_to_header["customer_code"], "MaKH")
        self.assertEqual(detection.field_to_header["loan_number"], "SoGiaiNgan")
        self.assertEqual(detection.field_to_header["group_code"], "MaToVayVon")
        self.assertEqual(detection.field_to_header["outstanding_balance"], "DuNo")
        self.assertEqual(detection.field_to_header["debt_group"], "NhomNo")
        self.assertEqual(detection.field_to_header["interest_rate"], "LaiSuat")
        self.assertFalse(detection.missing_required)

    def test_interest_report_missing_columns(self) -> None:
        source_path = Path(self.temporary_directory.name) / "missing.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(("MaKH", "TenKH"))
        workbook.save(source_path)
        workbook.close()

        with self.assertRaisesRegex(InterestReportError, "thiếu cột bắt buộc"):
            validate_interest_report_columns(source_path, ("MaKH", "MaToVayVon"))

    def test_report_period_dates_required(self) -> None:
        with self.assertRaisesRegex(InterestReportError, "Vui lòng nhập kỳ thu lãi"):
            create_interest_report(
                InterestReportRequest(
                    interest_file=Path("missing_sktl.xlsx"),
                    debt_file=Path("missing_skck.xlsx"),
                    output_path=Path("out.xlsx"),
                    from_date=None,  # type: ignore[arg-type]
                    to_date=date(2026, 6, 30),
                ),
                self.repository,
            )

    def test_report_period_from_date_before_to_date(self) -> None:
        with self.assertRaisesRegex(InterestReportError, "Từ ngày không được lớn hơn đến ngày"):
            create_interest_report(
                InterestReportRequest(
                    interest_file=Path("missing_sktl.xlsx"),
                    debt_file=Path("missing_skck.xlsx"),
                    output_path=Path("out.xlsx"),
                    from_date=date(2026, 7, 1),
                    to_date=date(2026, 6, 30),
                ),
                self.repository,
            )

    def test_interest_report_uses_group_commission_rate(self) -> None:
        output_path = self._create_sample_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = workbook["TongHopTheoTo"]
            headers = [cell.value for cell in sheet[1]]
            row = {headers[index]: value for index, value in enumerate(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))}
            self.assertEqual(row["MaTo"], "T001")
            self.assertEqual(row["HHCoBD"], 20_000)
            self.assertEqual(row["HHKhongBD"], 15_000)
            self.assertEqual(row["TongHH"], 35_000)
            self.assertEqual(row["HH_ToTruong"], 26_500)
        finally:
            workbook.close()

    def test_interest_report_uses_commission_rules(self) -> None:
        self.assertEqual(
            calculate_commission_pay_rate(
                0.96,
                0.0,
                CreditCommissionRuleSettings(),
            ),
            1.0,
        )
        self.assertEqual(
            calculate_commission_pay_rate(
                0.96,
                0.02,
                CreditCommissionRuleSettings(),
            ),
            0.0,
        )

    def test_commission_not_equal_total_interest_when_condition_is_50_percent(self) -> None:
        summary = self._commission_summary(
            lai_khong_bd=15_945_668,
            lai_co_bd=96_946_085,
        )

        commission = _commission_breakdown(summary, 0.5)

        self.assertEqual(commission["total"], 1_208_646)
        self.assertNotEqual(commission["total"], 112_891_753)

    def test_no_secured_commission_distribution(self) -> None:
        summary = self._commission_summary(lai_khong_bd=15_945_668, lai_co_bd=0)

        commission = _commission_breakdown(summary, 1.0)

        self.assertEqual(commission["unsecured_base"], 478_370)
        self.assertEqual(commission["unsecured"], 478_370)
        self.assertEqual(commission["no_secured_to_truong"], 382_696)
        self.assertEqual(commission["no_secured_cap_xa"], 62_188)
        self.assertEqual(commission["no_secured_cap_huyen"], 18_178)
        self.assertEqual(commission["no_secured_cap_tinh"], 11_959)
        self.assertEqual(commission["no_secured_cap_tw"], 3_349)
        self.assertNotEqual(commission["unsecured"], 15_945_668)

    def test_secured_commission_distribution(self) -> None:
        summary = self._commission_summary(lai_khong_bd=0, lai_co_bd=96_946_085)

        commission = _commission_breakdown(summary, 1.0)

        self.assertEqual(commission["secured_base"], 1_938_922)
        self.assertEqual(commission["secured"], 1_938_922)
        self.assertEqual(commission["secured_to_truong"], 1_745_030)
        self.assertEqual(commission["secured_cap_xa"], 193_892)
        self.assertNotEqual(commission["secured"], 96_946_085)

    def test_commission_condition_50_percent(self) -> None:
        summary = self._commission_summary(lai_khong_bd=15_945_668, lai_co_bd=0)

        commission = _commission_breakdown(summary, 0.5)

        self.assertEqual(commission["unsecured_base"], 478_370)
        self.assertEqual(commission["unsecured"], 239_185)
        self.assertEqual(commission["no_secured_to_truong"], 191_348)

    def test_commission_condition_bad_debt_zero(self) -> None:
        summary = self._commission_summary(lai_khong_bd=15_945_668, lai_co_bd=96_946_085)

        commission = _commission_breakdown(summary, 0.0)

        self.assertEqual(commission["total"], 0)
        self.assertEqual(commission["to_truong"], 0)
        self.assertEqual(commission["cap_xa"], 0)

    def test_commission_uses_group_specific_rate(self) -> None:
        summary = self._commission_summary(
            lai_khong_bd=1_000_000,
            lai_co_bd=0,
            rate=CreditGroupCommissionRate(
                ma_to="T001",
                no_secured_to_truong=60,
                no_secured_cap_xa=40,
                no_secured_cap_huyen=0,
                no_secured_cap_tinh=0,
                no_secured_cap_tw=0,
            ),
        )

        commission = _commission_breakdown(summary, 1.0)

        self.assertEqual(commission["unsecured_base"], 30_000)
        self.assertEqual(commission["no_secured_to_truong"], 18_000)
        self.assertEqual(commission["no_secured_cap_xa"], 12_000)

    def test_commission_base_rate_applied_before_condition_rate(self) -> None:
        summary = self._commission_summary(lai_khong_bd=1_000_000, lai_co_bd=1_000_000)

        commission = _commission_breakdown(summary, 0.5)

        self.assertEqual(commission["unsecured_base"], 30_000)
        self.assertEqual(commission["secured_base"], 20_000)
        self.assertEqual(commission["total"], 25_000)

    def test_distribution_applies_after_commission_amount(self) -> None:
        summary = self._commission_summary(lai_khong_bd=1_000_000, lai_co_bd=0)

        commission = _commission_breakdown(summary, 0.5)

        self.assertEqual(commission["unsecured"], 15_000)
        self.assertEqual(commission["no_secured_to_truong"], 12_000)
        self.assertNotEqual(commission["no_secured_to_truong"], 400_000)

    def test_interest_report_export_xlsx(self) -> None:
        output_path = self._create_sample_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            self.assertIn("T001", workbook.sheetnames)
            self.assertIn("TongHopTheoTo", workbook.sheetnames)
            self.assertIn("CanhBao", workbook.sheetnames)
            self.assertEqual(workbook["TongHopTheoTo"]["A1"].value, "MaTo")
        finally:
            workbook.close()

    def test_bangke_sorted_by_customer_code_column_c(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            records = self._bangke_records(workbook)
            customer_codes = [record["Mã khách hàng"] for record in records]
            self.assertEqual(customer_codes, sorted(customer_codes))
        finally:
            workbook.close()

    def test_bangke_does_not_include_extra_columns(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = self._main_bangke_sheet(workbook)
            headers = [sheet.cell(12, column).value for column in range(1, sheet.max_column + 1)]
            self.assertNotIn("HoaHongGoc", headers)
            self.assertNotIn("HoaHongThucNhan", headers)
            self.assertNotIn("TrangThaiKhoanVay", headers)
            self.assertIn("Hoa hồng thu lãi", headers)
        finally:
            workbook.close()

    def test_bangke_note_column_no_sktl_only_long_message(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            notes = [record.get("Ghi chú") for record in self._bangke_records(workbook)]
            self.assertNotIn("Có thu lãi trong kỳ nhưng không còn dư nợ cuối kỳ", notes)
        finally:
            workbook.close()

    def test_bangke_period_dates_written(self) -> None:
        output_path = self._create_sample_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = self._main_bangke_sheet(workbook)
            self.assertEqual(sheet["A6"].value, "(Từ ngày 01/01/2026 đến ngày 31/03/2026)")
            for record in self._bangke_records(workbook):
                self.assertIsNotNone(record["Thu lãi từ ngày"])
                self.assertIsNotNone(record["Thu lãi đến ngày"])
        finally:
            workbook.close()

    def test_report_period_dates_written_to_detail_rows_from_vba_source(self) -> None:
        output_path = self._create_period_source_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            record = self._bangke_records(workbook)[0]
            self.assertEqual(self._as_date(record["Thu lãi từ ngày"]), date(2026, 6, 1))
            self.assertEqual(self._as_date(record["Thu lãi đến ngày"]), date(2026, 6, 30))
        finally:
            workbook.close()

    def test_total_collected_label_uses_to_date(self) -> None:
        output_path = self._create_period_source_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = self._main_bangke_sheet(workbook)
            row_index = self._row_by_label(sheet, "Tổng lãi thực thu đến")
            self.assertEqual(sheet.cell(row_index, 2).value, "Tổng lãi thực thu đến 30/06/2026:")
        finally:
            workbook.close()

    def test_period_date_source_matches_vba_not_ui_from_date(self) -> None:
        output_path = self._create_period_source_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            record = self._bangke_records(workbook)[0]
            self.assertNotEqual(self._as_date(record["Thu lãi từ ngày"]), date(2026, 4, 1))
            self.assertEqual(self._as_date(record["Thu lãi từ ngày"]), date(2026, 6, 1))
        finally:
            workbook.close()

    def test_report_period_dates_use_existing_file_columns_when_available(self) -> None:
        output_path = self._create_period_source_interest_report(
            period_from=date(2026, 5, 5),
            period_to=date(2026, 5, 20),
        )
        workbook = load_workbook(output_path, data_only=True)
        try:
            record = self._bangke_records(workbook)[0]
            self.assertEqual(self._as_date(record["Thu lãi từ ngày"]), date(2026, 5, 5))
            self.assertEqual(self._as_date(record["Thu lãi đến ngày"]), date(2026, 5, 20))
        finally:
            workbook.close()

    def test_bangke_format_has_title_and_header(self) -> None:
        output_path = self._create_sample_interest_report()
        workbook = load_workbook(output_path)
        try:
            sheet = self._main_bangke_sheet(workbook)
            self.assertIn("BẢNG KÊ THU LÃI", sheet["A5"].value)
            self.assertEqual(sheet["A12"].value, "STT")
            self.assertTrue(sheet["A12"].font.bold)
            self.assertEqual(sheet["A12"].border.left.style, "thin")
            self.assertEqual(sheet["N12"].value, "Ghi chú")
        finally:
            workbook.close()

    def test_bangke_group_info_header_is_left_aligned(self) -> None:
        output_path = self._create_sample_interest_report()
        workbook = load_workbook(output_path)
        try:
            sheet = self._main_bangke_sheet(workbook)
            for address in ("B7", "C7", "B8", "C8", "B9", "B10"):
                self.assertEqual(sheet[address].alignment.horizontal, "left")
        finally:
            workbook.close()

    def test_bangke_branch_title_uses_database_branch_name(self) -> None:
        settings_database = AppSettingsDatabase(self.database_path)
        settings_database.save_branch_profile(
            BranchProfile(
                branch_name="Chi nhánh Lộc Phát Lâm Đồng",
                reporting_branch_name="Chi nhánh Đức Trọng Lâm Đồng",
            )
        )

        output_path = self._create_sample_interest_report()
        workbook = load_workbook(output_path)
        try:
            sheet = self._main_bangke_sheet(workbook)
            self.assertEqual(sheet["A3"].value, "CHI NHÁNH LỘC PHÁT LÂM ĐỒNG")
        finally:
            workbook.close()

    def test_summary_total_interest_number_format(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path)
        try:
            sheet = self._main_bangke_sheet(workbook)
            row_index = self._row_by_label(sheet, "Tổng lãi phải thu:")
            cell = sheet.cell(row_index, 4)
            self.assertIsInstance(cell.value, (int, float))
            self.assertEqual(cell.number_format, '#,##0')
            self.assertEqual(cell.alignment.horizontal, "right")
        finally:
            workbook.close()

    def test_summary_total_collected_number_format(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path)
        try:
            sheet = self._main_bangke_sheet(workbook)
            row_index = self._row_by_label(sheet, "Tổng lãi thực thu đến")
            cell = sheet.cell(row_index, 4)
            self.assertIsInstance(cell.value, (int, float))
            self.assertEqual(cell.number_format, '#,##0')
            self.assertEqual(cell.alignment.horizontal, "right")
        finally:
            workbook.close()

    def test_summary_percent_cells_still_percent_format(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path)
        try:
            sheet = self._main_bangke_sheet(workbook)
            interest_row = self._row_by_label(sheet, "Tổng lãi phải thu:")
            bad_debt_row = self._row_by_label(sheet, "Tổng lãi thực thu đến")
            self.assertEqual(sheet.cell(interest_row, 7).number_format, "0.00%")
            self.assertEqual(sheet.cell(bad_debt_row, 7).number_format, "0.00%")
        finally:
            workbook.close()

    def test_excel_summary_commission_values(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = self._main_bangke_sheet(workbook)
            lines = self._commission_lines(sheet)
            self.assertEqual(lines["*HOA HỒNG ĐỐI VỚI KHOẢN KHÔNG CÓ TSĐB (3%)"], 6_000)
            self.assertEqual(lines["* HOA HỒNG THỰC NHẬN (100%)"], 6_000)
            self.assertEqual(lines["Không BĐ - Tổ trưởng"], 4_800)
            self.assertEqual(lines["Không BĐ - Cấp xã"], 780)
            self.assertEqual(lines["*HOA HỒNG ĐỐI VỚI KHOẢN CÓ TSĐB (2%)"], 20_000)
            self.assertEqual(lines["Có BĐTS - Tổ trưởng"], 18_000)
            self.assertEqual(lines["Có BĐTS - Cấp xã"], 2_000)
        finally:
            workbook.close()

    def test_excel_detail_hoa_hong_thu_lai_column_uses_base_rate(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = self._main_bangke_sheet(workbook)
            rows = self._bangke_records(workbook)
            secured_row = next(item for item in rows if item["Mã khách hàng"] == "KH_JOIN")
            no_secured_row = next(item for item in rows if item["Mã khách hàng"] == "KH_ONLY")
            self.assertEqual(secured_row["Hoa hồng thu lãi"], 20_000)
            self.assertEqual(no_secured_row["Hoa hồng thu lãi"], 6_000)
            self.assertNotEqual(secured_row["Hoa hồng thu lãi"], secured_row["Lãi có bảo đảm"])
            self.assertNotEqual(no_secured_row["Hoa hồng thu lãi"], no_secured_row["Lãi không bảo đảm"])
            total_row = next(row for row in range(13, sheet.max_row + 1) if sheet.cell(row, 1).value == "Tổng cộng")
            self.assertEqual(sheet.cell(total_row, 13).value, 26_000)
        finally:
            workbook.close()

    def test_interest_report_log_includes_commission_base_and_net_amounts(self) -> None:
        _, result = self._create_sktl_only_interest_report()

        log_text = "\n".join(result.info_messages)

        self.assertIn("Lãi không BĐ: 200,000", log_text)
        self.assertIn("Tỷ lệ hoa hồng không BĐ: 3%", log_text)
        self.assertIn("Hoa hồng không BĐ thực nhận: 6,000", log_text)
        self.assertIn("Lãi có BĐTS: 1,000,000", log_text)
        self.assertIn("Tỷ lệ hoa hồng có BĐTS: 2%", log_text)
        self.assertIn("Hoa hồng có BĐTS thực nhận: 20,000", log_text)

    def test_interest_report_uses_group_commission_base_rates(self) -> None:
        output_path, result = self._create_sktl_only_interest_report(
            rate=CreditGroupCommissionRate(
                ma_to="T001",
                base_secured_rate=1.0,
                base_no_secured_rate=4.0,
            )
        )
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = self._main_bangke_sheet(workbook)
            rows = self._bangke_records(workbook)
            secured_row = next(item for item in rows if item["Mã khách hàng"] == "KH_JOIN")
            no_secured_row = next(item for item in rows if item["Mã khách hàng"] == "KH_ONLY")
            self.assertEqual(secured_row["Hoa hồng thu lãi"], 10_000)
            self.assertEqual(no_secured_row["Hoa hồng thu lãi"], 8_000)
            lines = self._commission_lines(sheet)
            self.assertEqual(lines["*HOA HỒNG ĐỐI VỚI KHOẢN KHÔNG CÓ TSĐB (4%)"], 8_000)
            self.assertEqual(lines["*HOA HỒNG ĐỐI VỚI KHOẢN CÓ TSĐB (1%)"], 10_000)
            log_text = "\n".join(result.info_messages)
            self.assertIn("Tỷ lệ hoa hồng không BĐ: 4%", log_text)
            self.assertIn("Tỷ lệ hoa hồng có BĐTS: 1%", log_text)
        finally:
            workbook.close()

    def test_interest_report_detail_commission_matches_tong_hop_theo_to(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T900", ten_to="Tổ 900", ten_to_truong="Tổ trưởng 900")
        )
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T900",
                base_secured_rate=2,
                base_no_secured_rate=4,
            )
        )
        interest_path = Path(self.temporary_directory.name) / "thu_lai_rounding.xlsx"
        debt_path = Path(self.temporary_directory.name) / "du_no_rounding.xlsx"
        output_path = Path(self.temporary_directory.name) / "BangKeThuLaiTo_Rounding.xlsx"
        rows = [
            {
                "MaKH": "KH_SEC",
                "TenKH": "Khách có bảo đảm",
                "SoGiaiNgan": "GN_SEC",
                "DuNo": 100_000_000,
                "SoLaiDaThuTrongKy": 72_619_528,
                "TyLeBaoBam": 1,
                "MaToVayVon": "T900",
                "NhomNo": "1",
                "LaiTon": 0,
                "LaiSuat": 12,
                "NgayTraLaiCuoiCung": date(2026, 3, 31),
            },
            {
                "MaKH": "KH_NO",
                "TenKH": "Khách không bảo đảm",
                "SoGiaiNgan": "GN_NO",
                "DuNo": 50_000_000,
                "SoLaiDaThuTrongKy": 32_794_854,
                "TyLeBaoBam": 0,
                "MaToVayVon": "T900",
                "NhomNo": "1",
                "LaiTon": 0,
                "LaiSuat": 12,
                "NgayTraLaiCuoiCung": date(2026, 3, 31),
            },
        ]
        self._write_interest_workbook(interest_path, rows=rows)
        self._write_interest_workbook(debt_path, rows=rows)

        create_interest_report(
            InterestReportRequest(
                interest_file=interest_path,
                debt_file=debt_path,
                output_path=output_path,
                from_date=date(2026, 1, 1),
                to_date=date(2026, 3, 31),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            summary = self._sheet_records(workbook["TongHopTheoTo"])[0]
            records = self._bangke_records(workbook)
            detail_total = sum(record["Hoa hồng thu lãi"] for record in records)
            total_row = next(
                row for row in range(13, workbook["T900"].max_row + 1)
                if workbook["T900"].cell(row, 1).value == "Tổng cộng"
            )

            self.assertEqual(summary["HHCoBD"], 1_452_391)
            self.assertEqual(summary["HHKhongBD"], 1_311_794)
            self.assertEqual(summary["TongHH"], 2_764_185)
            self.assertEqual(detail_total, summary["TongHH"])
            self.assertEqual(workbook["T900"].cell(total_row, 13).value, summary["TongHH"])
        finally:
            workbook.close()

    def test_interest_report_adjusts_detail_rounding_residue_to_group_total(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T901", ten_to="Tổ 901"))
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T901",
                base_secured_rate=2,
                base_no_secured_rate=4,
            )
        )
        interest_path = Path(self.temporary_directory.name) / "thu_lai_residue.xlsx"
        debt_path = Path(self.temporary_directory.name) / "du_no_residue.xlsx"
        output_path = Path(self.temporary_directory.name) / "BangKeThuLaiTo_Residue.xlsx"
        rows = [
            {
                "MaKH": f"KH{index}",
                "TenKH": f"Khách {index}",
                "SoGiaiNgan": f"GN{index}",
                "DuNo": 10_000_000,
                "SoLaiDaThuTrongKy": 10_931_618,
                "TyLeBaoBam": 0,
                "MaToVayVon": "T901",
                "NhomNo": "1",
                "LaiTon": 0,
                "LaiSuat": 12,
                "NgayTraLaiCuoiCung": date(2026, 3, 31),
            }
            for index in range(1, 4)
        ]
        self._write_interest_workbook(interest_path, rows=rows)
        self._write_interest_workbook(debt_path, rows=rows)

        create_interest_report(
            InterestReportRequest(
                interest_file=interest_path,
                debt_file=debt_path,
                output_path=output_path,
                from_date=date(2026, 1, 1),
                to_date=date(2026, 3, 31),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            summary = self._sheet_records(workbook["TongHopTheoTo"])[0]
            detail_total = sum(
                record["Hoa hồng thu lãi"] for record in self._bangke_records(workbook)
            )
            self.assertEqual(summary["HHKhongBD"], 1_311_794)
            self.assertEqual(detail_total, summary["TongHH"])
        finally:
            workbook.close()

    def test_interest_report_uses_group_custom_commission_rule(self) -> None:
        self.repository.save_commission_rule_settings(
            CreditCommissionRuleSettings(interest_pay_3=0)
        )

        output_path, result = self._create_sktl_only_interest_report(
            group_rule=CreditGroupCommissionRule(
                ma_to="T001",
                use_custom_rule=True,
                interest_pay_3=50,
            )
        )
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = self._main_bangke_sheet(workbook)
            lines = self._commission_lines(sheet)
            self.assertEqual(lines["* HOA HỒNG THỰC NHẬN (50%)"], 3_000)
            self.assertIn("Điều kiện chi: riêng theo tổ", "\n".join(result.info_messages))
        finally:
            workbook.close()

    def test_sktl_only_row_is_included(self) -> None:
        output_path, result = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            row = next(item for item in self._bangke_records(workbook) if item["Mã khách hàng"] == "KH_ONLY")
            self.assertEqual(row["Số giải ngân"], "GN_ONLY")
            self.assertIn("SKTL", "\n".join(result.warnings))
        finally:
            workbook.close()

    def test_sktl_only_row_outstanding_zero(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            row = next(item for item in self._bangke_records(workbook) if item["Mã khách hàng"] == "KH_ONLY")
            self.assertEqual(row["Dư nợ"], 0)
            self.assertNotIn("TrangThaiKhoanVay", row)
            self.assertEqual(row["Ghi chú"], None)
        finally:
            workbook.close()

    def test_sktl_only_row_counts_interest_collected(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            summary = self._sheet_records(workbook["TongHopTheoTo"])[0]
            self.assertEqual(summary["LaiCoBD"], 1_000_000)
            self.assertEqual(summary["LaiKhongBD"], 200_000)
        finally:
            workbook.close()

    def test_sktl_only_row_not_bad_debt(self) -> None:
        output_path, _ = self._create_sktl_only_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            summary = self._sheet_records(workbook["TongHopTheoTo"])[0]
            self.assertEqual(summary["TongDuNo"], 100_000_000)
            self.assertEqual(summary["TyLeNoXau"], 0)
        finally:
            workbook.close()

    def test_sktl_only_warning_summary_not_spam_log(self) -> None:
        output_path, result = self._create_sktl_only_interest_report()
        self.assertTrue(
            any("Có 1 khoản có trong SKTL nhưng không có trong SKCK" in warning for warning in result.warnings)
        )
        self.assertFalse(any("KH_ONLY/GN_ONLY" in warning for warning in result.warnings))

        workbook = load_workbook(output_path, data_only=True)
        try:
            detail = next(item for item in self._sheet_records(workbook["CanhBao"]) if item["Loai"] == "SKTL-only")
            self.assertEqual(detail["MaKH"], "KH_ONLY")
            self.assertEqual(detail["SoGiaiNgan"], "GN_ONLY")
            self.assertEqual(detail["LaiDaThu"], 200_000)
            self.assertEqual(detail["GhiChu"], "Có thu lãi trong kỳ nhưng không còn dư nợ cuối kỳ")
        finally:
            workbook.close()

    def test_remaining_interest_formula_matches_vba_if_columns_available(self) -> None:
        output_path = self._create_formula_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            row = self._bangke_records(workbook)[0]
            self.assertAlmostEqual(row["Lãi tồn"], 100_000_000 * 0.12 * 30 / 365)
        finally:
            workbook.close()

    def test_interest_receivable_formula_matches_vba_if_columns_available(self) -> None:
        output_path = self._create_formula_interest_report()
        workbook = load_workbook(output_path, data_only=True)
        try:
            row = self._bangke_records(workbook)[0]
            self.assertAlmostEqual(row["Lãi phải thu"], 1_000_000 + 100_000_000 * 0.12 * 30 / 365)
        finally:
            workbook.close()

    def test_sktl_only_remaining_interest_uses_vba_formula_if_columns_available(self) -> None:
        self.repository.save_group(
            CreditGroup(
                stt=1,
                ma_to="T001",
                ten_to="Tổ 001",
                ten_to_truong="Nguyễn Văn A",
            )
        )
        interest_path = Path(self.temporary_directory.name) / "sktl_only_formula.xlsx"
        debt_path = Path(self.temporary_directory.name) / "skck_empty.xlsx"
        output_path = Path(self.temporary_directory.name) / "BangKeThuLaiTo_SKTLOnlyFormula.xlsx"
        self._write_workbook(
            interest_path,
            headers=[
                "MaKH",
                "TenKH",
                "SoGiaiNgan",
                "DuNo",
                "SoLaiDaThuTrongKy",
                "TyLeBaoBam",
                "MaToVayVon",
                "LaiSuat",
                "NgayGiaiNgan",
                "NgayTraLaiCuoiCung",
            ],
            rows=[
                {
                    "MaKH": "KH_ONLY",
                    "TenKH": "Khách tất toán",
                    "SoGiaiNgan": "GN_ONLY",
                    "DuNo": 100_000_000,
                    "SoLaiDaThuTrongKy": 1_000_000,
                    "TyLeBaoBam": 1,
                    "MaToVayVon": "T001",
                    "LaiSuat": 12,
                    "NgayGiaiNgan": date(2026, 1, 1),
                    "NgayTraLaiCuoiCung": date(2026, 5, 31),
                }
            ],
        )
        self._write_interest_workbook(debt_path, rows=[])

        create_interest_report(
            InterestReportRequest(
                interest_file=interest_path,
                debt_file=debt_path,
                output_path=output_path,
                from_date=date(2026, 4, 1),
                to_date=date(2026, 6, 30),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            row = self._bangke_records(workbook)[0]
            self.assertEqual(row["Dư nợ"], 0)
            self.assertAlmostEqual(row["Lãi tồn"], 100_000_000 * 0.12 * 30 / 365)
            self.assertAlmostEqual(row["Lãi phải thu"], 1_000_000 + 100_000_000 * 0.12 * 30 / 365)
        finally:
            workbook.close()

    def test_custom_commission_rate_is_saved_per_ma_to(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        self.repository.save_group(
            CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002")
        )

        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                base_no_secured_rate=4,
                base_secured_rate=1,
                no_secured_to_truong=70,
                no_secured_cap_xa=15,
                no_secured_cap_huyen=5,
                no_secured_cap_tinh=5,
                no_secured_cap_tw=5,
                secured_to_truong=85,
                secured_cap_xa=10,
                secured_cap_huyen=2,
                secured_cap_tinh=2,
                secured_cap_tw=1,
            )
        )

        custom_rate = self.repository.get_or_create_commission_rate("T001")
        default_rate = self.repository.get_or_create_commission_rate("T002")
        self.assertEqual(custom_rate.base_no_secured_rate, 4.0)
        self.assertEqual(custom_rate.base_secured_rate, 1.0)
        self.assertEqual(custom_rate.no_secured_to_truong, 70.0)
        self.assertEqual(custom_rate.secured_to_truong, 85.0)
        self.assertEqual(default_rate.base_no_secured_rate, 3.0)
        self.assertEqual(default_rate.base_secured_rate, 2.0)
        self.assertEqual(default_rate.no_secured_to_truong, 80.0)
        self.assertEqual(default_rate.secured_to_truong, 90.0)

    def test_deactivate_group_is_soft_delete(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                base_no_secured_rate=4,
                base_secured_rate=1,
            )
        )

        self.repository.deactivate_group("T001")

        self.assertEqual(self.repository.list_groups(), [])
        groups = self.repository.list_groups(include_inactive=True)
        self.assertEqual(len(groups), 1)
        self.assertFalse(groups[0].active)
        rate = self.repository.get_or_create_commission_rate("T001")
        self.assertEqual(rate.base_no_secured_rate, 4.0)

    def test_reactivate_group(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        self.repository.soft_deactivate_group("T001")

        self.repository.reactivate_group("T001")

        group = self.repository.get_group("T001")
        self.assertIsNotNone(group)
        assert group is not None
        self.assertTrue(group.active)
        self.assertEqual([item.ma_to for item in self.repository.list_groups()], ["T001"])

    def test_delete_group_permanently(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(ma_to="T001", base_no_secured_rate=4)
        )
        self.repository.save_group_commission_rule(
            CreditGroupCommissionRule(
                ma_to="T001",
                use_custom_rule=True,
                interest_pay_3=50,
            )
        )

        self.repository.delete_group_permanently("T001")

        self.assertIsNone(self.repository.get_group("T001"))
        self.assertIsNone(self.repository.get_commission_rate("T001"))
        self.assertFalse(
            self.repository.get_group_commission_rule("T001").use_custom_rule
        )

    def test_group_custom_commission_rule_overrides_global_rule(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        self.repository.save_commission_rule_settings(
            CreditCommissionRuleSettings(interest_pay_3=0)
        )
        self.repository.save_group_commission_rule(
            CreditGroupCommissionRule(
                ma_to="T001",
                use_custom_rule=True,
                interest_pay_3=50,
            )
        )

        settings, uses_custom_rule = self.repository.commission_rule_for_group("T001")

        self.assertTrue(uses_custom_rule)
        self.assertEqual(settings.interest_pay_3, 50.0)

    def test_group_commission_rule_falls_back_to_global_rule(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        self.repository.save_commission_rule_settings(
            CreditCommissionRuleSettings(interest_pay_3=80)
        )

        settings, uses_custom_rule = self.repository.commission_rule_for_group("T001")

        self.assertFalse(uses_custom_rule)
        self.assertEqual(settings.interest_pay_3, 80.0)

    def test_disable_custom_rule_falls_back_to_global(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        self.repository.save_commission_rule_settings(
            CreditCommissionRuleSettings(interest_pay_3=80)
        )
        self.repository.save_group_commission_rule(
            CreditGroupCommissionRule(
                ma_to="T001",
                use_custom_rule=False,
                interest_pay_3=50,
            )
        )

        settings, uses_custom_rule = self.repository.get_effective_commission_rule("T001")

        self.assertFalse(uses_custom_rule)
        self.assertEqual(settings.interest_pay_3, 80.0)

    def test_outer_condition_tab_loads_custom_rule(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001", ten_to_truong="A")
        )
        self.repository.save_group_commission_rule(
            CreditGroupCommissionRule(
                ma_to="T001",
                use_custom_rule=True,
                interest_min_1=80,
                interest_max_1=90,
                interest_pay_1=50,
                interest_min_2=90,
                interest_max_2=95,
                interest_pay_2=90,
                interest_min_3=95,
                interest_pay_3=100,
                bad_debt_threshold=2,
                bad_debt_pay=0,
            )
        )

        window = CreditGroupManagementPlaceholderDialog(database_path=self.database_path)
        try:
            window._select_rule_group("T001")

            self.assertTrue(window.rule_use_custom_check.isChecked())
            self.assertEqual(window.rule_inputs["interest_min_1"].text(), "80")
            self.assertEqual(window.rule_inputs["interest_max_1"].text(), "90")
            self.assertEqual(window.rule_inputs["bad_debt_threshold"].text(), "2")
            self.assertIn("riêng", window.rule_mode_status.text())
        finally:
            window.close()

    def test_outer_condition_tab_loads_global_when_no_custom(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001", ten_to_truong="A")
        )
        self.repository.save_commission_rule_settings(
            CreditCommissionRuleSettings(interest_min_1=82, bad_debt_threshold=3)
        )

        window = CreditGroupManagementPlaceholderDialog(database_path=self.database_path)
        try:
            window._select_rule_group("T001")

            self.assertFalse(window.rule_use_custom_check.isChecked())
            self.assertEqual(window.rule_inputs["interest_min_1"].text(), "82")
            self.assertEqual(window.rule_inputs["bad_debt_threshold"].text(), "3")
            self.assertIn("mặc định chung", window.rule_mode_status.text())
        finally:
            window.close()

    def test_group_form_save_refreshes_parent_condition_tab(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001", ten_to_truong="A")
        )
        window = CreditGroupManagementPlaceholderDialog(database_path=self.database_path)
        try:
            window._select_rule_group("T001")
            self.repository.save_group_commission_rule(
                CreditGroupCommissionRule(
                    ma_to="T001",
                    use_custom_rule=True,
                    interest_min_1=80,
                    interest_max_1=90,
                    interest_pay_1=50,
                    interest_min_2=90,
                    interest_max_2=95,
                    interest_pay_2=90,
                    interest_min_3=95,
                    interest_pay_3=100,
                    bad_debt_threshold=2,
                    bad_debt_pay=0,
                )
            )

            window._load_groups()
            window._select_rule_group("T001")

            self.assertTrue(window.rule_use_custom_check.isChecked())
            self.assertEqual(window.rule_inputs["interest_min_1"].text(), "80")
        finally:
            window.close()

    def test_delete_requires_confirmation_logic(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        window = CreditGroupManagementPlaceholderDialog(database_path=self.database_path)
        try:
            window.groups_table.selectRow(0)
            with patch(
                "agribank_v3.features.credit.tovayvon.placeholder_windows.QMessageBox.question",
                return_value=QMessageBox.StandardButton.Yes,
            ), patch(
                "agribank_v3.features.credit.tovayvon.placeholder_windows.QInputDialog.getText",
                return_value=("SAI_MA", True),
            ), patch(
                "agribank_v3.features.credit.tovayvon.placeholder_windows.QMessageBox.warning"
            ):
                window._delete_selected_group_permanently()

            self.assertIsNotNone(self.repository.get_group("T001"))
        finally:
            window.close()

    def test_shared_group_selector_loads_rate_tab(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_group(CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002"))
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(ma_to="T002", base_no_secured_rate=4)
        )

        window = CreditGroupManagementPlaceholderDialog(database_path=self.database_path)
        try:
            window._select_commission_group("T002")

            self.assertEqual(window._selected_ma_to(), "T002")
            self.assertEqual(window.commission_inputs["base_no_secured_rate"].text(), "4")
            self.assertIn("T002", window.group_summary.text())
        finally:
            window.close()

    def test_shared_group_selector_loads_condition_tab(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_group(CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002"))
        self.repository.save_group_commission_rule(
            CreditGroupCommissionRule(
                ma_to="T002",
                use_custom_rule=True,
                interest_min_1=80,
                interest_max_1=90,
                interest_pay_1=50,
                interest_min_2=90,
                interest_max_2=95,
                interest_pay_2=90,
                interest_min_3=95,
                interest_pay_3=100,
                bad_debt_threshold=2,
                bad_debt_pay=0,
            )
        )

        window = CreditGroupManagementPlaceholderDialog(database_path=self.database_path)
        try:
            window._select_commission_group("T002")

            self.assertEqual(window._selected_rule_ma_to(), "T002")
            self.assertTrue(window.rule_use_custom_check.isChecked())
            self.assertEqual(window.rule_inputs["interest_min_1"].text(), "80")
        finally:
            window.close()

    def test_shared_group_selector_after_delete_clears_selection(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        window = CreditGroupManagementPlaceholderDialog(database_path=self.database_path)
        try:
            window._select_commission_group("T001")
            self.repository.delete_group_permanently("T001")

            window._load_groups()

            self.assertEqual(window._selected_ma_to(), "")
            self.assertIsNone(window.current_rate)
        finally:
            window.close()

    def test_no_duplicate_search_controls_inside_subtabs(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = CreditGroupManagementPlaceholderDialog(database_path=self.database_path)
        try:
            placeholders = [
                editor.placeholderText()
                for editor in window.findChildren(QLineEdit)
            ]

            self.assertEqual(
                placeholders.count("Nhập mã tổ, tên tổ, tổ trưởng, xã..."),
                1,
            )
        finally:
            window.close()

    def test_create_and_update_group_keeps_text_fields(self) -> None:
        self.repository.save_group(
            CreditGroup(
                stt=1,
                ma_to="T001",
                ten_to="Tổ 001",
                tk_to_truong="001234567890",
                so_dien_thoai="0912.345.678",
            )
        )

        self.repository.save_group(
            CreditGroup(
                stt=1,
                ma_to="T001",
                ten_to="Tổ 001 cập nhật",
                tk_to_truong="000000123456",
                so_dien_thoai="0987.000.111",
            )
        )

        group = self.repository.get_group("T001")
        self.assertIsNotNone(group)
        assert group is not None
        self.assertEqual(group.ten_to, "Tổ 001 cập nhật")
        self.assertEqual(group.tk_to_truong, "000000123456")
        self.assertEqual(group.so_dien_thoai, "0987.000.111")
        self.assertEqual(
            self.repository.get_or_create_commission_rate("T001").total_secured(),
            100.0,
        )

    def test_create_group_without_stt_assigns_next(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_group(CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002"))

        self.repository.save_group(CreditGroup(stt=0, ma_to="T003", ten_to="Tổ 003"))

        groups = self.repository.list_groups()
        self.assertEqual(
            [(group.ma_to, group.stt) for group in groups],
            [("T001", 1), ("T002", 2), ("T003", 3)],
        )

    def test_insert_group_with_duplicate_stt_reorders_sequence(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_group(CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002"))
        self.repository.save_group(CreditGroup(stt=3, ma_to="T003", ten_to="Tổ 003"))

        self.repository.save_group(CreditGroup(stt=2, ma_to="T004", ten_to="Tổ 004"))

        groups = self.repository.list_groups()
        self.assertEqual(
            [(group.ma_to, group.stt) for group in groups],
            [("T001", 1), ("T004", 2), ("T002", 3), ("T003", 4)],
        )

    def test_update_group_with_duplicate_stt_reorders_sequence(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_group(CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002"))
        self.repository.save_group(CreditGroup(stt=3, ma_to="T003", ten_to="Tổ 003"))

        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T003", ten_to="Tổ 003 cập nhật")
        )

        groups = self.repository.list_groups()
        self.assertEqual(
            [(group.ma_to, group.stt) for group in groups],
            [("T003", 1), ("T001", 2), ("T002", 3)],
        )
        self.assertEqual(groups[0].ten_to, "Tổ 003 cập nhật")

    def test_resequence_existing_duplicate_stt(self) -> None:
        now = "2026-01-01T00:00:00"
        with closing(sqlite3.connect(self.database_path)) as connection:
            connection.executescript(
                """
                INSERT INTO credit_groups(
                    ma_to, stt, ten_to, created_at, updated_at
                )
                VALUES
                    ('T001', 1, 'Tổ 001', '2026-01-01T00:00:01', '2026-01-01T00:00:01'),
                    ('T002', 1, 'Tổ 002', '2026-01-01T00:00:02', '2026-01-01T00:00:02'),
                    ('T003', 0, 'Tổ 003', '2026-01-01T00:00:03', '2026-01-01T00:00:03')
                """
            )
            connection.commit()

        self.repository.resequence_group_stt()

        groups = self.repository.list_groups()
        self.assertEqual(
            [(group.ma_to, group.stt) for group in groups],
            [("T001", 1), ("T002", 2), ("T003", 3)],
        )

    def test_invalid_commission_total_is_rejected(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )

        with self.assertRaises(CreditGroupRepositoryError):
            self.repository.save_commission_rate(
                CreditGroupCommissionRate(
                    ma_to="T001",
                    no_secured_to_truong=50,
                    no_secured_cap_xa=10,
                    no_secured_cap_huyen=10,
                    no_secured_cap_tinh=10,
                    no_secured_cap_tw=10,
                )
            )

        rate = self.repository.get_or_create_commission_rate("T001")
        self.assertEqual(rate.no_secured_to_truong, 80.0)

    def test_import_data_tvv_creates_default_commissions(self) -> None:
        source_path = Path(self.temporary_directory.name) / "DaTa_ToVayVon.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data_TVV"
        sheet.append(DATA_TVV_HEADERS)
        sheet.append(self._data_tvv_row(1, "T001", "Tổ 001"))
        sheet.append(self._data_tvv_row(2, "T002", "Tổ 002"))
        workbook.save(source_path)
        workbook.close()

        imported_count = self.repository.import_data_tvv(source_path)

        self.assertEqual(imported_count, 2)
        self.assertEqual(
            self.repository.get_or_create_commission_rate("T001").total_secured(),
            100.0,
        )
        self.assertEqual(
            self.repository.get_or_create_commission_rate("T002").total_no_secured(),
            100.0,
        )

    def test_export_data_tvv_preserves_original_22_columns_by_default(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                no_secured_to_truong=70,
                no_secured_cap_xa=15,
                no_secured_cap_huyen=5,
                no_secured_cap_tinh=5,
                no_secured_cap_tw=5,
                secured_to_truong=85,
                secured_cap_xa=10,
                secured_cap_huyen=2,
                secured_cap_tinh=2,
                secured_cap_tw=1,
            )
        )
        output_path = Path(self.temporary_directory.name) / "Data_TVV.xlsx"

        self.repository.export_data_tvv(output_path)

        workbook = load_workbook(output_path, read_only=True)
        try:
            headers = [cell.value for cell in next(workbook["Data_TVV"].iter_rows(max_row=1))]
        finally:
            workbook.close()
        self.assertEqual(headers, list(DATA_TVV_HEADERS))
        self.assertEqual(len(headers), 22)

    def test_create_data_tvv_template_has_46_columns(self) -> None:
        output_path = Path(self.temporary_directory.name) / "Mau_Data_TVV.xlsx"

        create_data_tvv_template(output_path)

        self.assertTrue(output_path.is_file())
        workbook = load_workbook(output_path)
        try:
            self.assertIn("Data_TVV", workbook.sheetnames)
            self.assertIn("HuongDan", workbook.sheetnames)
            headers = [
                cell.value
                for cell in next(workbook["Data_TVV"].iter_rows(max_row=1))
            ]
        finally:
            workbook.close()
        self.assertEqual(headers, list(DATA_TVV_TEMPLATE_HEADERS))
        self.assertEqual(headers[:22], list(DATA_TVV_HEADERS))
        self.assertEqual(headers[22:36], list(COMMISSION_EXPORT_HEADERS))
        self.assertEqual(headers[36:46], list(COMMISSION_RULE_EXPORT_HEADERS))
        self.assertEqual(len(headers), 46)

    def test_template_text_columns_are_formatted_as_text(self) -> None:
        output_path = Path(self.temporary_directory.name) / "Mau_Data_TVV.xlsx"

        create_data_tvv_template(output_path)

        workbook = load_workbook(output_path)
        try:
            sheet = workbook["Data_TVV"]
            header_to_column = {
                sheet.cell(row=1, column=column_index).value: column_index
                for column_index in range(1, len(DATA_TVV_TEMPLATE_HEADERS) + 1)
            }
            for header in DATA_TVV_TEXT_COLUMNS:
                column_index = header_to_column[header]
                self.assertEqual(
                    sheet.cell(row=2, column=column_index).number_format,
                    "@",
                )
                self.assertEqual(
                    sheet.cell(row=20, column=column_index).number_format,
                    "@",
                )
        finally:
            workbook.close()

    def test_import_created_template_no_crash(self) -> None:
        output_path = Path(self.temporary_directory.name) / "Mau_Data_TVV.xlsx"
        create_data_tvv_template(output_path)

        imported_count = self.repository.import_data_tvv(output_path)

        self.assertEqual(imported_count, 1)
        self.assertIsNotNone(self.repository.get_group("TVV001"))
        self.assertEqual(
            self.repository.get_or_create_commission_rate("TVV001").total_secured(),
            100.0,
        )

    def test_import_old_22_column_template_still_works(self) -> None:
        source_path = Path(self.temporary_directory.name) / "Old_Data_TVV.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data_TVV"
        sheet.append(DATA_TVV_HEADERS)
        sheet.append(self._data_tvv_row(1, "T001", "Tổ 001"))
        workbook.save(source_path)
        workbook.close()

        imported_count = self.repository.import_data_tvv(source_path)

        self.assertEqual(imported_count, 1)
        self.assertEqual(
            self.repository.get_or_create_commission_rate("T001").no_secured_to_truong,
            80.0,
        )

    def test_import_new_46_column_template_with_commission(self) -> None:
        source_path = Path(self.temporary_directory.name) / "New_Data_TVV.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data_TVV"
        sheet.append(DATA_TVV_TEMPLATE_HEADERS)
        sheet.append(
            self._data_tvv_row(1, "T001", "Tổ 001")
            + [4, 1, 70, 15, 5, 5, 5, 100, 85, 10, 2, 2, 1, 100]
            + [80, 88, 40, 88, 96, 85, 96, 100, 3, 0]
        )
        workbook.save(source_path)
        workbook.close()

        imported_count = self.repository.import_data_tvv(
            source_path,
            update_commission_rules=True,
        )

        self.assertEqual(imported_count, 1)
        rate = self.repository.get_or_create_commission_rate("T001")
        self.assertEqual(rate.base_no_secured_rate, 4.0)
        self.assertEqual(rate.base_secured_rate, 1.0)
        self.assertEqual(rate.no_secured_to_truong, 70.0)
        self.assertEqual(rate.secured_to_truong, 85.0)
        settings = self.repository.get_commission_rule_settings()
        self.assertEqual(settings.interest_min_1, 80.0)
        self.assertEqual(settings.bad_debt_threshold, 3.0)

    def test_invalid_commission_total_in_template_is_reported(self) -> None:
        source_path = Path(self.temporary_directory.name) / "Invalid_Data_TVV.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data_TVV"
        sheet.append(DATA_TVV_TEMPLATE_HEADERS)
        sheet.append(
            self._data_tvv_row(1, "T001", "Tổ 001")
            + [3, 2, 50, 10, 10, 10, 10, 90, 90, 10, 0, 0, 0, 100]
            + [85, 90, 50, 90, 95, 90, 95, 100, 2, 0]
        )
        workbook.save(source_path)
        workbook.close()

        with self.assertRaises(CreditGroupRepositoryError):
            self.repository.import_data_tvv(source_path)

    def test_export_with_commission_has_46_columns(self) -> None:
        self.repository.save_group(
            CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001")
        )
        output_path = Path(self.temporary_directory.name) / "Data_TVV_46.xlsx"

        self.repository.export_data_tvv(output_path, include_commission=True)

        workbook = load_workbook(output_path, read_only=True)
        try:
            headers = [
                cell.value
                for cell in next(workbook["Data_TVV"].iter_rows(max_row=1))
            ]
        finally:
            workbook.close()
        self.assertEqual(headers, list(DATA_TVV_TEMPLATE_HEADERS))
        self.assertEqual(len(headers), 46)

    def test_commission_rule_settings_have_vba_like_defaults(self) -> None:
        settings = self.repository.get_commission_rule_settings()

        self.assertEqual(settings.secured_base_rate, 2.0)
        self.assertEqual(settings.no_secured_base_rate, 3.0)
        self.assertEqual(settings.interest_min_1, 85.0)
        self.assertEqual(settings.interest_max_1, 90.0)
        self.assertEqual(settings.interest_pay_1, 50.0)
        self.assertEqual(settings.interest_min_2, 90.0)
        self.assertEqual(settings.interest_max_2, 95.0)
        self.assertEqual(settings.interest_pay_2, 90.0)
        self.assertEqual(settings.interest_min_3, 95.0)
        self.assertEqual(settings.interest_pay_3, 100.0)
        self.assertEqual(settings.bad_debt_threshold, 2.0)
        self.assertEqual(settings.bad_debt_pay, 0.0)

    def test_commission_rule_settings_are_saved_and_validated(self) -> None:
        self.repository.save_commission_rule_settings(
            CreditCommissionRuleSettings(
                secured_base_rate=1.5,
                no_secured_base_rate=2.5,
                interest_min_1=80,
                interest_max_1=88,
                interest_pay_1=40,
                interest_min_2=88,
                interest_max_2=96,
                interest_pay_2=85,
                interest_min_3=96,
                interest_pay_3=100,
                bad_debt_threshold=3,
                bad_debt_pay=0,
            )
        )

        settings = self.repository.get_commission_rule_settings()
        self.assertEqual(settings.secured_base_rate, 1.5)
        self.assertEqual(settings.no_secured_base_rate, 2.5)
        self.assertEqual(settings.interest_min_1, 80.0)
        self.assertEqual(settings.interest_pay_2, 85.0)
        self.assertEqual(settings.bad_debt_threshold, 3.0)

        with self.assertRaises(CreditGroupRepositoryError):
            self.repository.save_commission_rule_settings(
                CreditCommissionRuleSettings(
                    interest_min_1=90,
                    interest_max_1=85,
                    interest_pay_1=50,
                )
            )

    def test_payment_request_reads_tong_hop_theo_to(self) -> None:
        output_path = self._create_sample_interest_report()

        report_data = load_payment_report_data(output_path)

        self.assertEqual(len(report_data.rows), 1)
        self.assertEqual(report_data.rows[0].ma_to, "T001")
        self.assertEqual(report_data.rows[0].lai_co_bd, 1_000_000)
        self.assertEqual(report_data.period_from, "01/01/2026")
        self.assertEqual(report_data.period_to, "31/03/2026")

    def test_payment_template_default_path_exists_and_scans_placeholders(self) -> None:
        template_path = default_payment_template_path()

        placeholders = scan_word_placeholders(template_path)

        self.assertTrue(template_path.is_file())
        self.assertIn("[MaToVayVon]", placeholders)
        self.assertIn("[TongHH]", placeholders)
        self.assertIn("[HHKoBD]", placeholders)

    def test_word_replacement_preserves_ignorable_namespace_prefixes(self) -> None:
        template_path = default_payment_template_path()
        output_path = Path(self.temporary_directory.name) / "namespace_check.docx"

        replace_word_placeholders(template_path, output_path, {"MaToVayVon": "T001"})

        with ZipFile(output_path) as archive:
            xml = archive.read("word/document.xml").decode("utf-8")
        root_tag = xml[xml.find("<w:") : xml.find(">", xml.find("<w:"))]
        match = re.search(r'mc:Ignorable="([^"]+)"', xml)
        self.assertIsNotNone(match)
        assert match is not None
        for prefix in match.group(1).split():
            self.assertIn(f"xmlns:{prefix}=", root_tag)

    def test_word_placeholder_replacement_handles_split_runs(self) -> None:
        template_path = Path(self.temporary_directory.name) / "template.docx"
        output_path = Path(self.temporary_directory.name) / "output.docx"
        self._write_minimal_docx_template(
            template_path,
            paragraphs=("[MaToVayVon]", "[HH", "KoBD]"),
        )

        unmapped = replace_word_placeholders(
            template_path,
            output_path,
            {"MaToVayVon": "T001", "HHKoBD": "15.000"},
        )
        text = extract_docx_text(output_path)

        self.assertEqual(unmapped, set())
        self.assertIn("T001", text)
        self.assertIn("15.000", text)
        self.assertNotIn("[HHKoBD]", text)

    def test_payment_request_builds_vba_context(self) -> None:
        output_path = self._create_sample_interest_report()
        summary = load_payment_report_data(output_path).rows[0]

        context, warnings = build_payment_context(
            summary,
            self.repository,
            period_from="01/01/2026",
            period_to="31/03/2026",
            today=date(2026, 7, 16),
        )

        self.assertEqual(warnings, ())
        self.assertEqual(context["MaToVayVon"], "T001")
        self.assertEqual(context["KyThuLai"], "từ ngày 01/01/2026 đến ngày 31/03/2026")
        self.assertEqual(context["TL_KoBD"], "3%")
        self.assertEqual(context["TL_CoBD"], "2%")
        self.assertEqual(context["TongLaiThu"], "1.500.000")
        self.assertEqual(context["LaiKoBD"], "500.000")
        self.assertEqual(context["HHKoBD"], "15.000")
        self.assertEqual(context["HHCoBD"], "20.000")
        self.assertIn("đồng", context["TongHH_BC"])

    def test_payment_request_tl_kobd_uses_group_base_rate(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                base_no_secured_rate=2.5,
                base_secured_rate=2,
            )
        )

        context, warnings = build_payment_context(
            self._payment_summary("T001"),
            self.repository,
        )

        self.assertEqual(warnings, ())
        self.assertEqual(context["TL_KoBD"], "2,5%")
        self.assertNotEqual(context["TL_KoBD"], "3%")

    def test_payment_request_tl_cobd_uses_group_base_rate(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                base_no_secured_rate=3,
                base_secured_rate=1.8,
            )
        )

        context, warnings = build_payment_context(
            self._payment_summary("T001"),
            self.repository,
        )

        self.assertEqual(warnings, ())
        self.assertEqual(context["TL_CoBD"], "1,8%")
        self.assertNotEqual(context["TL_CoBD"], "2%")

    def test_payment_request_each_group_uses_own_rate(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_group(CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002"))
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                base_no_secured_rate=3,
                base_secured_rate=2,
            )
        )
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T002",
                base_no_secured_rate=2.5,
                base_secured_rate=1.8,
            )
        )

        context_1, _ = build_payment_context(self._payment_summary("T001"), self.repository)
        context_2, _ = build_payment_context(self._payment_summary("T002"), self.repository)

        self.assertEqual(context_1["TL_KoBD"], "3%")
        self.assertEqual(context_1["TL_CoBD"], "2%")
        self.assertEqual(context_2["TL_KoBD"], "2,5%")
        self.assertEqual(context_2["TL_CoBD"], "1,8%")

    def test_payment_request_rate_fallback_default(self) -> None:
        context, warnings = build_payment_context(
            self._payment_summary("UNKNOWN"),
            self.repository,
        )

        self.assertEqual(context["TL_KoBD"], "3%")
        self.assertEqual(context["TL_CoBD"], "2%")
        self.assertTrue(any("mặc định 3%/2%" in warning for warning in warnings))

    def test_payment_request_rate_fallback_from_summary_sheet(self) -> None:
        context, warnings = build_payment_context(
            self._payment_summary(
                "UNKNOWN",
                base_no_secured_rate=2.5,
                base_secured_rate=1.8,
            ),
            self.repository,
        )

        self.assertEqual(context["TL_KoBD"], "2,5%")
        self.assertEqual(context["TL_CoBD"], "1,8%")
        self.assertTrue(any("TongHopTheoTo" in warning for warning in warnings))

    def test_payment_request_does_not_use_distribution_rate_for_tl_kobd(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                base_no_secured_rate=2.5,
                no_secured_to_truong=80,
            )
        )

        context, _ = build_payment_context(self._payment_summary("T001"), self.repository)

        self.assertEqual(context["TL_KoBD"], "2,5%")
        self.assertNotEqual(context["TL_KoBD"], "80%")

    def test_payment_request_does_not_use_condition_pay_rate_for_tl_cobd(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                base_secured_rate=1.8,
            )
        )
        self.repository.save_commission_rule_settings(
            CreditCommissionRuleSettings(interest_pay_3=50)
        )

        context, _ = build_payment_context(self._payment_summary("T001"), self.repository)

        self.assertEqual(context["TL_CoBD"], "1,8%")
        self.assertNotEqual(context["TL_CoBD"], "50%")

    def test_payment_request_docx_replaces_tl_placeholders(self) -> None:
        report_path = self._create_sample_interest_report()
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                base_no_secured_rate=2.5,
                base_secured_rate=1.8,
                no_secured_to_truong=70,
                no_secured_cap_xa=10,
                no_secured_cap_huyen=10,
                no_secured_cap_tinh=5,
                no_secured_cap_tw=5,
                secured_to_truong=80,
                secured_cap_xa=20,
            )
        )
        template_path = Path(self.temporary_directory.name) / "payment_template_tl.docx"
        output_folder = Path(self.temporary_directory.name) / "KetQuaTL"
        self._write_minimal_docx_template(
            template_path,
            paragraphs=("[MaToVayVon]", "[TL_KoBD]", "[TL_CoBD]"),
        )

        result = export_payment_requests(
            report_path=report_path,
            template_path=template_path,
            output_folder=output_folder,
            repository=self.repository,
            ma_to="T001",
            today=date(2026, 7, 16),
        )
        text = extract_docx_text(result.output_paths[0])

        self.assertNotIn("[TL_KoBD]", text)
        self.assertNotIn("[TL_CoBD]", text)
        self.assertIn("2,5%", text)
        self.assertIn("1,8%", text)
        self.assertTrue(any("TL_KoBD=2,5%" in message for message in result.logs))

    def test_amount_to_vietnamese_words(self) -> None:
        self.assertEqual(amount_to_vietnamese_words(0), "không đồng")
        self.assertEqual(amount_to_vietnamese_words(1000), "một nghìn đồng")
        self.assertEqual(
            amount_to_vietnamese_words(2_417_292),
            "hai triệu bốn trăm mười bảy nghìn hai trăm chín mươi hai đồng",
        )
        self.assertEqual(
            amount_to_vietnamese_words(2_127_726),
            "hai triệu một trăm hai mươi bảy nghìn bảy trăm hai mươi sáu đồng",
        )

    def test_payment_request_exports_single_docx(self) -> None:
        report_path = self._create_sample_interest_report()
        template_path = Path(self.temporary_directory.name) / "payment_template.docx"
        output_folder = Path(self.temporary_directory.name) / "KetQua"
        self._write_minimal_docx_template(
            template_path,
            paragraphs=("[MaToVayVon]", "[TongHH]", "[HHKoBD]", "[TL_KoBD]"),
        )

        result = export_payment_requests(
            report_path=report_path,
            template_path=template_path,
            output_folder=output_folder,
            repository=self.repository,
            ma_to="T001",
            today=date(2026, 7, 16),
        )
        text = extract_docx_text(result.output_paths[0])

        self.assertEqual(len(result.output_paths), 1)
        self.assertTrue(result.output_paths[0].is_file())
        self.assertIn("T001", text)
        self.assertIn("35.000", text)
        self.assertIn("15.000", text)
        self.assertIn("3%", text)

    def test_payment_request_exports_all_docx(self) -> None:
        report_path = self._create_sample_interest_report()
        template_path = Path(self.temporary_directory.name) / "payment_template_all.docx"
        output_folder = Path(self.temporary_directory.name) / "KetQuaAll"
        self._write_minimal_docx_template(
            template_path,
            paragraphs=("[MaToVayVon]", "[TongHH]"),
        )

        result = export_payment_requests(
            report_path=report_path,
            template_path=template_path,
            output_folder=output_folder,
            repository=self.repository,
            export_all=True,
            today=date(2026, 7, 16),
        )

        self.assertEqual(len(result.output_paths), 1)
        self.assertTrue(result.output_paths[0].name.startswith("T001_"))
        self.assertFalse(result.output_paths[0].name.startswith("DeNghiThanhToan_"))

    def test_checkable_combo_select_all(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._sample_checkable_combo()

        combo.select_all()

        self.assertEqual(combo.get_selected_values(), ["T001", "T002", "T003"])
        self.assertEqual(combo.lineEdit().text(), "Tất cả tổ vay vốn")

    def test_checkable_combo_deselect_all(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._sample_checkable_combo()

        combo.select_all()
        combo.deselect_all()

        self.assertEqual(combo.get_selected_values(), [])
        self.assertEqual(combo.lineEdit().text(), "Chọn tổ vay vốn...")

    def test_checkable_combo_select_all_item_not_returned(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._sample_checkable_combo()

        combo._toggle_item(combo.model().index(0, 0))

        self.assertEqual(combo.get_selected_values(), ["T001", "T002", "T003"])
        self.assertNotIn("Tất cả", combo.get_selected_values())

    def test_checkable_combo_display_text(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._sample_checkable_combo()

        self.assertEqual(combo.lineEdit().text(), "Chọn tổ vay vốn...")
        combo.set_selected_values(["T001"])
        self.assertEqual(combo.lineEdit().text(), "T001 - Tổ 001")
        combo.set_selected_values(["T001", "T002"])
        self.assertEqual(combo.lineEdit().text(), "Đã chọn 2 tổ")
        combo.select_all()
        self.assertEqual(combo.lineEdit().text(), "Tất cả tổ vay vốn")

    def test_checkable_combo_does_not_close_on_item_check(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._sample_checkable_combo()

        combo._toggle_item(combo.model().index(1, 0))
        combo.hidePopup()

        self.assertFalse(combo._skip_next_hide)
        self.assertEqual(combo.get_selected_values(), ["T001"])

    def test_checkable_combo_opens_when_clicking_text_area(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._sample_checkable_combo()
        event = QEvent(QEvent.Type.MouseButtonRelease)

        with patch.object(combo, "showPopup") as show_popup:
            handled = combo.eventFilter(combo.lineEdit(), event)
            app.processEvents()

        self.assertTrue(handled)
        show_popup.assert_called_once()

    def test_interest_report_all_groups_default_checked(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        window = InterestReportWindow(database_path=self.database_path)

        self.assertTrue(window.all_groups_check.isChecked())
        self.assertFalse(window.group_combo.isEnabled())

    def test_interest_report_enable_group_multiselect_when_unchecked(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_group(CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002"))
        window = InterestReportWindow(database_path=self.database_path)

        window.all_groups_check.setChecked(False)
        window.group_combo.set_checked_data(["T001", "T002"])

        self.assertTrue(window.group_combo.isEnabled())
        self.assertIsInstance(window.group_combo, CheckableComboBox)
        self.assertEqual(window.group_combo.checked_data(), ["T001", "T002"])

    def test_interest_report_open_output_folder_uses_parent_folder(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        output_path = Path(self.temporary_directory.name) / "KetQua" / "BangKe.xlsx"
        window = InterestReportWindow(database_path=self.database_path)
        try:
            window.output_file_edit.setText(str(output_path))
            opened_urls: list[str] = []

            def fake_open_url(url: QUrl) -> bool:
                opened_urls.append(url.toLocalFile())
                return True

            with patch.object(QDesktopServices, "openUrl", side_effect=fake_open_url):
                window._open_output_folder()

            self.assertEqual(Path(opened_urls[0]), output_path.parent)
            self.assertTrue(output_path.parent.is_dir())
        finally:
            window.close()

    def test_interest_report_filters_selected_groups(self) -> None:
        output_path = self._create_multi_group_interest_report(
            selected_group_codes=("T001", "T003")
        )
        workbook = load_workbook(output_path, data_only=True)
        try:
            self.assertIn("T001", workbook.sheetnames)
            self.assertIn("T003", workbook.sheetnames)
            self.assertNotIn("T002", workbook.sheetnames)
            records = self._sheet_records(workbook["TongHopTheoTo"])
            self.assertEqual([record["MaTo"] for record in records], ["T001", "T003"])
        finally:
            workbook.close()

    def test_payment_request_all_groups_default_checked(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = PaymentRequestWindow(database_path=self.database_path)

        self.assertTrue(window.export_all_check.isChecked())
        self.assertFalse(window.group_combo.isEnabled())

    def test_payment_request_enable_multiselect_when_unchecked(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = PaymentRequestWindow(database_path=self.database_path)
        report_path = self._create_multi_group_interest_report()
        window.report_file_edit.setText(str(report_path))
        self.assertTrue(window._check_data(show_message=False))

        window.export_all_check.setChecked(False)
        window.group_combo.set_checked_data(["T001", "T002"])

        self.assertTrue(window.group_combo.isEnabled())
        self.assertIsInstance(window.group_combo, CheckableComboBox)
        self.assertEqual(window.group_combo.checked_data(), ["T001", "T002"])

    def test_payment_request_exports_selected_groups_only(self) -> None:
        report_path = self._create_multi_group_interest_report()
        template_path = Path(self.temporary_directory.name) / "payment_template_selected.docx"
        output_folder = Path(self.temporary_directory.name) / "KetQuaSelected"
        self._write_minimal_docx_template(
            template_path,
            paragraphs=("[MaToVayVon]", "[TongHH]"),
        )

        result = export_payment_requests(
            report_path=report_path,
            template_path=template_path,
            output_folder=output_folder,
            repository=self.repository,
            selected_group_codes=("T001", "T003"),
        )

        self.assertEqual(len(result.output_paths), 2)
        self.assertEqual(
            sorted(path.name.split("_")[0] for path in result.output_paths),
            ["T001", "T003"],
        )

    def test_payment_request_skips_ineligible_groups(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        report_path = self._write_payment_report_workbook(
            [
                self._payment_summary("T001"),
                self._payment_summary("T002", ty_le_chi=0, tong_hh=0),
                self._payment_summary("T003", canh_bao="Không đủ điều kiện chi"),
            ]
        )
        template_path = Path(self.temporary_directory.name) / "payment_template_skip.docx"
        output_folder = Path(self.temporary_directory.name) / "KetQuaSkip"
        self._write_minimal_docx_template(
            template_path,
            paragraphs=("[MaToVayVon]", "[TongHH]"),
        )

        result = export_payment_requests(
            report_path=report_path,
            template_path=template_path,
            output_folder=output_folder,
            repository=self.repository,
            export_all=True,
            today=date(2026, 7, 16),
        )

        self.assertEqual(len(result.output_paths), 1)
        self.assertEqual(result.output_paths[0].name, "T001_20260716.docx")
        self.assertFalse((output_folder / "T002_20260716.docx").exists())
        self.assertFalse((output_folder / "T003_20260716.docx").exists())
        self.assertTrue(any("Bỏ qua tổ T002" in message for message in result.logs))
        self.assertTrue(any("Bỏ qua tổ T003" in message for message in result.logs))
        self.assertTrue(any("đã tạo file: 1" in message for message in result.logs))

    def test_payment_request_selected_all_ineligible_raises_clear_message(self) -> None:
        report_path = self._write_payment_report_workbook(
            [
                self._payment_summary("T002", ty_le_chi=0, tong_hh=0),
                self._payment_summary("T003", canh_bao="Không đủ điều kiện chi"),
            ]
        )
        template_path = Path(self.temporary_directory.name) / "payment_template_none.docx"
        output_folder = Path(self.temporary_directory.name) / "KetQuaNone"
        self._write_minimal_docx_template(template_path, paragraphs=("[MaToVayVon]",))

        with self.assertRaisesRegex(
            PaymentRequestError,
            "Không có tổ nào đủ điều kiện chi hoa hồng",
        ):
            export_payment_requests(
                report_path=report_path,
                template_path=template_path,
                output_folder=output_folder,
                repository=self.repository,
                selected_group_codes=("T002", "T003"),
            )

        self.assertFalse(output_folder.exists())

    def test_payment_request_eligibility_summary_counts_rows(self) -> None:
        rows = (
            self._payment_summary("T001"),
            self._payment_summary("T002", ty_le_chi=0),
            self._payment_summary("T003", tong_hh=0),
            self._payment_summary("T004", canh_bao="Không đủ điều kiện chi"),
        )

        summary = analyze_payment_rows(rows)

        self.assertEqual(summary.total, 4)
        self.assertEqual([row.ma_to for row in summary.eligible], ["T001"])
        self.assertEqual([row.ma_to for row in summary.ineligible], ["T002", "T003", "T004"])
        self.assertEqual(payment_ineligible_reason(rows[1]), "TyLeChi = 0%")
        self.assertEqual(payment_ineligible_reason(rows[2]), "TongHH = 0")

    def test_payment_request_check_data_logs_eligible_and_ineligible_counts(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        report_path = self._write_payment_report_workbook(
            [
                self._payment_summary("T001"),
                self._payment_summary("T002", ty_le_chi=0, tong_hh=0),
                self._payment_summary("T003", canh_bao="Không đủ điều kiện chi"),
            ]
        )
        window = PaymentRequestWindow(database_path=self.database_path)
        try:
            window.report_file_edit.setText(str(report_path))
            self.assertTrue(window._check_data(show_message=False))
            log_text = window.log_edit.toPlainText()

            self.assertIn("- Tổng số tổ: 3", log_text)
            self.assertIn("- Đủ điều kiện chi: 1", log_text)
            self.assertIn("- Không đủ điều kiện chi: 2", log_text)
            self.assertIn("T002 - Tổ T002", log_text)
            self.assertIn("T003 - Tổ T003", log_text)
        finally:
            window.close()

    def test_payment_request_filename_without_prefix(self) -> None:
        report_path = self._create_multi_group_interest_report(group_codes=("5491LLG202100001",))
        template_path = Path(self.temporary_directory.name) / "payment_template_filename.docx"
        output_folder = Path(self.temporary_directory.name) / "KetQuaFilename"
        self._write_minimal_docx_template(
            template_path,
            paragraphs=("[MaToVayVon]", "[TongHH]"),
        )

        result = export_payment_requests(
            report_path=report_path,
            template_path=template_path,
            output_folder=output_folder,
            repository=self.repository,
            selected_group_codes=("5491LLG202100001",),
            today=date(2026, 7, 16),
        )

        self.assertEqual(result.output_paths[0].name, "5491LLG202100001_20260716.docx")

    def test_tovayvon_help_window_tabs(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = ToVayVonHelpWindow()
        try:
            self.assertEqual(window.windowTitle(), "Hướng dẫn tổ vay vốn - AgribankV3")
            self.assertEqual(window.tabs.count(), 3)
            self.assertEqual(window.tabs.tabText(0), "Hướng dẫn tạo bảng kê thu lãi")
            self.assertEqual(window.tabs.tabText(1), "Đề nghị thanh toán")
            self.assertEqual(window.tabs.tabText(2), "Đối chiếu tổ vay vốn")
        finally:
            window.close()

    def test_tovayvon_help_window_content(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = ToVayVonHelpWindow()
        try:
            step_cards = [
                widget
                for widget in window.findChildren(QFrame)
                if widget.objectName() == "GuideStepCard"
            ]
            self.assertEqual(len(step_cards), 9)
            all_text = "\n".join(label.text() for label in window.findChildren(QLabel))
            self.assertIn("Tổ vay vốn -> Quản lý tổ vay vốn", all_text)
            self.assertIn("Tổ vay vốn -> Đề nghị thanh toán hoa hồng tổ vay vốn", all_text)
            self.assertIn("Tổ vay vốn -> Đối chiếu dư nợ theo tổ vay vốn", all_text)
            self.assertIn("5491_ln01_20260630.csv", all_text)
        finally:
            window.close()

    def test_tovayvon_help_window_not_placeholder(self) -> None:
        self.assertNotIn(TOVAYVON_HELP_TITLE, CREDIT_TOVAYVON_PLACEHOLDER_TITLES)

    def test_debt_reconciliation_required_columns(self) -> None:
        source_path = Path(self.temporary_directory.name) / "du_no_missing.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_missing.xlsx"
        self._write_debt_workbook(
            source_path,
            headers=["MaKH", "SoGiaiNgan", "DuNo"],
            rows=[{"MaKH": "KH001", "SoGiaiNgan": "GN001", "DuNo": 100_000}],
        )

        detection = detect_debt_columns(source_path, self.repository)

        self.assertIn("group_code", detection.missing_required)
        with self.assertRaisesRegex(DebtReconciliationError, "thiếu cột bắt buộc"):
            create_debt_reconciliation(
                DebtReconciliationRequest(
                    input_file=source_path,
                    output_path=output_path,
                    reconciliation_date=date(2026, 7, 16),
                ),
                self.repository,
            )

    def test_debt_reconciliation_reads_group_code_as_text(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="00123", ten_to="Tổ 00123"))
        source_path = Path(self.temporary_directory.name) / "du_no_text_code.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_text_code.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[
                {
                    "MaKH": "KH001",
                    "TenKH": "Khách 001",
                    "SoGiaiNgan": "GN001",
                    "MaTo": "'00123",
                    "DuNo": 100_000,
                }
            ],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            detail = self._sheet_records(workbook[DETAIL_SHEET_NAME])
            self.assertEqual(detail[0]["MaTo"], "00123")
        finally:
            workbook.close()

    def test_debt_reconciliation_summary_by_group(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        source_path = Path(self.temporary_directory.name) / "du_no_summary.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_summary.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[
                {"MaKH": "KH001", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000, "INTEREST_AMOUNT": 1_000},
                {"MaKH": "KH002", "SoGiaiNgan": "GN002", "MaTo": "T001", "DuNo": 200_000, "INTEREST_AMOUNT": 2_000},
            ],
        )

        result = create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            summary = self._sheet_records(workbook[SUMMARY_SHEET_NAME])
            self.assertEqual(result.group_count, 1)
            self.assertEqual(summary[0]["MaTo"], "T001")
            self.assertEqual(summary[0]["SoKhachHang"], 2)
            self.assertEqual(summary[0]["SoKhoanVay"], 2)
            self.assertEqual(summary[0]["TongDuNo"], 300_000)
            self.assertEqual(summary[0]["TongLai"], 3_000)
        finally:
            workbook.close()

    def test_debt_reconciliation_missing_group_sheet(self) -> None:
        source_path = Path(self.temporary_directory.name) / "du_no_missing_group.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_missing_group.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[{"MaKH": "KH001", "SoGiaiNgan": "GN001", "MaTo": "", "DuNo": 100_000}],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            records = self._sheet_records(workbook[MISSING_GROUP_SHEET_NAME])
            self.assertEqual(records[0]["MaKH"], "KH001")
            self.assertIn("Thiếu mã tổ", records[0]["GhiChu"])
        finally:
            workbook.close()

    def test_debt_reconciliation_unknown_group_sheet(self) -> None:
        source_path = Path(self.temporary_directory.name) / "du_no_unknown_group.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_unknown_group.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[{"MaKH": "KH001", "SoGiaiNgan": "GN001", "MaTo": "UNKNOWN", "DuNo": 100_000}],
        )

        result = create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            records = self._sheet_records(workbook[UNKNOWN_GROUP_SHEET_NAME])
            self.assertEqual(records[0]["MaTo"], "UNKNOWN")
            self.assertTrue(any("UNKNOWN" in warning for warning in result.warnings))
        finally:
            workbook.close()

    def test_debt_reconciliation_group_without_balance_sheet(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        self.repository.save_group(CreditGroup(stt=2, ma_to="T002", ten_to="Tổ 002"))
        source_path = Path(self.temporary_directory.name) / "du_no_without_balance.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_without_balance.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[{"MaKH": "KH001", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000}],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            records = self._sheet_records(workbook[GROUP_WITHOUT_BALANCE_SHEET_NAME])
            self.assertEqual([record["MaTo"] for record in records], ["T002"])
        finally:
            workbook.close()

    def test_debt_reconciliation_bad_debt_ratio(self) -> None:
        self.assertEqual(normalize_debt_group("Nhóm 5"), 5)
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        source_path = Path(self.temporary_directory.name) / "du_no_bad_debt.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_bad_debt.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[
                {"MaKH": "KH001", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000, "NhomNo": "1"},
                {"MaKH": "KH002", "SoGiaiNgan": "GN002", "MaTo": "T001", "DuNo": 50_000, "NhomNo": "3"},
                {"MaKH": "KH003", "SoGiaiNgan": "GN003", "MaTo": "T001", "DuNo": 50_000, "NhomNo": "5"},
            ],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            summary = self._sheet_records(workbook[SUMMARY_SHEET_NAME])
            self.assertEqual(summary[0]["DuNoXau"], 100_000)
            self.assertEqual(summary[0]["TyLeNoXau"], 0.5)
        finally:
            workbook.close()

    def test_debt_reconciliation_filters_selected_groups(self) -> None:
        for index, ma_to in enumerate(("T001", "T002", "T003"), start=1):
            self.repository.save_group(CreditGroup(stt=index, ma_to=ma_to, ten_to=f"Tổ {index}"))
        source_path = Path(self.temporary_directory.name) / "du_no_selected.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_selected.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[
                {"MaKH": "KH001", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000},
                {"MaKH": "KH002", "SoGiaiNgan": "GN002", "MaTo": "T002", "DuNo": 200_000},
                {"MaKH": "KH003", "SoGiaiNgan": "GN003", "MaTo": "T003", "DuNo": 300_000},
            ],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
                selected_group_codes=("T001", "T003"),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            summary = self._sheet_records(workbook[SUMMARY_SHEET_NAME])
            self.assertEqual([record["MaTo"] for record in summary], ["T001", "T003"])
            self.assertEqual(self._sheet_records(workbook[GROUP_WITHOUT_BALANCE_SHEET_NAME]), [])
        finally:
            workbook.close()

    def test_debt_reconciliation_export_sheets_exist(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        source_path = Path(self.temporary_directory.name) / "du_no_sheets.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_sheets.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[{"MaKH": "KH001", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000}],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames[0], "T001")
            for sheet_name in (
                SUMMARY_SHEET_NAME,
                DETAIL_SHEET_NAME,
                MISSING_GROUP_SHEET_NAME,
                UNKNOWN_GROUP_SHEET_NAME,
                GROUP_WITHOUT_BALANCE_SHEET_NAME,
                WARNING_SHEET_NAME,
            ):
                self.assertIn(sheet_name, workbook.sheetnames)
        finally:
            workbook.close()

    def test_debt_reconciliation_vba_style_headers(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        source_path = Path(self.temporary_directory.name) / "du_no_vba_header.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_vba_header.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[{"MaKH": "KH001", "TenKH": "Nguyễn Văn A", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000}],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 6, 30),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=False)
        try:
            sheet = workbook["T001"]
            self.assertEqual(sheet["K1"].value, "Mẫu số: 22/ĐCN-CN")
            self.assertEqual(sheet["A7"].value, "BẢNG ĐỐI CHIẾU DƯ NỢ")
            self.assertEqual(sheet["A8"].value, "đến ngày 30 tháng 6 năm 2026")
            self.assertEqual(sheet["A9"].value, "Tổ vay vốn: T001")
            self.assertEqual(
                [sheet.cell(11, column).value for column in range(1, 12)],
                [
                    "TT",
                    "Họ tên khách hàng",
                    "Mã khách hàng",
                    "Số hợp đồng tín dụng",
                    "Số liệu tại Agribank",
                    None,
                    "Số liệu tại hồ sơ lưu khách hàng",
                    None,
                    "Chênh lệch",
                    None,
                    "Chữ ký ngưi vay",
                ],
            )
            self.assertEqual(
                [sheet.cell(12, column).value for column in range(1, 12)],
                [None, None, None, None, "Gốc", "Lãi đã thu", "Gốc", "Lãi đã thu", "Gốc", "Lãi đã thu", None],
            )
        finally:
            workbook.close()

    def test_debt_reconciliation_sheet_name_cleaned_like_vba(self) -> None:
        self.assertEqual(clean_debt_reconciliation_sheet_name("'Bad/Name:*?[]'"), "BadName")
        self.assertEqual(clean_debt_reconciliation_sheet_name("History"), "History_")
        self.assertEqual(clean_debt_reconciliation_sheet_name(""), "Group")

    def test_debt_reconciliation_group_by_grpno_vba_style(self) -> None:
        for index, ma_to in enumerate(("T001", "T002"), start=1):
            self.repository.save_group(CreditGroup(stt=index, ma_to=ma_to, ten_to=f"Tổ {index}"))
        source_path = Path(self.temporary_directory.name) / "du_no_vba_groups.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_vba_groups.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[
                {"MaKH": "KH001", "TenKH": "Khách 1", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000},
                {"MaKH": "KH002", "TenKH": "Khách 2", "SoGiaiNgan": "GN002", "MaTo": "T002", "DuNo": 200_000},
            ],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 6, 30),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=False)
        try:
            self.assertIn("T001", workbook.sheetnames)
            self.assertIn("T002", workbook.sheetnames)
            self.assertEqual(workbook["T001"]["B13"].value, "Khách 1")
            self.assertEqual(workbook["T002"]["B13"].value, "Khách 2")
        finally:
            workbook.close()

    def test_debt_reconciliation_aggregate_by_customer_loan_vba_style(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        source_path = Path(self.temporary_directory.name) / "du_no_vba_aggregate.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_vba_aggregate.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[
                {
                    "MaKH": "007",
                    "TenKH": "Khách 1",
                    "SoGiaiNgan": "GN001",
                    "MaTo": "T001",
                    "DuNo": 100_000,
                    "INTEREST_AMOUNT": 1_000,
                },
                {
                    "MaKH": "007",
                    "TenKH": "Khách 1",
                    "SoGiaiNgan": "GN001",
                    "MaTo": "T001",
                    "DuNo": 200_000,
                    "INTEREST_AMOUNT": 2_000,
                },
            ],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 6, 30),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=False)
        try:
            sheet = workbook["T001"]
            self.assertEqual(sheet["C13"].value, "'007")
            self.assertEqual(sheet["E13"].value, 300_000)
            self.assertEqual(sheet["F13"].value, 3_000)
            self.assertEqual(sheet["B14"].value, "Tổng cộng")
        finally:
            workbook.close()

    def test_debt_reconciliation_vba_style_totals_and_format(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        source_path = Path(self.temporary_directory.name) / "du_no_vba_format.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_vba_format.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[
                {"MaKH": "KH001", "TenKH": "Khách 1", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000, "INTEREST_AMOUNT": 1_000},
                {"MaKH": "KH002", "TenKH": "Khách 2", "SoGiaiNgan": "GN002", "MaTo": "T001", "DuNo": 200_000, "INTEREST_AMOUNT": 2_000},
            ],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 6, 30),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=False)
        try:
            sheet = workbook["T001"]
            self.assertEqual(sheet["E15"].value, "=SUM(E13:E14)")
            self.assertEqual(sheet["F15"].value, "=SUM(F13:F14)")
            self.assertTrue(sheet["A11"].font.bold)
            self.assertEqual(sheet["E13"].number_format, "#,##0")
            self.assertEqual(sheet["F15"].number_format, "#,##0")
            self.assertEqual(sheet["A13"].border.left.style, "thin")
            self.assertEqual(sheet["A15"].border.bottom.style, "double")
            self.assertEqual(sheet.row_dimensions[13].height, 35)
        finally:
            workbook.close()

    def test_debt_reconciliation_excel_number_formats(self) -> None:
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        source_path = Path(self.temporary_directory.name) / "du_no_format.xlsx"
        output_path = Path(self.temporary_directory.name) / "doi_chieu_format.xlsx"
        self._write_debt_workbook(
            source_path,
            rows=[
                {"MaKH": "KH001", "SoGiaiNgan": "GN001", "MaTo": "T001", "DuNo": 100_000, "NhomNo": "3"},
            ],
        )

        create_debt_reconciliation(
            DebtReconciliationRequest(
                input_file=source_path,
                output_path=output_path,
                reconciliation_date=date(2026, 7, 16),
            ),
            self.repository,
        )

        workbook = load_workbook(output_path, data_only=False)
        try:
            summary = workbook[SUMMARY_SHEET_NAME]
            self.assertEqual(summary.cell(2, 9).number_format, "#,##0")
            self.assertEqual(summary.cell(2, 16).number_format, "0.00%")
            detail = workbook[DETAIL_SHEET_NAME]
            self.assertEqual(detail.cell(2, 7).number_format, "#,##0")
        finally:
            workbook.close()

    def test_debt_reconciliation_all_groups_default_checked(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self.repository.save_group(CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"))
        window = DebtReconciliationWindow(database_path=self.database_path)
        try:
            self.assertTrue(window.all_groups_check.isChecked())
            self.assertFalse(window.group_combo.isEnabled())
            window.all_groups_check.setChecked(False)
            self.assertTrue(window.group_combo.isEnabled())
            self.assertIsInstance(window.group_combo, CheckableComboBox)
        finally:
            window.close()

    def _create_multi_group_interest_report(
        self,
        *,
        group_codes: tuple[str, ...] = ("T001", "T002", "T003"),
        selected_group_codes: tuple[str, ...] = (),
    ) -> Path:
        for index, ma_to in enumerate(group_codes, start=1):
            self.repository.save_group(
                CreditGroup(
                    stt=index,
                    ma_to=ma_to,
                    ten_to=f"Tổ {index:03d}",
                    ten_to_truong=f"Tổ trưởng {index:03d}",
                )
            )
        interest_path = Path(self.temporary_directory.name) / (
            "thu_lai_multi_" + "_".join(group_codes) + ".xlsx"
        )
        debt_path = Path(self.temporary_directory.name) / (
            "du_no_multi_" + "_".join(group_codes) + ".xlsx"
        )
        output_path = Path(self.temporary_directory.name) / (
            "BangKeThuLaiTo_Multi_" + "_".join(group_codes) + ".xlsx"
        )
        rows = [
            {
                "MaKH": f"KH{index:03d}",
                "TenKH": f"Khách {index:03d}",
                "SoGiaiNgan": f"GN{index:03d}",
                "DuNo": 100_000_000 + index,
                "SoLaiDaThuTrongKy": 1_000_000 + index,
                "TyLeBaoBam": 1,
                "MaToVayVon": ma_to,
                "NhomNo": "1",
                "LaiTon": 0,
                "LaiSuat": 12,
                "NgayTraLaiCuoiCung": date(2026, 3, 31),
            }
            for index, ma_to in enumerate(group_codes, start=1)
        ]
        self._write_interest_workbook(interest_path, rows=rows)
        self._write_interest_workbook(debt_path, rows=rows)
        create_interest_report(
            InterestReportRequest(
                interest_file=interest_path,
                debt_file=debt_path,
                output_path=output_path,
                from_date=date(2026, 1, 1),
                to_date=date(2026, 3, 31),
                selected_group_codes=selected_group_codes,
            ),
            self.repository,
        )
        return output_path

    @staticmethod
    def _payment_summary(
        ma_to: str,
        *,
        base_no_secured_rate: float = 0.0,
        base_secured_rate: float = 0.0,
        ty_le_chi: float = 1,
        tong_hh: float = 35_000,
        canh_bao: str = "",
    ) -> PaymentSummaryRow:
        return PaymentSummaryRow(
            ma_to=ma_to,
            ten_to=f"Tổ {ma_to}",
            ten_to_truong="Tổ trưởng",
            so_dong=2,
            tong_du_no=100_000_000,
            lai_co_bd=1_000_000,
            lai_khong_bd=500_000,
            lai_ton=0,
            lai_phai_thu=1_500_000,
            ty_le_thu_lai=1,
            ty_le_no_xau=0,
            ty_le_chi=ty_le_chi,
            hh_co_bd=20_000,
            hh_khong_bd=15_000,
            tong_hh=tong_hh,
            hh_to_truong=26_500,
            hh_cap_xa=5_000,
            hh_cap_huyen=1_500,
            hh_cap_tinh=1_250,
            hh_tw=750,
            canh_bao=canh_bao,
            base_no_secured_rate=base_no_secured_rate,
            base_secured_rate=base_secured_rate,
        )

    def _write_payment_report_workbook(
        self,
        rows: list[PaymentSummaryRow],
    ) -> Path:
        path = Path(self.temporary_directory.name) / "payment_report.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "TongHopTheoTo"
        headers = [
            "MaTo",
            "TenTo",
            "TenToTruong",
            "SoDong",
            "TongDuNo",
            "LaiCoBD",
            "LaiKhongBD",
            "LaiTon",
            "LaiPhaiThu",
            "TyLeThuLai",
            "TyLeNoXau",
            "TyLeChi",
            "HHCoBD",
            "HHKhongBD",
            "TongHH",
            "HH_ToTruong",
            "HH_CapXa",
            "HH_CapHuyen",
            "HH_CapTinh",
            "HH_TW",
            "CanhBao",
        ]
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    row.ma_to,
                    row.ten_to,
                    row.ten_to_truong,
                    row.so_dong,
                    row.tong_du_no,
                    row.lai_co_bd,
                    row.lai_khong_bd,
                    row.lai_ton,
                    row.lai_phai_thu,
                    row.ty_le_thu_lai,
                    row.ty_le_no_xau,
                    row.ty_le_chi,
                    row.hh_co_bd,
                    row.hh_khong_bd,
                    row.tong_hh,
                    row.hh_to_truong,
                    row.hh_cap_xa,
                    row.hh_cap_huyen,
                    row.hh_cap_tinh,
                    row.hh_tw,
                    row.canh_bao,
                ]
            )
        detail = workbook.create_sheet("T001")
        detail["A6"] = "(Từ ngày 01/01/2026 đến ngày 31/03/2026)"
        workbook.save(path)
        workbook.close()
        return path

    def _create_sample_interest_report(self) -> Path:
        self.repository.save_group(
            CreditGroup(
                stt=1,
                ma_to="T001",
                ten_to="Tổ 001",
                ten_to_truong="Nguyễn Văn A",
            )
        )
        self.repository.save_commission_rate(
            CreditGroupCommissionRate(
                ma_to="T001",
                no_secured_to_truong=70,
                no_secured_cap_xa=10,
                no_secured_cap_huyen=10,
                no_secured_cap_tinh=5,
                no_secured_cap_tw=5,
                secured_to_truong=80,
                secured_cap_xa=20,
                secured_cap_huyen=0,
                secured_cap_tinh=0,
                secured_cap_tw=0,
            )
        )
        interest_path = Path(self.temporary_directory.name) / "thu_lai.xlsx"
        debt_path = Path(self.temporary_directory.name) / "du_no.xlsx"
        output_path = Path(self.temporary_directory.name) / "BangKeThuLaiTo.xlsx"
        rows = [
            {
                "MaKH": "KH001",
                "TenKH": "Nguyễn Văn A",
                "SoGiaiNgan": "GN001",
                "DuNo": 100_000_000,
                "SoLaiDaThuTrongKy": 1_000_000,
                "LaiSuat": 12,
                "NgayTraLaiCuoiCung": date(2026, 3, 31),
                "TyLeBaoBam": 1,
                "MaToVayVon": "T001",
                "NhomNo": "1",
                "LaiTon": 0,
            },
            {
                "MaKH": "KH002",
                "TenKH": "Trần Thị B",
                "SoGiaiNgan": "GN002",
                "DuNo": 50_000_000,
                "SoLaiDaThuTrongKy": 500_000,
                "LaiSuat": 12,
                "NgayTraLaiCuoiCung": date(2026, 3, 31),
                "TyLeBaoBam": 0,
                "MaToVayVon": "T001",
                "NhomNo": "1",
                "LaiTon": 0,
            },
            {
                "MaKH": "KH003",
                "TenKH": "Lê Văn C",
                "SoGiaiNgan": "GN003",
                "DuNo": 10_000_000,
                "SoLaiDaThuTrongKy": 100_000,
                "LaiSuat": 12,
                "NgayTraLaiCuoiCung": date(2026, 3, 31),
                "TyLeBaoBam": 0,
                "MaToVayVon": "UNKNOWN",
                "NhomNo": "1",
                "LaiTon": 0,
            },
        ]
        self._write_interest_workbook(interest_path, rows=rows)
        self._write_interest_workbook(debt_path, rows=rows)
        result = create_interest_report(
            InterestReportRequest(
                interest_file=interest_path,
                debt_file=debt_path,
                output_path=output_path,
                from_date=date(2026, 1, 1),
                to_date=date(2026, 3, 31),
            ),
            self.repository,
        )
        self.assertEqual(result.group_count, 1)
        self.assertTrue(any("UNKNOWN" in warning for warning in result.warnings))
        return output_path

    def _create_period_source_interest_report(
        self,
        *,
        period_from: date | None = None,
        period_to: date | None = None,
    ) -> Path:
        self.repository.save_group(
            CreditGroup(
                stt=1,
                ma_to="T001",
                ten_to="Tổ 001",
                ten_to_truong="Nguyễn Văn A",
            )
        )
        interest_path = Path(self.temporary_directory.name) / "sktl_period.xlsx"
        debt_path = Path(self.temporary_directory.name) / "skck_period.xlsx"
        output_path = Path(self.temporary_directory.name) / "BangKeThuLaiTo_Period.xlsx"
        row = {
            "MaKH": "KH_PERIOD",
            "TenKH": "Khách kỳ thu lãi",
            "SoGiaiNgan": "GN_PERIOD",
            "DuNo": 100_000_000,
            "SoLaiDaThuTrongKy": 986_301.3698630137,
            "TyLeBaoBam": 1,
            "MaToVayVon": "T001",
            "NhomNo": "1",
            "LaiSuat": 12,
            "NgayTraLaiCuoiCung": date(2026, 6, 30),
        }
        if period_from is not None:
            row["ThuLaiTuNgay"] = period_from
        if period_to is not None:
            row["ThuLaiDenNgay"] = period_to
        self._write_interest_workbook(interest_path, rows=[row])
        self._write_interest_workbook(debt_path, rows=[row])
        create_interest_report(
            InterestReportRequest(
                interest_file=interest_path,
                debt_file=debt_path,
                output_path=output_path,
                from_date=date(2026, 4, 1),
                to_date=date(2026, 6, 30),
            ),
            self.repository,
        )
        return output_path

    @staticmethod
    def _commission_summary(
        *,
        lai_khong_bd: float,
        lai_co_bd: float,
        rate: CreditGroupCommissionRate | None = None,
    ) -> InterestGroupSummary:
        rows: list[InterestRow] = []
        if lai_khong_bd:
            rows.append(
                InterestRow(
                    ma_to="T001",
                    ma_kh="KH_KHONG_BD",
                    ten_kh="Không BĐ",
                    so_giai_ngan="GN1",
                    du_no=0,
                    ty_le_bao_bam=0,
                    so_lai_da_thu=lai_khong_bd,
                )
            )
        if lai_co_bd:
            rows.append(
                InterestRow(
                    ma_to="T001",
                    ma_kh="KH_CO_BD",
                    ten_kh="Có BĐTS",
                    so_giai_ngan="GN2",
                    du_no=0,
                    ty_le_bao_bam=100,
                    so_lai_da_thu=lai_co_bd,
                )
            )
        return InterestGroupSummary(
            group=CreditGroup(stt=1, ma_to="T001", ten_to="Tổ 001"),
            rate=rate or CreditGroupCommissionRate(ma_to="T001"),
            rows=rows,
        )

    def _create_sktl_only_interest_report(
        self,
        *,
        rate: CreditGroupCommissionRate | None = None,
        group_rule: CreditGroupCommissionRule | None = None,
    ):
        self.repository.save_group(
            CreditGroup(
                stt=1,
                ma_to="T001",
                ten_to="Tổ 001",
                ten_to_truong="Nguyễn Văn A",
            )
        )
        if rate is not None:
            self.repository.save_commission_rate(rate)
        if group_rule is not None:
            self.repository.save_group_commission_rule(group_rule)
        interest_path = Path(self.temporary_directory.name) / "sktl.xlsx"
        debt_path = Path(self.temporary_directory.name) / "skck.xlsx"
        output_path = Path(self.temporary_directory.name) / "BangKeThuLaiTo_SKTLOnly.xlsx"
        self._write_interest_workbook(
            interest_path,
            rows=[
                {
                    "MaKH": "KH_JOIN",
                    "TenKH": "Khách còn dư nợ",
                    "SoGiaiNgan": "GN_JOIN",
                    "SoLaiDaThuTrongKy": 1_000_000,
                    "TyLeBaoBam": 1,
                    "MaToVayVon": "T001",
                },
                {
                    "MaKH": "KH_ONLY",
                    "TenKH": "Khách tất toán",
                    "SoGiaiNgan": "GN_ONLY",
                    "SoLaiDaThuTrongKy": 200_000,
                    "TyLeBaoBam": 0,
                    "MaToVayVon": "T001",
                },
            ],
        )
        self._write_interest_workbook(
            debt_path,
            rows=[
                {
                    "MaKH": "KH_JOIN",
                    "TenKH": "Khách còn dư nợ",
                    "SoGiaiNgan": "GN_JOIN",
                    "DuNo": 100_000_000,
                    "SoLaiDaThuTrongKy": 0,
                    "TyLeBaoBam": 1,
                    "MaToVayVon": "T001",
                    "NhomNo": "1",
                    "LaiTon": 0,
                }
            ],
        )
        result = create_interest_report(
            InterestReportRequest(
                interest_file=interest_path,
                debt_file=debt_path,
                output_path=output_path,
                from_date=date(2026, 4, 1),
                to_date=date(2026, 6, 30),
            ),
            self.repository,
        )
        return output_path, result

    def _create_formula_interest_report(self) -> Path:
        self.repository.save_group(
            CreditGroup(
                stt=1,
                ma_to="T001",
                ten_to="Tổ 001",
                ten_to_truong="Nguyễn Văn A",
            )
        )
        interest_path = Path(self.temporary_directory.name) / "sktl_formula.xlsx"
        debt_path = Path(self.temporary_directory.name) / "skck_formula.xlsx"
        output_path = Path(self.temporary_directory.name) / "BangKeThuLaiTo_Formula.xlsx"
        self._write_interest_workbook(
            interest_path,
            rows=[
                {
                    "MaKH": "KH001",
                    "TenKH": "Nguyễn Văn A",
                    "SoGiaiNgan": "GN001",
                    "SoLaiDaThuTrongKy": 1_000_000,
                    "TyLeBaoBam": 1,
                    "MaToVayVon": "T001",
                }
            ],
        )
        self._write_workbook(
            debt_path,
            headers=[
                "MaKH",
                "TenKH",
                "SoGiaiNgan",
                "DuNo",
                "TyLeBaoBam",
                "MaToVayVon",
                "NhomNo",
                "LaiSuat",
                "NgayGiaiNgan",
                "NgayTraLaiCuoiCung",
            ],
            rows=[
                {
                    "MaKH": "KH001",
                    "TenKH": "Nguyễn Văn A",
                    "SoGiaiNgan": "GN001",
                    "DuNo": 100_000_000,
                    "TyLeBaoBam": 1,
                    "MaToVayVon": "T001",
                    "NhomNo": "1",
                    "LaiSuat": 12,
                    "NgayGiaiNgan": date(2026, 1, 1),
                    "NgayTraLaiCuoiCung": date(2026, 5, 31),
                }
            ],
        )
        create_interest_report(
            InterestReportRequest(
                interest_file=interest_path,
                debt_file=debt_path,
                output_path=output_path,
                from_date=date(2026, 4, 1),
                to_date=date(2026, 6, 30),
            ),
            self.repository,
        )
        return output_path

    def _write_debt_workbook(
        self,
        path: Path,
        *,
        rows: list[dict[str, object]],
        headers: list[str] | None = None,
    ) -> None:
        self._write_workbook(
            path,
            headers=headers
            or [
                "MaKH",
                "TenKH",
                "SoGiaiNgan",
                "MaTo",
                "DuNo",
                "NhomNo",
                "INTEREST_AMOUNT",
                "Ngày giải ngân",
                "Ngày đến hạn",
            ],
            rows=rows,
        )

    @staticmethod
    def _write_interest_workbook(
        path: Path,
        *,
        rows: list[dict[str, object]],
    ) -> None:
        headers = [
            "MaKH",
            "TenKH",
            "SoGiaiNgan",
            "DuNo",
            "SoLaiDaThuTrongKy",
            "SoGocDaThuTrongKy",
            "TyLeBaoBam",
            "MaToVayVon",
            "NhomNo",
            "LaiTon",
            "LaiSuat",
            "NgayTraGocCuoiCung",
            "NgayTraLaiCuoiCung",
            "ThuLaiTuNgay",
            "ThuLaiDenNgay",
        ]
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _write_workbook(
        path: Path,
        *,
        headers: list[str],
        rows: list[dict[str, object]],
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _write_minimal_docx_template(path: Path, *, paragraphs: tuple[str, ...]) -> None:
        namespace = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        body: list[str] = []
        index = 0
        while index < len(paragraphs):
            if (
                paragraphs[index].startswith("[")
                and index + 1 < len(paragraphs)
                and paragraphs[index + 1].endswith("]")
            ):
                body.append(
                    "<w:p><w:r><w:t>"
                    + paragraphs[index]
                    + "</w:t></w:r><w:r><w:t>"
                    + paragraphs[index + 1]
                    + "</w:t></w:r></w:p>"
                )
                index += 2
            else:
                body.append(f"<w:p><w:r><w:t>{paragraphs[index]}</w:t></w:r></w:p>")
                index += 1
        xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<w:document xmlns:w="{namespace}"><w:body>'
            + "".join(body)
            + "</w:body></w:document>"
        )
        with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr("word/document.xml", xml)

    @staticmethod
    def _sample_checkable_combo() -> CheckableComboBox:
        combo = CheckableComboBox()
        combo.add_check_item("T001 - Tổ 001", "T001")
        combo.add_check_item("T002 - Tổ 002", "T002")
        combo.add_check_item("T003 - Tổ 003", "T003")
        return combo

    @staticmethod
    def _sheet_records(sheet) -> list[dict[str, object]]:
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "") for value in rows[0]]
        return [
            {
                headers[index]: value
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            for row in rows[1:]
            if any(value is not None for value in row)
        ]

    @staticmethod
    def _main_bangke_sheet(workbook):
        for sheet_name in workbook.sheetnames:
            if sheet_name not in {"TongHopTheoTo", "CanhBao"}:
                return workbook[sheet_name]
        raise AssertionError("Không tìm thấy sheet bảng kê chính.")

    @classmethod
    def _bangke_records(cls, workbook) -> list[dict[str, object]]:
        sheet = cls._main_bangke_sheet(workbook)
        header_row = 12
        headers = [str(sheet.cell(header_row, column).value or "") for column in range(1, 15)]
        records: list[dict[str, object]] = []
        row_index = header_row + 1
        while row_index <= sheet.max_row:
            if sheet.cell(row_index, 1).value == "Tổng cộng":
                break
            if isinstance(sheet.cell(row_index, 1).value, int):
                records.append(
                    {
                        headers[column - 1]: sheet.cell(row_index, column).value
                        for column in range(1, 15)
                        if headers[column - 1]
                    }
                )
            row_index += 1
        return records

    @staticmethod
    def _commission_lines(sheet) -> dict[str, object]:
        result: dict[str, object] = {}
        section = ""
        for row in range(1, sheet.max_row + 1):
            label_b = str(sheet.cell(row, 2).value or "").strip()
            label_c = str(sheet.cell(row, 3).value or "").strip()
            value = sheet.cell(row, 10).value
            if label_b.startswith("*HOA HỒNG ĐỐI VỚI KHOẢN KHÔNG"):
                section = "Không BĐ"
                result[label_b] = value
            elif label_b.startswith("*HOA HỒNG ĐỐI VỚI KHOẢN CÓ"):
                section = "Có BĐTS"
                result[label_b] = value
            elif label_b.startswith("* HOA HỒNG THỰC NHẬN") and label_b not in result:
                result[label_b] = value
            elif label_c and section:
                result[f"{section} - {label_c}"] = value
        return result

    @staticmethod
    def _row_by_label(sheet, label_prefix: str) -> int:
        for row in range(1, sheet.max_row + 1):
            if str(sheet.cell(row, 2).value or "").startswith(label_prefix):
                return row
        raise AssertionError(f"Không tìm thấy dòng bắt đầu bằng {label_prefix!r}.")

    @staticmethod
    def _as_date(value: object) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    @staticmethod
    def _data_tvv_row(stt: int, ma_to: str, ten_to: str) -> list[str | int]:
        return [
            stt,
            ma_to,
            ten_to,
            f"{ten_to} đầy đủ",
            "Xã A",
            f"TT{stt:03d}",
            "Nguyễn Văn A",
            "Ấp 1",
            "123456789",
            "0900000000",
            "Hội nông dân",
            "987654321",
            "Tổ chức A",
            "Huyện A",
            "111",
            "Tỉnh A",
            "222",
            "Trung ương",
            "333",
            "Có",
            "TW",
            "Tỉnh",
        ]


if __name__ == "__main__":
    unittest.main()
