from __future__ import annotations

from contextlib import closing
import csv
from dataclasses import replace
from datetime import date
import hashlib
import json
import os
from pathlib import Path
import sqlite3
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch
import zipfile

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from PySide6.QtCore import QPoint, Qt
from PySide6.QtWidgets import QApplication, QComboBox, QFileDialog, QLineEdit, QMessageBox, QPushButton, QRadioButton, QScrollArea, QSizePolicy, QStyleOptionViewItem, QWidget

from agribank_v3.features.catalog import SECTIONS
from agribank_v3.features.credit.summary.credit_report import (
    CREDIT_QUALITY_DISPLAY_NAME,
    CUSTOMER_TYPE_LEGAL,
    CUSTOMER_TYPE_PERSONAL,
    GROUP_CREDIT_QUALITY,
    GROUP_CUSTOMER_TYPE,
    GROUP_DECREE55,
    GROUP_INDUSTRY,
    GROUP_SUMMARY,
    GROUP_TERM_STRUCTURE,
    REPORT_GROUP_LABELS,
    VIEW_COMPARE_PERIODS,
    VIEW_CURRENT_PERIOD,
    CreditReportFilters,
    CreditReportRepository,
    compare_report_snapshots,
    classify_industry,
    classify_term_from_account,
    get_report_snapshot,
)
from agribank_v3.features.credit.summary.dashboard_repository import NimDashboardRepository
from agribank_v3.features.credit.summary.dashboard_charts import DashboardBranchComparisonChart, branch_bar_values, branch_period_pair_values
from agribank_v3.features.credit.summary.dashboard_export import (
    SHEET_BRANCH,
    SHEET_DETAIL,
    SHEET_GROWTH,
    SHEET_OVERVIEW,
    DashboardNimExportService,
    export_dashboard_rows,
)
from agribank_v3.features.credit.summary.dashboard_service import DashboardFilters, build_nim_dashboard
from agribank_v3.features.credit.summary.dashboard_window import (
    BRANCH_MODE_CURRENT,
    BRANCH_MODE_PERIOD_COMPARE,
    DETAIL_MODE_OFFICE,
    OVERVIEW_MODE_ENDPOINTS,
    NimDashboardWindow,
)
from agribank_v3.features.credit.summary.database import CREDIT_SUMMARY_DATABASE_NAME, credit_summary_database_path
from agribank_v3.features.credit.summary.menu import SUMMARY_FEATURES
from agribank_v3.features.credit.summary.models import (
    CreditLimitRow,
    LOAN_COMPARE_TITLE,
    DashboardData,
    DashboardMetric,
    NormalizedLoanRow,
    REPORT_DATA_TITLE,
    REPORT_SUMMARY_TITLE,
    SummaryDataType,
    SummaryError,
)
from agribank_v3.features.credit.summary.report_data_menu import (
    REPORT_DATA_FEATURES,
    REPORT_SUMMARY_ROUTE,
)
from agribank_v3.features.credit.summary.customer.routes import CUSTOMER_DATA_TITLE
from agribank_v3.features.credit.summary.ln01 import parse_ln01_bytes, normalize_credit_group_code
from agribank_v3.features.credit.summary.group_lending.export_service import GroupLendingExportService
from agribank_v3.features.credit.summary.group_lending.models import (
    ASSOCIATION_UNKNOWN,
    DETAIL_BY_GROUP,
    GROUP_STATUS_NOT_DECLARED,
    GroupLendingFilters,
    SUMMARY_BY_ASSOCIATION,
)
from agribank_v3.features.credit.summary.group_lending.repository import GroupLendingRepository
from agribank_v3.features.credit.summary.group_lending.service import GroupLendingService
from agribank_v3.features.credit.summary.report_window import ReportSummaryWindow
from agribank_v3.features.credit.tovayvon.models import ASSOCIATION_FARMERS_UNION, ASSOCIATION_OTHER, ASSOCIATION_WOMENS_UNION, CreditGroup
from agribank_v3.features.credit.tovayvon.repository import CreditGroupRepository
from agribank_v3.features.credit.summary.credit_limit import (
    CreditLimitExcelBatchStore,
    credit_limit_row_to_excel_values,
    excel_record_to_credit_limit_row,
    normalize_credit_limit_row,
)
from agribank_v3.features.credit.summary.credit_limit.models import DATA_SHEET_NAME, META_SHEET_NAME
from agribank_v3.features.credit.summary.nim_ui_config import NIM_DN_UI_CONFIG, NIM_NV_UI_CONFIG
from agribank_v3.features.credit.summary.officer_history.export import SHEET_NAMES, export_analysis_rows
from agribank_v3.features.credit.summary.officer_history.models import (
    METRIC_AVERAGE_RATE,
    METRIC_BALANCE,
    METRIC_BALANCE_GROWTH,
    METRIC_NIM_AFTER,
    METRIC_NIM_BEFORE,
    HistoryFilters,
    HistoryPoint,
)
from agribank_v3.features.credit.summary.officer_history.repository import OfficerHistoryRepository, officer_key
from agribank_v3.features.credit.summary.officer_history.service import (
    build_multiple_officer_comparison,
    build_officer_branch_comparison,
    build_officer_growth_history,
    build_officer_overview,
)
from agribank_v3.features.credit.summary.officer_history.window import OfficerHistoryDialog
from agribank_v3.features.credit.summary.officer_history.widgets import OfficerMultiSelectCombo
from agribank_v3.features.credit.summary.reports import (
    CREDIT_LIMIT_VBA_HEADERS,
    LOAN_COMPARE_VBA_HEADERS,
    NIM_DN_DETAIL_HEADERS,
    NIM_NV_DETAIL_HEADERS,
    export_credit_limit_view_report,
    export_rows,
)
from agribank_v3.features.credit.summary.repository import SummaryRepository
from agribank_v3.features.credit.summary.regression import compare_workbooks
from agribank_v3.features.credit.summary.services import (
    Ln01DuplicateDecision,
    Ln01ImportCoordinator,
    build_officer_history,
    compare_loan_balances,
    export_officer_history_excel,
    import_credit_limit_file,
    import_nim_dn,
    import_nim_nv,
)
from agribank_v3.features.credit.summary.ln01_import_dialog import _dialog_actions
from agribank_v3.features.credit.summary.windows import (
    CreditLimitOfficerChart,
    CreditLimitTab,
    DeleteNimPeriodDialog,
    LoanCompareTab,
    NimTrendChart,
    NimTab,
    SummaryMaintenanceDialog,
    _chart_expired,
    _chart_expiring,
    _chart_total,
    _chart_total_limit,
    _chart_total_outstanding,
    _credit_limit_chart_tooltip,
    _credit_limit_row_value,
    _display_officer_name,
    _format_money_vn,
    _format_percent_vn,
    _integer_ticks,
    _populate_combo,
)
from agribank_v3.features.settings.unit_directory.models import (
    AppUnitSettings,
    BranchDirectoryEntry,
    OfficeDirectoryEntry,
    TRANSACTION_OFFICE,
)
from agribank_v3.ui.main_window import MainWindow
from agribank_v3.ui.components.controls import (
    configure_combo_popup_width,
    configure_filter_combo_width,
    danger_button,
    primary_button,
    recommended_button_width,
    recommended_control_height,
    secondary_button,
)
from agribank_v3.ui.components.kpi import CompactKpiCard, MetricGrid
from agribank_v3.update.db_migrations import MigrationSpec, apply_migrations


GROUP_ORDER_FOR_TESTS = (
    GROUP_SUMMARY,
    GROUP_TERM_STRUCTURE,
    GROUP_CUSTOMER_TYPE,
    GROUP_CREDIT_QUALITY,
    GROUP_DECREE55,
    GROUP_INDUSTRY,
)


class CreditSummaryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.qt_app = QApplication.instance() or QApplication([])
        self.temporary_directory = TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.database_path = self.root / "DuLieuV3.db"
        self.repository = SummaryRepository(self.database_path)

    def tearDown(self) -> None:
        self.temporary_directory.cleanup()

    def _write_ln01_file(self, name: str = "5491_ln01_20260427.csv") -> Path:
        path = self.root / name
        headers = [f"H{i}" for i in range(63)]
        headers[0] = "BRCD"
        headers[1] = "CUSTSEQ"
        headers[2] = "CUSTNM"
        headers[3] = "TAI_KHOAN"
        headers[5] = "DU_NO"
        headers[14] = "APPRSEQ"
        headers[15] = "APPROVED_DATE"
        headers[17] = "APPROVED_AMOUNT"
        headers[18] = "EXPIRY_DATE"
        headers[27] = "OFFICER"
        headers[35] = "ADDR1"
        headers[49] = "GRPNO"
        headers[62] = "CREDIT_LINE_YPE"

        def row(
            customer: str,
            contract: str,
            expiry: str,
            outstanding: str,
            *,
            credit_type: str = "Line of Credit",
            group_code: str = "5491LLG202100001",
        ) -> list[str]:
            values = [""] * 63
            values[0] = "5491"
            values[1] = customer
            values[2] = f"Khach {customer}"
            values[3] = f"TK{customer}"
            values[5] = outstanding
            values[14] = contract
            values[15] = "20240101"
            values[17] = "1000"
            values[18] = expiry
            values[27] = "CB1"
            values[35] = "Dia chi"
            values[49] = group_code
            values[62] = credit_type
            return values

        rows = [
            row("KH01", "HD01", "20240115", "100"),
            row("KH02", "HD02", "20240210", "200"),
            row("KH03", "HD03", "20240510", "300"),
            row("KH04", "HD04", "20240210", "400", credit_type="Term Loan"),
            row("KH01", "HD01", "20240115", "50"),
        ]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(rows)
        return path

    def _write_report_ln01_file(self, name: str = "5491_ln01_20260430.csv", *, scale: float = 1.0) -> Path:
        path = self.root / name
        headers = [f"H{i}" for i in range(77)]
        headers[0] = "BRCD"
        headers[1] = "CUSTSEQ"
        headers[2] = "CUSTNM"
        headers[3] = "TAI_KHOAN"
        headers[5] = "DU_NO"
        headers[14] = "APPRSEQ"
        headers[15] = "APPRDT"
        headers[17] = "APPRAMT"
        headers[18] = "APPRMATDT"
        headers[27] = "OFFICER_NAME"
        headers[31] = "CUSTOMER_TYPE_CODE"
        headers[35] = "ADDR1"
        headers[43] = "SECURED_PERCENT"
        headers[44] = "NHOM_NO"
        headers[49] = "GRPNO"
        headers[62] = "CREDIT_LINE_YPE"
        headers[76] = "MA_NGANH_KT"

        def row(
            customer: str,
            account: str,
            contract: str,
            balance: str,
            customer_type: str,
            debt_group: str,
            secured_percent: str,
            industry_code: str,
            group_code: str = "5491LLG202100001",
        ) -> list[str]:
            values = [""] * 77
            values[0] = "5491"
            values[1] = customer
            values[2] = f"Khach {customer}"
            values[3] = account
            values[5] = balance
            values[14] = contract
            values[15] = "20240101"
            values[17] = "1000"
            values[18] = "20241231"
            values[27] = "[CB01] Nguyen Van A"
            values[31] = customer_type
            values[35] = "Dia chi"
            values[43] = secured_percent
            values[44] = debt_group
            values[49] = group_code
            values[62] = "Line of Credit"
            values[76] = industry_code
            return values

        rows = [
            row("0001", "211001", "HD01", str(int(100 * scale)), "100", "01", "0", "10104", "5491LLG202100001"),
            row("0002", "212001", "HD02", str(int(200 * scale)), "999", "02", "10", "100101", "5491LLG202100001"),
            row("0003", "213001", "HD03", str(int(300 * scale)), "570", "03", "0,00", "812345", "5491LLG202100002"),
            row("0001", "211001", "HD01", str(int(50 * scale)), "100", "01", "0", "10104", "5491LLG202100001"),
            row("0004", "999001", "HD04", "0", "999", "", "5", "999999", ""),
        ]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(rows)
        return path

    def _seed_credit_groups(self) -> None:
        repository = CreditGroupRepository(self.database_path)
        repository.save_group(
            CreditGroup(
                ma_to="5491LLG202100001",
                ten_to="To vay von 001",
                xa="Xa 1",
                ten_to_truong="To truong 1",
                association_type=ASSOCIATION_FARMERS_UNION,
                active=True,
            )
        )
        repository.save_group(
            CreditGroup(
                ma_to="5491LLG202100002",
                ten_to="To vay von 002",
                xa="Xa 2",
                ten_to_truong="To truong 2",
                association_type=ASSOCIATION_WOMENS_UNION,
                active=True,
            )
        )

    def _write_ln01_group_alias_file(self, name: str = "5491_ln01_alias_20260430.csv") -> Path:
        path = self.root / name
        headers = [f"H{i}" for i in range(63)]
        headers[0] = "BRCD"
        headers[1] = "CUSTSEQ"
        headers[2] = "CUSTNM"
        headers[3] = "TAI_KHOAN"
        headers[5] = "DU_NO"
        headers[10] = "MaToVayVon"
        headers[14] = "APPRSEQ"
        headers[15] = "APPRDT"
        headers[17] = "APPRAMT"
        headers[18] = "APPRMATDT"
        headers[27] = "OFFICER_NAME"
        headers[35] = "ADDR1"
        headers[49] = "AX_NOT_GROUP"
        headers[62] = "CREDIT_LINE_YPE"
        values = [""] * 63
        values[0] = "5491"
        values[1] = "0001"
        values[2] = "Khach alias"
        values[3] = "211001"
        values[5] = "100"
        values[10] = "ALIAS-GROUP-001"
        values[14] = "HD01"
        values[15] = "20240101"
        values[17] = "1000"
        values[18] = "20241231"
        values[27] = "[CB01] Nguyen Van A"
        values[35] = "Dia chi"
        values[49] = "WRONG-FALLBACK-GROUP"
        values[62] = "Line of Credit"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerow(values)
        return path

    def test_credit_statement_card_renamed_to_report_data(self) -> None:
        titles = [feature.title for feature in SECTIONS["Tín dụng"]]
        self.assertIn(REPORT_DATA_TITLE, titles)
        self.assertNotIn("Sao kê tín dụng", titles)

    def test_report_child_menu_has_summary_report(self) -> None:
        titles = [feature.title for feature in REPORT_DATA_FEATURES]
        self.assertEqual(titles, [REPORT_SUMMARY_TITLE])
        self.assertNotIn("Chức năng 2", titles)
        self.assertEqual(REPORT_SUMMARY_ROUTE, "credit.report_summary")

    def test_report_data_card_opens_child_menu(self) -> None:
        window = MainWindow()
        try:
            window.open_feature(REPORT_DATA_TITLE)
            self.assertIsNotNone(window.report_data_page)
            self.assertIs(window.pages.currentWidget(), window.report_data_page)
        finally:
            window.close()

    def test_report_child_menu_does_not_invent_other_feature_names(self) -> None:
        titles = [feature.title for feature in REPORT_DATA_FEATURES]
        self.assertNotIn("Chức năng 2", titles)
        self.assertNotIn("Chức năng 3", titles)
        self.assertNotIn("Chức năng 4", titles)

    def test_summary_report_single_instance_and_non_modal(self) -> None:
        window = MainWindow()
        try:
            first = window.open_report_summary_window()
            second = window.open_report_summary_window()
            self.assertIs(first, second)
            self.assertFalse(first.isModal())
        finally:
            if window._report_summary_window is not None:
                window._report_summary_window.close()
            window.close()

    def test_report_window_has_group_lending_tab_in_order(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertEqual(
                [window.tabs.tabText(index) for index in range(window.tabs.count())],
                ["Tổng quan", "Cho vay qua tổ", "Lịch sử import", "Cài đặt phân loại"],
            )
            removed = {
                "Cơ cấu thời hạn",
                "Loại hình khách hàng",
                "Chất lượng dư nợ",
                "Cho vay Nghị định 55",
                "Dư nợ theo ngành kinh tế",
                "So sánh các kỳ",
            }
            self.assertTrue(removed.isdisjoint({window.tabs.tabText(index) for index in range(window.tabs.count())}))
        finally:
            window.close()

    def test_report_window_modes_and_group_options(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertTrue(window.mode_button_group.exclusive())
            self.assertTrue(window.current_mode_radio.isChecked())
            self.assertFalse(window.from_period_combo.isEnabled())
            self.assertFalse(window.to_period_combo.isEnabled())
            self.assertTrue(window.period_combo.isEnabled())
            window.compare_mode_radio.setChecked(True)
            self.assertTrue(window.from_period_combo.isEnabled())
            self.assertTrue(window.to_period_combo.isEnabled())
            self.assertFalse(window.period_combo.isEnabled())
            self.assertEqual(set(window.group_buttons), set(GROUP_ORDER_FOR_TESTS))
            self.assertEqual(window.group_buttons[GROUP_CREDIT_QUALITY].text(), CREDIT_QUALITY_DISPLAY_NAME)
            visible_text = " ".join(
                [window.tabs.tabText(index) for index in range(window.tabs.count())]
                + [button.text() for button in window.group_buttons.values()]
            )
            self.assertNotIn("Chất lượng dư nợ", visible_text)
        finally:
            window.close()

    def test_report_view_mode_is_outside_overview_tab(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertIsNotNone(window.view_mode_widget)
            self.assertIsNot(window.view_mode_widget.parentWidget(), window.overview_tab)
            self.assertNotIn(window.current_mode_radio, window.overview_tab.findChildren(QRadioButton))
            self.assertNotIn(window.compare_mode_radio, window.overview_tab.findChildren(QRadioButton))
        finally:
            window.close()

    def test_report_overview_has_no_view_mode_widget(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertIsNone(window.overview_tab.findChild(QWidget, "ReportViewModeWidget"))
            overview_radio_texts = {button.text() for button in window.overview_tab.findChildren(QRadioButton)}
            self.assertNotIn("Kỳ hiện tại", overview_radio_texts)
            self.assertNotIn("So sánh các kỳ", overview_radio_texts)
        finally:
            window.close()

    def test_global_customer_search_removed(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertFalse(hasattr(window, "search_edit"))
            placeholders = {edit.placeholderText() for edit in window.findChildren(QLineEdit)}
            self.assertNotIn("Tìm mã hoặc tên khách hàng", placeholders)
        finally:
            window.close()

    def test_report_header_has_view_mode_widget(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            area = window.findChild(QWidget, "ReportSummaryFilterArea")
            self.assertIsNotNone(area)
            self.assertIs(window.view_mode_widget.parentWidget(), area)
            self.assertIsNotNone(area.findChild(QWidget, "ReportViewModeWidget"))
        finally:
            window.close()

    def test_view_mode_replaces_search_position(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            area = window.findChild(QWidget, "ReportSummaryFilterArea")
            flow = area.layout()
            self.assertIs(flow.itemAt(flow.count() - 1).widget(), window.view_mode_widget)
            self.assertEqual(window.view_mode_widget.objectName(), "ReportViewModeWidget")
        finally:
            window.close()

    def test_view_mode_radio_group_is_exclusive(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertTrue(window.mode_button_group.exclusive())
            window.compare_mode_radio.setChecked(True)
            self.assertFalse(window.current_mode_radio.isChecked())
            self.assertTrue(window.compare_mode_radio.isChecked())
        finally:
            window.close()

    def test_current_period_is_default(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertEqual(window._view_mode(), VIEW_CURRENT_PERIOD)
            self.assertTrue(window.current_mode_radio.isChecked())
        finally:
            window.close()

    def test_current_mode_enables_report_period(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.current_mode_radio.setChecked(True)
            self.assertTrue(window.period_combo.isEnabled())
        finally:
            window.close()

    def test_current_mode_disables_from_period(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.current_mode_radio.setChecked(True)
            self.assertFalse(window.from_period_combo.isEnabled())
        finally:
            window.close()

    def test_current_mode_disables_to_period(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.current_mode_radio.setChecked(True)
            self.assertFalse(window.to_period_combo.isEnabled())
        finally:
            window.close()

    def test_compare_mode_enables_from_period(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.compare_mode_radio.setChecked(True)
            self.assertTrue(window.from_period_combo.isEnabled())
        finally:
            window.close()

    def test_compare_mode_enables_to_period(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.compare_mode_radio.setChecked(True)
            self.assertTrue(window.to_period_combo.isEnabled())
        finally:
            window.close()

    def test_compare_mode_disables_report_period(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.compare_mode_radio.setChecked(True)
            self.assertFalse(window.period_combo.isEnabled())
        finally:
            window.close()

    def test_view_mode_applies_to_overview_tab(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.tabs.setCurrentWidget(window.overview_tab)
            with patch.object(window, "reload_overview") as reload_overview, patch.object(window.group_lending_tab, "refresh") as group_refresh:
                window.compare_mode_radio.setChecked(True)
            reload_overview.assert_called_once()
            group_refresh.assert_not_called()
        finally:
            window.close()

    def test_view_mode_applies_to_group_lending_tab(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.tabs.setCurrentWidget(window.group_lending_tab)
            with patch.object(window, "reload_overview") as reload_overview, patch.object(window.group_lending_tab, "refresh") as group_refresh:
                window.compare_mode_radio.setChecked(True)
            group_refresh.assert_called_once()
            reload_overview.assert_not_called()
        finally:
            window.close()

    def test_view_mode_does_not_refresh_import_history(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.tabs.setCurrentIndex(2)
            with patch.object(window.repository, "import_history_rows") as import_history_rows:
                window.compare_mode_radio.setChecked(True)
            import_history_rows.assert_not_called()
        finally:
            window.close()

    def test_view_mode_does_not_refresh_classification_settings(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            window.tabs.setCurrentWidget(window.rules_tab)
            with patch.object(window.repository, "customer_type_rule_rows") as rule_rows:
                window.compare_mode_radio.setChecked(True)
            rule_rows.assert_not_called()
        finally:
            window.close()

    def test_report_export_reads_global_view_mode(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        captured: dict[str, object] = {}
        try:
            window.tabs.setCurrentWidget(window.overview_tab)
            with patch.object(window, "reload_overview"), patch.object(window.group_lending_tab, "refresh"):
                window.compare_mode_radio.setChecked(True)
            with patch.object(QFileDialog, "getSaveFileName", return_value=(str(self.root / "report.xlsx"), "Excel (*.xlsx)")):
                with patch.object(
                    window.repository,
                    "export_workbook",
                    side_effect=lambda path, *, filters, view_mode: captured.update({"filters": filters, "view_mode": view_mode}) or path,
                ):
                    with patch.object(window, "_run_background", side_effect=lambda _title, fn, _done: fn(lambda _message: None)):
                        window.export_excel()
            self.assertEqual(captured["view_mode"], VIEW_COMPARE_PERIODS)
            self.assertEqual(captured["filters"].search, "")
        finally:
            window.close()

    def test_group_lending_export_reads_global_view_mode(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        captured: dict[str, object] = {}
        try:
            window.tabs.setCurrentWidget(window.group_lending_tab)
            with patch.object(window, "reload_overview"), patch.object(window.group_lending_tab, "refresh"):
                window.compare_mode_radio.setChecked(True)
            def fake_export(*args, **kwargs):
                path = args[-1] if args else self.root / "groups.xlsx"
                captured.update(
                    {
                        "view_mode": kwargs.get("view_mode"),
                        "filters": kwargs.get("filters"),
                        "display_mode": kwargs.get("display_mode"),
                    }
                )
                return path
            with patch.object(QFileDialog, "getSaveFileName", return_value=(str(self.root / "groups.xlsx"), "Excel (*.xlsx)")):
                with patch.object(
                    GroupLendingExportService,
                    "export",
                    side_effect=fake_export,
                ):
                    with patch.object(QMessageBox, "information", return_value=None), patch.object(QMessageBox, "warning", return_value=None):
                        window.export_excel()
            self.assertEqual(captured["view_mode"], VIEW_COMPARE_PERIODS)
        finally:
            window.close()

    def test_group_lending_search_box_still_exists(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertEqual(window.group_lending_tab.search_edit.placeholderText(), "Tìm mã hoặc tên tổ")
        finally:
            window.close()

    def test_global_search_signal_removed(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            self.assertFalse(hasattr(window, "search_edit"))
            self.assertEqual(window._filters().search, "")
        finally:
            window.close()

    def test_report_filter_layout_has_no_horizontal_scroll(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            area = window.findChild(QWidget, "ReportSummaryFilterArea")
            self.assertIsNotNone(area)
            self.assertFalse(hasattr(area, "horizontalScrollBar"))
            self.assertTrue(area.layout().hasHeightForWidth())
        finally:
            window.close()

    def test_report_filter_layout_wraps_on_narrow_width(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            area = window.findChild(QWidget, "ReportSummaryFilterArea")
            flow = area.layout()
            self.assertGreater(flow.heightForWidth(640), flow.heightForWidth(1800))
        finally:
            window.close()

    def test_mode_widget_text_not_clipped(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            for button in (window.current_mode_radio, window.compare_mode_radio):
                required_width = button.fontMetrics().horizontalAdvance(button.text()) + 32
                self.assertGreaterEqual(button.minimumWidth(), required_width)
            self.assertGreaterEqual(window.view_mode_widget.minimumWidth(), window.current_mode_radio.minimumWidth() + window.compare_mode_radio.minimumWidth())
        finally:
            window.close()

    def test_group_lending_tab_options_are_exclusive_and_default_detail(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            tab = window.group_lending_tab
            self.assertTrue(tab.option_button_group.exclusive())
            self.assertTrue(tab.detail_radio.isChecked())
            self.assertEqual(tab.detail_radio.property("group_lending_mode"), DETAIL_BY_GROUP)
            self.assertEqual(tab.association_radio.property("group_lending_mode"), SUMMARY_BY_ASSOCIATION)
            self.assertNotIn("Chart", {child.__class__.__name__ for child in tab.findChildren(QWidget)})
        finally:
            window.close()

    def _assert_combo_text_has_room(self, combo: QComboBox, text: str | None = None) -> None:
        visible_text = text if text is not None else combo.itemText(0)
        required_width = combo.fontMetrics().horizontalAdvance(visible_text) + 54
        self.assertGreaterEqual(combo.minimumWidth(), required_width, visible_text)
        self.assertEqual(combo.toolTip().splitlines()[0], combo.currentText())
        line_edit = combo.lineEdit()
        self.assertIsInstance(line_edit, QLineEdit)
        self.assertEqual(line_edit.cursorPosition(), 0)

    def test_report_filter_combo_minimum_widths(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            expected = (
                (window.from_period_combo, 132),
                (window.to_period_combo, 132),
                (window.period_combo, 136),
                (window.branch_combo, 168),
                (window.office_combo, 184),
                (window.customer_type_combo, 176),
                (window.debt_group_combo, 152),
                (window.term_combo, 166),
                (window.officer_combo, 170),
            )
            for combo, minimum_width in expected:
                self.assertGreaterEqual(combo.minimumWidth(), minimum_width)
        finally:
            window.close()

    def test_report_period_combo_text_visible(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            for combo in (window.from_period_combo, window.to_period_combo, window.period_combo):
                self._assert_combo_text_has_room(combo)
        finally:
            window.close()

    def test_report_branch_combo_text_visible(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            for combo in (window.branch_combo, window.office_combo):
                self._assert_combo_text_has_room(combo)
        finally:
            window.close()

    def test_report_combo_popup_width_includes_full_text_and_tooltip(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            combo = window.office_combo
            long_text = "Phòng giao dịch Trung tâm Agribank huyện rất dài"
            combo.addItem(long_text, "pgd-long")
            configure_filter_combo_width(combo, minimum_width=184, maximum_width=250)
            combo.setCurrentIndex(combo.findData("pgd-long"))
            self.qt_app.processEvents()
            self.assertEqual(combo.toolTip(), long_text)
            self.assertGreaterEqual(combo.view().minimumWidth(), combo.fontMetrics().horizontalAdvance(long_text) + 40)
            self.assertEqual(combo.view().horizontalScrollBarPolicy(), Qt.ScrollBarPolicy.ScrollBarAsNeeded)
            self.assertEqual(combo.lineEdit().cursorPosition(), 0)
        finally:
            window.close()

    def test_group_lending_filter_combo_text_visible(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            tab = window.group_lending_tab
            for combo in (tab.branch_combo, tab.office_combo, tab.association_combo, tab.status_combo, tab.officer_combo):
                self._assert_combo_text_has_room(combo)
        finally:
            window.close()

    def test_report_and_group_filter_layouts_wrap_when_narrow(self) -> None:
        window = ReportSummaryWindow(self.database_path)
        try:
            for object_name in ("ReportSummaryFilterArea", "GroupLendingFilterArea"):
                area = window.findChild(QWidget, object_name)
                self.assertIsNotNone(area)
                flow = area.layout()
                self.assertTrue(flow.hasHeightForWidth())
                self.assertGreaterEqual(flow.heightForWidth(640), flow.heightForWidth(1800))
            self.assertEqual(window.group_lending_tab.search_edit.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
        finally:
            window.close()

    def test_customer_data_menu_route_still_exists(self) -> None:
        self.assertIn(CUSTOMER_DATA_TITLE, [feature.title for feature in SECTIONS["Tín dụng"]])

    def test_ln01_parser_reads_grpno(self) -> None:
        path = self._write_ln01_file()

        rows, source_count = parse_ln01_bytes(path, path.read_bytes(), period="2026-04")

        self.assertEqual(source_count, 5)
        self.assertTrue(hasattr(rows[0], "group_code"))
        self.assertEqual(rows[0].group_code, "5491LLG202100001")

    def test_ln01_parser_grpno_alias_preferred_over_ax_fallback(self) -> None:
        path = self._write_ln01_group_alias_file()

        rows, _source_count = parse_ln01_bytes(path, path.read_bytes(), period="2026-04")

        self.assertEqual(rows[0].group_code, "ALIAS-GROUP-001")

    def test_ln01_group_code_normalization_rules(self) -> None:
        self.assertEqual(normalize_credit_group_code(" 5491LLG202100003 "), "5491LLG202100003")
        self.assertEqual(normalize_credit_group_code("'5491LLG202100003"), "5491LLG202100003")
        self.assertEqual(normalize_credit_group_code("549100003.0"), "549100003")
        self.assertIsNone(normalize_credit_group_code(""))
        self.assertIsNone(normalize_credit_group_code(None))

    def test_ln01_missing_grpno_header_warns_but_imports_other_reports(self) -> None:
        path = self._write_report_ln01_file()
        with path.open(encoding="utf-8") as handle:
            rows = list(csv.reader(handle))
        rows[0][49] = "AX_NOT_GROUP"
        for row in rows[1:]:
            row[49] = ""
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerows(rows)

        result = import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            loan_count = connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0]
            run = connection.execute(
                "SELECT group_code_stats_json FROM credit_import_runs WHERE period = '2026-04' AND source_type = 'LN01'"
            ).fetchone()
        stats = json.loads(run["group_code_stats_json"])
        self.assertIn("không có header GRPNO", result.message)
        self.assertEqual(loan_count, 4)
        self.assertFalse(stats["has_group_code_column"])
        self.assertEqual(stats["normalized_loans_with_group_code"], 0)

    def test_credit_loan_period_has_group_code_and_migration_is_idempotent(self) -> None:
        legacy_root = self.root / "legacy"
        legacy_root.mkdir()
        main_database = legacy_root / "DuLieuV3.db"
        credit_database = legacy_root / "Credit.db"
        with closing(sqlite3.connect(credit_database)) as connection:
            with connection:
                connection.executescript(
                    """
                    CREATE TABLE credit_import_runs (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        period TEXT NOT NULL,
                        source_type TEXT NOT NULL,
                        source_file_name TEXT NOT NULL DEFAULT '',
                        source_sha256 TEXT NOT NULL DEFAULT '',
                        source_file_size INTEGER NOT NULL DEFAULT 0,
                        source_row_count INTEGER NOT NULL DEFAULT 0,
                        normalized_loan_count INTEGER NOT NULL DEFAULT 0,
                        customer_count INTEGER NOT NULL DEFAULT 0,
                        ln01_total_balance REAL NOT NULL DEFAULT 0,
                        card_balance REAL NOT NULL DEFAULT 0,
                        status TEXT NOT NULL DEFAULT 'success',
                        warning_count INTEGER NOT NULL DEFAULT 0,
                        message TEXT NOT NULL DEFAULT '',
                        imported_by TEXT NOT NULL DEFAULT '',
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL,
                        duration_ms INTEGER NOT NULL DEFAULT 0
                    );
                    CREATE TABLE credit_customer_master (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        branch_code TEXT NOT NULL,
                        customer_sequence TEXT NOT NULL,
                        customer_code TEXT NOT NULL DEFAULT '',
                        customer_name TEXT NOT NULL DEFAULT '',
                        latest_customer_type_code TEXT NOT NULL DEFAULT '',
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL,
                        UNIQUE(branch_code, customer_sequence)
                    );
                    CREATE TABLE credit_loan_period (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        period TEXT NOT NULL,
                        customer_id INTEGER NOT NULL,
                        branch_code TEXT NOT NULL,
                        loan_key TEXT NOT NULL,
                        account_number TEXT NOT NULL DEFAULT '',
                        approval_sequence TEXT NOT NULL DEFAULT '',
                        outstanding_balance REAL NOT NULL DEFAULT 0,
                        customer_type_code TEXT NOT NULL DEFAULT '',
                        debt_group_code TEXT NOT NULL DEFAULT 'UNKNOWN',
                        secured_percent REAL,
                        industry_code TEXT NOT NULL DEFAULT '',
                        officer_code TEXT NOT NULL DEFAULT '',
                        officer_name TEXT NOT NULL DEFAULT '',
                        term_category TEXT NOT NULL DEFAULT 'UNKNOWN',
                        source_run_id INTEGER NOT NULL,
                        source_row_count INTEGER NOT NULL DEFAULT 1,
                        created_at TEXT NOT NULL,
                        UNIQUE(period, branch_code, loan_key)
                    );
                    INSERT INTO credit_import_runs(id, period, source_type, source_file_name, created_at, updated_at)
                    VALUES (1, '2026-03', 'LN01', 'legacy.csv', '2026-01-01T00:00:00', '2026-01-01T00:00:00');
                    INSERT INTO credit_customer_master(id, branch_code, customer_sequence, customer_code, customer_name, created_at, updated_at)
                    VALUES (1, '5491', '0001', '54910001', 'Khach 0001', '2026-01-01T00:00:00', '2026-01-01T00:00:00');
                    INSERT INTO credit_loan_period(
                        id, period, customer_id, branch_code, loan_key, account_number, approval_sequence,
                        outstanding_balance, customer_type_code, debt_group_code, officer_code, officer_name,
                        term_category, source_run_id, created_at
                    )
                    VALUES (
                        1, '2026-03', 1, '5491', '1:211001:HD01', '211001', 'HD01',
                        100, '100', '01', 'CB01', 'Nguyen Van A', 'SHORT', 1, '2026-01-01T00:00:00'
                    );
                    """
                )

        repository = CreditReportRepository(main_database)
        repository.ensure_schema()

        with closing(repository.connect()) as connection:
            columns = {row["name"] for row in connection.execute("PRAGMA table_info(credit_loan_period)").fetchall()}
            indexes = {row["name"] for row in connection.execute("PRAGMA index_list(credit_loan_period)").fetchall()}
            row = connection.execute("SELECT group_code FROM credit_loan_period WHERE id = 1").fetchone()
            run = connection.execute("SELECT group_code_stats_json FROM credit_import_runs WHERE id = 1").fetchone()
        self.assertIn("group_code", columns)
        self.assertIn("idx_credit_loan_period_period_group_customer", indexes)
        self.assertIsNone(row["group_code"])
        self.assertEqual(run["group_code_stats_json"], "{}")
        self.assertTrue(any((legacy_root / "backups").glob("Credit-before-*.db")))

    def test_ln01_import_writes_group_code_and_metadata(self) -> None:
        self._seed_credit_groups()
        path = self._write_report_ln01_file()

        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            rows = {
                row["account_number"]: row
                for row in connection.execute(
                    """
                    SELECT account_number, outstanding_balance, group_code
                    FROM credit_loan_period
                    WHERE period = '2026-04'
                    """
                ).fetchall()
            }
            run = connection.execute(
                "SELECT group_code_stats_json FROM credit_import_runs WHERE period = '2026-04' AND source_type = 'LN01'"
            ).fetchone()
        stats = json.loads(run["group_code_stats_json"])
        self.assertEqual(rows["211001"]["group_code"], "5491LLG202100001")
        self.assertEqual(rows["211001"]["outstanding_balance"], 150)
        self.assertIsNone(rows["999001"]["group_code"])
        self.assertTrue(stats["has_group_code_column"])
        self.assertEqual(stats["source_rows_with_group_code"], 4)
        self.assertEqual(stats["normalized_loans_with_group_code"], 3)
        self.assertEqual(stats["balance_with_group_code"], 650)
        self.assertEqual(stats["matched_group_code_count"], 2)
        self.assertEqual(stats["unknown_group_code_count"], 0)

    def test_ln01_overwrite_period_replaces_group_code(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20260430.csv")
        second_path = self._write_report_ln01_file("5491_ln01_20260430_v2.csv")
        second_path.write_text(
            second_path.read_text(encoding="utf-8").replace("5491LLG202100001", "5491LLG202100003"),
            encoding="utf-8",
        )
        import_credit_limit_file(self.repository, first_path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        import_credit_limit_file(
            self.repository,
            second_path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.CREATE_BOTH,
        )

        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            codes = {
                str(row["group_code"])
                for row in connection.execute(
                    "SELECT group_code FROM credit_loan_period WHERE period = '2026-04' AND group_code IS NOT NULL"
                ).fetchall()
            }
            loan_count = connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0]
        self.assertEqual(loan_count, 4)
        self.assertEqual(codes, {"5491LLG202100002", "5491LLG202100003"})

    def test_hmhethan_workbook_preserves_group_code(self) -> None:
        path = self._write_ln01_file()

        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        batch = self.repository.list_credit_limit_batches()[0]
        workbook = load_workbook(batch.file_path, data_only=True)
        try:
            sheet = workbook[DATA_SHEET_NAME]
            headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
            group_column = headers.index("group_code") + 1
            self.assertEqual(sheet.cell(2, group_column).value, "5491LLG202100001")
        finally:
            workbook.close()

    def test_group_lending_detail_joins_credit_groups_by_ma_to(self) -> None:
        self._seed_credit_groups()
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        result = GroupLendingRepository(self.database_path).get_group_lending_snapshot(
            "2026-04",
            GroupLendingFilters(period="2026-04"),
        )

        rows = {row.group_code: row for row in result.rows}
        kpis = {kpi.label: kpi.value for kpi in result.kpis}
        self.assertEqual(rows["5491LLG202100001"].group_name, "To vay von 001")
        self.assertEqual(rows["5491LLG202100001"].association_type, ASSOCIATION_FARMERS_UNION)
        self.assertEqual(rows["5491LLG202100001"].member_count, 2)
        self.assertEqual(rows["5491LLG202100001"].loan_count, 2)
        self.assertEqual(rows["5491LLG202100001"].total_balance, 350)
        self.assertEqual(rows["5491LLG202100002"].association_type, ASSOCIATION_WOMENS_UNION)
        self.assertEqual(kpis["Số tổ có dư nợ"], 2)
        self.assertEqual(kpis["Tổng số tổ viên duy nhất"], 3)
        self.assertEqual(kpis["Tổng lượt tổ viên theo tổ"], 3)
        self.assertEqual(kpis["Tổng dư nợ cho vay qua tổ"], 650)

    def test_group_lending_unknown_group_is_preserved(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        result = GroupLendingRepository(self.database_path).get_group_lending_snapshot(
            "2026-04",
            GroupLendingFilters(period="2026-04"),
        )

        rows = {row.group_code: row for row in result.rows}
        self.assertEqual(rows["5491LLG202100001"].status, GROUP_STATUS_NOT_DECLARED)
        self.assertEqual(rows["5491LLG202100001"].group_name, "Chưa khai báo")
        self.assertEqual(rows["5491LLG202100001"].association_type, ASSOCIATION_UNKNOWN)
        self.assertEqual(sum(row.total_balance for row in result.rows), 650)

    def test_group_lending_blank_group_code_is_excluded_and_warns_old_period(self) -> None:
        path = self._write_report_ln01_file()
        path.write_text(
            path.read_text(encoding="utf-8")
            .replace("5491LLG202100001", "")
            .replace("5491LLG202100002", ""),
            encoding="utf-8",
        )
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        result = GroupLendingRepository(self.database_path).get_group_lending_snapshot(
            "2026-04",
            GroupLendingFilters(period="2026-04"),
        )

        self.assertEqual(result.rows, ())
        self.assertIn("Kỳ 2026-04 chưa có dữ liệu mã tổ vay vốn", " ".join(result.notes))
        self.assertEqual(result.diagnostics["no_group_balance"], 650)

    def test_group_lending_association_summary_and_dynamic_directory_updates(self) -> None:
        self._seed_credit_groups()
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        repository = GroupLendingRepository(self.database_path)
        filters = GroupLendingFilters(period="2026-04")
        before = {row.association_type: row for row in repository.get_association_summary("2026-04", filters).rows}
        CreditGroupRepository(self.database_path).save_group(
            CreditGroup(
                ma_to="5491LLG202100002",
                ten_to="To vay von 002 updated",
                xa="Xa 2",
                ten_to_truong="To truong 2",
                association_type=ASSOCIATION_OTHER,
                association_other_name="Hoi khac",
                active=True,
            )
        )
        after = {row.group_code: row for row in repository.get_group_lending_snapshot("2026-04", filters).rows}

        self.assertEqual(before[ASSOCIATION_FARMERS_UNION].total_balance, 350)
        self.assertEqual(before[ASSOCIATION_WOMENS_UNION].total_balance, 300)
        self.assertEqual(after["5491LLG202100002"].group_name, "To vay von 002 updated")
        self.assertEqual(after["5491LLG202100002"].association_type, ASSOCIATION_OTHER)

    def test_group_lending_compare_is_direct_full_outer_join(self) -> None:
        self._seed_credit_groups()
        first_path = self._write_report_ln01_file("5491_ln01_20260331.csv")
        first_path.write_text(
            first_path.read_text(encoding="utf-8").replace("5491LLG202100002", ""),
            encoding="utf-8",
        )
        second_path = self._write_report_ln01_file("5491_ln01_20260430.csv", scale=2.0)
        import_credit_limit_file(self.repository, first_path, period="2026-03", warn_days=30, reference_date=date(2024, 1, 20))
        import_credit_limit_file(self.repository, second_path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        result = GroupLendingRepository(self.database_path).compare_groups(
            "2026-03",
            "2026-04",
            GroupLendingFilters(from_period="2026-03", to_period="2026-04"),
        )

        rows = {row.group_code: row for row in result.rows}
        self.assertEqual(rows["5491LLG202100001"].balance_from, 350)
        self.assertEqual(rows["5491LLG202100001"].balance_to, 700)
        self.assertEqual(rows["5491LLG202100001"].balance_growth_rate, 100)
        self.assertEqual(rows["5491LLG202100002"].balance_from, 0)
        self.assertEqual(rows["5491LLG202100002"].balance_to, 600)
        self.assertEqual(rows["5491LLG202100002"].movement_category, "Tổ mới có dư nợ")

    def test_group_lending_customer_in_multiple_groups_diagnostics_and_unique_kpi(self) -> None:
        self._seed_credit_groups()
        path = self._write_report_ln01_file()
        with path.open(encoding="utf-8") as handle:
            rows = list(csv.reader(handle))
        rows[3][1] = "0001"
        rows[3][2] = "Khach 0001"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerows(rows)
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        result = GroupLendingRepository(self.database_path).get_group_lending_snapshot(
            "2026-04",
            GroupLendingFilters(period="2026-04"),
        )
        kpis = {kpi.label: kpi.value for kpi in result.kpis}

        self.assertEqual(kpis["Tổng số tổ viên duy nhất"], 2)
        self.assertEqual(kpis["Tổng lượt tổ viên theo tổ"], 3)
        self.assertEqual(result.diagnostics["multi_group_customer_count"], 1)
        self.assertEqual(result.diagnostics["credit_card_scope"], "excluded_without_group_code")

    def test_group_lending_export_modes(self) -> None:
        self._seed_credit_groups()
        first_path = self._write_report_ln01_file("5491_ln01_20260331.csv", scale=0.5)
        second_path = self._write_report_ln01_file("5491_ln01_20260430.csv", scale=1.0)
        import_credit_limit_file(self.repository, first_path, period="2026-03", warn_days=30, reference_date=date(2024, 1, 20))
        import_credit_limit_file(self.repository, second_path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        exporter = GroupLendingExportService(GroupLendingService(self.database_path))
        cases = (
            (VIEW_CURRENT_PERIOD, DETAIL_BY_GROUP, "ChoVayQuaTo", "current_detail.xlsx"),
            (VIEW_CURRENT_PERIOD, SUMMARY_BY_ASSOCIATION, "TongHopTheoHoi", "current_association.xlsx"),
            (VIEW_COMPARE_PERIODS, DETAIL_BY_GROUP, "SoSanhTheoTo", "compare_detail.xlsx"),
            (VIEW_COMPARE_PERIODS, SUMMARY_BY_ASSOCIATION, "SoSanhTheoHoi", "compare_association.xlsx"),
        )

        for view_mode, display_mode, sheet_name, file_name in cases:
            output = exporter.export(
                self.root / file_name,
                filters=GroupLendingFilters(period="2026-04", from_period="2026-03", to_period="2026-04"),
                view_mode=view_mode,
                display_mode=display_mode,
            )
            workbook = load_workbook(output, data_only=True)
            try:
                self.assertIn(sheet_name, workbook.sheetnames)
                self.assertIn("ThongTin", workbook.sheetnames)
                self.assertGreaterEqual(workbook[sheet_name].max_row, 2)
                metadata = {
                    workbook["ThongTin"].cell(row, 1).value: workbook["ThongTin"].cell(row, 2).value
                    for row in range(2, workbook["ThongTin"].max_row + 1)
                }
                self.assertEqual(metadata["Nguồn GRPNO"], "LN01.GRPNO -> credit_groups.ma_to")
            finally:
                workbook.close()

    def test_ln01_shared_import_reads_file_once_and_updates_both_outputs(self) -> None:
        path = self._write_ln01_file()
        calls: list[Path] = []
        original = Path.read_bytes

        def counting_read_bytes(target: Path) -> bytes:
            if Path(target) == path:
                calls.append(path)
            return original(target)

        with patch.object(Path, "read_bytes", counting_read_bytes):
            result = import_credit_limit_file(
                self.repository,
                path,
                warn_days=30,
                reference_date=date(2024, 1, 20),
            )

        self.assertEqual(calls, [path])
        self.assertEqual(result.row_count, 2)
        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)
        report_repository = CreditReportRepository(self.database_path)
        with closing(report_repository.connect()) as connection:
            loan_count = connection.execute("SELECT COUNT(*) FROM credit_loan_period").fetchone()[0]
            total_balance = connection.execute("SELECT SUM(outstanding_balance) FROM credit_loan_period").fetchone()[0]
        self.assertEqual(loan_count, 4)
        self.assertEqual(total_balance, 1050)

    def test_credit_limit_import_requires_period(self) -> None:
        path = self._write_ln01_file("ln01_no_period.csv")

        with self.assertRaisesRegex(SummaryError, "Không xác định được kỳ"):
            import_credit_limit_file(
                self.repository,
                path,
                warn_days=30,
                reference_date=date(2024, 1, 20),
            )

    def test_credit_limit_period_inferred_from_source_filename(self) -> None:
        path = self._write_ln01_file("5491_ln01_20251231.csv")

        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )

        batch = self.repository.list_credit_limit_batches()[0]
        self.assertEqual(batch.period, "2025-12")
        self.assertEqual(batch.branch_code, "5491")

    def test_credit_limit_period_not_based_on_import_time(self) -> None:
        path = self._write_ln01_file("5491_ln01_20251231.csv")

        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2026, 8, 4),
        )

        self.assertEqual(self.repository.list_credit_limit_batches()[0].period, "2025-12")

    def test_credit_limit_metadata_contains_period(self) -> None:
        path = self._write_ln01_file("5491_ln01_20251231.csv")

        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )

        batch = self.repository.list_credit_limit_batches()[0]
        self.assertEqual(batch.period, "2025-12")
        self.assertTrue(batch.is_active)
        self.assertFalse(batch.period_missing)

    def test_credit_limit_filename_contains_period(self) -> None:
        path = self._write_ln01_file("5491_ln01_20251231.csv")

        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )

        file_name = self.repository.list_credit_limit_batches()[0].file_name
        self.assertTrue(file_name.startswith("HMHETHAN_2025-12_5491_"))

    def test_credit_limit_combo_displays_period(self) -> None:
        path = self._write_ln01_file("5491_ln01_20251231.csv")
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))
        tab = CreditLimitTab(self.repository)
        self.addCleanup(tab.deleteLater)

        tab.reload_filters()

        self.assertEqual(tab.batch_filter.itemText(0), "Tất cả kỳ")
        self.assertIn("2025-12", tab.batch_filter.itemText(1))
        self.assertNotIn(path.name, tab.batch_filter.itemText(1))

    def test_credit_limit_table_displays_period(self) -> None:
        path = self._write_ln01_file("5491_ln01_20251231.csv")
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        rows = self.repository.query_credit_limits(page_size=20).rows

        self.assertEqual(rows[0]["period"], "2025-12")

    def test_credit_limit_all_periods_uses_one_active_batch_per_period(self) -> None:
        store = CreditLimitExcelBatchStore(self.database_path)
        row = self._sample_credit_limit_row()
        store.create_batch(
            source_path=self.root / "5491_ln01_20251231.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="1" * 64,
            source_file_size=1,
            duplicate_policy="new",
        )
        store.create_batch(
            source_path=self.root / "5491_ln01_20251231_v2.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="2" * 64,
            source_file_size=1,
            duplicate_policy="new",
            period="2025-12",
            branch_code="5491",
        )

        self.assertEqual(len(store.list_all_batches()), 2)
        self.assertEqual(len(store.list_batches()), 1)
        self.assertEqual(store.query_credit_limits(page_size=20).total_rows, 1)

    def test_credit_limit_same_period_offers_overwrite(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20251231.csv", scale=1.0)
        second_path = self._write_report_ln01_file("5491_ln01_20251231_v2.csv", scale=2.0)
        import_credit_limit_file(self.repository, first_path, period="2025-12", warn_days=30, reference_date=date(2024, 1, 20))

        prepared = Ln01ImportCoordinator(self.repository).prepare_import(
            second_path,
            period="2025-12",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )

        self.assertTrue(prepared.context.resolution.requires_confirmation)
        self.assertIn(Ln01DuplicateDecision.OVERWRITE_BOTH, prepared.context.resolution.allowed_decisions)
        self.assertIn(Ln01DuplicateDecision.OVERWRITE_CREDIT, prepared.context.resolution.allowed_decisions)

    def test_credit_limit_overwrite_does_not_create_duplicate_active_batch(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20251231.csv", scale=1.0)
        second_path = self._write_report_ln01_file("5491_ln01_20251231_v2.csv", scale=2.0)
        import_credit_limit_file(self.repository, first_path, period="2025-12", warn_days=30, reference_date=date(2024, 1, 20))

        import_credit_limit_file(
            self.repository,
            second_path,
            period="2025-12",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.OVERWRITE_BOTH,
        )

        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)

    def test_credit_limit_overwrite_archives_previous_version(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20251231.csv", scale=1.0)
        second_path = self._write_report_ln01_file("5491_ln01_20251231_v2.csv", scale=2.0)
        import_credit_limit_file(self.repository, first_path, period="2025-12", warn_days=30, reference_date=date(2024, 1, 20))

        import_credit_limit_file(
            self.repository,
            second_path,
            period="2025-12",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.OVERWRITE_BOTH,
        )

        history_files = list((self.database_path.parent / "HMHETHAN" / "History").glob("*.xlsx"))
        self.assertEqual(len(history_files), 1)

    def test_credit_limit_legacy_batch_period_migration(self) -> None:
        path = self._write_ln01_file("5491_ln01_20251231.csv")
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))
        batch = self.repository.list_credit_limit_batches()[0]
        workbook = load_workbook(batch.file_path)
        try:
            sheet = workbook[META_SHEET_NAME]
            for row in sheet.iter_rows(min_row=2, max_col=2):
                if row[0].value == "period":
                    row[1].value = ""
                if row[0].value == "branch_code":
                    row[1].value = ""
            workbook.save(batch.file_path)
        finally:
            workbook.close()
        self.repository.credit_limit_store.invalidate_cache()

        migrated = self.repository.list_credit_limit_batches()[0]

        self.assertEqual(migrated.period, "2025-12")
        self.assertEqual(migrated.branch_code, "5491")

    def test_credit_limit_period_matches_credit_database_period(self) -> None:
        path = self._write_ln01_file("5491_ln01_20251231.csv")

        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            row = connection.execute(
                "SELECT period FROM credit_import_runs WHERE source_type = 'LN01'"
            ).fetchone()
        self.assertEqual(row["period"], "2025-12")

    def test_report_total_balance_includes_card_and_card_is_short_term(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        report_repository = CreditReportRepository(self.database_path)
        card_row = NormalizedLoanRow(
            period="2026-04",
            source_file="5491_FTPLN_20260430.csv",
            source_row_number=2,
            branch_code="5491",
            trctcd="00",
            transaction_office="Hoi so",
            customer_sequence="0001",
            customer_code="54910001",
            customer_name="Khach 0001",
            customer_type="CN",
            ftp_code="DN15",
            balance=25,
            ftp=0,
            interest_rate=0,
            ftp_adjustment=0,
            officer_code="CB01",
            officer_name="Nguyen Van A",
            debt_group_code="01",
            debt_group_number=1,
            debt_group_category="NORMAL",
            has_valid_debt_group=True,
        )
        report_repository.save_credit_card_projection(
            period="2026-04",
            file_name="5491_FTPLN_20260430.csv",
            source_sha256="card-sha",
            source_file_size=100,
            source_row_count=1,
            rows=[card_row],
        )

        summary = report_repository.overall_summary(CreditReportFilters(period="2026-04"))
        self.assertEqual(summary["ln01_total_balance"], 650)
        self.assertEqual(summary["credit_card_balance"], 25)
        self.assertEqual(summary["total_balance"], 675)
        term_rows = report_repository.term_structure_rows(CreditReportFilters(period="2026-04"))
        short = next(row for row in term_rows if row["Loại thời hạn"] == "Ngắn hạn")
        self.assertEqual(short["Dư nợ thẻ"], 25)
        self.assertEqual(short["Tổng dư nợ"], 175)

    def test_report_customer_count_distinct_and_excludes_zero_balance(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        summary = CreditReportRepository(self.database_path).overall_summary(CreditReportFilters(period="2026-04"))
        self.assertEqual(summary["customer_count"], 3)

    def test_report_customer_type_rules_and_ratios(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        repository = CreditReportRepository(self.database_path)
        rows = repository.customer_type_rows(CreditReportFilters(period="2026-04"))
        by_label = {row["Loại khách hàng"]: row for row in rows}
        self.assertEqual(by_label["Cá nhân"]["Tổng dư nợ"], 450)
        self.assertEqual(by_label["Pháp nhân"]["Tổng dư nợ"], 200)
        repository.save_personal_type_rule("999")
        rows = repository.customer_type_rows(CreditReportFilters(period="2026-04"))
        by_label = {row["Loại khách hàng"]: row for row in rows}
        self.assertEqual(by_label["Cá nhân"]["Tổng dư nợ"], 650)

    def test_report_debt_group_decree55_and_industry_rules(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        repository = CreditReportRepository(self.database_path)
        debt_rows = {row["Nhóm nợ"]: row for row in repository.debt_group_rows(CreditReportFilters(period="2026-04"))}
        self.assertEqual(debt_rows["Nợ nhóm 1"]["Tổng dư nợ"], 150)
        self.assertEqual(debt_rows["Nợ nhóm 2 - Nợ cần chú ý"]["Tổng dư nợ"], 200)
        self.assertEqual(debt_rows["Nợ nhóm 3-5 - Nợ xấu"]["Tổng dư nợ"], 300)
        decree55 = repository.decree55_rows(CreditReportFilters(period="2026-04"))[0]
        self.assertEqual(decree55["Tổng dư nợ"], 450)
        industry = {row["Nhóm ngành"]: row for row in repository.industry_rows(CreditReportFilters(period="2026-04"))}
        self.assertEqual(industry["Cho vay ứng dụng công nghệ cao"]["Tổng dư nợ"], 150)
        self.assertEqual(industry["Cho vay bán buôn, bán lẻ"]["Tổng dư nợ"], 200)
        self.assertEqual(industry["Cho vay bất động sản"]["Tổng dư nợ"], 300)

    def test_report_classification_helpers(self) -> None:
        self.assertEqual(classify_term_from_account("211001"), "SHORT")
        self.assertEqual(classify_term_from_account("212001"), "MEDIUM")
        self.assertEqual(classify_term_from_account("251001"), "MEDIUM")
        self.assertEqual(classify_term_from_account("213001"), "LONG")
        self.assertEqual(classify_term_from_account("252001"), "LONG")
        self.assertEqual(classify_term_from_account("999001"), "UNKNOWN")
        self.assertEqual(classify_industry("10104"), "HIGH_TECH")
        self.assertEqual(classify_industry("100101"), "WHOLESALE_RETAIL")
        self.assertEqual(classify_industry("101301"), "WHOLESALE_RETAIL")
        self.assertEqual(classify_industry("1001"), "OTHER")
        self.assertEqual(classify_industry("812345"), "REAL_ESTATE")
        self.assertEqual(classify_industry("240117"), "REAL_ESTATE")

    def test_report_compare_summary_between_selected_periods(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20260331.csv", scale=0.5)
        second_path = self._write_report_ln01_file("5491_ln01_20260430.csv", scale=1.0)
        import_credit_limit_file(
            self.repository,
            first_path,
            period="2026-03",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        import_credit_limit_file(
            self.repository,
            second_path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        repository = CreditReportRepository(self.database_path)
        filters = CreditReportFilters(from_period="2026-03", to_period="2026-04")
        from_snapshot = get_report_snapshot(repository, "2026-03", filters)
        to_snapshot = get_report_snapshot(repository, "2026-04", filters)
        result = compare_report_snapshots(from_snapshot, to_snapshot, GROUP_SUMMARY)
        rows = {row.get("Chỉ tiêu"): row for row in result.rows}
        self.assertEqual(rows["Tổng dư nợ"].get("Giá trị Từ kỳ"), 325)
        self.assertEqual(rows["Tổng dư nợ"].get("Giá trị Đến kỳ"), 650)
        self.assertEqual(rows["Tổng dư nợ"].get("Tăng/giảm tuyệt đối"), 325)
        self.assertEqual(rows["Tổng dư nợ"].get("Tăng trưởng (%)"), 100)

    def test_report_compare_same_period(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        repository = CreditReportRepository(self.database_path)
        filters = CreditReportFilters(from_period="2026-04", to_period="2026-04")
        snapshot = get_report_snapshot(repository, "2026-04", filters)
        result = compare_report_snapshots(snapshot, snapshot, GROUP_SUMMARY)
        rows = {row.get("Chỉ tiêu"): row for row in result.rows}
        self.assertEqual(rows["Tổng dư nợ"].get("Tăng/giảm tuyệt đối"), 0)
        self.assertEqual(rows["Tổng dư nợ"].get("Tăng trưởng (%)"), 0)
        self.assertIn("Từ kỳ và Đến kỳ đang giống nhau.", result.notes)

    def test_report_export_current_mode_sheets(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        output = self.root / "report.xlsx"
        CreditReportRepository(self.database_path).export_workbook(
            output,
            filters=CreditReportFilters(period="2026-04", from_period="2026-04", to_period="2026-04"),
            view_mode=VIEW_CURRENT_PERIOD,
        )
        workbook = load_workbook(output, data_only=True)
        try:
            self.assertEqual(
                workbook.sheetnames,
                [
                    "TongHop",
                    "CoCauThoiHan",
                    "LoaiHinhKhachHang",
                    "ChatLuongTinDung",
                    "NghiDinh55",
                    "NganhKinhTe",
                    "ThongTin",
                ],
            )
        finally:
            workbook.close()

    def test_report_export_compare_mode_sheets(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20260331.csv", scale=0.5)
        second_path = self._write_report_ln01_file("5491_ln01_20260430.csv", scale=1.0)
        import_credit_limit_file(
            self.repository,
            first_path,
            period="2026-03",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        import_credit_limit_file(
            self.repository,
            second_path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        output = self.root / "report_compare.xlsx"
        CreditReportRepository(self.database_path).export_workbook(
            output,
            filters=CreditReportFilters(from_period="2026-03", to_period="2026-04"),
            view_mode=VIEW_COMPARE_PERIODS,
        )
        workbook = load_workbook(output, data_only=True)
        try:
            self.assertEqual(
                workbook.sheetnames,
                [
                    "SoSanhTongHop",
                    "SoSanhCoCauThoiHan",
                    "SoSanhLoaiHinhKH",
                    "SoSanhChatLuongTinDung",
                    "SoSanhNghiDinh55",
                    "SoSanhNganhKinhTe",
                    "ThongTin",
                ],
            )
            self.assertNotIn("ChatLuongDuNo", workbook.sheetnames)
        finally:
            workbook.close()

    def test_credit_delete_period_removes_period_rows_and_import_metadata(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        repository = CreditReportRepository(self.database_path)
        result = repository.delete_report_period("2026-04")
        self.assertEqual(result["loan_rows"], 4)
        self.assertEqual(result["runs"], 1)
        self.assertEqual(result["files"], 1)
        with closing(repository.connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period").fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_card_period").fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs").fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_files").fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_customer_master").fetchone()[0], 0)

    def test_credit_delete_period_keeps_other_periods_and_referenced_customers(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20260331.csv", scale=0.5)
        second_path = self._write_report_ln01_file("5491_ln01_20260430.csv", scale=1.0)
        import_credit_limit_file(self.repository, first_path, period="2026-03", warn_days=30, reference_date=date(2024, 1, 20))
        import_credit_limit_file(self.repository, second_path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        repository = CreditReportRepository(self.database_path)
        repository.delete_report_period("2026-03")
        with closing(repository.connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-03'").fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 4)
            self.assertGreater(connection.execute("SELECT COUNT(*) FROM credit_customer_master").fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-03'").fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-04'").fetchone()[0], 1)

    def test_credit_delete_period_transaction_rollback(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        repository = CreditReportRepository(self.database_path)
        with patch.object(repository, "_delete_orphan_customers", side_effect=RuntimeError("boom")):
            with self.assertRaises(RuntimeError):
                repository.delete_report_period("2026-04")
        with closing(repository.connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 4)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-04'").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_files WHERE period = '2026-04'").fetchone()[0], 1)

    def test_credit_delete_period_does_not_delete_hmhethan_or_nim_data(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        self._write_nim_dn_file("5491_FTPLN_20260430.csv", ["5491,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)
        CreditReportRepository(self.database_path).delete_report_period("2026-04")
        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-04"), 1)

    def test_credit_diagnostics_reports_freelist_and_reclaimable_size(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        repository = CreditReportRepository(self.database_path)
        repository.delete_report_period("2026-04")
        diagnostics = repository.diagnostics()
        self.assertIn("freelist_count", diagnostics)
        self.assertEqual(diagnostics["reclaimable_bytes"], diagnostics["freelist_count"] * diagnostics["page_size"])
        self.assertEqual(diagnostics["tables"]["credit_loan_period"], 0)
        self.assertEqual(diagnostics["tables"]["credit_customer_master"], 0)

    def test_credit_compact_database_creates_backup_and_preserves_schema_rules(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        repository = CreditReportRepository(self.database_path)
        repository.delete_report_period("2026-04")
        result = repository.compact_database(backup_directory=self.root / "backup")
        self.assertTrue(result["vacuum"])
        self.assertTrue(Path(result["backup_path"]).is_file())
        self.assertEqual(result["integrity_check"], "ok")
        with closing(repository.connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_schema_migrations").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_customer_type_rules").fetchone()[0], 2)

    def test_credit_same_period_overwrite_no_duplicate_loans_or_customers(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20260430.csv", scale=1.0)
        second_path = self._write_report_ln01_file("5491_ln01_20260430_v2.csv", scale=2.0)
        import_credit_limit_file(
            self.repository,
            first_path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        import_credit_limit_file(
            self.repository,
            second_path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_policy="new",
            overwrite_report_period=True,
        )
        repository = CreditReportRepository(self.database_path)
        with closing(repository.connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 4)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_customer_master").fetchone()[0], 4)
            self.assertEqual(connection.execute("SELECT SUM(outstanding_balance) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 1300)
            row = connection.execute(
                "SELECT source_row_count FROM credit_loan_period WHERE period = '2026-04' AND account_number = '211001'"
            ).fetchone()
            self.assertEqual(row[0], 2)

    def test_ln01_no_hmhethan_no_credit_creates_both(self) -> None:
        path = self._write_report_ln01_file()

        result = import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )

        self.assertEqual(result.row_count, 0)
        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)
        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 4)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-04'").fetchone()[0], 1)

    def test_ln01_existing_hmhethan_missing_credit_allows_recreate(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        CreditReportRepository(self.database_path).delete_report_period("2026-04")

        result = import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        self.assertIn("Đã tạo lại dữ liệu báo cáo kỳ 2026-04", result.message)
        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 4)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-04' AND source_type = 'LN01'").fetchone()[0], 1)

    def test_ln01_existing_hmhethan_missing_credit_reuses_batch(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        original_batch = self.repository.list_credit_limit_batches()[0]
        original_path = original_batch.file_path
        original_hash = hashlib.sha256(original_path.read_bytes()).hexdigest()
        CreditReportRepository(self.database_path).delete_report_period("2026-04")

        result = import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        batch = self.repository.list_credit_limit_batches()[0]
        self.assertEqual(batch.file_path, original_path)
        self.assertEqual(batch.batch_id, original_batch.batch_id)
        self.assertEqual(hashlib.sha256(batch.file_path.read_bytes()).hexdigest(), original_hash)
        self.assertEqual(result.batch_id, original_batch.batch_id)

    def test_ln01_existing_hmhethan_missing_credit_does_not_create_duplicate_excel(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        CreditReportRepository(self.database_path).delete_report_period("2026-04")

        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)
        files = list((self.database_path.parent / "HMHETHAN").glob("*.xlsx"))
        self.assertEqual(len(files), 1)

    def test_ln01_existing_hmhethan_existing_credit_offers_overwrite(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        coordinator = Ln01ImportCoordinator(self.repository)

        prepared = coordinator.prepare_import(path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        self.assertTrue(prepared.context.resolution.requires_confirmation)
        self.assertEqual(prepared.context.resolution.default_decision, Ln01DuplicateDecision.OVERWRITE_CREDIT)
        self.assertIn(Ln01DuplicateDecision.OVERWRITE_BOTH, prepared.context.resolution.allowed_decisions)

    def test_ln01_overwrite_credit_keeps_hmhethan_file(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        batch = self.repository.list_credit_limit_batches()[0]
        mtime = batch.file_path.stat().st_mtime_ns
        digest = hashlib.sha256(batch.file_path.read_bytes()).hexdigest()

        result = import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.OVERWRITE_CREDIT,
        )

        self.assertIn("Batch Hạn mức hiện có không bị thay đổi", result.message)
        self.assertEqual(batch.file_path.stat().st_mtime_ns, mtime)
        self.assertEqual(hashlib.sha256(batch.file_path.read_bytes()).hexdigest(), digest)
        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 4)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-04' AND source_type = 'LN01'").fetchone()[0], 1)

    def test_ln01_overwrite_both_replaces_hmhethan_atomically(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        batch = self.repository.list_credit_limit_batches()[0]
        os.utime(batch.file_path, (1_700_000_000, 1_700_000_000))
        old_mtime = batch.file_path.stat().st_mtime_ns

        result = import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.OVERWRITE_BOTH,
        )

        self.assertIn("Đã ghi đè dữ liệu báo cáo và batch Hạn mức kỳ 2026-04", result.message)
        replacement = self.repository.list_credit_limit_batches()[0]
        self.assertEqual(replacement.file_path, batch.file_path)
        self.assertNotEqual(replacement.file_path.stat().st_mtime_ns, old_mtime)
        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            row = connection.execute("SELECT action FROM credit_action_log ORDER BY id DESC LIMIT 1").fetchone()
        self.assertEqual(row["action"], "OVERWRITE_BOTH")

    def test_ln01_missing_hmhethan_existing_credit_can_create_batch_only(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        batch = self.repository.list_credit_limit_batches()[0]
        self.repository.delete_credit_limit_batch(batch.batch_id)

        result = import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.CREATE_HMHE_THAN_ONLY,
        )

        self.assertIn("Dữ liệu Credit.db không thay đổi", result.message)
        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)
        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-04' AND source_type = 'LN01'").fetchone()[0], 1)
            row = connection.execute("SELECT action FROM credit_action_log ORDER BY id DESC LIMIT 1").fetchone()
        self.assertEqual(row["action"], "CREATE_HMHEthan_ONLY")

    def test_ln01_same_period_different_sha_requires_confirmation(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20260430.csv", scale=1.0)
        second_path = self._write_report_ln01_file("5491_ln01_20260430_v2.csv", scale=2.0)
        import_credit_limit_file(self.repository, first_path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        with self.assertRaisesRegex(SummaryError, "file LN01 khác|Cần xác nhận"):
            import_credit_limit_file(self.repository, second_path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)
        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            self.assertEqual(connection.execute("SELECT SUM(outstanding_balance) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 650)

    def test_ln01_cancel_changes_nothing(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        before_batches = len(self.repository.list_credit_limit_batches())

        result = import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.CANCEL,
        )

        self.assertEqual(result.row_count, 0)
        self.assertEqual(len(self.repository.list_credit_limit_batches()), before_batches)
        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-04' AND source_type = 'LN01'").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 4)

    def test_replace_ln01_period_preserves_credit_card_rows(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        card_rows = [
            NormalizedLoanRow(
                period="2026-04",
                source_file="5491_FTPLN_20260430.csv",
                source_row_number=2,
                branch_code="5491",
                trctcd="00",
                transaction_office="Hoi so",
                customer_sequence="0001",
                customer_code="54910001",
                customer_name="Khach 0001",
                customer_type="CN",
                ftp_code="DN15",
                balance=25,
                ftp=0,
                interest_rate=0,
                ftp_adjustment=0,
                officer_code="CB01",
                officer_name="Nguyen Van A",
                debt_group_code="01",
                debt_group_number=1,
                debt_group_category="NORMAL",
                has_valid_debt_group=True,
            ),
            NormalizedLoanRow(
                period="2026-04",
                source_file="5491_FTPLN_20260430.csv",
                source_row_number=3,
                branch_code="5491",
                trctcd="00",
                transaction_office="Hoi so",
                customer_sequence="0099",
                customer_code="54910099",
                customer_name="Card Only",
                customer_type="CN",
                ftp_code="DN15",
                balance=75,
                ftp=0,
                interest_rate=0,
                ftp_adjustment=0,
                officer_code="CB99",
                officer_name="Card Officer",
                debt_group_code="01",
                debt_group_number=1,
                debt_group_category="NORMAL",
                has_valid_debt_group=True,
            ),
        ]
        report_repository = CreditReportRepository(self.database_path)
        report_repository.save_credit_card_projection(
            period="2026-04",
            file_name="5491_FTPLN_20260430.csv",
            source_sha256="card-sha",
            source_file_size=100,
            source_row_count=2,
            rows=card_rows,
        )

        import_credit_limit_file(
            self.repository,
            path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.OVERWRITE_CREDIT,
        )

        summary = report_repository.overall_summary(CreditReportFilters(period="2026-04"))
        self.assertEqual(summary["ln01_total_balance"], 650)
        self.assertEqual(summary["credit_card_balance"], 100)
        self.assertEqual(summary["total_balance"], 750)
        with closing(report_repository.connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_card_period WHERE period = '2026-04'").fetchone()[0], 2)
            self.assertIsNotNone(
                connection.execute(
                    "SELECT 1 FROM credit_customer_master WHERE customer_code = '54910099'"
                ).fetchone()
            )

    def test_replace_ln01_period_does_not_append_and_updates_metadata(self) -> None:
        first_path = self._write_report_ln01_file("5491_ln01_20260430.csv", scale=1.0)
        second_path = self._write_report_ln01_file("5491_ln01_20260430_v2.csv", scale=2.0)
        import_credit_limit_file(self.repository, first_path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        import_credit_limit_file(
            self.repository,
            second_path,
            period="2026-04",
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_decision=Ln01DuplicateDecision.CREATE_BOTH,
        )

        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 4)
            self.assertEqual(connection.execute("SELECT SUM(outstanding_balance) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 1300)
            row = connection.execute("SELECT source_file_name FROM credit_import_runs WHERE period = '2026-04' AND source_type = 'LN01'").fetchone()
        self.assertEqual(row["source_file_name"], second_path.name)

    def test_overwrite_both_failure_restores_old_batch(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        batch = self.repository.list_credit_limit_batches()[0]
        original_hash = hashlib.sha256(batch.file_path.read_bytes()).hexdigest()

        with patch.object(CreditReportRepository, "log_ln01_import_action", side_effect=RuntimeError("boom")):
            with self.assertRaises(SummaryError):
                import_credit_limit_file(
                    self.repository,
                    path,
                    period="2026-04",
                    warn_days=30,
                    reference_date=date(2024, 1, 20),
                    duplicate_decision=Ln01DuplicateDecision.OVERWRITE_BOTH,
                )

        self.assertEqual(hashlib.sha256(batch.file_path.read_bytes()).hexdigest(), original_hash)
        with closing(CreditReportRepository(self.database_path).connect()) as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM credit_import_runs WHERE period = '2026-04' AND source_type = 'LN01'").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT SUM(outstanding_balance) FROM credit_loan_period WHERE period = '2026-04'").fetchone()[0], 650)

    def test_duplicate_dialog_has_explicit_actions(self) -> None:
        path = self._write_report_ln01_file()
        import_credit_limit_file(self.repository, path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))
        prepared = Ln01ImportCoordinator(self.repository).prepare_import(path, period="2026-04", warn_days=30, reference_date=date(2024, 1, 20))

        actions = _dialog_actions(prepared.context)
        labels = [item[0] for item in actions]
        decisions = [item[1] for item in actions]

        self.assertIn("Ghi đè dữ liệu kỳ báo cáo", labels)
        self.assertIn("Ghi đè cả dữ liệu kỳ và batch Hạn mức", labels)
        self.assertIn(Ln01DuplicateDecision.OVERWRITE_CREDIT, decisions)
        self.assertIn(Ln01DuplicateDecision.OVERWRITE_BOTH, decisions)
        self.assertIn(Ln01DuplicateDecision.CANCEL, decisions)

    def test_import_from_credit_limit_uses_same_duplicate_logic(self) -> None:
        names = self._code_names(CreditLimitTab.import_file)
        self.assertIn("prepare_import", names)
        self.assertIn("execute_prepared_import", names)

    def test_import_from_report_summary_uses_same_duplicate_logic(self) -> None:
        names = self._code_names(ReportSummaryWindow.import_ln01)
        self.assertIn("prepare_import", names)
        self.assertIn("execute_prepared_import", names)

    def _code_names(self, function: object) -> set[str]:
        code = function.__code__
        names = set(code.co_names)
        for item in code.co_consts:
            if hasattr(item, "co_names"):
                names.update(item.co_names)
        return names

    def _sample_credit_limit_row(self) -> CreditLimitRow:
        return CreditLimitRow(
            customer_code="00123",
            customer_name="Khach 00123",
            contract_number="HD001",
            approved_date=date(2024, 1, 1),
            approved_amount=1_000_000,
            outstanding_balance=250_000,
            expiry_date=date(2024, 1, 15),
            address="Dia chi",
            officer="CB1",
            note="",
            days_to_expiry=None,
            status="",
            branch_code="5491",
            account_number="0000123456",
            credit_line_type="Line of Credit",
            source_row_count=1,
        )

    def _set_sample_nim_kpis(self, tab: NimTab) -> list[CompactKpiCard]:
        tab.metrics.set_data(
            DashboardData(
                metrics=(
                    DashboardMetric(tab.ui_config.total_balance_label, "28.056.021.646.654"),
                    DashboardMetric("NIM trước ĐC", "2,31%"),
                    DashboardMetric("NIM sau ĐC", "2,40%"),
                    DashboardMetric("Lãi suất bình quân", "8,01%"),
                )
            )
        )
        return tab.metrics.findChildren(CompactKpiCard)

    def _assert_button_text_fits(self, button: QPushButton, text: str) -> None:
        button.setText(text)
        required = recommended_control_height(button)
        self.assertGreaterEqual(button.minimumHeight(), required)
        self.assertGreaterEqual(max(button.sizeHint().height(), button.minimumHeight()), required)
        self.assertEqual(button.maximumHeight(), 16777215)

    def _assert_button_width_fits(self, button: QPushButton) -> None:
        self.assertGreaterEqual(max(button.minimumWidth(), button.sizeHint().width()), recommended_button_width(button))

    def _nim_combo_with_long_officer(self) -> QComboBox:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)
        combo = tab.officer_filter
        _populate_combo(
            combo,
            [("Nguyễn Văn A Phòng GD Trung tâm", "540000321")],
        )
        return combo

    def test_nim_dn_and_nv_import_use_vba_formulas(self) -> None:
        dn = self.root / "5491_FTPLN_20260331.csv"
        dn.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,CB1,00,1000,,CN",
                ]
            ),
            encoding="utf-8",
        )
        nv = self.root / "5491_FTPDP_20260331.csv"
        nv.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBHD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,CB2,00,1000,,TC",
                ]
            ),
            encoding="utf-8",
        )

        dn_output = self.root / "BaoCaoNIM_CSDL.xlsx"
        nv_output = self.root / "BaoCaoNIM_NV_CSDL.xlsx"
        result_dn = import_nim_dn(self.repository, self.root, export_path=dn_output)
        result_nv = import_nim_nv(self.repository, self.root, export_path=nv_output)

        self.assertEqual(result_dn.row_count, 1)
        self.assertEqual(result_nv.row_count, 1)
        self.assertEqual(result_dn.output_path, dn_output)
        self.assertEqual(result_nv.output_path, nv_output)
        page_dn = self.repository.query_nim(SummaryDataType.NIM_DN)
        self.assertEqual(page_dn.total_rows, 1)
        row_dn = page_dn.rows[0]
        self.assertEqual(row_dn["period"], "2026-03")
        self.assertAlmostEqual(float(row_dn["average_rate"]), 10.0)
        self.assertAlmostEqual(float(row_dn["nim_before"]), 8.0)
        self.assertAlmostEqual(float(row_dn["nim_after"]), 7.0)

        page_nv = self.repository.query_nim(SummaryDataType.NIM_NV)
        row_nv = page_nv.rows[0]
        self.assertAlmostEqual(float(row_nv["nim_before"]), -8.0)
        self.assertAlmostEqual(float(row_nv["nim_after"]), -7.0)
        workbook_dn = load_workbook(dn_output, data_only=True)
        try:
            self.assertEqual(workbook_dn.sheetnames, ["Cache_Nim", "Báo Cáo NIM DN", "Sheet1", "Sheet2", "Sheet3"])
            self.assertEqual(workbook_dn["Cache_Nim"].sheet_state, "hidden")
            self.assertEqual(
                [workbook_dn["Cache_Nim"].cell(1, col).value for col in range(1, 9)],
                list(NIM_DN_DETAIL_HEADERS),
            )
            self.assertEqual(
                [workbook_dn["Báo Cáo NIM DN"].cell(15, col).value for col in range(1, 9)],
                list(NIM_DN_DETAIL_HEADERS),
            )
        finally:
            workbook_dn.close()
        workbook_nv = load_workbook(nv_output, data_only=True)
        try:
            self.assertEqual(workbook_nv.sheetnames, ["Cache_Nim_NV", "Báo Cáo NIM NV", "Sheet1", "Sheet2", "Sheet3"])
            self.assertEqual(workbook_nv["Cache_Nim_NV"].sheet_state, "hidden")
            self.assertEqual(
                [workbook_nv["Cache_Nim_NV"].cell(1, col).value for col in range(1, 8)],
                list(NIM_NV_DETAIL_HEADERS),
            )
            self.assertEqual(
                [workbook_nv["Báo Cáo NIM NV"].cell(15, col).value for col in range(1, 8)],
                list(NIM_NV_DETAIL_HEADERS),
            )
        finally:
            workbook_nv.close()

    def test_nim_branch_combo_uses_unit_directory(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Động", short_name="CN Động")
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)

        labels = [tab.branch_filter.itemText(index) for index in range(tab.branch_filter.count())]

        self.assertIn("6501 - CN Động", labels)
        self.assertEqual(tab.branch_filter.itemData(labels.index("6501 - CN Động")), "6501")

    def test_nim_office_combo_uses_unit_directory(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Động", short_name="CN Động")
        )
        self.repository.unit_directory.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="6501",
                trctcd="07",
                office_code="",
                office_name="Phòng giao dịch Kiểm thử",
                short_name="PGD Kiểm thử",
                office_type=TRANSACTION_OFFICE,
            )
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,07,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)

        labels = [tab.transaction_office_filter.itemText(index) for index in range(tab.transaction_office_filter.count())]

        self.assertIn("6501-07 - PGD Kiểm thử", labels)
        self.assertEqual(tab.transaction_office_filter.itemData(labels.index("6501-07 - PGD Kiểm thử")), "6501-07")

    def test_table_branch_name_changes_after_directory_update(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Cũ", short_name="CN Cũ")
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Mới", short_name="CN Mới")
        )

        row = self.repository.query_nim(SummaryDataType.NIM_DN).rows[0]

        self.assertEqual(row["branch"], "6501 - CN Mới")

    def test_officer_history_uses_current_branch_name_after_directory_update(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Cũ", short_name="CN Cũ")
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,[650100001] Nguyễn Văn A,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Mới", short_name="CN Mới")
        )

        history = build_officer_history(
            self.repository,
            SummaryDataType.NIM_DN,
            officer="[650100001] Nguyễn Văn A",
            branch="6501 - CN Mới",
            transaction_office="Hội sở",
        )

        self.assertEqual(history.current_period, "2026-03")
        self.assertAlmostEqual(history.current_balance, 1000.0)

    def test_export_uses_current_branch_name(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Xuất", short_name="CN Xuất")
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Xuất đổi", short_name="CN Xuất đổi")
        )
        data = build_nim_dashboard(
            NimDashboardRepository(self.repository),
            SummaryDataType.NIM_DN,
            DashboardFilters(branch="6501"),
        )
        output = self.root / "nim-dashboard.xlsx"

        DashboardNimExportService(data).export_detail(output)
        workbook = load_workbook(output, data_only=True)
        try:
            sheet = workbook[NIM_DN_UI_CONFIG.dashboard_sheets["detail"]]
            names = [sheet.cell(row, 2).value for row in range(2, sheet.max_row + 1)]
        finally:
            workbook.close()

        self.assertIn("6501 - CN Xuất đổi", names)

    def test_export_reflects_branch_rename_without_reimport(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Export cũ", short_name="CN Export cũ")
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Export mới", short_name="CN Export mới")
        )

        data = build_nim_dashboard(
            NimDashboardRepository(self.repository),
            SummaryDataType.NIM_DN,
            DashboardFilters(branch="6501"),
        )
        output = self.root / "nim-export-rename.xlsx"
        DashboardNimExportService(data).export_detail(output)
        workbook = load_workbook(output, data_only=True)
        try:
            sheet = workbook[NIM_DN_UI_CONFIG.dashboard_sheets["detail"]]
            exported_names = [sheet.cell(row, 2).value for row in range(2, sheet.max_row + 1)]
        finally:
            workbook.close()

        self.assertIn("6501 - CN Export mới", exported_names)

    def test_unit_directory_changed_refreshes_open_windows(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh đang mở", short_name="CN đang mở")
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        app = QApplication.instance() or QApplication([])
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(window.close)

        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh đã đổi", short_name="CN đã đổi")
        )
        app.processEvents()

        labels = [window.branch_combo.itemText(index) for index in range(window.branch_combo.count())]
        self.assertIn("6501 - CN đã đổi", labels)

    def test_unit_directory_changed_preserves_selected_code(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh chọn", short_name="CN chọn")
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        app = QApplication.instance() or QApplication([])
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(window.close)
        index = window.branch_combo.findData("6501")
        self.assertGreaterEqual(index, 0)
        window.branch_combo.setCurrentIndex(index)

        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh chọn mới", short_name="CN chọn mới")
        )
        app.processEvents()

        self.assertEqual(window.branch_combo.currentData(), "6501")

    def test_no_runtime_legacy_mapping_lookup(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Triển khai", short_name="CN Triển khai")
        )
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)

        row = self.repository.query_nim(SummaryDataType.NIM_DN).rows[0]

        self.assertEqual(row["branch"], "6501 - CN Triển khai")

    def test_import_unknown_branch_does_not_fail(self) -> None:
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])

        result = import_nim_dn(self.repository, self.root)

        self.assertEqual(result.row_count, 1)
        self.assertIsNotNone(self.repository.unit_directory.get_branch("6501"))

    def test_import_unknown_office_does_not_fail(self) -> None:
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,03,1000,,CN"])

        result = import_nim_dn(self.repository, self.root)

        self.assertEqual(result.row_count, 1)
        self.assertIsNotNone(self.repository.unit_directory.get_office("6501", "03"))

    def test_unknown_branch_placeholder_or_warning(self) -> None:
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,00,1000,,CN"])

        result = import_nim_dn(self.repository, self.root)

        self.assertIn("6501", result.message)
        self.assertEqual(self.repository.unit_directory.get_branch_display_name("6501"), "6501 - CN chưa khai báo")

    def test_unknown_office_not_assigned_wrong_name(self) -> None:
        self._write_nim_dn_file("6501_FTPLN_20260331.csv", ["6501,2,10,1,CB1,03,1000,,CN"])
        import_nim_dn(self.repository, self.root)

        self.assertEqual(self.repository.unit_directory.get_office_display_name("6501", "03"), "6501-03 - PGD 03")

    def test_nim_dn_result_unchanged(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260331.csv", ["5491,2,10,1,CB1,00,1000,,CN"])

        import_nim_dn(self.repository, self.root)
        row = self.repository.query_nim(SummaryDataType.NIM_DN).rows[0]

        self.assertAlmostEqual(float(row["average_rate"]), 10.0)
        self.assertAlmostEqual(float(row["nim_before"]), 8.0)
        self.assertAlmostEqual(float(row["nim_after"]), 7.0)

    def test_nim_nv_result_unchanged(self) -> None:
        self._write_nim_nv_file("5491_FTPDP_20260331.csv", ["5491,2,10,1,CB1,00,1000,,CN"])

        import_nim_nv(self.repository, self.root)
        row = self.repository.query_nim(SummaryDataType.NIM_NV).rows[0]

        self.assertAlmostEqual(float(row["nim_before"]), -8.0)
        self.assertAlmostEqual(float(row["nim_after"]), -7.0)

    def test_nim_ui_display_format_helpers(self) -> None:
        self.assertEqual(_display_officer_name("[540000321] Nguyễn Văn A"), "Nguyễn Văn A")
        self.assertEqual(_display_officer_name("Nguyễn Văn B"), "Nguyễn Văn B")
        self.assertEqual(_format_money_vn(1234567890), "1.234.567.890")
        self.assertEqual(_format_percent_vn(7.15), "7,15%")
        self.assertEqual(_format_percent_vn(2.34), "2,34%")
        self.assertEqual(_format_percent_vn(0.58), "0,58%")

    def test_summary_buttons_use_shared_style(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tabs = (
            NimTab(self.repository, SummaryDataType.NIM_DN),
            LoanCompareTab(self.repository),
            CreditLimitTab(self.repository),
        )
        try:
            for tab in tabs:
                buttons = [button for button in tab.findChildren(QPushButton) if button.text()]
                self.assertTrue(buttons)
                self.assertTrue(
                    all(button.objectName() in {"PrimaryButton", "SecondaryButton", "DangerButton", "KpiToggleButton"} for button in buttons),
                    [f"{button.text()}={button.objectName()}" for button in buttons],
                )
        finally:
            for tab in tabs:
                tab.deleteLater()

    def test_nim_dn_toolbar_has_no_customer_data_button(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)
        button_texts = [button.text() for button in tab.findChildren(QPushButton) if button.text()]
        self.assertNotIn("Dữ liệu khách hàng", button_texts)

    def test_nim_dn_toolbar_has_no_debt_group_analysis_button(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)
        button_texts = [button.text() for button in tab.findChildren(QPushButton) if button.text()]
        self.assertNotIn("Phân tích nhóm nợ", button_texts)

    def test_nim_dn_dashboard_button_still_exists(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)
        button_texts = [button.text() for button in tab.findChildren(QPushButton) if button.text()]
        self.assertIn("Mở Dashboard", button_texts)

    def test_nim_dn_remaining_buttons_have_visible_text(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)
        for button in [button for button in tab.findChildren(QPushButton) if button.text()]:
            self.assertGreaterEqual(max(button.sizeHint().width(), button.minimumWidth()), button.fontMetrics().horizontalAdvance(button.text()) + 18)

    def test_nim_nv_toolbar_unchanged(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_NV)
        self.addCleanup(tab.deleteLater)
        button_texts = {button.text() for button in tab.findChildren(QPushButton) if button.text()}
        self.assertTrue(
            {
                "Import thư mục",
                "Xuất",
                "Làm mới",
                "Sao lưu",
                "Khôi phục",
                "Bảo trì dữ liệu",
                "Xóa kỳ dữ liệu",
                "Mở Dashboard",
            }.issubset(button_texts)
        )
        self.assertNotIn("Dữ liệu khách hàng", button_texts)
        self.assertNotIn("Phân tích nhóm nợ", button_texts)

    def test_nim_dashboard_apply_button_text_visible(self) -> None:
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        try:
            self._assert_button_width_fits(window.apply_button)
        finally:
            window.close()
            window.deleteLater()

    def test_nim_dashboard_clear_button_text_visible(self) -> None:
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        try:
            self._assert_button_width_fits(window.clear_button)
        finally:
            window.close()
            window.deleteLater()

    def test_nim_dashboard_export_button_text_visible(self) -> None:
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        try:
            self._assert_button_width_fits(window.export_all_button)
        finally:
            window.close()
            window.deleteLater()

    def test_nim_dashboard_close_button_text_visible(self) -> None:
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        try:
            self._assert_button_width_fits(window.close_button)
        finally:
            window.close()
            window.deleteLater()

    def test_nim_dashboard_toolbar_wraps(self) -> None:
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        try:
            toolbar = window.findChild(QWidget, "NimDashboardToolbar")
            self.assertIsNotNone(toolbar)
            layout = toolbar.layout()
            self.assertTrue(layout.hasHeightForWidth())
            self.assertGreater(layout.heightForWidth(520), layout.heightForWidth(1600))
        finally:
            window.close()
            window.deleteLater()

    def test_nim_dashboard_has_no_horizontal_scroll(self) -> None:
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        try:
            toolbar = window.findChild(QWidget, "NimDashboardToolbar")
            self.assertIsNotNone(toolbar)
            self.assertEqual(toolbar.findChildren(QScrollArea), [])
        finally:
            window.close()
            window.deleteLater()

    def test_nim_source_toolbar_buttons_visible(self) -> None:
        for data_type in (SummaryDataType.NIM_DN, SummaryDataType.NIM_NV):
            tab = NimTab(self.repository, data_type)
            try:
                for button in [button for button in tab.findChildren(QPushButton) if button.text()]:
                    self._assert_button_width_fits(button)
            finally:
                tab.deleteLater()

    def test_nim_source_toolbar_wraps(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_NV)
        try:
            action_area = tab.findChild(QWidget, "SummaryDataActionArea")
            self.assertIsNotNone(action_area)
            layout = action_area.layout()
            self.assertTrue(layout.hasHeightForWidth())
            self.assertGreater(layout.heightForWidth(360), layout.heightForWidth(1800))
        finally:
            tab.deleteLater()

    def test_nim_dn_removed_duplicate_buttons_stay_removed(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            button_texts = [button.text() for button in tab.findChildren(QPushButton) if button.text()]
            self.assertNotIn("Dữ liệu khách hàng", button_texts)
            self.assertNotIn("Phân tích nhóm nợ", button_texts)
            self.assertIn("Mở Dashboard", button_texts)
        finally:
            tab.deleteLater()

    def test_nim_nv_toolbar_business_actions_unchanged(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_NV)
        try:
            button_texts = {button.text() for button in tab.findChildren(QPushButton) if button.text()}
            self.assertTrue(
                {
                    "Import thư mục",
                    "Xuất",
                    "Làm mới",
                    "Sao lưu",
                    "Khôi phục",
                    "Bảo trì dữ liệu",
                    "Xóa kỳ dữ liệu",
                    "Mở Dashboard",
                }.issubset(button_texts)
            )
            self.assertNotIn("Dữ liệu khách hàng", button_texts)
            self.assertNotIn("Phân tích nhóm nợ", button_texts)
        finally:
            tab.deleteLater()

    def test_action_button_size_uses_font_metrics(self) -> None:
        buttons = [
            primary_button("Áp dụng"),
            secondary_button("Xóa lọc"),
            secondary_button("Xuất toàn bộ"),
            secondary_button("Khôi phục"),
        ]
        try:
            for button in buttons:
                self._assert_button_width_fits(button)
        finally:
            for button in buttons:
                button.deleteLater()

    def test_action_button_descenders_not_clipped(self) -> None:
        buttons = [
            primary_button("Áp dụng"),
            secondary_button("Phòng giao dịch"),
            secondary_button("Khôi phục"),
            secondary_button("Đóng"),
        ]
        try:
            for button in buttons:
                self.assertGreaterEqual(button.minimumHeight(), recommended_control_height(button))
                self.assertGreaterEqual(max(button.sizeHint().height(), button.minimumHeight()), recommended_control_height(button))
        finally:
            for button in buttons:
                button.deleteLater()

    def test_summary_data_action_toolbar_wraps_when_narrow(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)
        action_area = tab.findChild(QWidget, "SummaryDataActionArea")
        self.assertIsNotNone(action_area)
        flow = action_area.layout()
        self.assertTrue(flow.hasHeightForWidth())
        self.assertGreaterEqual(flow.heightForWidth(360), flow.heightForWidth(1800))

    def test_summary_comboboxes_use_shared_style(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tabs = (
            NimTab(self.repository, SummaryDataType.NIM_DN),
            LoanCompareTab(self.repository),
            CreditLimitTab(self.repository),
        )
        try:
            combos = [combo for tab in tabs for combo in tab.findChildren(QComboBox)]
            self.assertTrue(combos)
            self.assertTrue(all(combo.objectName() == "AgribankComboBox" for combo in combos))
        finally:
            for tab in tabs:
                tab.deleteLater()

    def test_credit_limit_uses_unit_directory(self) -> None:
        self.repository.unit_directory.save_branch(
            BranchDirectoryEntry(branch_code="6501", branch_name="Chi nhánh Hạn mức", short_name="CN Hạn mức")
        )
        self.repository.unit_directory.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="6501",
                trctcd="00",
                office_code="6501-00",
                office_name="Hội sở Hạn mức",
                short_name="Hội sở Hạn mức",
                office_type="HEAD_OFFICE",
            )
        )
        self.repository.unit_directory.save_settings(
            AppUnitSettings(home_branch_code="6501", default_office_code="6501-00")
        )
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = CreditLimitTab(self.repository)
        try:
            tab.reload()
            self.assertIs(tab.repository.unit_directory, self.repository.unit_directory)
            self.assertEqual(tab.repository.unit_directory.get_home_branch().branch_code, "6501")
        finally:
            tab.deleteLater()

    def test_nim_kpi_cards_use_shared_compact_component(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tabs = (
            NimTab(self.repository, SummaryDataType.NIM_DN),
            NimTab(self.repository, SummaryDataType.NIM_NV),
        )
        try:
            for tab in tabs:
                tab.metrics.set_data(
                    DashboardData(
                        metrics=(
                            DashboardMetric(tab.ui_config.total_balance_label, "28.056.021.646.654"),
                            DashboardMetric("NIM trước ĐC", "2,31%"),
                            DashboardMetric("NIM sau ĐC", "2,11%"),
                        )
                    )
                )
                self.assertIsInstance(tab.metrics, MetricGrid)
                cards = tab.metrics.findChildren(CompactKpiCard)
                self.assertEqual(len(cards), 3)
                self.assertTrue(all(card.objectName().startswith("CompactKpiCard") for card in cards))
                self.assertTrue(all(card.maximumHeight() <= 80 for card in cards))
        finally:
            for tab in tabs:
                tab.deleteLater()

    def test_nim_kpi_money_is_compact_with_full_tooltip(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            tab.metrics.set_data(
                DashboardData(
                    metrics=(DashboardMetric("Tổng dư nợ", "28.056.021.646.654"),)
                )
            )
            card = tab.metrics.findChildren(CompactKpiCard)[0]
            self.assertIn("tỷ", card.value_label.text())
            self.assertNotEqual(card.value_label.text(), "28.056.021.646.654")
            self.assertIn("28.056.021.646.654 đồng", card.toolTip())
        finally:
            tab.deleteLater()

    def test_nim_kpi_percentage_display_preserves_scale(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            tab.metrics.set_data(
                DashboardData(
                    metrics=(
                        DashboardMetric("NIM trước ĐC", "2,31%"),
                        DashboardMetric("Lãi suất bình quân", "7,15%"),
                    )
                )
            )
            values = [card.value_label.text() for card in tab.metrics.findChildren(CompactKpiCard)]
            self.assertEqual(values, ["2,31%", "7,15%"])
            self.assertNotIn("231,00%", values)
            self.assertNotIn("715,00%", values)
        finally:
            tab.deleteLater()

    def test_nim_kpi_empty_value_shows_no_data(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_NV)
        try:
            tab.metrics.set_data(DashboardData(metrics=(DashboardMetric("Tổng nguồn vốn", ""),)))
            card = tab.metrics.findChildren(CompactKpiCard)[0]
            self.assertEqual(card.value_label.text(), "—")
            self.assertIn("Không có dữ liệu", card.toolTip())
        finally:
            tab.deleteLater()

    def test_officer_history_info_uses_shared_kpi_grid(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._write_nim_dn_file(
            "5491_FTPLN_20260131.csv",
            ["5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN"],
        )
        import_nim_dn(self.repository, self.root)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, "[540000321] Nguyễn Văn A")
        try:
            self.assertIsInstance(dialog.info_metrics, MetricGrid)
            titles = [card.title_label.text() for card in dialog.info_metrics.findChildren(CompactKpiCard)]
            self.assertIn("CBTD", titles)
            self.assertIn("Tổng dư nợ kỳ hiện tại", titles)
            self.assertIn("NIM sau ĐC hiện tại", titles)
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_nim_combos_use_shared_popup_and_preserve_data(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            combo = tab.officer_filter
            _populate_combo(combo, [("Nguyễn Văn A", "[540000321] Nguyễn Văn A")])
            self.assertEqual(combo.objectName(), "AgribankComboBox")
            self.assertEqual(combo.view().objectName(), "AgribankComboPopup")
            self.assertEqual(combo.view().textElideMode(), Qt.TextElideMode.ElideNone)
            self.assertEqual(combo.itemData(1), "[540000321] Nguyễn Văn A")
            changes = {"count": 0}
            combo.currentIndexChanged.connect(lambda _index: changes.__setitem__("count", changes["count"] + 1))
            combo.setCurrentIndex(1)
            self.assertEqual(changes["count"], 1)
        finally:
            tab.deleteLater()

    def test_officer_history_combos_use_shared_popup(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._write_nim_dn_file(
            "5491_FTPLN_20260131.csv",
            ["5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN"],
        )
        import_nim_dn(self.repository, self.root)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, "[540000321] Nguyễn Văn A")
        try:
            combos = dialog.findChildren(QComboBox)
            self.assertTrue(combos)
            self.assertTrue(all(combo.objectName() == "AgribankComboBox" for combo in combos))
            self.assertTrue(all(combo.view().objectName() == "AgribankComboPopup" for combo in combos))
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_shared_buttons_have_font_metric_safe_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        buttons = [
            primary_button("Áp dụng cập nhật"),
            secondary_button("Nhập dữ liệu"),
            danger_button("Xóa dữ liệu"),
        ]
        try:
            for button in buttons:
                self.assertGreaterEqual(button.minimumHeight(), recommended_control_height(button))
                self.assertEqual(button.maximumHeight(), 16777215)
        finally:
            for button in buttons:
                button.deleteLater()

    def test_shared_button_height_handles_scaled_font(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = primary_button("Đồng ý")
        try:
            font = button.font()
            font.setPointSize(max(font.pointSize() + 5, 17))
            button.setFont(font)
            button.setMinimumHeight(recommended_control_height(button))
            self.assertGreaterEqual(button.minimumHeight(), recommended_control_height(button))
        finally:
            button.deleteLater()

    def test_button_size_hint_stable_when_pressed_and_focused(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = secondary_button("Áp dụng")
        try:
            before = button.sizeHint()
            button.setFocus(Qt.FocusReason.OtherFocusReason)
            button.setDown(True)
            app.processEvents()
            self.assertEqual(button.sizeHint(), before)
            button.setDown(False)
        finally:
            button.deleteLater()

    def test_nim_debt_uses_shared_compact_kpi_card(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            cards = self._set_sample_nim_kpis(tab)
            self.assertIsInstance(tab.metrics, MetricGrid)
            self.assertTrue(cards)
            self.assertTrue(all(isinstance(card, CompactKpiCard) for card in cards))
        finally:
            tab.deleteLater()

    def test_nim_funding_uses_shared_compact_kpi_card(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_NV)
        try:
            cards = self._set_sample_nim_kpis(tab)
            self.assertIsInstance(tab.metrics, MetricGrid)
            self.assertTrue(cards)
            self.assertTrue(all(isinstance(card, CompactKpiCard) for card in cards))
        finally:
            tab.deleteLater()

    def test_nim_kpi_money_is_compact(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            cards = self._set_sample_nim_kpis(tab)
            balance_card = next(card for card in cards if card.title_label.text() == "Tổng dư nợ")
            self.assertEqual(balance_card.value_label.text(), "28.056,02 tỷ")
        finally:
            tab.deleteLater()

    def test_nim_kpi_money_tooltip_has_full_value(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            cards = self._set_sample_nim_kpis(tab)
            balance_card = next(card for card in cards if card.title_label.text() == "Tổng dư nợ")
            self.assertIn("Tổng dư nợ", balance_card.toolTip())
            self.assertIn("28.056.021.646.654 đồng", balance_card.toolTip())
        finally:
            tab.deleteLater()

    def test_nim_kpi_percentage_format(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            cards = self._set_sample_nim_kpis(tab)
            values = {card.title_label.text(): card.value_label.text() for card in cards}
            self.assertEqual(values["NIM trước ĐC"], "2,31%")
            self.assertEqual(values["NIM sau ĐC"], "2,40%")
            self.assertEqual(values["Lãi suất bình quân"], "8,01%")
        finally:
            tab.deleteLater()

    def test_nim_kpi_no_data_state(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            tab.metrics.set_data(DashboardData(metrics=(DashboardMetric("Tổng dư nợ", ""),)))
            card = tab.metrics.findChildren(CompactKpiCard)[0]
            self.assertEqual(card.value_label.text(), "—")
            self.assertNotIn("None", card.toolTip())
            self.assertNotIn("nan", card.toolTip().casefold())
        finally:
            tab.deleteLater()

    def test_nim_kpi_grid_responsive(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            tab.metrics.set_metrics([("KPI " + str(index), index, "count") for index in range(8)])
            tab.metrics.resize(1600, 120)
            tab.metrics.resizeEvent(None)
            self.assertEqual(tab.metrics.main_column_count(), 8)
            tab.metrics.resize(700, 220)
            tab.metrics.resizeEvent(None)
            self.assertEqual(tab.metrics.main_column_count(), 3)
        finally:
            tab.deleteLater()

    def test_nim_kpi_values_unchanged_after_style_refactor(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._write_nim_dn_file(
            "5491_FTPLN_20260131.csv",
            ["5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN"],
        )
        import_nim_dn(self.repository, self.root)
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            dashboard = tab.query_dashboard()
            before = tuple((metric.label, metric.value) for metric in dashboard.metrics)
            tab.metrics.set_data(dashboard)
            after = tuple((metric.label, metric.value) for metric in dashboard.metrics)
            self.assertEqual(after, before)
            self.assertIn(("Tổng dư nợ", "1.000"), before)
            self.assertIn(("NIM sau ĐC", "7,00%"), before)
        finally:
            tab.deleteLater()

    def test_nim_combos_use_shared_agribank_style(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        dashboard = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN, parent=tab)
        try:
            combos = tab.findChildren(QComboBox) + dashboard.findChildren(QComboBox)
            self.assertTrue(combos)
            self.assertTrue(all(combo.objectName() == "AgribankComboBox" for combo in combos))
            self.assertTrue(all(combo.view().objectName() == "AgribankComboPopup" for combo in combos))
        finally:
            dashboard.close()
            dashboard.deleteLater()
            tab.deleteLater()

    def test_nim_combo_closed_height_compact(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._nim_combo_with_long_officer()
        self.assertGreaterEqual(combo.minimumHeight(), 30)
        self.assertLessEqual(combo.maximumHeight(), 34)

    def test_nim_combo_popup_item_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._nim_combo_with_long_officer()
        option = QStyleOptionViewItem()
        height = combo.view().itemDelegate().sizeHint(option, combo.model().index(0, 0)).height()
        self.assertGreaterEqual(height, 24)
        self.assertLessEqual(height, 32)

    def test_nim_combo_selected_item_contrast(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._nim_combo_with_long_officer()
        style = combo.view().styleSheet()
        self.assertIn("selection-background-color", style)
        self.assertIn("selection-color: #202020", style)
        self.assertIn("174, 28, 63", style)

    def test_nim_combo_popup_full_officer_name(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._nim_combo_with_long_officer()
        long_text = combo.itemText(1)
        width = configure_combo_popup_width(combo, minimum_popup_width=260)
        self.assertGreaterEqual(width, combo.fontMetrics().horizontalAdvance(long_text) + 40)

    def test_nim_combo_item_data_unchanged(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._nim_combo_with_long_officer()
        self.assertEqual(combo.itemData(1), "540000321")

    def test_nim_combo_signals_unchanged(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = self._nim_combo_with_long_officer()
        changes = {"count": 0}
        combo.currentIndexChanged.connect(lambda _index: changes.__setitem__("count", changes["count"] + 1))
        combo.setCurrentIndex(1)
        self.assertEqual(changes["count"], 1)

    def test_button_height_fits_font_metrics(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = primary_button("Áp dụng cập nhật Đồng ý")
        try:
            self._assert_button_text_fits(button, button.text())
        finally:
            button.deleteLater()

    def test_button_apply_text_not_clipped(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = primary_button("Áp dụng")
        try:
            self._assert_button_text_fits(button, "Áp dụng")
        finally:
            button.deleteLater()

    def test_button_descender_g_not_clipped(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = secondary_button("Tăng trưởng")
        try:
            self._assert_button_text_fits(button, "Tăng trưởng g")
        finally:
            button.deleteLater()

    def test_button_descender_p_not_clipped(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = secondary_button("Nhập dữ liệu")
        try:
            self._assert_button_text_fits(button, "Nhập p q y")
        finally:
            button.deleteLater()

    def test_button_vietnamese_dot_below_not_clipped(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = danger_button("Khôi phục")
        try:
            self._assert_button_text_fits(button, "dụng động nhập ạ ặ ậ ệ ị ọ ộ ợ ự")
        finally:
            button.deleteLater()

    def test_button_primary_secondary_same_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        primary = primary_button("Áp dụng")
        secondary = secondary_button("Xóa lọc")
        try:
            self.assertEqual(primary.minimumHeight(), secondary.minimumHeight())
        finally:
            primary.deleteLater()
            secondary.deleteLater()

    def test_button_focus_does_not_change_geometry(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = primary_button("Đồng ý")
        try:
            before = button.sizeHint()
            button.setFocus(Qt.FocusReason.OtherFocusReason)
            app.processEvents()
            self.assertEqual(button.sizeHint(), before)
        finally:
            button.deleteLater()

    def test_button_pressed_text_does_not_shift(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = secondary_button("Cập nhật")
        try:
            before = button.sizeHint()
            button.setDown(True)
            app.processEvents()
            self.assertEqual(button.sizeHint(), before)
            button.setDown(False)
        finally:
            button.deleteLater()

    def test_button_no_conflicting_fixed_max_height(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = primary_button("Áp dụng")
        try:
            self.assertEqual(button.maximumHeight(), 16777215)
            stylesheet = Path("src/agribank_v3/ui/styles.py").read_text(encoding="utf-8")
            self.assertNotRegex(stylesheet, r"(?s)QPushButton[^{]*\{[^}]*max-height")
        finally:
            button.deleteLater()

    def test_toolbar_button_text_visible_at_125_percent(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = secondary_button("Phòng giao dịch")
        try:
            font = button.font()
            font.setPointSize(max(font.pointSize() + 2, 13))
            button.setFont(font)
            button.setMinimumHeight(recommended_control_height(button))
            self._assert_button_text_fits(button, "Phòng giao dịch")
        finally:
            button.deleteLater()

    def test_toolbar_button_text_visible_at_150_percent(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        button = secondary_button("Đề nghị thanh toán")
        try:
            font = button.font()
            font.setPointSize(max(font.pointSize() + 5, 16))
            button.setFont(font)
            button.setMinimumHeight(recommended_control_height(button))
            self._assert_button_text_fits(button, "Đề nghị thanh toán")
        finally:
            button.deleteLater()

    def test_maintenance_delete_batch_button_not_expanding(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dialog = SummaryMaintenanceDialog(self.repository)
        try:
            self.assertNotEqual(dialog.delete_batch_button.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
            self.assertLessEqual(dialog.delete_batch_button.sizeHint().width(), 240)
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_maintenance_batch_row_compact(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dialog = SummaryMaintenanceDialog(self.repository)
        try:
            self.assertEqual(dialog.batch_row.spacing(), 8)
            self.assertGreaterEqual(dialog.batch_combo.minimumWidth(), 360)
            self.assertEqual(dialog.batch_combo.sizePolicy().horizontalPolicy(), QSizePolicy.Policy.Expanding)
            self.assertEqual(dialog.batch_row.indexOf(dialog.delete_batch_button), 1)
            self.assertIsNotNone(dialog.batch_row.itemAt(2).spacerItem())
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_maintenance_actions_use_two_rows(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dialog = SummaryMaintenanceDialog(self.repository)
        try:
            self.assertEqual(dialog.data_actions_row.count(), 4)
            self.assertEqual(dialog.system_actions_row.count(), 3)
            self.assertEqual(dialog.system_actions_flow.count(), 3)
            self.assertIs(dialog.data_actions_row.itemAt(0).widget(), dialog.refresh_button)
            self.assertIs(dialog.system_actions_flow.itemAt(0).widget(), dialog.optimize_button)
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_maintenance_close_button_aligned_right(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dialog = SummaryMaintenanceDialog(self.repository)
        try:
            self.assertEqual(dialog.system_actions_row.indexOf(dialog.close_button), dialog.system_actions_row.count() - 1)
            self.assertIsNotNone(dialog.system_actions_row.itemAt(dialog.system_actions_row.count() - 2).spacerItem())
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_maintenance_buttons_do_not_clip_text(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dialog = SummaryMaintenanceDialog(self.repository)
        try:
            buttons = dialog.findChildren(QPushButton)
            self.assertTrue(buttons)
            for button in buttons:
                self.assertGreaterEqual(button.minimumHeight(), recommended_control_height(button))
                self.assertEqual(button.maximumHeight(), 16777215)
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_maintenance_layout_responsive(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dialog = SummaryMaintenanceDialog(self.repository)
        try:
            dialog.resize(760, 320)
            app.processEvents()
            self.assertLessEqual(dialog.minimumWidth(), 760)
            self.assertTrue(dialog.data_actions_row.hasHeightForWidth())
            self.assertGreater(
                dialog.data_actions_row.heightForWidth(420),
                dialog.data_actions_row.heightForWidth(900),
            )
            self.assertTrue(dialog.system_actions_flow.hasHeightForWidth())
            self.assertLessEqual(dialog.system_actions_row.sizeHint().width(), dialog.width())
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_maintenance_no_horizontal_overflow(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dialog = SummaryMaintenanceDialog(self.repository)
        try:
            dialog.resize(max(760, dialog.minimumWidth()), 320)
            app.processEvents()
            available = dialog.width() - dialog.layout().contentsMargins().left() - dialog.layout().contentsMargins().right()
            self.assertLessEqual(dialog.batch_row.sizeHint().width(), available)
            for index in range(dialog.data_actions_row.count()):
                self.assertLessEqual(dialog.data_actions_row.itemAt(index).sizeHint().width(), available)
            self.assertLessEqual(dialog.system_actions_row.sizeHint().width(), available)
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_menu_label_is_customer_increase_decrease_comparison(self) -> None:
        self.assertEqual(LOAN_COMPARE_TITLE, "So sánh tăng giảm khách hàng")
        titles = [feature.title for feature in SUMMARY_FEATURES]
        self.assertIn("So sánh tăng giảm khách hàng", titles)
        self.assertNotIn("So sánh dư nợ", titles)

    def test_loan_compare_headers_are_vietnamese(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._seed_loan_compare_rows()
        tab = LoanCompareTab(self.repository)
        try:
            tab.reload()
            headers = [tab.table.horizontalHeaderItem(index).text() for index in range(tab.table.columnCount())]
            self.assertEqual(
                headers,
                [
                    "Mã KH",
                    "Tên khách hàng",
                    "Dư nợ kỳ trước",
                    "Dư nợ kỳ này",
                    "Tăng/giảm",
                    "Loại KH",
                    "Cán bộ QL",
                    "Địa chỉ",
                ],
            )
        finally:
            tab.deleteLater()

    def test_loan_compare_money_display_uses_dot_separator(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._seed_loan_compare_rows()
        tab = LoanCompareTab(self.repository)
        try:
            tab.reload()
            values = [tab.table.item(0, column).text() for column in (2, 3, 4)]
            self.assertEqual(values, ["13.500.000.000", "0", "-13.500.000.000"])
            for column in (2, 3, 4):
                self.assertEqual(tab.table.item(0, column).textAlignment(), Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            ui_rows = tab._export_rows()
            output = export_rows(ui_rows, self.root / "loan-ui.xlsx", title=tab.title, sheet_name=tab.title)
            workbook = load_workbook(output, data_only=True)
            try:
                worksheet = workbook[tab.title[:31]]
                self.assertEqual(worksheet["A3"].value, "000123")
                self.assertEqual(worksheet["A3"].number_format, "@")
                self.assertEqual(worksheet["C3"].value, 13500000000)
                self.assertEqual(worksheet["C3"].number_format, "#,##0")
            finally:
                workbook.close()
        finally:
            tab.deleteLater()

    def test_delete_nim_dn_period(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260131.csv", ["5491,2,10,1,CB1,00,1000,,CN"])
        self._write_nim_dn_file("5491_FTPLN_20260228.csv", ["5491,2,10,1,CB1,00,2000,,CN"])
        import_nim_dn(self.repository, self.root)

        info = self.repository.delete_nim_period(SummaryDataType.NIM_DN, "2026-01")

        self.assertEqual(info["row_count"], 1)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-01"), 0)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-02"), 1)

    def test_delete_nim_nv_period(self) -> None:
        self._write_nim_nv_file("5491_FTPDP_20260131.csv", ["5491,10,2,1,CB1,00,1000,,CN"])
        self._write_nim_nv_file("5491_FTPDP_20260228.csv", ["5491,10,2,1,CB1,00,2000,,CN"])
        import_nim_nv(self.repository, self.root)

        info = self.repository.delete_nim_period(SummaryDataType.NIM_NV, "2026-01")

        self.assertEqual(info["row_count"], 1)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_NV, "2026-01"), 0)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_NV, "2026-02"), 1)

    def test_delete_nim_period_does_not_delete_other_type(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260131.csv", ["5491,2,10,1,CB1,00,1000,,CN"])
        self._write_nim_nv_file("5491_FTPDP_20260131.csv", ["5491,10,2,1,CB2,00,2000,,CN"])
        import_nim_dn(self.repository, self.root)
        import_nim_nv(self.repository, self.root)

        self.repository.delete_nim_period(SummaryDataType.NIM_DN, "2026-01")

        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-01"), 0)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_NV, "2026-01"), 1)

    def test_delete_nim_period_transaction_rollback(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260131.csv", ["5491,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)

        with patch.object(self.repository, "log_action", side_effect=RuntimeError("boom")):
            with self.assertRaises(RuntimeError):
                self.repository.delete_nim_period(SummaryDataType.NIM_DN, "2026-01")

        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-01"), 1)
        self.assertEqual(len(self.repository.list_batches(SummaryDataType.NIM_DN)), 1)

    def test_delete_nim_period_refreshes_available_periods(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._write_nim_dn_file("5491_FTPLN_20260131.csv", ["5491,2,10,1,CB1,00,1000,,CN"])
        self._write_nim_dn_file("5491_FTPLN_20260228.csv", ["5491,2,10,1,CB1,00,2000,,CN"])
        import_nim_dn(self.repository, self.root)
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            self.repository.delete_nim_period(SummaryDataType.NIM_DN, "2026-01")
            tab.reload()
            values = [tab.period_filter.itemData(index) for index in range(tab.period_filter.count())]
            self.assertNotIn("2026-01", values)
            self.assertIn("2026-02", values)
        finally:
            tab.deleteLater()

    def test_credit_summary_database_created(self) -> None:
        self.assertEqual(self.repository.database_path.name, CREDIT_SUMMARY_DATABASE_NAME)
        self.assertTrue(self.repository.database_path.is_file())
        with closing(sqlite3.connect(self.repository.database_path)) as connection:
            table = connection.execute(
                """
                SELECT 1
                FROM sqlite_master
                WHERE type = 'table' AND name = 'credit_summary_schema_migrations'
                """
            ).fetchone()
        self.assertIsNotNone(table)

    def test_summary_repository_uses_separate_database(self) -> None:
        expected = credit_summary_database_path(self.database_path)
        self.assertEqual(self.repository.database_path, expected)
        self.assertNotEqual(self.repository.database_path, self.database_path)

    def test_migrate_existing_summary_data_preserves_rows(self) -> None:
        legacy_path = self.root / "legacy" / "DuLieuV3.db"
        self._create_legacy_summary_database(legacy_path)

        migrated = SummaryRepository(legacy_path)

        self.assertEqual(migrated.database_path, legacy_path.parent / CREDIT_SUMMARY_DATABASE_NAME)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-03", repository=migrated), 1)
        self.assertEqual(len(migrated.list_batches(SummaryDataType.NIM_DN)), 1)

    def test_migrate_summary_data_idempotent(self) -> None:
        legacy_path = self.root / "legacy-idempotent" / "DuLieuV3.db"
        self._create_legacy_summary_database(legacy_path)

        first = SummaryRepository(legacy_path)
        second = SummaryRepository(legacy_path)

        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-03", repository=first), 1)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-03", repository=second), 1)

    def test_duplicate_import_hash_is_detected(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260131.csv", ["5491,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)

        with self.assertRaises(SummaryError):
            import_nim_dn(self.repository, self.root)

    def test_nim_import_stores_aggregated_rows_only(self) -> None:
        rows = [
            f"5491,2,10,1,[540000{index % 4:03d}] CBTD {index % 4},00,100,,{'CN' if index % 2 == 0 else 'TC'}"
            for index in range(1000)
        ]
        self._write_nim_dn_file("5491_FTPLN_20260331.csv", rows)

        import_nim_dn(self.repository, self.root)

        self.assertEqual(self._count_raw_nim_rows(SummaryDataType.NIM_DN, "2026-03"), 0)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-03"), 4)
        with closing(self.repository.connect()) as connection:
            source_rows = int(
                connection.execute(
                    """
                    SELECT SUM(source_row_count)
                    FROM nim_period_summary
                    WHERE data_type = ? AND period = ?
                    """,
                    (SummaryDataType.NIM_DN.value, "2026-03"),
                ).fetchone()[0]
                or 0
            )
        self.assertEqual(source_rows, 1000)

    def test_nim_aggregate_matches_raw_totals(self) -> None:
        self._write_nim_dn_file(
            "5491_FTPLN_20260331.csv",
            [
                "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                "5491,4,8,1,[540000321] Nguyễn Văn A,00,3000,,TC",
            ],
        )

        import_nim_dn(self.repository, self.root)

        with closing(self.repository.connect()) as connection:
            row = connection.execute(
                """
                SELECT
                    SUM(balance) AS balance,
                    SUM(interest_rate_numerator) AS interest_rate_numerator,
                    SUM(numerator_before) AS numerator_before,
                    SUM(numerator_after) AS numerator_after,
                    SUM(source_row_count) AS source_row_count
                FROM nim_period_summary
                WHERE data_type = ? AND period = ?
                """,
                (SummaryDataType.NIM_DN.value, "2026-03"),
            ).fetchone()
        self.assertAlmostEqual(float(row["balance"]), 4000.0)
        self.assertAlmostEqual(float(row["interest_rate_numerator"]), 34000.0)
        self.assertAlmostEqual(float(row["numerator_before"]), 20000.0)
        self.assertAlmostEqual(float(row["numerator_after"]), 16000.0)
        self.assertEqual(int(row["source_row_count"]), 2)

    def test_all_customer_types_calculated_on_query(self) -> None:
        self._write_nim_dn_file(
            "5491_FTPLN_20260331.csv",
            [
                "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                "5491,4,8,1,[540000321] Nguyễn Văn A,00,3000,,TC",
            ],
        )
        import_nim_dn(self.repository, self.root)

        with closing(self.repository.connect()) as connection:
            saved_total_rows = int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM nim_period_summary
                    WHERE customer_type = ?
                    """,
                    ("[Tất cả KH]",),
                ).fetchone()[0]
                or 0
            )
        page = self.repository.query_nim(SummaryDataType.NIM_DN)

        self.assertEqual(saved_total_rows, 0)
        self.assertEqual(page.total_rows, 1)
        self.assertEqual(page.rows[0]["customer_type"], "Tất cả")
        self.assertAlmostEqual(float(page.rows[0]["balance"]), 4000.0)
        self.assertAlmostEqual(float(page.rows[0]["average_rate"]), 8.5)

    def test_weighted_nim_not_average_percent(self) -> None:
        self._write_nim_dn_file(
            "5491_FTPLN_20260331.csv",
            [
                "5491,2,10,0,[540000321] Nguyễn Văn A,00,100,,CN",
                "5491,1,2,0,[540000321] Nguyễn Văn A,00,300,,TC",
            ],
        )
        import_nim_dn(self.repository, self.root)

        row = self.repository.query_nim(SummaryDataType.NIM_DN).rows[0]

        self.assertAlmostEqual(float(row["nim_before"]), 2.75)
        self.assertNotAlmostEqual(float(row["nim_before"]), 4.5)

    def test_existing_raw_migration_preserves_totals(self) -> None:
        legacy_path = self.root / "legacy-raw" / "DuLieuV3.db"
        self._create_legacy_summary_database(legacy_path)
        with closing(sqlite3.connect(legacy_path)) as connection:
            batch_id = int(connection.execute("SELECT id FROM summary_import_history LIMIT 1").fetchone()[0])
            connection.execute(
                """
                INSERT INTO nim_details(
                    batch_id, data_type, period, branch_code, branch_name, trctcd,
                    transaction_office, customer_type, officer, balance, interest_rate,
                    ftp_rate, adjustment_rate, numerator_before, numerator_after,
                    average_rate_numerator, source_file, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    SummaryDataType.NIM_DN.value,
                    "2026-03",
                    "5491",
                    "5491 - CN Lộc Phát",
                    "00",
                    "Hội sở",
                    "Cá nhân (CN)",
                    "CB1",
                    3000,
                    8,
                    4,
                    1,
                    12000,
                    9000,
                    24000,
                    "legacy.csv",
                    "2026-03-31T08:00:00+07:00",
                ),
            )
            connection.commit()

        migrated = SummaryRepository(legacy_path)
        verification = migrated.verify_nim_summary_totals()

        self.assertTrue(verification["matches"])
        self.assertEqual(verification["raw_rows"], 2)
        self.assertEqual(verification["summary_source_rows_for_raw"], 2)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-03", repository=migrated), 1)

    def test_same_officer_has_one_row_per_customer_type_period(self) -> None:
        self._write_nim_dn_file(
            "5491_FTPLN_20260331.csv",
            [
                "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                "5491,2,11,1,[540000321] Nguyễn Văn A,00,2000,,CN",
                "5491,4,8,1,[540000321] Nguyễn Văn A,00,3000,,TC",
                "5491,4,9,1,[540000321] Nguyễn Văn A,00,4000,,TC",
            ],
        )
        import_nim_dn(self.repository, self.root)

        with closing(self.repository.connect()) as connection:
            rows = connection.execute(
                """
                SELECT customer_type, COUNT(*) AS count_rows, SUM(source_row_count) AS source_rows
                FROM nim_period_summary
                WHERE data_type = ? AND period = ? AND officer_code = ?
                GROUP BY customer_type
                ORDER BY customer_type
                """,
                (SummaryDataType.NIM_DN.value, "2026-03", "540000321"),
            ).fetchall()

        self.assertEqual([(row["customer_type"], row["count_rows"]) for row in rows], [("Cá nhân (CN)", 1), ("Tổ chức (TC)", 1)])
        self.assertEqual(sum(int(row["source_rows"]) for row in rows), 4)

    def test_dashboard_uses_summary_table(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260331.csv", ["5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)

        data = build_nim_dashboard(
            NimDashboardRepository(self.repository),
            SummaryDataType.NIM_DN,
            DashboardFilters(customer_type="Cá nhân (CN)"),
        )

        self.assertEqual(self._count_raw_nim_rows(SummaryDataType.NIM_DN, "2026-03"), 0)
        self.assertAlmostEqual(data.period_rows[0].balance, 1000.0)

    def test_officer_history_uses_summary_table(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260331.csv", ["5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)

        history = build_officer_history(
            self.repository,
            SummaryDataType.NIM_DN,
            officer="[540000321] Nguyễn Văn A",
        )

        self.assertEqual(self._count_raw_nim_rows(SummaryDataType.NIM_DN, "2026-03"), 0)
        self.assertEqual(len(history.points), 1)
        self.assertAlmostEqual(history.current_balance, 1000.0)

    def test_export_vba_result_unchanged(self) -> None:
        self._write_nim_dn_file(
            "5491_FTPLN_20260331.csv",
            [
                "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                "5491,4,8,1,[540000321] Nguyễn Văn A,00,3000,,TC",
            ],
        )
        output = self.root / "BaoCaoNIM_CSDL.xlsx"

        import_nim_dn(self.repository, self.root, export_path=output)

        workbook = load_workbook(output, data_only=True)
        try:
            cache = workbook["Cache_Nim"]
            customer_types = [cache.cell(row_index, 4).value for row_index in range(2, 5)]
            self.assertEqual(customer_types, ["Cá nhân (CN)", "[Tất cả KH]", "Tổ chức (TC)"])
            self.assertAlmostEqual(cache.cell(3, 6).value, 8.5)
            self.assertAlmostEqual(cache.cell(3, 7).value, 5.0)
            self.assertAlmostEqual(cache.cell(3, 8).value, 4.0)
        finally:
            workbook.close()

    def test_delete_period_summary_rows(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260331.csv", ["5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)

        info = self.repository.delete_nim_period(SummaryDataType.NIM_DN, "2026-03")

        self.assertEqual(info["row_count"], 1)
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-03"), 0)
        self.assertEqual(self._count_raw_nim_rows(SummaryDataType.NIM_DN, "2026-03"), 0)

    def test_duplicate_file_hash_not_imported_twice(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260331.csv", ["5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)

        with self.assertRaises(SummaryError):
            import_nim_dn(self.repository, self.root)

        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-03"), 1)
        self.assertEqual(self._count_raw_nim_rows(SummaryDataType.NIM_DN, "2026-03"), 0)

    def test_same_file_hash_different_period_is_allowed(self) -> None:
        self._write_nim_dn_file("5414_FTPLN_20260331.csv", [])
        self._write_nim_dn_file("5414_FTPLN_20260420.csv", [])

        result = import_nim_dn(self.repository, self.root)

        self.assertEqual(result.row_count, 0)
        self.assertEqual(len(self.repository.list_batches(SummaryDataType.NIM_DN)), 2)

    def test_database_optimize_does_not_change_data(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260131.csv", ["5491,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        before = self._count_nim_rows(SummaryDataType.NIM_DN, "2026-01")

        result = self.repository.optimize_database(vacuum=True)

        self.assertTrue(result["vacuum"])
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-01"), before)

    def test_summary_backup_restore(self) -> None:
        self._write_nim_dn_file("5491_FTPLN_20260131.csv", ["5491,2,10,1,CB1,00,1000,,CN"])
        import_nim_dn(self.repository, self.root)
        backup = self.repository.backup_database(self.root / "summary-backup.zip")
        with zipfile.ZipFile(backup) as archive:
            self.assertIn(CREDIT_SUMMARY_DATABASE_NAME, archive.namelist())
        self.repository.delete_nim_period(SummaryDataType.NIM_DN, "2026-01")
        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-01"), 0)

        self.repository.restore_database(backup)

        self.assertEqual(self._count_nim_rows(SummaryDataType.NIM_DN, "2026-01"), 1)

    def test_loan_compare_customer_code_preserves_leading_zero(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._seed_loan_compare_rows()
        tab = LoanCompareTab(self.repository)
        try:
            tab.reload()
            self.assertEqual(tab.table.item(0, 0).text(), "000123")
        finally:
            tab.deleteLater()

    def test_loan_compare_category_display_is_vietnamese(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._seed_loan_compare_rows()
        tab = LoanCompareTab(self.repository)
        try:
            tab.reload()
            categories = {
                tab.table.item(row, 5).text()
                for row in range(tab.table.rowCount())
            }
            self.assertIn("Khách hàng tất toán", categories)
            self.assertIn("Khách hàng vay tăng", categories)
            labels = [tab.category_filter.itemText(index) for index in range(tab.category_filter.count())]
            self.assertEqual(labels[0], "Tất cả loại")
            self.assertIn("Khách hàng vay tăng", labels)
        finally:
            tab.deleteLater()

    def test_nim_tab_groups_all_customer_types_by_officer_only(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        dn = self.root / "5491_FTPLN_20260331.csv"
        dn.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                    "5491,4,8,1,[540000321] Nguyễn Văn A,00,3000,,TC",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        try:
            all_types = tab.query_page()
            self.assertEqual(all_types.total_rows, 1)
            row = all_types.rows[0]
            self.assertEqual(row["officer"], "[540000321] Nguyễn Văn A")
            self.assertEqual(row["customer_type"], "Tất cả")
            self.assertAlmostEqual(float(row["balance"]), 4000.0)
            self.assertAlmostEqual(float(row["average_rate"]), 8.5)
            self.assertAlmostEqual(float(row["nim_before"]), 5.0)
            self.assertAlmostEqual(float(row["nim_after"]), 4.0)
            self.assertFalse(hasattr(tab, "rate_input"))
            self.assertNotIn("Số dòng", [metric.label for metric in tab.query_dashboard().metrics])
            tab._render_table(all_types.rows)
            tab._restore_column_widths()
            total_width = sum(tab.table.columnWidth(index) for index in range(tab.table.columnCount()))
            self.assertLessEqual(total_width, tab.table.viewport().width() + 2)
            for column in (5, 6, 7, 8):
                self.assertEqual(tab.table.item(0, column).textAlignment(), Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

            tab.customer_type_filter.setCurrentIndex(tab.customer_type_filter.findData("Cá nhân (CN)"))
            personal = tab.query_page()
            self.assertEqual(personal.total_rows, 1)
            self.assertEqual(personal.rows[0]["customer_type"], "Cá nhân (CN)")
            self.assertAlmostEqual(float(personal.rows[0]["balance"]), 1000.0)
            self.assertAlmostEqual(float(personal.rows[0]["average_rate"]), 10.0)
        finally:
            tab.deleteLater()

    def test_nim_dashboard_weighted_branch_metrics(self) -> None:
        first = self.root / "5491_FTPLN_20260131.csv"
        first.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,100,,CN",
                    "5491,1,2,0.5,[540000322] Nguyễn Văn B,00,300,,CN",
                    "5400,1,5,0,[540000323] Nguyễn Văn C,00,200,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        data = build_nim_dashboard(
            NimDashboardRepository(self.repository),
            SummaryDataType.NIM_DN,
            DashboardFilters(customer_type="Cá nhân (CN)", metric=METRIC_NIM_BEFORE),
        )

        branch_rows = {row.branch: row for row in data.branch_rows}
        self.assertAlmostEqual(branch_rows["5491 - CN Lộc Phát"].nim_before, 2.75)
        self.assertNotAlmostEqual(branch_rows["5491 - CN Lộc Phát"].nim_before, 4.5)
        self.assertEqual(data.kpis[0].label, "Số chi nhánh")
        self.assertEqual(data.kpis[0].value, "2")

    def test_nim_dashboard_all_customer_types_grouped_for_detail(self) -> None:
        path = self.root / "5491_FTPLN_20260131.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                    "5491,4,8,1,[540000322] Nguyễn Văn B,00,3000,,TC",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        data = build_nim_dashboard(
            NimDashboardRepository(self.repository),
            SummaryDataType.NIM_DN,
            DashboardFilters(),
        )

        self.assertEqual(len(data.detail_rows), 1)
        self.assertEqual(data.detail_rows[0].customer_type, "Tất cả")
        self.assertAlmostEqual(data.detail_rows[0].balance, 4000.0)
        self.assertAlmostEqual(data.detail_rows[0].average_rate, 8.5)

    def test_nim_dashboard_window_is_modeless(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        try:
            self.assertFalse(window.isModal())
            self.assertEqual(window.windowModality(), Qt.WindowModality.NonModal)
            self.assertEqual(window.tabs.count(), 4)
        finally:
            window.close()
            window.deleteLater()

    def test_export_overview_tab(self) -> None:
        data = self._seed_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))
        output = DashboardNimExportService(data).export_overview_by_period(self.root / "overview.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            worksheet = workbook[SHEET_OVERVIEW]
            self.assertEqual([worksheet.cell(1, col).value for col in range(1, 8)], [
                "Kỳ",
                "Tổng dư nợ",
                "Lãi suất bình quân",
                "NIM trước ĐC",
                "NIM sau ĐC",
                "Tăng/giảm dư nợ tuyệt đối",
                "Tăng trưởng dư nợ (%)",
            ])
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(worksheet.auto_filter.ref, worksheet.dimensions)
        finally:
            workbook.close()

    def test_export_branch_comparison_tab(self) -> None:
        data = self._seed_dashboard_data(DashboardFilters(period_from="2026-02", period_to="2026-02", customer_type="Cá nhân (CN)"))
        output = DashboardNimExportService(data, metric=METRIC_NIM_AFTER).export_branch_comparison(self.root / "branch.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            worksheet = workbook[SHEET_BRANCH]
            headers = [worksheet.cell(1, col).value for col in range(1, 9)]
            self.assertEqual(headers[-2:], ["Chỉ tiêu đang chọn", "Giá trị chỉ tiêu"])
            branches = [worksheet.cell(row, 2).value for row in range(2, worksheet.max_row + 1)]
            self.assertEqual(set(branches), {"5400 - CN Lâm Đồng", "5491 - CN Lộc Phát"})
            self.assertEqual(worksheet.cell(2, 8).number_format, '0.00"%"')
        finally:
            workbook.close()

    def test_export_growth_tab(self) -> None:
        data = self._seed_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))
        output = DashboardNimExportService(data).export_growth(self.root / "growth.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            worksheet = workbook[SHEET_GROWTH]
            self.assertEqual([worksheet.cell(1, col).value for col in range(1, 10)], [
                "Kỳ",
                "Tên chi nhánh",
                "Phòng GD",
                "Loại KH",
                "Dư nợ",
                "Tăng/giảm dư nợ tuyệt đối",
                "Tăng trưởng dư nợ (%)",
                "Biến động NIM trước ĐC",
                "Biến động NIM sau ĐC",
            ])
            self.assertIn("5491 - CN Lộc Phát", [worksheet.cell(row, 2).value for row in range(2, worksheet.max_row + 1)])
        finally:
            workbook.close()

    def test_export_detail_tab(self) -> None:
        data = self._seed_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))
        output = DashboardNimExportService(data).export_detail(self.root / "detail.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            worksheet = workbook[SHEET_DETAIL]
            self.assertEqual(worksheet.max_row - 1, len(data.detail_rows))
            self.assertEqual(worksheet.cell(1, 10).value, "Tăng/giảm dư nợ tuyệt đối")
        finally:
            workbook.close()

    def test_export_all_tabs(self) -> None:
        data = self._seed_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))
        output = DashboardNimExportService(data, metric=METRIC_NIM_AFTER).export_all_tabs(self.root / "all.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, [SHEET_OVERVIEW, SHEET_BRANCH, SHEET_GROWTH, SHEET_DETAIL])
        finally:
            workbook.close()

    def test_export_uses_current_filters(self) -> None:
        data = self._seed_dashboard_data(
            DashboardFilters(
                period_from="2026-02",
                period_to="2026-02",
                branch="5491 - CN Lộc Phát",
                customer_type="Cá nhân (CN)",
            )
        )
        output = DashboardNimExportService(data).export_all_tabs(self.root / "filtered.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            detail = workbook[SHEET_DETAIL]
            self.assertEqual(detail.max_row, 2)
            self.assertEqual(detail["A2"].value, "2026-02")
            self.assertEqual(detail["B2"].value, "5491 - CN Lộc Phát")
            self.assertEqual(detail["D2"].value, "Cá nhân (CN)")
        finally:
            workbook.close()

    def test_branch_chart_uses_full_branch_names(self) -> None:
        data = self._seed_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))
        bars = branch_bar_values(data.branch_rows, METRIC_BALANCE)

        names = [name for _period, name, _value in bars]
        self.assertIn("5400 - CN Lâm Đồng", names)
        self.assertIn("5491 - CN Lộc Phát", names)
        self.assertFalse(any("..." in name for name in names))

    def test_multiline_header_size(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        try:
            window.detail_table.setColumnWidth(8, 70)
            header = window.detail_table.horizontalHeader()
            header.refresh_height()
            self.assertGreater(header.height(), 38)
        finally:
            window.close()
            window.deleteLater()

    def test_percentage_export_format(self) -> None:
        data = self._seed_dashboard_data(DashboardFilters(period_from="2026-02", period_to="2026-02", branch="5491 - CN Lộc Phát", customer_type="Cá nhân (CN)"))
        output = DashboardNimExportService(data).export_overview_by_period(self.root / "percent.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            worksheet = workbook[SHEET_OVERVIEW]
            self.assertAlmostEqual(worksheet["C2"].value, 8.0)
            self.assertEqual(worksheet["C2"].number_format, '0.00"%"')
            self.assertNotAlmostEqual(worksheet["C2"].value, 0.08)
        finally:
            workbook.close()

    def _seed_dashboard_mode_data(self, filters: DashboardFilters | None = None):
        self.repository.unit_directory.save_office(
            OfficeDirectoryEntry(
                id=None,
                branch_code="5491",
                trctcd="01",
                office_code="5491-01",
                office_name="Phòng giao dịch Đức Trọng",
                short_name="PGD Đức Trọng",
                office_type=TRANSACTION_OFFICE,
            )
        )
        rows_by_file = {
            "5491_FTPLN_20260131.csv": [
                "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                "5491,1,8,0,[540000322] Nguyễn Văn B,01,3000,,CN",
                "5400,1,6,0,[540000323] Nguyễn Văn C,00,2000,,CN",
            ],
            "5491_FTPLN_20260228.csv": [
                "5491,2,9,1,[540000321] Nguyễn Văn A,00,1500,,CN",
                "5491,1,7,0,[540000322] Nguyễn Văn B,01,3500,,CN",
                "5400,1,6,0,[540000323] Nguyễn Văn C,00,2500,,CN",
            ],
            "5491_FTPLN_20260331.csv": [
                "5491,2,8,1,[540000321] Nguyễn Văn A,00,2000,,CN",
                "5491,1,6,0,[540000322] Nguyễn Văn B,01,4000,,CN",
                "5400,1,6,0,[540000323] Nguyễn Văn C,00,3000,,CN",
            ],
        }
        for filename, rows in rows_by_file.items():
            self._write_nim_dn_file(filename, rows)
        import_nim_dn(self.repository, self.root)
        return build_nim_dashboard(
            NimDashboardRepository(self.repository),
            SummaryDataType.NIM_DN,
            filters or DashboardFilters(period_from="2026-01", period_to="2026-03", customer_type="Cá nhân (CN)"),
        )

    def _dashboard_mode_window(self, filters: DashboardFilters | None = None) -> NimDashboardWindow:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._seed_dashboard_mode_data(filters)
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(window.close)
        self.addCleanup(window.deleteLater)
        for combo, value in (
            (window.period_from_combo, "2026-01"),
            (window.period_to_combo, "2026-03"),
            (window.customer_type_combo, "Cá nhân (CN)"),
        ):
            index = combo.findData(value)
            if index >= 0:
                combo.setCurrentIndex(index)
        window.reload()
        return window

    def test_nim_debt_chart_legend_does_not_overlap_x_axis(self) -> None:
        chart = NimTrendChart()
        self.addCleanup(chart.deleteLater)
        chart.resize(640, 260)
        chart.set_trend((("2026-01", 2.0, 3.0, 8.0), ("2026-02", 2.5, 3.5, 8.5)))
        self.assertFalse(chart.legend_overlaps_x_axis())

    def test_nim_debt_chart_has_metric_radio_group(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)
        radios = [radio.text() for radio in tab.findChildren(QRadioButton)]
        self.assertIn("NIM", radios)
        self.assertIn("Lãi suất bình quân", radios)

    def test_nim_debt_chart_default_mode_is_nim(self) -> None:
        tab = NimTab(self.repository, SummaryDataType.NIM_DN)
        self.addCleanup(tab.deleteLater)
        self.assertEqual(tab.chart.metric_mode, "nim")
        self.assertTrue(tab.chart_nim_radio.isChecked())

    def test_nim_debt_chart_nim_mode_has_two_series(self) -> None:
        chart = NimTrendChart()
        self.addCleanup(chart.deleteLater)
        chart.set_trend((("2026-01", 2.0, 3.0, 8.0),))
        self.assertEqual(len(chart._series_payload()), 2)

    def test_nim_debt_chart_rate_mode_has_one_series(self) -> None:
        chart = NimTrendChart()
        self.addCleanup(chart.deleteLater)
        chart.set_trend((("2026-01", 2.0, 3.0, 8.0),))
        chart.set_metric_mode("rate")
        self.assertEqual(len(chart._series_payload()), 1)
        self.assertEqual(chart._series_payload()[0][0], "Lãi suất bình quân")

    def test_nim_debt_chart_switch_mode_reuses_loaded_data(self) -> None:
        chart = NimTrendChart()
        self.addCleanup(chart.deleteLater)
        data = (("2026-01", 2.0, 3.0, 8.0), ("2026-02", 2.2, 3.2, 8.2))
        chart.set_trend(data)
        before = chart.points_data
        chart.set_metric_mode("rate")
        chart.set_metric_mode("nim")
        self.assertEqual(chart.points_data, before)

    def test_nim_debt_chart_tooltip_matches_mode(self) -> None:
        chart = NimTrendChart()
        self.addCleanup(chart.deleteLater)
        chart.set_trend((("2026-07", 2.53, 2.61, 8.59),))
        self.assertIn("NIM sau ĐC", "\n".join(label for label, _color, _values in chart._series_payload()))
        chart.set_metric_mode("rate")
        self.assertEqual([label for label, _color, _values in chart._series_payload()], ["Lãi suất bình quân"])

    def test_chart_legend_layout_at_125_percent_dpi(self) -> None:
        chart = NimTrendChart()
        self.addCleanup(chart.deleteLater)
        chart.resize(800, 325)
        chart.set_trend(tuple((f"2026-{month:02d}", 2.0, 2.1, 8.0) for month in range(1, 8)))
        self.assertFalse(chart.legend_overlaps_x_axis())

    def test_chart_legend_layout_at_150_percent_dpi(self) -> None:
        chart = NimTrendChart()
        self.addCleanup(chart.deleteLater)
        chart.resize(960, 390)
        chart.set_trend(tuple((f"2026-{month:02d}", 2.0, 2.1, 8.0) for month in range(1, 8)))
        self.assertFalse(chart.legend_overlaps_x_axis())

    def test_dashboard_overview_legend_does_not_overlap_axis(self) -> None:
        window = self._dashboard_mode_window()
        window.overview_chart.resize(800, 320)
        self.assertFalse(window.overview_chart.legend_overlaps_x_axis())

    def test_period_table_has_all_periods_mode(self) -> None:
        window = self._dashboard_mode_window()
        self.assertEqual(window.overview_table_mode, "all")

    def test_period_table_has_endpoint_comparison_mode(self) -> None:
        window = self._dashboard_mode_window()
        window._overview_mode_changed(OVERVIEW_MODE_ENDPOINTS)
        self.assertEqual(window.overview_table_mode, OVERVIEW_MODE_ENDPOINTS)

    def test_period_table_all_periods_returns_intermediate_periods(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).overview_by_period_rows()
        self.assertEqual([row["Kỳ"] for row in rows], ["2026-01", "2026-02", "2026-03"])

    def test_period_table_has_all_periods_mode_returns_intermediate_periods_alias(self) -> None:
        self.test_period_table_all_periods_returns_intermediate_periods()

    def test_period_table_endpoint_mode_returns_only_from_and_to(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).overview_endpoint_rows()
        self.assertEqual([row["Kỳ"] for row in rows], ["2026-01", "2026-03"])

    def test_period_table_same_from_to_returns_one_row(self) -> None:
        data = self._seed_dashboard_mode_data(DashboardFilters(period_from="2026-02", period_to="2026-02", customer_type="Cá nhân (CN)"))
        rows = DashboardNimExportService(data).overview_endpoint_rows()
        self.assertEqual([row["Kỳ"] for row in rows], ["2026-02"])
        self.assertIsNone(rows[0]["Tăng/giảm dư nợ tuyệt đối"])

    def test_endpoint_growth_compares_to_from_period(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).overview_endpoint_rows()
        self.assertEqual(rows[-1]["Tăng/giảm dư nợ tuyệt đối"], 3000.0)
        self.assertAlmostEqual(rows[-1]["Tăng trưởng dư nợ (%)"], 50.0)

    def test_period_table_export_respects_mode(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).overview_endpoint_rows()
        output = export_dashboard_rows(rows, self.root / "overview-mode.xlsx", sheet_name=SHEET_OVERVIEW, metadata=[("Chế độ bảng", "So sánh Từ kỳ và Đến kỳ")])
        workbook = load_workbook(output, data_only=True)
        try:
            self.assertEqual(workbook[SHEET_OVERVIEW].max_row, 3)
            self.assertEqual(workbook["ThongTin"]["A2"].value, "Chế độ bảng")
        finally:
            workbook.close()

    def test_period_chart_still_shows_all_periods_in_endpoint_table_mode(self) -> None:
        window = self._dashboard_mode_window()
        window._overview_mode_changed(OVERVIEW_MODE_ENDPOINTS)
        self.assertEqual(len(window.overview_chart.series[0].values), 3)
        self.assertEqual(window.overview_table.rowCount(), 2)

    def test_branch_comparison_has_two_modes(self) -> None:
        window = self._dashboard_mode_window()
        labels = [radio.text() for radio in window.findChildren(QRadioButton)]
        self.assertIn("So sánh Từ kỳ đến kỳ", labels)
        self.assertIn("Kỳ hiện tại", labels)

    def test_branch_comparison_default_is_from_to(self) -> None:
        window = self._dashboard_mode_window()
        self.assertEqual(window.branch_compare_mode, BRANCH_MODE_PERIOD_COMPARE)

    def test_branch_from_to_returns_two_period_values(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).branch_period_comparison_rows()
        self.assertIn("Dư nợ Từ kỳ", rows[0])
        self.assertIn("Dư nợ Đến kỳ", rows[0])

    def test_branch_current_mode_uses_to_period(self) -> None:
        window = self._dashboard_mode_window()
        window._branch_mode_changed(BRANCH_MODE_CURRENT)
        periods = {window.branch_table.item(row, 0).text() for row in range(window.branch_table.rowCount())}
        self.assertEqual(periods, {"2026-03"})

    def test_branch_from_to_debt_change(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data, metric=METRIC_BALANCE).branch_period_comparison_rows()
        loc_phat = [row for row in rows if row["Mã chi nhánh"] == "5491"][0]
        self.assertEqual(loc_phat["Tăng/giảm tuyệt đối"], 2000.0)
        self.assertAlmostEqual(loc_phat["Tăng trưởng (%)"], 50.0)

    def test_branch_from_to_nim_change_percentage_points(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data, metric=METRIC_NIM_AFTER).branch_period_comparison_rows()
        loc_phat = [row for row in rows if row["Mã chi nhánh"] == "5491"][0]
        self.assertIn("NIM sau ĐC Từ kỳ", loc_phat)
        self.assertAlmostEqual(loc_phat["Thay đổi (điểm %)"], -2.0)

    def test_branch_from_to_rate_change_percentage_points(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data, metric=METRIC_AVERAGE_RATE).branch_period_comparison_rows()
        loc_phat = [row for row in rows if row["Mã chi nhánh"] == "5491"][0]
        self.assertIn("Lãi suất bình quân Từ kỳ", loc_phat)
        self.assertAlmostEqual(loc_phat["Thay đổi (điểm %)"], -1.833333333333333)

    def test_branch_current_chart_sorted_by_metric(self) -> None:
        data = self._seed_dashboard_mode_data()
        bars = branch_bar_values(data.branch_rows, METRIC_BALANCE)
        self.assertGreaterEqual(bars[0][2] or 0, bars[1][2] or 0)

    def test_branch_comparison_export_respects_mode(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data, metric=METRIC_NIM_AFTER).branch_period_comparison_rows()
        output = export_dashboard_rows(rows, self.root / "branch-mode.xlsx", sheet_name=SHEET_BRANCH, metadata=[("Chế độ bảng", "So sánh Từ kỳ đến kỳ")])
        workbook = load_workbook(output, data_only=True)
        try:
            headers = [workbook[SHEET_BRANCH].cell(1, column).value for column in range(1, workbook[SHEET_BRANCH].max_column + 1)]
            self.assertIn("Thay đổi (điểm %)", headers)
            self.assertEqual(workbook["ThongTin"]["B2"].value, "So sánh Từ kỳ đến kỳ")
        finally:
            workbook.close()

    def test_branch_chart_uses_horizontal_layout(self) -> None:
        data = self._seed_dashboard_mode_data()
        pairs, from_period, to_period = branch_period_pair_values(data.branch_rows, METRIC_BALANCE, from_period="2026-01", to_period="2026-03")
        chart = DashboardBranchComparisonChart()
        self.addCleanup(chart.deleteLater)
        chart.set_pairs(pairs, value_kind="money", metric_label="Dư nợ", from_period=from_period, to_period=to_period)
        self.assertEqual(chart.orientation, "horizontal_grouped")

    def test_branch_long_names_do_not_overlap(self) -> None:
        self.repository.unit_directory.save_branch(BranchDirectoryEntry(branch_code="5491", branch_name="Chi nhánh Lộc Phát Lâm Đồng tên rất dài", short_name=""))
        data = self._seed_dashboard_mode_data()
        pairs, _from_period, _to_period = branch_period_pair_values(data.branch_rows, METRIC_BALANCE, from_period="2026-01", to_period="2026-03")
        self.assertTrue(any("tên rất dài" in label for _code, label, _from_value, _to_value in pairs))

    def test_detail_table_has_branch_and_office_modes(self) -> None:
        window = self._dashboard_mode_window()
        labels = [radio.text() for radio in window.findChildren(QRadioButton)]
        self.assertIn("Tổng hợp theo chi nhánh", labels)
        self.assertIn("Chi tiết theo Hội sở/PGD", labels)

    def test_detail_branch_mode_one_row_per_branch_period(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).detail_branch_rows()
        self.assertEqual(len([row for row in rows if row["Kỳ"] == "2026-01" and row["Mã chi nhánh"] == "5491"]), 1)
        self.assertNotIn("Hội sở/Phòng GD", rows[0])

    def test_detail_branch_mode_weighted_average_rate(self) -> None:
        data = self._seed_dashboard_mode_data()
        row = [item for item in DashboardNimExportService(data).detail_branch_rows() if item["Kỳ"] == "2026-01" and item["Mã chi nhánh"] == "5491"][0]
        self.assertAlmostEqual(row["Lãi suất bình quân"], 8.5)

    def test_detail_branch_mode_weighted_nim_before(self) -> None:
        data = self._seed_dashboard_mode_data()
        row = [item for item in DashboardNimExportService(data).detail_branch_rows() if item["Kỳ"] == "2026-01" and item["Mã chi nhánh"] == "5491"][0]
        self.assertAlmostEqual(row["NIM trước ĐC"], 7.25)

    def test_detail_branch_mode_weighted_nim_after(self) -> None:
        data = self._seed_dashboard_mode_data()
        row = [item for item in DashboardNimExportService(data).detail_branch_rows() if item["Kỳ"] == "2026-01" and item["Mã chi nhánh"] == "5491"][0]
        self.assertAlmostEqual(row["NIM sau ĐC"], 7.0)

    def test_detail_office_mode_one_row_per_office_period(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).detail_office_rows()
        self.assertEqual(len([row for row in rows if row["Kỳ"] == "2026-01" and row["Mã chi nhánh"] == "5491"]), 2)

    def test_detail_office_mode_dynamic_office_name(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).detail_office_rows()
        self.assertIn("PGD Đức Trọng", {row["Hội sở/Phòng GD"] for row in rows})

    def test_detail_office_mode_head_office_type(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).detail_office_rows()
        head = [row for row in rows if row["Mã đơn vị"] == "5491-00"][0]
        self.assertEqual(head["Loại đơn vị"], "Hội sở")

    def test_detail_office_mode_pgd_type(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).detail_office_rows()
        pgd = [row for row in rows if row["Mã đơn vị"] == "5491-01"][0]
        self.assertEqual(pgd["Loại đơn vị"], "Phòng giao dịch")

    def test_detail_growth_compares_same_entity(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).detail_office_rows()
        pgd_march = [row for row in rows if row["Kỳ"] == "2026-03" and row["Mã đơn vị"] == "5491-01"][0]
        self.assertEqual(pgd_march["Tăng/giảm dư nợ tuyệt đối"], 500.0)
        self.assertAlmostEqual(pgd_march["Tăng trưởng dư nợ (%)"], 14.2857142857)

    def test_detail_export_respects_grouping_mode(self) -> None:
        data = self._seed_dashboard_mode_data()
        rows = DashboardNimExportService(data).detail_branch_rows()
        output = export_dashboard_rows(rows, self.root / "detail-mode.xlsx", sheet_name=SHEET_DETAIL, metadata=[("Chế độ bảng", "Tổng hợp theo chi nhánh")])
        workbook = load_workbook(output, data_only=True)
        try:
            headers = [workbook[SHEET_DETAIL].cell(1, column).value for column in range(1, workbook[SHEET_DETAIL].max_column + 1)]
            self.assertNotIn("Hội sở/Phòng GD", headers)
            self.assertEqual(workbook["ThongTin"]["B2"].value, "Tổng hợp theo chi nhánh")
        finally:
            workbook.close()

    def test_detail_table_uses_limit_offset(self) -> None:
        self._seed_dashboard_mode_data()
        repository = NimDashboardRepository(self.repository)
        filters = DashboardFilters(period_from="2026-01", period_to="2026-03", customer_type="Cá nhân (CN)").as_query_filters()
        page = repository.detail_summary(SummaryDataType.NIM_DN, filters, limit=1, offset=1)
        self.assertEqual(len(page), 1)
        self.assertGreater(repository.detail_summary_count(SummaryDataType.NIM_DN, filters), 1)

    def test_officer_multi_select_combo_suppresses_popup_close_on_row_toggle(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = OfficerMultiSelectCombo()
        try:
            combo.set_officers([officer_key("[540000321] Nguyễn Văn A"), officer_key("[540000322] Nguyễn Văn B")])
            combo._toggle_item(combo.model().index(1, 0))
            self.assertEqual([item.code for item in combo.selected_officers()], ["540000321"])
            combo._suppress_next_hide = True
            combo.hidePopup()
            self.assertFalse(combo._suppress_next_hide)
        finally:
            combo.deleteLater()

    def test_select_more_than_8_officers_is_allowed(self) -> None:
        officers = self._seed_many_nim_dn_officers(12)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, officers[0])
        try:
            dialog.officer_selector.select_all()
            with patch("agribank_v3.features.credit.summary.officer_history.window.QMessageBox.warning") as warning:
                dialog.apply_officer_comparison()
            self.assertFalse(warning.called)
            self.assertEqual(len({row.officer.code for row in dialog.compare_rows}), 12)
            self.assertEqual(dialog.compare_status_label.text(), "Đã chọn 12 cán bộ. Biểu đồ đang hiển thị 8 cán bộ theo từng trang.")
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_officer_chart_pagination(self) -> None:
        officers = self._seed_many_nim_dn_officers(18)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, officers[0])
        try:
            dialog.officer_selector.select_all()
            dialog.apply_officer_comparison()
            self.assertEqual(dialog._compare_total_pages(), 3)
            self.assertEqual(len(dialog.officer_compare_chart.series), 8)
            self.assertEqual(dialog.compare_page_label.text(), "Trang 1/3")
            dialog.next_compare_chart_page()
            self.assertEqual(len(dialog.officer_compare_chart.series), 8)
            self.assertEqual(dialog.compare_page_label.text(), "Trang 2/3")
            dialog.next_compare_chart_page()
            self.assertEqual(len(dialog.officer_compare_chart.series), 2)
            self.assertEqual(dialog.compare_page_label.text(), "Trang 3/3")
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_chart_page_does_not_limit_table_data(self) -> None:
        officers = self._seed_many_nim_dn_officers(18)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, officers[0])
        try:
            dialog.officer_selector.select_all()
            dialog.apply_officer_comparison()
            self.assertEqual(len(dialog.officer_compare_chart.series), 8)
            self.assertEqual(dialog.officer_compare_table.rowCount(), 18)
            dialog.next_compare_chart_page()
            self.assertEqual(len(dialog.officer_compare_chart.series), 8)
            self.assertEqual(dialog.officer_compare_table.rowCount(), 18)
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_export_contains_all_selected_officers(self) -> None:
        officers = self._seed_many_nim_dn_officers(18)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, officers[0])
        try:
            dialog.officer_selector.select_all()
            dialog.apply_officer_comparison()
            dialog.tabs.setCurrentIndex(2)
            tab_key, rows = dialog._current_export_rows()
            output = export_analysis_rows(rows, self.root / "compare_all.xlsx", tab_key=tab_key, metadata=dialog._export_metadata(tab_key))

            workbook = load_workbook(output, data_only=True)
            try:
                worksheet = workbook["SoSanhCBTD"]
                exported = {worksheet.cell(row, 2).value for row in range(2, worksheet.max_row + 1)}
                self.assertEqual(len(exported), 18)
                self.assertIn("ThongTin", workbook.sheetnames)
                self.assertEqual(workbook["ThongTin"]["B3"].value, 18)
            finally:
                workbook.close()
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_top_n_officers(self) -> None:
        officers = self._seed_many_nim_dn_officers(18)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, officers[0])
        try:
            dialog.officer_selector.select_all()
            dialog.apply_officer_comparison()
            dialog.compare_mode_combo.setCurrentIndex(dialog.compare_mode_combo.findText("Top N cán bộ"))
            dialog.compare_top_combo.setCurrentIndex(dialog.compare_top_combo.findData(10))
            dialog._render_officer_compare_chart()

            labels = [series.label for series in dialog.officer_compare_chart.series]
            self.assertEqual(len(labels), 10)
            self.assertEqual(labels[0], "Cán bộ 18")
            self.assertEqual(labels[-1], "Cán bộ 09")
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_select_all_officers(self) -> None:
        officers = self._seed_many_nim_dn_officers(18)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, officers[0])
        try:
            dialog.officer_selector.select_all()
            self.assertEqual(dialog.officer_selector.selected_count(), 18)
            dialog.apply_officer_comparison()
            self.assertEqual(len({row.officer.code for row in dialog.compare_rows}), 18)
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_officer_popup_stays_open(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = OfficerMultiSelectCombo()
        try:
            combo.set_officers([officer_key("[540000321] Nguyễn Văn A"), officer_key("[540000322] Nguyễn Văn B")])
            combo._toggle_item(combo.model().index(1, 0))
            combo._suppress_next_hide = True
            combo.hidePopup()
            self.assertFalse(combo._suppress_next_hide)
            self.assertEqual([item.code for item in combo.selected_officers()], ["540000321"])
        finally:
            combo.deleteLater()

    def test_page_navigation_does_not_query_database_again(self) -> None:
        officers = self._seed_many_nim_dn_officers(18)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_DN, officers[0])
        try:
            dialog.officer_selector.select_all()
            dialog.apply_officer_comparison()
            with patch("agribank_v3.features.credit.summary.officer_history.window.build_multiple_officer_comparison") as query:
                dialog.next_compare_chart_page()
                dialog.previous_compare_chart_page()
            self.assertFalse(query.called)
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_nim_nv_more_than_8_officers(self) -> None:
        officers = self._seed_many_nim_nv_officers(12)
        dialog = self._open_officer_history_dialog(SummaryDataType.NIM_NV, officers[0])
        try:
            dialog.officer_selector.select_all()
            dialog.apply_officer_comparison()
            self.assertEqual(len({row.officer.code for row in dialog.compare_rows}), 12)
            self.assertEqual(len(dialog.officer_compare_chart.series), 8)
            self.assertEqual(dialog.officer_compare_table.rowCount(), 12)
            self.assertEqual(dialog.compare_page_label.text(), "Trang 1/2")
        finally:
            dialog.close()
            dialog.deleteLater()

    def test_nim_nv_main_table_responsive_columns(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            ["5491,10,2,1,[540500321] Nguyễn Văn A,00,1000,,CN"],
        )
        import_nim_nv(self.repository, self.root)

        tab = NimTab(self.repository, SummaryDataType.NIM_NV)
        try:
            rows = tab.query_page().rows
            tab._render_table(rows)
            tab._restore_column_widths()
            headers = [tab.table.horizontalHeaderItem(index).text() for index in range(tab.table.columnCount())]
            self.assertEqual(
                headers,
                ["Kỳ", "Tên chi nhánh", "Phòng GD", "Loại KH", "Người quản lý NV", "Số dư nguồn vốn", "NIM trước ĐC", "NIM sau ĐC"],
            )
            self.assertNotIn("Lãi suất bình quân", headers)
            self.assertIn("Mở Dashboard", [button.text() for button in tab.findChildren(QPushButton)])
            self.assertEqual([metric.label for metric in tab.query_dashboard().metrics], ["Tổng nguồn vốn", "NIM trước ĐC", "NIM sau ĐC"])
            self.assertLessEqual(sum(tab.table.columnWidth(index) for index in range(tab.table.columnCount())), tab.table.viewport().width() + 2)
            self.assertEqual(tab.table.item(0, 0).textAlignment(), Qt.AlignmentFlag.AlignCenter)
            for column in (5, 6, 7):
                self.assertEqual(tab.table.item(0, column).textAlignment(), Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        finally:
            tab.deleteLater()

    def test_nim_nv_all_customer_types_grouped_per_officer(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            [
                "5491,2,10,1,[540500321] Nguyễn Văn A,00,1000,,CN",
                "5491,4,8,1,[540500321] Nguyễn Văn A,00,3000,,TC",
            ],
        )
        import_nim_nv(self.repository, self.root)

        tab = NimTab(self.repository, SummaryDataType.NIM_NV)
        try:
            all_types = tab.query_page()
            self.assertEqual(all_types.total_rows, 1)
            row = all_types.rows[0]
            self.assertEqual(row["officer"], "[540500321] Nguyễn Văn A")
            self.assertEqual(row["customer_type"], "Tất cả")
            self.assertAlmostEqual(float(row["balance"]), 4000.0)
            self.assertAlmostEqual(float(row["nim_before"]), -5.0)
            self.assertAlmostEqual(float(row["nim_after"]), -4.0)
            self.assertNotIn("average_rate", row)

            tab.customer_type_filter.setCurrentIndex(tab.customer_type_filter.findData("Cá nhân (CN)"))
            personal = tab.query_page()
            self.assertEqual(personal.total_rows, 1)
            self.assertEqual(personal.rows[0]["customer_type"], "Cá nhân (CN)")
            self.assertAlmostEqual(float(personal.rows[0]["balance"]), 1000.0)
            self.assertAlmostEqual(float(personal.rows[0]["nim_before"]), -8.0)
            self.assertAlmostEqual(float(personal.rows[0]["nim_after"]), -7.0)
        finally:
            tab.deleteLater()

    def test_nim_nv_officer_history(self) -> None:
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            [
                "5491,2,10,1,[540500321] Nguyễn Văn A,00,1000,,CN",
                "5491,4,8,1,[540500321] Nguyễn Văn A,00,3000,,TC",
            ],
        )
        self._write_nim_nv_file(
            "5491_FTPDP_20260228.csv",
            ["5491,4,8,1,[540500321] Nguyễn Văn A,00,2000,,CN"],
        )
        import_nim_nv(self.repository, self.root)

        overview = build_officer_overview(
            self.repository,
            SummaryDataType.NIM_NV,
            officer_code="540500321",
            officer="[540500321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
        )

        self.assertEqual([point.period for point in overview.points], ["2026-01", "2026-02"])
        self.assertEqual(overview.officer.display_name, "Nguyễn Văn A")
        self.assertAlmostEqual(overview.points[0].balance, 4000.0)
        self.assertAlmostEqual(overview.points[0].nim_before, -5.0)
        self.assertAlmostEqual(overview.points[0].nim_after, -4.0)
        self.assertAlmostEqual(overview.current_balance, 2000.0)
        self.assertAlmostEqual(overview.current_nim_before, -4.0)
        self.assertAlmostEqual(overview.current_nim_after, -3.0)

    def test_nim_nv_growth_history(self) -> None:
        history_points = (
            HistoryPoint("2026-01", 1000.0, 0.0, -8.0, -7.0),
            HistoryPoint("2026-02", 1500.0, 0.0, -5.0, -4.0),
            HistoryPoint("2026-03", 1200.0, 0.0, -4.0, -3.0),
        )

        growth = build_officer_growth_history(history_points)

        self.assertIsNone(growth[0].delta)
        self.assertIsNone(growth[0].growth_percent)
        self.assertAlmostEqual(growth[1].delta, 500.0)
        self.assertAlmostEqual(growth[1].growth_percent, 50.0)
        self.assertAlmostEqual(growth[2].delta, -300.0)
        self.assertAlmostEqual(growth[2].growth_percent, -20.0)

    def test_nim_nv_zero_previous_balance(self) -> None:
        history_points = (
            HistoryPoint("2026-01", 0.0, 0.0, 0.0, 0.0),
            HistoryPoint("2026-02", 1000.0, 0.0, -8.0, -7.0),
        )

        growth = build_officer_growth_history(history_points)

        self.assertAlmostEqual(growth[1].delta, 1000.0)
        self.assertIsNone(growth[1].growth_percent)

    def test_nim_nv_multiple_officer_comparison(self) -> None:
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            [
                "5491,10,2,1,[540500321] Nguyễn Văn A,00,1000,,CN",
                "5491,8,4,1,[540500322] Nguyễn Văn B,00,3000,,CN",
            ],
        )
        import_nim_nv(self.repository, self.root)

        rows = build_multiple_officer_comparison(
            self.repository,
            SummaryDataType.NIM_NV,
            officers=[
                officer_key("[540500321] Nguyễn Văn A"),
                officer_key("[540500322] Nguyễn Văn B"),
            ],
            metric=METRIC_BALANCE,
            branch="5491 - CN Lộc Phát",
        )

        values = {(row.officer.code, row.period): row.value for row in rows}
        self.assertEqual(values[("540500321", "2026-01")], 1000)
        self.assertEqual(values[("540500322", "2026-01")], 3000)

    def test_nim_nv_branch_weighted_nim(self) -> None:
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            [
                "5491,2,10,1,[540500321] Nguyễn Văn A,00,100,,CN",
                "5491,1,2,0.5,[540500322] Nguyễn Văn B,00,300,,CN",
            ],
        )
        import_nim_nv(self.repository, self.root)

        _series, rows = build_officer_branch_comparison(
            self.repository,
            SummaryDataType.NIM_NV,
            officer_code="540500321",
            officer="[540500321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
            metric=METRIC_NIM_BEFORE,
            filters=HistoryFilters(customer_type="Cá nhân (CN)"),
        )

        branch_row = [row for row in rows if row.officer.display_name == "5491 - CN Lộc Phát"][0]
        self.assertAlmostEqual(branch_row.value or 0, -2.75)
        self.assertNotAlmostEqual(branch_row.value or 0, -4.5)

    def test_nim_nv_dashboard_overview(self) -> None:
        data = self._seed_nim_nv_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))

        rows = DashboardNimExportService(data).overview_by_period_rows()

        self.assertEqual(data.ui_config, NIM_NV_UI_CONFIG)
        self.assertEqual([metric.label for metric in data.kpis], ["Số chi nhánh", "Tổng nguồn vốn", "NIM trước ĐC bình quân", "NIM sau ĐC bình quân", "Tăng trưởng nguồn vốn kỳ gần nhất"])
        self.assertEqual(list(rows[0].keys()), ["Kỳ", "Tổng nguồn vốn", "NIM trước ĐC", "NIM sau ĐC", "Tăng/giảm nguồn vốn tuyệt đối", "Tăng trưởng nguồn vốn (%)"])
        self.assertNotIn("Lãi suất bình quân", rows[0])

    def test_nim_nv_dashboard_branch_comparison(self) -> None:
        data = self._seed_nim_nv_dashboard_data(DashboardFilters(period_from="2026-02", period_to="2026-02", customer_type="Cá nhân (CN)", metric=METRIC_NIM_AFTER))
        rows = DashboardNimExportService(data, metric=METRIC_NIM_AFTER).branch_comparison_rows()

        self.assertEqual(list(rows[0].keys()), ["Kỳ", "Tên chi nhánh", "Số dư nguồn vốn", "NIM trước ĐC", "NIM sau ĐC", "Chỉ tiêu đang chọn", "Giá trị chỉ tiêu"])
        self.assertEqual({row["Tên chi nhánh"] for row in rows}, {"5400 - CN Lâm Đồng", "5491 - CN Lộc Phát"})
        self.assertEqual({row["Chỉ tiêu đang chọn"] for row in rows}, {"NIM sau ĐC"})
        self.assertFalse(any("..." in name for _period, name, _value in branch_bar_values(data.branch_rows, METRIC_BALANCE)))

    def test_nim_nv_dashboard_growth(self) -> None:
        data = self._seed_nim_nv_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))

        rows = DashboardNimExportService(data).growth_rows()

        self.assertIn("Số dư nguồn vốn", rows[0])
        self.assertIn("Tăng/giảm nguồn vốn tuyệt đối", rows[0])
        self.assertIn("Tăng trưởng nguồn vốn (%)", rows[0])
        self.assertNotIn("Dư nợ", rows[0])
        growth_values = sorted(row["Tăng trưởng nguồn vốn (%)"] for row in rows if row["Tăng trưởng nguồn vốn (%)"] is not None)
        self.assertEqual(growth_values, [25.0, 50.0])

    def test_nim_nv_dashboard_detail(self) -> None:
        data = self._seed_nim_nv_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))

        rows = DashboardNimExportService(data).detail_rows()

        self.assertEqual(list(rows[0].keys()), ["Kỳ", "Tên chi nhánh", "Phòng GD", "Loại KH", "Số dư nguồn vốn", "NIM trước ĐC", "NIM sau ĐC", "Tăng trưởng nguồn vốn (%)", "Tăng/giảm nguồn vốn tuyệt đối"])
        self.assertNotIn("Lãi suất bình quân", rows[0])
        self.assertEqual({row["Loại KH"] for row in rows}, {"Cá nhân (CN)"})

    def test_nim_nv_export_each_tab(self) -> None:
        data = self._seed_nim_nv_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))
        service = DashboardNimExportService(data, metric=METRIC_NIM_AFTER)

        exports = {
            "overview": service.export_overview_by_period(self.root / "nv_overview.xlsx"),
            "branch": service.export_branch_comparison(self.root / "nv_branch.xlsx"),
            "growth": service.export_growth(self.root / "nv_growth.xlsx"),
            "detail": service.export_detail(self.root / "nv_detail.xlsx"),
        }

        for tab_key, output in exports.items():
            workbook = load_workbook(output, data_only=True)
            try:
                worksheet = workbook[NIM_NV_UI_CONFIG.dashboard_sheets[tab_key]]
                headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
                self.assertNotIn("Dư nợ", headers)
                self.assertNotIn("Lãi suất bình quân", headers)
                self.assertEqual(worksheet.freeze_panes, "A2")
                self.assertEqual(worksheet.auto_filter.ref, worksheet.dimensions)
            finally:
                workbook.close()

    def test_nim_nv_export_all_tabs(self) -> None:
        data = self._seed_nim_nv_dashboard_data(DashboardFilters(customer_type="Cá nhân (CN)"))
        output = DashboardNimExportService(data, metric=METRIC_NIM_AFTER).export_all_tabs(self.root / "nim_nv_all.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["TongQuanNIMNV", "SoSanhChiNhanhNIMNV", "TangTruongNIMNV", "BangDuLieuNIMNV"])
            self.assertEqual(workbook["BangDuLieuNIMNV"].cell(1, 5).value, "Số dư nguồn vốn")
        finally:
            workbook.close()

    def test_nim_nv_multiselect_popup_stays_open(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        combo = OfficerMultiSelectCombo(placeholder="Chọn cán bộ", counter_label="cán bộ")
        try:
            combo.set_officers([officer_key("[540500321] Nguyễn Văn A"), officer_key("[540500322] Nguyễn Văn B")])
            self.assertEqual(combo.lineEdit().text(), "Chọn cán bộ")
            combo._toggle_item(combo.model().index(1, 0))
            self.assertEqual([item.code for item in combo.selected_officers()], ["540500321"])
            combo._toggle_item(combo.model().index(2, 0))
            self.assertEqual(combo.lineEdit().text(), "2 cán bộ đã chọn")
            combo._suppress_next_hide = True
            combo.hidePopup()
            self.assertFalse(combo._suppress_next_hide)
            self.assertEqual([item.code for item in combo.selected_officers()], ["540500321", "540500322"])
        finally:
            combo.deleteLater()

    def test_nim_nv_multiline_header(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        window = NimDashboardWindow(self.repository, SummaryDataType.NIM_NV)
        try:
            self.assertEqual(window.windowTitle(), "Dashboard NIM nguồn vốn - AgribankV3")
            headers = [window.detail_table.horizontalHeaderItem(index).text() for index in range(window.detail_table.columnCount())]
            self.assertIn("Số dư nguồn vốn", headers)
            self.assertNotIn("Dư nợ", headers)
            window.detail_table.setColumnWidth(8, 70)
            header = window.detail_table.horizontalHeader()
            header.refresh_height()
            self.assertGreater(header.height(), 38)
        finally:
            window.close()
            window.deleteLater()

    def test_nim_nv_non_modal_windows(self) -> None:
        app = QApplication.instance() or QApplication([])
        _ = app
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            ["5491,10,2,1,[540500321] Nguyễn Văn A,00,1000,,CN"],
        )
        import_nim_nv(self.repository, self.root)

        tab = NimTab(self.repository, SummaryDataType.NIM_NV)
        dashboard = NimDashboardWindow(self.repository, SummaryDataType.NIM_NV, parent=tab)
        history = OfficerHistoryDialog(
            self.repository,
            SummaryDataType.NIM_NV,
            officer="[540500321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
            transaction_office="Hội sở",
            customer_type="Cá nhân (CN)",
            parent=tab,
        )
        try:
            tab.open_dashboard()
            self.assertFalse(dashboard.isModal())
            self.assertEqual(dashboard.windowModality(), Qt.WindowModality.NonModal)
            self.assertFalse(history.isModal())
            self.assertEqual(history.windowModality(), Qt.WindowModality.NonModal)
            self.assertEqual(len(tab.dashboard_windows), 1)
            self.assertFalse(tab.dashboard_windows[0].isModal())
        finally:
            for dialog in list(tab.dashboard_windows):
                dialog.close()
                dialog.deleteLater()
            history.close()
            history.deleteLater()
            dashboard.close()
            dashboard.deleteLater()
            tab.deleteLater()

    def test_nim_nv_percentage_not_multiplied_twice(self) -> None:
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            ["5491,10,7.684322,0,[540500321] Nguyễn Văn A,00,1000,,CN"],
        )
        import_nim_nv(self.repository, self.root)
        data = build_nim_dashboard(
            NimDashboardRepository(self.repository),
            SummaryDataType.NIM_NV,
            DashboardFilters(customer_type="Cá nhân (CN)"),
        )
        output = DashboardNimExportService(data).export_overview_by_period(self.root / "nv_percent.xlsx")

        workbook = load_workbook(output, data_only=True)
        try:
            worksheet = workbook["TongQuanNIMNV"]
            self.assertAlmostEqual(worksheet["C2"].value, 2.315678)
            self.assertEqual(worksheet["C2"].number_format, '0.00"%"')
            self.assertNotAlmostEqual(worksheet["C2"].value, 0.02315678)
        finally:
            workbook.close()

    def _seed_dashboard_data(self, filters: DashboardFilters) -> object:
        january = self.root / "5491_FTPLN_20260131.csv"
        february = self.root / "5491_FTPLN_20260228.csv"
        january.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                    "5400,1,6,0,[540000322] Nguyễn Văn B,00,2000,,CN",
                    "5491,3,12,1,[540000323] Nguyễn Văn C,00,7000,,TC",
                ]
            ),
            encoding="utf-8",
        )
        february.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,3,8,1,[540000321] Nguyễn Văn A,00,1500,,CN",
                    "5400,1,6,0,[540000322] Nguyễn Văn B,00,2500,,CN",
                    "5491,3,12,1,[540000323] Nguyễn Văn C,00,9000,,TC",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)
        return build_nim_dashboard(NimDashboardRepository(self.repository), SummaryDataType.NIM_DN, filters)

    def _write_nim_dn_file(self, filename: str, rows: list[str]) -> Path:
        path = self.root / filename
        path.write_text(
            "\n".join(["BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP", *rows]),
            encoding="utf-8",
        )
        return path

    def _write_nim_nv_file(self, filename: str, rows: list[str]) -> Path:
        path = self.root / filename
        path.write_text(
            "\n".join(["BRCD,FTP,INTRT,MUCFTPDC,CBHD,TRCTCD,LDRBAL,TRREF,CUSTTP", *rows]),
            encoding="utf-8",
        )
        return path

    def _seed_many_nim_dn_officers(self, count: int) -> list[str]:
        officers = [f"[540{i:06d}] Cán bộ {i:02d}" for i in range(1, count + 1)]
        self._write_nim_dn_file(
            "5491_FTPLN_20260131.csv",
            [
                f"5491,2,10,1,{officer},00,{index * 1000},,CN"
                for index, officer in enumerate(officers, start=1)
            ],
        )
        import_nim_dn(self.repository, self.root)
        return officers

    def _seed_many_nim_nv_officers(self, count: int) -> list[str]:
        officers = [f"[550{i:06d}] Cán bộ NV {i:02d}" for i in range(1, count + 1)]
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            [
                f"5491,10,2,1,{officer},00,{index * 1000},,CN"
                for index, officer in enumerate(officers, start=1)
            ],
        )
        import_nim_nv(self.repository, self.root)
        return officers

    def _open_officer_history_dialog(self, data_type: SummaryDataType, officer: str) -> OfficerHistoryDialog:
        app = QApplication.instance() or QApplication([])
        _ = app
        return OfficerHistoryDialog(
            self.repository,
            data_type,
            officer=officer,
            branch="5491 - CN Lộc Phát",
            transaction_office="Hội sở",
            customer_type="Cá nhân (CN)",
        )

    def _seed_loan_compare_rows(self) -> None:
        previous = self.root / "loan_previous.csv"
        previous.write_text(
            "\n".join(
                [
                    "CUSTSEQ,CUSTNM,DU_NO,ADDR1,OFFICER_NAME",
                    "000123,Khach tat toan,13500000000,Dia chi rat dai khong lam cot qua rong,CB1",
                    "000124,Khach tang,1000000000,Dia chi 2,CB2",
                ]
            ),
            encoding="utf-8",
        )
        current = self.root / "loan_current.csv"
        current.write_text(
            "\n".join(
                [
                    "CUSTSEQ,CUSTNM,DU_NO,ADDR1,OFFICER_NAME",
                    "000124,Khach tang,1500000000,Dia chi 2,CB2",
                ]
            ),
            encoding="utf-8",
        )
        compare_loan_balances(self.repository, previous, current)

    def _count_nim_rows(
        self,
        data_type: SummaryDataType,
        period: str,
        *,
        repository: SummaryRepository | None = None,
    ) -> int:
        repo = repository or self.repository
        with closing(repo.connect()) as connection:
            return int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM nim_period_summary
                    WHERE data_type = ? AND period = ?
                    """,
                    (data_type.value, period),
                ).fetchone()[0]
                or 0
            )

    def _count_raw_nim_rows(
        self,
        data_type: SummaryDataType,
        period: str,
        *,
        repository: SummaryRepository | None = None,
    ) -> int:
        repo = repository or self.repository
        with closing(repo.connect()) as connection:
            return int(
                connection.execute(
                    """
                    SELECT COUNT(*)
                    FROM nim_details
                    WHERE data_type = ? AND period = ?
                    """,
                    (data_type.value, period),
                ).fetchone()[0]
                or 0
            )

    def _create_legacy_summary_database(self, database_path: Path) -> None:
        database_path.parent.mkdir(parents=True, exist_ok=True)
        with closing(sqlite3.connect(database_path)) as connection:
            connection.row_factory = sqlite3.Row
            apply_migrations(
                connection,
                (MigrationSpec(version="0.1.6", description="legacy summary schema"),),
                update_root=self.root,
            )
            now = "2026-03-31T08:00:00+07:00"
            cursor = connection.execute(
                """
                INSERT INTO summary_import_history(
                    data_type, period, source_path, file_name, imported_by,
                    row_count, duration_ms, version, status, message, source_hash,
                    created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, 1, 'success', ?, ?, ?, ?)
                """,
                (
                    SummaryDataType.NIM_DN.value,
                    "2026-03",
                    "legacy.csv",
                    "legacy.csv",
                    "tester",
                    1,
                    10,
                    "legacy",
                    "legacy-hash",
                    now,
                    now,
                ),
            )
            batch_id = int(cursor.lastrowid)
            connection.execute(
                """
                INSERT INTO nim_details(
                    batch_id, data_type, period, branch_code, branch_name, trctcd,
                    transaction_office, customer_type, officer, balance, interest_rate,
                    ftp_rate, adjustment_rate, numerator_before, numerator_after,
                    average_rate_numerator, source_file, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    SummaryDataType.NIM_DN.value,
                    "2026-03",
                    "5491",
                    "5491 - CN Lộc Phát",
                    "00",
                    "Hội sở",
                    "Cá nhân (CN)",
                    "CB1",
                    1000,
                    10,
                    2,
                    1,
                    8000,
                    7000,
                    10000,
                    "legacy.csv",
                    now,
                ),
            )
            connection.commit()

    def _seed_nim_nv_dashboard_data(self, filters: DashboardFilters) -> object:
        self._write_nim_nv_file(
            "5491_FTPDP_20260131.csv",
            [
                "5491,10,2,1,[540500321] Nguyễn Văn A,00,1000,,CN",
                "5400,6,1,0,[540500322] Nguyễn Văn B,00,2000,,CN",
                "5491,12,3,1,[540500323] Nguyễn Văn C,00,7000,,TC",
            ],
        )
        self._write_nim_nv_file(
            "5491_FTPDP_20260228.csv",
            [
                "5491,8,3,1,[540500321] Nguyễn Văn A,00,1500,,CN",
                "5400,6,1,0,[540500322] Nguyễn Văn B,00,2500,,CN",
                "5491,12,3,1,[540500323] Nguyễn Văn C,00,9000,,TC",
            ],
        )
        import_nim_nv(self.repository, self.root)
        return build_nim_dashboard(NimDashboardRepository(self.repository), SummaryDataType.NIM_NV, filters)

    def test_officer_history_no_data(self) -> None:
        history = build_officer_history(
            self.repository,
            SummaryDataType.NIM_DN,
            officer="[540000999] Không Có",
        )

        self.assertEqual(history.points, ())
        self.assertEqual(history.current_period, "")
        self.assertEqual(history.current_balance, 0)

    def test_officer_history_single_period(self) -> None:
        path = self.root / "5491_FTPLN_20260131.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        rows = self.repository.get_officer_history(SummaryDataType.NIM_DN, officer="[540000321] Nguyễn Văn A")
        history = build_officer_history(
            self.repository,
            SummaryDataType.NIM_DN,
            officer="[540000321] Nguyễn Văn A",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(len(history.points), 1)
        self.assertEqual(history.current_period, "2026-01")
        self.assertAlmostEqual(history.current_balance, 1000.0)
        self.assertAlmostEqual(history.current_average_rate, 10.0)
        self.assertAlmostEqual(history.current_nim_before, 8.0)
        self.assertAlmostEqual(history.current_nim_after, 7.0)

    def test_officer_history_multiple_periods_sorted_and_weighted(self) -> None:
        january = self.root / "5491_FTPLN_20260131.csv"
        february = self.root / "5491_FTPLN_20260228.csv"
        january.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                    "5491,4,8,1,[540000321] Nguyễn Văn A,00,3000,,TC",
                ]
            ),
            encoding="utf-8",
        )
        february.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,4,8,1,[540000321] Nguyễn Văn A,00,2000,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        history = build_officer_history(
            self.repository,
            SummaryDataType.NIM_DN,
            officer="[540000321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
            transaction_office="Hội sở",
        )

        self.assertEqual([point.period for point in history.points], ["2026-01", "2026-02"])
        self.assertAlmostEqual(history.points[0].balance, 4000.0)
        self.assertAlmostEqual(history.points[0].average_rate, 8.5)
        self.assertAlmostEqual(history.points[0].nim_before, 5.0)
        self.assertAlmostEqual(history.points[0].nim_after, 4.0)
        self.assertAlmostEqual(history.points[1].balance, 2000.0)
        self.assertAlmostEqual(history.current_nim_after, 3.0)

    def test_officer_history_export_excel(self) -> None:
        path = self.root / "5491_FTPLN_20260131.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)
        history = build_officer_history(
            self.repository,
            SummaryDataType.NIM_DN,
            officer="[540000321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
            transaction_office="Hội sở",
            customer_type="Cá nhân (CN)",
        )

        output_path = export_officer_history_excel(history, self.root / "history.xlsx")

        workbook = load_workbook(output_path, data_only=True)
        try:
            worksheet = workbook["LichSu_NIM_CBTD"]
            self.assertEqual(worksheet["A1"].value, "Lịch sử NIM CBTD")
            self.assertEqual(worksheet["B3"].value, "[540000321] Nguyễn Văn A")
            self.assertEqual([worksheet.cell(8, col).value for col in range(1, 6)], ["Kỳ", "Dư nợ", "Lãi suất bình quân", "NIM trước ĐC", "NIM sau ĐC"])
            self.assertEqual(worksheet["A9"].value, "2026-01")
            self.assertEqual(worksheet["B9"].value, 1000)
            self.assertEqual(worksheet["C9"].value, 10)
            self.assertEqual(worksheet["C9"].number_format, '0.00"%"')
        finally:
            workbook.close()

    def test_officer_overview_has_nim_before_and_after(self) -> None:
        january = self.root / "5491_FTPLN_20260131.csv"
        february = self.root / "5491_FTPLN_20260228.csv"
        january.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                ]
            ),
            encoding="utf-8",
        )
        february.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,4,8,1,[540000321] Nguyễn Văn A,00,2000,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        overview = build_officer_overview(
            self.repository,
            SummaryDataType.NIM_DN,
            officer_code="540000321",
            officer="[540000321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
        )

        self.assertEqual(overview.current_period, "2026-02")
        self.assertAlmostEqual(overview.current_nim_before, 4.0)
        self.assertAlmostEqual(overview.current_nim_after, 3.0)
        self.assertAlmostEqual(overview.current_average_rate, 8.0)

    def test_officer_growth_history(self) -> None:
        points = (
            build_officer_overview(
                self.repository,
                SummaryDataType.NIM_DN,
                officer="[none]",
            ).points
        )
        self.assertEqual(points, ())
        from agribank_v3.features.credit.summary.officer_history.models import HistoryPoint

        history_points = (
            HistoryPoint("2026-01", 1000.0, 10.0, 8.0, 7.0),
            HistoryPoint("2026-02", 1500.0, 9.0, 5.0, 4.0),
            HistoryPoint("2026-03", 1200.0, 8.0, 4.0, 3.0),
        )

        growth = build_officer_growth_history(history_points)

        self.assertIsNone(growth[0].delta)
        self.assertIsNone(growth[0].growth_percent)
        self.assertAlmostEqual(growth[1].delta, 500.0)
        self.assertAlmostEqual(growth[1].growth_percent, 50.0)
        self.assertAlmostEqual(growth[2].delta, -300.0)
        self.assertAlmostEqual(growth[2].growth_percent, -20.0)

    def test_officer_growth_zero_previous_balance(self) -> None:
        from agribank_v3.features.credit.summary.officer_history.models import HistoryPoint

        history_points = (
            HistoryPoint("2026-01", 0.0, 0.0, 0.0, 0.0),
            HistoryPoint("2026-02", 1000.0, 10.0, 8.0, 7.0),
        )

        growth = build_officer_growth_history(history_points)

        self.assertAlmostEqual(growth[1].delta, 1000.0)
        self.assertIsNone(growth[1].growth_percent)

    def test_multiple_officer_comparison(self) -> None:
        path = self.root / "5491_FTPLN_20260131.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                    "5491,4,8,1,[540000322] Nguyễn Văn B,00,3000,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        rows = build_multiple_officer_comparison(
            self.repository,
            SummaryDataType.NIM_DN,
            officers=[
                officer_key("[540000321] Nguyễn Văn A"),
                officer_key("[540000322] Nguyễn Văn B"),
            ],
            metric=METRIC_BALANCE,
            branch="5491 - CN Lộc Phát",
        )

        values = {(row.officer.code, row.period): row.value for row in rows}
        self.assertEqual(values[("540000321", "2026-01")], 1000)
        self.assertEqual(values[("540000322", "2026-01")], 3000)

    def test_duplicate_officer_names_use_officer_code(self) -> None:
        path = self.root / "5491_FTPLN_20260131.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                    "5491,4,8,1,[540000999] Nguyễn Văn A,00,3000,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        first = build_officer_overview(
            self.repository,
            SummaryDataType.NIM_DN,
            officer_code="540000321",
            officer="[540000321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
        )
        second = build_officer_overview(
            self.repository,
            SummaryDataType.NIM_DN,
            officer_code="540000999",
            officer="[540000999] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
        )

        self.assertEqual(first.officer.display_name, "Nguyễn Văn A")
        self.assertEqual(second.officer.display_name, "Nguyễn Văn A")
        self.assertAlmostEqual(first.current_balance, 1000.0)
        self.assertAlmostEqual(second.current_balance, 3000.0)
        comparison = build_multiple_officer_comparison(
            self.repository,
            SummaryDataType.NIM_DN,
            officers=[
                officer_key("[540000321] Nguyễn Văn A"),
                officer_key("[540000999] Nguyễn Văn A"),
            ],
            metric=METRIC_BALANCE,
            branch="5491 - CN Lộc Phát",
        )
        self.assertEqual(sorted(row.officer.code for row in comparison), ["540000321", "540000999"])

    def test_branch_nim_uses_weighted_formula(self) -> None:
        path = self.root / "5491_FTPLN_20260131.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,100,,CN",
                    "5491,1,2,0.5,[540000322] Nguyễn Văn B,00,300,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        _series, rows = build_officer_branch_comparison(
            self.repository,
            SummaryDataType.NIM_DN,
            officer_code="540000321",
            officer="[540000321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
            metric=METRIC_NIM_BEFORE,
            filters=HistoryFilters(customer_type="Cá nhân (CN)"),
        )

        branch_row = [row for row in rows if row.officer.display_name == "5491 - CN Lộc Phát"][0]
        self.assertAlmostEqual(branch_row.value or 0, 2.75)
        self.assertNotAlmostEqual(branch_row.value or 0, 4.5)

    def test_branch_average_rate_uses_weighted_formula(self) -> None:
        path = self.root / "5491_FTPLN_20260131.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,100,,CN",
                    "5491,1,2,0.5,[540000322] Nguyễn Văn B,00,300,,CN",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        _series, rows = build_officer_branch_comparison(
            self.repository,
            SummaryDataType.NIM_DN,
            officer_code="540000321",
            officer="[540000321] Nguyễn Văn A",
            branch="5491 - CN Lộc Phát",
            metric=METRIC_AVERAGE_RATE,
            filters=HistoryFilters(customer_type="Cá nhân (CN)"),
        )

        branch_row = [row for row in rows if row.officer.display_name == "5491 - CN Lộc Phát"][0]
        self.assertAlmostEqual(branch_row.value or 0, 4.0)
        self.assertNotAlmostEqual(branch_row.value or 0, 6.0)

    def test_all_customer_types_grouped_per_officer_period(self) -> None:
        path = self.root / "5491_FTPLN_20260131.csv"
        path.write_text(
            "\n".join(
                [
                    "BRCD,FTP,INTRT,MUCFTPDC,CBTD,TRCTCD,LDRBAL,TRREF,CUSTTP",
                    "5491,2,10,1,[540000321] Nguyễn Văn A,00,1000,,CN",
                    "5491,4,8,1,[540000321] Nguyễn Văn A,00,3000,,TC",
                ]
            ),
            encoding="utf-8",
        )
        import_nim_dn(self.repository, self.root)

        rows = OfficerHistoryRepository(self.repository).get_multiple_officer_history(
            SummaryDataType.NIM_DN,
            officers=[officer_key("[540000321] Nguyễn Văn A")],
            branch="5491 - CN Lộc Phát",
            filters=HistoryFilters(),
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["customer_type"], "Tất cả")
        self.assertAlmostEqual(float(rows[0]["balance"]), 4000.0)

    def test_export_each_analysis_tab(self) -> None:
        sample_rows = [{"Kỳ": "2026-01", "Giá trị": 1.23}]

        for tab_key, sheet_name in SHEET_NAMES.items():
            output = export_analysis_rows(sample_rows, self.root / f"{tab_key}.xlsx", tab_key=tab_key)
            workbook = load_workbook(output, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, [sheet_name])
                self.assertEqual(workbook[sheet_name]["A2"].value, "2026-01")
            finally:
                workbook.close()

    def test_loan_compare_preserves_vba_classification(self) -> None:
        previous = self.root / "2003.csv"
        previous.write_text(
            "\n".join(
                [
                    "CUSTSEQ,CUSTNM,DU_NO,ADDR1,OFFICER_NAME",
                    "B,Khach B,100,,CB1",
                    "A,Khach A,100,,CB1",
                    "C,Khach C,100,,CB2",
                ]
            ),
            encoding="utf-8",
        )
        current = self.root / "2010.csv"
        current.write_text(
            "\n".join(
                [
                    "CUSTSEQ,CUSTNM,DU_NO,ADDR1,OFFICER_NAME",
                    "A,Khach A,150,,CB1",
                    "B,Khach B,0,,CB1",
                    "D,Khach D,70,,CB3",
                ]
            ),
            encoding="utf-8",
        )

        output_path = self.root / "BaoCaoTangGiamKH.xlsx"
        result = compare_loan_balances(self.repository, previous, current, export_path=output_path)

        self.assertEqual(result.row_count, 4)
        self.assertEqual(result.output_path, output_path)
        rows = {row["customer_code"]: row for row in self.repository.query_loan_compare(page_size=20).rows}
        self.assertEqual(rows["A"]["category"], "Khach hang vay tang")
        self.assertEqual(rows["B"]["category"], "Khach hang tat toan")
        self.assertEqual(rows["C"]["category"], "Khach hang tat toan")
        self.assertEqual(rows["D"]["category"], "Khach hang vay moi")
        workbook = load_workbook(output_path, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["SoSanh_DuNo", "Sheet2", "Sheet3"])
            worksheet = workbook["SoSanh_DuNo"]
            self.assertEqual([worksheet.cell(1, col).value for col in range(1, 7)], list(LOAN_COMPARE_VBA_HEADERS))
            self.assertEqual([worksheet.cell(row, 1).value for row in range(2, 6)], ["B", "A", "C", "D"])
            self.assertEqual(worksheet["C2"].number_format, "#,##0")
            self.assertEqual(worksheet["D2"].number_format, "#,##0")
        finally:
            workbook.close()

    def test_credit_limit_import_filters_ln01_like_vba(self) -> None:
        path = self.root / "5491_ln01_20260427.csv"
        headers = [f"H{i}" for i in range(63)]
        headers[0] = "BRCD"
        headers[1] = "CUSTSEQ"
        headers[2] = "CUSTNM"
        headers[3] = "TAI_KHOAN"
        headers[62] = "CREDIT_LINE_YPE"
        expired = [""] * 63
        expired[1] = "KH01"
        expired[2] = "Khach 01"
        expired[5] = "100"
        expired[14] = "HD01"
        expired[15] = "20240101"
        expired[17] = "1000"
        expired[18] = "20240115"
        expired[27] = "CB1"
        expired[35] = "Dia chi"
        expired[62] = "Line of Credit"
        soon = expired.copy()
        soon[1] = "KH02"
        soon[14] = "HD02"
        soon[18] = "20240210"
        path.write_text(
            ",".join(headers) + "\n" + ",".join(expired) + "\n" + ",".join(soon),
            encoding="utf-8",
        )

        output_path = self.root / "BaoCaoHanMucHetHan.xlsx"
        result = import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
            export_path=output_path,
        )

        self.assertEqual(result.row_count, 2)
        self.assertEqual(result.output_path, output_path)
        rows = {row["contract_number"]: row for row in self.repository.query_credit_limits(page_size=20).rows}
        self.assertEqual(rows["HD01"]["status"], "Đã hết hạn")
        self.assertEqual(rows["HD02"]["status"], "Sắp hết hạn")
        workbook = load_workbook(output_path, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["HanMuc_HetHan"])
            worksheet = workbook["HanMuc_HetHan"]
            self.assertEqual([worksheet.cell(1, col).value for col in range(1, 11)], list(CREDIT_LIMIT_VBA_HEADERS))
            self.assertEqual(worksheet["D2"].number_format, "dd/mm/yyyy")
            self.assertEqual(worksheet["E2"].number_format, "#,##0")
            self.assertEqual(worksheet["G2"].number_format, "dd/mm/yyyy")
            self.assertEqual(worksheet["J2"].value, "Hợp đồng hạn mức tín dụng ã quá hạn đến thời đểm hiện tại")
            self.assertEqual(worksheet["J3"].value, "Hợp đồng tín dụng dến hạn trong vòng 30 ngày tới theo thời đểm hiện tại")
        finally:
            workbook.close()

    def test_credit_limit_import_writes_excel_batch_not_detail_database(self) -> None:
        path = self._write_ln01_file()
        database_size = self.repository.database_path.stat().st_size

        result = import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )

        self.assertEqual(result.row_count, 2)
        with closing(self.repository.connect()) as database:
            batch_count = database.execute(
                "SELECT COUNT(*) FROM summary_import_history WHERE data_type = ?",
                (SummaryDataType.CREDIT_LIMIT.value,),
            ).fetchone()[0]
            detail_count = database.execute("SELECT COUNT(*) FROM credit_limit_details").fetchone()[0]
        self.assertEqual(batch_count, 0)
        self.assertEqual(detail_count, 0)
        self.assertEqual(self.repository.database_path.stat().st_size, database_size)
        batches = self.repository.list_credit_limit_batches()
        self.assertEqual(len(batches), 1)
        self.assertTrue(batches[0].file_path.is_file())
        workbook = load_workbook(batches[0].file_path, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, [META_SHEET_NAME, DATA_SHEET_NAME])
            worksheet = workbook[DATA_SHEET_NAME]
            self.assertEqual(worksheet.max_row, 4)
            hd01 = [worksheet.cell(row, 7).value for row in range(2, worksheet.max_row + 1)].index("HD01") + 2
            self.assertEqual(worksheet.cell(hd01, 11).value, 150)
            self.assertEqual(worksheet.cell(hd01, 15).value, 2)
        finally:
            workbook.close()

    def test_credit_limit_query_status_is_dynamic_from_reference_date(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )

        january_rows = {
            row["contract_number"]: row
            for row in self.repository.query_credit_limits(page_size=20, reference_date=date(2024, 1, 20), warn_days=30).rows
        }
        march_rows = {
            row["contract_number"]: row
            for row in self.repository.query_credit_limits(page_size=20, reference_date=date(2024, 3, 20), warn_days=30).rows
        }

        self.assertEqual(january_rows["HD01"]["status"], "Đã hết hạn")
        self.assertEqual(january_rows["HD02"]["status"], "Sắp hết hạn")
        self.assertNotIn("HD03", january_rows)
        self.assertEqual(march_rows["HD01"]["status"], "Đã hết hạn")
        self.assertEqual(march_rows["HD02"]["status"], "Đã hết hạn")
        self.assertNotIn("HD03", march_rows)

    def test_credit_limit_duplicate_detection_requires_confirmation_by_default(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )

        with self.assertRaises(SummaryError):
            import_credit_limit_file(
                self.repository,
                path,
                warn_days=30,
                reference_date=date(2024, 1, 20),
            )
        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
            duplicate_policy="new",
        )

        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)

    def test_credit_limit_query_no_data_returns_empty_page(self) -> None:
        result = self.repository.query_credit_limits(page_size=20)

        self.assertEqual(result.total_rows, 0)
        self.assertEqual(result.rows, [])

    def test_credit_limit_backup_restore_roundtrip_uses_hmhethan_files(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        batch = self.repository.list_credit_limit_batches()[0]
        backup = self.repository.backup_credit_limit_storage(self.root / "hmhethan.zip")

        deleted = self.repository.delete_credit_limit_batch(batch.batch_id)
        self.assertTrue(Path(deleted).is_file())
        self.assertEqual(self.repository.query_credit_limits(page_size=20).total_rows, 0)
        restored = self.repository.restore_credit_limit_storage(backup, conflict_policy="skip")

        self.assertEqual(restored["restored"], 1)
        self.assertEqual(self.repository.query_credit_limits(page_size=20).total_rows, 2)

    def test_credit_limit_migrate_legacy_batches_keeps_legacy_tables(self) -> None:
        legacy_batch_id = self.repository.create_batch(
            SummaryDataType.CREDIT_LIMIT,
            period="2024-01-20",
            source_path=self.root / "legacy_ln01.csv",
            imported_by="tester",
            row_count=1,
            duration_ms=1,
            message="legacy",
            source_hash="",
        )
        self.repository.save_credit_limit_rows(
            legacy_batch_id,
            [
                CreditLimitRow(
                    customer_code="KH01",
                    customer_name="Khach 01",
                    contract_number="HD01",
                    approved_date=date(2024, 1, 1),
                    approved_amount=1000,
                    outstanding_balance=100,
                    expiry_date=date(2024, 1, 15),
                    address="Dia chi",
                    officer="CB1",
                    note="Hợp đồng hạn mức tín dụng ã quá hạn đến thời đểm hiện tại",
                    days_to_expiry=-5,
                    status="Đã hết hạn",
                )
            ],
            reference_date="2024-01-20",
            warn_days=30,
            min_limit=0,
        )

        result = self.repository.migrate_legacy_credit_limit_batches()

        self.assertEqual(result, {"migrated": 1, "skipped": 0})
        self.assertEqual(len(self.repository.list_credit_limit_batches()), 1)
        with closing(self.repository.connect()) as database:
            self.assertEqual(database.execute("SELECT COUNT(*) FROM summary_import_history").fetchone()[0], 1)
            self.assertEqual(database.execute("SELECT COUNT(*) FROM credit_limit_details").fetchone()[0], 1)
        self.assertEqual(self.repository.query_credit_limits(page_size=20).total_rows, 1)

    def test_credit_limit_maintenance_ignores_temp_and_reports_invalid_files(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(
            self.repository,
            path,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        storage = self.repository.credit_limit_storage_status().storage_path
        (storage / "~$temp.xlsx").write_text("", encoding="utf-8")
        (storage / "invalid.xlsx").write_text("not an excel file", encoding="utf-8")

        status = self.repository.credit_limit_storage_status()

        self.assertEqual(status.valid_files, 1)
        self.assertEqual(status.invalid_files, 1)
        self.assertEqual(status.temporary_files, 1)

    def test_credit_limit_excel_store_accepts_credit_limit_row(self) -> None:
        store = CreditLimitExcelBatchStore(self.database_path)
        row = self._sample_credit_limit_row()

        metadata = store.create_batch(
            source_path=self.root / "sample_ln01.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="a" * 64,
            source_file_size=1,
        )

        self.assertTrue(metadata.file_path.is_file())
        self.assertEqual(store.query_credit_limits(page_size=20).total_rows, 1)

    def test_credit_limit_excel_writer_does_not_call_get_on_dataclass(self) -> None:
        store = CreditLimitExcelBatchStore(self.database_path)
        row = self._sample_credit_limit_row()

        metadata = store.create_batch(
            source_path=self.root / "sample_ln01.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="b" * 64,
            source_file_size=1,
        )

        self.assertTrue(metadata.file_path.exists())

    def test_credit_limit_row_to_excel_values(self) -> None:
        row = self._sample_credit_limit_row()

        values = credit_limit_row_to_excel_values(row=row, batch_id="batch-1")

        self.assertEqual(values[0], "batch-1")
        self.assertEqual(values[2], "00123")
        self.assertEqual(values[6], "HD001")
        self.assertEqual(values[8], 1_000_000)
        self.assertEqual(values[10], 250_000)

    def test_credit_limit_excel_values_preserve_field_order(self) -> None:
        row = self._sample_credit_limit_row()

        values = credit_limit_row_to_excel_values(row=row, batch_id="batch-1")

        self.assertEqual(values[:5], ["batch-1", "5491", "00123", "Khach 00123", "0000123456"])

    def test_credit_limit_batch_id_passed_separately(self) -> None:
        row = self._sample_credit_limit_row()

        values = credit_limit_row_to_excel_values(row=row, batch_id="external-batch")

        self.assertEqual(values[0], "external-batch")
        self.assertFalse(hasattr(row, "batch_id"))

    def test_credit_limit_metadata_not_read_from_row(self) -> None:
        row = self._sample_credit_limit_row()

        values = credit_limit_row_to_excel_values(row=row, batch_id="metadata-batch")

        self.assertEqual(values[0], "metadata-batch")

    def test_credit_limit_excel_reader_returns_credit_limit_rows(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        batch = self.repository.list_credit_limit_batches()[0]
        rows = self.repository.credit_limit_store._rows_from_file(batch.file_path)

        self.assertTrue(rows)
        self.assertIsInstance(rows[0], CreditLimitRow)

    def test_credit_limit_mapping_adapter_returns_credit_limit_row(self) -> None:
        row = normalize_credit_limit_row(
            {
                "customer_code": "001",
                "customer_name": "Khach",
                "approval_sequence": "HD",
                "approval_date": "2024-01-01",
                "approved_limit": "1.000.000",
                "maturity_date": "2024-01-15",
                "outstanding_balance": "70.000.000",
                "officer_name": "CB1",
            }
        )

        self.assertIsInstance(row, CreditLimitRow)
        self.assertEqual(row.approved_amount, 1_000_000)
        self.assertEqual(row.outstanding_balance, 70_000_000)

    def test_credit_limit_invalid_row_type_raises_clear_error(self) -> None:
        with self.assertRaisesRegex(TypeError, "Unsupported credit limit row type"):
            normalize_credit_limit_row(object())

    def test_credit_limit_import_ln01_creates_valid_xlsx(self) -> None:
        path = self._write_ln01_file()

        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        batch = self.repository.list_credit_limit_batches()[0]
        workbook = load_workbook(batch.file_path, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, [META_SHEET_NAME, DATA_SHEET_NAME])
            worksheet = workbook[DATA_SHEET_NAME]
            self.assertEqual(worksheet.cell(2, 8).is_date, True)
            self.assertIsInstance(worksheet.cell(2, 9).value, int | float)
            self.assertIsInstance(worksheet.cell(2, 11).value, int | float)
        finally:
            workbook.close()

    def test_credit_limit_import_no_attribute_error(self) -> None:
        path = self._write_ln01_file()

        result = import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        self.assertEqual(result.row_count, 2)

    def test_credit_limit_import_failure_cleans_temp_file(self) -> None:
        path = self._write_ln01_file()

        with patch.object(CreditLimitExcelBatchStore, "_write_data_sheet", side_effect=AttributeError("boom")):
            with self.assertRaisesRegex(SummaryError, "Không thể tạo batch"):
                import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        temp_files = list((self.database_path.parent / "HMHETHAN" / "Temp").glob("*"))
        self.assertEqual(temp_files, [])

    def test_credit_limit_import_failure_does_not_create_batch(self) -> None:
        path = self._write_ln01_file()

        with patch.object(CreditLimitExcelBatchStore, "_write_data_sheet", side_effect=AttributeError("boom")):
            with self.assertRaises(SummaryError):
                import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        self.assertEqual(self.repository.list_credit_limit_batches(), [])

    def test_credit_limit_import_failure_does_not_write_database(self) -> None:
        path = self._write_ln01_file()
        size_before = self.repository.database_path.stat().st_size

        with patch.object(CreditLimitExcelBatchStore, "_write_data_sheet", side_effect=AttributeError("boom")):
            with self.assertRaises(SummaryError):
                import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        with closing(self.repository.connect()) as database:
            self.assertEqual(database.execute("SELECT COUNT(*) FROM summary_import_history WHERE data_type = ?", (SummaryDataType.CREDIT_LIMIT.value,)).fetchone()[0], 0)
            self.assertEqual(database.execute("SELECT COUNT(*) FROM credit_limit_details").fetchone()[0], 0)
        self.assertEqual(self.repository.database_path.stat().st_size, size_before)

    def test_credit_limit_import_button_reenabled_after_error(self) -> None:
        tab = CreditLimitTab(self.repository)
        self.addCleanup(tab.deleteLater)
        tab.import_button.setEnabled(False)

        with patch.object(QMessageBox, "warning", return_value=None):
            tab._import_failed(SummaryError("x"))

        self.assertTrue(tab.import_button.isEnabled())

    def test_credit_limit_kpi_accepts_credit_limit_rows(self) -> None:
        store = CreditLimitExcelBatchStore(self.database_path)
        row = self._sample_credit_limit_row()
        store.create_batch(
            source_path=self.root / "sample_ln01.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="c" * 64,
            source_file_size=1,
        )

        dashboard = store.dashboard_credit_limits(reference_date=date(2024, 1, 20))

        self.assertEqual({metric.label: metric.value for metric in dashboard.metrics}["Tổng dư nợ"], "250.000")

    def test_credit_limit_filter_accepts_credit_limit_rows(self) -> None:
        row = self._sample_credit_limit_row()

        selected = self.repository.credit_limit_store.filter_rows(
            batch_id=None,
            min_limit=0,
            warn_days=30,
            reference_date=date(2024, 1, 20),
        )
        self.assertEqual(selected, [])
        self.assertEqual(row.customer_code, "00123")

    def test_credit_limit_table_model_accepts_credit_limit_rows(self) -> None:
        row = self._sample_credit_limit_row()

        self.assertEqual(_credit_limit_row_value(row, "outstanding_balance"), 250_000)

    def test_credit_limit_export_accepts_credit_limit_rows(self) -> None:
        tab = CreditLimitTab(self.repository)
        self.addCleanup(tab.deleteLater)
        tab.current_rows = [self._sample_credit_limit_row()]

        rows = tab._export_rows()

        self.assertEqual(rows[0]["Tổng dư nợ HĐTD"], 250_000)

    def test_credit_limit_export_money_columns_are_numeric(self) -> None:
        tab = CreditLimitTab(self.repository)
        self.addCleanup(tab.deleteLater)
        tab.current_rows = [self._sample_credit_limit_row()]

        output = export_rows(tab._export_rows(), self.root / "credit-limit-ui.xlsx", title=tab.title, sheet_name=tab.title)

        workbook = load_workbook(output, data_only=True)
        try:
            worksheet = workbook[tab.title]
            headers = [worksheet.cell(2, column).value for column in range(1, worksheet.max_column + 1)]
            approved_column = headers.index("Hạn mức TD") + 1
            outstanding_column = headers.index("Tổng dư nợ HĐTD") + 1
            self.assertEqual(worksheet.cell(3, approved_column).value, 1_000_000)
            self.assertEqual(worksheet.cell(3, outstanding_column).value, 250_000)
            self.assertEqual(worksheet.cell(3, approved_column).number_format, "#,##0")
            self.assertEqual(worksheet.cell(3, outstanding_column).number_format, "#,##0")
        finally:
            workbook.close()

    def test_credit_limit_export_includes_summary_metrics(self) -> None:
        tab = CreditLimitTab(self.repository)
        self.addCleanup(tab.deleteLater)
        tab.current_rows = [self._sample_credit_limit_row()]
        tab.current_dashboard = DashboardData(
            metrics=(
                DashboardMetric("HĐTD đã hết hạn", "1", "expired"),
                DashboardMetric("HĐTD sắp hết hạn", "0", "expiring"),
                DashboardMetric("Tổng HĐTD cảnh báo", "1", "total"),
                DashboardMetric("Tổng hạn mức", "1.000.000", "limit"),
                DashboardMetric("Tổng dư nợ", "250.000", "balance"),
            )
        )

        output = export_credit_limit_view_report(
            tab._export_rows(),
            tab.current_dashboard.metrics,
            self.root / "credit-limit-summary.xlsx",
            title=tab.title,
            sheet_name=tab.title,
        )

        workbook = load_workbook(output, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["TongHop", tab.title])
            summary = workbook["TongHop"]
            metrics = {summary.cell(row, 1).value: summary.cell(row, 2).value for row in range(3, summary.max_row + 1)}
            self.assertEqual(metrics["Tổng dư nợ"], 250_000)
            self.assertEqual(metrics["Tổng hạn mức"], 1_000_000)
            self.assertEqual(metrics["Tổng HĐTD cảnh báo"], 1)
        finally:
            workbook.close()

    def test_credit_limit_legacy_mapping_normalized_once(self) -> None:
        mapping = {"customer_code": "001", "approval_sequence": "HD", "approved_limit": 100, "maturity_date": "2024-01-15"}

        row = excel_record_to_credit_limit_row(mapping)

        self.assertIsInstance(row, CreditLimitRow)

    def test_credit_limit_total_outstanding_kpi(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        metrics = {metric.label: metric.value for metric in self.repository.dashboard_credit_limits(reference_date=date(2024, 1, 20)).metrics}

        self.assertEqual(metrics["Tổng dư nợ"], "350")

    def test_credit_limit_total_outstanding_uses_outstanding_balance(self) -> None:
        row = self._sample_credit_limit_row()
        store = CreditLimitExcelBatchStore(self.database_path)
        store.create_batch(
            source_path=self.root / "sample_ln01.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="d" * 64,
            source_file_size=1,
        )

        metrics = {metric.label: metric.value for metric in store.dashboard_credit_limits(reference_date=date(2024, 1, 20)).metrics}

        self.assertEqual(metrics["Tổng dư nợ"], "250.000")

    def test_credit_limit_total_limit_uses_approved_limit(self) -> None:
        row = self._sample_credit_limit_row()
        store = CreditLimitExcelBatchStore(self.database_path)
        store.create_batch(
            source_path=self.root / "sample_ln01.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="e" * 64,
            source_file_size=1,
        )

        metrics = {metric.label: metric.value for metric in store.dashboard_credit_limits(reference_date=date(2024, 1, 20)).metrics}

        self.assertEqual(metrics["Tổng hạn mức"], "1.000.000")

    def test_credit_limit_outstanding_sum_all_filtered_rows(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=120, reference_date=date(2024, 1, 20))

        metrics = {metric.label: metric.value for metric in self.repository.dashboard_credit_limits(reference_date=date(2024, 1, 20), warn_days=120).metrics}

        self.assertEqual(metrics["Tổng dư nợ"], "650")

    def test_credit_limit_outstanding_not_limited_to_current_page(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=120, reference_date=date(2024, 1, 20))

        page = self.repository.query_credit_limits(page_size=1, reference_date=date(2024, 1, 20), warn_days=120)
        metrics = {metric.label: metric.value for metric in self.repository.dashboard_credit_limits(reference_date=date(2024, 1, 20), warn_days=120).metrics}

        self.assertEqual(len(page.rows), 1)
        self.assertEqual(metrics["Tổng dư nợ"], "650")

    def test_credit_limit_zero_outstanding_displays_zero(self) -> None:
        row = self._sample_credit_limit_row()
        row = CreditLimitRow(**{**{field: getattr(row, field) for field in row.__dataclass_fields__}, "outstanding_balance": 0})
        store = CreditLimitExcelBatchStore(self.database_path)
        store.create_batch(
            source_path=self.root / "zero_ln01.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="f" * 64,
            source_file_size=1,
        )

        metrics = {metric.label: metric.value for metric in store.dashboard_credit_limits(reference_date=date(2024, 1, 20)).metrics}

        self.assertEqual(metrics["Tổng dư nợ"], "0")

    def test_credit_limit_missing_outstanding_field_displays_na(self) -> None:
        store = CreditLimitExcelBatchStore(self.database_path)
        row = self._sample_credit_limit_row()
        metadata = store.create_batch(
            source_path=self.root / "missing-outstanding.csv",
            rows=(row,),
            accepted_rows=(row,),
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="1" * 64,
            source_file_size=1,
        )
        workbook = load_workbook(metadata.file_path)
        try:
            worksheet = workbook[DATA_SHEET_NAME]
            headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
            worksheet.delete_cols(headers.index("outstanding_balance") + 1)
            workbook.save(metadata.file_path)
        finally:
            workbook.close()
        store.invalidate_cache()

        metric = {metric.label: metric for metric in store.dashboard_credit_limits(reference_date=date(2024, 1, 20)).metrics}["Tổng dư nợ"]

        self.assertEqual(metric.value, "—")
        self.assertEqual(metric.detail, "Batch này không có dữ liệu dư nợ HĐTD.")

    def test_credit_limit_numeric_outstanding_from_excel(self) -> None:
        row = excel_record_to_credit_limit_row({"outstanding_balance": 70_000_000})

        self.assertEqual(row.outstanding_balance, 70_000_000)

    def test_credit_limit_string_outstanding_normalized(self) -> None:
        row = excel_record_to_credit_limit_row({"outstanding_balance": "70.000.000"})

        self.assertEqual(row.outstanding_balance, 70_000_000)

    def test_credit_limit_total_outstanding_updates_after_filter(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        metrics = {metric.label: metric.value for metric in self.repository.dashboard_credit_limits(status="Đã hết hạn", reference_date=date(2024, 1, 20)).metrics}

        self.assertEqual(metrics["Tổng dư nợ"], "150")

    def test_credit_limit_total_outstanding_updates_reference_date(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        metrics = {metric.label: metric.value for metric in self.repository.dashboard_credit_limits(reference_date=date(2024, 3, 20), warn_days=30).metrics}

        self.assertEqual(metrics["Tổng dư nợ"], "350")

    def test_credit_limit_chart_title(self) -> None:
        self.assertEqual(CreditLimitOfficerChart.title, "Top CBTD theo số HĐTD cảnh báo")

    def test_credit_limit_chart_aggregates_by_officer_code(self) -> None:
        store = CreditLimitExcelBatchStore(self.database_path)
        row = self._sample_credit_limit_row()
        rows = (
            replace(row, officer="Nguyễn Văn A", officer_code="540000001", contract_number="HD001"),
            replace(row, officer="Nguyễn Văn A", officer_code="540000002", contract_number="HD002"),
        )
        store.create_batch(
            source_path=self.root / "officer-code.csv",
            rows=rows,
            accepted_rows=rows,
            reference_date=date(2024, 1, 20),
            min_limit=0,
            warn_days=30,
            source_file_sha256="2" * 64,
            source_file_size=1,
        )

        payload = store.dashboard_credit_limits(reference_date=date(2024, 1, 20)).pies

        self.assertEqual({item[0] for item in payload}, {"540000001", "540000002"})
        self.assertEqual(len(payload), 2)

    def test_credit_limit_chart_expired_count(self) -> None:
        payload = ("", "CB1", 21, 5, 26, 100, 50)

        self.assertEqual(_chart_expired(payload), 21)

    def test_credit_limit_chart_expiring_count(self) -> None:
        payload = ("", "CB1", 21, 5, 26, 100, 50)

        self.assertEqual(_chart_expiring(payload), 5)

    def test_credit_limit_chart_total_warning_count(self) -> None:
        payload = ("", "CB1", 21, 5, 26, 100, 50)

        self.assertEqual(_chart_total(payload), 26)

    def test_credit_limit_chart_total_limit(self) -> None:
        payload = ("", "CB1", 21, 5, 26, 12_500_000_000, 4_200_000_000)

        self.assertEqual(_chart_total_limit(payload), 12_500_000_000)

    def test_credit_limit_chart_total_outstanding(self) -> None:
        payload = ("", "CB1", 21, 5, 26, 12_500_000_000, 4_200_000_000)

        self.assertEqual(_chart_total_outstanding(payload), 4_200_000_000)

    def test_credit_limit_chart_top_limit(self) -> None:
        chart = CreditLimitOfficerChart()
        self.addCleanup(chart.deleteLater)

        chart.set_values(tuple(("", f"CB{i}", 1, 0, 1, 0, 0) for i in range(20)))

        self.assertEqual(len(chart.values), 10)

    def test_credit_limit_chart_integer_axis(self) -> None:
        self.assertEqual(_integer_ticks(4), [0, 1, 2, 3, 4])
        self.assertEqual(_integer_ticks(26), [0, 10, 20, 30])

    def test_credit_limit_chart_total_labels(self) -> None:
        payload = ("", "CB1", 1, 2, 3, 0, 0)

        self.assertIn("Tổng cảnh báo: 3", _credit_limit_chart_tooltip(payload))

    def test_credit_limit_chart_long_officer_names(self) -> None:
        long_name = "Nguyễn Văn Cán Bộ Tín Dụng Tên Rất Dài"
        payload = ("", long_name, 1, 2, 3, 0, 0)

        self.assertIn(long_name, _credit_limit_chart_tooltip(payload))

    def test_credit_limit_chart_unknown_officer_group(self) -> None:
        path = self._write_ln01_file()
        content = path.read_text(encoding="utf-8").replace("CB1", "")
        path.write_text(content, encoding="utf-8")
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        payload = self.repository.dashboard_credit_limits(reference_date=date(2024, 1, 20)).pies[0]

        self.assertEqual(payload[1], "Không xác định CBTD")

    def test_credit_limit_chart_empty_state(self) -> None:
        chart = CreditLimitOfficerChart()
        self.addCleanup(chart.deleteLater)

        chart.set_values(())

        self.assertEqual(chart.values, ())

    def test_credit_limit_chart_tooltip_on_hover(self) -> None:
        chart = CreditLimitOfficerChart()
        self.addCleanup(chart.deleteLater)

        chart._show_payload_tooltip(("", "CB1", 1, 2, 3, 100, 50), "Đã hết hạn", QPoint(10, 10))

        self.assertTrue(chart.tooltip_is_visible())

    def test_credit_limit_chart_tooltip_officer_name(self) -> None:
        self.assertIn("Cán bộ: CB1", _credit_limit_chart_tooltip(("", "CB1", 1, 2, 3, 100, 50)))

    def test_credit_limit_chart_tooltip_expired_count(self) -> None:
        self.assertIn("HĐTD đã hết hạn: 1", _credit_limit_chart_tooltip(("", "CB1", 1, 2, 3, 100, 50)))

    def test_credit_limit_chart_tooltip_expiring_count(self) -> None:
        self.assertIn("HĐTD sắp hết hạn: 2", _credit_limit_chart_tooltip(("", "CB1", 1, 2, 3, 100, 50)))

    def test_credit_limit_chart_tooltip_total_warning(self) -> None:
        self.assertIn("Tổng cảnh báo: 3", _credit_limit_chart_tooltip(("", "CB1", 1, 2, 3, 100, 50)))

    def test_credit_limit_chart_tooltip_total_limit(self) -> None:
        self.assertIn("Tổng hạn mức: 100 đồng", _credit_limit_chart_tooltip(("", "CB1", 1, 2, 3, 100, 50)))

    def test_credit_limit_chart_tooltip_total_outstanding(self) -> None:
        self.assertIn("Tổng dư nợ: 50 đồng", _credit_limit_chart_tooltip(("", "CB1", 1, 2, 3, 100, 50)))

    def test_credit_limit_chart_tooltip_persists_while_hovered(self) -> None:
        chart = CreditLimitOfficerChart()
        self.addCleanup(chart.deleteLater)

        chart._show_payload_tooltip(("", "CB1", 1, 2, 3, 100, 50), "Sắp hết hạn", QPoint(10, 10))

        self.assertTrue(chart.tooltip_is_visible())

    def test_credit_limit_chart_tooltip_hides_on_leave(self) -> None:
        chart = CreditLimitOfficerChart()
        self.addCleanup(chart.deleteLater)
        chart._show_payload_tooltip(("", "CB1", 1, 2, 3, 100, 50), "", QPoint(10, 10))

        chart.hide_tooltip()

        self.assertFalse(chart.tooltip_is_visible())

    def test_credit_limit_chart_tooltip_stays_inside_screen(self) -> None:
        chart = CreditLimitOfficerChart()
        chart.resize(240, 180)
        self.addCleanup(chart.deleteLater)

        chart._show_payload_tooltip(("", "CB1", 1, 2, 3, 100, 50), "", QPoint(230, 170))

        self.assertLessEqual(chart.tooltip_label.x() + chart.tooltip_label.width(), chart.width())
        self.assertLessEqual(chart.tooltip_label.y() + chart.tooltip_label.height(), chart.height())

    def test_credit_limit_kpi_chart_table_same_filter(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        page = self.repository.query_credit_limits(status="Đã hết hạn", reference_date=date(2024, 1, 20))
        dashboard = self.repository.dashboard_credit_limits(status="Đã hết hạn", reference_date=date(2024, 1, 20))

        self.assertEqual(page.total_rows, 1)
        self.assertEqual(dashboard.pies[0][4], 1)
        self.assertEqual({metric.label: metric.value for metric in dashboard.metrics}["Tổng HĐTD cảnh báo"], "1")

    def test_credit_limit_batch_change_refreshes_all(self) -> None:
        self.assertEqual(self.repository.dashboard_credit_limits().metrics[2].value, "0")

    def test_credit_limit_status_change_refreshes_all(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        self.assertEqual(self.repository.dashboard_credit_limits(status="Sắp hết hạn", reference_date=date(2024, 1, 20)).metrics[2].value, "1")

    def test_credit_limit_officer_change_refreshes_all(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        self.assertEqual(self.repository.dashboard_credit_limits(officer="CB1", reference_date=date(2024, 1, 20)).pies[0][1], "CB1")

    def test_credit_limit_search_refreshes_all(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        dashboard = self.repository.dashboard_credit_limits(search="HD02", reference_date=date(2024, 1, 20))

        self.assertEqual(dashboard.metrics[2].value, "1")

    def test_credit_limit_reference_date_refreshes_all(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        dashboard = self.repository.dashboard_credit_limits(reference_date=date(2024, 3, 20), warn_days=30)

        self.assertEqual(dashboard.metrics[0].value, "2")

    def test_credit_limit_warning_days_refreshes_all(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))

        dashboard = self.repository.dashboard_credit_limits(reference_date=date(2024, 1, 20), warn_days=120)

        self.assertEqual(dashboard.metrics[2].value, "3")

    def test_credit_limit_minimum_limit_refreshes_all(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=120, reference_date=date(2024, 1, 20))

        dashboard = self.repository.dashboard_credit_limits(reference_date=date(2024, 1, 20), warn_days=120, min_limit=999)

        self.assertEqual(dashboard.metrics[2].value, "3")

    def test_credit_limit_refresh_reads_excel_once(self) -> None:
        path = self._write_ln01_file()
        import_credit_limit_file(self.repository, path, warn_days=30, reference_date=date(2024, 1, 20))
        batch = self.repository.list_credit_limit_batches()[0]

        self.repository.query_credit_limits(reference_date=date(2024, 1, 20))
        cache_size = len(self.repository.credit_limit_store._row_cache)
        self.repository.dashboard_credit_limits(reference_date=date(2024, 1, 20))

        self.assertEqual(len(self.repository.credit_limit_store._row_cache), cache_size)
        self.assertIn(batch.file_path, self.repository.credit_limit_store._row_cache)

    def test_export_rows_supports_excel_csv_pdf(self) -> None:
        rows = [{"A": "x", "B": 1}]
        xlsx = export_rows(rows, self.root / "out.xlsx", title="Test")
        csv_path = export_rows(rows, self.root / "out.csv", title="Test")
        pdf = export_rows(rows, self.root / "out.pdf", title="Test")

        self.assertTrue(xlsx.is_file())
        self.assertTrue(csv_path.read_text(encoding="utf-8-sig").startswith("A,B"))
        self.assertTrue(pdf.read_bytes().startswith(b"%PDF"))

    def test_backup_and_restore_roundtrip(self) -> None:
        backup = self.repository.backup_database(self.root / "backup.zip")

        self.assertTrue(backup.is_file())
        safety = self.repository.restore_database(backup)
        self.assertTrue(safety.is_file())

    def test_workbook_compare_reports_cell_differences(self) -> None:
        from openpyxl import Workbook

        vba = self.root / "vba.xlsx"
        python = self.root / "python.xlsx"
        for path, value in ((vba, "A"), (python, "B")):
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "SoSanh_DuNo"
            worksheet["A1"] = value
            workbook.save(path)

        result = compare_workbooks(vba, python, self.root / "diff.xlsx")

        self.assertFalse(result.same)
        self.assertGreaterEqual(result.diff_count, 1)
        self.assertTrue(result.report_path.is_file())


if __name__ == "__main__":
    unittest.main()
